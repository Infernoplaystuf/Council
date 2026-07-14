# ============================================================
# council_gui_engine.py  —  Council main GUI
# ============================================================
# Optional dependencies:
#   pip install paramiko        # SSH support for remote nodes
#   pip install sounddevice soundfile   # microphone STT input
#   pip install faster-whisper  # local speech-to-text
#   pip install pyttsx3         # text-to-speech playback
# ============================================================

from __future__ import annotations

import _windll_bootstrap  # noqa: F401  — Windows: route llama-cpp to torch's CUDA DLLs

import json as _json
import os
# Demo siloing: prevent the council from pulling cross-session memory into prompts.
# Memory files stay on disk; they just are not injected during the demo.
os.environ.setdefault('COUNCIL_DEMO_SILO', '1')
# Single-voice mode: only the Writer speaks. No specialists, no multi-role
# panel deliberation surfaced to the user. The Writer still gets every piece
# of injected context (files, vault matches, analyst results).
os.environ.setdefault('COUNCIL_SINGLE_VOICE', '1')


def _single_voice_mode():
    return os.environ.get('COUNCIL_SINGLE_VOICE', '').strip().lower() in ('1', 'true', 'yes', 'on')
import queue
import re as _re
import subprocess
import sys
import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

import council_engine as ce
# apothecary_engine is lazily imported inside _build_apoth_tab so
# consumer builds (no --advanced / COUNCIL_ADVANCED=1) never load
# the SSH provisioning code or its dependencies. Keep `ae` as None
# at module scope so any stray reference fails loudly in dev rather
# than silently working only in advanced mode.
ae = None  # type: ignore[assignment]
import branding
# Set the Windows AppUserModelID *before* any Tk root is constructed, so the
# taskbar uses our cog+flame icon instead of the host (Spyder/python.exe)
# icon. No-op on macOS / Linux.
branding.set_app_user_model_id()
import onboarding
import task_memory as _task_memory_mod
import specialists as _spec
import crash_reporter
import licensing
import activation_dialog
import updater
import splash
import data_index

# ── Agent modules (graceful optional imports) ─────────────────
try:
    import coder_agent as ca
    _CODER_AGENT_OK = True
except Exception:
    _CODER_AGENT_OK = False

try:
    import intern_agent as ia
    _INTERN_AGENT_OK = True
except Exception:
    _INTERN_AGENT_OK = False

try:
    import vault_agent as va
    _VAULT_AGENT_OK = True
except Exception:
    va = None
    _VAULT_AGENT_OK = False

try:
    import sage_agent as sa
    _SAGE_OK = True
except Exception:
    sa = None
    _SAGE_OK = False

try:
    import vault_scraper as vs
    _SCRAPER_OK = True
except Exception:
    vs = None
    _SCRAPER_OK = False

try:
    import dream3d_primer as d3p
    _DREAM3D_OK = True
except Exception:
    _DREAM3D_OK = False

try:
    import dream3d_council_patch as d3d_patch
    _D3D_PATCH_OK = True
except Exception:
    _D3D_PATCH_OK = False

try:
    import vault_rag as vr
    _RAG_OK = True
except Exception:
    _RAG_OK = False

class _LazyModule:
    """Imports the real module on FIRST attribute access, deferring heavy graph
    deps (plotly + matplotlib, ~seconds) out of app startup. The Grapher tab
    builds with widgets only; gd/ge/gp are touched solely on user plot actions,
    so this moves those imports off the startup path with zero call-site
    changes (`ge.foo` transparently loads graph_engine on first use)."""
    __slots__ = ("_lm_name", "_lm_mod")

    def __init__(self, name):
        self._lm_name = name
        self._lm_mod = None

    def __getattr__(self, attr):
        mod = self._lm_mod          # a slot — no __getattr__ recursion
        if mod is None:
            import importlib
            mod = importlib.import_module(self._lm_name)
            self._lm_mod = mod
        return getattr(mod, attr)


try:
    import importlib.util as _ilu
    _GRAPHER_OK = all(_ilu.find_spec(_m) is not None
                      for _m in ("graph_data", "graph_engine",
                                 "graph_personality"))
except Exception:
    _GRAPHER_OK = False
if _GRAPHER_OK:
    # Lazy proxies — the underlying modules (and plotly/matplotlib) load on
    # first use, not at import.
    gd = _LazyModule("graph_data")
    ge = _LazyModule("graph_engine")
    gp = _LazyModule("graph_personality")

try:
    import tkinterweb
    _TKWEB_OK = True
except Exception:
    _TKWEB_OK = False

# ============================================================
# Mode flags
# ============================================================
# Advanced mode exposes power-user tabs (IDE, Agents, Nodes, Apothecary,
# Vault Health, Librarian snapshots). Enabled via:
#   • CLI flag:    python council_gui_engine.py --advanced
#   • Env var:     COUNCIL_ADVANCED=1
# Off by default for the polished customer experience.
_ADVANCED_MODE = (
    os.environ.get("COUNCIL_ADVANCED", "").lower() in ("1", "true", "yes")
    or "--advanced" in sys.argv
)

# ============================================================
# Agent event types
# ============================================================

@dataclass
class AgentEvent:
    who: str
    kind: str   # "thought" | "action" | "observation" | "final" | "token" | "phase"
    text: str


@dataclass
class AgentContext:
    user_text: str
    shared: Dict[str, Any] = field(default_factory=dict)


ToolFn = Callable[[Dict[str, Any]], Tuple[bool, str, Dict[str, Any]]]
_TOOL_JSON_RE = _re.compile(r"\{.*\}", _re.DOTALL)


# ============================================================
# Helpers
# ============================================================

def _extract_tool_calls(text: str) -> List[Dict[str, Any]]:
    calls: List[Dict[str, Any]] = []
    for m in _TOOL_JSON_RE.finditer(text):
        blob = m.group(0).strip()
        try:
            obj = _json.loads(blob)
        except Exception:
            continue
        if isinstance(obj, dict) and "tool" in obj:
            calls.append({"tool": obj.get("tool"), "args": obj.get("args", {})})
        elif isinstance(obj, dict) and "tool_calls" in obj and isinstance(obj["tool_calls"], list):
            for tc in obj["tool_calls"]:
                if isinstance(tc, dict) and "tool" in tc:
                    calls.append({"tool": tc.get("tool"), "args": tc.get("args", {})})
    return calls


def _extract_code_block(text: str) -> str:
    m = _re.search(r"```(?:python)?\s*(.*?)```", text, flags=_re.DOTALL | _re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _safe_script_basename(name: str) -> str:
    name = (name or "").strip()
    name = _re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._-")
    if not name:
        name = "script"
    if len(name) > 60:
        name = name[:60]
    if name.lower().endswith(".py"):
        name = name[:-3]
    return name


def _parse_script_json(text: str) -> Tuple[str, str]:
    m = _re.search(r"\{.*\}", text, flags=_re.DOTALL)
    if not m:
        return "", ""
    try:
        obj = _json.loads(m.group(0))
    except Exception:
        return "", ""
    fn = str(obj.get("filename", "")).strip()
    code = str(obj.get("code", "")).strip()
    return fn, code


# ── Route → (panel, synth) mapping ──────────────────────────────────────────
# Maps the judge's route decision to the most appropriate council panel.
# Coding tasks frontload Coder + Intern + Skeptic (production hardening).
# Writing/explanation tasks frontload Writer + Intern + Artist.
# Design tasks frontload Artist. Fallback: broad general panel.
_PANEL_FOR_ROUTE: Dict[str, tuple] = {
    "chat":       (["writer",     "peasant"],             "writer"),  # pure conversation
    "writer":     (["writer",     "intern",  "peasant"],  "writer"),  # research/docs
    "ide":        (["coder", "intern",  "skeptic"],  "writer"),  # code tasks
    "artist":     (["artist",     "writer",  "intern" ],  "writer"),  # visual/UI
    "intern":     (["intern",     "writer",  "peasant"],  "writer"),  # planning
    "coder": (["coder", "intern",  "skeptic"],  "writer"),  # architecture
    "peasant":    (["writer",     "peasant"           ],  "writer"),  # simple explain
    "sage":       (["sage",       "writer",  "peasant"],  "writer"),  # domain knowledge
    "strategist": (["strategist", "coder", "skeptic"], "writer"), # planning/strategy
    "content":    (["content",    "writer",  "strategist"], "writer"),  # content creation
    "director":   (["director",   "writer",  "content"],   "writer"),  # style analysis + scripting
    "_default":   (["writer",     "intern",  "peasant"],   "writer"),  # fallback
}


# ── Code-block filter for conversational responses ───────────────────────────
import re as _re

_CODE_REQUEST_SIGNALS = {
    "write", "create", "make", "build", "generate", "implement",
    "modify", "edit", "update", "fix", "refactor", "patch", "change",
    "script", "function", "class", "program", "code", ".py", ".sh",
    ".bat", "def ", "import ", "```",
}

def _user_wants_code(query: str) -> bool:
    """Return True only if the user explicitly asked for code."""
    q = query.lower()
    # Must have at least one action word AND one code-context word together
    action_words  = {"write", "create", "make", "build", "generate", "implement",
                     "modify", "edit", "update", "fix", "refactor", "patch"}
    code_words    = {"script", "function", "class", "program", "code",
                     ".py", ".sh", ".bat", "def ", "import "}
    has_action = any(w in q for w in action_words)
    has_code   = any(w in q for w in code_words)
    # Also treat explicit technical markers as code requests
    explicit = any(m in query for m in ("```", "def ", "import ", ".py", ".sh", ".bat"))
    return (has_action and has_code) or explicit


def _strip_code_blocks(text: str) -> str:
    """Remove all fenced code blocks from a response."""
    # Remove triple-backtick blocks (with or without language tag)
    cleaned = _re.sub(r"```[\w]*\n?[\s\S]*?```", "", text, flags=_re.MULTILINE)
    # Remove lines that are just indented code (4-space indent used as code)
    # Only remove if 3+ consecutive indented lines (a real code block, not a quote)
    lines = cleaned.split("\n")
    out, run = [], 0
    for line in lines:
        if line.startswith("    ") and line.strip():
            run += 1
        else:
            if run >= 3:
                # drop the accumulated indented block
                for _ in range(run):
                    if out:
                        out.pop()
            run = 0
            out.append(line)
    if run < 3:
        pass  # trailing indented block was short, already in out via append
    cleaned = "\n".join(out)
    # Collapse 3+ blank lines to 2
    cleaned = _re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    return cleaned



# ---- File injection -------------------------------------------------------
# Detects file paths in the user message, reads them, and prepends their
# contents so every council member sees the actual data rather than
# hallucinating about it.

def _is_wsl() -> bool:
    """True when running inside WSL (any flavour). Used to pick a higher
    default Tk scaling factor — WSLg always reports 96 DPI even when the
    Windows host is at 150 % / 200 % scaling, so Tkinter widgets render
    tiny without an explicit multiplier."""
    try:
        if "microsoft" in (os.uname().release or "").lower():  # type: ignore[attr-defined]
            return True
    except Exception:
        pass
    try:
        with open("/proc/version", "r", encoding="utf-8", errors="ignore") as fh:
            return "microsoft" in fh.read().lower()
    except Exception:
        return False


_FILE_PATH_RE = _re.compile(r'[a-zA-Z]:[/\\]\S+|/\S+')
_FILE_READ_CHAR_LIMIT = 12000  # total chars per file in the injected block
# Extracts the file/folder name out of an injection cost label like
# "[VAULT MATCH: sales.csv]" / "[FILE: x]" / "[FOLDER: y]" — used per-label in
# the per-turn provenance source assembly (hoisted out of that loop).
_SOURCE_LABEL_RE = _re.compile(
    r"\[(?:VAULT MATCH(?:\s*\(pinned\))?|FILE|FOLDER):\s*(.+?)\]")


# ----- Token-aware injection helpers ---------------------------------------
# These exist because llama-cpp-python silently clips any prompt that exceeds
# `n_ctx`, which is the #1 cause of "the model hallucinated even though I gave
# it the file." Instead of letting the runtime do silent truncation, we:
#   • tag every block with its token cost so the model sees what it's getting
#   • cap each block at n_ctx/8 tokens (head + tail with a visible marker)
#   • drop the lowest-priority blocks (vault matches) when the assembled
#     prompt would still exceed the safe input budget
# All callers go through `_inject_file_contents`, which orchestrates this.

_BLOCK_HEADER_RE = _re.compile(r"\s*~[\d,]+\s+tokens\s*")


from functools import lru_cache as _lru_cache


@_lru_cache(maxsize=4096)
def _estimate_block_tokens_cached(block_text: str) -> int:
    """Inner cacheable form. Tokenisation is pure with respect to its
    input string — the same block always counts to the same number of
    tokens for a given model, and the model isn't hot-swapped during
    a session. 4096 entries × few hundred bytes each ~= 1-2 MB cap."""
    try:
        import council_engine as _ce
        return _ce.estimate_tokens(block_text or "")
    except Exception:
        return max(1, (len(block_text or "") + 3) // 4)


def _estimate_block_tokens(block_text: str) -> int:
    """Cheap token estimator — uses the loaded llama tokenizer when
    available, falls back to chars/4. Never raises.

    Wrapper around the lru_cache'd impl so callers can pass non-string
    inputs (None, bytes from a misbehaving caller) without poisoning
    the cache with un-hashable keys.
    """
    if not isinstance(block_text, str):
        block_text = "" if block_text is None else str(block_text)
    return _estimate_block_tokens_cached(block_text)


def _tag_block_header(block_text: str, token_count: int) -> str:
    """Stamp `  ~N tokens` into the [LABEL: ...] header on the first line so
    the model — and the user, via `context info` — can see how many tokens
    each injection costs. Idempotent: re-tagging strips any prior tag first.

    Recognises both `[LABEL]` and `[LABEL: name]` openings. If the first
    line isn't bracketed, we synthesize an `[INJECTION ~N tokens]` header.
    """
    if not block_text:
        return block_text
    nl = block_text.find('\n')
    first = block_text[:nl] if nl >= 0 else block_text
    rest = block_text[nl:] if nl >= 0 else ''
    tag = f"  ~{token_count:,} tokens"
    if first.lstrip().startswith('[') and ']' in first:
        cleaned = _BLOCK_HEADER_RE.sub("", first, count=1)
        close = cleaned.find(']')
        if close > 0:
            tagged = cleaned[:close] + tag + cleaned[close:]
        else:
            tagged = cleaned + tag
    else:
        tagged = f"[INJECTION{tag}]\n" + first
    return tagged + rest


def _smart_truncate_text(text: str, max_chars: int) -> tuple:
    """Trim long text to head + tail with a visible elision marker.

    Returns (trimmed, was_truncated). The marker tells the model what
    fraction is missing so it doesn't pretend it saw the whole file —
    that's the whole point vs llama-cpp's silent tail-clip.
    """
    if not text or len(text) <= max_chars:
        return text, False
    half = max(200, (max_chars - 120) // 2)
    head = text[:half]
    tail = text[-half:]
    skipped = len(text) - len(head) - len(tail)
    marker = (
        f"\n... [truncated: {skipped:,} characters elided from the middle — "
        f"ask for specific lines or use the analyst for exact answers] ...\n"
    )
    return head + marker + tail, True


def _text_peek(path, max_bytes: int = 4096) -> str:
    """First ~max_bytes of a file decoded as text, or a binary note. Bounded."""
    p = Path(path)
    try:
        data = p.read_bytes()[:max_bytes]
    except Exception as e:
        return f"(could not read: {e})"
    try:
        txt = data.decode("utf-8")
    except Exception:
        try:
            txt = data.decode("utf-8", errors="replace")
        except Exception:
            try:
                return f"(binary file, {p.stat().st_size} bytes)"
            except Exception:
                return "(binary file)"
    return txt


def _data_preview_text(path, max_rows: int = 50):
    """Return ``(schema_text, rows_text)`` for a quick, MODEL-FREE preview of a
    data file — column dtypes + the first ``max_rows`` rows. Bounded reads.
    Best-effort across CSV/TSV/parquet/Excel/JSON; falls back to a short text
    peek for anything non-tabular. Pure + UI-free so it's unit-testable."""
    p = Path(path)
    suf = p.suffix.lower()
    try:
        import pandas as pd
    except Exception:
        return ("(pandas unavailable — showing a raw text peek.)", _text_peek(p))
    df = None
    try:
        if suf in (".csv", ".tsv", ".txt"):
            sep = "\t" if suf == ".tsv" else None
            df = pd.read_csv(p, nrows=max_rows, sep=sep, engine="python")
        elif suf in (".parquet", ".pq"):
            df = pd.read_parquet(p).head(max_rows)
        elif suf in (".xlsx", ".xls"):
            df = pd.read_excel(p, nrows=max_rows)
        elif suf == ".json":
            try:
                df = pd.read_json(p).head(max_rows)
            except Exception:
                df = None
    except Exception as e:
        return (f"Could not parse {p.name} as a table: {e}", _text_peek(p))
    if df is not None:
        try:
            schema_lines = [f"{c}: {t}"
                            for c, t in df.dtypes.astype(str).items()]
            schema = (f"{len(df.columns)} column(s) (dtypes from the first "
                      f"{max_rows} rows):\n\n" + "\n".join(schema_lines))
            rows = df.to_string(index=False, max_rows=max_rows, max_cols=40)
            return (schema, rows)
        except Exception:
            pass
    return ("Not a recognised table format — showing a raw text peek.",
            _text_peek(p))


# Discoverability: example prompts shown by the "What can I ask?" panel.
# Grouped (category, prompt, hint). Hints explain WHY a prompt is useful and
# which capability it exercises — most of these are instant direct routes.
_COUNCIL_EXAMPLES = [
    ("Counting & summaries",
     "how many files are in data_in",
     "Instant exact count + by-type breakdown (no model)."),
    ("Counting & summaries",
     "give me a data summary of the files",
     "One row per file: rows, columns, types (no model)."),
    ("Counting & summaries",
     "summary of stats for the files",
     "Per-column min/max/mean across the files (no model)."),
    ("Find & explore",
     "find files containing Job Blue",
     "Deterministic search by name + indexed content; click to preview."),
    ("Find & explore",
     "look up the value 12345 across files",
     "Cross-file value search — which files mention it, and where."),
    ("Charts",
     "chart sales by month",
     "Finds the right CSV and plots it for you."),
    ("Projects (Collections)",
     "summarise the Job Blue files",
     "Answer about a whole project once you've saved it as a Collection "
     "(Vault tab → Collections)."),
    ("Column analytics",
     "column stats in sales.csv",
     "Per-column min/max/mean/median/std/sum + zero counts, with mean & "
     "median BOTH including and excluding zeros (no model)."),
    ("Column analytics",
     "missing data in sales.csv",
     "Nulls per column + how many rows are fully complete (no model)."),
    ("Column analytics",
     "duplicates in sales.csv",
     "How many exact duplicate rows, with a few examples (no model)."),
    ("Column analytics",
     "top values in sales.csv",
     "Most frequent values in each column — a fast frequency table (no model)."),
    ("Column analytics",
     "correlations in sales.csv",
     "Strongest pairwise correlations between numeric columns (no model)."),
    ("Column analytics",
     "mean of revenue in all csvs in data_in and save to a csv",
     "Compute a stat (mean/sum/min/max/median/std/count) of one column across "
     "every CSV in a folder; add 'save to a csv/text file' to write the result "
     "into the vault's data_out/reports/ (no model)."),
    ("Images",
     "image stats of layer_0345.png",
     "Pixel statistics (brightness, contrast, per-channel, dominant colours) "
     "of one image, or 'image stats in <folder>' for a whole-folder rollup "
     "(no model). Also: 'ocr <image>' to read text inside it."),
    ("Images",
     "count features in layer_0345.png expecting 12",
     "Detect + count discrete features/objects (pores, spatter, spots) in an "
     "image, check against an expected number, and save an annotated image with "
     "each one boxed + numbered (classical CV, no model). Use 'bright'/'dark' "
     "to pick feature polarity."),
    ("Build a tool",
     "create a tool that flags rows where quantity is 0",
     "Opens the Tool Creation tab; the local model writes the tool, the "
     "sandbox validates it (read-only), and it's saved (UNREVIEWED) for reuse."),
    ("Find in files",
     "which files mention bacon",
     "Layered vault search — file summaries first, then a deeper scan of the "
     "actual text; app-state files are excluded."),
    ("When an answer is weak",
     "(click ⤓ Defer to Vault)",
     "Save what the model couldn't do; run it from the Vault tab, then "
     "re-ask — the saved result is reused."),
]


def _build_answer_report_md(question, answer, table, sources):
    """Render a council answer as a Markdown report (question, answer, result
    table, sources). Pure + UI-free so it's unit-testable. ``sources`` may be
    Path objects or strings."""
    lines = ["# Council answer", ""]
    q = (question or "").strip()
    if q:
        lines += ["## Question", "", q, ""]
    lines += ["## Answer", "", (answer or "").strip(), ""]
    t = (table or "").strip()
    if t:
        lines += ["## Result table", "", "```", t, "```", ""]
    srcs = list(sources or [])
    if srcs:
        lines += ["## Sources", ""]
        for s in srcs:
            name = getattr(s, "name", None) or Path(str(s)).name
            lines.append(f"- {name}")
        lines.append("")
    return "\n".join(lines)


# ── Filename wildcard patterns ──────────────────────────────────────────────
# Users reference files by shape, not spelling: "job_####" means "job_ then any
# four characters" (job_1234, job_0087, job_ab12), and "report_*" means "report_
# then anything". The resolvers below used pure substring matching, so `#`/`*`
# were treated as literal characters and never matched. `_compile_name_pattern`
# turns such a token into a safe, anchored, case-insensitive regex:
#   #  → any single character        (the user's "any 4 characters")
#   *  → any run of characters (incl. empty)
#   ?  → any single character
# It returns None when the token has no `#`/`*` wildcard, so callers keep their
# plain-substring behaviour for ordinary names. Pure stdlib, fully offline. The
# generated regex has no nested quantifiers, so there is no catastrophic-
# backtracking risk regardless of user input.
_NAME_WILDCARD_CHARS = ("#", "*")


def _compile_name_pattern(token):
    """Compile a filename-wildcard token to a case-insensitive ``re.Pattern``,
    or return ``None`` when ``token`` contains no ``#``/``*`` wildcard."""
    token = (token or "").strip().strip("'\"`")
    if not token:
        return None
    if not any(c in token for c in _NAME_WILDCARD_CHARS):
        return None
    parts = []
    for ch in token:
        if ch == "#" or ch == "?":
            parts.append(".")          # any single character
        elif ch == "*":
            parts.append(".*")         # any run (incl. empty)
        else:
            parts.append(_re.escape(ch))
    try:
        return _re.compile("".join(parts), _re.IGNORECASE)
    except _re.error:
        return None


def _name_matches_pattern(pat, filename: str) -> bool:
    """True when ``filename`` matches the compiled pattern. Anchored: the
    pattern must span the whole basename OR the whole stem (so ``job_####``
    matches ``job_1234.csv`` via the stem and ``job_####.csv`` via the name)."""
    if pat is None or not filename:
        return False
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    return bool(pat.fullmatch(filename) or pat.fullmatch(stem))


def _search_vault_filenames(in_dir, term, limit: int = 200):
    """Files under ``in_dir`` whose path/name contains every word of ``term``
    (case-insensitive), OR whose basename matches a ``#``/``*`` wildcard
    pattern in ``term`` (e.g. ``job_####``). App-generated output dirs are
    skipped. Returns a list of (abs_path, reason). Pure + UI-free so it's
    unit-testable."""
    import os as _os
    skip = {"derived", "deferred_results", "converted_mongo", "__pycache__",
            ".vault_index", ".stats_cache", "conversation_logs", ".git"}
    # Wildcard mode: match the compiled pattern against each basename. This
    # takes precedence because re.findall(r"[a-z0-9]+", ...) below would
    # silently drop `#`/`*`/`_` and collapse "job_####" to just "job".
    pat = _compile_name_pattern(term)
    out = []
    if pat is not None:
        try:
            for dp, dn, fn in _os.walk(str(in_dir)):
                dn[:] = [d for d in dn if d not in skip and not d.startswith(".")]
                for f in fn:
                    if f.startswith("."):
                        continue
                    if _name_matches_pattern(pat, f):
                        out.append((_os.path.join(dp, f), "pattern match"))
                        if len(out) >= limit:
                            return out
        except Exception:
            pass
        return out
    words = [w for w in _re.findall(r"[a-z0-9]+", (term or "").lower())
             if len(w) > 0]
    if not words:
        return []
    try:
        for dp, dn, fn in _os.walk(str(in_dir)):
            dn[:] = [d for d in dn if d not in skip and not d.startswith(".")]
            for f in fn:
                if f.startswith("."):
                    continue
                full = _os.path.join(dp, f)
                full_lc = full.lower()   # lower once per file, not per word
                if all(w in full_lc for w in words):
                    out.append((full, "name match"))
                    if len(out) >= limit:
                        return out
    except Exception:
        pass
    return out


def _coach_for_error(msg: str):
    """Map a raw error / traceback string to plain-language guidance plus a
    one-click fix. Returns ``dict(plain, action_label, action)`` for a
    recognised failure, or ``None`` for an unrecognised one (the caller then
    shows the raw error). Pure + UI-free so it's unit-testable.

    Actions the GUI knows how to run: ``"engine"`` (open Engine settings),
    ``"models"`` (open the Models tab), ``"cpu"`` (force CPU + retry).
    """
    m = (msg or "").lower()
    # 1) Context-window overflow — the most common confusing failure on a
    #    small model / 4K-ctx box ("exceeds max tokens").
    if (("exceed" in m and ("context" in m or "token" in m))
            or "requested tokens" in m
            or "exceeds max tokens" in m
            or "context window" in m):
        return {
            "plain": ("The question plus its data was larger than the model's "
                      "context window. Raise the max context in Engine "
                      "settings, or ask about fewer files at once."),
            "action_label": "⚙ Open Engine settings",
            "action": "engine",
        }
    # 2) GPU / CUDA failure — VRAM exhaustion, a driver fault, or a prior
    #    core-dump sentinel. Degrade to CPU (slower but reliable).
    if ("cuda" in m or "cublas" in m or "ggml_cuda" in m
            or "device-side assert" in m
            or ("gpu" in m and "memory" in m)):
        return {
            "plain": ("The GPU run failed — usually not enough VRAM. Switch "
                      "the model to CPU (slower but reliable) and retry."),
            "action_label": "Switch to CPU and retry",
            "action": "cpu",
        }
    # 3) Model not loaded / missing file.
    if (("llama" in m and ("failed" in m or "could not" in m or "load" in m))
            or "no such file" in m
            or "model not found" in m
            or "council_gguf_path" in m
            or ("gguf" in m and "not found" in m)):
        return {
            "plain": ("The model could not be loaded. Choose a model file, or "
                      "download one that fits your hardware."),
            "action_label": "🇺🇸 Open Models",
            "action": "models",
        }
    # 4) Generic out-of-memory (CPU RAM).
    if ("memoryerror" in m or "out of memory" in m
            or "cannot allocate" in m):
        return {
            "plain": ("Ran out of memory. Try asking about fewer or smaller "
                      "files, or switch to a smaller model."),
            "action_label": "🇺🇸 Open Models",
            "action": "models",
        }
    return None


def _smart_truncate_block_to_tokens(block_text: str, max_tokens: int) -> str:
    """Truncate a rendered injection block to fit a per-block token cap.

    Uses `estimate_tokens` and falls back to char-budget trimming. Always
    leaves a visible marker so the model knows content was elided.
    """
    cur = _estimate_block_tokens(block_text)
    if cur <= max_tokens:
        return block_text
    # ~4 chars/token works for English; over-estimate a touch by using 4.
    target_chars = max(800, max_tokens * 4)
    trimmed, _ = _smart_truncate_text(block_text, target_chars)
    return trimmed


def _sample_rows_head_mid_tail(rows: list, max_rows: int) -> tuple:
    """Sample a long row list into head + middle + tail slices.

    Used by CSV / TSV / Parquet / Excel renderers so the model sees the
    SHAPE of all three regions (a CSV where column dtypes change after
    row 5,000 is now visible) instead of just the first N rows.

    Returns (head_rows, mid_rows, tail_rows, skipped_before_mid,
    skipped_between_mid_and_tail). Each slice is roughly max_rows/3 long.
    """
    n = len(rows)
    if n <= max_rows:
        return rows, [], [], 0, 0
    slice_size = max(3, max_rows // 3)
    head = rows[:slice_size]
    tail = rows[-slice_size:]
    mid_center = n // 2
    mid_start = max(slice_size, mid_center - slice_size // 2)
    mid_end = min(n - slice_size, mid_start + slice_size)
    mid = rows[mid_start:mid_end]
    skipped_before_mid = mid_start - len(head)
    skipped_between = (n - len(tail)) - mid_end
    return head, mid, tail, skipped_before_mid, skipped_between


def _render_dataframe_block(kind_label, df, max_rows=20):
    """Render a small DataFrame as a CSV-shaped text block. Used by the
    injection layer for TSV / Parquet / gzipped CSV so the model sees
    the same structure regardless of source format.

    When the frame has more than `max_rows` rows, samples HEAD + MIDDLE +
    TAIL (with explicit skip markers) so the model sees the SHAPE of the
    full file — not just the top rows. This matters for files where
    column dtypes change deep in the data, or where the last rows hold
    totals/summaries.
    """
    headers = [str(c) for c in df.columns]
    total = len(df)
    rows = [r.tolist() for _, r in df.iterrows()]

    def _fmt(r):
        cells = []
        for v in r:
            cv = str(v).replace('\n', ' ').strip()
            if len(cv) > 80:
                cv = cv[:77] + '...'
            cells.append(cv)
        return '  ' + ' | '.join(cells)

    lines = [
        f"({kind_label} — already parsed into plain text below. "
        f"Read each row as if it were a CSV row.)",
        f"Columns ({len(headers)}): " + ', '.join(headers),
    ]
    if total <= max_rows:
        # Note: `total` is the count this renderer received — callers often
        # pass a .head(200)'d slice of a larger source, so we don't claim
        # this is the "full file." Just report what we can see.
        lines.append(f"Total rows shown: {total} (of {total} loaded)")
        lines.append("Sample rows:")
        for r in rows:
            lines.append(_fmt(r))
    else:
        head, mid, tail, sk_b, sk_a = _sample_rows_head_mid_tail(rows, max_rows)
        shown = len(head) + len(mid) + len(tail)
        lines.append(
            f"Total rows shown: {shown} of {total} loaded "
            f"(HEAD + MIDDLE + TAIL sampled — middle rows elided)"
        )
        lines.append("Sample rows (head):")
        for r in head:
            lines.append(_fmt(r))
        if mid:
            lines.append(f"  ... [skipped {sk_b:,} rows] ...")
            lines.append("Sample rows (middle):")
            for r in mid:
                lines.append(_fmt(r))
        if tail:
            lines.append(f"  ... [skipped {sk_a:,} rows] ...")
            lines.append("Sample rows (tail):")
            for r in tail:
                lines.append(_fmt(r))
    return '\n'.join(lines)


def _render_sqlite_block(p):
    """Render a SQLite database as table name + columns + row count + a
    small sample from each table. Read-only connection."""
    import sqlite3 as _sq
    lines = [
        "(SQLite database — already parsed into plain text below. "
        "Each table is shown with its columns and a few sample rows.)",
    ]
    try:
        con = _sq.connect(f"file:{p}?mode=ro", uri=True)
        try:
            cur = con.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "ORDER BY name LIMIT 20"
            )
            names = [r[0] for r in cur.fetchall() if r and r[0]]
            lines.append(f"Tables ({len(names)}): " + ', '.join(names))
            lines.append("")
            for tname in names[:10]:
                qname = '"' + tname.replace('"', '""') + '"'
                try:
                    cur = con.execute(f"PRAGMA table_info({qname})")
                    cols = [r[1] for r in cur.fetchall() if r and r[1]]
                except Exception:
                    cols = []
                try:
                    cur = con.execute(f"SELECT COUNT(*) FROM {qname}")
                    nrows = cur.fetchone()[0]
                except Exception:
                    nrows = None
                lines.append(f'Table "{tname}" ({nrows if nrows is not None else "?"} rows):')
                lines.append("  columns: " + ', '.join(cols))
                try:
                    cur = con.execute(f"SELECT * FROM {qname} LIMIT 5")
                    sample = cur.fetchall()
                    if sample:
                        lines.append("  rows:")
                        for r in sample:
                            cells = []
                            for v in r:
                                cv = str(v).replace('\n', ' ').strip()
                                if len(cv) > 80:
                                    cv = cv[:77] + '...'
                                cells.append(cv)
                            lines.append('    ' + ' | '.join(cells))
                except Exception:
                    pass
                lines.append("")
        finally:
            con.close()
    except Exception as exc:
        lines.append(f"(could not open database: {exc!r})")
    return '\n'.join(lines).rstrip()


def _detect_excel_header_rows(p, sheet_name, max_check=3):
    """Look at the first `max_check` rows of `sheet_name` via openpyxl
    and detect whether the workbook uses MERGED CELLS in the top rows
    (the common "group headers above sub-headers" pattern). Returns
    the integer number of header rows to feed pandas' header= parameter.

    Returns 1 (single-row header) when no merging is detected. Returns
    2 when row 1 has any merged cells spanning multiple columns
    (interpreted as a group-header row above the actual column names).
    Returns 3 if rows 1 AND 2 both have merging.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=False, data_only=True)
    except Exception:
        return 1
    try:
        if isinstance(sheet_name, int) and 0 <= sheet_name < len(wb.worksheets):
            ws = wb.worksheets[sheet_name]
        elif isinstance(sheet_name, str) and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
        merged_rows = set()
        for rng in ws.merged_cells.ranges:
            if rng.max_col - rng.min_col >= 1:  # spans multiple columns
                for r in range(rng.min_row, min(rng.max_row, max_check) + 1):
                    merged_rows.add(r)
        if not merged_rows:
            return 1
        # If row 1 is merged, we have at least 2 header rows (the merged
        # group + the column-name row below). Extend if row 2 also has
        # merges that span columns.
        max_merged = max(merged_rows)
        return max(2, min(max_merged + 1, max_check))
    finally:
        try: wb.close()
        except Exception: pass


def _flatten_multi_header(df, sep=" / "):
    """Collapse a MultiIndex column header into single strings.

    pandas read_excel(header=[0,1]) yields tuples like ("Site A", "energy");
    we render them as "Site A / energy" for the prompt and as a single
    column name for the analyst. Empty group labels (Unnamed: 0_level_0)
    get cleaned up.
    """
    import pandas as _pd_fi
    if not isinstance(df.columns, _pd_fi.MultiIndex):
        return df
    new_cols = []
    for tup in df.columns:
        parts = [str(x) for x in tup
                 if x is not None and not str(x).startswith("Unnamed:")]
        new_cols.append(sep.join(parts) if parts else "(unnamed)")
    df = df.copy()
    df.columns = new_cols
    return df


def _render_excel_block(p, char_limit):
    """Render an .xlsx/.xls workbook with awareness of merged-cell
    headers — the common 'top row groups, second row sub-columns'
    layout. Without this, pandas mangles the columns to 'Unnamed: N'
    and the model can't tell what each value represents.
    """
    import pandas as _pd_fi
    xl = _pd_fi.ExcelFile(str(p))
    lines = [
        "(Excel workbook — already parsed into plain text below. "
        "Read each sheet as if it were a CSV table. Merged-cell group "
        "headers are joined with ' / ' so each column has a unique name.)",
        "Total sheets: " + str(len(xl.sheet_names)),
        "",
    ]
    sample_per_sheet = 10
    for sname in xl.sheet_names[:8]:
        header_rows = _detect_excel_header_rows(p, sname)
        try:
            if header_rows >= 2:
                df = xl.parse(sname, header=list(range(header_rows)),
                              nrows=sample_per_sheet * 2)
                df = _flatten_multi_header(df)
                merged_note = (f" (DETECTED {header_rows} header rows — "
                               f"top rows have merged cells, joined with ' / ')")
            else:
                df = xl.parse(sname, nrows=sample_per_sheet * 2)
                merged_note = ""
        except Exception as _ex:
            lines.append(f'Sheet "{sname}": read error: {_ex}')
            lines.append("")
            continue
        headers = [str(c) for c in df.columns]
        lines.append(f'Sheet "{sname}"{merged_note} '
                     f'({len(df)} rows shown, {len(headers)} columns):')
        lines.append('  columns: ' + ', '.join(headers))
        lines.append('  rows:')
        for _, row in df.head(sample_per_sheet).iterrows():
            cells = []
            for v in row.tolist():
                cv = str(v).replace('\n', ' ').strip()
                if len(cv) > 80:
                    cv = cv[:77] + '...'
                cells.append(cv)
            lines.append('    ' + ' | '.join(cells))
        lines.append("")
    content = '\n'.join(lines).rstrip()
    # Visible head+tail trim so the model sees the start AND end of the
    # workbook listing rather than silently losing the last sheets.
    if len(content) > char_limit:
        content, _ = _smart_truncate_text(content, char_limit)
    return content


def _short_count(n: int) -> str:
    """Compact integer rendering used in the folder summary's per-file
    schema preview. Keeps lines short enough that 47 files fit in
    a token budget."""
    if n is None:
        return "?"
    try:
        n = int(n)
    except (TypeError, ValueError):
        return "?"
    if abs(n) >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if abs(n) >= 1_000:
        return f"{n / 1_000:.1f}K"
    return str(n)


def _schema_preview_for_file(rel_path: Path, full_path: Path,
                              vault_index, include_columns: bool) -> str:
    """One-line schema preview for a single file. Looks up the file's
    record in the vault index and renders the columns / row count /
    sheet names / JSON keys. Falls back to a size-only line when the
    file isn't in the index (e.g. a .gguf, a binary blob).

    ``include_columns`` controls whether to spell out the first ~8
    column names. We turn this OFF for files past the per-folder
    detail cap so a 200-file folder still fits in a token budget.

    The fix for "model can see files exist but can't read data inside":
    the old compact renderer showed JUST the filename, so the model had
    no way to know what columns or how many rows each CSV had. Now
    every line carries the schema — model can answer questions like
    "what columns does orders.csv have?" purely from the folder block.
    """
    rec = None
    if vault_index is not None:
        try:
            rec = vault_index.records.get(str(full_path))
            # Fallback — index keys often use resolved paths
            if rec is None:
                rec = vault_index.records.get(str(full_path.resolve()))
        except Exception:
            rec = None

    # No index record → just filename + size
    if not isinstance(rec, dict):
        try:
            sz = full_path.stat().st_size
            sz_str = (f"{sz/1024/1024:.1f} MB"
                      if sz >= 1024*1024
                      else (f"{sz/1024:.0f} KB" if sz >= 1024
                            else f"{sz} B"))
            return f"  {rel_path}  ·  {sz_str}"
        except Exception:
            return f"  {rel_path}"

    rtype = rec.get("type") or "?"

    if rtype in ("csv", "tsv", "csv.gz", "parquet"):
        rows = _short_count(rec.get("rows"))
        cols_full = rec.get("headers") or []
        ncols = len(cols_full)
        suffix = f"  ·  {rtype} · {rows} rows × {ncols} cols"
        if include_columns and cols_full:
            shown = cols_full[:8]
            extra = (f", +{ncols - 8}" if ncols > 8 else "")
            suffix += f" [{', '.join(map(str, shown))}{extra}]"
        return f"  {rel_path}{suffix}"

    if rtype == "excel":
        sheets = rec.get("sheets") or []
        suffix = f"  ·  excel · {len(sheets)} sheet{'s' if len(sheets) != 1 else ''}"
        if include_columns and sheets:
            sheet_names = [str(s.get("sheet", "?")) for s in sheets[:5]]
            extra = (f", +{len(sheets) - 5}" if len(sheets) > 5 else "")
            suffix += f" [{', '.join(sheet_names)}{extra}]"
        return f"  {rel_path}{suffix}"

    if rtype in ("json", "d3dpipeline", "bson"):
        keys = rec.get("keys") or []
        nkeys = len(keys)
        suffix = f"  ·  {rtype} · {nkeys} key{'s' if nkeys != 1 else ''}"
        if include_columns and keys:
            shown = keys[:8]
            extra = (f", +{nkeys - 8}" if nkeys > 8 else "")
            suffix += f" [{', '.join(map(str, shown))}{extra}]"
        return f"  {rel_path}{suffix}"

    if rtype in ("sqlite", "duckdb"):
        tables = rec.get("tables") or []
        suffix = f"  ·  {rtype} · {len(tables)} table{'s' if len(tables) != 1 else ''}"
        if include_columns and tables:
            table_names = [str(t.get("table", "?")) for t in tables[:5]]
            extra = (f", +{len(tables) - 5}" if len(tables) > 5 else "")
            suffix += f" [{', '.join(table_names)}{extra}]"
        return f"  {rel_path}{suffix}"

    if rtype == "image":
        w = rec.get("width")
        h = rec.get("height")
        if w and h:
            return f"  {rel_path}  ·  image · {w}×{h}"
        return f"  {rel_path}  ·  image"

    # Plain text / source code / etc.
    if rec.get("rows"):
        return f"  {rel_path}  ·  {rtype} · {_short_count(rec.get('rows'))} lines"
    return f"  {rel_path}  ·  {rtype}"


def _render_folder_summary_compact(folder, max_files=120, max_chars=4500,
                                    *, vault_index=None,
                                    detailed_top_n: int = 40):
    """Compact folder summary — counts + subfolder breakdown + filename
    list WITH per-file schema previews pulled from the vault index.

    Previously this rendered filenames ONLY. The model saw "orders.csv
    exists" but had no way to see the columns or row count, which led
    to the "model can see files exist but can't read data inside"
    reports — perfectly true from the model's perspective, since the
    block we injected literally only contained filenames.

    Now each line carries a one-line schema preview:
        orders.csv     ·  csv · 1.2K rows × 8 cols [order_id, customer, total, ...]
        customers.csv  ·  csv · 503 rows × 5 cols  [id, name, email, joined, segment]
        config.json    ·  json · 12 keys [settings, sources, outputs, ...]
        report.xlsx    ·  excel · 3 sheets [Q1, Q2, Q3]

    The top ``detailed_top_n`` files get the full schema with column
    names; the remainder get a shorter `name · type · rows × cols`
    line so a 500-file folder still fits in a few hundred lines.

    Used by the vault-trigger auto-injection so a small-context model
    (e.g. 4 K window) doesn't get its entire prompt budget eaten by one
    rich folder block. About 5-10× smaller than _render_folder_for_injection
    on the same vault.

    Returns ``None`` when the folder is missing — caller treats as a
    no-op and skips the block.
    """
    folder = Path(folder)
    if not folder.exists() or not folder.is_dir():
        return None

    # Lazy-fetch the vault index if the caller didn't pass one. The
    # injector usually has it in hand already; standalone callers
    # (the wizard / debug paths) get it via the existing helper.
    if vault_index is None:
        try:
            vault_index = _get_vault_index()
        except Exception:
            vault_index = None
        if vault_index is not None:
            # Make sure the index actually reflects what's on disk —
            # delta-aware so this is cheap if nothing changed.
            try:
                vault_index.rebuild()
            except Exception:
                pass

    try:
        import conversation_logger as _cl
    except Exception:
        _cl = None

    SKIP_DIRS = {"__pycache__", ".git", ".venv", "venv", "node_modules"}
    # Higher than the rich renderer's 5K — we don't parse files, so
    # walking 10K dir entries is still cheap.
    SCAN_LIMIT = 10000

    from collections import defaultdict as _dd
    files: list = []          # list of (rel_path, full_path)
    subfolders: dict = _dd(int)
    by_suffix: dict = _dd(int)
    scanned = 0
    walk_truncated = False
    for p in folder.rglob("*"):
        scanned += 1
        if scanned > SCAN_LIMIT:
            walk_truncated = True
            break
        if not p.is_file():
            continue
        try:
            rel = p.relative_to(folder)
        except ValueError:
            continue
        if any(part in SKIP_DIRS or part.startswith(".")
               for part in rel.parts[:-1]):
            continue
        if _cl is not None:
            try:
                if _cl.is_protected_path(p, folder.parent):
                    continue
            except Exception:
                pass
        if len(rel.parts) > 1:
            subfolders[rel.parts[0]] += 1
        suf = p.suffix.lower() or "(no ext)"
        by_suffix[suf] += 1
        files.append((rel, p))

    total = len(files)
    total_label = f"{total}" + ("+" if walk_truncated else "")
    lines = [f"[FOLDER SUMMARY: {folder.name or folder}]"]
    lines.append(f"Total files: {total_label}"
                 + (f" (walked first {SCAN_LIMIT:,} entries — more may exist)"
                    if walk_truncated else ""))

    if subfolders:
        lines.append("")
        lines.append("Subfolder file counts:")
        for sub, count in sorted(subfolders.items(),
                                  key=lambda x: (-x[1], x[0]))[:30]:
            lines.append(f"  {sub}/: {count} file{'s' if count != 1 else ''}")
        if len(subfolders) > 30:
            lines.append(f"  ... ({len(subfolders) - 30} more subfolders)")

    if by_suffix:
        type_bits = [f"{s}: {c}" for s, c
                     in sorted(by_suffix.items(), key=lambda x: -x[1])[:12]]
        lines.append("")
        lines.append("Files by type: " + ", ".join(type_bits))

    if files:
        lines.append("")
        shown = files[:max_files]
        lines.append(
            f"Files (first {len(shown)} of {total}) — schema preview per "
            f"file so you can answer 'what's in X' without a follow-up "
            f"tool call:")
        for idx, (rel_path, full_path) in enumerate(shown):
            include_cols = (idx < detailed_top_n)
            lines.append(_schema_preview_for_file(
                rel_path, full_path, vault_index,
                include_columns=include_cols,
            ))
        if total > max_files:
            lines.append(f"  ... ({total - max_files} more files not shown — "
                         f"ask about specific files or subfolders.)")

    text = "\n".join(lines)
    if len(text) > max_chars:
        text, _ = _smart_truncate_text(text, max_chars)
    return text


def _render_folder_for_injection(folder, max_files=40, max_chars=12000):
    """Build a comprehensive [FOLDER: ...] injection block listing every
    file in the folder. For tabular files, include headers + first 3
    rows so the model sees the SHAPE of every file, not just top-5
    vault-search hits.

    Honors PROTECTED_SUBDIRS: files under conversation_logs/ /
    conversations/ are never included.
    """
    folder = Path(folder)
    if not folder.exists() or not folder.is_dir():
        return None
    try:
        import conversation_logger as _cl
    except Exception:
        _cl = None

    lines = [f"[FOLDER: {folder}]"]
    lines.append("(All files in this folder are listed below — already "
                 "scanned. Treat this as the COMPLETE inventory of what's "
                 "available; do not invent files that aren't shown here.)")
    lines.append("")

    # Collect files (recursive, skip hidden/build/protected).
    # Bound the walk: previously `for p in folder.rglob("*")` traversed
    # the entire tree even on folders with 50k+ entries (the user only
    # ever sees 40). On corporate file shares this is multi-second
    # latency for nothing. SCAN_LIMIT caps the walk; if hit, we tell
    # the model the listing is partial so it doesn't claim to have
    # seen everything.
    SKIP_DIRS = {"__pycache__", ".git", ".venv", "venv", "node_modules"}
    SCAN_LIMIT = 5000   # entries walked, not files kept
    files = []
    scanned = 0
    walk_truncated = False
    for p in folder.rglob("*"):
        scanned += 1
        if scanned > SCAN_LIMIT:
            walk_truncated = True
            break
        if not p.is_file():
            continue
        if any(part in SKIP_DIRS or part.startswith(".")
               for part in p.relative_to(folder).parts[:-1]):
            continue
        if _cl is not None:
            try:
                if _cl.is_protected_path(p, folder.parent):
                    continue
            except Exception:
                pass
        try:
            size = p.stat().st_size
        except Exception:
            continue
        files.append((p, size))
    files.sort(key=lambda fs: (fs[0].suffix.lower(), fs[0].name.lower()))

    if not files:
        lines.append("(folder is empty)")
        return "\n".join(lines)

    # Truncate if too many
    truncated_n = max(0, len(files) - max_files)
    files_shown = files[:max_files]
    total_label = f"{len(files)}" + ("+ " if walk_truncated else "")
    suffix = ""
    if walk_truncated:
        suffix = (f" — walked first {SCAN_LIMIT:,} entries; the folder "
                  f"contains more files not scanned")
    elif truncated_n:
        suffix = f" (showing first {max_files})"
    lines.append(f"Total files: {total_label}{suffix}")
    lines.append("")

    # Per-file summary — short for tabular, name-only for binary
    for p, size in files_shown:
        rel = p.relative_to(folder)
        suf = p.suffix.lower()
        size_kb = size / 1024
        size_str = f"{size_kb:>8.1f} KB" if size_kb < 1024 else f"{size_kb/1024:>7.1f} MB"
        head = f"  {size_str}  {rel}"
        try:
            if suf == ".csv":
                import pandas as _pd
                df = _pd.read_csv(p, nrows=3)
                cols = ", ".join(str(c) for c in df.columns[:12])
                more = "" if len(df.columns) <= 12 else f" + {len(df.columns)-12} more"
                lines.append(head + "  columns: " + cols + more)
            elif suf in (".xlsx", ".xls", ".xlsm"):
                import pandas as _pd
                xl = _pd.ExcelFile(p)
                lines.append(head + f"  sheets: {', '.join(xl.sheet_names[:5])}"
                             + (f" (+{len(xl.sheet_names)-5})" if len(xl.sheet_names) > 5 else ""))
            elif suf == ".json":
                import json as _json
                try:
                    obj = _json.loads(p.read_text(encoding="utf-8", errors="replace")[:2048] + "}")
                except Exception:
                    obj = None
                if isinstance(obj, dict):
                    keys = ", ".join(str(k) for k in list(obj.keys())[:8])
                    lines.append(head + f"  keys: {keys}")
                else:
                    lines.append(head)
            elif suf in (".txt", ".md", ".rst", ".log", ".yaml", ".yml", ".toml", ".ini"):
                snippet = p.read_text(encoding="utf-8", errors="replace")[:80].replace("\n", " ")
                lines.append(head + f"  preview: {snippet}...")
            else:
                lines.append(head)
        except Exception:
            lines.append(head)

        if sum(len(ln) for ln in lines) > max_chars:
            lines.append("  ... (folder listing truncated)")
            break

    if truncated_n:
        lines.append(f"  ... ({truncated_n} additional files not shown — "
                     f"ask about specific files or subfolders to see them)")
    lines.append("[END FOLDER]")
    return "\n".join(lines)


def _render_duckdb_block(p):
    """DuckDB database — read-only listing of tables + columns + samples."""
    lines = [
        "(DuckDB database — already parsed into plain text below. "
        "Each table is shown with its columns and a few sample rows.)",
    ]
    try:
        import duckdb as _duckdb
    except Exception as exc:
        return ("(DuckDB file — install duckdb with `pip install duckdb` "
                f"to view: {exc!r})")
    try:
        con = _duckdb.connect(str(p), read_only=True)
        try:
            tnames = [r[0] for r in con.execute(
                "SELECT table_name FROM duckdb_tables() "
                "ORDER BY table_name LIMIT 20"
            ).fetchall() if r and r[0]]
            lines.append(f"Tables ({len(tnames)}): " + ', '.join(tnames))
            lines.append("")
            for tname in tnames[:10]:
                qname = '"' + tname.replace('"', '""') + '"'
                try:
                    rows = con.execute(f"DESCRIBE {qname}").fetchall()
                    cols = [r[1] for r in rows if r and r[1]]
                except Exception:
                    cols = []
                try:
                    nrows = con.execute(f"SELECT COUNT(*) FROM {qname}").fetchone()[0]
                except Exception:
                    nrows = None
                lines.append(f'Table "{tname}" ({nrows if nrows is not None else "?"} rows):')
                lines.append("  columns: " + ', '.join(cols))
                try:
                    sample = con.execute(
                        f"SELECT * FROM {qname} LIMIT 5"
                    ).fetchall()
                    if sample:
                        lines.append("  rows:")
                        for r in sample:
                            cells = []
                            for v in r:
                                cv = str(v).replace('\n', ' ').strip()
                                if len(cv) > 80:
                                    cv = cv[:77] + '...'
                                cells.append(cv)
                            lines.append('    ' + ' | '.join(cells))
                except Exception:
                    pass
                lines.append("")
        finally:
            con.close()
    except Exception as exc:
        lines.append(f"(could not open database: {exc!r})")
    return '\n'.join(lines).rstrip()


def _render_bson_block(p):
    """MongoDB BSON dump — list field names and a few sample documents."""
    lines = [
        "(MongoDB BSON file — already parsed into plain text below. "
        "Each line below is one document.)",
    ]
    try:
        import bson as _bson
    except Exception as exc:
        return ("(BSON file — install pymongo with `pip install pymongo` "
                f"to view: {exc!r})")
    try:
        with open(p, "rb") as fh:
            data = fh.read()
        docs = _bson.decode_all(data)
    except Exception as exc:
        return f"(BSON read failed: {exc!r})"
    lines.append(f"Total documents: {len(docs)}")
    # Field-name summary
    keys = set()
    for d in docs[:200]:
        if isinstance(d, dict):
            keys.update(d.keys())
    lines.append("Fields: " + ", ".join(sorted(map(str, keys))[:30]))
    lines.append("")
    lines.append("Sample documents:")
    import json as _j
    for i, d in enumerate(docs[:10]):
        try:
            line = _j.dumps(d, default=str, ensure_ascii=False)
        except Exception:
            line = repr(d)
        if len(line) > 300:
            line = line[:297] + "..."
        lines.append(f"  {i+1}. {line}")
    return '\n'.join(lines)


# Small memo so a file read once during injection isn't re-opened and
# re-parsed a second time in the same turn (the provenance pass reads the same
# explicit paths). Keyed on (resolved path, mtime_ns, size) — a single stat()
# is far cheaper than re-parsing a 50k-row CSV, and the key invalidates when
# the file changes across turns. None results (missing/protected/dir) are NOT
# cached, so protection is always re-checked.
_FILE_INJECT_CACHE: dict = {}
_FILE_INJECT_CACHE_MAX = 64


def _read_file_for_injection(path_str):
    """Memoizing wrapper over ``_read_file_for_injection_uncached`` (see it for
    the real read + the protected-path guard). Returns byte-identical blocks;
    falls through to the uncached read whenever the path can't be stat'd."""
    try:
        p = Path(str(path_str).strip())
        st = p.stat()
        key = (str(p.resolve()).lower(), st.st_mtime_ns, st.st_size)
    except Exception:
        return _read_file_for_injection_uncached(path_str)
    hit = _FILE_INJECT_CACHE.get(key)
    if hit is not None:
        return hit
    block = _read_file_for_injection_uncached(path_str)
    if block is not None:
        if len(_FILE_INJECT_CACHE) >= _FILE_INJECT_CACHE_MAX:
            _FILE_INJECT_CACHE.clear()
        _FILE_INJECT_CACHE[key] = block
    return block


def _read_file_for_injection_uncached(path_str):
    """Read a file into a compact prompt block. CSVs get headers + sample
    rows; large rows are abbreviated so the header line always survives the
    model's context window.

    HARD GUARD: refuses to read anything under a protected vault subdir
    (conversation_logs, etc.). The model must never see those files.
    """
    try:
        import csv as _csv_fi
        p = Path(path_str.strip())
        if not p.exists() or not p.is_file():
            return None
        # Defense-in-depth: even if the user pastes a path inside
        # conversation_logs/, we refuse to inject it.
        try:
            import conversation_logger as _cl
            if _cl.is_protected_path(p, VAULT_DIR):
                return None
        except Exception:
            pass
        suffix = p.suffix.lower()
        if suffix == '.csv':
            # Bound the read: previously this did `list(reader)` of the
            # entire file, which on a 500MB / 5M-row CSV materialises
            # hundreds of MB of Python tuples and can OOM the GUI. The
            # injection only needs ~24 rows (head + middle + tail), so
            # we cap reading at 50k rows. For larger files the model
            # sees the shape of the first 50k rows and is told the
            # actual file is bigger — the analyst is the right tool
            # for anything that needs full-file counts.
            import itertools as _it
            CSV_ROW_HARD_CAP = 50_000
            rows: list = []
            csv_truncated = False
            try:
                with open(p, newline='', encoding='utf-8', errors='replace') as fh:
                    reader = _csv_fi.reader(fh)
                    # Read up to cap + 1 so we can detect overflow
                    for i, row in enumerate(_it.islice(reader, CSV_ROW_HARD_CAP + 1)):
                        if i >= CSV_ROW_HARD_CAP:
                            csv_truncated = True
                            break
                        rows.append(row)
            except Exception:
                # Fall through with whatever we managed to read
                pass
            if not rows:
                return None
            header = rows[0]
            data_rows = rows[1:]
            total_rows = len(data_rows)
            if csv_truncated:
                # Surface the cap to the model — without this it would
                # report "total rows: 49,999" as if that were the file's
                # actual size, which is a subtle hallucination.
                total_rows_str = f"{total_rows:,}+ (file exceeds the {CSV_ROW_HARD_CAP:,}-row injection cap; ask the analyst for exact full-file counts)"
            else:
                total_rows_str = f"{total_rows:,}"

            lines = ['Columns (' + str(len(header)) + '): ' + ', '.join(header)]
            lines.append('Total data rows: ' + total_rows_str)

            def _fmt(r):
                cells = []
                for c in r:
                    c = str(c).replace('\n', ' ').strip()
                    if len(c) > 80:
                        c = c[:77] + '...'
                    cells.append(c)
                return ' | '.join(cells)

            sample_target = 24
            if total_rows <= sample_target:
                if total_rows:
                    lines.append(f'Sample rows ({total_rows}, full file):')
                    for r in data_rows:
                        lines.append(_fmt(r))
            else:
                head, mid, tail, sk_b, sk_a = _sample_rows_head_mid_tail(
                    data_rows, sample_target,
                )
                shown = len(head) + len(mid) + len(tail)
                lines.append(
                    f'Sample rows ({shown} of {total_rows} — HEAD + MIDDLE + '
                    f'TAIL sampled; middle rows elided):'
                )
                lines.append('--- head ---')
                for r in head:
                    lines.append(_fmt(r))
                if mid:
                    lines.append(f'... [skipped {sk_b:,} rows] ...')
                    lines.append('--- middle ---')
                    for r in mid:
                        lines.append(_fmt(r))
                if tail:
                    lines.append(f'... [skipped {sk_a:,} rows] ...')
                    lines.append('--- tail ---')
                    for r in tail:
                        lines.append(_fmt(r))
            content = '\n'.join(lines)
            if len(content) > _FILE_READ_CHAR_LIMIT:
                content, _ = _smart_truncate_text(content, _FILE_READ_CHAR_LIMIT)
        elif suffix in ('.tsv',):
            # TSV — same rendering as CSV, just tab-separated
            import pandas as _pd_fi
            df = _pd_fi.read_csv(p, sep='\t', nrows=200, on_bad_lines='skip')
            content = _render_dataframe_block('TSV', df)
        elif suffix == '.parquet':
            try:
                import pandas as _pd_fi
                df = _pd_fi.read_parquet(p).head(200)
                content = _render_dataframe_block('Parquet', df)
            except Exception as _ex:
                content = (f"(Parquet file — could not read: {_ex!r}. "
                           f"Install pyarrow with `pip install pyarrow`.)")
        elif suffix == '.gz' and p.stem.lower().endswith('.csv'):
            import pandas as _pd_fi
            df = _pd_fi.read_csv(p, nrows=200, on_bad_lines='skip', compression='infer')
            content = _render_dataframe_block('Gzipped CSV', df)
        elif suffix in ('.db', '.sqlite', '.sqlite3'):
            content = _render_sqlite_block(p)
        elif suffix == '.duckdb':
            content = _render_duckdb_block(p)
        elif suffix == '.bson':
            content = _render_bson_block(p)
        elif suffix in ('.xlsx', '.xls', '.xlsm'):
            content = _render_excel_block(p, _FILE_READ_CHAR_LIMIT)
        else:
            # Read up to 2× the budget so we can show head + tail with a
            # visible elision marker (silent tail-clip caused the model
            # to invent the "missing" portion).
            with open(p, encoding='utf-8', errors='replace') as fh:
                raw = fh.read(_FILE_READ_CHAR_LIMIT * 2)
            if len(raw) > _FILE_READ_CHAR_LIMIT:
                content, _ = _smart_truncate_text(raw, _FILE_READ_CHAR_LIMIT)
            else:
                content = raw
        if not content.strip():
            return None
        return '[FILE: ' + p.name + ']\n' + content + '\n[END FILE]'
    except Exception:
        return None


def _extract_file_paths(text):
    """Find file paths in the user message. Protected paths (conversation
    logs) are filtered before they reach injection — see PROTECTED_SUBDIRS."""
    import unicodedata as _ud
    clean_chars = []
    for ch in text:
        if ch in ('\\', '/', '.', '~', ':'):
            clean_chars.append(ch)
        elif _ud.category(ch) in ('Pi', 'Pf', 'Po', 'Ps', 'Pe') or ch in ('<', '>'):
            clean_chars.append(' ')
        else:
            clean_chars.append(ch)
    clean = ''.join(clean_chars)
    paths = []
    seen = set()
    for m in _FILE_PATH_RE.finditer(clean):
        candidate = m.group(0).strip()
        while candidate and not (candidate[-1].isalnum() or candidate[-1] in '\\/'):
            candidate = candidate[:-1]
        if candidate and candidate not in seen:
            seen.add(candidate)
            paths.append(candidate)
    # HARD GUARD — drop any path under a protected vault subdir before
    # it can reach the injection layer. The model must never see files
    # under conversation_logs/.
    try:
        import conversation_logger as _cl
        paths = [p for p in paths if not _cl.is_protected_path(p, VAULT_DIR)]
    except Exception:
        pass
    return paths


# Phrase substrings that mark a CONCEPTUAL search query — "show me
# files that …", "find scripts about …", etc. When matched, the
# injector defaults to search-headers mode (compact one-line entries
# per match) rather than full [VAULT MATCH] blocks. The model
# typically wants to acknowledge what's there before zooming into
# specific files, and headers let it see far more matches per token.
_SEARCH_HEADERS_PHRASES = (
    "find files", "find any file", "find every", "find a file",
    "find scripts", "find pipelines", "find docs",
    "look through", "look for files",
    "search through", "search for files", "search for any",
    "show me files", "show me the files", "list files", "list of files",
    "which files", "what files", "any files",
    "files that contain", "files with", "files about",
    "scripts that", "pipelines that",
)


def _wants_search_headers(text: str) -> bool:
    """Heuristic: does this query look like 'show me what's there' rather
    than 'tell me about X specifically'? Conceptual-search queries get
    compact one-line headers instead of full blocks so the model can see
    more matches per token. Manual override via
    COUNCIL_FORCE_FULL_VAULT_BLOCKS=1 (or =yes/true/on) skips this and
    keeps the legacy full-block behaviour."""
    if os.environ.get("COUNCIL_FORCE_FULL_VAULT_BLOCKS", "").strip().lower() in (
        "1", "true", "yes", "on"
    ):
        return False
    t = (text or "").lower()
    return any(p in t for p in _SEARCH_HEADERS_PHRASES)


def _build_search_header_block(rec, score=None) -> str:
    """A single-line header for a vault match. ~30 tokens.

    Format: ``name  ·  type, R rows  ·  topics: a, b, c  ·  score 4.2``
    Used by search-headers mode in place of a full [VAULT MATCH] block.

    Semantic-only matches (no keyword hit) get a leading "[semantic] "
    tag so the model knows the file matched by meaning, not by query
    terms appearing in it.
    """
    name = rec.get("name") or "?"
    rtype = rec.get("type") or "?"
    bits = [rtype]
    if rtype in ("csv", "tsv", "csv.gz", "parquet"):
        rows = rec.get("rows")
        if isinstance(rows, int):
            bits.append(f"{rows:,} rows")
        n_cols = len(rec.get("headers") or [])
        if n_cols:
            bits.append(f"{n_cols} cols")
    elif rtype == "excel":
        sheets = rec.get("sheets") or []
        bits.append(f"{len(sheets)} sheets")
    elif rtype in ("sqlite", "duckdb"):
        tables = rec.get("tables") or []
        bits.append(f"{len(tables)} tables")
    topics = rec.get("topics") or []
    sem_tag = "[semantic] " if rec.get("_semantic_only") else ""
    parts = [f"{sem_tag}{name}", ", ".join(bits)]
    if topics:
        parts.append("topics: " + ", ".join(str(t) for t in topics[:4]))
    if score is not None:
        parts.append(f"score {score:.1f}")
    return "  ·  ".join(parts)


def _inject_file_contents(user_text, analyst_block=None, n_ctx=None,
                           task_memo_block=None, pinned_files=None):
    """Public entry point — wraps `_inject_file_contents_impl` in a
    defensive try/except so unexpected exceptions during injection
    (vault index corruption, network-share disconnect mid-walk,
    broken pdf parser, etc.) degrade to "no injection" rather than
    crashing `_send` and leaving the transcript hung on the user's
    last typed line.
    """
    try:
        return _inject_file_contents_impl(
            user_text, analyst_block=analyst_block, n_ctx=n_ctx,
            task_memo_block=task_memo_block,
            pinned_files=pinned_files,
        )
    except Exception as _top_e:
        import sys as _sys_dbg
        import traceback as _tb
        print(f"[inject] unexpected top-level exception: {_top_e!r}",
              file=_sys_dbg.stderr)
        _tb.print_exc(file=_sys_dbg.stderr)
        # Build a minimal breakdown so the caller doesn't crash on
        # `_injection_breakdown.get("costs", [])`.
        try:
            import council_engine as _ce_safe
            n_ctx_safe = int(n_ctx if n_ctx is not None else _ce_safe.get_n_ctx())
        except Exception:
            n_ctx_safe = int(n_ctx or 4096)
        breakdown = {
            "costs": [],
            "dropped": [],
            "n_ctx": n_ctx_safe,
            "remaining": n_ctx_safe,
            "per_block_cap": 2048,
            "running": 0,
            "user_text_tokens": 0,
            "injection_error": repr(_top_e),
        }
        # Optionally surface the failure as a synthetic NO_DATA-style
        # block so the model is told context retrieval failed and
        # refuses to invent values from training memory.
        warn = (
            "[INJECTION FAILURE — the vault / file readers raised an "
            "unexpected error while gathering context for this query.]\n"
            f"  error: {_top_e!r}\n"
            "[Treat this turn as if NO data has been provided. Do NOT "
            "invent specific values, file names, or row counts.]"
        )
        return (warn + "\n\n" + (user_text or "")), {}, breakdown


def _inject_file_contents_impl(user_text, analyst_block=None, n_ctx=None,
                                task_memo_block=None, pinned_files=None):
    """Augment the user message with file/vault context before deliberation.

    Returns ``(augmented_text, fuzzy_matches, breakdown)`` where:
      • ``breakdown["costs"]`` — list of ``(label, token_cost)`` for each
        block that made it into the assembled prompt (in order).
      • ``breakdown["dropped"]`` — list of ``(label, token_cost)`` for
        blocks that exceeded the cumulative budget. Only droppable-class
        blocks (vault matches) ever appear here.
      • ``breakdown["n_ctx"]`` / ``["remaining"]`` — the window the
        assembly was sized against.

    Priority order (lower number = higher priority; ties impossible):
      0. ``[NO DATA AVAILABLE]`` marker (must always surface)
      1. ``[TASK MEMO]`` — RAM-resident sticky note (goal/constraints/
         forbidden) so small models don't forget the original question
         once the context window fills up with file blocks
      2. ``[ANALYST RESULT]`` — the computed answer
      3. ``[FILE: ...]`` — explicit user-pasted paths
      4. ``[FOLDER: ...]`` — explicit user-pasted directory paths
      5. ``[VAULT MATCH: ...]`` — speculative search hits (droppable)

    Per-block cap: each block is head/tail-trimmed to roughly ``n_ctx//8``
    tokens before assembly. Cumulative cap: vault-match blocks stop being
    added once the running total would exceed the safe input budget
    (``n_ctx`` minus ``max(256, n_ctx*0.25)`` reply reserve, minus the
    typed user text and a 500-token writer-prompt overhead).

    ``analyst_block`` is optional and injected at priority 2 when given.
    When explicit paths are present, the vault search is skipped entirely
    (#4 — explicit paths take precedence; the user can re-ask without a
    path if they want fuzzy vault matches).
    """
    import sys as _sys_dbg
    print('[DEBUG inject] called with: ' + repr(user_text[:80]), file=_sys_dbg.stderr)

    # -- Budgets ------------------------------------------------------------
    try:
        import council_engine as _ce
        n_ctx = int(n_ctx if n_ctx is not None else _ce.get_n_ctx())
    except Exception:
        n_ctx = int(n_ctx or 4096)
    # Per-block cap is a *safety net* against pathological inputs (e.g. a
    # 1MB file that slipped past the renderer's own char-limit). The
    # renderers already cap themselves at ~12KB (~3,000 tokens), so the
    # cap floor of 2048 tokens means well-behaved blocks pass through
    # untouched while a runaway block still gets trimmed.
    # Per-block cap. Was max(2048, n_ctx // 4) which floored at 2048 for
    # n_ctx <= 8192 — meaning a single non-droppable block could eat 50%
    # of a 4K context, and several stacked could push past the window
    # entirely (the assembly only drops blocks at DROPPABLE_FROM and
    # above). Scale linearly with n_ctx so a 4K-ctx model gets ~512-token
    # blocks while a 32K-ctx model still gets 5K-token blocks.
    per_block_cap = max(512, n_ctx // 6)
    reply_reserve = max(256, int(n_ctx * 0.25))
    safe_input = max(1, n_ctx - reply_reserve)
    user_cost = _estimate_block_tokens(user_text)
    writer_overhead = 500   # rough Writer system prompt allowance
    remaining = max(256, safe_input - user_cost - writer_overhead)

    # Priority bands — lower = higher priority
    # VAULT_SUMMARY lists EVERY matching file by name (with brief
    # metadata: type, row count, topics) so the model knows the full
    # set even when individual VAULT MATCH content blocks get dropped
    # to fit the budget. Slotted ABOVE individual matches so the
    # filename list always reaches the model. Not droppable; tiny in
    # tokens (~15/file × ~50 files = ~750 tokens worst case).
    (PRIO_NODATA, PRIO_TASK_MEMO, PRIO_ANALYST,
     PRIO_EXPLICIT, PRIO_FOLDER,
     PRIO_VAULT_SUMMARY, PRIO_VAULT) = 0, 1, 2, 3, 4, 5, 6
    DROPPABLE_FROM = PRIO_VAULT   # only individual vault matches are droppable

    explicit_paths = _extract_file_paths(user_text)
    print('[DEBUG inject] explicit paths: ' + str(explicit_paths), file=_sys_dbg.stderr)

    candidates = []   # list of (priority, label, content)
    fuzzy_matches = {}
    missing_paths = []

    # -- Task memo (priority 1) ----------------------------------------------
    # A short RAM-resident sticky note carrying the user's original goal +
    # constraints + forbidden actions. Re-injected on every turn so even
    # if later context blocks push the user's typed message past the
    # window's tail, the model still sees what was asked. ~60-100 tokens.
    if task_memo_block:
        candidates.append((PRIO_TASK_MEMO, "[TASK MEMO]", task_memo_block))

    # -- Analyst result (priority 2) ----------------------------------------
    if analyst_block:
        candidates.append((PRIO_ANALYST, "[ANALYST RESULT]", analyst_block))

    # -- Explicit paths (priority 3 or 4) -----------------------------------
    # Rich folder rendering (per-file column previews / sheet names / JSON
    # key heads) is ~12 KB worst-case; compact rendering (counts + filename
    # list) is ~3.5 KB. On a tight 4 K-ctx model, picking rich for a pasted
    # folder eats the whole window. We pick adaptively:
    #   • n_ctx >= 16384 → rich (the original behaviour)
    #   • n_ctx <  16384 → compact (matches the auto-injection path's choice)
    # The user can always paste an individual file path inside the folder
    # to get full rich rendering for that one file.
    _use_compact_folder = n_ctx < 16384
    for path_str in explicit_paths:
        p_check = Path(path_str.strip())
        # Directory → folder listing block (priority 4, not droppable).
        if p_check.is_dir():
            if _use_compact_folder:
                folder_block = _render_folder_summary_compact(p_check)
                lbl_prefix = "[FOLDER SUMMARY: "
            else:
                folder_block = _render_folder_for_injection(p_check)
                lbl_prefix = "[FOLDER: "
            if folder_block:
                candidates.append((
                    PRIO_FOLDER, f"{lbl_prefix}{p_check.name or path_str}]",
                    folder_block,
                ))
            continue
        snippet = _read_file_for_injection(path_str)
        if snippet:
            candidates.append((
                PRIO_EXPLICIT, f"[FILE: {p_check.name or path_str}]", snippet,
            ))
        else:
            missing_paths.append(path_str)

    # -- Vault folder context (priority 4) ----------------------------------
    # When the user mentions "the vault", "my vault", "vault folder", etc.
    # but didn't paste an explicit path, inject a COMPACT data_in/ summary
    # (counts + subfolder breakdown + filename list — no per-file column
    # previews) so the model knows what files exist. Without this, vault
    # search may return zero matches for generic queries ("show me what's
    # in the vault") and the model defaults to "I don't have access to
    # your filesystem".
    #
    # Skipped when:
    #   • the user pasted an explicit path (FILE/FOLDER block already
    #     covers their intent)
    #   • the analyst block is present (analyst has authoritative counts;
    #     adding a second full listing wastes the prompt budget — this
    #     was the crash trigger on the 4K-ctx laptop: "how many files in
    #     data_in" → analyst answers, then injection ALSO rendered the
    #     full folder = 3K+ tokens of duplicate context.)
    #
    # Compact renderer caps at ~3.5 KB even on 1000-file vaults, so a
    # 4 K-ctx model has headroom for the user text + reply reserve.
    #
    # Rebuild the vault index ONCE per turn when we'll consult it (no explicit
    # paths). The folder-summary render(s) below and the vault-search branch
    # each used to trigger their own rebuild() on the same singleton — 2-3 full
    # rglob walks + a stat() per file per turn. rebuild() is delta-aware and
    # idempotent, so one consolidated call yields byte-identical folder + search
    # blocks. Passed into the renders via vault_index= so they skip re-walking.
    _vault_idx = None
    if not explicit_paths:
        _vault_idx = _get_vault_index()
        if _vault_idx is not None:
            try:
                import council_engine as _ce_tim0
                with _ce_tim0._TimingScope("vault.rebuild"):
                    _vault_idx.rebuild()
            except Exception as _rb_exc:
                print('[DEBUG inject] vault index rebuild failed: '
                      + repr(_rb_exc), file=_sys_dbg.stderr)
    if (_vault_search_keywords(user_text)
            and not explicit_paths
            and analyst_block is None):
        try:
            import data_index as _di
            vault_data_in = _di.input_dir(VAULT_DIR)
        except Exception:
            vault_data_in = VAULT_DIR
        try:
            vault_folder_block = _render_folder_summary_compact(
                vault_data_in, vault_index=_vault_idx)
        except Exception as _e:
            print('[DEBUG inject] vault folder render failed: ' + repr(_e),
                  file=_sys_dbg.stderr)
            vault_folder_block = None
        if vault_folder_block:
            vf_label = f"[FOLDER SUMMARY: {Path(vault_data_in).name or str(vault_data_in)}]"
            already_have_vault_folder = any(
                lbl == vf_label for _prio, lbl, _content in candidates
            )
            if not already_have_vault_folder:
                candidates.append((PRIO_FOLDER, vf_label, vault_folder_block))
                print('[DEBUG inject] vault folder injected: '
                      + str(vault_data_in), file=_sys_dbg.stderr)

        # Sub-folder reference: user said something like "look in the
        # projects subfolder" or "files in Q3_2024". Compact-render
        # that subfolder so the user gets focused context.
        try:
            import vault_analyst as _va_sub
            sub = _va_sub.resolve_subfolder_hint(user_text, vault_data_in)
        except Exception:
            sub = None
        if sub is not None and sub != vault_data_in:
            try:
                sub_block = _render_folder_summary_compact(
                    sub, vault_index=_vault_idx)
            except Exception:
                sub_block = None
            if sub_block:
                sub_label = f"[FOLDER SUMMARY: {sub.name or str(sub)}]"
                already = any(lbl == sub_label
                              for _prio, lbl, _content in candidates)
                if not already:
                    candidates.append((PRIO_FOLDER, sub_label, sub_block))
                    print('[DEBUG inject] vault subfolder injected: '
                          + str(sub), file=_sys_dbg.stderr)

    # -- Vault search (priority 5, droppable) -------------------------------
    # #4: when the user pasted explicit paths, trust them and skip vault
    # search entirely. The vault hits competed with the explicit file for
    # budget and frequently caused the explicit file to get truncated.
    folder_scope = _detect_folder_scope(user_text)
    if not explicit_paths:
        idx = _vault_idx   # rebuilt once at the top of this region
        if idx is not None:
            try:
                # When analyst already answered (#7), reduce the vault-match
                # pull from 5 to 1 — the analyst's CSV is the authoritative
                # source and 5 fuzzy matches just consume budget.
                #
                # K and TAIL_K both scale with n_ctx so a 4 K-ctx model
                # doesn't get 5 × ~400-token VAULT MATCH blocks (2 KB just
                # in matches, half its window). 8 K → 3 / 30, 16 K → 5 / 45.
                if n_ctx <= 4096:
                    base_k, base_tail = 2, 15
                elif n_ctx <= 8192:
                    base_k, base_tail = 3, 30
                else:
                    base_k, base_tail = 5, 45
                k = 1 if analyst_block else base_k
                # Semantic expansion via the local model: when a query
                # term isn't in the vault's vocab, the index calls the
                # model to ask "which of these vocab tokens belong in
                # the queried category?" The model decides per-vault
                # (so "metals" returns "promethium" only if it's
                # actually in this user's files). Cached on disk; one
                # call per novel concept ever.
                def _semantic_llm_call(prompt: str) -> str:
                    import council_engine as _ce_sem
                    return _ce_sem.local_chat(
                        messages=[{"role": "user", "content": prompt}],
                        temperature=0.0,
                        num_predict=180,
                        timeout=45,
                    )
                # We pull k + TAIL_K total hits. The top-k get FULL
                # [VAULT MATCH] content blocks; the rest go into a
                # single compact [VAULT SEARCH SUMMARY] block that
                # lists every matching file by name + brief metadata.
                # This is the fix for the "model only sees 5 files
                # even though 47 match" failure mode: the summary is
                # ~15 tokens per file, so 50 files cost ~750 tokens
                # vs ~2000 tokens for 5 full blocks. The model now
                # KNOWS the full set of matching files even when only
                # a few get full content.
                TAIL_K = base_tail
                # NOTE: `_ce_tim` was referenced here but never imported
                # (introduced by the "behavior-preserving" perf batch c613e75),
                # so this whole block raised NameError on EVERY turn and the
                # outer except swallowed it — vault search was silently dead.
                # Import the alias locally, matching _ce_sem / _ce_pack nearby.
                import council_engine as _ce_tim
                with _ce_tim._TimingScope("vault.search"):
                    all_hits, fuzzy_matches = idx.search(
                        user_text, k=k + TAIL_K, folder=folder_scope,
                        llm_call=_semantic_llm_call,
                    )
                # Drop any hit that's already explicit (won't happen
                # in practice — we skip vault search entirely when
                # explicit_paths is non-empty — but safe to keep the
                # filter for future call sites).
                all_hits = [(s, r) for s, r in all_hits
                            if r.get("path") not in explicit_paths]
                full_hits = all_hits[:k]
                tail_hits = all_hits[k:]

                print('[DEBUG inject] vault hits: total='
                      + str(len(all_hits))
                      + ' full=' + str(len(full_hits))
                      + ' tail=' + str(len(tail_hits)),
                      file=_sys_dbg.stderr)
                if fuzzy_matches:
                    print('[DEBUG inject] fuzzy: ' + repr(fuzzy_matches),
                          file=_sys_dbg.stderr)

                # Decide between full [VAULT MATCH] blocks and compact
                # search-headers, based on intent. Conceptual queries
                # ("show me files about X", "find scripts that …") get
                # one-line headers: the model sees a wider set of
                # matches per token and can ask follow-up "tell me more
                # about file Y" in the next turn — those references
                # will trigger the zoom-in via pinned_files (see
                # CouncilConsole._send wire-up).
                use_search_headers = _wants_search_headers(user_text)

                # Pinned files from a previous turn — the user's last
                # answer referenced these by name, so promote them out
                # of headers mode into full blocks even when the
                # current query looks conceptual. Pins expire after a
                # few turns to avoid bloat (managed by the caller).
                pinned_set = {str(n).lower() for n in (pinned_files or [])}

                try:
                    import council_engine as _ce_pack
                    _count_tokens = _ce_pack.estimate_tokens
                except Exception:
                    _count_tokens = None
                vault_block_budget = max(256, remaining // 2)

                if use_search_headers and not pinned_set:
                    # Pure headers mode — no per-file rich blocks.
                    # Pack everything we found into one compact block.
                    header_lines = ["[VAULT MATCHES (headers — ask about a "
                                     "file by name to zoom in)]"]
                    for score, rec in all_hits:
                        header_lines.append("  • " + _build_search_header_block(rec, score))
                    headers_block = "\n".join(header_lines)
                    candidates.append((
                        PRIO_VAULT, "[VAULT MATCHES — headers]", headers_block,
                    ))
                    print('[DEBUG inject] search-headers mode: '
                          + str(len(all_hits)) + ' headers',
                          file=_sys_dbg.stderr)
                else:
                    # Full-block mode (legacy + pinned-zoom path).
                    # Split hits into "pinned" (always full) and
                    # "everyone else" (headers if search-headers mode).
                    pinned_recs = []
                    other_full_hits = []
                    for score, rec in full_hits:
                        if str(rec.get("name", "")).lower() in pinned_set:
                            pinned_recs.append(rec)
                        else:
                            other_full_hits.append((score, rec))

                    packed_blocks, pack_diag = idx.assemble_match_blocks(
                        pinned_recs + [rec for _s, rec in other_full_hits],
                        budget_tokens=vault_block_budget,
                        count_tokens=_count_tokens,
                    )
                    print('[DEBUG inject] vault match assembly: '
                          + repr(pack_diag), file=_sys_dbg.stderr)
                    ordered_recs = pinned_recs + [r for _s, r in other_full_hits]
                    for rec, block in zip(ordered_recs, packed_blocks):
                        name = rec.get("name") or "?"
                        label_prefix = ("[VAULT MATCH (pinned): "
                                        if str(name).lower() in pinned_set
                                        else "[VAULT MATCH: ")
                        candidates.append((PRIO_VAULT, f"{label_prefix}{name}]", block))

                # Summary block — emitted ONLY when not in headers mode,
                # because the headers block already covers the same
                # ground (compact per-file listing). When tail_hits is
                # empty the summary is redundant; when search-headers
                # is on, ALL hits already render as headers, so the
                # tail summary is also redundant.
                if tail_hits and not use_search_headers:
                    summary_block = _build_vault_search_summary(
                        all_hits, full_count=len(full_hits),
                    )
                    candidates.append((
                        PRIO_VAULT_SUMMARY,
                        "[VAULT SEARCH SUMMARY]",
                        summary_block,
                    ))
            except Exception as _e:
                print('[DEBUG inject] vault search failed: ' + repr(_e),
                      file=_sys_dbg.stderr)

    # -- Cell-value matches (compact, high-priority) ------------------------
    # The vault search above is filename / header / topic / embedding biased,
    # so a value that lives INSIDE a file (a cell in row 400, a code, a name)
    # frequently isn't surfaced — the app's core "can't find data inside my
    # files" complaint. Here we run the deterministic cell scanner
    # (data_index.search_value — pure pandas/dict, NO model call, so the
    # single serialized-inference lock is untouched) over the query's content
    # terms and fold real hits into the context as a compact block. Slotted at
    # PRIO_VAULT_SUMMARY: above droppable vault matches (these are ACTUAL cell
    # hits, more authoritative than fuzzy matches) and kept small by self-cap.
    if not explicit_paths and analyst_block is None:
        try:
            _val_terms = _content_query_terms(user_text)
            _di_inst = _get_data_index() if _val_terms else None
            if _di_inst is not None:
                try:
                    _di_inst.refresh()
                except Exception:
                    pass
                _val_merged: dict = {}
                for _t in _val_terms:
                    try:
                        _vhits = _di_inst.search_value(_t, max_per_file=5)
                    except Exception:
                        _vhits = []
                    for _h in _vhits:
                        _ent = _val_merged.setdefault(_h["path"], {
                            "file": _h["file"], "cols": [], "rows": [],
                            "terms": set(),
                        })
                        _ent["terms"].add(_t)
                        for _c in _h["column_hits"]:
                            if _c not in _ent["cols"]:
                                _ent["cols"].append(_c)
                        for _r in _h["rows"]:
                            if len(_ent["rows"]) < 3:
                                _ent["rows"].append(_r)
                if _val_merged:
                    # Most terms matched + most rows first; cap files for size.
                    _ranked = sorted(
                        _val_merged.values(),
                        key=lambda e: (-len(e["terms"]), -len(e["rows"])),
                    )[:4]
                    _vlines = ["[VALUE MATCHES — rows inside your vault files "
                               "that contain the search terms]"]
                    for _ent in _ranked:
                        _vlines.append(
                            f"  {_ent['file']}  (matched in column(s): "
                            f"{', '.join(_ent['cols'][:6])})"
                        )
                        for _r in _ent["rows"]:
                            _cells = "  |  ".join(
                                f"{_k}={_v}" for _k, _v in _r.items()
                                if _v not in (None, "", "nan")
                            )
                            _vlines.append("    " + _cells[:160])
                    candidates.append(
                        (PRIO_VAULT_SUMMARY, "[VALUE MATCHES]", "\n".join(_vlines))
                    )
                    print('[DEBUG inject] value matches: '
                          + str(len(_ranked)) + ' file(s) from terms '
                          + repr(_val_terms), file=_sys_dbg.stderr)
        except Exception as _ve:
            print('[DEBUG inject] value search failed: ' + repr(_ve),
                  file=_sys_dbg.stderr)

    # -- Dataset overview (expert-mode grounding) ---------------------------
    # A compact, always-current map of ALL of data_in so the model knows the
    # whole dataset it is meant to be an expert on; the retrieval blocks above
    # then supply the specifics. Model-free + cached (rebuilds only when data_in
    # changes), size-scaled to n_ctx, and gated to real data questions.
    if (not explicit_paths and analyst_block is None
            and _content_query_terms(user_text)):
        try:
            import dataset_digest as _dd
            _digest_cap = (1200 if n_ctx <= 4096
                           else 2500 if n_ctx <= 8192 else 4000)
            _digest = _dd.get_digest(VAULT_DIR, max_chars=_digest_cap)
            if _digest:
                candidates.append((
                    PRIO_VAULT_SUMMARY, "[DATASET OVERVIEW]",
                    "[DATASET OVERVIEW — every data file currently in the "
                    "vault; ask about any of them]\n" + _digest))
                print('[DEBUG inject] dataset overview injected ('
                      + str(len(_digest)) + ' chars)', file=_sys_dbg.stderr)
        except Exception as _de:
            print('[DEBUG inject] dataset overview failed: ' + repr(_de),
                  file=_sys_dbg.stderr)

    # -- NO DATA marker (priority 1, never dropped) -------------------------
    # Critical defense against cross-machine hallucination: when the user
    # mentioned a path that doesn't exist locally, surface the gap loudly
    # so the Writer's ABSOLUTE RULE clause kicks in and refuses values.
    if missing_paths:
        miss_block = (
            "[NO DATA AVAILABLE — the user's message referenced these "
            "file paths, but they do NOT exist on this machine:]"
        )
        for mp in missing_paths:
            miss_block += "\n  - " + str(mp)
        miss_block += (
            "\n[Refuse to give specific values for these files. Tell "
            "the user the file is not present on this machine. Do NOT "
            "invent column names, row counts, or values from memory.]"
        )
        candidates.append((PRIO_NODATA, "[NO DATA AVAILABLE]", miss_block))

    # -- Assemble: priority sort, per-block cap, cumulative cap, tag --------
    candidates.sort(key=lambda t: t[0])

    placed: list = []   # list of (prio, label, capped, cost)
    dropped = []
    running = 0
    # NO_DATA / TASK_MEMO / ANALYST are exempt from the per-block cap:
    #   • NO_DATA is tiny by construction.
    #   • TASK_MEMO is intentionally short (~80 tokens) and trimming it
    #     would defeat the "remember the original question" guarantee.
    #   • ANALYST is already capped at ~4 KB by format_result_for_prompt;
    #     trimming further could elide the answer the user asked for.
    # VAULT_SUMMARY is uncapped on the per-block axis (it scales with hits
    # already) but DOES participate in cumulative-budget eviction below.
    UNCAPPED_PRIOS = {PRIO_NODATA, PRIO_TASK_MEMO, PRIO_ANALYST,
                      PRIO_VAULT_SUMMARY}
    # Truly undroppable — these must reach the model no matter what.
    # Everything else is sacrificeable in reverse-priority order if the
    # cumulative budget gets blown by stacked non-droppable blocks (the
    # old assembly let those overflow the window unconditionally).
    HARD_KEEP = {PRIO_NODATA, PRIO_TASK_MEMO, PRIO_ANALYST}
    for (prio, label, content) in candidates:
        if prio in UNCAPPED_PRIOS:
            capped = content
        else:
            capped = _smart_truncate_block_to_tokens(content, per_block_cap)
        cost = _estimate_block_tokens(capped)
        # First-pass cumulative gate — only droppable blocks are filtered
        # here. Non-droppable ones go into `placed` unconditionally and
        # the rescue pass below handles overflow.
        if prio >= DROPPABLE_FROM and running + cost > remaining:
            dropped.append((label, cost))
            continue
        placed.append((prio, label, capped, cost))
        running += cost

    # -- Cumulative-overflow rescue ------------------------------------------
    # The first pass enforces budget only for droppable blocks. If multiple
    # non-droppable blocks (EXPLICIT / FOLDER / VAULT_SUMMARY) stack up on
    # a small context window, `running` can still exceed `remaining` — the
    # crash mode reported on the 4 K-ctx laptop. Walk back over the placed
    # blocks in REVERSE priority order (lowest priority first), evicting
    # until the budget fits. HARD_KEEP blocks are never evicted.
    if running > remaining:
        # Before evicting an overflow block, try to CONDENSE it to fit —
        # chunk it and keep the parts most relevant to the user's task
        # (the [TASK MEMO]) instead of losing the whole block. This is the
        # "extend the model's effective context" path: a big file or wide
        # vault-match set survives in digest form on a small window.
        # Deterministic by default (no latency); COUNCIL_CONDENSE_LLM=1
        # switches to a model map-reduce. Disable with
        # COUNCIL_CONDENSE_OVERFLOW=0 to restore plain eviction.
        _condense_on = os.environ.get(
            "COUNCIL_CONDENSE_OVERFLOW", "1").strip().lower() \
            not in ("0", "false", "no", "off")
        _cond = None
        if _condense_on:
            try:
                import context_condenser as _cond
            except Exception:
                _cond = None
        _cond_llm = None
        if _cond is not None and os.environ.get(
                "COUNCIL_CONDENSE_LLM", "").strip().lower() in (
                "1", "true", "yes", "on"):
            def _cond_llm(_p):
                import council_engine as _cc
                return _cc.local_chat(
                    messages=[{"role": "user", "content": _p}],
                    temperature=0.0, num_predict=256, timeout=60)
        _task_src = task_memo_block or user_text or ""

        evictable_order = sorted(
            range(len(placed)),
            key=lambda i: (-placed[i][0], -i),
        )
        keep_mask = [True] * len(placed)
        for i in evictable_order:
            if running <= remaining:
                break
            prio_i, label_i, _capped_i, cost_i = placed[i]
            if prio_i in HARD_KEEP:
                continue
            # How much room this block may occupy if it's the one that fits.
            headroom = remaining - (running - cost_i)
            if _cond is not None and headroom >= 64:
                try:
                    condensed = _cond.condense_to_fit(
                        _capped_i, headroom, task=_task_src,
                        estimate_tokens=_estimate_block_tokens,
                        llm_call=_cond_llm)
                except Exception as _cx:
                    print('[inject] condense failed: ' + repr(_cx),
                          file=_sys_dbg.stderr)
                    condensed = None
                if condensed:
                    new_cost = _estimate_block_tokens(condensed)
                    if new_cost < cost_i and new_cost <= headroom:
                        placed[i] = (prio_i, label_i, condensed, new_cost)
                        running -= (cost_i - new_cost)
                        print(f'[inject] condensed {label_i}: '
                              f'{cost_i}->{new_cost} tok', file=_sys_dbg.stderr)
                        continue
            keep_mask[i] = False
            dropped.append((label_i + " (budget overflow)", cost_i))
            running -= cost_i
        placed = [b for i, b in enumerate(placed) if keep_mask[i]]

    # Re-sort by priority (eviction can reorder via mask removal — keep
    # the assembled order priority-stable for the model).
    placed.sort(key=lambda t: t[0])

    final_blocks = []
    per_block_costs = []
    for prio, label, capped, cost in placed:
        tagged = _tag_block_header(capped, cost)
        final_blocks.append(tagged)
        per_block_costs.append((label, cost))

    breakdown = {
        "costs":    per_block_costs,
        "dropped":  dropped,
        "n_ctx":    n_ctx,
        "remaining": remaining,
        "per_block_cap": per_block_cap,
        "running":  running,
        "user_text_tokens": user_cost,
    }

    if not final_blocks:
        return user_text, fuzzy_matches, breakdown
    return ('\n\n'.join(final_blocks) + '\n\n' + user_text,
            fuzzy_matches, breakdown)


_FOLDER_RE = _re.compile(
    r"(?:in(?:side)?|under|within|from)\s+(?:my|the)?\s*"
    r"(?:folder|directory|dir|subfolder)\s+"
    r"['\"]?([A-Za-z0-9_\-./\\]+)['\"]?",
    _re.IGNORECASE,
)
_FOLDER_KEYWORD_RE = _re.compile(
    r"\bfolder[: ]+['\"]?([A-Za-z0-9_\-./\\]+)['\"]?",
    _re.IGNORECASE,
)


def _build_vault_search_summary(all_hits, full_count: int) -> str:
    """Format a compact [VAULT SEARCH SUMMARY] block listing every
    matching file by name + brief metadata. ``full_count`` is how many
    of the hits also got an individual [VAULT MATCH] block above this
    one — so the summary can say "showing full content for the top N".

    Per-row format (keeps total cost ~15 tokens per file):
        N.  <name>   [type, R rows]   topics: t1, t2, t3
    or for non-tabular:
        N.  <name>   [type, K keys]   topics: ...

    The block always carries the TOTAL count and a hint that the
    user can ask about a specific file to get its full content.
    """
    total = len(all_hits)
    lines: list = ["[VAULT SEARCH SUMMARY]"]
    lines.append(f"Total files matching the query: {total}")
    if full_count > 0:
        lines.append(
            f"Full content for the top {full_count} match"
            f"{'es' if full_count != 1 else ''} is shown in the "
            f"[VAULT MATCH] blocks below. The remaining "
            f"{total - full_count} match"
            f"{'es are' if total - full_count != 1 else ' is'} "
            f"listed by filename only — ask about a specific file "
            f"by name to see its content."
        )
    else:
        lines.append("Each entry is listed by filename only — ask "
                     "about a specific file by name to see its content.")
    lines.append("")
    lines.append("Files matched (relevance-ordered):")

    for i, (_score, rec) in enumerate(all_hits, start=1):
        name = rec.get("name") or "?"
        rtype = rec.get("type") or "?"
        # Compact metadata that varies by type.
        meta_bits: list = [rtype]
        if rtype in ("csv", "tsv", "csv.gz", "parquet"):
            rows = rec.get("rows")
            if isinstance(rows, int):
                meta_bits.append(f"{rows:,} rows")
            n_cols = len(rec.get("headers") or [])
            if n_cols:
                meta_bits.append(f"{n_cols} cols")
        elif rtype == "excel":
            sheets = rec.get("sheets") or []
            meta_bits.append(f"{len(sheets)} sheets")
        elif rtype in ("sqlite", "duckdb"):
            tables = rec.get("tables") or []
            meta_bits.append(f"{len(tables)} tables")
        elif rtype in ("json", "d3dpipeline", "bson"):
            keys = rec.get("keys") or []
            if keys:
                meta_bits.append(f"{len(keys)} keys")
            if rec.get("indexing_tier") == "sampled_head_tail":
                meta_bits.append("sampled")

        topics = rec.get("topics") or []
        topic_str = ""
        if topics:
            preview = ", ".join(str(t) for t in topics[:5])
            if len(topics) > 5:
                preview += f", +{len(topics) - 5}"
            topic_str = f"  topics: {preview}"

        meta = ", ".join(meta_bits)
        lines.append(f"  {i:>3}. {name}   [{meta}]{topic_str}")

    lines.append("[END VAULT SEARCH SUMMARY]")
    return "\n".join(lines)


def _detect_folder_scope(text):
    """If the user references a folder by name, return that folder substring."""
    for rx in (_FOLDER_KEYWORD_RE, _FOLDER_RE):
        m = rx.search(text or "")
        if m:
            cand = m.group(1).strip().strip("'\"`")
            if "." not in Path(cand).name or Path(cand).suffix == "":
                return cand
    return None


_VAULT_TRIGGER_PHRASES = (
    "look through", "search through", "find files", "find every", "find any",
    "every file", "any file with", "all files", "across files",
    "in my vault", "in the vault", "from the vault", "the vault folder",
    "my vault folder", "vault folder", "vault directory", "the vault",
    "data_in", "data folder", "input folder",
    "what's in the vault", "whats in the vault", "what is in the vault",
    "files that contain", "files with", "which files",
    "list files", "list of files", "show files", "show me the files",
    "show me what's", "scan", "index",
)


def _vault_search_keywords(text):
    t = (text or "").lower()
    return any(p in t for p in _VAULT_TRIGGER_PHRASES)


_VAULT_INDEX_INSTANCE = None


def _get_vault_index():
    """Lazy-init the vault index. Returns None on import/setup failure."""
    global _VAULT_INDEX_INSTANCE
    if _VAULT_INDEX_INSTANCE is not None:
        return _VAULT_INDEX_INSTANCE
    try:
        import vault_index as _vi
        _VAULT_INDEX_INSTANCE = _vi.VaultIndex(VAULT_DIR)
    except Exception as _e:
        import sys as _sys_dbg
        print('[VaultIndex] init failed: ' + repr(_e), file=_sys_dbg.stderr)
        _VAULT_INDEX_INSTANCE = None
    return _VAULT_INDEX_INSTANCE


_DATA_INDEX_INSTANCE = None


def _register_data_index(di) -> None:
    """Let module-level helpers (the context injector's value-search stage)
    reuse the SAME DataIndex the console built + refreshes, instead of
    constructing a duplicate one."""
    global _DATA_INDEX_INSTANCE
    _DATA_INDEX_INSTANCE = di


def _get_data_index():
    """Return a DataIndex over the vault's data_in/ + bundled samples.

    Prefers the instance the console registered (already warm + refreshed);
    otherwise lazily builds one so headless / agent contexts still get value
    search. Returns None on failure — callers must degrade gracefully.
    """
    global _DATA_INDEX_INSTANCE
    if _DATA_INDEX_INSTANCE is not None:
        return _DATA_INDEX_INSTANCE
    try:
        import data_index as _di
        _DATA_INDEX_INSTANCE = _di.DataIndex(
            search_roots=[
                _di.input_dir(VAULT_DIR),
                _di.bundled_samples_dir(),
            ],
            write_root=_di.output_dir(VAULT_DIR),
        )
    except Exception as _e:
        import sys as _sys_dbg
        print('[DataIndex] module init failed: ' + repr(_e), file=_sys_dbg.stderr)
        _DATA_INDEX_INSTANCE = None
    return _DATA_INDEX_INSTANCE


# Stop-words for the value-search stage's content-term extraction. Module-level
# because the console's _query_keywords is a method, unavailable to the
# module-level injector. Broad by design: we want proper-noun / ID-like tokens
# likely to appear as CELL VALUES, not generic verbs or aggregate words (which
# describe an OPERATION on the data, not a value stored in it).
_VALUE_QUERY_STOPS = frozenset({
    "the", "a", "an", "of", "for", "by", "in", "on", "at", "and", "or", "to",
    "with", "what", "which", "who", "whose", "how", "many", "much", "do",
    "does", "did", "is", "are", "was", "were", "have", "has", "had", "from",
    "this", "that", "these", "those", "all", "any", "show", "tell", "give",
    "find", "list", "look", "up", "lookup", "search", "me", "my", "our",
    "your", "us", "i", "about", "across", "between", "into", "per", "over",
    "under", "most", "least", "more", "less", "than", "when", "where", "why",
    "value", "values", "row", "rows", "record", "records", "file", "files",
    "data", "column", "columns", "field", "fields", "get", "see", "please",
    "name", "named", "called",
    # Aggregate / analytic words describe an operation, not a stored value.
    "average", "avg", "total", "sum", "count", "mean", "median", "mode",
    "maximum", "minimum", "max", "min", "percentage", "percent", "number",
    "amount", "compare", "trend", "group", "grouped",
})


def _content_query_terms(text, *, max_terms: int = 6):
    """Content terms likely to be CELL VALUES: strip punctuation + stop-words,
    keep tokens length >= 3 (or ANY token containing a digit — an ID/code).
    Deduped, order-preserving, capped at ``max_terms``."""
    out: List[str] = []
    seen: Set[str] = set()
    for raw in (text or "").split():
        t = raw.strip(".,!?;:()[]{}\"'`").lower()
        if not t or t in _VALUE_QUERY_STOPS:
            continue
        if len(t) < 3 and not any(ch.isdigit() for ch in t):
            continue
        if t in seen:
            continue
        seen.add(t)
        out.append(t)
        if len(out) >= max_terms:
            break
    return out


# ---- Data analyst step ----------------------------------------------------
# Computational questions ("how many", "what percentage", "average X") deserve
# real numbers, not the model's best guess from a sample. This step asks a
# local model to write pandas code calling our helpers, executes it in a
# locked-down sandbox, and returns the result as text the Writer can quote.

def _run_analyst_step(query):
    """Public entry point — wraps `_run_analyst_step_impl` in a defensive
    try/except so that any unexpected exception (network share dropped
    mid-call, malformed vault state, OOM mid-tokenize, etc.) degrades
    cleanly to "no analyst" rather than crashing `_send`. The caller's
    transcript stays interactive instead of hanging on "computing…".
    """
    try:
        return _run_analyst_step_impl(query)
    except Exception as _top_e:
        import sys as _sys_dbg
        import traceback as _tb
        print(f"[analyst] unexpected top-level exception: {_top_e!r}",
              file=_sys_dbg.stderr)
        _tb.print_exc(file=_sys_dbg.stderr)
        return None, None, []


def _run_analyst_step_impl(query):
    """If `query` looks computational, generate pandas code via a local model
    and execute it sandboxed against the vault's data_in/ folder.

    Returns ``(block, error_msg, notices)``:
      • ``(None, None, [])``             — query isn't computational.
      • ``(success_block, None, [n…])``  — code executed; notices is a
        list of short human-friendly strings about how the analyst
        resolved filename / subfolder references (so the user can spot
        wrong matches: "I asked about sales.csv but the analyst picked
        Sales_Q3_2024.csv").
      • ``(failure_block, msg, [n…])``   — analyst was attempted but
        failed. The failure block carries an explicit refusal directive
        so the Writer doesn't invent a number; ``msg`` is a one-line
        summary the caller posts to the transcript.
    """
    import sys as _sys_dbg
    notices: list = []
    try:
        import vault_analyst as _va
    except Exception as _e:
        print('[analyst] import failed: ' + repr(_e), file=_sys_dbg.stderr)
        return None, None, notices

    # ── Direct-route: PRECOMPUTED ANSWER from a deferred task ────────
    # If the user re-asks something they previously deferred to the Vault
    # tab and ran there, answer FROM that saved result instead of
    # recomputing — that is the whole point of "Defer to Vault": ask again
    # later, get the answer.
    #
    # This runs BEFORE the looks_computational gate on purpose: a re-asked
    # deferred question must surface its saved answer even when the phrasing
    # doesn't read as "computational" (e.g. "bigger summary of sales.csv"
    # has no compute keyword and would otherwise be dropped here).
    # Reuse goes through the DerivedStore, which only returns a result whose
    # SOURCES ARE UNCHANGED — so a precomputed average is never served after
    # the underlying CSVs were edited (the staleness fix).
    try:
        import derived_results as _dr
        _ans = _dr.DerivedStore(VAULT_DIR).find_fresh(query)
    except Exception:
        _ans = None
    if _ans is not None and _ans.output:
        try:
            import pandas as _pd_pre
            _rp = Path(_ans.output)
            _rdf = _pd_pre.read_csv(_rp)
            try:
                _n_ctx_pre = ce.get_n_ctx()
            except Exception:
                _n_ctx_pre = 4096
            _ptable = _va.format_result_for_prompt(
                _rdf, max_rows=300, max_chars=12000,
                max_tokens=max(150, _n_ctx_pre // 4),
                count_tokens=ce.estimate_tokens)
            block = (f"[ANALYST RESULT — precomputed (sources unchanged)]\n"
                     f"# This was computed earlier ({_ans.operation or _ans.label}) "
                     f"and its source files are UNCHANGED, so answer from this "
                     f"saved result ({_rp.name}); do NOT recompute.\n"
                     f"{_ptable}")
            notices.append(
                f"Reused a fresh precomputed result — {_rp.name}.")
            try:
                notices.append("__ANALYST_TABLE__:" + _rdf.to_string(
                    index=False, max_rows=80, max_cols=20))
            except Exception:
                pass
            notices.append(
                "__ANALYST_ANSWER__:"
                f"Reused a saved result that is still current "
                f"({_rp.name}) — its source files are unchanged, so this "
                f"was not recomputed (see the table above).")
            # Source files behind this answer → provenance chips.
            try:
                import json as _js
                _srcs = list(getattr(_ans, "sources", []) or [])
                _srcs.append(str(_rp))
                notices.append("__ANALYST_SOURCES__:" + _js.dumps(_srcs))
            except Exception:
                pass
            return block, None, notices
        except Exception as _pe:
            print('[analyst] precomputed-answer load failed: ' + repr(_pe),
                  file=_sys_dbg.stderr)
            # fall through to normal routing if the saved file is unreadable

    # ── Direct-route: a saved COLLECTION named in the query ─────────
    # "show me Job Blue" / "summarise the Job Blue files" — if the query
    # names a saved collection, tell the council exactly which files make up
    # that project so it answers about the whole set (and only that set).
    # Runs before the looks_computational gate ("show me X" isn't a compute
    # phrase). The Vault tab's Collections panel is where the set is built.
    try:
        import vault_collections as _vc_q
        _coll = _vc_q.CollectionStore(VAULT_DIR).find_in_text(query)
    except Exception:
        _coll = None
    if _coll is not None and len(_coll.name.strip()) >= 3 and _coll.files:
        _flist = "\n".join(f"  - {f}" for f in _coll.files[:200])
        _more = (f"\n  …(+{len(_coll.files) - 200} more)"
                 if len(_coll.files) > 200 else "")
        block = (f"[ANALYST RESULT — collection “{_coll.name}”]\n"
                 f"# “{_coll.name}” is a saved collection the user grouped "
                 f"together: {len(_coll.files)} file(s). When answering about "
                 f"“{_coll.name}”, use ONLY these files:\n{_flist}{_more}")
        notices.append(
            f"Recognised the “{_coll.name}” collection — {len(_coll.files)} file(s).")
        notices.append(
            "__ANALYST_ANSWER__:"
            f"“{_coll.name}” is a saved collection of {len(_coll.files)} "
            f"file(s):\n{_flist}{_more}")
        try:
            import json as _js
            notices.append("__ANALYST_SOURCES__:" + _js.dumps(
                list(_coll.files or [])))
        except Exception:
            pass
        return block, None, notices

    # No precomputed answer / collection — only continue into the data
    # routes when the question actually looks computational.
    if not _va.looks_computational(query):
        return None, None, notices

    try:
        allowed_folders = [data_index.input_dir(VAULT_DIR)]
    except Exception:
        allowed_folders = [VAULT_DIR]

    # ── Direct-intent shortcut for "true data summary" queries ──────
    # These map deterministically to folder_data_summary() — no model
    # codegen needed. Saves ~2-5 s per call AND removes a class of
    # failure mode (the model picking the wrong helper). The trigger
    # phrases below are a subset of _COMPUTE_KEYWORDS that ONLY make
    # sense as "summarise everything"; anything ambiguous still goes
    # through the model.
    _DIRECT_SUMMARY_TRIGGERS = (
        "true data summary",
        "data summary",
        "summary of the files",
        "summary of files",
        "summarize the files",
        "summarize files",
        "describe the files",
        "overview of files",
        "overview of the files",
        "overview of the data",
        "inventory of files",
        "file inventory",
        "what's in this folder",
        "what is in this folder",
        "schema of the files",
        "schemas of",
        "profile the data",
        "profile this folder",
    )
    qlower = (query or "").lower()

    # ── Direct-route: FILE COUNT ("how many files in data_in") ──────
    # A trivial question that must NOT go through model code-gen — the
    # code-gen prompt is ~3.5K tokens and overflows a small context
    # window ("exceeds max tokens" crash on the 4K-ctx machine). Answer
    # it deterministically with a cheap census (no file reads, tiny
    # prompt). Guarded against row/record/column intents, which need
    # per-file work, not a file count.
    _FILE_COUNT_TRIGGERS = (
        "how many files", "how many data files", "how many csv",
        "how many csvs", "how many spreadsheets", "how many documents",
        "how many json", "number of files", "count of files",
        "count the files", "file count", "total files", "how many files are",
    )
    if (any(t in qlower for t in _FILE_COUNT_TRIGGERS)
            and not any(x in qlower for x in
                        ("row", "record", "column", "value", "cell"))):
        try:
            _csub = _va.resolve_subfolder_hint(query, allowed_folders[0])
        except Exception:
            _csub = None
        _ctarget = _csub if _csub is not None else allowed_folders[0]
        _cscope = (f" — scope: {_csub.relative_to(allowed_folders[0])}"
                   if _csub is not None else "")
        try:
            counts = _va.folder_file_counts(_ctarget)
        except Exception as _ce_exc:
            print('[analyst] folder_file_counts failed: ' + repr(_ce_exc),
                  file=_sys_dbg.stderr)
            counts = None
        if counts is not None:
            by_ext = counts.get("by_ext", {})
            breakdown = ", ".join(
                f"{n} {ext}" for ext, n in
                sorted(by_ext.items(), key=lambda kv: -kv[1])) or "no files"
            folder_name = (str(_csub.relative_to(allowed_folders[0]))
                           if _csub is not None else "data_in")
            block = (f"[ANALYST RESULT — file count{_cscope}]\n"
                     f"# Direct census (no model code-gen, no file reads).\n"
                     f"{folder_name} contains {counts.get('total', 0)} file(s) "
                     f"across {counts.get('folders', 0)} subfolder(s).\n"
                     f"By type: {breakdown}")
            notices.append(
                f"Analyst direct-routed to file count"
                + (f" on subfolder {folder_name}" if _csub is not None else "")
                + f" — {counts.get('total', 0)} file(s).")
            notices.append(
                "__ANALYST_ANSWER__:"
                f"{folder_name} contains {counts.get('total', 0)} file(s) "
                f"across {counts.get('folders', 0)} subfolder(s).\n"
                f"By type: {breakdown}")
            return block, None, notices
        # Census failed → fall through to the model path.

    if any(phrase in qlower for phrase in _DIRECT_SUMMARY_TRIGGERS):
        # Honour subfolder hints exactly like the model path would
        # (we resolve scope_folder below for the model branch too).
        try:
            sub = _va.resolve_subfolder_hint(query, allowed_folders[0])
        except Exception:
            sub = None
        target_folders = [sub] if sub is not None else allowed_folders
        scope_str = (f" — scope: {sub.relative_to(allowed_folders[0])}"
                      if sub is not None else "")

        # Query-report cache: an identical summary over an UNCHANGED set
        # of files is served from disk instantly. The key fingerprints
        # each input file's mtime, so adding / editing / removing a file
        # changes the key and forces a fresh compute — correct staleness.
        _qc = None
        _qkey = None
        _csv_paths = None
        try:
            import stats_cache as _sc
            _qc = _sc.QueryReportCache(VAULT_DIR)
            # Keep the raw Path list so the helper below doesn't walk for CSVs
            # a second time (it's the same file set that built the cache key).
            _csv_paths = _va.list_csv_files(target_folders)
            _inputs = [str(p) for p in _csv_paths]
            _qkey = _qc.make_key("folder_data_summary" + scope_str, _inputs)
        except Exception:
            _qc = None

        rep = None
        _hit = _qc.get(_qkey) if (_qc and _qkey) else None
        if _hit is not None:
            rep = _hit.get("report")

        if rep is None:
            try:
                result_df = _va.folder_data_summary(
                    target_folders, csv_files=_csv_paths)
            except Exception as _e:
                print('[analyst] direct folder_data_summary failed: '
                      + repr(_e), file=_sys_dbg.stderr)
                result_df = None
            if result_df is not None and not result_df.empty:
                try:
                    _n_ctx_ds = ce.get_n_ctx()
                except Exception:
                    _n_ctx_ds = 4096
                ds_max_tokens = max(150, _n_ctx_ds // 4)
                table_text = _va.format_result_for_prompt(
                    result_df, max_rows=250, max_chars=12000,
                    max_tokens=ds_max_tokens, count_tokens=ce.estimate_tokens,
                )
                try:
                    _user_render = result_df.to_string(index=False,
                                                        max_rows=80, max_cols=20)
                except Exception:
                    _user_render = ""
                rep = {"table_text": table_text, "user_render": _user_render,
                       "n_files": int(len(result_df))}
                if _qc and _qkey:
                    try:
                        _qc.put(_qkey, query="folder_data_summary" + scope_str,
                                report=rep)
                    except Exception:
                        pass

        if rep is not None:
            block = (f"[ANALYST RESULT — folder_data_summary{scope_str}]\n"
                     f"# Direct call (no model code-gen). One row per file.\n"
                     f"{rep.get('table_text', '')}")
            notices.append(
                f"Analyst direct-routed to folder_data_summary"
                + (f" on subfolder {sub.relative_to(allowed_folders[0])}"
                   if sub is not None else "")
                + f" — {rep.get('n_files', 0)} file(s) profiled"
                + (" (cached)" if _hit is not None else "") + "."
            )
            # Surface the table to the transcript via the same
            # __ANALYST_TABLE__ payload the model path uses.
            if rep.get("user_render"):
                notices.append("__ANALYST_TABLE__:" + rep["user_render"])
            notices.append(
                "__ANALYST_ANSWER__:"
                f"Here is a per-file data summary of {rep.get('n_files', 0)} "
                f"file(s) — one row per file (see the table above).")
            return block, None, notices
        # Fall through to the model path if the direct call returned
        # nothing (empty folder, hint resolved to non-existent path, etc).

    # ── Direct-route: STATS summary over a folder ───────────────────
    # "summary of stats", "column statistics", "min/max/mean of the
    # files" etc. used to fall through to the MODEL, which would write
    # arbitrary pandas — and over 200+ CSVs a `pd.concat([read_csv(f)
    # for f in files])` blows memory and crashes the app. Route these
    # to folder_column_stats(), which reads ONLY from the precomputed
    # stats cache (streaming compute on a miss, one file at a time) so
    # peak memory is bounded no matter how many files there are.
    _STATS_SUMMARY_TRIGGERS = (
        "summary of stats", "summary of the stats", "stats summary",
        "statistics summary", "summary statistics", "statistical summary",
        "column stats", "column statistics", "stats of the files",
        "stats for the files", "statistics of the files",
        "statistics for the files", "stats of the folder",
        "stats on the files", "min max mean", "min/max/mean",
        "summarize the stats", "summarise the stats", "summarize stats",
        "give me stats", "compute stats", "calculate stats",
        "stats across", "statistics across",
    )
    if any(phrase in qlower for phrase in _STATS_SUMMARY_TRIGGERS):
        try:
            _ssub = _va.resolve_subfolder_hint(query, allowed_folders[0])
        except Exception:
            _ssub = None
        _starget = [_ssub] if _ssub is not None else allowed_folders
        _sscope = (f" — scope: {_ssub.relative_to(allowed_folders[0])}"
                   if _ssub is not None else "")

        # Query-report cache, fingerprinted on each input file's mtime —
        # an unchanged folder serves the prior stats table instantly.
        _sqc = _sqkey = None
        _scsv_paths = None
        try:
            import stats_cache as _sc2
            _sqc = _sc2.QueryReportCache(VAULT_DIR)
            # Reuse this CSV Path list below instead of re-walking for it.
            _scsv_paths = _va.list_csv_files(_starget)
            _sinputs = [str(p) for p in _scsv_paths]
            _sqkey = _sqc.make_key("folder_column_stats" + _sscope, _sinputs)
        except Exception:
            _sqc = None

        srep = None
        _shit = _sqc.get(_sqkey) if (_sqc and _sqkey) else None
        if _shit is not None:
            srep = _shit.get("report")

        if srep is None:
            try:
                _sdf = _va.folder_column_stats(
                    VAULT_DIR, _starget, csv_files=_scsv_paths)
            except Exception as _se:
                print('[analyst] direct folder_column_stats failed: '
                      + repr(_se), file=_sys_dbg.stderr)
                _sdf = None
            if _sdf is not None and not _sdf.empty:
                try:
                    _n_ctx_st = ce.get_n_ctx()
                except Exception:
                    _n_ctx_st = 4096
                _st_max_tokens = max(150, _n_ctx_st // 4)
                _stable = _va.format_result_for_prompt(
                    _sdf, max_rows=400, max_chars=12000,
                    max_tokens=_st_max_tokens, count_tokens=ce.estimate_tokens,
                )
                try:
                    _srender = _sdf.to_string(index=False,
                                              max_rows=120, max_cols=20)
                except Exception:
                    _srender = ""
                _nfiles = int(_sdf["file"].nunique()) if "file" in _sdf else 0
                srep = {"table_text": _stable, "user_render": _srender,
                        "n_files": _nfiles}
                if _sqc and _sqkey:
                    try:
                        _sqc.put(_sqkey,
                                 query="folder_column_stats" + _sscope,
                                 report=srep)
                    except Exception:
                        pass

        if srep is not None:
            block = (f"[ANALYST RESULT — folder_column_stats{_sscope}]\n"
                     f"# Direct call (no model code-gen). Stats served from "
                     f"the precomputed cache — one row per (file, column).\n"
                     f"{srep.get('table_text', '')}")
            notices.append(
                "Analyst direct-routed to folder_column_stats"
                + (f" on subfolder {_ssub.relative_to(allowed_folders[0])}"
                   if _ssub is not None else "")
                + f" — stats for {srep.get('n_files', 0)} file(s)"
                + (" (cached)" if _shit is not None else "") + "."
            )
            if srep.get("user_render"):
                notices.append("__ANALYST_TABLE__:" + srep["user_render"])
            notices.append(
                "__ANALYST_ANSWER__:"
                f"Here are per-column statistics across "
                f"{srep.get('n_files', 0)} file(s) — one row per "
                f"(file, column) (see the table above).")
            return block, None, notices
        # Empty result → fall through to the model path.

    # ── Upgrade A+B: pre-resolve filename + subfolder hints ──────────
    # Before sending the prompt to the model, look for explicit
    # filename references ("sales.csv", quoted strings) and folder
    # references ("in the test_data folder") in the user's question.
    # Resolve them against the actual on-disk inventory so the model
    # sees the real paths instead of guessing — guesses are how
    # FileNotFoundError / wrong-file-grabbed failures used to creep in.
    base_folder = allowed_folders[0] if allowed_folders else None
    scope_folder = None
    if base_folder is not None:
        try:
            scope_folder = _va.resolve_subfolder_hint(query, base_folder)
        except Exception as _e:
            print('[analyst] subfolder resolve failed: ' + repr(_e),
                  file=_sys_dbg.stderr)
    if scope_folder is not None:
        allowed_folders = [scope_folder]
        try:
            rel = scope_folder.relative_to(base_folder) if base_folder else scope_folder
        except Exception:
            rel = scope_folder
        notices.append(f"Analyst scoped to subfolder: {rel}")
        print('[analyst] scope restricted to: ' + str(scope_folder),
              file=_sys_dbg.stderr)

    filename_hints_pairs = []
    try:
        filename_hints_pairs = _va.resolve_filename_hints(query, allowed_folders)
    except Exception as _e:
        print('[analyst] filename hint resolve failed: ' + repr(_e),
              file=_sys_dbg.stderr)
    filename_hints_text = ""
    if filename_hints_pairs:
        # Build a one-line summary for the transcript so the user can spot
        # a wrong match before reading the answer. Suppress trivial
        # resolutions where the token == filename — those look like
        # typos to the user ("Filename hints: 'sales.csv' → sales.csv")
        # and add no information. Only surface non-trivial resolutions
        # and "no match" cases.
        bits: list = []
        for tok, resolved in filename_hints_pairs:
            if resolved is None:
                bits.append(f"'{tok}' → no match")
            elif tok.lower() != resolved.name.lower():
                # Non-trivial resolution — the user said one thing, the
                # resolver picked something different. This is the case
                # they need to see.
                bits.append(f"'{tok}' → {resolved.name}")
            # else: exact match between user token and filename → silent
        if bits:
            notices.append("Filename hints: " + "; ".join(bits))

        filename_hints_text = _va.format_filename_hints(
            filename_hints_pairs,
            base_folder=allowed_folders[0] if allowed_folders else None,
        )
        print('[analyst] filename hints:\n' + filename_hints_text,
              file=_sys_dbg.stderr)

    inventory = _va.preview_csv_inventory(allowed_folders, max_files=15, max_cols=30)
    prompt = _va.build_pandas_code_prompt(
        query, allowed_folders, inventory,
        filename_hints=filename_hints_text or None,
        subfolder_scope=scope_folder,
    )

    try:
        import council_engine as _ce
        # Backend-agnostic helper — works in both Ollama and GGUF modes.
        with _ce._TimingScope("analyst.codegen"):
            raw = _ce.local_chat(
                messages=[{'role': 'user', 'content': prompt}],
                temperature=0.0,
                num_predict=600,
                timeout=90,
            )
    except Exception as _e:
        msg = f"code generation failed: {_e!r}"
        print('[analyst] ' + msg, file=_sys_dbg.stderr)
        _record_app_failure("analyst.codegen_error", "council_gui_engine",
                            msg, context={"query": query[:200]})
        return _build_analyst_failure_block("(no code generated)", msg), msg, notices

    code = _va.extract_python_code(raw)
    if not code.strip():
        msg = "the model produced no executable pandas code"
        print('[analyst] empty code from model', file=_sys_dbg.stderr)
        _record_app_failure("analyst.empty_code", "council_gui_engine",
                            msg, context={"query": query[:200]})
        return _build_analyst_failure_block("(empty)", msg), msg, notices
    print('[analyst] generated code (first 300):\n' + code[:300], file=_sys_dbg.stderr)

    # Hand the analyst a way to BUILD its own tools when a capability is
    # missing. save_app_tool validates via the sandbox and persists to the
    # vault's App_Built_tools/ (UNREVIEWED); run/list let it reuse them. These
    # are curated app callables — NOT passed to app-built tools themselves, so
    # a tool can't spawn more tools.
    def _save_app_tool(name, description, code):
        import app_built_tools as _abt
        ok, msg, _ = _abt.save_tool(name, description, code,
                                    author="council", vault_dir=VAULT_DIR)
        return msg

    def _run_app_tool(name, args=None):
        import app_built_tools as _abt
        _df, _msg = _abt.run_tool(name, args or {},
                                  allowed_folders=allowed_folders,
                                  vault_dir=VAULT_DIR)
        return _df if _df is not None else _msg

    def _list_app_tools():
        import app_built_tools as _abt
        return _abt.list_tools(vault_dir=VAULT_DIR)

    _analyst_extra = {
        "save_app_tool": _save_app_tool,
        "run_app_tool":  _run_app_tool,
        "list_app_tools": _list_app_tools,
    }
    with _ce._TimingScope("analyst.exec"):
        result_df, log = _va.execute_pandas_code(
            code, allowed_folders, extra_globals=_analyst_extra)
    if result_df is None:
        # Extract the first traceback line for a concise transcript message.
        # The full log keeps the entire traceback for debugging in the block.
        first_err = _summarise_analyst_error(log)
        print('[analyst] exec failed: ' + log[:400], file=_sys_dbg.stderr)
        _record_app_failure("analyst.exec_error", "vault_analyst",
                            first_err, detail=log,
                            context={"query": query[:200]})
        return _build_analyst_failure_block(code, log), first_err, notices

    # Budget-scaled rendering. The analyst block is exempt from the
    # per-block cap (UNCAPPED_PRIOS) and below DROPPABLE_FROM, but
    # giving it 12 KB worth of rows on a 4 K-ctx machine still eats the
    # whole window. Cap at 25 % of n_ctx so a 4 K window gets ~1 K
    # tokens for the analyst result and a 32 K window gets ~8 K.
    # The renderer keeps head + tail and elides the middle, which
    # preserves the summary rows pandas appends (Total / Mean / …)
    # that the model needs to give a correct answer.
    try:
        _n_ctx_a = ce.get_n_ctx()
    except Exception:
        _n_ctx_a = 4096
    analyst_max_tokens = max(150, _n_ctx_a // 4)
    table_text = _va.format_result_for_prompt(
        result_df,
        max_rows=250, max_chars=12000,
        max_tokens=analyst_max_tokens,
        count_tokens=ce.estimate_tokens,
    )

    # Build a separate USER-FACING render of the full result (or as
    # much as fits in a transcript box) AND save the complete DataFrame
    # to vault/analyst_results/ so nothing is ever silently lost. The
    # caller posts the transcript render to the chat AND mentions the
    # saved-file path so the user can re-open the result later.
    user_table_lines: list = []
    full_table_path = None
    try:
        nrows = len(result_df)
        ncols = len(result_df.columns)
        user_table_lines.append(
            f"Analyst result: {nrows:,} row{'s' if nrows != 1 else ''} "
            f"× {ncols} column{'s' if ncols != 1 else ''}"
        )
        # Render up to 500 rows directly into the transcript. Beyond
        # that, save a CSV and tell the user where it is — Tk text
        # widgets get sluggish above a few thousand rows of text.
        TRANSCRIPT_DISPLAY_ROWS = 500
        if nrows <= TRANSCRIPT_DISPLAY_ROWS:
            user_table_lines.append(
                result_df.to_string(index=False, max_colwidth=80)
            )
        else:
            user_table_lines.append(
                result_df.head(TRANSCRIPT_DISPLAY_ROWS)
                          .to_string(index=False, max_colwidth=80)
            )
            user_table_lines.append(
                f"\n... ({nrows - TRANSCRIPT_DISPLAY_ROWS:,} more rows "
                f"omitted from this transcript view; full table saved "
                f"to file — see path below.)"
            )

        # Save the complete DataFrame as CSV under
        # vault/analyst_results/<timestamp>_<rows>x<cols>.csv. Users
        # can open it from disk to verify or share the full data.
        try:
            from datetime import datetime as _dt
            stamp = _dt.now().strftime("%Y%m%d_%H%M%S")
            results_dir = VAULT_DIR / "analyst_results"
            results_dir.mkdir(parents=True, exist_ok=True)
            full_table_path = results_dir / f"{stamp}_{nrows}x{ncols}.csv"
            result_df.to_csv(full_table_path, index=False)
            user_table_lines.append(
                f"\nFull table saved to: {full_table_path}"
            )
        except Exception as _save_exc:
            print(f"[analyst] could not save full table: {_save_exc!r}",
                  file=_sys_dbg.stderr)
            full_table_path = None
    except Exception as _render_exc:
        print(f"[analyst] could not build user-facing table: "
              f"{_render_exc!r}", file=_sys_dbg.stderr)
        user_table_lines = []
    user_table = "\n".join(user_table_lines) if user_table_lines else ""

    if user_table:
        notices.append("__ANALYST_TABLE__:" + user_table)

    block = (
        '[ANALYST RESULT — computed from real CSV data]\n'
        + 'pandas code:\n' + code.strip() + '\n\n'
        + 'output:\n' + table_text + '\n'
        + '[END ANALYST]'
    )
    return block, None, notices


def _summarise_analyst_error(log: str) -> str:
    """Extract a short human-readable summary from a Python traceback.

    Returns the last `ExceptionType: message` line if present (most
    useful for the user), else the first 200 chars of the log.
    """
    if not log:
        return "analyst execution failed"
    # Find the final "ExceptionType: message" line in the traceback.
    last = None
    for line in log.splitlines():
        line = line.strip()
        if line and ":" in line and not line.startswith(("File ", "  ", "Traceback")):
            # Plausible exception line — keep the latest one.
            last = line
    if last:
        return last[:200]
    return log.strip().splitlines()[0][:200] if log.strip() else "analyst failed"


def _record_app_failure(kind: str, subsystem: str, message: str,
                        detail: str = "",
                        context: "Optional[dict]" = None) -> None:
    """Best-effort failure capture into the self-improvement loop.

    Forwards to agent_logs.record_failure (append-only JSONL in the
    vault). The FailureAnalyzer aggregates recurring signatures into
    human-reviewed improvement proposals in the Agent panel. Import-
    guarded and exception-proof so the failing code path is never made
    worse by the act of recording it."""
    try:
        import agent_logs as _al
        _al.record_failure(kind, subsystem, message,
                           detail=detail, context=context)
    except Exception:
        pass


def _build_analyst_failure_block(code: str, error_log: str) -> str:
    """Build an [ANALYST RESULT — FAILED] block that tells the model the
    computation was attempted but failed, and explicitly refuses any
    invented numeric answer. This is the critical anti-hallucination
    fallback: without it, the model would silently produce a freeform
    answer that looks confident but invents numbers from training data.
    """
    code_part = code.strip() or "(no code)"
    err_part = (error_log or "").strip() or "(no traceback captured)"
    if len(err_part) > 1200:
        err_part = err_part[:1200] + "\n... (truncated)"
    return (
        "[ANALYST RESULT — COMPUTATION FAILED]\n"
        "The analyst attempted a deterministic pandas calculation but the "
        "sandbox raised an error. The numeric answer the user asked for "
        "is NOT AVAILABLE.\n\n"
        "pandas code attempted:\n"
        + code_part + "\n\n"
        "error:\n"
        + err_part + "\n\n"
        "[INSTRUCTION TO THE WRITER: Do NOT invent a numeric answer to "
        "compensate. Tell the user the deterministic computation failed, "
        "quote the error type briefly, and suggest they rephrase the "
        "question or check that the file/column they mentioned exists. "
        "Do NOT pull values from training-data memory of similarly-named "
        "datasets.]\n"
        "[END ANALYST]"
    )


_CODE_ROUTES = {"ide", "coder", "intern"}

def _filter_final(text: str, route: str, user_query: str) -> str:
    """
    Strip code blocks from the council's final answer if:
    - The route is not a code route, AND
    - The user did not explicitly ask for code.
    Applied after Writer synthesis, before the result reaches the UI.
    """
    if _user_wants_code(user_query):
        return text                          # user asked for code explicitly
    stripped = _strip_code_blocks(text)
    if stripped != text:
        # Append a soft note so the user knows they can ask for code
        stripped = stripped.rstrip() + (
            "\n\n*(Ask me to write or modify a script if you need code.)*"
            if len(text) - len(stripped) > 80 else ""
        )
    return stripped
# ─────────────────────────────────────────────────────────────────────────────

def _detect_latex_request(text: str) -> bool:
    """Return True if the user explicitly requested LaTeX output."""
    t = text.lower()
    _latex_phrases = [
        "in latex", "as latex", "latex document", "latex format",
        "write latex", "write in latex", "using latex", "latex output",
        "as a latex", "latex file", "tex file", "write a latex",
        "latex report", "latex essay", "latex paper",
    ]
    return any(ph in t for ph in _latex_phrases)


def _wrap_latex(title: str, body: str) -> str:
    """Wrap plain text content in a minimal LaTeX document."""
    import re as _re
    # Escape common LaTeX special chars in the body
    _escapes = {
        "&": r"\&", "%": r"\%", "$": r"\$", "#": r"\#",
        "_": r"\_", "{": r"\{", "}": r"\}", "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    safe = body
    for ch, esc in _escapes.items():
        safe = safe.replace(ch, esc)
    # Convert markdown-ish bold **text** → \textbf{text}
    safe = _re.sub(r"\*\*(.+?)\*\*", r"\\textbf{\1}", safe)
    # Preserve line breaks as paragraph breaks
    safe = safe.replace("\n\n", "\n\n\\medskip\n\n")

    return (
        "\\documentclass[12pt,a4paper]{article}\n"
        "\\usepackage[utf8]{inputenc}\n"
        "\\usepackage[T1]{fontenc}\n"
        "\\usepackage{lmodern}\n"
        "\\usepackage{microtype}\n"
        "\\usepackage[margin=2.5cm]{geometry}\n"
        "\\usepackage{parskip}\n"
        "\\title{" + title + "}\n"
        "\\author{Council AI}\n"
        "\\date{\\today}\n"
        "\\begin{document}\n"
        "\\maketitle\n\n"
        + safe + "\n\n"
        "\\end{document}\n"
    )


def _detect_user_question(text: str) -> str:
    """
    Check if a personality response contains a direct question aimed at the user
    (not rhetorical, not Peasant-style cross-examination).
    Returns the question sentence if found, empty string otherwise.
    """
    import re as _re
    # Look for lines that end with ? and contain user-directed phrasing
    _user_markers = [
        "could you", "can you", "would you", "do you", "what is your",
        "what are your", "please clarify", "please provide", "i need to know",
        "could you clarify", "could you provide", "could you tell",
        "what do you mean", "what exactly", "which do you prefer",
        "which would you", "how do you want", "what would you like",
        "do you have", "do you want", "are you looking for",
    ]
    sentences = _re.split(r"(?<=[.!?])\s+", text)
    for s in sentences:
        s_low = s.lower().strip()
        if s.strip().endswith("?") and any(m in s_low for m in _user_markers):
            return s.strip()
    return ""


def _panel_for_route(route: str) -> tuple:
    """Return (panel_list, synth_role) for the given judge route."""
    return _PANEL_FOR_ROUTE.get(route, _PANEL_FOR_ROUTE["_default"])


def peasant_cross_exam(
    peasant_model,
    *,
    candidate_role: str,
    candidate_text: str,
    user_text: str,
    prior_qa: Optional[List[Dict[str, str]]] = None,
    query_mode: str = "",
) -> str:
    """
    Cross-examine a candidate response as the Peasant.

    Passes the code/answer as extra_context so the full Peasant system
    prompt applies — meaning questions are specific to THIS code, not generic.

    prior_qa is a list of {"q": "...", "a": "..."} dicts accumulated across
    all earlier Peasant turns in this deliberation.  When present they are
    injected as a hard DO-NOT-REPEAT block so the Peasant cannot recycle
    questions that have already been asked and answered.
    """
    has_code = any(marker in candidate_text for marker in
                   ["def ", "class ", "import ", "```", "for ", "while ", "if "])
    content_label = "CODE" if has_code else "RESPONSE"

    parts = [
        f"ORIGINAL REQUEST:\n{user_text}\n",
        f"{candidate_role.upper()} {content_label} TO REVIEW:\n{candidate_text}\n",
    ]

    if prior_qa:
        lines = [
            "━━━ QUESTIONS YOU HAVE ALREADY ASKED THIS SESSION ━━━",
            "Do NOT ask any of these again, even in paraphrased form.",
            "Do NOT ask questions whose answers are already contained below.",
            "",
        ]
        for i, item in enumerate(prior_qa, 1):
            lines.append(f"[{i}] Q: {item['q']}")
            if item.get("a"):
                lines.append(f"    A: {item['a']}")
            lines.append("")
        lines.append("━━━ END OF PRIOR Q&A ━━━")
        parts.append("\n".join(lines))

    _mode_instruction = ""
    if query_mode == "conversational":
        _mode_instruction = (
            "⚠ CONVERSATIONAL MODE: The user asked a conversational question, not for code.\n"
            "Do NOT ask about error handling, imports, types, or code structure.\n"
            "Ask whether the explanation is accurate, clear, complete, and actually answers "
            "what the user asked.\n\n"
        )
    elif query_mode == "technical":
        _mode_instruction = (
            "⚠ TECHNICAL MODE: Focus on code correctness, edge cases, and robustness.\n\n"
        )
    if _mode_instruction:
        parts.append(_mode_instruction)

    parts.append(
        "Your task: identify NEW specific problems, edge cases, or dangerous assumptions "
        "in the above that have NOT already been raised. Ask questions tied to specific "
        "lines, variable names, or behaviours you can see in this exact code — "
        "not generic questions, and not anything already covered above."
    )

    extra_context = "\n\n".join(parts)

    prompt = (
        f"Review the {candidate_role} {content_label.lower()} above and ask your NEW questions now. "
        "Every question must reference something specific you can see in that code, "
        "and must not duplicate any question from the prior Q&A list above."
    )

    return peasant_model.respond(prompt, extra_context=extra_context)


def _looks_like_two_questions(text: str) -> bool:
    """
    Returns True if the Peasant response looks like it followed the format.
    Accepts either the old Q1/Q2 format or the new format with DANGEROUS ASSUMPTION.
    """
    t = text.lower()
    has_questions = ("q1:" in t) and ("q2:" in t)
    # Also accept if model gave specific feedback even without strict Q1/Q2 labels
    has_question_marks = t.count("?") >= 2
    return has_questions or (has_question_marks and len(text) > 80)


def _peasant_quality_score(
    text: str,
    candidate_text: str,
    prior_qa: Optional[List[Dict[str, str]]] = None,
) -> Dict[str, Any]:
    """
    Score Peasant output on four axes; return dict with 0-4 total.

    Axes:
      questions  -- at least 2 question marks present
      length     -- at least 80 chars
      specific   -- references at least 3 words from the candidate text
      non_repeat -- no prior question shares >60% 4-gram overlap
    """
    import re as _re
    scores: Dict[str, bool] = {}

    scores["questions"] = text.count("?") >= 2
    scores["length"]    = len(text.strip()) >= 80

    cand_words = set(w.lower() for w in _re.findall(r"\w{4,}", candidate_text))
    resp_words = set(w.lower() for w in _re.findall(r"\w{4,}", text))
    scores["specific"] = len(cand_words & resp_words) >= 3

    if prior_qa:
        def _ngrams(s: str, n: int = 4) -> set:
            ws = s.lower().split()
            return set(tuple(ws[i:i+n]) for i in range(len(ws) - n + 1))
        resp_ng = _ngrams(text)
        overlap_ok = True
        for qa in prior_qa:
            prev_ng = _ngrams(qa.get("q", ""))
            if prev_ng and resp_ng:
                overlap = len(resp_ng & prev_ng) / max(len(prev_ng), 1)
                if overlap > 0.6:
                    overlap_ok = False
                    break
        scores["non_repeat"] = overlap_ok
    else:
        scores["non_repeat"] = True

    total = sum(scores.values())
    return {"total": total, "max": 4, "axes": scores}


# ============================================================
# ModelAgent  (with streaming token callback)
# ============================================================

class ModelAgent:
    def __init__(
        self,
        display_name: str,
        personality_model: Any,
        *,
        tools: Dict[str, ToolFn] | None = None,
        enable_tools: bool = False,
        max_tool_steps: int = 3,
        token_callback: Optional[Callable[[str, str], None]] = None,
    ):
        self.display_name = display_name
        self.model = personality_model
        self.tools = tools or {}
        self.enable_tools = enable_tools
        self.max_tool_steps = max_tool_steps
        # token_callback(who, token) — called for each streamed token
        self.token_callback = token_callback

    def _compose_prompt(self, ctx: AgentContext) -> str:
        parts: List[str] = []

        # Inject query mode so every personality knows what type of response to give.
        # This overrides the code-centric defaults baked into each system prompt.
        _mode = ctx.shared.get("query_mode", "")
        if _mode == "conversational":
            parts.append(
                "QUERY MODE: CONVERSATIONAL\n"
                "The user is having a conversation — NOT asking for code.\n"
                "Respond entirely in natural prose. Do NOT write code, scripts, or "
                "technical implementations unless the user explicitly asked for them.\n"
                "Focus on explaining, discussing, or answering the question directly."
            )
        elif _mode == "technical":
            parts.append(
                "QUERY MODE: TECHNICAL\n"
                "The user wants working code or a technical solution.\n"
                "Prioritise correctness, completeness, and runnability."
            )

        cands = ctx.shared.get("candidates", {})
        rank = ctx.shared.get("judge_ranking", "")
        critique = ctx.shared.get("judge_critique", "")
        tool_payloads = ctx.shared.get("tool_payloads", {})
        discussion = ctx.shared.get("discussion_transcript", "")

        if cands:
            parts.append("CANDIDATE ANSWERS + PEASANT QUESTIONS + REBUTTALS:")
            for role, data in cands.items():
                parts.append(f"--- {role} ---")
                parts.append("ANSWER:")
                parts.append(data.get("answer", ""))
                if pq := data.get("peasant_q", ""):
                    parts.append(f"PEASANT QUESTIONS:\n{pq}")
                if rb := data.get("rebuttal", ""):
                    parts.append(f"REBUTTAL:\n{rb}")
                if disc := data.get("discussion", ""):
                    parts.append(f"DISCUSSION LOG:\n{disc}")
                parts.append("")

        if discussion:
            parts.append(f"FULL DISCUSSION (condensed):\n{discussion}\n")
        if rank:
            # Inject explicit winner label — models respond far better to a clear
            # directive than to having to parse the winner out of embedded JSON.
            try:
                import json as _cjson
                _robj = _cjson.loads(rank) if isinstance(rank, str) else rank
                _winner = _robj.get("winner", "")
                if _winner and _winner != "unknown":
                    parts.append(f"WINNING CANDIDATE: {_winner} — synthesise primarily from this answer.")
            except Exception:
                pass
            parts.append(f"JUDGE RANKING (JSON):\n{rank}\n")
        if critique:
            parts.append(f"JUDGE CRITIQUE:\n{critique}\n")
        required_changes = ctx.shared.get("required_changes", [])
        if required_changes:
            parts.append("REQUIRED CHANGES -- YOU MUST ADDRESS EVERY ITEM BELOW:")
            for _rc in required_changes:
                parts.append("  - " + _rc)
            parts.append("")
        adv_challenge = ctx.shared.get("adversarial_challenge", "")
        if adv_challenge:
            parts.append(
                "ADVERSARIAL CHALLENGE (you MUST explicitly rebut this in your answer):\n"
                + adv_challenge + "\n"
            )
        if tool_payloads:
            parts.append("PRIOR TOOL OUTPUTS:")
            for k, v in tool_payloads.items():
                parts.append(f"- {k}: {str(v)[:900]}")
            parts.append("")

        # Repeat mode reminder as the LAST thing before user request.
        # Models anchor to recency — the instruction at the bottom wins over code seen above.
        _mode_bottom = ctx.shared.get("query_mode", "")
        if _mode_bottom == "conversational":
            parts.append(
                "⚠ REMINDER: This is a CONVERSATIONAL query. "
                "Do NOT write code. Respond in prose only. "
                "Ignore any code in the candidate answers above — it should not have been there."
            )
        elif _mode_bottom == "technical":
            parts.append(
                "⚠ REMINDER: This is a TECHNICAL query. "
                "Prioritise working, complete code."
            )

        parts.append(f"USER REQUEST:\n{ctx.user_text}")

        if self.enable_tools and self.tools:
            tool_list = ", ".join(sorted(self.tools.keys()))
            parts += [
                "",
                f"TOOLS AVAILABLE: {tool_list}",
                "To use a tool, output ONLY JSON: {\"tool\":\"name\",\"args\":{...}}",
                "Otherwise write a normal answer.",
            ]
        return "\n".join(parts)

    def _make_token_cb(self) -> Optional[Callable[[str], None]]:
        if self.token_callback is None:
            return None
        who = self.display_name
        cb = self.token_callback
        def _cb(token: str):
            cb(who, token)
        return _cb

    def act(self, ctx: AgentContext) -> List[AgentEvent]:
        events: List[AgentEvent] = []
        prompt = self._compose_prompt(ctx)

        if not (self.enable_tools and self.tools):
            events.append(AgentEvent(self.display_name, "thought", "Generating response…"))
            text = self.model.respond(prompt, token_callback=self._make_token_cb())
            return [AgentEvent(self.display_name, "final", text)]

        events.append(AgentEvent(self.display_name, "thought", "Calling model backend…"))
        text = self.model.respond(prompt, token_callback=self._make_token_cb())

        for _ in range(self.max_tool_steps):
            calls = _extract_tool_calls(text)
            if not calls:
                events.append(AgentEvent(self.display_name, "final", text))
                return events

            events.append(AgentEvent(self.display_name, "action", f"Tool calls requested ({len(calls)})."))
            obs_lines: List[str] = []
            payloads: Dict[str, Any] = {}

            for i, call in enumerate(calls, start=1):
                tool_name = str(call.get("tool", "")).strip()
                args = call.get("args", {})
                if tool_name not in self.tools:
                    obs_lines.append(f"[{i}] ERROR: unknown tool '{tool_name}'")
                    continue
                if not isinstance(args, dict):
                    obs_lines.append(f"[{i}] ERROR: args must be a dict")
                    continue
                ok, msg, payload = self.tools[tool_name](args)
                obs_lines.append(f"[{i}] {tool_name}: {'OK' if ok else 'FAIL'}\n{msg}")
                if payload:
                    payloads[f"{tool_name}_{i}"] = payload

            ctx.shared.setdefault("tool_payloads", {}).update(payloads)
            obs_text = "\n\n".join(obs_lines).strip() or "(no tool output)"
            events.append(AgentEvent(self.display_name, "observation", obs_text))

            followup = (
                f"TOOL RESULTS:\n{obs_text}\n\n"
                "Now produce the best possible answer (no tool JSON unless more tools needed)."
            )
            events.append(AgentEvent(self.display_name, "thought", "Calling model (post-tool)…"))
            text = self.model.respond(followup, token_callback=self._make_token_cb())

        events.append(AgentEvent(self.display_name, "final", text))
        return events


# ============================================================
# DeliberationOrchestrator  (yields events live via callback)
# ============================================================

class DeliberationOrchestrator:
    """
    Runs the full deliberation loop and emits events live through
    an `event_callback(AgentEvent)` as they occur.
    """

    def __init__(
        self,
        *,
        judge_model: Any,
        agents: Dict[str, ModelAgent],
        max_rounds: int = 2,
        debate_turns: int = 2,
        event_callback: Optional[Callable[[AgentEvent], None]] = None,
        clarification_cb: Optional[Callable[[str, str], None]] = None,
        pause_event: Optional[threading.Event] = None,
        answer_getter: Optional[Callable[[], str]] = None,
    ):
        self.judge = judge_model
        self.agents = agents
        self.max_rounds = max_rounds
        self.debate_turns = max(1, int(debate_turns))
        self.event_callback = event_callback or (lambda e: None)
        # Clarification pause support
        self._clarification_cb = clarification_cb   # fn(who, question) → shows UI
        self._pause_event      = pause_event         # threading.Event to wait on
        self._answer_getter    = answer_getter        # fn() → str answer

    def _emit(self, event: AgentEvent) -> None:
        self.event_callback(event)

    def _phase(self, label: str) -> None:
        self._emit(AgentEvent("Orchestrator", "phase", f"▶ {label}"))

    def run(self, user_text: str, *, panel: List[str], synth: str = "writer",
            extra_ctx: Optional[Dict[str, Any]] = None) -> List[AgentEvent]:
        ctx = AgentContext(user_text=user_text)
        if extra_ctx:
            ctx.shared.update(extra_ctx)
        all_events: List[AgentEvent] = []

        # Guard: synth must exist in agents — fall back to writer or first panel member
        if synth not in self.agents:
            synth = "writer" if "writer" in self.agents else (panel[0] if panel else synth)

        # Accumulates every Peasant question across all rounds and cross-fire
        # turns so the model is never shown a blank slate and cannot re-ask
        # something already covered.  Each entry: {"q": <text>, "a": ""}
        _peasant_qa_log: List[Dict[str, str]] = []

        def _log_peasant_questions(qtxt: str) -> None:
            import re as _re
            parts = _re.split(r"(?:^|\n)(?:Q\d+[:.)]|\d+[.)]\ +|\[\d+\]\ *)", qtxt)
            questions = [p.strip() for p in parts if p.strip() and "?" in p]
            if not questions:
                questions = [p.strip() for p in qtxt.split("\n\n") if "?" in p]
            if not questions:
                questions = [qtxt.strip()]
            for q in questions:
                _peasant_qa_log.append({"q": q, "a": ""})

        def emit(ev: AgentEvent):
            all_events.append(ev)
            self._emit(ev)

        for r in range(self.max_rounds):
            # ── Check pause at start of each round ──────────────────
            # If a clarification is pending, wait here before any
            # new model calls fire. This ensures the whole round
            # waits, not just the individual candidate step.
            if self._pause_event and not self._pause_event.is_set():
                self._pause_event.wait(timeout=300)

            self._phase(f"Round {r+1}/{self.max_rounds} — Candidate generation")

            candidates: Dict[str, Dict[str, str]] = {}
            discussion_lines: List[str] = []

            # 1) Candidates + Peasant cross-exam
            for key in panel:
                self._phase(f"{key.capitalize()} — drafting answer")
                evs = self.agents[key].act(ctx)
                for ev in evs:
                    emit(ev)
                answer = next((e.text for e in reversed(evs) if e.kind == "final"), "")
                # Strip code from candidate answers on conversational routes
                # so they don't contaminate what other panel members read.
                _qmode = ctx.shared.get("query_mode", "")
                _stored_answer = answer
                if _qmode == "conversational":
                    _stored_answer = _strip_code_blocks(answer)
                # ── #8 Self-reported confidence ─────────────────────────────
                # Ask each candidate to rate their own confidence 1-10.
                # A single cheap token call — models are usually well-calibrated
                # at distinguishing "I'm guessing" from "I'm certain".
                _self_conf = 5  # default if call fails
                try:
                    _conf_raw = self.agents[key].model.respond(
                        "Rate your confidence in the answer you just gave, 1–10. "
                        "Reply with ONLY the single digit — no words, no punctuation.\n\n"
                        f"YOUR ANSWER (first 400 chars):\n{answer[:400]}",
                        max_tokens=5,
                    ).strip()
                    _self_conf = int(_conf_raw[0]) if _conf_raw and _conf_raw[0].isdigit() else 5
                    _self_conf = max(1, min(10, _self_conf))
                except Exception:
                    pass
                candidates[key] = {
                    "answer": _stored_answer,
                    "peasant_q": "", "rebuttal": "", "discussion": "",
                    "self_confidence": _self_conf,
                }
                if _self_conf <= 4:
                    emit(AgentEvent(key.capitalize(), "observation",
                                   f"⚠ Self-confidence: {_self_conf}/10 — answer may be weak"))
                else:
                    emit(AgentEvent(key.capitalize(), "observation",
                                   f"Confidence: {_self_conf}/10"))
                discussion_lines.append(f"{key.upper()} CANDIDATE [conf:{_self_conf}/10]:\n{_stored_answer}\n")

                # ── Clarification pause ──────────────────────────────
                # If a non-Peasant personality asked the user a direct question,
                # pause deliberation and wait for the user to answer.
                if key != "peasant" and self._clarification_cb and self._pause_event:
                    _q = _detect_user_question(_stored_answer)
                    if _q:
                        self._pause_event.clear()  # pause
                        self._clarification_cb(key.capitalize(), _q)
                        # Block the worker thread until user answers (5 min max)
                        self._pause_event.wait(timeout=300)
                        _user_answer = self._answer_getter() if self._answer_getter else ""
                        if _user_answer and not _user_answer.startswith("[User skipped"):
                            _clarif_note = (f"\n\nUSER CLARIFICATION for {key}:\n"
                                           f"  Q: {_q}\n  A: {_user_answer}\n")
                            user_text = user_text + _clarif_note
                            ctx.user_text = user_text
                            discussion_lines.append(_clarif_note)

                if key != "peasant" and "peasant" in self.agents:
                    self._phase(f"Peasant — cross-examining {key}")
                    _pexam_mode = ctx.shared.get("query_mode", "")
                    qtxt = peasant_cross_exam(
                        self.agents["peasant"].model,
                        candidate_role=key, candidate_text=answer, user_text=user_text,
                        prior_qa=_peasant_qa_log if _peasant_qa_log else None,
                        query_mode=_pexam_mode,
                    )
                    _pq_score = _peasant_quality_score(qtxt, answer, _peasant_qa_log)
                    if not _looks_like_two_questions(qtxt):
                        # Reformat existing answer rather than full regeneration — cheaper
                        _reformat_prompt = (
                            "Your response below is good but needs exactly two questions "
                            "labelled Q1: and Q2:. Reformat it now — keep the same ideas, "
                            "just add Q1: and Q2: labels and make sure each ends with '?'.\n\n"
                            f"YOUR RESPONSE:\n{qtxt}"
                        )
                        qtxt = self.agents["peasant"].model.respond(
                            _reformat_prompt, max_tokens=300)
                        _pq_score = _peasant_quality_score(qtxt, answer, _peasant_qa_log)
                        if not _looks_like_two_questions(qtxt):
                            _axes = ", ".join(
                                k + ("=✓" if v else "=✗")
                                for k, v in _pq_score["axes"].items()
                            )
                            emit(AgentEvent("Peasant", "observation",
                                "⚠ Quality low after reformat ("
                                + str(_pq_score["total"]) + "/4: " + _axes + ")"))
                    _log_peasant_questions(qtxt)
                    candidates[key]["peasant_q"] = qtxt
                    _stag = " [q:" + str(_pq_score["total"]) + "/4]"
                    ev = AgentEvent("Peasant", "observation",
                                   f"Questions about {key}" + _stag + ":\n" + qtxt)
                    emit(ev)
                    discussion_lines.append(f"PEASANT → {key}:\n{qtxt}\n")

                ctx.shared["candidates"] = candidates
                ctx.shared["discussion_transcript"] = "\n".join(discussion_lines[-40:])

            # 2) Rebuttals
            if self._pause_event and not self._pause_event.is_set():
                self._pause_event.wait(timeout=300)
            self._phase("Rebuttal round")
            for key in panel:
                if key == "peasant" or key not in candidates:
                    continue
                other_roles = [r for r in candidates if r != key]
                debate_lines = [
                    "DEBATE CONTEXT:",
                    f"User request:\n{user_text}\n",
                    f"Your original answer ({key}):\n{candidates[key].get('answer','')}\n",
                ]
                if my_pq := candidates[key].get("peasant_q", ""):
                    debate_lines.append(f"Peasant questions about YOUR answer:\n{my_pq}\n")
                for rr in other_roles:
                    debate_lines.append(f"Other candidate ({rr}):\n{candidates[rr].get('answer','')}\n")
                    if pq := candidates[rr].get("peasant_q", ""):
                        debate_lines.append(f"Peasant questions about {rr}:\n{pq}\n")
                _rb_mode = ctx.shared.get("query_mode", "")
                _rb_mode_line = (
                    "⚠ MODE: CONVERSATIONAL — rebuttal must be in prose only, no code.\n"
                    if _rb_mode == "conversational" else
                    "⚠ MODE: TECHNICAL — focus on code correctness and completeness.\n"
                    if _rb_mode == "technical" else ""
                )
                debate_lines += [
                    _rb_mode_line,
                    "INSTRUCTIONS:",
                    "- Write a rebuttal/improvement note.",
                    "- Explicitly state disagreements.",
                    "- Address Peasant questions.",
                    "- Propose concrete fixes.",
                    "- Keep under 12 bullet points.",
                    "- Do NOT introduce code unless this is a TECHNICAL query.",
                ]
                extra_context = "\n".join(debate_lines)
                self._phase(f"{key.capitalize()} — rebuttal")
                rebuttal_text = self.agents[key].model.respond(
                    "Produce your rebuttal now.", extra_context=extra_context,
                    token_callback=self.agents[key]._make_token_cb(),
                    max_tokens=600,  # rebuttals must be concise bullets, not essays
                )
                candidates[key]["rebuttal"] = rebuttal_text
                ev = AgentEvent(key.capitalize(), "observation", f"Rebuttal:\n{rebuttal_text}")
                emit(ev)
                discussion_lines.append(f"{key.upper()} REBUTTAL:\n{rebuttal_text}\n")

                # ── Back-fill Peasant QA answers (Change 8) ────────────
                # The candidate's rebuttal IS their answer to Peasant's questions.
                # Fill the "a" slot in _peasant_qa_log so that in cross-fire,
                # Peasant sees what was already answered and can go deeper.
                if _peasant_qa_log:
                    peasant_qs_for_key = candidates[key].get("peasant_q", "")
                    for qa_entry in _peasant_qa_log:
                        if not qa_entry.get("a") and qa_entry["q"][:60] in peasant_qs_for_key:
                            qa_entry["a"] = rebuttal_text[:400].strip()
                ctx.shared["candidates"] = candidates
                ctx.shared["discussion_transcript"] = "\n".join(discussion_lines[-60:])

            # 3) Cross-fire
            if self._pause_event and not self._pause_event.is_set():
                self._pause_event.wait(timeout=300)
            self._phase(f"Cross-fire — {self.debate_turns} turns")
            for turn in range(1, self.debate_turns + 1):
                for key in panel:
                    if key == "peasant" or key not in candidates:
                        continue
                    _cf_mode = ctx.shared.get("query_mode", "")
                    _cf_mode_note = (
                        "⚠ CONVERSATIONAL mode: respond in prose only, no code.\n"
                        if _cf_mode == "conversational" else
                        "⚠ TECHNICAL mode: focus on code quality and correctness.\n"
                        if _cf_mode == "technical" else ""
                    )
                    extra_context = (
                        f"CROSS-FIRE CONTEXT — Turn {turn}/{self.debate_turns}\n\n"
                        + _cf_mode_note +
                        "Rules:\n"
                        "- Write ONE short message.\n"
                        "- Include: AGREE: ... | DISAGREE: ... | ADD: ...\n"
                        "- Address Peasant questions about your answer.\n"
                        "- Keep under 10 lines.\n\n"
                        f"Discussion so far:\n{ctx.shared.get('discussion_transcript','')}\n"
                    )
                    self._phase(f"{key.capitalize()} — cross-fire T{turn}")
                    msg = self.agents[key].model.respond(
                        "Post your cross-fire message now.", extra_context=extra_context,
                        token_callback=self.agents[key]._make_token_cb(),
                        max_tokens=400,  # cross-fire must be tight — 10 lines max
                    )
                    candidates[key]["discussion"] = (
                        candidates[key].get("discussion", "") + f"\nTURN {turn}:\n{msg}\n"
                    ).strip()
                    ev = AgentEvent(key.capitalize(), "observation", f"Cross-fire T{turn}:\n{msg}")
                    emit(ev)
                    discussion_lines.append(f"{key.upper()} CROSS-FIRE T{turn}:\n{msg}\n")

                    if "peasant" in self.agents:
                        self._phase(f"Peasant — questions after {key} T{turn}")
                        _cf_pmode = ctx.shared.get("query_mode", "")
                        pq = peasant_cross_exam(
                            self.agents["peasant"].model,
                            candidate_role=f"{key} (T{turn})", candidate_text=msg, user_text=user_text,
                            prior_qa=_peasant_qa_log if _peasant_qa_log else None,
                            query_mode=_cf_pmode,
                        )
                        _cf_score = _peasant_quality_score(pq, msg, _peasant_qa_log)
                        if not _looks_like_two_questions(pq):
                            # Reformat rather than regenerate — same ideas, proper labels
                            _cf_reformat = (
                                "Your response below is good but needs exactly two questions "
                                "labelled Q1: and Q2:. Reformat it now — keep the same ideas, "
                                "just add Q1: and Q2: labels and make sure each ends with '?'.\n\n"
                                f"YOUR RESPONSE:\n{pq}"
                            )
                            pq = self.agents["peasant"].model.respond(
                                _cf_reformat, max_tokens=300)
                            _cf_score = _peasant_quality_score(pq, msg, _peasant_qa_log)
                            if not _looks_like_two_questions(pq):
                                _axes = ", ".join(
                                    k + ("=✓" if v else "=✗")
                                    for k, v in _cf_score["axes"].items()
                                )
                                emit(AgentEvent("Peasant", "observation",
                                    "⚠ CF quality low after reformat ("
                                    + str(_cf_score["total"]) + "/4: " + _axes + ")"))
                        _log_peasant_questions(pq)
                        _cftag = " [q:" + str(_cf_score["total"]) + "/4]"
                        pev = AgentEvent("Peasant", "observation",
                                        f"Cross-fire questions after {key} T{turn}" + _cftag + ":\n" + pq)
                        emit(pev)
                        discussion_lines.append(f"PEASANT → {key} T{turn}:\n{pq}\n")

                    ctx.shared["candidates"] = candidates
                    ctx.shared["discussion_transcript"] = "\n".join(discussion_lines[-80:])

            # 4) Judge ranks
            self._phase("Judge — ranking candidates")
            rank_json = self.judge.rank_candidates(user_text, candidates)
            ctx.shared["judge_ranking"] = rank_json
            try:
                import json as _rj
                ctx.shared["judge_confidence"] = int(_rj.loads(rank_json).get("confidence", 0))
            except Exception:
                ctx.shared["judge_confidence"] = 0
            ev = AgentEvent("Judge", "observation", f"Ranking:\n{rank_json}")
            emit(ev)

            # 4a) Low-confidence gap logging ─────────────────────────────────
            # Roles that reported self-confidence ≤4 are flagged so the
            # Librarian wishlist captures what vault data would have helped.
            for _lc_role, _lc_data in candidates.items():
                if _lc_data.get("self_confidence", 10) <= 4:
                    try:
                        _lc_topic = f"{_lc_role} answer to: {user_text[:80]}"
                        _lc_reason = (
                            f"{_lc_role} self-reported confidence "
                            f"{_lc_data['self_confidence']}/10 — vault data on this topic "
                            "would have strengthened the answer"
                        )
                        ctx.shared.setdefault("_low_conf_gaps", []).append(
                            {"who": _lc_role, "topic": _lc_topic, "reason": _lc_reason}
                        )
                    except Exception:
                        pass

            # 4b) Peasant adversarial challenge (optional)
            _adversarial_challenge = ""
            if (ctx.shared.get("peasant_adversarial", False)
                    and "peasant" in self.agents):
                try:
                    import json as _aj
                    _robj = _aj.loads(rank_json)
                    _winner_role = _robj.get("winner", "")
                    _winner_ans = candidates.get(_winner_role, {}).get("answer", "")
                except Exception:
                    _winner_role, _winner_ans = "", ""
                if _winner_role and _winner_ans:
                    self._phase("Peasant — adversarial challenge")
                    _adv_ctx = (
                        "USER REQUEST:\n" + user_text + "\n\n"
                        "WINNING CANDIDATE: " + _winner_role + "\n"
                        "WINNING ANSWER:\n" + _winner_ans + "\n\n"
                        "Your task: argue AGAINST this answer. Identify the single most\n"
                        "dangerous flaw, edge case, or false assumption.\n"
                        "Be specific and adversarial. Do NOT offer improvements.\n"
                        "Format: CHALLENGE: <your strongest objection in 3-6 sentences>"
                    )
                    _adversarial_challenge = self.agents["peasant"].model.respond(
                        "State your adversarial challenge now.",
                        extra_context=_adv_ctx,
                        max_tokens=300,
                    )
                    ctx.shared["adversarial_challenge"] = _adversarial_challenge
                    ctx.shared["adversarial_target"] = _winner_role
                    emit(AgentEvent("Peasant", "observation",
                                   "Adversarial: " + _adversarial_challenge))

            # 5) Writer synthesizes
            self._phase("Writer — synthesizing final answer")
            synth_evs = self.agents[synth].act(ctx)
            for ev in synth_evs:
                emit(ev)
            synth_final = next((e.text for e in reversed(synth_evs) if e.kind == "final"), "")
            # T1-D: Track per-round Writer output, emit unified diff on round 2+
            _round_outputs = ctx.shared.setdefault("_round_outputs", [])
            _round_outputs.append(synth_final)
            if len(_round_outputs) >= 2:
                import difflib as _dl
                _prev_r = len(_round_outputs) - 1
                _curr_r = len(_round_outputs)
                _prev_lines = _round_outputs[-2].splitlines(keepends=True)
                _curr_lines = _round_outputs[-1].splitlines(keepends=True)
                _diff_lines = list(_dl.unified_diff(
                    _prev_lines, _curr_lines,
                    fromfile="round_" + str(_prev_r),
                    tofile="round_" + str(_curr_r),
                    lineterm="",
                ))
                if _diff_lines:
                    _diff_text = "".join(_diff_lines[:80])
                    _diff_label = "r" + str(_prev_r) + " -> r" + str(_curr_r)
                    _diff_msg = "Round diff (" + _diff_label + "):\n" + _diff_text
                    emit(AgentEvent("Orchestrator", "observation", _diff_msg))


            # 6) Judge critiques
            self._phase("Judge — critiquing synthesis")
            critique = self.judge.critique(user_text, synth_final, extra_context=f"Ranking:\n{rank_json}", query_mode=ctx.shared.get("query_mode", ""))
            ctx.shared["judge_critique"] = critique
            ev = AgentEvent("Judge", "observation", critique)
            emit(ev)

            if "Verdict: PASS" in critique:
                self._phase("✓ Verdict: PASS — deliberation complete")
                break

            # ── Confidence-gated early exit ──────────────────────────────
            # Even on NEEDS_WORK, if Judge confidence is very high (≥8/10)
            # and this is the final round, skip re-deliberation — the answer
            # is probably good enough and more rounds won't help much.
            _conf = ctx.shared.get("judge_confidence", 0)
            _is_last_round = (r == self.max_rounds - 1)
            if _conf >= 8 and _is_last_round:
                self._phase(
                    f"✓ High confidence ({_conf}/10) — accepting answer despite NEEDS_WORK"
                )
                break

            # ── Confidence-gated extra round ─────────────────────────────
            # If confidence is very low (≤2/10) on round 1, allow an extra
            # round beyond max_rounds — the answer needs more work.
            if _conf <= 2 and r == 0 and self.max_rounds < 3:
                self._phase(
                    f"⚠ Low confidence ({_conf}/10) — adding extra deliberation round"
                )
                self.max_rounds = 3

            else:
                # T2-C: Parse REQUIRED_CHANGES for targeted round-2 Writer brief
                _changes = self.judge.__class__.parse_required_changes(critique)
                if _changes:
                    ctx.shared["required_changes"] = _changes
                    _chg_txt = "\n".join("- " + c for c in _changes)
                    emit(AgentEvent("Judge", "observation",
                                   "Required changes for next round:\n" + _chg_txt))

        # Expose shared context so caller can retrieve low-confidence gaps etc.
        self._last_ctx = ctx
        return all_events


# ============================================================
# Vault search tool
# ============================================================

def _librarian_brief(
    rag,
    query: str,
    *,
    log_cb=None,
    max_chars: int = 5500,
) -> dict:
    """
    Run a multi-angle vault search before deliberation and return a structured
    briefing dict for injecting into Council personality prompts.

    Returns:
        raw     — full briefing text for most roles
        peasant — shorter targeted briefing for the Peasant role
        summary — one-line summary of what was found
        sources — list of filenames that contributed context
        found   — True if any relevant content was found
    """
    empty = {"raw": "", "peasant": "", "summary": "", "sources": [], "found": False}
    q = (query or "").strip()
    if not q:
        return empty

    import re as _re
    stop = {"what", "when", "where", "who", "why", "how", "is", "are", "the",
            "a", "an", "in", "on", "of", "to", "do", "does", "can", "i",
            "my", "me", "for", "and", "or", "with", "this", "that"}
    keywords = [w for w in _re.findall(r"[a-zA-Z]{3,}", q.lower()) if w not in stop]
    angles = [q] + keywords[:3]

    all_matches: list = []
    sources_seen: set = set()

    for angle in angles:
        if log_cb:
            log_cb("RAG search: " + repr(angle))
        try:
            if hasattr(rag, "search"):
                results = rag.search(angle, n_results=4)
                # RAGResult may be a custom object, not a plain list.
                # Normalise to a list of dicts regardless of return type.
                if results is None:
                    rows = []
                elif isinstance(results, (list, tuple)):
                    rows = list(results)
                elif hasattr(results, "documents"):
                    # ChromaDB QueryResult style: .documents, .metadatas, .ids
                    docs  = results.documents  or [[]]
                    metas = results.metadatas  or [[]]
                    ids   = results.ids        or [[]]
                    # ChromaDB returns list-of-lists (one per query)
                    docs  = docs[0]  if docs  and isinstance(docs[0],  list) else docs
                    metas = metas[0] if metas and isinstance(metas[0], list) else metas
                    ids   = ids[0]   if ids   and isinstance(ids[0],   list) else ids
                    rows  = [
                        {"text": d, "file": (m or {}).get("source", i)}
                        for d, m, i in zip(docs, metas, ids)
                    ]
                elif hasattr(results, "__iter__"):
                    rows = list(results)
                else:
                    # Unknown RAGResult type — try treating as single result
                    rows = [results]

                for r in rows:
                    if isinstance(r, dict):
                        fname = r.get("file") or r.get("source") or r.get("id") or "unknown"
                        text  = r.get("text") or r.get("document") or r.get("content") or ""
                    elif hasattr(r, "__dict__"):
                        d     = vars(r)
                        fname = d.get("file") or d.get("source") or d.get("id") or "unknown"
                        text  = d.get("text") or d.get("document") or d.get("content") or ""
                    else:
                        fname, text = "unknown", str(r)[:500]
                    if text and fname not in sources_seen:
                        sources_seen.add(fname)
                        all_matches.append({"file": fname, "excerpt": text[:500]})
            else:
                vault_dir = getattr(rag, "vault_dir", None)
                if vault_dir:
                    from pathlib import Path as _Path
                    res = _vault_search_impl(_Path(vault_dir), angle)
                    for m in res.get("matches", []):
                        if m["file"] not in sources_seen:
                            sources_seen.add(m["file"])
                            all_matches.append(m)
        except Exception as e:
            if log_cb:
                log_cb("RAG angle error (" + repr(angle) + "): " + str(e))

    if not all_matches:
        return empty

    sections = []
    for m in all_matches[:8]:
        fname   = m.get("file", "?")
        excerpt = (m.get("excerpt") or m.get("text") or "").strip()
        if excerpt:
            sections.append("[" + fname + "]\n" + excerpt)

    raw = "VAULT CONTEXT:\n" + "\n\n".join(sections)
    if len(raw) > max_chars:
        raw = raw[:max_chars] + "\n...[truncated]"

    peasant_sections = sections[:3]
    peasant = (
        "VAULT CONTEXT (use to ask targeted follow-up questions):\n"
        + "\n\n".join(peasant_sections)
    )
    if len(peasant) > max_chars // 2:
        peasant = peasant[:max_chars // 2] + "\n...[truncated]"

    summary = "Found " + str(len(all_matches)) + " relevant vault excerpt(s) from: " + ", ".join(list(sources_seen)[:4])

    return {
        "raw":     raw,
        "peasant": peasant,
        "summary": summary,
        "sources": sorted(sources_seen),
        "found":   True,
    }

def _vault_search_impl(vault_dir: Path, query: str, *, max_files: int = 80) -> Dict[str, Any]:
    q = (query or "").strip()
    if not q:
        return {"query": q, "matches": []}
    qlow = q.lower()
    matches = []
    files = sorted(vault_dir.iterdir(), key=lambda p: p.name.lower())
    scanned = 0
    for p in files:
        if scanned >= max_files or not p.is_file():
            continue
        scanned += 1
        try:
            if p.stat().st_size > 250_000:
                continue
            text = p.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue
        tlow = text.lower()
        idx = tlow.find(qlow)
        if idx == -1:
            continue
        start = max(0, idx - 140)
        end = min(len(text), idx + 260)
        excerpt = text[start:end].replace("\n", " ")
        matches.append({"file": p.name, "excerpt": excerpt})
        if len(matches) >= 12:
            break
    return {"query": q, "matches": matches, "scanned": scanned}


def _make_tools(runner: ce.LocalRunner, librarian: ce.Librarian, vault_dir: Path) -> Dict[str, ToolFn]:
    def run_python(args: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
        code = str(args.get("code", "")).strip()
        if not code:
            return False, "No code provided.", {}
        rc, out, err, path = runner.run_code(code, filename_hint=str(args.get("filename", "scratch.py")), timeout_s=int(args.get("timeout_s", 120)))
        return True, f"rc={rc}\n--- stdout ---\n{out}\n--- stderr ---\n{err}\nfile={path}", {"rc": rc, "stdout": out, "stderr": err, "path": str(path)}

    def vault_save(args: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
        name = str(args.get("name", "note.txt")).strip() or "note.txt"
        content = str(args.get("content", ""))
        if not content:
            return False, "No content provided.", {}
        p = librarian.save_text(name, content)
        return True, f"Saved to vault as '{p.name}'.", {"path": str(p)}

    def vault_list(args: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
        items = librarian.list_items()
        return True, "\n".join(items) if items else "(empty)", {"items": items}

    def vault_read(args: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
        name = str(args.get("name", "")).strip()
        if not name:
            return False, "Provide {'name': 'file.txt'}", {}
        txt = librarian.read_text(name)
        return True, txt, {"name": name}

    def vault_search(args: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
        res = _vault_search_impl(vault_dir, str(args.get("query", "")))
        lines = [f"Vault search: {res.get('query','')}"]
        for m in res.get("matches", []):
            lines.append(f"- {m['file']}: {m['excerpt']}")
        if not res.get("matches"):
            lines.append("(no matches)")
        return True, "\n".join(lines), res

    return {"run_python": run_python, "vault_save": vault_save,
            "vault_list": vault_list, "vault_read": vault_read, "vault_search": vault_search}


# ============================================================
# App paths / config
# ============================================================

STORE_PASSWORDS = True

# COUNCIL_APP_DIR  — root for all Council state (.council/ by default)
# COUNCIL_VAULT_ROOT — explicit vault folder (overrides APP_DIR/vault).
# Letting an analyst point at a per-client folder (or a OneDrive sync) without
# editing source is the difference between "I have to copy data in" and "I can
# just open my project share".
APP_DIR = Path(os.environ.get("COUNCIL_APP_DIR", "")).expanduser().resolve() \
    if os.environ.get("COUNCIL_APP_DIR") else (Path.home() / ".council")
APP_DIR.mkdir(parents=True, exist_ok=True)


# ── Vault Git clone helper (used by Vault Manager tab) ────────────────────────

def _vmgr_clone_repo(
    url: str,
    *,
    vault_dir: Path,
    subfolder: str | None = None,
    branch: str | None = None,
    depth: int = 1,
    log_cb=None,
) -> Path:
    """
    Clone or update a GitHub repo and copy indexable files into vault_dir.
    log_cb(msg) is called with progress strings for the GUI log.
    Returns the destination vault subfolder Path.
    """
    import re as _re
    import shutil as _shutil
    import subprocess as _sp

    def _log(m):
        if log_cb:
            log_cb(m)
        else:
            print(m)

    INDEXABLE = {
        ".py", ".md", ".txt", ".json", ".yaml", ".yml",
        ".html", ".rst", ".csv", ".log", ".toml", ".ini",
    }
    SKIP_DIRS  = {".git", ".github", "__pycache__", "node_modules",
                  ".tox", "dist", "build", ".venv", "venv", "env",
                  ".eggs", ".mypy_cache", ".pytest_cache"}
    SKIP_FILES = {".gitignore", ".gitattributes", ".gitmodules",
                  "poetry.lock", "package-lock.json", "yarn.lock",
                  "Pipfile.lock", ".DS_Store"}
    MAX_BYTES  = 500_000

    if not subfolder:
        name = url.rstrip("/").rstrip(".git").rsplit("/", 1)[-1]
        subfolder = _re.sub(r"[^A-Za-z0-9._-]", "_", name) or "repo"

    clone_dir = vault_dir / ".git_clones" / subfolder
    dest_dir  = vault_dir / subfolder

    if clone_dir.exists():
        _log(f"Updating existing clone: {subfolder}")
        r = _sp.run(["git", "pull"], cwd=str(clone_dir),
                    capture_output=True, text=True, timeout=120)
        _log(r.stdout.strip() or r.stderr.strip() or "Already up to date.")
    else:
        clone_dir.parent.mkdir(parents=True, exist_ok=True)
        cmd = ["git", "clone", f"--depth={depth}"]
        if branch:
            cmd += ["--branch", branch]
        cmd += [url, str(clone_dir)]
        _log(f"Cloning {url} …")
        r = _sp.run(cmd, capture_output=True, text=True, timeout=300)
        if r.returncode != 0:
            raise RuntimeError(r.stderr.strip() or "git clone failed")
        _log(f"Cloned to {clone_dir.name}")

    dest_dir.mkdir(parents=True, exist_ok=True)
    copied = skipped = 0
    for src in clone_dir.rglob("*"):
        if not src.is_file():
            continue
        parts = set(src.relative_to(clone_dir).parts)
        if parts & SKIP_DIRS or src.name in SKIP_FILES:
            skipped += 1
            continue
        if src.suffix.lower() not in INDEXABLE:
            skipped += 1
            continue
        try:
            if src.stat().st_size > MAX_BYTES:
                skipped += 1
                continue
        except OSError:
            skipped += 1
            continue
        dst = dest_dir / src.relative_to(clone_dir)
        dst.parent.mkdir(parents=True, exist_ok=True)
        _shutil.copy2(src, dst)
        copied += 1

    _log(f"Copied {copied} files → vault/{subfolder}  ({skipped} skipped)")
    return dest_dir


# ── Shared import filters (zip + folder import both use these) ──────────
# Widened from the original code-repo allow-list to include the DATA formats
# the analyst actually reads (Excel / Parquet / SQLite / DuckDB / BSON / TSV
# / NDJSON / images / PDF), so importing a zip or folder of data isn't lossy.
_IMPORT_INDEXABLE = {
    # text / code / config
    ".py", ".md", ".txt", ".json", ".yaml", ".yml", ".html", ".rst",
    ".csv", ".log", ".toml", ".ini", ".xml", ".cfg", ".conf", ".tex",
    ".r", ".m", ".ipynb",
    # tabular / structured data the analyst reads
    ".tsv", ".xlsx", ".xls", ".xlsm", ".parquet", ".feather", ".orc",
    ".arrow", ".db", ".sqlite", ".sqlite3", ".duckdb", ".bson",
    ".jsonl", ".ndjson", ".gz",
    # images (parsed for metadata / vision)
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tif", ".tiff",
    # documents
    ".pdf",
}
_IMPORT_SKIP_DIRS = {"__pycache__", "node_modules", ".git", ".venv", "venv",
                     "dist", "build", ".eggs", ".tox", ".idea", ".vscode"}


def _import_max_bytes() -> int:
    """Max size of a SINGLE file kept on import — a runaway guard, NOT a data
    limit. Default 1 GiB (vs the old 500 KB, which silently dropped any real
    data file). Override with COUNCIL_IMPORT_MAX_MB; set it very high to
    effectively disable the cap."""
    import os as _os
    ov = _os.environ.get("COUNCIL_IMPORT_MAX_MB", "").strip()
    if ov:
        try:
            return max(1, int(ov)) * 1024 * 1024
        except ValueError:
            pass
    return 1024 * 1024 * 1024


def _vmgr_extract_zip(
    zip_path: Path,
    *,
    vault_dir: Path,
    subfolder: str | None = None,
    log_cb=None,
) -> tuple:
    """
    Extract a zip archive into a vault subfolder, keeping only indexable files.
    Returns (dest_dir, copied_count, skipped_count).
    """
    import zipfile
    import shutil as _shutil
    import re as _re

    def _log(m):
        if log_cb: log_cb(m)
        else: print(m)

    INDEXABLE = _IMPORT_INDEXABLE
    SKIP_DIRS = _IMPORT_SKIP_DIRS
    MAX_BYTES = _import_max_bytes()

    if not subfolder:
        subfolder = zip_path.stem
    subfolder = _re.sub(r"[^A-Za-z0-9._-]", "_", subfolder) or "import"
    dest_dir = vault_dir / subfolder
    dest_dir.mkdir(parents=True, exist_ok=True)
    _dest_root = dest_dir.resolve()   # Zip Slip containment boundary

    if not zipfile.is_zipfile(zip_path):
        raise ValueError(f"{zip_path.name} is not a valid zip file")

    copied = skipped = 0
    with zipfile.ZipFile(zip_path, "r") as zf:
        members = [m for m in zf.infolist() if not m.filename.endswith("/")]
        _log(f"  {len(members)} files in archive")

        # Detect common top-level prefix to strip (e.g. "repo-main/")
        all_parts = [Path(m.filename).parts for m in members]
        strip_prefix = ""
        if all_parts and len(set(p[0] for p in all_parts if p)) == 1:
            strip_prefix = all_parts[0][0]

        for member in members:
            parts = Path(member.filename).parts
            if any(p in SKIP_DIRS for p in parts): skipped += 1; continue
            if any(p.startswith(".") for p in parts): skipped += 1; continue
            if Path(member.filename).suffix.lower() not in INDEXABLE: skipped += 1; continue
            if member.file_size > MAX_BYTES:
                _log(f"  SKIP (too large {member.file_size//1024}KB): {member.filename}")
                skipped += 1; continue

            # Strip the common prefix so files land at vault/subfolder/file, not vault/subfolder/repo-main/file
            rel_parts = parts[1:] if (strip_prefix and parts and parts[0] == strip_prefix) else parts
            if not rel_parts:
                rel_parts = (Path(member.filename).name,)
            dest_file = dest_dir / Path(*rel_parts)
            # ── Zip Slip guard ──────────────────────────────────────────
            # zf.open()+manual write bypasses ZipFile.extractall()'s built-in
            # sanitisation, so a crafted entry with an ABSOLUTE path (pathlib
            # resets the join) or ../ / symlink parts could write OUTSIDE the
            # target. Refuse anything that doesn't resolve inside dest_dir.
            try:
                _resolved = dest_file.resolve()
                _resolved.relative_to(_dest_root)
            except (ValueError, OSError):
                _log(f"  SKIP (unsafe path escapes target): {member.filename}")
                skipped += 1
                continue
            dest_file = _resolved
            dest_file.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member) as src, open(dest_file, "wb") as dst:
                _shutil.copyfileobj(src, dst)
            copied += 1

    return dest_dir, copied, skipped


def _vmgr_copy_folder(
    src: Path,
    *,
    vault_dir: Path,
    subfolder: str | None = None,
    log_cb=None,
) -> tuple:
    """
    Copy a local folder into the vault, keeping only indexable files.
    Returns (dest_dir, copied_count, skipped_count).
    """
    import shutil as _shutil
    import re as _re

    def _log(m):
        if log_cb: log_cb(m)
        else: print(m)

    INDEXABLE = _IMPORT_INDEXABLE
    SKIP_DIRS = _IMPORT_SKIP_DIRS
    MAX_BYTES = _import_max_bytes()

    if not subfolder:
        subfolder = src.name
    subfolder = _re.sub(r"[^A-Za-z0-9._-]", "_", subfolder) or "import"
    dest_dir = vault_dir / subfolder
    dest_dir.mkdir(parents=True, exist_ok=True)

    copied = skipped = 0
    for src_file in src.rglob("*"):
        if not src_file.is_file(): continue
        rel = src_file.relative_to(src)
        if any(p in SKIP_DIRS for p in rel.parts): skipped += 1; continue
        if any(p.startswith(".") for p in rel.parts): skipped += 1; continue
        if src_file.suffix.lower() not in INDEXABLE: skipped += 1; continue
        try:
            if src_file.stat().st_size > MAX_BYTES: skipped += 1; continue
        except OSError:
            skipped += 1; continue
        dest_file = dest_dir / rel
        dest_file.parent.mkdir(parents=True, exist_ok=True)
        _shutil.copy2(src_file, dest_file)
        copied += 1

    _log(f"  Copied {copied} files from {src.name}")
    return dest_dir, copied, skipped


# ── All persistent data lives under VAULT_DIR ─────────────────
# ~/.council/vault/
#   conversations/             ← per-session chat history
#   memory/                    ← per-role persistent memory
#   logs/                      ← council.log and session logs
#   workspace/                 ← code runner scratch files
#   graph_output/              ← grapher exports
#   .chromadb/                 ← ChromaDB vector index
#   .git_clones/               ← cloned reference repos
#   node_registry.json         ← SSH node registry
#   personality_backends.json  ← model pins
VAULT_DIR            = Path(os.environ.get("COUNCIL_VAULT_ROOT", "")).expanduser().resolve() \
    if os.environ.get("COUNCIL_VAULT_ROOT") else (APP_DIR / "vault")
# Repo root — directory holding this file (council_gui_engine.py).
# Used by the Changelog tab to invoke `git log`/`git show` against the
# actual checkout the user is running, regardless of CWD.
_REPO_ROOT           = Path(__file__).resolve().parent
VERDICT_HISTORY_PATH = VAULT_DIR / "verdict_history.jsonl"
VAULT_DIR.mkdir(parents=True, exist_ok=True)

LOG_PATH      = VAULT_DIR / "logs" / "council.log"
WORKSPACE_DIR = VAULT_DIR / "workspace"
TMP_DIR       = VAULT_DIR / "tmp"
REGISTRY_PATH = VAULT_DIR / "node_registry.json"
PINS_PATH            = VAULT_DIR / "personality_backends.json"
INSTRUCTIONS_PATH    = VAULT_DIR / "council_instructions.json"
CONTENT_STYLE_PATH   = VAULT_DIR / "content_style.json"

# Ensure subdirs exist on startup
for _d in (LOG_PATH.parent, WORKSPACE_DIR, TMP_DIR):
    _d.mkdir(parents=True, exist_ok=True)


def _migrate_old_paths_to_vault() -> None:
    """
    One-time migration: move data from old scattered locations into vault.
    Safe to run repeatedly — skips anything already moved.
    """
    import shutil

    migrations = [
        # (old_path,                          new_path,           is_dir)
        (APP_DIR / "council.log",             LOG_PATH,           False),
        (APP_DIR / "node_registry.json",      REGISTRY_PATH,      False),
        (APP_DIR / "personality_backends.json", PINS_PATH,        False),
        (APP_DIR / "workspace",               WORKSPACE_DIR,      True),
        (APP_DIR / "graph_output",            VAULT_DIR / "graph_output", True),
        (APP_DIR / ".chromadb",               VAULT_DIR / ".chromadb",    True),
        # dream3d docs scraped next to the script
        (Path(__file__).parent / "vault" / "dream3d_docs",
         VAULT_DIR / "dream3d_docs", True),
    ]

    moved = []
    for old, new, is_dir in migrations:
        if not old.exists() or old == new:
            continue
        if new.exists():
            # destination already has content — don't overwrite
            continue
        try:
            new.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(old), str(new))
            moved.append(f"  {old.name} → vault/{new.relative_to(VAULT_DIR)}")
        except Exception as e:
            print(f"[Migration] Could not move {old.name}: {e}")

    if moved:
        print("[Migration] Moved old data into vault:")
        for m in moved:
            print(m)


# Run migration silently on startup
try:
    _migrate_old_paths_to_vault()
except Exception:
    pass


# ============================================================
# Colour / tag constants for the transcript
# ============================================================

ROLE_COLORS = {
    "User":        "#4fc3f7",   # light blue
    "Judge":       "#ef9a9a",   # red-ish
    "Writer":      "#a5d6a7",   # green
    "Coder": "#ce93d8",   # purple
    "Intern":      "#ffe082",   # yellow
    "Peasant":     "#ffcc80",   # orange
    "Artist":      "#f48fb1",   # pink
    "Orchestrator":"#b0bec5",   # grey
    "Librarian":   "#80cbc4",   # teal
    "Apothecary":  "#bcaaa4",   # brown-ish
}

PHASE_COLOR   = "#78909c"
TOKEN_COLOR   = "#e0e0e0"
DEFAULT_COLOR = "#cfd8dc"


# ============================================================
# CouncilConsole  (main GUI)
# ============================================================

# ============================================================
# Council Instruction Manager
# ============================================================
# Persists a list of named instructions to vault/council_instructions.json
# Each entry: {"id": str, "name": str, "text": str, "active": bool}
# ============================================================

# ============================================================
# Content Style Manager
# ============================================================
# Persists cross-session learning for the Content Creator:
#   - What hooks/structures worked for this creator
#   - Audience and tone preferences
#   - Script templates for different video types
# Stored in vault/content_style.json
# ============================================================

# Built-in script templates — saved to vault on first run if missing
DEFAULT_SCRIPT_TEMPLATES = {
    "explainer": {
        "name": "Explainer / Educational",
        "description": "Teaching the viewer something clearly",
        "structure": [
            "HOOK (0:00-0:30): Lead with the surprising fact or the payoff — why should they care?",
            "SETUP (0:30-1:30): Define the problem or concept in plain language.",
            "SECTION 1 (1:30-4:00): Core concept — one main idea, explained with an example.",
            "SECTION 2 (4:00-6:30): Deeper dive or second angle — add nuance or a complication.",
            "SECTION 3 (6:30-8:30): Practical application — what does the viewer do with this?",
            "RECAP (8:30-9:30): Summarise the 3 key points in one sentence each.",
            "CTA/OUTRO (9:30-10:00): Call to action. Natural, not forced.",
        ]
    },
    "comedy_retrospective": {
        "name": "Comedy / Retrospective",
        "description": "Looking back at something with humour and self-awareness",
        "structure": [
            "HOOK (0:00-0:30): Most absurd or funniest moment first — then 'let me explain'.",
            "CONTEXT (0:30-1:30): Set the scene. Who were you, why did this exist?",
            "THE THING (1:30-5:00): Walk through it with running commentary. Lean into the absurdity.",
            "TURNING POINT (5:00-7:00): The moment you realise how unhinged it was.",
            "REFLECTION (7:00-9:00): What you'd do differently / what it says about that time.",
            "CALLBACK (9:00-9:45): Return to the opening hook with new context.",
            "OUTRO (9:45-10:00): Short, punchy. Leave them on a laugh.",
        ]
    },
    "project_showcase": {
        "name": "Project Showcase / Build Log",
        "description": "Showing off something you built",
        "structure": [
            "HOOK (0:00-0:30): Show the final result first. Make them want to know how.",
            "THE PROBLEM (0:30-1:30): Why did you build this? What was broken or missing?",
            "THE APPROACH (1:30-3:00): How you decided to solve it — options you considered.",
            "BUILD SECTION 1 (3:00-5:30): First major step — keep it visual, show don't tell.",
            "BUILD SECTION 2 (5:30-8:00): Second step — include a failure or pivot if there was one.",
            "RESULT (8:00-9:00): Show it working. Be honest about limitations.",
            "WHAT I LEARNED (9:00-9:45): One or two genuine takeaways.",
            "OUTRO (9:45-10:00): What's next. CTA.",
        ]
    },
    "powerpoint_roast": {
        "name": "PowerPoint / Document Roast",
        "description": "Revisiting old work with comedic commentary",
        "structure": [
            "HOOK (0:00-0:30): Title slide reveal — just let the title land.",
            "INTRO (0:30-1:30): What was this, when did you make it, why does it exist?",
            "SLIDE BY SLIDE (1:30-7:30): Work through it. Commentary on each slide. Don't rush.",
            "HIGHLIGHT (7:30-8:30): The single most unhinged moment. Give it space.",
            "VERDICT (8:30-9:30): Would past-you have been proud? Were you right?",
            "OUTRO (9:30-10:00): Tease the next one or invite viewers to share their worst.",
        ]
    },
    "tutorial": {
        "name": "Tutorial / How-To",
        "description": "Teaching the viewer to do something step by step",
        "structure": [
            "HOOK (0:00-0:30): Show the finished result — what they'll be able to do.",
            "REQUIREMENTS (0:30-1:30): What they need before starting. Be specific.",
            "STEP 1 (1:30-3:30): First step — clear, slow, no assumptions.",
            "STEP 2 (3:30-5:30): Second step — mention common mistakes here.",
            "STEP 3 (5:30-7:30): Third step — most tutorials lose people here, be extra clear.",
            "TROUBLESHOOTING (7:30-8:30): Top 2-3 things that go wrong and how to fix them.",
            "RESULT (8:30-9:30): Show the working result. Recap the steps briefly.",
            "OUTRO (9:30-10:00): Where to go next. CTA.",
        ]
    },
}


class ContentStyleManager:
    """
    Persists cross-session learning for the Content Creator personality.
    Stores style preferences, what worked, audience notes, and script templates.
    """

    def __init__(self, path: Path):
        self.path = path
        self._data: Dict = {}
        self._load()
        self._ensure_templates()

    def _load(self) -> None:
        try:
            if self.path.exists():
                self._data = _json.loads(self.path.read_text(encoding="utf-8"))
        except Exception:
            self._data = {}

    def _save(self) -> None:
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.path.write_text(
                _json.dumps(self._data, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass

    def _ensure_templates(self) -> None:
        """Write default templates to vault if not already present."""
        if "templates" not in self._data:
            self._data["templates"] = DEFAULT_SCRIPT_TEMPLATES
            self._save()

    # ── Style memory ─────────────────────────────────────────────────────

    def add_style_note(self, note: str, category: str = "general") -> None:
        """Record something that worked or a style preference."""
        notes = self._data.setdefault("style_notes", [])
        notes.append({
            "note": note.strip(),
            "category": category,
            "ts": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        })
        # Keep last 50 notes
        self._data["style_notes"] = notes[-50:]
        self._save()

    def set_audience(self, description: str) -> None:
        self._data["audience"] = description.strip()
        self._save()

    def set_tone(self, description: str) -> None:
        self._data["tone"] = description.strip()
        self._save()

    def get_style_notes(self, category: str = "") -> List[Dict]:
        notes = self._data.get("style_notes", [])
        if category:
            notes = [n for n in notes if n.get("category") == category]
        return notes

    # ── Templates ────────────────────────────────────────────────────────

    def get_templates(self) -> Dict:
        return self._data.get("templates", DEFAULT_SCRIPT_TEMPLATES)

    def add_template(self, key: str, name: str, description: str,
                     structure: List[str]) -> None:
        templates = self._data.setdefault("templates", {})
        templates[key] = {"name": name, "description": description, "structure": structure}
        self._save()

    def best_template_for(self, query: str) -> Optional[Dict]:
        """Return the most relevant template based on query keywords."""
        q = query.lower()
        templates = self.get_templates()
        _signals = {
            "powerpoint_roast":    ["powerpoint", "ppt", "slides", "presentation", "college", "roast"],
            "comedy_retrospective":["funny", "unhinged", "comedy", "absurd", "joke", "old", "past"],
            "project_showcase":    ["project", "build", "built", "made", "showcase", "raspberry", "council", "ai"],
            "tutorial":            ["tutorial", "how to", "how do", "step by step", "guide"],
            "explainer":           ["explain", "what is", "why does", "overview", "introduction"],
        }
        best_key = None
        best_score = 0
        for key, signals in _signals.items():
            score = sum(1 for s in signals if s in q)
            if score > best_score and key in templates:
                best_score = score
                best_key = key
        if best_key:
            return {"key": best_key, **templates[best_key]}
        return None

    # ── Context block for injection ──────────────────────────────────────

    def build_context_block(self, query: str) -> str:
        """Build a context string to prepend to the Content Creator's prompt."""
        parts: List[str] = []

        audience = self._data.get("audience", "")
        tone     = self._data.get("tone", "")
        if audience or tone:
            aud_lines = []
            if audience:
                aud_lines.append(f"  Audience: {audience}")
            if tone:
                aud_lines.append(f"  Tone: {tone}")
            parts.append("CREATOR PROFILE:\n" + "\n".join(aud_lines))

        notes = self.get_style_notes()
        if notes:
            note_lines = [f"  • [{n['category']}] {n['note']}" for n in notes[-10:]]
            parts.append("CONTENT STYLE MEMORY (what has worked for this creator):\n"
                         + "\n".join(note_lines))

        template = self.best_template_for(query)
        if template:
            struct_lines = ["  " + s for s in template.get("structure", [])]
            parts.append(
                f"SCRIPT TEMPLATE — {template['name']} ({template['description']}):\n"
                + "\n".join(struct_lines)
                + "\n\nUse this structure as your scaffold. Adapt timing to actual content length."
            )

        return "\n\n".join(parts) if parts else ""


class InstructionManager:
    """Persistent, toggleable list of council-wide instructions."""

    def __init__(self, path: Path):
        self.path = path
        self._instructions: List[Dict] = []
        self._load()

    def _load(self) -> None:
        try:
            if self.path.exists():
                data = _json.loads(self.path.read_text(encoding="utf-8"))
                self._instructions = data if isinstance(data, list) else []
        except Exception:
            self._instructions = []

    def _save(self) -> None:
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.path.write_text(
                _json.dumps(self._instructions, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass

    def add(self, name: str, text: str) -> Dict:
        import uuid as _uuid
        entry = {
            "id":     _uuid.uuid4().hex[:8],
            "name":   name.strip() or text[:40].strip(),
            "text":   text.strip(),
            "active": True,
        }
        self._instructions.append(entry)
        self._save()
        return entry

    def toggle(self, entry_id: str) -> bool:
        """Flip active state. Returns new state."""
        for e in self._instructions:
            if e["id"] == entry_id:
                e["active"] = not e["active"]
                self._save()
                return e["active"]
        return False

    def remove(self, entry_id: str) -> None:
        self._instructions = [e for e in self._instructions if e["id"] != entry_id]
        self._save()

    def update_text(self, entry_id: str, new_text: str) -> None:
        for e in self._instructions:
            if e["id"] == entry_id:
                e["text"] = new_text.strip()
                self._save()
                return

    def all(self) -> List[Dict]:
        return list(self._instructions)

    def active_text(self) -> str:
        """Return all active instructions joined, ready to inject into context."""
        parts = [e["text"] for e in self._instructions if e.get("active")]
        return "\n".join(parts)

    def active_count(self) -> int:
        return sum(1 for e in self._instructions if e.get("active"))


class CouncilConsole(tk.Tk):
    def __init__(self):
        super().__init__()
        # Hide immediately, before any widgets are built. Otherwise Tk
        # paints the half-constructed root while the ~15 tabs are being
        # assembled — the "small version appears while loading then
        # disappears" flicker. main() reveals the finished window after
        # the splash (the geometry/scaling calls below are unaffected by
        # being withdrawn). Reveal is guaranteed by main()'s splash
        # on_done + an independent backstop timer.
        try:
            self.withdraw()
        except Exception:
            pass
        self.title(branding.window_title())
        branding.apply_window_icon(self)
        self.geometry("1150x820")
        self.configure(bg="#1a1414")

        # ── UI scaling ──────────────────────────────────────────────
        # Tk's default scaling factor renders text small on a few
        # common configurations: WSLg on Windows (always 96 DPI even
        # when the host is 1.5×/2× scaled), HiDPI Linux without
        # GDK_SCALE, and Windows native on a 4K monitor where Tkinter
        # ignores the system DPI setting. The fonts in this codebase
        # are hardcoded to size 9-11 — instead of editing dozens of
        # widget creation sites, we lean on Tk's global scaling
        # multiplier which applies to ALL widgets uniformly.
        #
        # COUNCIL_UI_SCALE env var lets the user pin a value; otherwise
        # we auto-detect a sensible default (1.5 on WSL, 1.3 on Linux,
        # OS default on Windows native).
        try:
            scale_env = os.environ.get("COUNCIL_UI_SCALE", "").strip()
            if scale_env:
                self._ui_scale = float(scale_env)
            elif _is_wsl():
                # WSLg always reports 96 DPI; users routinely want
                # 1.5-2.0 for a comfortable read on a 4K Windows host.
                self._ui_scale = 1.5
            elif sys.platform.startswith("linux"):
                self._ui_scale = 1.3
            else:
                self._ui_scale = float(self.tk.call("tk", "scaling"))
        except Exception:
            self._ui_scale = 1.0
        try:
            self.tk.call("tk", "scaling", self._ui_scale)
        except Exception:
            pass

        # Ctrl+= / Ctrl+- (and Ctrl+0 to reset) bump the scaling
        # multiplier at runtime — same convention as browsers. The new
        # value persists for the session; pin it via COUNCIL_UI_SCALE
        # to make it stick across launches.
        self.bind_all("<Control-equal>",   lambda _e: self._adjust_ui_scale(+0.1))
        self.bind_all("<Control-plus>",    lambda _e: self._adjust_ui_scale(+0.1))
        self.bind_all("<Control-KP_Add>",  lambda _e: self._adjust_ui_scale(+0.1))
        self.bind_all("<Control-minus>",   lambda _e: self._adjust_ui_scale(-0.1))
        self.bind_all("<Control-KP_Subtract>", lambda _e: self._adjust_ui_scale(-0.1))
        self.bind_all("<Control-Key-0>",   lambda _e: self._reset_ui_scale())

        # Refresh the title bar with the chosen n_ctx + source as soon
        # as the model loads. We poll at startup (n_ctx_status returns a
        # preview before the model is loaded) and then on every chat
        # turn (via _refresh_title_with_n_ctx). The user always sees the
        # window size in a persistent location without needing to type
        # 'context info'.
        self._base_title = branding.window_title()
        self.after(500, self._refresh_title_with_n_ctx)

        # ── Loading splash (covers the whole construction) ───────────
        # Bring the spinning-cog splash up NOW, before the heavy
        # personality/engine/agent wiring and the ~15-tab build below.
        # Manual mode = no auto-dismiss; we pump one frame per heavy step
        # (the Tk loop is blocked during synchronous construction, so the
        # cog can't spin on its own) and main() dismisses + reveals once
        # construction is done and a minimum display time has elapsed.
        import time as _time
        self._splash = None
        self._splash_started = _time.monotonic()
        # Under Spyder/IPython the host's Qt loop + in-process threads make
        # the Tk splash-pump and the off-main-thread torch load unsafe, so
        # we detect that here and take the simpler path (no splash, reveal
        # now, no RAG auto-thread). Normal .exe / CLI launches are
        # unaffected.
        self._interactive_host = _under_interactive_host()
        if not self._interactive_host:
            try:
                import splash as _splash_mod
                self._splash = _splash_mod.show_splash(self, manual=True)
                self._splash.pump()
            except Exception as _e:
                print(f"[Splash] could not start loading screen: {_e!r}")
                self._splash = None
        else:
            print("[startup] Interactive host (Spyder/IPython) detected — "
                  "splash disabled; window shows immediately.", flush=True)

        self.ui_q: queue.Queue = queue.Queue()
        # Pause/resume for personality clarification requests
        self._pause_event = threading.Event()
        self._pause_event.set()  # starts unpaused
        self._clarification_answer: str = ""

        self.vault_dir = VAULT_DIR
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.prior_session_id: Optional[str] = None
        self.convo_store = ce.ConversationStore(VAULT_DIR / "conversations")
        # Per-session debug log under vault/conversation_logs/.
        # CRITICAL: this folder is in PROTECTED_SUBDIRS and is never read
        # by the model — see conversation_logger.py for the full guarantee.
        import conversation_logger as _cl
        self.conv_logger = _cl.ConversationLogger(VAULT_DIR)
        # Provenance tracker — in-memory record of injected file content
        # per turn so we can answer "where did this value come from?" with
        # an exact row reference (or flag it as hallucinated).
        import provenance as _prov
        self.provenance = _prov.ProvenanceTracker(max_turns=20)
        try:
            self.conv_logger.start_session(self.session_id)
        except Exception as _e:
            print(f"[ConvLogger] start_session failed: {_e!r}")
        # Window-close handler so the log gets a clean session_end marker.
        self.protocol("WM_DELETE_WINDOW", self._on_app_close)
        # Periodic background flush so logs survive crashes mid-session.
        self.after(30_000, self._periodic_log_flush)
        # Load persisted backend selection (Ollama vs. GGUF, model choice)
        # BEFORE the engine builds its model dispatcher. Applies to env vars
        # so council_engine.DEFAULT_MODELS picks the right values.
        _backend_cfg = self._load_backend_settings()
        # Fast direct answers: when the analyst can answer deterministically
        # (file count, summaries, a fresh derived result, a collection), skip
        # the multi-role deliberation and reply instantly. Default ON; the
        # Engine settings dialog persists the user's choice.
        self._fast_answers_enabled = bool(
            (_backend_cfg or {}).get("fast_answers", True))
        try:
            ce.refresh_backend_config()
        except Exception:
            pass
        # Personal Specialists — config-only registry, persists to vault/specialists.json
        # Seeds 3 defaults (Sales / Inventory / Customer) on first run.
        self.specialists = _spec.SpecialistRegistry(VAULT_DIR)

        # ── Read/write split for user data ─────────────────────────
        # Inputs land in vault/data_in/ (read-only by the app — it
        # never overwrites or deletes anything in there). Derived
        # outputs go to vault/data_out/. The DataIndex constructor
        # validates the two never overlap.
        data_index.init_data_dirs(VAULT_DIR)
        # Sweep: an earlier version of the migration helper copied
        # app-internal config files (specialists.json, node_registry.json,
        # etc.) into data_in/. Clean them out here so the dropdown stays
        # showing only real user data.
        try:
            cleaned = data_index.cleanup_misplaced_internals(VAULT_DIR)
            if cleaned:
                print(f"[DataIndex] Removed {len(cleaned)} stray app-config "
                      f"file(s) from data_in/")
        except Exception as e:
            print(f"[DataIndex] Cleanup skipped: {e}")
        # One-time migration: copy any loose user CSV/JSON at the vault
        # root into data_in/ so they're discoverable. Originals stay
        # put — we never silently move user data. App-internal config
        # filenames are excluded.
        try:
            migrated = data_index.migrate_loose_vault_files(VAULT_DIR)
            if migrated:
                print(f"[DataIndex] Copied {len(migrated)} loose data file(s) "
                      f"from vault root into data_in/")
        except Exception as e:
            print(f"[DataIndex] Migration skipped: {e}")

        self.data_index = data_index.DataIndex(
            search_roots=[
                data_index.input_dir(VAULT_DIR),
                data_index.bundled_samples_dir(),
            ],
            write_root=data_index.output_dir(VAULT_DIR),
        )
        # Share this warm instance with the module-level context injector so
        # its cell-value search stage reuses the same (refreshed) index.
        _register_data_index(self.data_index)
        self.librarian = ce.Librarian(VAULT_DIR, LOG_PATH)
        self.runner = ce.LocalRunner(WORKSPACE_DIR)
        self.speech = ce.SpeechToText(model_size="base")
        self.dispatcher = ce.build_dispatcher()

        pins = ce.load_personality_pins(PINS_PATH)
        # Persistent council-wide instructions — managed as a list
        self._instr_mgr       = InstructionManager(INSTRUCTIONS_PATH)
        self._content_style  = ContentStyleManager(CONTENT_STYLE_PATH)

        self.personalities = ce.build_personalities(
            pins=pins, vault_dir=VAULT_DIR, session_id=self.session_id,
            trace=True, dispatcher=self.dispatcher,
            prior_session_id=self.prior_session_id,
        )
        self._unpack_personalities()

        # ── Sage agent (vault knowledge + gap detection) ────────────
        self.sage_agent_obj = None
        if _SAGE_OK and self.sage is not None:
            _sage_kb = sa.SageKnowledge(VAULT_DIR / "sage_knowledge")
            self.sage_agent_obj = sa.SageAgent(
                model=self.sage,
                knowledge=_sage_kb,
                on_gap=lambda q, r: self.ui_q.put((
                    "agent_phase", "sage_gap",
                    f"⚠ Sage gap: {q[:80]} — {r}"
                )) if hasattr(self, "ui_q") else None,
            )
            # Inject Sage system prompt
            self.sage.system_prompt = sa.SAGE_SYSTEM_PROMPT

        # ── Dream3D primer injection ───────────────────────────
        if _DREAM3D_OK:
            try:
                d3p.inject_dream3d_context(self.personalities)
            except Exception as e:
                print(f"[Dream3D] Primer injection failed: {e}")

        # ── Dream3D domain patch ───────────────────────────────
        if _D3D_PATCH_OK:
            d3d_patch.patch_personalities(self.personalities, VAULT_DIR)
        else:
            print("[Council] dream3d_council_patch.py not found — Dream3D expertise not injected")

        # NOTE: self.apoth is no longer initialised here — it moved into
        # _build_apoth_tab so that consumer builds (which don't build the
        # tab) never pay the import or init cost for the Apothecary
        # engine. The advanced-mode branch in _build_ui calls
        # _build_apoth_tab which sets self.apoth before the tab is added.

        # RAM-resident "task memo" — a short sticky note carrying the
        # user's original goal + constraints + forbidden actions across
        # every turn this session. Re-injected at the TOP of the prompt
        # on every query so small models (4K-8K context) don't forget
        # what was asked once vault/file context pushes the user message
        # past the window's tail. See task_memory.py for the design.
        self.task_memory = _task_memory_mod.TaskMemory()

        self.current_script_name = "script"
        self._stream_buffers: Dict[str, str] = {}  # role -> partial streamed text
        # Set while ≥1 token was inserted into the stream box during the
        # current _poll_ui_queue drain. The expensive see("end") (which
        # forces a full line-geometry recalc) is deferred to ONCE per
        # drain instead of once per token — at 100+ tok/s that turns
        # ~100 layout passes/sec into ~20, eliminating streaming jank.
        self._stream_box_dirty = False
        self._node_refresh_id = None

        # ── Agents ────────────────────────────────────────────
        def _agent_event_cb(phase: str, msg: str):
            self.ui_q.put(("agent_phase", phase, msg))

        if _CODER_AGENT_OK:
            self.coder_agent = ca.CoderAgent(
                personality_model=self.coder,
                runner=self.runner,
                max_attempts=8,
                event_callback=_agent_event_cb,
            )
        else:
            self.coder_agent = None

        if _INTERN_AGENT_OK:
            self.intern_agent = ia.InternAgent(
                personality_model=self.intern,
                event_callback=_agent_event_cb,
                max_research_pages=3,
            )
        else:
            self.intern_agent = None

        self._pump_splash()

        # ── RAG ───────────────────────────────────────────────
        if _RAG_OK:
            if self._interactive_host:
                # Off-main-thread torch/CUDA init segfaults the kernel under
                # Spyder/IPython, so construct eagerly on the MAIN thread here
                # (the background index thread never starts under an interactive
                # host, so deferring construction would leave self.rag = None).
                self.rag = vr.VaultRAG(
                    vault_dir=VAULT_DIR, chroma_dir=VAULT_DIR / ".chromadb")
                print("[startup] RAG auto-index skipped under interactive "
                      "host (avoids off-main-thread torch crash). Use the "
                      "Vault tab to build it if you need semantic search.",
                      flush=True)
            else:
                # Construct AND index on the background thread: the ~6.7s
                # SentenceTransformer load + chromadb PersistentClient open used
                # to run here on the main thread, blocking the window from
                # appearing. Every self.rag consumer already guards on
                # `if self.rag`, so None during the brief build window is safe.
                self.rag = None
                threading.Thread(target=self._init_rag_index,
                                 daemon=True).start()
        else:
            self.rag = None

        self._pump_splash()
        self._build_ui()
        self._apply_dark_theme()
        self._pump_splash()
        self.after(100, self._poll_ui_queue)
        self.after(2000, self._refresh_nodes_async)   # initial node probe
        self._start_config_watcher()                   # T1-E: hot-reload pins.json
        # Startup chain (sequenced so dialogs don't pile up):
        #   1) Onboarding wizard (first launch only)
        #   2) Licensing check (every launch — opens activation dialog if
        #      trial expired and no license)
        #   3) Crash recovery prompt (if the previous run died mid-flight)
        # 500ms delay lets the main window draw before the first modal.
        self.after(500, lambda: onboarding.run_if_needed(
            self, VAULT_DIR,
            on_complete=lambda _ok: self.after(400, self._check_license_status),
        ))

        # Incremental data-stats precompute, deferred + on a daemon thread
        # so it never competes with startup or the model. CPU/IO only (no
        # GPU), streamed in chunks. Slow the first sweep on a big vault,
        # near-instant after — only unprocessed files are touched, so it
        # keeps pace as data grows. Disable with COUNCIL_STATS_PRECOMPUTE=0.
        if os.environ.get("COUNCIL_STATS_PRECOMPUTE", "1").strip().lower() \
                not in ("0", "false", "no", "off"):
            self.after(8000, lambda: threading.Thread(
                target=self._init_stats_index, daemon=True).start())

        # Interactive host: no splash will reveal us, so show the finished
        # window now (it was withdrawn at the top of __init__ to avoid the
        # construction flicker). main() does the splash-timed reveal for
        # the normal launch path.
        if self._interactive_host:
            try:
                self.deiconify()
                self.lift()
            except Exception:
                pass

    def _unpack_personalities(self):
        # Required core personalities — missing any of these is a config error
        _required = ("judge", "writer", "peasant", "intern", "coder", "artist")
        _missing = [r for r in _required if r not in self.personalities]
        if _missing:
            raise RuntimeError(
                f"Missing required personalities: {', '.join(_missing)}. "
                f"Check personality_config.yaml and personality_backends.json."
            )
        self.judge                = self.personalities["judge"]
        self.writer               = self.personalities["writer"]
        self.peasant              = self.personalities["peasant"]
        self.intern               = self.personalities["intern"]
        self.coder                = self.personalities["coder"]
        self.artist               = self.personalities["artist"]
        self.skeptic              = self.personalities.get("skeptic")
        self.sage                 = self.personalities.get("sage")
        self.strategist           = self.personalities.get("strategist")
        # Librarian personality — interprets vault search results for other agents
        self.librarian_personality = self.personalities.get("librarian")
        self.content              = self.personalities.get("content")
        self.director             = self.personalities.get("director")
        self.algorithm            = self.personalities.get("algorithm")
        self.coach                = self.personalities.get("coach")

    # ============================
    # Licensing & trial gate
    # ============================

    def _check_license_status(self):
        """
        Status check at startup. In DEMO_MODE this short-circuits to
        "personal use — no gates"; in product mode it can open a
        modal activation dialog. Crash recovery + update check run
        either way.
        """
        try:
            import device_fingerprint
            fp = device_fingerprint.compute(VAULT_DIR)
            status = licensing.get_status(VAULT_DIR, fingerprint=fp)
        except Exception as e:
            print(f"[License] Status check failed: {e}")
            self.after(400, self._check_crash_recovery)
            self.after(2000, self._kick_update_check)
            return

        self._license_status = status

        # Update the badge in the UI if the Council action bar already exists
        self._refresh_license_badge()

        # DEMO_MODE: never block, never show the activation dialog.
        if not getattr(branding, "DEMO_MODE", False):
            if status["status"] in (licensing.STATUS_TRIAL_EXPIRED,
                                    licensing.STATUS_LICENSE_EXPIRED,
                                    licensing.STATUS_INVALID_LICENSE,
                                    licensing.STATUS_NEEDS_ACTIVATION):
                activation_dialog.open_activation_dialog(
                    self, VAULT_DIR,
                    on_status_change=self._on_license_status_change,
                    blocking=True,
                )
        # Crash recovery runs whether or not the user activated
        self.after(400, self._check_crash_recovery)
        # Update check runs in the background — never blocks startup.
        # DEMO_MODE forces the manifest URL empty so this is a fast no-op.
        self.after(2000, self._kick_update_check)

    def _kick_update_check(self):
        """Spawn a background update check. Notification fires later if needed."""
        def _on_update(info: dict):
            # Bounce to UI thread
            self.after(0, lambda i=info: self._show_update_notification(i))
        try:
            updater.check_async(VAULT_DIR, _on_update, delay_seconds=2.0)
        except Exception as e:
            print(f"[Updater] Check failed silently: {e}")

    def _show_update_notification(self, info: dict):
        """Non-blocking update toast — user can act on it any time."""
        from tkinter import messagebox
        version = info.get("version", "(unknown)")
        notes_url = info.get("release_notes_url", "")
        choice = messagebox.askyesnocancel(
            "Update available",
            f"{branding.PRODUCT_NAME} {version} is available "
            f"(you're on {branding.VERSION}).\n\n"
            f"Yes  — open the download page in your browser\n"
            f"No   — skip this version (you'll be reminded for the next)\n"
            f"Cancel — remind me again next launch\n\n"
            f"Note: {branding.PRODUCT_NAME} works fully offline. Updates "
            f"are optional.",
            parent=self,
        )
        if choice is True:
            target = info.get("platform_url") or info.get("download_url") or notes_url
            try:
                import webbrowser
                webbrowser.open(target)
            except Exception:
                pass
        elif choice is False:
            updater.skip_version(VAULT_DIR, version)

    def _on_license_status_change(self, new_status: dict):
        """Callback after the user activates/deactivates from the dialog."""
        self._license_status = new_status
        self._refresh_license_badge()

    def _refresh_license_badge(self):
        """Update the small status indicator in the Council action bar."""
        if not hasattr(self, "_license_badge_var"):
            return
        st = getattr(self, "_license_status", None) or {}
        self._license_badge_var.set(st.get("message", ""))

    def _can_run_deliberation(self) -> bool:
        """Gate: returns True when new deliberations are permitted."""
        st = getattr(self, "_license_status", None)
        if not st:
            return True   # status not yet computed — fail-open during startup
        return bool(st.get("can_use_full_features", True))

    # ============================
    # Crash recovery
    # ============================

    def _check_crash_recovery(self):
        """
        On startup, check for sessions that were in-flight when the app last
        exited (sentinel `.active` file present). Offer to resume the most
        recent one. Older orphans are cleared silently.
        """
        try:
            orphans = self.convo_store.find_orphaned_sessions()
        except Exception:
            return
        if not orphans:
            return

        # Most recent orphan is candidate; older ones get cleared.
        target = orphans[0]
        for stale in orphans[1:]:
            try:
                self.convo_store.clear_orphan(stale.get("session_id", ""))
            except Exception:
                pass

        sid = target.get("session_id", "")
        query = target.get("query", "") or "(no query recorded)"
        started = target.get("started_at", "")

        from tkinter import messagebox
        choice = messagebox.askyesnocancel(
            "Recover unfinished session?",
            f"Council appears to have closed during a deliberation.\n\n"
            f"Started:  {started}\n"
            f"Question: {query[:160]}\n\n"
            f"Yes  → Load that session and continue\n"
            f"No   → Discard the unfinished work\n"
            f"Cancel → Decide later (will ask again next launch)",
            parent=self,
        )
        if choice is True:
            # Load the orphaned session as the active one
            try:
                self.session_id = sid
                if hasattr(self, "_load_session_into_transcript"):
                    self._load_session_into_transcript(sid)
                self._append_transcript(
                    "Council",
                    f"⏪ Resumed session {sid}. Re-send your last question if "
                    f"you want the panel to continue.",
                    "observation",
                )
            except Exception as e:
                print(f"[CrashRecovery] Failed to resume {sid}: {e}")
                self.convo_store.clear_orphan(sid)
        elif choice is False:
            self.convo_store.clear_orphan(sid)
        # Cancel → leave sentinel; next launch asks again.

    # ============================
    # Dark theme
    # ============================

    def _apply_dark_theme(self):
        """
        Apply the active theme palette from branding.get_theme(). Pulls
        every value from the central palette so a recolour is a one-file
        edit.
        """
        t = branding.get_theme("dark")
        bg       = t["bg"]
        fg       = t["fg"]
        abg      = t["panel_bg"]      # frame / widget bg
        ibg      = t["input_bg"]      # text-entry bg
        sel      = t["selection_bg"]
        bord     = t["border"]
        accent   = t["accent"]
        # Expose to the rest of the GUI for inline tk.Frame / canvas
        # widgets that aren't ttk-styled.
        self._theme = t

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", background=bg, foreground=fg, fieldbackground=ibg,
                         insertcolor=fg, troughcolor=abg, bordercolor=bord)
        style.configure("TNotebook", background=bg, borderwidth=0)
        style.configure("TNotebook.Tab", background=abg, foreground=fg, padding=[10, 4])
        style.map("TNotebook.Tab",
                  background=[("selected", sel)],
                  foreground=[("selected", fg)])
        style.configure("TFrame", background=bg)
        style.configure("TLabel", background=bg, foreground=fg)
        style.configure("TLabelframe", background=bg, foreground=fg, bordercolor=bord)
        style.configure("TLabelframe.Label", background=bg, foreground=fg)
        style.configure("TButton", background=abg, foreground=fg, relief="flat",
                        padding=4, bordercolor=bord)
        style.map("TButton",
                  background=[("active", sel), ("pressed", accent)])
        style.configure("TCheckbutton", background=bg, foreground=fg)
        style.configure("TRadiobutton",  background=bg, foreground=fg)
        style.configure("TEntry",  fieldbackground=ibg, foreground=fg, insertcolor=fg)
        style.configure("TCombobox", fieldbackground=ibg, foreground=fg)
        style.configure("TScrollbar", background=abg, troughcolor=ibg, bordercolor=bord)
        style.configure("Treeview", background=ibg, foreground=fg, fieldbackground=ibg,
                         rowheight=24, bordercolor=bord)
        style.map("Treeview", background=[("selected", sel)])
        style.configure("Treeview.Heading", background=abg, foreground=fg,
                        bordercolor=bord)
        self.configure(bg=bg)

    def _make_text(self, parent, **kwargs) -> tk.Text:
        defaults = dict(
            bg="#231a1a", fg="#d4d4d4", insertbackground="#d4d4d4",
            selectbackground="#5a3030", relief="flat", bd=0,
            font=("Consolas", 10),
        )
        defaults.update(kwargs)
        return tk.Text(parent, **defaults)

    # ============================
    # UI construction
    # ============================

    def _init_rag_index(self):
        """Construct (if needed) + index the vault RAG on a background thread at
        startup, so the heavy SentenceTransformer/chromadb init doesn't block
        the window. Safe to call again from the Vault tab's re-index button."""
        try:
            if self.rag is None and _RAG_OK:
                self.rag = vr.VaultRAG(
                    vault_dir=VAULT_DIR, chroma_dir=VAULT_DIR / ".chromadb")
            if self.rag:
                stats = self.rag.index()
                self.ui_q.put(("agent_phase", "rag_index",
                               f"Vault indexed: {stats.total_files} files, "
                               f"{stats.total_chunks} chunks ({stats.backend})"))
        except Exception as e:
            self.ui_q.put(("agent_phase", "rag_index", f"RAG index error: {e}"))

    def _build_stats_index(self, *, on_progress=None) -> dict:
        """Precompute column stats for any UNPROCESSED CSVs under data_in
        (incremental — only files not already cached at their current
        mtime). Pure pandas / CPU, streamed in chunks so it's memory-safe
        on big files. Slow the first time, near-instant afterwards;
        new files added later are picked up on the next call. Returns the
        sweep counts ({seen, processed, already_current})."""
        import stats_cache as _sc
        import vault_analyst as _va
        try:
            data_in = data_index.input_dir(VAULT_DIR)
        except Exception:
            data_in = VAULT_DIR
        cache = _sc.StatsCache(VAULT_DIR)
        return cache.process_unprocessed(
            data_in, list_files=_va.list_csv_files, on_progress=on_progress)

    def _init_stats_index(self):
        """Background-thread entry: incremental stats sweep with the
        result posted to the activity feed. Failures are swallowed — a
        stats precompute must never break the app."""
        try:
            res = self._build_stats_index()
            if res.get("processed"):
                self.ui_q.put(("agent_phase", "stats_index",
                               f"Data stats: processed {res['processed']} new "
                               f"file(s); {res['already_current']} already "
                               f"cached ({res['seen']} CSVs total)."))
        except Exception as e:
            self.ui_q.put(("agent_phase", "stats_index", f"stats index error: {e}"))

    def _pump_splash(self):
        """Advance the loading-cog one frame during blocking construction.
        No-op once the splash is gone. Cheap (one frame + redraw, no
        event dispatch) so peppering build steps with it is safe."""
        sp = getattr(self, "_splash", None)
        if sp is not None:
            sp.pump()

    def _build_ui(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=6, pady=6)

        # ── Customer-facing flow (in order of how a session naturally goes) ──
        # 1) Ask a question     → Council (entrypoint — most users start here)
        # 2) See the data       → Grapher (Council pulls relevant files in)
        # 3) Tune the experts   → Personal Specialists
        # 4) Second opinion     → Lens
        # 5) Re-visit           → Sessions
        # 6) Manage data        → Vault
        # 7) Speech I/O         → Speech
        # Pump the loading-cog after each tab so it visibly turns while
        # the Tk event loop is blocked building widgets.
        for _build in (self._build_council_tab,
                       self._build_dream3d_tab,
                       self._build_grapher_tab,
                       self._build_specialists_tab,
                       self._build_model_finder_tab,
                       self._build_lens_tab,
                       self._build_sessions_tab,
                       self._build_vault_manager_tab,
                       self._build_agent_jobs_tab,
                       self._build_tool_forge_tab,
                       self._build_speech_tab,
                       self._build_changelog_tab,
                       self._build_diagnostics_tab):
            _build()
            self._pump_splash()

        # ── Advanced / admin tabs ──
        # Hidden by default unless explicitly enabled. Power users and
        # support staff can launch with --advanced or set
        # COUNCIL_ADVANCED=1 to expose the IDE, Agents panel, Librarian
        # snapshots, Nodes, Vault Health, and Apothecary tabs.
        if _ADVANCED_MODE:
            for _build in (self._build_ide_tab,
                           self._build_librarian_tab,
                           self._build_agents_tab,
                           self._build_nodes_tab,
                           self._build_vault_health_tab,
                           self._build_apoth_tab):
                _build()
                self._pump_splash()

    # ---- Backend strip (model backend selector) ----

    _BACKEND_SETTINGS_FILENAME = "backend_settings.json"

    def _backend_settings_path(self):
        return VAULT_DIR / self._BACKEND_SETTINGS_FILENAME

    def _load_backend_settings(self):
        """Read persisted GGUF model path from vault/backend_settings.json and
        apply to env so council_engine picks it up at import.

        Env-var precedence: if COUNCIL_GGUF_PATH is ALREADY set in the
        environment (typically by run-windows.bat / run-linux.sh / a
        shell export), we DO NOT clobber it with the persisted JSON.
        This matches the documented precedence in
        ``onboarding.load_gguf_path`` (env wins over JSON). Previously
        the JSON unconditionally won, which broke users who moved their
        model file and overrode the launch path via the env var — the
        stale JSON path silently took effect and Llama() failed to
        find the model.
        """
        p = self._backend_settings_path()
        if not p.exists():
            return {}
        try:
            import json as _j
            data = _j.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
        if data.get("gguf_path") and not os.environ.get("COUNCIL_GGUF_PATH", "").strip():
            os.environ["COUNCIL_GGUF_PATH"] = str(data["gguf_path"])
        # Apply persisted engine knobs (set from the Engine settings dialog)
        # to the environment, with the same "env always wins" precedence —
        # a shell/launcher export overrides the saved value.
        for _key, _env in (("n_ctx", "COUNCIL_GGUF_N_CTX"),
                           ("gpu_layers", "COUNCIL_GGUF_GPU_LAYERS"),
                           ("embed_device", "COUNCIL_EMBED_DEVICE")):
            _val = data.get(_key)
            if _val not in (None, "") and not os.environ.get(_env, "").strip():
                os.environ[_env] = str(_val)
        return data

    def _save_backend_settings(self):
        """Persist the user's GGUF path. MERGES with the existing file so
        the wizard's informational keys (model_id, model_org) survive a
        Browse-driven model switch. Vision-specific keys (clip_path)
        are EXPLICITLY cleared because the Browse button selects a
        single text-only .gguf — if the user wants vision, they go
        through the wizard which knows to pair a model with an mmproj.
        Leaving a stale clip_path would trigger the vision-attach
        retry warning on every subsequent load.
        """
        import json as _j
        path = self._backend_settings_path()
        existing: dict = {}
        if path.exists():
            try:
                existing = _j.loads(path.read_text(encoding="utf-8"))
                if not isinstance(existing, dict):
                    existing = {}
            except Exception:
                existing = {}
        existing["gguf_path"] = self._gguf_path_var.get()
        # Clear vision state — Browse is text-only by contract.
        existing.pop("clip_path", None)
        try:
            path.write_text(_j.dumps(existing, indent=2), encoding="utf-8")
        except Exception as _e:
            print(f"[backend strip] could not save settings: {_e}")

    def _build_backend_strip(self, parent):
        """Top-of-tab row: just a GGUF file label + Browse button. Ollama
        support was removed; the council only runs on local .gguf files."""
        strip = ttk.Frame(parent)
        strip.pack(fill="x", padx=6, pady=(6, 0))

        ttk.Label(strip, text="GGUF model:").pack(side="left", padx=(0, 4))
        self._gguf_path_var = tk.StringVar(
            value=os.environ.get("COUNCIL_GGUF_PATH", "")
        )
        self._gguf_path_label = ttk.Label(
            strip, textvariable=self._gguf_path_var, foreground="#888",
            anchor="w",
        )
        self._gguf_path_label.pack(side="left", padx=(0, 6), fill="x", expand=True)
        ttk.Button(strip, text="Browse...",
                   command=self._browse_gguf_file).pack(side="left")
        ttk.Button(strip, text="⚙ Engine",
                   command=self._open_engine_settings).pack(side="left", padx=(6, 0))

    def _open_engine_settings(self):
        """Dialog to set the local-engine env knobs from the app: max
        context (COUNCIL_GGUF_N_CTX), GPU layers (COUNCIL_GGUF_GPU_LAYERS),
        and embedding device (COUNCIL_EMBED_DEVICE). Applied live + persisted
        so they survive a restart — no shell exports needed."""
        import tkinter as tk
        from tkinter import ttk, messagebox
        win = tk.Toplevel(self)
        win.title("Engine settings")
        win.transient(self)
        win.resizable(False, False)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, foreground="#888", justify="left",
                  text="Tune the local model engine. Blank = auto / default.\n"
                       "Context + GPU layers reload the model on the next "
                       "message.").grid(row=0, column=0, columnspan=2,
                                        sticky="w", pady=(0, 10))

        ttk.Label(frm, text="Max context (n_ctx):").grid(
            row=1, column=0, sticky="w", pady=3)
        nctx_var = tk.StringVar(value=os.environ.get("COUNCIL_GGUF_N_CTX", ""))
        ttk.Entry(frm, textvariable=nctx_var, width=12).grid(
            row=1, column=1, sticky="w")
        ttk.Label(frm, foreground="#888",
                  text="e.g. 8192, 16384, 32768. Blank = auto-detect "
                       "(falls back to 4096).").grid(
            row=2, column=1, sticky="w")

        ttk.Label(frm, text="GPU layers:").grid(
            row=3, column=0, sticky="w", pady=3)
        gpu_var = tk.StringVar(
            value=os.environ.get("COUNCIL_GGUF_GPU_LAYERS", ""))
        ttk.Entry(frm, textvariable=gpu_var, width=12).grid(
            row=3, column=1, sticky="w")
        ttk.Label(frm, foreground="#888",
                  text="99 = all on GPU, 0 = CPU only. Blank = default (99).").grid(
            row=4, column=1, sticky="w")

        ttk.Label(frm, text="Embedding device:").grid(
            row=5, column=0, sticky="w", pady=3)
        cur_embed = os.environ.get("COUNCIL_EMBED_DEVICE", "").strip() or "auto"
        embed_var = tk.StringVar(value=cur_embed)
        ttk.Combobox(frm, textvariable=embed_var, width=10, state="readonly",
                     values=["auto", "cpu", "cuda"]).grid(
            row=5, column=1, sticky="w")
        ttk.Label(frm, foreground="#888",
                  text="cpu avoids GPU VRAM contention (recommended on WSL).").grid(
            row=6, column=1, sticky="w")

        fast_var = tk.BooleanVar(
            value=bool(getattr(self, "_fast_answers_enabled", True)))
        ttk.Checkbutton(
            frm, text="Fast direct answers (skip council for instant answers)",
            variable=fast_var).grid(row=7, column=0, columnspan=2,
                                    sticky="w", pady=(8, 0))
        ttk.Label(frm, foreground="#888", justify="left",
                  text="File counts, summaries, derived results and "
                       "collections answer instantly with no model call.\n"
                       "Use “⤢ Expand with council” on any answer for a "
                       "fuller discussion.").grid(
            row=8, column=0, columnspan=2, sticky="w")

        status = tk.StringVar(value="")
        ttk.Label(frm, textvariable=status, foreground="#a6e3a1").grid(
            row=9, column=0, columnspan=2, sticky="w", pady=(8, 0))

        def _apply():
            n = nctx_var.get().strip()
            g = gpu_var.get().strip()
            e = embed_var.get().strip()
            for label, val in (("Max context", n), ("GPU layers", g)):
                if val and not val.isdigit():
                    messagebox.showwarning(
                        "Invalid value",
                        f"{label} must be a whole number (or left blank).")
                    return
            self._apply_engine_settings(
                n_ctx=n, gpu_layers=g,
                embed_device=("" if e == "auto" else e),
                fast_answers=bool(fast_var.get()))
            status.set("Applied + saved. Model reloads on the next message.")

        btns = ttk.Frame(frm)
        btns.grid(row=10, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(btns, text="Apply", command=_apply).pack(side="right")
        ttk.Button(btns, text="Close", command=win.destroy).pack(
            side="right", padx=6)

    def _apply_engine_settings(self, *, n_ctx: str = "", gpu_layers: str = "",
                               embed_device: str = "",
                               fast_answers: bool = None):
        """Set the engine env knobs (or clear them when blank), persist to
        backend_settings.json, and reset the engine so context / GPU-layer
        changes take effect on the next inference."""
        def _set(env: str, val: str):
            if val:
                os.environ[env] = str(val)
            else:
                os.environ.pop(env, None)
        _set("COUNCIL_GGUF_N_CTX", n_ctx)
        _set("COUNCIL_GGUF_GPU_LAYERS", gpu_layers)
        _set("COUNCIL_EMBED_DEVICE", embed_device)
        if fast_answers is not None:
            self._fast_answers_enabled = bool(fast_answers)

        import json as _j
        path = self._backend_settings_path()
        existing: dict = {}
        if path.exists():
            try:
                existing = _j.loads(path.read_text(encoding="utf-8"))
                if not isinstance(existing, dict):
                    existing = {}
            except Exception:
                existing = {}
        existing["n_ctx"] = n_ctx
        existing["gpu_layers"] = gpu_layers
        existing["embed_device"] = embed_device
        if fast_answers is not None:
            existing["fast_answers"] = bool(fast_answers)
        try:
            path.write_text(_j.dumps(existing, indent=2), encoding="utf-8")
        except Exception as _e:
            print(f"[engine settings] could not save: {_e}")

        # Reset the GGUF singleton so n_ctx / gpu_layers are re-read on the
        # next call (they're applied at model load, not per-inference).
        try:
            import council_engine as _ce
            # The user explicitly chose a GPU-layer setting — clear any
            # GPU-crash sentinel so a GPU retry is actually attempted
            # (the auto-CPU fallback only kicks in after an UNconfirmed
            # crash, not after a deliberate setting change).
            try:
                _ce.gpu_clear_attempt()
            except Exception:
                pass
            _ce.refresh_backend_config()
        except Exception as _e:
            print(f"[engine settings] backend refresh failed: {_e}")

    def _browse_gguf_file(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select a .gguf model file",
            filetypes=[("GGUF model files", "*.gguf"), ("All files", "*.*")],
        )
        if not path:
            return
        self._gguf_path_var.set(path)
        os.environ["COUNCIL_GGUF_PATH"] = path
        try:
            import council_engine as _ce
            _ce.refresh_backend_config()
        except Exception as _e:
            print(f"[backend strip] refresh failed: {_e}")
        self._save_backend_settings()
        self._append_transcript(
            "Council",
            f"GGUF model set: {Path(path).name}",
            "observation",
        )

    # ---- Pipeline intent handler ----

    # Patterns matched against the user's first message line.
    _PIPELINE_SHOW_RE = _re.compile(
        r"^\s*(?:show|display|view|render|render\s+the|view\s+the)"
        r"(?:\s+(?:me|the))?\s+pipelines?\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_LIST_RE = _re.compile(
        r"^\s*(?:list|show|what)\s+(?:are\s+)?"
        r"(?:my\s+|the\s+)?(?:available\s+)?pipelines?\??\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_MODIFY_RE = _re.compile(
        r"^\s*(?:modify|edit|change|update|tweak)"
        r"(?:\s+(?:the|my))?\s+pipelines?\s+(.+?)\s+(?:to|so\s+that|so|with|by|using)\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_EXPLAIN_RE = _re.compile(
        r"^\s*(?:explain|describe|what\s+does)"
        r"(?:\s+(?:the|my))?\s+pipelines?\s+(.+?)\s*(?:do)?\s*[.?!]?\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_VALIDATE_RE = _re.compile(
        r"^\s*(?:validate|check|verify)"
        r"(?:\s+(?:the|my))?\s+pipelines?\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_GRAPH_RE = _re.compile(
        r"^\s*(?:graph|data\s*flow|dependencies\s+of)"
        r"(?:\s+(?:the|my|for))?\s+pipelines?\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_CREATE_RE = _re.compile(
        r"^\s*(?:create|generate|make|write)"
        r"(?:\s+(?:me|a))?\s+(?:new\s+)?pipelines?\s+(?:that\s+|to\s+|which\s+)?(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _PIPELINE_EXPORT_RE = _re.compile(
        r"^\s*(?:export|save)\s+pipelines?\s+(.+?)\s+(?:as|to)\s+"
        r"(?:markdown|md|a\s+markdown\s+file)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _DOWNLOAD_HF_RE = _re.compile(
        r"^\s*(?:download|fetch|pull)\s+(?:from\s+)?(?:huggingface\s+|hf\s+)?"
        r"([A-Za-z0-9_\-./]+)\s+(\S+\.gguf)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _PEEK_FILE_RE = _re.compile(
        r"^\s*(?:peek|preview|peek\s+at|preview\s+of)\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _COMPARE_PIPELINES_RE = _re.compile(
        r"^\s*(?:compare|diff)\s+pipelines?\s+(.+?)\s+(?:and|vs\.?|versus|to|with)\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )

    def _handle_pipeline_intent(self, user_text: str) -> bool:
        """Detect and handle pipeline show/list/modify intents.

        Returns True if the message was a pipeline command (in which case
        the normal council flow is skipped). Returns False otherwise — the
        message falls through to the regular deliberation.
        """
        if not user_text:
            return False
        single_line = user_text.split("\n", 1)[0]

        # List
        if self._PIPELINE_LIST_RE.match(single_line):
            self._pipeline_list_response()
            return True

        # Show / display
        m = self._PIPELINE_SHOW_RE.match(single_line)
        if m:
            name = m.group(1).strip().strip("'\"`")
            self._pipeline_show_response(name)
            return True

        # Modify
        m = self._PIPELINE_MODIFY_RE.match(single_line)
        if m:
            name = m.group(1).strip().strip("'\"`")
            change = m.group(2).strip().strip("'\"`.")
            self._pipeline_modify_response(name, change, user_text)
            return True

        # Explain
        m = self._PIPELINE_EXPLAIN_RE.match(single_line)
        if m:
            self._pipeline_explain_response(m.group(1).strip().strip("'\"`"))
            return True

        # Validate
        m = self._PIPELINE_VALIDATE_RE.match(single_line)
        if m:
            self._pipeline_validate_response(m.group(1).strip().strip("'\"`"))
            return True

        # Dependency graph
        m = self._PIPELINE_GRAPH_RE.match(single_line)
        if m:
            self._pipeline_graph_response(m.group(1).strip().strip("'\"`"))
            return True

        # Create (generation)
        m = self._PIPELINE_CREATE_RE.match(single_line)
        if m:
            description = m.group(1).strip().strip("'\"`.")
            self._pipeline_create_response(description, user_text)
            return True

        # Export to markdown
        m = self._PIPELINE_EXPORT_RE.match(single_line)
        if m:
            self._pipeline_export_markdown(m.group(1).strip().strip("'\"`"))
            return True

        # HF download (handled here since it's chat-driven and Dream3D-related
        # in spirit — pulling a GGUF model)
        m = self._DOWNLOAD_HF_RE.match(single_line)
        if m:
            repo, filename = m.group(1).strip(), m.group(2).strip()
            self._hf_download_response(repo, filename)
            return True

        # Peek at a file — preview WITHOUT injecting into the model prompt
        m = self._PEEK_FILE_RE.match(single_line)
        if m:
            self._peek_file_response(m.group(1).strip().strip("'\"`"))
            return True

        # Compare two pipelines
        m = self._COMPARE_PIPELINES_RE.match(single_line)
        if m:
            self._compare_pipelines_response(
                m.group(1).strip().strip("'\"`"),
                m.group(2).strip().strip("'\"`"),
            )
            return True

        return False

    def _peek_file_response(self, target: str):
        """Show file contents in the transcript ONLY — never injected into
        a model prompt. Lets the user verify a file's structure for free."""
        # Resolve the target: explicit path, then vault search
        p = Path(target).expanduser()
        if not p.is_file():
            # Try to find under vault by name
            import pipeline_scanner as _ps  # piggy-back on its scan helpers
            from vault_analyst import list_data_files, list_parquet_files
            try:
                candidates = list(list_data_files(VAULT_DIR)) \
                           + list(list_parquet_files(VAULT_DIR))
            except Exception:
                candidates = []
            q = target.lower()
            matches = [c for c in candidates if q in c.name.lower()]
            if matches:
                p = matches[0]
        if not p.is_file():
            self._append_transcript(
                "Writer", f"No file found for '{target}'.", "final",
            )
            self._set_status("● idle")
            return

        # Defense-in-depth: never peek into protected paths
        try:
            import conversation_logger as _cl
            if _cl.is_protected_path(p, VAULT_DIR):
                self._append_transcript(
                    "Writer",
                    f"'{p.name}' is in a protected vault folder — "
                    f"not viewable.",
                    "final",
                )
                self._set_status("● idle")
                return
        except Exception:
            pass

        block = _read_file_for_injection(str(p))
        if not block:
            self._append_transcript(
                "Writer",
                f"Couldn't preview {p.name} — unsupported format or read error.",
                "final",
            )
        else:
            # Strip the [FILE: ...] / [END FILE] wrapper — that exists to
            # mark prompt injection; for a transcript preview it's noise.
            inner = block
            if inner.startswith("[FILE:"):
                inner = inner.split("\n", 1)[1] if "\n" in inner else inner
            if inner.endswith("[END FILE]"):
                inner = inner.rsplit("\n", 1)[0]
            self._append_transcript(
                "Writer",
                f"Preview of {p.name} (not sent to the model):\n\n{inner}",
                "final",
            )
        self._set_status("● idle")

    def _compare_pipelines_response(self, a_name: str, b_name: str):
        import pipeline_scanner as _ps
        a = _ps.find_pipeline_by_name(VAULT_DIR, a_name)
        b = _ps.find_pipeline_by_name(VAULT_DIR, b_name)
        if not a or not b:
            missing = [n for n, pl in [(a_name, a), (b_name, b)] if pl is None]
            self._append_transcript(
                "Writer",
                f"Could not find: {', '.join(missing)}. Type 'list pipelines'.",
                "final",
            )
            self._set_status("● idle")
            return
        diff = _ps.compare_pipelines(a, b)
        self._append_transcript("Writer", diff, "final")
        self._set_status("● idle")

    def _pipeline_export_markdown(self, name: str):
        import pipeline_scanner as _ps
        pl = _ps.find_pipeline_by_name(VAULT_DIR, name)
        if not pl:
            self._append_transcript(
                "Writer", f"No pipeline matching '{name}'.", "final",
            )
            self._set_status("● idle")
            return
        try:
            md = _ps.export_pipeline_to_markdown(pl)
            out_dir = data_index.output_dir(VAULT_DIR)
            out_path = out_dir / f"{Path(pl.name).stem}.md"
            n = 2
            while out_path.exists():
                out_path = out_dir / f"{Path(pl.name).stem}_v{n}.md"
                n += 1
            out_path.write_text(md, encoding="utf-8")
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Markdown export failed: {exc!r}", "final",
            )
            self._set_status("● idle")
            return
        try:
            rel = out_path.relative_to(VAULT_DIR)
        except Exception:
            rel = out_path
        self._append_transcript(
            "Writer", f"Exported {pl.name} -> {rel}", "final",
        )
        self._set_status("● idle")

    def _hf_download_response(self, repo: str, filename: str):
        """Download a GGUF model from Hugging Face in a worker thread."""
        try:
            import hf_download as _hf
        except Exception as exc:
            self._append_transcript(
                "Writer", f"hf_download module failed to import: {exc!r}", "final",
            )
            return
        if not _hf.hf_cli_available():
            self._append_transcript(
                "Writer",
                "huggingface_hub is not available. Install with:\n"
                "  pip install huggingface_hub\n"
                "(or `conda install -c conda-forge huggingface_hub`).",
                "final",
            )
            return

        dest = VAULT_DIR / "models"
        self._append_transcript(
            "Council",
            f"Downloading {filename} from {repo} into {dest} ...",
            "observation",
        )
        self._set_status("● downloading…", "#cba6f7")

        def _progress(line: str):
            # Filter noisy progress-bar updates — keep only meaningful lines
            if any(s in line.lower() for s in ("downloading", "to ", "error", "%")):
                self.after(0, lambda l=line: self._append_transcript(
                    "Workflow", "  " + l[:200], "observation",
                ))

        def _worker():
            try:
                ok, msg, path = _hf.download_gguf(
                    repo, filename, dest_dir=dest, on_progress=_progress,
                )
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"Download crashed: {exc!r}", "final"),
                    self._set_status("● idle"),
                ))
                return

            def _done():
                if ok and path:
                    self._append_transcript(
                        "Writer",
                        f"{msg}\n\nClick 'Browse...' at the top of the Council "
                        f"tab to point the model loader at this file, or run:\n"
                        f"  set COUNCIL_GGUF_PATH={path}",
                        "final",
                    )
                elif ok:
                    self._append_transcript("Writer", msg, "final")
                else:
                    self._append_transcript("Writer",
                                            f"Download failed: {msg}", "final")
                self._set_status("● idle")
            self.after(0, _done)

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _pipeline_explain_response(self, name: str):
        import pipeline_scanner as _ps
        import pipeline_editor as _pe
        pl = _ps.find_pipeline_by_name(VAULT_DIR, name)
        if not pl:
            self._append_transcript(
                "Writer",
                f"No pipeline matching '{name}'. Type 'list pipelines'.",
                "final",
            )
            self._set_status("● idle")
            return
        self._append_transcript(
            "Council", f"Asking model to describe {pl.name}…", "observation",
        )
        self._set_status("● describing…", "#fab387")

        def _worker():
            try:
                desc = _pe.pipeline_to_natural_language(pl)
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer", f"description failed: {exc!r}", "final"),
                    self._set_status("● idle"),
                ))
                return
            self.after(0, lambda: (
                self._append_transcript("Writer",
                                        f"{pl.name}:\n\n{desc}", "final"),
                self._set_status("● idle"),
            ))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _pipeline_validate_response(self, name: str):
        import pipeline_scanner as _ps
        pl = _ps.find_pipeline_by_name(VAULT_DIR, name)
        if not pl:
            self._append_transcript(
                "Writer", f"No pipeline matching '{name}'.", "final",
            )
            self._set_status("● idle")
            return
        issues = _ps.validate_pipeline_params(pl)
        if not issues:
            self._append_transcript(
                "Writer",
                f"{pl.name}: no issues found in {len(pl.steps)} step"
                f"{'s' if len(pl.steps)!=1 else ''} (against known simplnx schema).",
                "final",
            )
        else:
            lines = [f"{pl.name}: {len(issues)} issue(s):"]
            for i in issues:
                lines.append(f"  • {i}")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _pipeline_graph_response(self, name: str):
        import pipeline_scanner as _ps
        pl = _ps.find_pipeline_by_name(VAULT_DIR, name)
        if not pl:
            self._append_transcript(
                "Writer", f"No pipeline matching '{name}'.", "final",
            )
            self._set_status("● idle")
            return
        text = _ps.pipeline_dependency_graph(pl)
        self._append_transcript("Writer", text, "final")
        self._set_status("● idle")

    def _pipeline_create_response(self, description: str, full_request: str):
        import pipeline_editor as _pe

        self._append_transcript(
            "Council",
            f"Generating new pipeline from: {description}",
            "observation",
        )
        self._set_status("● generating…", "#fab387")

        # Suggest a filename — pull the first 3-4 content tokens
        tokens = [t for t in _re.split(r"\W+", description) if len(t) >= 3][:4]
        suggested = "_".join(t.lower() for t in tokens) or "generated_pipeline"

        def _worker():
            try:
                path, log = _pe.generate_pipeline_from_description(
                    description, VAULT_DIR, suggested_name=suggested,
                )
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"generation failed: {exc!r}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            if not path:
                self.after(0, lambda: (
                    self._append_transcript("Writer",
                                            f"couldn't generate pipeline: {log}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            try:
                rel = path.relative_to(VAULT_DIR)
            except Exception:
                rel = path

            def _done():
                self._append_transcript(
                    "Writer",
                    f"Saved new pipeline: {rel}\n\nReview with: "
                    f"show pipeline {path.name}",
                    "final",
                )
                if hasattr(self, "_dream3d_refresh_pipelines"):
                    try: self._dream3d_refresh_pipelines()
                    except Exception: pass
                self._set_status("● idle")
            self.after(0, _done)

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _pipeline_list_response(self):
        import pipeline_scanner as _ps
        in_dir = _ps.vault_pipelines_in_dir(VAULT_DIR)
        pipelines = _ps.scan_pipelines(in_dir)
        if not pipelines:
            self._append_transcript(
                "Writer",
                f"No pipelines found in {in_dir}. Drop .py simplnx scripts or "
                ".dream3d files in there and I'll see them.",
                "final",
            )
        else:
            lines = [f"Found {len(pipelines)} pipeline"
                     f"{'s' if len(pipelines) != 1 else ''} in vault/pipelines/in/:"]
            for pl in pipelines:
                note = f" ({pl.format}, {len(pl.steps)} step{'s' if len(pl.steps)!=1 else ''})"
                lines.append(f"  • {pl.name}{note}")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _pipeline_show_response(self, name: str):
        import pipeline_scanner as _ps
        pl = _ps.find_pipeline_by_name(VAULT_DIR, name)
        if not pl:
            self._append_transcript(
                "Writer",
                f"No pipeline matching '{name}' under vault/pipelines/in/. "
                f"Type 'list pipelines' to see what's available.",
                "final",
            )
        else:
            rendered = _ps.render_pipeline(pl)
            self._append_transcript("Writer", rendered, "final")
        self._set_status("● idle")

    def _pipeline_modify_response(self, name: str, change: str, full_request: str):
        import pipeline_scanner as _ps

        pl = _ps.find_pipeline_by_name(VAULT_DIR, name)
        if not pl:
            self._append_transcript(
                "Writer",
                f"No pipeline matching '{name}' to modify. Type 'list "
                f"pipelines' to see what's available.",
                "final",
            )
            return

        self._append_transcript(
            "Council", f"Modifying {pl.name}: {change}", "observation",
        )
        # Show the original first so the user sees what they're starting from
        self._append_transcript("Writer", _ps.render_pipeline(pl), "observation")
        self._set_status("● editing pipeline…", "#fab387")

        # Run the model call + edit application in a worker so the UI
        # stays responsive. The model can take 10+ seconds on slower GGUFs.
        def _worker(pl=pl, full_request=full_request):
            try:
                import pipeline_editor as _pe_w
                result = _pe_w.modify_pipeline_by_request(
                    pl.path, full_request, VAULT_DIR,
                )
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"Pipeline edit failed: {exc!r}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            self.after(0, lambda: self._pipeline_modify_finish(result))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _pipeline_modify_finish(self, result):
        """Main-thread continuation after a pipeline modification worker finishes."""
        if not result.success:
            self._append_transcript(
                "Writer",
                f"Couldn't apply that change:\n  {result.error}",
                "final",
            )
            if result.log:
                self._append_transcript(
                    "Council",
                    "Partial log:\n  " + "\n  ".join(result.log),
                    "observation",
                )
            self._set_status("● idle")
            return

        try:
            rel = result.new_path.relative_to(VAULT_DIR)
        except Exception:
            rel = result.new_path

        log_lines = [f"Saved new version: {rel}", "", "Edits applied:"]
        for line in result.log:
            log_lines.append(f"  • {line}")
        self._append_transcript("Writer", "\n".join(log_lines), "final")
        # Refresh the Dream3D pipeline list if the tab is built.
        if hasattr(self, "_dream3d_refresh_pipelines"):
            try:
                self._dream3d_refresh_pipelines()
            except Exception:
                pass
        self._set_status("● idle")

    # ---- Vault ergonomic intent handler ----

    _STATS_RE = _re.compile(
        r"^\s*(?:vault\s+stats|stats|show\s+(?:vault\s+)?stats|"
        r"what'?s\s+in\s+(?:my|the)\s+vault)\s*\??\s*$", _re.IGNORECASE,
    )
    _DUPES_RE = _re.compile(
        # Match the vault-duplicate-finder intent. End-anchored so that
        # "duplicate this" or "make a duplicate of X" do NOT trigger the
        # duplicates report — those are unrelated requests that happen
        # to start with the word. Accept "duplicate" + (s) optional,
        # optionally followed by "file" / "files".
        r"^\s*(?:find|show|list)?\s*duplicate(?:s|\s+files?)?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _HISTORY_SEARCH_RE = _re.compile(
        r"^\s*(?:search|find\s+in)\s+history\s+(?:for\s+)?['\"]?(.+?)['\"]?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _RECENT_QUERIES_RE = _re.compile(
        r"^\s*(?:recent\s+(?:queries|questions)|"
        r"what\s+have\s+i\s+asked|"
        r"my\s+recent\s+(?:queries|questions))\s*\??\s*$", _re.IGNORECASE,
    )
    _BUILD_SEMANTIC_RE = _re.compile(
        r"^\s*(?:build|generate|refresh)\s+(?:the\s+)?"
        r"(?:semantic\s+|llm\s+|description\s+|smart\s+)?(?:vault\s+)?"
        r"(?:index|descriptions?|summaries)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    # Topics-only build: skips the prose description for non-tabular
    # files, generating only the keyword topics. ~3x faster than the
    # full description path on CPU-only inference. Tabular files (CSV,
    # Excel, etc.) get schema-based descriptions either way — those
    # don't go through the model.
    _BUILD_TOPICS_ONLY_RE = _re.compile(
        r"^\s*(?:build|generate|refresh)\s+(?:the\s+)?"
        r"(?:vault\s+|file\s+)?"
        r"(?:topics?(?:\s+only)?|tags|keywords|categories?)"
        r"(?:\s+only)?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _BUILD_EMBEDDINGS_RE = _re.compile(
        r"^\s*(?:build|generate|refresh|rebuild)\s+(?:the\s+)?"
        r"(?:vector\s+|embedding|embeddings|semantic\s+vector)\s*"
        r"(?:index|cache|embeddings?)?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _TREE_RE = _re.compile(
        r"^\s*(?:tree|show\s+tree|folder\s+tree|directory\s+tree)"
        r"(?:\s+(?:of|in|for)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _GREP_RE = _re.compile(
        r"^\s*(?:grep|search\s+files|find\s+text|find\s+in\s+files)"
        r"\s+(?:for\s+)?['\"]?(.+?)['\"]?"
        r"(?:\s+(?:in|under|inside|within)\s+(.+?))?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _FIND_COLUMN_RE = _re.compile(
        r"^\s*(?:find|which|what)\s+(?:files\s+(?:have|contain|with)|"
        r"(?:csvs?|files?)\s+(?:have|with))\s+(?:a\s+|the\s+)?"
        r"['\"]?(.+?)['\"]?\s+column\s*\??\s*$",
        _re.IGNORECASE,
    )
    _RECENT_FILES_RE = _re.compile(
        r"^\s*(?:recent\s+files|what(?:'s|\s+is)\s+new|"
        r"files?\s+(?:changed|modified)\s+(?:recently|in\s+the\s+last))"
        r"(?:\s+(?:in|under)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _ROMAN_RE = _re.compile(
        r"^\s*(?:find|list|show|count)\s+(?:roman\s+numerals?|roman)"
        r"(?:\s+(?:in|under|inside|within)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _COMPARE_SCHEMAS_RE = _re.compile(
        r"^\s*(?:compare|diff)\s+schemas?\s+(?:of\s+)?(.+?)\s+(?:and|vs\.?|versus|to|with)\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _COLUMN_TYPES_RE = _re.compile(
        r"^\s*(?:infer|detect|show|what\s+are)\s+(?:the\s+)?(?:column\s+)?types?"
        r"(?:\s+(?:in|of|for)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    # ── Quick per-file analytics (model-free) ────────────────────────────
    # Requires the word "column(s)" so this never grabs the folder-level
    # "summary of stats for the files" route handled elsewhere.
    _COLUMN_STATS_RE = _re.compile(
        r"^\s*(?:column\s+(?:stats|statistics|summary)"
        r"|summari[sz]e\s+(?:the\s+)?(?:data\s+in\s+(?:the\s+)?)?columns?"
        r"|describe\s+(?:the\s+)?columns?"
        r"|stats?\s+(?:for|of|on)\s+(?:the\s+)?columns?)"
        r"(?:\s+(?:in|of|for|on)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _MISSING_DATA_RE = _re.compile(
        r"^\s*(?:missing\s+(?:data|values)|null\s+(?:report|counts?)|nulls?|"
        r"completeness)"
        r"(?:\s+(?:in|of|for|report\s+for)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _DUPLICATES_RE = _re.compile(
        r"^\s*(?:find\s+|count\s+|show\s+|check\s+(?:for\s+)?)?"
        r"(?:duplicates?|duplicate\s+rows?|dupes?)"
        r"(?:\s+(?:in|of|for)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _TOP_VALUES_RE = _re.compile(
        r"^\s*(?:top\s+values?|value\s+counts?|most\s+(?:common|frequent)\s+values?|"
        r"value\s+frequenc(?:y|ies)|frequenc(?:y|ies))"
        r"(?:\s+(?:in|of|for)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _CORRELATIONS_RE = _re.compile(
        r"^\s*(?:correlations?|correlate|corr|what\s+correlates)"
        r"(?:\s+(?:in|of|for|between\s+(?:the\s+)?columns\s+(?:in|of))\s+(.+?))?"
        r"\s*\??\s*$",
        _re.IGNORECASE,
    )
    # "image stats of <file>" / "pixel stats in <folder>" / "analyze images in X"
    _IMAGE_STATS_RE = _re.compile(
        r"^\s*(?:(?:image|picture|photo|pixel)\s+(?:stats|statistics|analysis|"
        r"summary)|analy[sz]e\s+(?:the\s+)?images?|summari[sz]e\s+(?:the\s+)?"
        r"images?)\s+(?:for|of|in|on|under)\s+(.+?)\s*\??\s*$",
        _re.IGNORECASE,
    )
    # "ocr <image>" / "read text in <image>" / "extract text from <image>"
    _OCR_RE = _re.compile(
        r"^\s*(?:ocr|read\s+(?:the\s+)?text\s+(?:in|from|on)|extract\s+text\s+"
        r"(?:from|in))\s+(.+?)\s*\??\s*$",
        _re.IGNORECASE,
    )
    # "count 12 features in <image>" / "detect objects in <image> expecting 12"
    # / "count dark pores in <image>". Groups: leadN, polarity, target, trailN.
    _DETECT_RE = _re.compile(
        r"^\s*(?:detect|count|find)\s+(?:and\s+count\s+)?"
        r"(?:(\d+)\s+)?"
        r"(?:(bright|light|dark)\s+)?"
        r"(?:objects?|features?|blobs?|particles?|pores?|spots?|holes?|"
        r"defects?|dots?|voids?)"
        r"\s+(?:in|on|of)\s+(.+?)"
        r"(?:\s+(?:expecting|expect|should\s+(?:have|be)|~|=)\s*(\d+))?"
        r"\s*\??\s*$",
        _re.IGNORECASE,
    )
    # "mean of <column> in <folder>" (+ optional "and save to <file>").
    # Groups: (1) aggregation word, (2) column, (3) the rest (folder phrase,
    # possibly with a trailing save clause parsed by _SAVE_CLAUSE_RE).
    _FOLDER_AGG_RE = _re.compile(
        r"^\s*(?:get|calculate|calc|compute|find|give\s+me|show(?:\s+me)?|"
        r"what(?:'s|\s+is|\s+are)|tell\s+me)?\s*(?:the\s+)?"
        r"(mean|average|avg|sum|total|min(?:imum)?|max(?:imum)?|median|"
        r"std(?:ev)?|standard\s+deviation|count)\b"
        r"\s+(?:value\s+)?(?:(?:of|for)\s+)?(?:the\s+)?(?:column\s+)?"
        r"['\"]?(.+?)['\"]?"
        r"\s+(?:column\s+)?(?:in|across|over|from|within)\s+(.+?)\s*$",
        _re.IGNORECASE,
    )
    # Layered vault content search: "search all files in the vault for X",
    # "which files mention X", "find references to X". Distinct from _GREP_RE
    # (a folder-scoped grep) — this requires the all-files / vault / which-files
    # framing so bare "search files for X in Y" still routes to grep.
    _VAULT_SEARCH_RE = _re.compile(
        r"^\s*(?:search\s+(?:through\s+)?(?:all\s+(?:the\s+)?|the\s+|my\s+)"
        r"(?:files|documents?|data)(?:\s+in\s+(?:the\s+)?vault)?\s+for"
        r"|search\s+(?:the\s+)?vault\s+for"
        r"|which\s+files\s+(?:mention|reference|contain|include|talk\s+about)"
        r"|what\s+files\s+(?:mention|reference|contain|include)"
        r"|find\s+(?:all\s+)?(?:references?\s+to|mentions?\s+of))"
        r"\s+(?:the\s+(?:term|word|phrase|value)\s+)?['\"]?(.+?)['\"]?\s*\??\s*$",
        _re.IGNORECASE,
    )
    # "create a tool that ..." / "make a tool to ..." — route to Tool Creation.
    _TOOL_CREATE_RE = _re.compile(
        r"^\s*(?:create|make|build|write|generate|forge)\s+(?:me\s+)?"
        r"(?:a\s+|an\s+)?(?:new\s+)?(?:tool|function|script)\s+"
        r"(?:that|to|which|for|called)\s+(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    # "list app tools" / "app tools" / "what tools have you built"
    _APP_TOOLS_RE = _re.compile(
        r"^\s*(?:list\s+|show\s+|what\s+(?:are\s+the\s+)?)?"
        r"(?:app[\s_-]?built\s+tools?|app\s+tools?|built\s+tools?|"
        r"self[\s-]?built\s+tools?|custom\s+tools?)"
        r"(?:\s+(?:have\s+you\s+built|available|do\s+i\s+have))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    # Trailing "…and save it to a csv file called foo.csv" clause.
    # Groups: (1) optional format (csv/tsv/text/txt), (2) optional filename.
    _SAVE_CLAUSE_RE = _re.compile(
        r"\s+(?:and\s+|then\s+)?(?:save|write|export|output|dump|put)\s+"
        r"(?:it|them|this|that|the\s+results?|the\s+findings?|the\s+output)?\s*"
        r"(?:out\s+)?(?:to|as|into|in)?\s*(?:a\s+|an\s+)?"
        r"(csv|tsv|text|txt|plain\s+text)?\s*(?:file)?\s*"
        r"(?:(?:named|called|titled)\s+)?"
        r"['\"]?([\w\-. ]+?)?['\"]?\s*[.!]?\s*$",
        _re.IGNORECASE,
    )
    _MONEY_RE_CHAT = _re.compile(
        r"^\s*(?:find|list|show)\s+(?:money|currency|prices?|amounts?|dollar\s+amounts?)"
        r"(?:\s+(?:in|under|inside|within)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _CONTEXT_INFO_RE = _re.compile(
        # Matches: "context info", "context window", "show context", "n_ctx",
        # "token budget", "how big is the context", "what is the context window"
        r"^\s*(?:"
        r"context\s+(?:info|window|size|budget|status)"
        r"|show\s+context(?:\s+window|\s+info)?"
        r"|n[_\s]?ctx"
        r"|token\s+budget"
        r"|how\s+(?:big|large)\s+is\s+(?:the\s+)?context(?:\s+window)?"
        r"|what(?:'s|\s+is)\s+(?:the\s+)?context(?:\s+window|\s+size)?"
        r")\s*\??\s*$",
        _re.IGNORECASE,
    )

    # ── Task-memo meta-commands ───────────────────────────────────────
    # "show memo" / "what's the task memo" — render the current memo
    # without re-condensing (useful for debugging a misread intent).
    # "reset memo" / "forget memo" / "new task" — drop the memo entirely
    # so the NEXT user query starts fresh instead of inheriting the
    # previous goal/constraints.
    _SHOW_MEMO_RE = _re.compile(
        r"^\s*(?:"
        r"show\s+(?:task\s+)?memo"
        r"|(?:task\s+)?memo\s+status"
        r"|what(?:'s|\s+is)\s+(?:the\s+)?(?:task\s+)?memo"
        r"|display\s+(?:task\s+)?memo"
        r")\s*\??\s*$",
        _re.IGNORECASE,
    )
    _RESET_MEMO_RE = _re.compile(
        r"^\s*(?:"
        r"reset\s+(?:task\s+)?memo"
        r"|forget\s+(?:task\s+)?memo"
        r"|clear\s+(?:task\s+)?memo"
        r"|new\s+task"
        r"|new\s+question"
        r"|start\s+over"
        r")\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )

    # ── Learned-synonym cache inspection / reset ─────────────────────
    # The vault index caches every semantic expansion the model
    # produced ("metals" -> ["promethium", "adamantium"]) in
    # vault/semantic_cache.json. These chat intents let the user see
    # what the model learned about their vault, and reset the cache
    # if they want to start over (e.g. they corrected the model's
    # interpretation and want it re-run with different vocab).
    _SHOW_LEARNED_RE = _re.compile(
        r"^\s*(?:"
        r"show\s+learned\s+(?:synonyms?|categories?)"
        r"|(?:learned\s+)?synonyms?(?:\s+status)?"
        r"|show\s+(?:vault\s+)?categories?"
        r"|what\s+has\s+the\s+model\s+learned"
        r")\s*\??\s*$",
        _re.IGNORECASE,
    )
    _CLEAR_LEARNED_RE = _re.compile(
        r"^\s*(?:"
        r"clear\s+learned\s+(?:synonyms?|categories?|cache)"
        r"|reset\s+learned\s+(?:synonyms?|categories?|cache)"
        r"|forget\s+learned\s+(?:synonyms?|categories?)"
        r"|clear\s+semantic\s+cache"
        r"|reset\s+semantic\s+cache"
        r")\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )

    # Dependency / diagnostics chat intent — same info the Diagnostics
    # tab shows, available via chat for users who prefer to type. Useful
    # for sharing the full system snapshot when reporting a bug
    # (the chat output goes into the transcript export).
    _DIAGNOSTICS_RE = _re.compile(
        r"^\s*(?:"
        r"check\s+dependencies"
        r"|check\s+deps"
        r"|what(?:'s|\s+is)\s+missing"
        r"|missing\s+dependencies"
        r"|missing\s+features"
        r"|system\s+status"
        r"|system\s+diagnostics"
        r"|diagnose"
        r"|run\s+diagnostics"
        r"|show\s+(?:dependencies|deps|diagnostics)"
        r"|what\s+optional\s+(?:packages|features|dependencies)"
        r")\s*\??\s*$",
        _re.IGNORECASE,
    )

    _ELEMENT_RANKING_RE = _re.compile(
        # Matches phrasings like:
        #   what is the most common (atomic) element [in <where>]
        #   most common element(s) in data_in
        #   top atomic elements in vault
        #   rank atomic elements
        #   list top 5 elements
        # Trailing \s+ is inside each verb branch (rather than after the
        # alternation) so "list top 5 elements" doesn't need a double
        # space between "top" and "5".
        r"^\s*(?:what\s+(?:is|are)\s+(?:the\s+)?)?"
        r"(?:most\s+common\s+|top\s+|rank\s+|list\s+(?:top\s+)?)"
        r"(?:(\d+)\s+)?"                              # optional count
        r"(?:atomic\s+|chemical\s+)?elements?"
        r"(?:\s+(?:in|of|under|inside|within)\s+(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _LIST_FOLDERS_RE = _re.compile(
        # Matches phrasings like:
        #   list (the) subfolders [in <where>]
        #   list the subfolders in the data_in folder within the vault
        #   show (me) folders (in|of|under|inside|within) <where>
        #   what folders are in <where>?
        #   what (are) (the) subfolders
        # Captures <where> as group 1 (may be None when no location given).
        r"^\s*(?:list|show|what)\s+"
        r"(?:me\s+|are\s+)?(?:the\s+)?"
        r"(?:sub)?(?:folders?|directories|dirs)"
        r"(?:\s+(?:are\s+)?(?:in|of|under|inside|within)\s+"
        r"(?:the\s+)?(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _LIST_FILES_RE = _re.compile(
        # "list files in X" / "show files in X" / "what files are in X"
        # Captures <where> as group 1.
        r"^\s*(?:list|show|what)\s+"
        r"(?:me\s+|are\s+)?(?:the\s+|all\s+(?:of\s+)?(?:the\s+)?)?"
        r"files?"
        r"(?:\s+(?:are\s+)?(?:in|of|under|inside|within)\s+"
        r"(?:the\s+)?(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _LOOK_AT_FILES_RE = _re.compile(
        # "look at (all) (the) files in X" — triggers the full folder
        # injection so the model sees a per-file summary.
        r"^\s*(?:look\s+at|inspect|scan|read)\s+"
        r"(?:all\s+(?:of\s+)?(?:the\s+)?|the\s+)?files?"
        r"(?:\s+(?:in|under|inside|within)\s+(?:the\s+)?(.+?))?\s*\??\s*$",
        _re.IGNORECASE,
    )
    # Permissive "listing intent" detector — catches any phrasing that
    # asks for a list of files/contents/items in a folder, even if it
    # doesn't strictly match _LIST_FILES_RE. Used as a SAFETY NET so the
    # model never gets a chance to paraphrase a folder listing — the
    # deterministic `_list_files_response` always wins when this fires.
    # The pattern is intentionally loose; we cross-check with an actual
    # folder path before bypassing the model, so false-positives are
    # harmless.
    _LISTING_INTENT_RE = _re.compile(
        r"(?:"
            # Strict listing verbs only — these unambiguously ask for a
            # bulleted file list, not a summary or analysis.
            r"\b(?:list|enumerate|dump)\b"
            # "show / display / print (files|everything|contents)" —
            # only listing when the object is itself listing-shaped.
            r"|\b(?:show|display|print)\s+"
            r"(?:me\s+|all\s+(?:of\s+)?(?:the\s+)?|the\s+|every\s+|each\s+)?"
            r"(?:files?|folders?|directories|subfolders?|"
            r"contents|everything|every\s+file)\b"
            # "give me a (list|listing|rundown|inventory) of..."
            r"|\bgive\s+me\s+(?:a|the)\s+"
            r"(?:list|listing|rundown|inventory)\b"
            # "what's in / what is in / whats in" (anchored)
            r"|\b(?:what(?:'s|\s+is)|whats)\s+(?:in|inside|within|under)\b"
            # "what files / what folders / what subfolders"
            r"|\bwhat\s+(?:files?|folders?|directories|subfolders?)\b"
            # polite wrappers — only for strict listing verbs
            r"|\b(?:can|could|would|will)\s+you\s+"
            r"(?:please\s+)?(?:list|enumerate|dump)\b"
            r"|\bplease\s+(?:list|enumerate|dump)\b"
        r")",
        _re.IGNORECASE,
    )

    # Trailing-noise words to strip from the captured target — users say
    # things like "data_in folder within the vault" and we only want
    # "data_in".
    _FOLDER_NOISE_RE = _re.compile(
        r"\s+(?:folder|directory|subfolder|subdirectory|"
        r"within\s+(?:the\s+)?vault|inside\s+(?:the\s+)?vault|"
        r"of\s+(?:the\s+)?vault|in\s+(?:the\s+)?vault|"
        r"the\s+vault|vault)\b.*$",
        _re.IGNORECASE,
    )
    _EXPORT_TRANSCRIPT_RE = _re.compile(
        r"^\s*(?:export|save)\s+(?:the\s+)?(?:current\s+)?transcript"
        r"(?:\s+as\s+(?:markdown|md))?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _SCHEMA_DOC_RE = _re.compile(
        r"^\s*(?:generate|build|make|export)\s+(?:a\s+)?schema\s+(?:doc(?:ument)?|"
        r"documentation)\s+(?:for\s+|of\s+)?(.+?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _QUALITY_RE = _re.compile(
        r"^\s*(?:quality\s+check|check\s+(?:data\s+)?quality|"
        r"data\s+quality(?:\s+check)?|"
        r"audit\s+(?:data\s+)?(?:quality\s+of\s+)?)\s*(.*?)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _WHERE_VALUE_RE = _re.compile(
        r"^\s*(?:where\s+(?:did|does|is)|trace|verify|cite|find\s+the\s+source\s+of)"
        r"\s+(?:that\s+|the\s+)?(?:value\s+)?['\"]?(.+?)['\"]?"
        r"(?:\s+(?:come\s+)?from|\s+in\s+(?:the\s+)?data)?\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _SHOW_CONTEXT_RE = _re.compile(
        r"^\s*(?:show|what(?:'s| is)?)\s+(?:last\s+|the\s+last\s+|recent\s+)?"
        r"(?:context|injection|injected\s+(?:data|content)|what\s+(?:you|the\s+model)\s+saw)"
        r"\s*\??\s*$",
        _re.IGNORECASE,
    )
    _VERIFY_LAST_RE = _re.compile(
        r"^\s*(?:verify|check|audit)\s+(?:the\s+)?(?:last|previous|that)\s+"
        r"(?:answer|response|reply)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )
    _LIST_SQL_CONNS_RE = _re.compile(
        r"^\s*(?:list|show)\s+(?:my\s+|the\s+)?sql\s+connections?\s*\??\s*$",
        _re.IGNORECASE,
    )
    _ADD_SQL_CONN_RE = _re.compile(
        r"^\s*(?:add|save|register)\s+sql\s+connection\s+(\S+)\s+(\S+)\s*[.!?]?\s*$",
        _re.IGNORECASE,
    )

    def _handle_vault_tools_intent(self, user_text: str) -> bool:
        if not user_text:
            return False
        single_line = user_text.split("\n", 1)[0]
        import vault_tools as _vt

        if self._STATS_RE.match(single_line):
            stats = _vt.vault_stats(VAULT_DIR)
            self._append_transcript("Writer", _vt.format_vault_stats(stats), "final")
            return True

        if self._CONTEXT_INFO_RE.match(single_line):
            self._context_info_response()
            return True

        if self._SHOW_MEMO_RE.match(single_line):
            self._show_task_memo_response()
            return True

        if self._RESET_MEMO_RE.match(single_line):
            self._reset_task_memo_response()
            return True

        if self._SHOW_LEARNED_RE.match(single_line):
            self._show_learned_synonyms_response()
            return True

        if self._CLEAR_LEARNED_RE.match(single_line):
            self._clear_learned_synonyms_response()
            return True

        if self._DIAGNOSTICS_RE.match(single_line):
            self._diagnostics_chat_response()
            return True

        if self._DUPES_RE.match(single_line):
            groups = _vt.find_duplicate_files(VAULT_DIR)
            self._append_transcript("Writer", _vt.format_duplicates(groups), "final")
            return True

        m = self._HISTORY_SEARCH_RE.match(single_line)
        if m:
            query = m.group(1).strip()
            hits = _vt.query_history_search(VAULT_DIR, query, limit=10)
            self._append_transcript("Writer", _vt.format_history_hits(hits), "final")
            return True

        if self._RECENT_QUERIES_RE.match(single_line):
            hits = _vt.recent_queries(VAULT_DIR, n=10)
            if not hits:
                self._append_transcript("Writer", "No past queries found.", "final")
            else:
                lines = [f"Last {len(hits)} user queries:"]
                for h in hits:
                    snippet = h["text"][:140].replace("\n", " ")
                    lines.append(f"  [{h['ts']}]  {snippet}")
                self._append_transcript("Writer", "\n".join(lines), "final")
            return True

        if self._BUILD_SEMANTIC_RE.match(single_line):
            self._build_semantic_index_response()
            return True

        if self._BUILD_TOPICS_ONLY_RE.match(single_line):
            self._build_topics_only_response()
            return True

        if self._BUILD_EMBEDDINGS_RE.match(single_line):
            self._build_embeddings_response()
            return True

        m = self._LIST_FOLDERS_RE.match(single_line)
        if m:
            target = (m.group(1) or "").strip().strip("'\"`")
            # Strip trailing filler ("X folder within the vault" -> "X")
            target = self._FOLDER_NOISE_RE.sub("", target).strip()
            self._list_folders_response(target)
            return True

        m = self._LIST_FILES_RE.match(single_line)
        if m:
            target = self._FOLDER_NOISE_RE.sub("", (m.group(1) or "").strip().strip("'\"`")).strip()
            self._list_files_response(target)
            return True

        m = self._LOOK_AT_FILES_RE.match(single_line)
        if m:
            target = self._FOLDER_NOISE_RE.sub("", (m.group(1) or "").strip().strip("'\"`")).strip()
            self._look_at_files_response(target)
            return True

        m = self._ELEMENT_RANKING_RE.match(single_line)
        if m:
            n = int(m.group(1)) if m.group(1) else 10
            target = (m.group(2) or "").strip().strip("'\"`")
            target = self._FOLDER_NOISE_RE.sub("", target).strip()
            self._element_ranking_response(target, top_n=n)
            return True

        # Filesystem helpers — tree, grep, column search, recent
        m = self._TREE_RE.match(single_line)
        if m:
            target = self._FOLDER_NOISE_RE.sub("", (m.group(1) or "").strip().strip("'\"`")).strip()
            self._tree_response(target)
            return True
        m = self._VAULT_SEARCH_RE.match(single_line)
        if m:
            self._vault_term_search_response(m.group(1).strip().strip("'\"`"))
            return True
        m = self._GREP_RE.match(single_line)
        if m:
            self._grep_response(m.group(1).strip().strip("'\"`"),
                                 (m.group(2) or "").strip().strip("'\"`"))
            return True
        m = self._FIND_COLUMN_RE.match(single_line)
        if m:
            self._find_column_response(m.group(1).strip().strip("'\"`"))
            return True
        m = self._RECENT_FILES_RE.match(single_line)
        if m:
            target = self._FOLDER_NOISE_RE.sub("", (m.group(1) or "").strip().strip("'\"`")).strip()
            self._recent_files_response(target)
            return True
        # Pattern searches — roman, money
        m = self._ROMAN_RE.match(single_line)
        if m:
            target = self._FOLDER_NOISE_RE.sub("", (m.group(1) or "").strip().strip("'\"`")).strip()
            self._roman_response(target)
            return True
        m = self._MONEY_RE_CHAT.match(single_line)
        if m:
            target = self._FOLDER_NOISE_RE.sub("", (m.group(1) or "").strip().strip("'\"`")).strip()
            self._money_response(target)
            return True
        m = self._COMPARE_SCHEMAS_RE.match(single_line)
        if m:
            self._compare_schemas_response(
                m.group(1).strip().strip("'\"`"),
                m.group(2).strip().strip("'\"`"),
            )
            return True
        m = self._COLUMN_TYPES_RE.match(single_line)
        if m:
            target = (m.group(1) or "").strip().strip("'\"`")
            self._column_types_response(target)
            return True
        m = self._COLUMN_STATS_RE.match(single_line)
        if m:
            self._column_stats_response((m.group(1) or "").strip().strip("'\"`"))
            return True
        m = self._MISSING_DATA_RE.match(single_line)
        if m:
            self._missing_data_response((m.group(1) or "").strip().strip("'\"`"))
            return True
        m = self._DUPLICATES_RE.match(single_line)
        if m:
            self._duplicates_response((m.group(1) or "").strip().strip("'\"`"))
            return True
        m = self._TOP_VALUES_RE.match(single_line)
        if m:
            self._top_values_response((m.group(1) or "").strip().strip("'\"`"))
            return True
        m = self._CORRELATIONS_RE.match(single_line)
        if m:
            self._correlations_response((m.group(1) or "").strip().strip("'\"`"))
            return True
        m = self._IMAGE_STATS_RE.match(single_line)
        if m:
            self._image_stats_response(m.group(1).strip().strip("'\"`"))
            return True
        m = self._OCR_RE.match(single_line)
        if m:
            self._ocr_response(m.group(1).strip().strip("'\"`"))
            return True
        m = self._DETECT_RE.match(single_line)
        if m:
            self._detect_features_response(
                (m.group(2) or "").strip(),
                m.group(3).strip().strip("'\"`"),
                (m.group(4) or m.group(1)))
            return True
        m = self._FOLDER_AGG_RE.match(single_line)
        if m:
            self._folder_agg_response(m.group(1), m.group(2), m.group(3))
            return True
        if self._APP_TOOLS_RE.match(single_line):
            self._app_tools_response()
            return True
        m = self._TOOL_CREATE_RE.match(single_line)
        if m:
            _task = m.group(1).strip().strip("'\"`")
            self._append_transcript(
                "Council",
                f"Opening the Tool Creation tab to build a tool: {_task}",
                "observation")
            self._forge_route(_task)
            return True

        if self._EXPORT_TRANSCRIPT_RE.match(single_line):
            self._export_transcript_response()
            return True

        m = self._SCHEMA_DOC_RE.match(single_line)
        if m:
            self._schema_doc_response(m.group(1).strip().strip("'\"`"))
            return True

        m = self._QUALITY_RE.match(single_line)
        if m:
            target = (m.group(1) or "").strip().strip("'\"`")
            self._quality_check_response(target)
            return True

        # Provenance / verification intents — these consult the in-session
        # ProvenanceTracker, not the model. Fast and authoritative.
        if self._SHOW_CONTEXT_RE.match(single_line):
            self._show_last_context_response()
            return True

        if self._VERIFY_LAST_RE.match(single_line):
            self._verify_last_answer_response()
            return True

        m = self._WHERE_VALUE_RE.match(single_line)
        if m:
            value = m.group(1).strip().strip("'\"`")
            self._where_value_response(value)
            return True

        if self._LIST_SQL_CONNS_RE.match(single_line):
            self._list_sql_connections_response()
            return True

        m = self._ADD_SQL_CONN_RE.match(single_line)
        if m:
            self._add_sql_connection_response(m.group(1), m.group(2))
            return True

        return False

    def _list_sql_connections_response(self):
        import vault_analyst as _va
        conns = _va.list_sql_connections(VAULT_DIR)
        if not conns:
            self._append_transcript(
                "Writer",
                "No saved SQL connections.\nAdd one with:\n"
                "  add sql connection myname postgresql://user:${PGPASS}@host/db",
                "final",
            )
        else:
            lines = [f"{len(conns)} saved SQL connection(s):"]
            for name, url in sorted(conns.items()):
                # Mask anything that looks like a password in the URL preview
                masked = _re.sub(r"://[^/@:]+:([^@]+)@", "://USER:***@", url)
                lines.append(f"  • {name}  ->  {masked}")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _add_sql_connection_response(self, name: str, url: str):
        import vault_analyst as _va
        try:
            _va.save_sql_connection(VAULT_DIR, name, url)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Couldn't save connection: {exc!r}", "final",
            )
            self._set_status("● idle")
            return
        masked = _re.sub(r"://[^/@:]+:([^@]+)@", "://USER:***@", url)
        self._append_transcript(
            "Writer",
            f"Saved SQL connection '{name}' = {masked}\n"
            f"Use it from the analyst with read_sql_table(vault, "
            f"'{name}', 'tablename') or sql_query(vault, '{name}', 'SELECT ...').",
            "final",
        )
        self._set_status("● idle")

    def _show_last_context_response(self):
        """Print the [FILE: ...] blocks the model saw in the last turn."""
        prov = getattr(self, "provenance", None)
        if not prov:
            self._append_transcript("Writer", "Provenance is not initialized.", "final")
            self._set_status("● idle")
            return
        turn = prov.last_turn()
        if not turn:
            self._append_transcript(
                "Writer", "No previous turn recorded yet.", "final",
            )
            self._set_status("● idle")
            return
        if not turn.injected_files:
            self._append_transcript(
                "Writer",
                f"Turn #{turn.turn_id} had NO injected files — the model was "
                f"answering from its own training, not from your data.",
                "final",
            )
            self._set_status("● idle")
            return
        lines = [f"Context from turn #{turn.turn_id} ({turn.timestamp}):"]
        for ib in turn.injected_files:
            lines.append("")
            lines.append(f"--- {ib.file_name} ({ib.file_path}) ---")
            # Cap rendering so the transcript doesn't explode
            block = ib.block
            if len(block) > 3000:
                block = block[:3000] + "\n... (truncated)"
            lines.append(block)
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _verify_last_answer_response(self):
        """Extract numeric values from the most recent model reply and
        check each against the injection. Flags any that aren't found —
        those are the most likely hallucinations."""
        prov = getattr(self, "provenance", None)
        if not prov or not prov.last_turn():
            self._append_transcript("Writer", "No previous answer to verify.", "final")
            self._set_status("● idle")
            return
        report = prov.verify_response("")
        if not report["checked"]:
            self._append_transcript(
                "Writer",
                "No numeric values found in the last answer to verify.",
                "final",
            )
            self._set_status("● idle")
            return
        lines = [f"Verifying {len(report['checked'])} numeric value(s) "
                 f"from the last answer:"]
        for hit in report["found"]:
            w = hit["where"]
            lines.append(
                f"  ✓ {hit['value']} — found in {w['file_name']}, "
                f"line {w['line_index']+1}"
            )
        for v in report["missing"]:
            lines.append(
                f"  ✗ {v} — NOT FOUND in the injected files. "
                f"Likely hallucinated or computed indirectly."
            )
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _where_value_response(self, value: str):
        prov = getattr(self, "provenance", None)
        if not prov:
            self._append_transcript("Writer", "Provenance is not initialized.", "final")
            self._set_status("● idle")
            return
        hits = prov.search_value(value, max_turns_back=5, max_hits=8)
        if not hits:
            self._append_transcript(
                "Writer",
                f"No match for {value!r} in the last 5 turns of injected "
                f"context.\n\nThat usually means one of three things:\n"
                f"  1. The value was computed by the model from numbers in "
                f"the file (so the literal string isn't there).\n"
                f"  2. The value was hallucinated — the model produced it "
                f"without any source.\n"
                f"  3. The file is in the vault but wasn't injected this "
                f"session (try 'show pipeline <path>' or pass the path "
                f"explicitly to bring it into context).",
                "final",
            )
            self._set_status("● idle")
            return
        lines = [f"Found {len(hits)} occurrence(s) of {value!r}:"]
        for h in hits:
            lines.append("")
            lines.append(f"  Turn #{h['turn_id']} — {h['file_name']} "
                         f"(line {h['line_index']+1}, {h['match_kind']} match)")
            lines.append("  Context:")
            for snip_line in h["context_snippet"].split("\n"):
                lines.append("    " + snip_line)
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _quality_check_response(self, target: str):
        """Run detect_data_quality_issues against a CSV (specific file) or
        every CSV in the vault (when target is empty)."""
        import vault_analyst as _va
        if not target:
            # Whole-vault sweep
            issues = _va.detect_data_quality_issues_per_csv(VAULT_DIR)
            if issues.empty:
                self._append_transcript(
                    "Writer",
                    "No quality issues detected across the vault's CSV files.",
                    "final",
                )
            else:
                lines = [f"Found {len(issues)} issue(s) across vault CSVs:"]
                for _, row in issues.head(40).iterrows():
                    lines.append(
                        f"  [{row.get('severity','?')}] "
                        f"{row.get('csv','?')} :: "
                        f"{row.get('column','')} :: "
                        f"{row.get('kind','?')} — {row.get('message','')}"
                    )
                if len(issues) > 40:
                    lines.append(f"  ... ({len(issues) - 40} more)")
                self._append_transcript("Writer", "\n".join(lines), "final")
            self._set_status("● idle")
            return

        # Single file
        p = Path(target).expanduser()
        if not p.is_file():
            from vault_analyst import list_csv_files
            try:
                matches = [c for c in list_csv_files(VAULT_DIR)
                           if target.lower() in c.name.lower()]
            except Exception:
                matches = []
            if matches:
                p = matches[0]
        if not p.is_file():
            self._append_transcript(
                "Writer", f"No file found for '{target}'.", "final",
            )
            self._set_status("● idle")
            return
        try:
            import pandas as _pd
            df = _va.read_table(p)
            issues = _va.detect_data_quality_issues(df)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Quality check failed: {exc!r}", "final",
            )
            self._set_status("● idle")
            return
        if issues.empty:
            self._append_transcript(
                "Writer", f"{p.name}: no quality issues detected.", "final",
            )
        else:
            lines = [f"{p.name}: {len(issues)} issue(s) found:"]
            for _, row in issues.head(30).iterrows():
                lines.append(
                    f"  [{row.get('severity','?')}] "
                    f"{row.get('column','')} :: "
                    f"{row.get('kind','?')} — {row.get('message','')}"
                )
            if len(issues) > 30:
                lines.append(f"  ... ({len(issues) - 30} more)")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _export_transcript_response(self):
        import vault_tools as _vt
        try:
            out = _vt.export_transcript_as_markdown(VAULT_DIR, self.session_id)
        except Exception as exc:
            self._append_transcript("Writer", f"Export failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        if not out:
            self._append_transcript(
                "Writer",
                "Nothing to export yet — this session has no recorded turns.",
                "final",
            )
        else:
            try:
                rel = out.relative_to(VAULT_DIR)
            except Exception:
                rel = out
            self._append_transcript(
                "Writer", f"Transcript exported to {rel}", "final",
            )
        self._set_status("● idle")

    def _schema_doc_response(self, target: str):
        """Generate a schema doc for a CSV (path or name). Saves to data_out."""
        p = Path(target).expanduser()
        if not p.is_file():
            from vault_analyst import list_csv_files
            try:
                matches = [c for c in list_csv_files(VAULT_DIR)
                           if target.lower() in c.name.lower()]
            except Exception:
                matches = []
            if matches:
                p = matches[0]
        if not p.is_file() or p.suffix.lower() != ".csv":
            self._append_transcript(
                "Writer",
                f"Could not find a CSV matching '{target}'.",
                "final",
            )
            self._set_status("● idle")
            return
        try:
            import vault_analyst as _va
            md = _va.schema_doc_from_csv(p)
            out_dir = data_index.output_dir(VAULT_DIR)
            out_path = out_dir / f"schema_{p.stem}.md"
            n = 2
            while out_path.exists():
                out_path = out_dir / f"schema_{p.stem}_v{n}.md"
                n += 1
            out_path.write_text(md, encoding="utf-8")
        except Exception as exc:
            self._append_transcript("Writer", f"Schema doc failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        try:
            rel = out_path.relative_to(VAULT_DIR)
        except Exception:
            rel = out_path
        self._append_transcript(
            "Writer", f"Schema doc for {p.name} -> {rel}", "final",
        )
        self._set_status("● idle")

    def _resolve_folder_target(self, target: str) -> Path:
        """Shared resolver — explicit absolute path, vault-relative path,
        or empty -> data_in/. Used by tree / grep / recent / roman / money
        / element-ranking / list-subfolders handlers."""
        if not target:
            try:
                import data_index
                return data_index.input_dir(VAULT_DIR)
            except Exception:
                return VAULT_DIR / "data_in"
        p = Path(target).expanduser()
        if p.is_absolute() and p.exists():
            return p
        candidate = VAULT_DIR / target
        return candidate if candidate.exists() else p

    def _tree_response(self, target: str):
        import vault_tools as _vt
        root = self._resolve_folder_target(target)
        if not root.exists() or not root.is_dir():
            self._append_transcript("Writer", f"Folder not found: {root}", "final")
        else:
            text = _vt.tree(root, max_depth=3, show_files=False)
            self._append_transcript("Writer", text, "final")
        self._set_status("● idle")

    def _vault_term_search_response(self, term: str):
        """Layered vault search for a term: file SUMMARIES / keywords first,
        then a DEEPER scan of file contents — over the user's data area only.
        App-state files (question_history.json, agent_jobs.json, indices) and
        conversation logs are excluded via conversation_logger.is_protected_path.
        """
        term = (term or "").strip().strip("'\"`")
        if not term:
            self._append_transcript(
                "Writer", "What term should I search the files for?", "final")
            self._set_status("● idle")
            return
        import data_index as _di
        import vault_tools as _vt
        try:
            import conversation_logger as _cl
        except Exception:
            _cl = None

        def _protected(pth) -> bool:
            if _cl is None or not pth:
                return False
            try:
                return _cl.is_protected_path(pth, VAULT_DIR)
            except Exception:
                return False

        self._set_status("● searching…", "#cba6f7")
        data_root = _di.input_dir(VAULT_DIR)

        # 1) Summaries / descriptions / keywords via the vault index (fast, no
        #    file reads). This is the "check the summaries first" layer.
        summary: list = []          # (name, type)
        summary_names: set = set()
        idx = _get_vault_index()
        if idx is not None:
            try:
                idx.rebuild()
            except Exception:
                pass
            try:
                all_hits, _fz = idx.search(term, k=25)
            except Exception:
                all_hits = []
            for _score, rec in all_hits:
                pth = rec.get("path")
                if not pth or _protected(pth):
                    continue
                nm = rec.get("name") or Path(str(pth)).name
                if nm in summary_names:
                    continue
                summary_names.add(nm)
                summary.append((nm, rec.get("type", "?")))

        # 2) Deeper scan of the actual file text (the "then a deeper look"
        #    layer). find_files_containing_text already skips protected paths.
        content: list = []          # (name, snippet)
        content_names: set = set()
        try:
            hits = _vt.find_files_containing_text(data_root, term, max_hits=300)
        except Exception:
            hits = []
        for h in hits:
            nm = Path(str(h.get("path", ""))).name
            if not nm or nm in summary_names or nm in content_names:
                continue
            if _protected(h.get("path")):
                continue
            content_names.add(nm)
            content.append((nm, (h.get("context") or "").strip()[:100]))

        if not summary and not content:
            self._append_transcript(
                "Writer",
                f"No files reference {term!r} — checked the file summaries and "
                f"then the full text of files under {data_root.name}/.", "final")
            self._set_status("● idle")
            return
        lines = [f"Files referencing {term!r}:"]
        if summary:
            lines.append("")
            lines.append(f"Matched in file summary / keywords ({len(summary)}):")
            for nm, ty in summary[:40]:
                lines.append(f"  • {nm}  [{ty}]")
        if content:
            lines.append("")
            lines.append(f"Found deeper in file contents ({len(content)}):")
            for nm, snip in content[:40]:
                lines.append(f"  • {nm}" + (f"  →  {snip}" if snip else ""))
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _grep_response(self, query: str, target: str):
        import vault_tools as _vt
        # Default to the user's DATA area (data_in), NOT the raw vault root —
        # the vault root holds app-state files (question_history.json,
        # agent_jobs.json, indices) and the conversation_logs/ the model must
        # never read; rooting a whole-vault grep there returned those instead
        # of real files. _resolve_folder_target('') already resolves to data_in.
        root = self._resolve_folder_target(target)
        if not root.exists():
            self._append_transcript("Writer", f"Folder not found: {root}", "final")
            self._set_status("● idle")
            return
        hits = _vt.find_files_containing_text(root, query, max_hits=50)
        if not hits:
            self._append_transcript(
                "Writer",
                f"No matches for {query!r} in any text file under {root.name}/.",
                "final",
            )
        else:
            lines = [f"Found {len(hits)} match(es) for {query!r}:"]
            for h in hits:
                lines.append(f"  {h['path']}:{h['line']}  →  {h['context'][:120]}")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _find_column_response(self, column: str):
        import vault_tools as _vt
        hits = _vt.find_files_with_column(VAULT_DIR, column)
        if not hits:
            self._append_transcript(
                "Writer", f"No CSV/Excel file under the vault has a column matching {column!r}.",
                "final",
            )
        else:
            lines = [f"{len(hits)} file(s)/sheet(s) have a {column!r} column:"]
            for h in hits:
                where = h['path']
                if h.get('sheet'):
                    where += f"  (sheet: {h['sheet']})"
                cols = ", ".join(h['matched_columns'])
                lines.append(f"  {where}  →  {cols}")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _recent_files_response(self, target: str):
        import vault_tools as _vt
        root = self._resolve_folder_target(target) if target else VAULT_DIR
        files = _vt.recent_files(root, since_days=7, limit=20)
        if not files:
            self._append_transcript(
                "Writer", f"No files modified in the last 7 days under {root.name}/.",
                "final",
            )
        else:
            lines = [f"Files modified in the last 7 days ({len(files)} shown):"]
            for f in files:
                size_kb = f['size'] / 1024
                lines.append(f"  {f['iso']}  {size_kb:>8.1f} KB  {f['path']}")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _roman_response(self, target: str):
        import vault_tools as _vt
        root = self._resolve_folder_target(target)
        df = _vt.find_roman_numerals(root, top_n=20)
        if df is None or len(df) == 0:
            self._append_transcript(
                "Writer", f"No Roman numerals (length ≥ 2) found under {root.name}/.", "final",
            )
            self._set_status("● idle")
            return
        # Split into confident vs ambiguous-English-abbreviation
        confident = df[df['count'] > 0]
        ambiguous = df[df['ambiguous_count'] > 0]
        lines = [f"Roman numerals under {root.name}/ (single-letter matches skipped):"]
        if len(confident):
            lines.append("")
            lines.append("Confident matches:")
            lines.append(f"  {'roman':<8}{'integer':>8}{'count':>8}{'files':>8}")
            for _, r in confident.iterrows():
                lines.append(f"  {r['roman']:<8}{int(r['integer']):>8}"
                             f"{int(r['count']):>8}{int(r['files']):>8}")
        else:
            lines.append("  (no non-ambiguous Roman numerals found)")
        if len(ambiguous):
            lines.append("")
            lines.append("Likely false positives (common English abbreviations):")
            lines.append(f"  {'token':<8}{'as roman':>10}{'count':>8}{'note':<25}")
            notes = {
                "ML":"machine learning", "MD":"M.D./Maryland", "MC":"Master of Ceremonies",
                "MV":"music video", "MI":"Michigan", "DC":"Washington DC",
                "DI":"digital input", "DL":"download/driver licence",
                "LI":"Long Island", "LV":"Las Vegas", "CL":"chlorine",
                "CM":"centimeter", "CD":"compact disc", "XL":"extra large",
            }
            for _, r in ambiguous.iterrows():
                lines.append(f"  {r['roman']:<8}{int(r['integer']):>10}"
                             f"{int(r['ambiguous_count']):>8}"
                             f"  {notes.get(r['roman'], '')}")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _resolve_file_target(self, target: str) -> Optional[Path]:
        """Resolve target string to an existing CSV/Excel file. Supports
        ``#``/``*`` wildcard patterns (e.g. ``job_####`` → job_1234.csv)."""
        p = Path(target).expanduser()
        if p.is_absolute() and p.is_file():
            return p
        try:
            import vault_analyst as _va
            pat = _compile_name_pattern(target)
            matches = []
            for c in _va.list_csv_files(VAULT_DIR) + _va.list_excel_files(VAULT_DIR):
                if pat is not None:
                    if _name_matches_pattern(pat, c.name):
                        matches.append(c)
                elif target.lower() in c.name.lower():
                    matches.append(c)
            if matches:
                # Deterministic pick when a pattern spans several files.
                return sorted(matches, key=lambda m: m.name)[0] if pat is not None else matches[0]
        except Exception:
            pass
        candidate = VAULT_DIR / target
        return candidate if candidate.is_file() else None

    def _compare_schemas_response(self, a: str, b: str):
        import vault_analyst as _va
        pa = self._resolve_file_target(a)
        pb = self._resolve_file_target(b)
        if not pa or not pb:
            missing = [n for n, p in [(a, pa), (b, pb)] if p is None]
            self._append_transcript(
                "Writer", f"Could not resolve: {', '.join(missing)}",
                "final",
            )
            self._set_status("● idle")
            return
        try:
            df = _va.compare_schemas(pa, pb)
        except Exception as exc:
            self._append_transcript("Writer",
                                    f"Schema compare failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        lines = [f"Schema diff: {pa.name}  vs  {pb.name}"]
        for _, row in df.iterrows():
            tag = {
                "in_both":      "  =", "only_in_a":   "- A",
                "only_in_b":    "+ B", "type_changed": "  ~",
            }.get(row['status'], "  ?")
            extra = ""
            if row['status'] == 'type_changed':
                extra = f"  ({row['dtype_a']} -> {row['dtype_b']})"
            lines.append(f"  {tag}  {row['column']}{extra}")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _column_types_response(self, target: str):
        import vault_analyst as _va
        p = self._resolve_file_target(target) if target else None
        if not p:
            self._append_transcript(
                "Writer",
                f"Could not resolve a file from {target!r}.", "final",
            )
            self._set_status("● idle")
            return
        try:
            df = _va.column_type_inferences(p)
        except Exception as exc:
            self._append_transcript("Writer",
                                    f"Type inference failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        lines = [f"Inferred column types for {p.name}:"]
        lines.append(f"  {'column':<35}{'dtype':<12}{'inferred kind':<25}"
                     f"{'non-null':>10}{'unique':>10}")
        for _, row in df.iterrows():
            lines.append(
                f"  {str(row['column'])[:34]:<35}{str(row['dtype'])[:11]:<12}"
                f"{str(row['inferred_kind'])[:24]:<25}"
                f"{int(row['non_null']):>10}"
                f"{int(row['unique']):>10}"
            )
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _resolve_data_file_or_hint(self, target: str):
        """Resolve `target` to a CSV/Excel file, or return (None, message)
        naming a few available data files to try. Shared by the quick-analytics
        commands. Supports the #/​* filename wildcards via _resolve_file_target."""
        p = self._resolve_file_target(target) if target else None
        if p:
            return p, None
        try:
            import vault_analyst as _va
            avail = (_va.list_csv_files(VAULT_DIR)
                     + _va.list_excel_files(VAULT_DIR))[:6]
            names = ", ".join(a.name for a in avail) if avail else \
                "(no CSV/Excel files found in the vault)"
        except Exception:
            names = "(could not list vault files)"
        lead = (f"Couldn't find a data file matching {target!r}."
                if target else
                "Name a data file — e.g. 'column stats in sales.csv'.")
        return None, lead + "\n  Available: " + names

    @staticmethod
    def _fmt_stat(v) -> str:
        """Compact numeric formatting for stat tables; '' for None / NaN."""
        if v is None:
            return ""
        try:
            f = float(v)
        except (TypeError, ValueError):
            return str(v)
        if f != f:                       # NaN
            return ""
        if f == int(f) and abs(f) < 1e15:
            return f"{int(f)}"
        return f"{f:.4g}"

    def _column_stats_response(self, target: str):
        """Per-column descriptive stats (with & without zeros). No model."""
        import vault_analyst as _va
        p, hint = self._resolve_data_file_or_hint(target)
        if not p:
            self._append_transcript("Writer", hint, "final")
            self._set_status("● idle")
            return
        try:
            df = _va.column_stats(p)
        except Exception as exc:
            self._append_transcript("Writer",
                                    f"Column stats failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        if "error" in df.columns:
            self._append_transcript(
                "Writer", f"Could not read {p.name}: {df.iloc[0]['error']}",
                "final")
            self._set_status("● idle")
            return
        fs = self._fmt_stat
        lines = [f"Column summary for {p.name}:"]
        num = df[df["kind"] == "numeric"]
        txt = df[df["kind"] == "text"]
        if len(num):
            lines.append("")
            lines.append("Numeric columns "
                         "(mean/median shown as all | excluding-zeros):")
            lines.append(
                f"  {'column':<22}{'n':>7}{'nulls':>7}{'zeros':>7}"
                f"{'min':>10}{'max':>10}{'mean':>12}{'mean≠0':>12}"
                f"{'median':>10}{'med≠0':>10}{'std':>10}{'sum':>12}")
            for _, r in num.iterrows():
                lines.append(
                    f"  {str(r['column'])[:21]:<22}"
                    f"{int(r['count']):>7}{int(r['nulls']):>7}{int(r['zeros']):>7}"
                    f"{fs(r['min']):>10}{fs(r['max']):>10}"
                    f"{fs(r['mean']):>12}{fs(r['mean_nonzero']):>12}"
                    f"{fs(r['median']):>10}{fs(r['median_nonzero']):>10}"
                    f"{fs(r['std']):>10}{fs(r['sum']):>12}")
        if len(txt):
            lines.append("")
            lines.append("Text / other columns:")
            lines.append(f"  {'column':<28}{'n':>8}{'nulls':>8}"
                         f"{'unique':>8}   top value")
            for _, r in txt.iterrows():
                top = (f"{r['top']} (×{int(r['top_count'])})"
                       if r['top'] else "")
                lines.append(
                    f"  {str(r['column'])[:27]:<28}"
                    f"{int(r['count']):>8}{int(r['nulls']):>8}"
                    f"{int(r['unique']):>8}   {str(top)[:40]}")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _missing_data_response(self, target: str):
        """Per-column nulls + fully-complete row count. No model."""
        import vault_analyst as _va
        p, hint = self._resolve_data_file_or_hint(target)
        if not p:
            self._append_transcript("Writer", hint, "final")
            self._set_status("● idle")
            return
        try:
            rep = _va.missing_data_report(p)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Missing-data report failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        total = rep["total_rows"]
        lines = [
            f"Missing-data report for {p.name}:",
            f"  {rep['complete_rows']} of {total} row(s) fully complete "
            f"({rep['complete_pct']}%).",
            "",
            f"  {'column':<32}{'nulls':>10}{'null %':>9}{'non-null':>10}",
        ]
        for c in rep["columns"]:
            lines.append(
                f"  {str(c['column'])[:31]:<32}{int(c['nulls']):>10}"
                f"{c['null_pct']:>8}%{int(c['non_null']):>10}")
        if not rep["columns"]:
            lines.append("  (no columns)")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _duplicates_response(self, target: str):
        """Exact duplicate-row count + a few examples. No model."""
        import vault_analyst as _va
        p, hint = self._resolve_data_file_or_hint(target)
        if not p:
            self._append_transcript("Writer", hint, "final")
            self._set_status("● idle")
            return
        try:
            rep = _va.duplicate_rows_report(p)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Duplicate scan failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        n = rep["duplicate_rows"]
        lines = [f"Duplicate-row scan for {p.name}:"]
        if n == 0:
            lines.append(f"  No exact duplicate rows — all "
                         f"{rep['total_rows']} row(s) are unique.")
        else:
            lines.append(
                f"  {n} duplicate row(s) (redundant copies) of "
                f"{rep['total_rows']} total; {rep['unique_rows']} unique.")
            sample = rep["sample"]
            if len(sample):
                lines.append("")
                lines.append("  Examples:")
                cols = list(sample.columns)
                for _, row in sample.iterrows():
                    cells = "  |  ".join(f"{c}={row[c]}" for c in cols)
                    lines.append("    " + cells[:180])
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _top_values_response(self, target: str):
        """Most frequent values per column (frequency table). No model."""
        import vault_analyst as _va
        p, hint = self._resolve_data_file_or_hint(target)
        if not p:
            self._append_transcript("Writer", hint, "final")
            self._set_status("● idle")
            return
        try:
            rep = _va.top_values_per_column(p, top_n=5)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Top-values failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        lines = [f"Most frequent values per column in {p.name}:"]
        for c in rep["columns"]:
            vals = ", ".join(f"{v} (×{ct})" for v, ct in c["values"]) \
                or "(all null)"
            lines.append(f"  {str(c['column'])[:30]:<30} "
                         f"[{c['unique']} unique]  {vals[:120]}")
        if rep.get("truncated"):
            lines.append("  … (more columns not shown)")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _correlations_response(self, target: str):
        """Strongest numeric correlations between columns. No model."""
        import vault_analyst as _va
        p, hint = self._resolve_data_file_or_hint(target)
        if not p:
            self._append_transcript("Writer", hint, "final")
            self._set_status("● idle")
            return
        try:
            df = _va.numeric_correlations(p, top_n=15)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Correlation scan failed: {exc!r}", "final")
            self._set_status("● idle")
            return
        if df is None or len(df) == 0:
            self._append_transcript(
                "Writer",
                f"{p.name} has fewer than two numeric columns with variation "
                "— nothing to correlate.", "final")
            self._set_status("● idle")
            return
        lines = [f"Strongest numeric correlations in {p.name}:",
                 f"  {'column A':<24}{'column B':<24}{'corr':>8}"]
        for _, r in df.iterrows():
            lines.append(f"  {str(r['col_a'])[:23]:<24}"
                         f"{str(r['col_b'])[:23]:<24}"
                         f"{float(r['corr']):>8.3f}")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _resolve_image_target(self, target: str):
        """Resolve a target string to (folder, file). Returns (Path|None,
        Path|None): a directory of images, or a single image file."""
        import image_stats as _ims
        import data_index as _di
        t = (target or "").strip().strip("'\"`")
        try:
            din = _di.input_dir(VAULT_DIR)
        except Exception:
            din = VAULT_DIR / "data_in"
        if not t:
            return (din if din.exists() else None), None
        p = Path(t).expanduser()
        if p.is_absolute():
            if p.is_dir():
                return p, None
            if p.is_file():
                return None, p
        for base in (din, VAULT_DIR):
            cand = base / t
            if cand.is_dir():
                return cand, None
            if cand.is_file():
                return None, cand
        # bare image name anywhere under data_in
        try:
            for q in din.rglob("*"):
                if (q.is_file() and q.name.lower() == t.lower()
                        and q.suffix.lower() in _ims._IMAGE_SUFFIXES):
                    return None, q
        except Exception:
            pass
        return None, None

    def _image_stats_response(self, target: str):
        """Per-image pixel statistics, or a folder rollup. Model-free."""
        import image_stats as _ims
        root, fpath = self._resolve_image_target(target)
        if root is None and fpath is None:
            self._append_transcript(
                "Writer", f"Could not find an image or image folder for "
                f"{target!r}.", "final")
            self._set_status("● idle")
            return
        self._set_status("● analysing images…", "#cba6f7")
        if fpath is not None:
            s = _ims.image_pixel_stats(fpath)
            if "error" in s:
                self._append_transcript("Writer", s["error"], "final")
                self._set_status("● idle")
                return
            lines = [f"Pixel statistics for {s['file']}:",
                     f"  {s['width']}x{s['height']} px  ({s['megapixels']} MP, "
                     f"{s['mode']}, {s['format']}, {s.get('size_kb')} KB)",
                     f"  brightness (0-255): {s['brightness']}   "
                     f"contrast: {s['contrast']}"]
            for c, cs in s.get("channels", {}).items():
                lines.append(f"  {c}: mean {cs['mean']:>6}  std {cs['std']:>6}  "
                             f"min {cs['min']:>3}  max {cs['max']:>3}")
            dc = s.get("dominant_colors") or []
            if dc:
                lines.append("  dominant colours: " + ", ".join(
                    f"rgb{tuple(x['rgb'])} ({int(x['fraction']*100)}%)"
                    for x in dc[:5]))
            self._append_transcript("Writer", "\n".join(lines), "final")
            self._set_status("● idle")
            return
        rep = _ims.aggregate_image_folder(root)
        if rep.get("count", 0) == 0:
            self._append_transcript(
                "Writer", f"No images found under {root.name}/.", "final")
            self._set_status("● idle")
            return
        if "error" in rep:
            self._append_transcript("Writer", str(rep["error"]), "final")
            self._set_status("● idle")
            return

        def _fmt(d):
            return (f"mean {d['mean']}  (min {d['min']} – max {d['max']})"
                    if isinstance(d, dict) else "n/a")
        lines = [f"Image folder stats for {root.name}/ "
                 f"({rep['count']} image(s)):",
                 "  formats: " + ", ".join(f"{k} x{v}"
                                           for k, v in rep["by_format"].items()),
                 f"  width:  {_fmt(rep.get('width'))}",
                 f"  height: {_fmt(rep.get('height'))}",
                 f"  brightness: {_fmt(rep.get('brightness'))}",
                 f"  contrast:   {_fmt(rep.get('contrast'))}",
                 f"  darkest: {rep.get('darkest')}   "
                 f"brightest: {rep.get('brightest')}"]
        if rep.get("truncated"):
            lines.append("  (capped at 500 images)")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _detect_features_response(self, polarity: str, target: str, expected):
        """Detect + count discrete features/objects in an image (classical CV,
        no model), compare to an expected count, and write an annotated image
        with each counted feature boxed + numbered to data_out/annotated/."""
        import feature_detect as _fd
        import data_index as _di
        _root, fpath = self._resolve_image_target(target)
        if fpath is None:
            self._append_transcript(
                "Writer",
                f"Point me at a single image file — I couldn't resolve "
                f"{target!r} to one image. Try 'count features in "
                "layer_0345.png'.", "final")
            self._set_status("● idle")
            return
        pol = (polarity or "auto").lower()
        if pol == "light":
            pol = "bright"
        exp = None
        if expected is not None and str(expected).strip():
            try:
                exp = int(str(expected).strip())
            except Exception:
                exp = None
        self._set_status("● detecting features…", "#cba6f7")

        def _worker():
            try:
                out_dir = _di.output_dir(VAULT_DIR) / "annotated"
            except Exception:
                out_dir = None
            try:
                r = _fd.detect_and_count_features(
                    fpath, polarity=pol, expected=exp, out_dir=out_dir)
            except Exception as exc:
                r = {"error": f"detection failed: {exc!r}"}

            def _apply():
                if "error" in r:
                    self._append_transcript("Writer", str(r["error"]), "final")
                    self._set_status("● idle")
                    return
                lines = [f"Detected {r['count']} {r['polarity']} feature(s) in "
                         f"{r['file']}  (threshold {r['threshold']}, "
                         f"min area {r['min_area']} px)."]
                if "expected" in r:
                    if r.get("matches_expected"):
                        lines.append(f"  ✓ matches the expected {r['expected']}.")
                    else:
                        d = r["difference"]
                        lines.append(
                            f"  ✗ expected {r['expected']}, found {r['count']} "
                            f"({'+' if d > 0 else ''}{d}). If off, try the other "
                            "polarity (bright/dark) or a clearer image.")
                if r.get("annotated_image"):
                    lines.append(f"  Annotated image → {r['annotated_image']}")
                feats = r.get("features") or []
                if feats:
                    sizes = sorted((f["area"] for f in feats), reverse=True)[:8]
                    lines.append("  largest feature areas (px): "
                                 + ", ".join(str(s) for s in sizes))
                self._append_transcript("Writer", "\n".join(lines), "final")
                self._set_status("● idle")
            self.after(0, _apply)

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _ocr_response(self, target: str):
        """Extract text rendered inside an image (charts, scanned tables)."""
        import image_stats as _ims
        _root, fpath = self._resolve_image_target(target)
        if fpath is None:
            self._append_transcript(
                "Writer", f"Could not find an image named {target!r}.", "final")
            self._set_status("● idle")
            return
        self._set_status("● reading image text…", "#cba6f7")
        rep = _ims.ocr_image(fpath)
        if "error" in rep:
            self._append_transcript("Writer", str(rep["error"]), "final")
        elif not rep.get("text"):
            self._append_transcript(
                "Writer", f"No text found in {rep.get('file', target)}.", "final")
        else:
            self._append_transcript(
                "Writer",
                f"Text extracted from {rep['file']} ({rep['chars']} chars):\n\n"
                + rep["text"], "final")
        self._set_status("● idle")

    def _clean_folder_phrase(self, s: str) -> str:
        """Strip 'all csvs in' / 'the files in' prefixes and folder/vault noise
        from a folder phrase, leaving a folder target ('' -> data_in)."""
        s = (s or "").strip().strip("'\"`")
        s = _re.sub(
            r"^(?:all\s+|the\s+)*(?:csv\s+files?|csvs?|excels?|xlsx?|"
            r"spreadsheets?|data\s+files?|files?|data)\s+"
            r"(?:files?\s+)?(?:in|under|within|inside|from|of)\s+",
            "", s, flags=_re.IGNORECASE)
        s = _re.sub(r"^(?:all\s+|the\s+)+", "", s, flags=_re.IGNORECASE).strip()
        s = self._FOLDER_NOISE_RE.sub("", s).strip()
        return s

    def _folder_agg_response(self, agg_word, column, rest):
        """Compute one aggregation (mean/sum/min/max/median/std/count) of a
        column across every CSV/Excel in a folder, and optionally write the
        result to a CSV/TSV/TXT under the vault output folder. No model."""
        import vault_analyst as _va
        agg_word = (agg_word or "mean").strip()
        column = (column or "").strip().strip("'\"`")
        rest = (rest or "").strip()

        # 1) Detect + strip a trailing "…save/write/export to <file>" clause.
        save = False
        save_fmt = ""
        out_name = None
        m = self._SAVE_CLAUSE_RE.search(rest)
        if m and (m.group(1) or m.group(2)):
            save = True
            save_fmt = (m.group(1) or "").strip().lower()
            out_name = (m.group(2) or "").strip() or None
            rest = rest[:m.start()].strip()

        # 2) Optional "excluding zeros" modifier.
        _zero_re = _re.compile(
            r"[,\s]*(?:exclud\w*|without|ignor\w*|not?\s+counting|no)\s+"
            r"(?:the\s+)?zero(?:e?s)?\b", _re.IGNORECASE)
        exclude_zeros = bool(_zero_re.search(column) or _zero_re.search(rest))
        column = _zero_re.sub("", column).strip().strip("'\"`")
        rest = _zero_re.sub("", rest).strip()

        # 3) Resolve the folder.
        root = self._resolve_folder_target(self._clean_folder_phrase(rest))
        if not root.exists():
            self._append_transcript("Writer", f"Folder not found: {root}",
                                    "final")
            self._set_status("● idle")
            return

        canon = _va.canonical_agg(agg_word) or "mean"
        self._set_status("● calculating…", "#cba6f7")
        try:
            res = _va.folder_column_aggregate(
                root, column, canon, exclude_zeros=exclude_zeros)
        except Exception as exc:
            self._append_transcript("Writer",
                                    f"Calculation failed: {exc!r}", "final")
            self._set_status("● idle")
            return

        fs = self._fmt_stat
        zsuffix = " (excluding zeros)" if exclude_zeros else ""
        body = [f"{canon} of '{column}'{zsuffix} across "
                f"{res['files_scanned']} file(s) in {root.name or root}:"]
        if res["per_file"]:
            body.append(f"  {'file':<40}{'n':>8}{canon:>16}")
            for r in res["per_file"]:
                body.append(f"  {str(r['file'])[:39]:<40}"
                            f"{int(r['n']):>8}{fs(r['value']):>16}")
            body.append("")
            body.append(f"  OVERALL {canon} (n={res['overall_n']}): "
                        f"{fs(res['overall'])}")
        else:
            body.append(f"  No CSV/Excel under {root.name or root} has a "
                        f"column matching '{column}'.")
        if res["missing"] and res["per_file"]:
            body.append(f"  ({len(res['missing'])} file(s) had no matching "
                        f"column)")
        if res["truncated"]:
            body.append("  (scan capped at 200 files)")

        lines = list(body)
        if save and res["per_file"]:
            lines.append("")
            lines.append(self._save_stat_report(
                res, column, root, save_fmt, out_name, body))
        elif save:
            lines.append("")
            lines.append("  (nothing to save — no matching column found)")

        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _save_stat_report(self, res, column, root, save_fmt, out_name, body):
        """Write an aggregate result to a CSV/TSV/TXT under the vault output
        folder (data_out/reports/). Returns a status line for the transcript.
        Uses data_index.safe_write_path, which refuses to touch input data."""
        kind = "csv"
        if save_fmt in ("text", "txt", "plain text"):
            kind = "txt"
        elif save_fmt == "tsv":
            kind = "tsv"
        if out_name and "." in out_name:
            ext = out_name.rsplit(".", 1)[1].lower()
            kind = {"txt": "txt", "text": "txt", "tsv": "tsv",
                    "csv": "csv"}.get(ext, kind)
        base = out_name or f"{res['agg']}_{column}_{root.name or 'vault'}"
        if "." in base:
            base = base.rsplit(".", 1)[0]
        base = _re.sub(r"[^\w\-]+", "_", base).strip("_") or "stat_report"
        ext = {"txt": ".txt", "tsv": ".tsv", "csv": ".csv"}[kind]
        fname = base + ext
        try:
            outp = self.data_index.safe_write_path(fname, subfolder="reports")
        except Exception as exc:
            return f"  ⚠ Could not save the report: {exc!r}"
        try:
            if kind in ("csv", "tsv"):
                import pandas as _pd
                recs = [{"file": r["file"], "column": r["matched_column"],
                         "n": r["n"], res["agg"]: r["value"]}
                        for r in res["per_file"]]
                recs.append({"file": "OVERALL", "column": column,
                             "n": res["overall_n"], res["agg"]: res["overall"]})
                _pd.DataFrame(recs).to_csv(
                    outp, index=False, sep=("\t" if kind == "tsv" else ","))
            else:
                outp.write_text("\n".join(body) + "\n", encoding="utf-8")
        except Exception as exc:
            return f"  ⚠ Could not write {outp.name}: {exc!r}"
        return f"  ✓ Saved to: {outp}"

    def _app_tools_response(self):
        """List the self-authored (app-built) tools the models have created."""
        try:
            import app_built_tools as _abt
            tools = _abt.list_tools(vault_dir=VAULT_DIR)
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Could not list app-built tools: {exc!r}", "final")
            self._set_status("● idle")
            return
        if not tools:
            self._append_transcript(
                "Writer",
                "No app-built tools yet. The council and agent create these "
                "automatically when a needed capability is missing; each is "
                "sandbox-validated and saved UNREVIEWED under App_Built_tools/ "
                "in your vault.", "final")
            self._set_status("● idle")
            return
        lines = [f"App-built tools ({len(tools)}) — UNREVIEWED, may be "
                 f"inaccurate:"]
        for t in tools:
            lines.append(
                f"  • {t.get('name')}  [{t.get('author', '?')}]"
                f"  — {(t.get('description') or '')[:80]}")
        lines.append("")
        lines.append(f"Location: {_abt.tools_dir(VAULT_DIR)}")
        lines.append("(You can delete any file there to remove a tool.)")
        self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _money_response(self, target: str):
        import vault_tools as _vt
        root = self._resolve_folder_target(target)
        hits = _vt.find_money_amounts(root, max_hits=50)
        if not hits:
            self._append_transcript(
                "Writer", f"No currency amounts found in text files under {root.name}/.", "final",
            )
        else:
            lines = [f"Found {len(hits)} currency amount(s):"]
            for h in hits[:30]:
                lines.append(f"  {h['amount']:<20}  {h['path']}")
            if len(hits) > 30:
                lines.append(f"  ... ({len(hits)-30} more)")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _element_ranking_response(self, target: str, *, top_n: int = 10):
        """Resolve `target` to a folder, tally atomic elements found in
        every text/Excel file under it, and report the top N."""
        import vault_tools as _vt
        if not target:
            try:
                import data_index
                root = data_index.input_dir(VAULT_DIR)
            except Exception:
                root = VAULT_DIR / "data_in"
        else:
            p = Path(target).expanduser()
            if p.is_absolute() and p.exists():
                root = p
            else:
                candidate = VAULT_DIR / target
                root = candidate if candidate.exists() else p

        if not root.exists() or not root.is_dir():
            self._append_transcript(
                "Writer", f"Folder not found: {root}", "final",
            )
            self._set_status("● idle")
            return

        self._append_transcript(
            "Council",
            f"Scanning {root} for atomic-element mentions "
            f"(proper names + case-sensitive symbols)...",
            "observation",
        )
        self._set_status("● scanning…", "#cba6f7")

        def _worker():
            try:
                df = _vt.find_atomic_elements_in_folder(root)
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"element scan failed: {exc!r}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            text = _vt.format_element_ranking(df, top_n=top_n)
            self.after(0, lambda: (
                self._append_transcript("Writer", text, "final"),
                self._set_status("● idle"),
            ))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _list_files_response(self, target: str):
        """Deterministic file listing — every file in <folder> + size,
        recursive. No model involvement; no risk of hallucination."""
        root = self._resolve_folder_target(target)
        if not root.exists() or not root.is_dir():
            self._append_transcript("Writer", f"Folder not found: {root}", "final")
            self._set_status("● idle")
            return
        try:
            import conversation_logger as _cl
        except Exception:
            _cl = None
        SKIP_DIRS = {"__pycache__", ".git", ".venv", "venv", "node_modules"}
        files = []
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if any(part in SKIP_DIRS or part.startswith(".")
                   for part in p.relative_to(root).parts[:-1]):
                continue
            if _cl is not None:
                try:
                    if _cl.is_protected_path(p, root.parent):
                        continue
                except Exception:
                    pass
            try:
                files.append((p, p.stat().st_size))
            except Exception:
                continue
        files.sort(key=lambda fs: str(fs[0]).lower())
        if not files:
            self._append_transcript("Writer",
                                    f"{root}/ contains no files.", "final")
        else:
            lines = [f"Files in {root}/ ({len(files)} total):"]
            for p, size in files[:200]:
                rel = p.relative_to(root)
                size_kb = size / 1024
                size_str = (f"{size_kb:>8.1f} KB" if size_kb < 1024
                            else f"{size_kb/1024:>7.1f} MB")
                lines.append(f"  {size_str}  {rel}")
            if len(files) > 200:
                lines.append(f"  ... ({len(files) - 200} more — refine with a "
                             f"subfolder or 'look at files in <subfolder>')")
            self._append_transcript("Writer", "\n".join(lines), "final")
        self._set_status("● idle")

    def _look_at_files_response(self, target: str):
        """Inject the folder block then let the Writer answer the user's
        question with full visibility into every file's columns/sheets."""
        root = self._resolve_folder_target(target)
        if not root.exists() or not root.is_dir():
            self._append_transcript("Writer", f"Folder not found: {root}", "final")
            self._set_status("● idle")
            return
        block = _render_folder_for_injection(root)
        if not block:
            self._append_transcript("Writer",
                                    f"Could not inspect {root}.", "final")
            self._set_status("● idle")
            return
        # For the no-model case, just render it directly to the transcript.
        # Strip the wrapper since [FOLDER:] tags are prompt-injection-only.
        text = block
        if text.startswith("[FOLDER:"):
            text = text.split("\n", 1)[1] if "\n" in text else text
        if text.endswith("[END FOLDER]"):
            text = text.rsplit("\n", 1)[0]
        self._append_transcript("Writer",
                                f"Inspection of {root}/ (no model used):\n\n{text}",
                                "final")
        self._set_status("● idle")

    def _list_folders_response(self, target: str):
        """Resolve `target` (a path, a vault-relative folder name, or
        empty for `data_in/`) and list its immediate subfolders."""
        import vault_tools as _vt
        if not target:
            # Default: data_in/ — the canonical user-data scope
            try:
                import data_index
                root = data_index.input_dir(VAULT_DIR)
            except Exception:
                root = VAULT_DIR / "data_in"
        else:
            p = Path(target).expanduser()
            if p.is_absolute() and p.exists():
                root = p
            else:
                # Try interpreting as a vault-relative path
                candidate = VAULT_DIR / target
                root = candidate if candidate.exists() else p

        if not root.exists() or not root.is_dir():
            self._append_transcript(
                "Writer",
                f"Folder not found: {root}",
                "final",
            )
            self._set_status("● idle")
            return

        folders = _vt.list_subfolders(root, max_depth=2)
        text = _vt.format_subfolder_listing(root, folders)
        self._append_transcript("Writer", text, "final")
        self._set_status("● idle")

    def _show_task_memo_response(self):
        """Print the current task memo to the transcript without
        re-condensing or modifying it. Useful when the user wants to
        spot-check the condenser's interpretation of an earlier query.
        """
        memo = self.task_memory.current()
        if memo is None or memo.is_empty():
            self._append_transcript(
                "Writer",
                "No task memo set yet. The next non-meta question will "
                "create one. Type 'reset memo' any time to clear it.",
                "final",
            )
            return
        lines: list = [
            "Current task memo",
            "─────────────────",
            f"goal: {memo.goal}",
        ]
        if memo.constraints:
            lines.append("constraints:")
            for c in memo.constraints:
                lines.append(f"  • {c}")
        if memo.forbidden:
            lines.append("forbidden:")
            for f in memo.forbidden:
                lines.append(f"  • {f}")
        lines.append("")
        lines.append(
            "Source query: " + (memo.raw_query[:160]
                                + ("…" if len(memo.raw_query) > 160 else ""))
        )
        if memo.is_extension:
            lines.append("(This memo extended a previous one — constraints "
                         "inherited from the prior turn.)")
        lines.append("")
        lines.append(
            "Type 'reset memo' to start fresh on the next question."
        )
        self._append_transcript("Writer", "\n".join(lines), "final")

    def _diagnostics_chat_response(self):
        """Render the dependency-check report into the transcript as a
        Writer-final message. Same content as the Diagnostics tab —
        useful for sharing in bug reports because chat output flows
        into the transcript export.
        """
        try:
            import dependency_check as _dc
            text = _dc.render_as_text()
        except Exception as exc:
            text = f"dependency_check failed: {exc!r}"
        self._append_transcript("Writer", text, "final")

    def _show_learned_synonyms_response(self):
        """Render every (term -> [tokens-from-vault]) pair the semantic
        expansion layer has cached so far. Lets the user spot bad
        categorizations ("ah, the model thinks 'denim' is a metal —
        let me clear that").
        """
        idx = _get_vault_index()
        if idx is None:
            self._append_transcript(
                "Writer", "Vault index unavailable.", "final",
            )
            return
        try:
            cache = idx._load_semantic_cache()
        except Exception as exc:
            self._append_transcript(
                "Writer", f"Could not read learned synonyms: {exc!r}", "final",
            )
            return
        entries = cache.get("entries", {}) if isinstance(cache, dict) else {}
        if not entries:
            self._append_transcript(
                "Writer",
                "No learned synonyms yet. The model builds these "
                "automatically the first time you search for a category "
                "word that isn't literally in your files "
                "(e.g. 'metals', 'fabrics', 'weapons'). After the first "
                "such search the result is cached on disk.\n\n"
                "Cache file: vault/semantic_cache.json",
                "final",
            )
            return
        lines: list = [
            "Learned semantic categories",
            "───────────────────────────",
            "(The model decided which of your vault's actual tokens "
            "belong in each category. Cached per-vault on disk.)",
            "",
        ]
        for term in sorted(entries):
            expansions = entries[term] or []
            if expansions:
                preview = ", ".join(expansions[:10])
                if len(expansions) > 10:
                    preview += f", … (+{len(expansions) - 10} more)"
                lines.append(f"  {term}  ->  {preview}")
            else:
                lines.append(f"  {term}  ->  (no matches in this vault)")
        lines.append("")
        lines.append(
            f"Vocab hash: {cache.get('vocab_hash', '(unknown)')}"
        )
        lines.append(
            "Type 'clear learned synonyms' to wipe the cache. The "
            "next category search will recompute against the current vocab."
        )
        self._append_transcript("Writer", "\n".join(lines), "final")

    def _clear_learned_synonyms_response(self):
        """Wipe the semantic-expansion cache file. The next category
        query will recompute fresh against the current vocab."""
        idx = _get_vault_index()
        if idx is None:
            self._append_transcript(
                "Writer", "Vault index unavailable.", "final",
            )
            return
        try:
            cache_path = idx._semantic_cache_path()
        except Exception:
            self._append_transcript(
                "Writer", "Could not locate semantic cache file.", "final",
            )
            return
        if cache_path.exists():
            try:
                cache_path.unlink()
                self._append_transcript(
                    "Writer",
                    "Learned synonyms cleared. The next category-shaped "
                    "search will ask the model fresh.",
                    "final",
                )
                return
            except Exception as exc:
                self._append_transcript(
                    "Writer", f"Could not delete cache: {exc!r}", "final",
                )
                return
        self._append_transcript(
            "Writer", "No learned-synonym cache to clear.", "final",
        )

    def _reset_task_memo_response(self):
        """Drop the current task memo. Next user query starts fresh —
        no constraints / forbidden are inherited."""
        had_memo = not self.task_memory.current().is_empty()
        self.task_memory.reset()
        if had_memo:
            self._append_transcript(
                "Writer",
                "Task memo cleared. The next question will start a new "
                "memo from scratch — no constraints carried over.",
                "final",
            )
        else:
            self._append_transcript(
                "Writer",
                "No task memo to clear.",
                "final",
            )

    def _context_info_response(self):
        """Report current context-window configuration and last-query usage.

        Helps the user decide whether to raise `COUNCIL_GGUF_N_CTX`. We
        report:
          • current n_ctx (env var)
          • model's advertised maximum (from GGUF metadata if loaded)
          • last assembled prompt's token count (if any has been sent)
          • a one-line "safe input" budget (n_ctx minus the reply reserve)
          • the recommended way to raise the cap

        This is a Writer-final message so it doesn't trigger the council
        deliberation pipeline — the user just wants to see the numbers.
        """
        try:
            import council_engine as _ce
            n_ctx = _ce.get_n_ctx()
            max_ctx = _ce.get_model_max_context()
            loaded = (max_ctx is not None)
        except Exception as exc:
            self._append_transcript(
                "Writer",
                f"Could not read context info: {exc!r}",
                "final",
            )
            return

        lines: list[str] = ["Context-window status"]
        lines.append(f"  • Configured n_ctx:   {n_ctx:,} tokens "
                     f"(set via COUNCIL_GGUF_N_CTX)")
        if max_ctx:
            lines.append(f"  • Model advertises:   {max_ctx:,} tokens max")
            if max_ctx > n_ctx:
                head = max_ctx - n_ctx
                lines.append(f"    → You have {head:,} tokens of headroom "
                             f"to raise the cap.")
        else:
            lines.append("  • Model advertises:   (unknown — model not "
                         "loaded yet, ask any question first)")

        # Reply reserve mirrors context_budget_report()
        reply_reserve = max(256, int(n_ctx * 0.25))
        safe_input = max(1, n_ctx - reply_reserve)
        lines.append(f"  • Reply reserve:      ~{reply_reserve:,} tokens "
                     f"(25% of n_ctx, floor 256)")
        lines.append(f"  • Safe input budget:  ~{safe_input:,} tokens "
                     f"before the reply gets squeezed")

        last = getattr(self, "_last_context_budget", None)
        if isinstance(last, dict) and last.get("input_tokens"):
            used = last["input_tokens"]
            pct = last.get("pct_of_window", 0.0)
            tag = "exact" if last.get("tokenizer") == "exact" else "estimated"
            lines.append("")
            lines.append(f"Last query: ~{used:,} tokens ({tag}, "
                         f"{pct:.0f}% of window)")
            if last.get("over_window"):
                lines.append("  ⚠ Exceeded n_ctx — the tail was silently "
                             "truncated. Raise the cap and try again.")
            elif last.get("over_safe"):
                lines.append("  ⚠ Ate into the reply reserve — answers will "
                             "be short.")

        # Per-injection-block breakdown — shows where the budget went last
        # turn. Lets the user see e.g. "the CSV ate 3,200 tokens; the vault
        # match was only 400" and decide whether to raise n_ctx or use a
        # smaller subset of the data.
        bd = getattr(self, "_last_injection_breakdown", None)
        if isinstance(bd, dict) and (bd.get("costs") or bd.get("dropped")):
            lines.append("")
            lines.append("Last injection breakdown:")
            ut = bd.get("user_text_tokens") or 0
            cap = bd.get("per_block_cap") or 0
            running = bd.get("running") or 0
            remaining = bd.get("remaining") or 0
            lines.append(f"  per-block cap: ~{cap:,} tokens  ·  "
                         f"injection budget: ~{remaining:,} tokens  ·  "
                         f"typed text: ~{ut:,} tokens")
            for label, cost in bd.get("costs") or []:
                lines.append(f"    {label}  ~{cost:,} tokens")
            lines.append(f"  total injected: ~{running:,} tokens")
            if bd.get("dropped"):
                lines.append("  dropped (budget exceeded):")
                for label, cost in bd["dropped"]:
                    lines.append(f"    {label}  ~{cost:,} tokens  (dropped)")

        # Task-memo status — show whether the RAM-resident sticky note
        # is active so the user sees what's being re-injected on every
        # turn (and can `reset memo` if they want to start fresh).
        try:
            memo = self.task_memory.current()
        except Exception:
            memo = None
        if memo is not None and not memo.is_empty():
            lines.append("")
            lines.append("Task memo (re-injected every turn):")
            lines.append(f"  goal: {memo.goal[:120]}"
                         + ("…" if len(memo.goal) > 120 else ""))
            if memo.constraints:
                lines.append(f"  constraints: {len(memo.constraints)} "
                             f"({'; '.join(memo.constraints[:2])}"
                             + ('; …' if len(memo.constraints) > 2 else '')
                             + ')')
            if memo.forbidden:
                lines.append(f"  forbidden:   {len(memo.forbidden)}")
            lines.append("  (Type 'show memo' for full text, "
                         "'reset memo' to clear.)")

        lines.append("")
        lines.append("To raise the cap before launch:")
        lines.append("  set COUNCIL_GGUF_N_CTX=16384   (Windows cmd)")
        lines.append("  $env:COUNCIL_GGUF_N_CTX=\"16384\"  (PowerShell)")
        lines.append("  export COUNCIL_GGUF_N_CTX=16384  (bash)")
        lines.append("Doubling n_ctx roughly doubles KV-cache RAM. If the "
                     "GPU runs out, drop COUNCIL_GGUF_GPU_LAYERS to spill "
                     "to CPU.")

        self._append_transcript("Writer", "\n".join(lines), "final")

    def _build_embeddings_response(self):
        """Kick off the vector embedding build over the vault index.

        Uses a sentence-transformer (~80 MB) to make retrieval semantic
        rather than keyword-only. First call may take a moment to load
        the model; subsequent builds reuse the cached vectors and only
        re-embed records whose mtime changed.
        """
        idx = _get_vault_index()
        if idx is None:
            self._append_transcript(
                "Writer", "Vault index unavailable.", "final",
            )
            return
        try:
            idx.rebuild()
        except Exception:
            pass
        emb = idx.embeddings()
        if emb is None:
            self._append_transcript(
                "Writer",
                "sentence-transformers is not available. Install with:\n"
                "  pip install sentence-transformers\n"
                "Then re-run 'build embeddings'.",
                "final",
            )
            return
        self._append_transcript(
            "Council",
            f"Embedding {len(idx.records)} files with model "
            f"{emb.model_name}. First run may download the model "
            f"(~80 MB). Subsequent rebuilds only touch changed files.",
            "observation",
        )
        self._set_status("● embedding…", "#cba6f7")

        def _worker():
            def _progress(i, total, name):
                if i % 10 == 0 or i == total:
                    self.after(0, lambda: self._set_status(
                        f"● embedding {i}/{total}…", "#cba6f7"
                    ))
            try:
                n = idx.build_embeddings(on_progress=_progress)
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"embedding build failed: {exc!r}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            stats = emb.stats()
            self.after(0, lambda: (
                self._append_transcript(
                    "Writer",
                    f"Vector index ready — {stats['vectors']} files embedded "
                    f"({stats['dim']}-dim, {stats['size_kb']} KB on disk). "
                    f"Semantic search will now blend into vault queries.",
                    "final",
                ),
                self._set_status("● idle"),
            ))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _build_semantic_index_response(self):
        """Kick off the LLM description pass over the vault index."""
        idx = _get_vault_index()
        if idx is None:
            self._append_transcript(
                "Writer", "Vault index is unavailable.", "final",
            )
            return
        # Ensure the keyword index is up-to-date first
        try:
            idx.rebuild()
        except Exception:
            pass
        pending = sum(1 for r in idx.records.values()
                      if not r.get("description"))
        if pending == 0:
            self._append_transcript(
                "Writer",
                f"All {len(idx.records)} indexed files already have semantic "
                f"descriptions. (Type 'refresh descriptions' to regenerate.)",
                "final",
            )
            self._set_status("● idle")
            return
        self._append_transcript(
            "Council",
            f"Building semantic descriptions for {pending} files. This runs "
            f"in the background; the chat stays usable. Each file takes "
            f"~3-10 seconds.",
            "observation",
        )
        self._set_status("● indexing…", "#cba6f7")

        def _worker():
            def _progress(i, total, name):
                if i % 5 == 0 or i == total:
                    self.after(0, lambda: self._set_status(
                        f"● indexing {i}/{total}…", "#cba6f7"
                    ))
            try:
                n = idx.generate_descriptions(on_progress=_progress)
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"semantic index failed: {exc!r}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            self.after(0, lambda: (
                self._append_transcript(
                    "Writer",
                    f"Semantic index complete — {n} files described.",
                    "final",
                ),
                self._set_status("● idle"),
            ))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _build_topics_only_response(self):
        """Faster description build — generates ONLY the keyword topics
        (no prose summary). On CPU-only inference this is roughly 3x
        faster than the full description path on the same record set,
        because the model emits ~24 tokens instead of ~80 per file.

        Tabular files (CSV / Excel / Parquet / SQLite / DuckDB) get
        their description + topics from the schema with zero model
        calls regardless — that's pure dict work.
        """
        idx = _get_vault_index()
        if idx is None:
            self._append_transcript(
                "Writer", "Vault index is unavailable.", "final",
            )
            return
        try:
            idx.rebuild()
        except Exception:
            pass
        pending = sum(1 for r in idx.records.values()
                      if not r.get("topics"))
        if pending == 0:
            self._append_transcript(
                "Writer",
                f"All {len(idx.records)} indexed files already have "
                f"topic keywords. (Type 'refresh topics' to regenerate.)",
                "final",
            )
            self._set_status("● idle")
            return
        self._append_transcript(
            "Council",
            f"Building topic keywords for {pending} files. Tabular "
            f"files are instant; JSON/text files batch through the "
            f"model in groups of 4 for ~3x faster generation.",
            "observation",
        )
        self._set_status("● topics…", "#cba6f7")

        def _worker():
            def _progress(i, total, name):
                if i % 5 == 0 or i == total:
                    self.after(0, lambda: self._set_status(
                        f"● topics {i}/{total}…", "#cba6f7"
                    ))
            try:
                n = idx.generate_descriptions(
                    topics_only=True, batch_size=4, on_progress=_progress,
                )
            except Exception as exc:
                self.after(0, lambda exc=exc: (
                    self._append_transcript("Writer",
                                            f"topic build failed: {exc!r}",
                                            "final"),
                    self._set_status("● idle"),
                ))
                return
            self.after(0, lambda: (
                self._append_transcript(
                    "Writer",
                    f"Topic keywords complete — {n} files tagged.",
                    "final",
                ),
                self._set_status("● idle"),
            ))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    # ---- Workflow intent handler ----

    _WORKFLOW_RE = _re.compile(
        r"^\s*run\s+(?:the\s+)?workflow\b", _re.IGNORECASE,
    )

    def _handle_workflow_intent(self, user_text: str) -> bool:
        """Detect 'run workflow ...' commands and execute them in a worker."""
        if not user_text:
            return False
        single_line = user_text.split("\n", 1)[0]
        if not self._WORKFLOW_RE.match(single_line):
            return False

        import workflow_runner as _wr
        spec = _wr.parse_workflow_request(single_line, VAULT_DIR)
        if not spec.pipeline_paths:
            self._append_transcript(
                "Writer",
                "I couldn't resolve any pipelines from that workflow. Try "
                "`list pipelines` to see what's available, then `run "
                "workflow <name1>, <name2>, ...`",
                "final",
            )
            return True

        mode_label = {
            "linear":   "linear (each pipeline once)",
            "per_file": f"per-file over {spec.input_dir} (pattern: {spec.pattern})",
            "per_step": f"per-step over {spec.input_dir} (pattern: {spec.pattern})",
        }.get(spec.mode, spec.mode)

        names = "\n  ".join(p.name for p in spec.pipeline_paths)
        self._append_transcript(
            "Council",
            f"Starting workflow ({mode_label})\n  {names}",
            "observation",
        )
        self._set_status("● workflow…", "#fab387")

        def _post_transcript(who: str, text: str, tag: str) -> None:
            # Tk widget calls must happen on the main thread.
            self.after(0, lambda: self._append_transcript(who, text, tag))

        def _post_status(label: str, color: Optional[str] = None) -> None:
            self.after(0, lambda: self._set_status(label, color))

        def worker():
            def on_step(step: "_wr.StepResult") -> None:
                if step.success:
                    msg = (f"  [ok ]  #{step.step_index} {step.pipeline_name} "
                           f"({step.input_label}) — {step.duration_s:.1f}s")
                else:
                    msg = (f"  [FAIL]  #{step.step_index} {step.pipeline_name} "
                           f"({step.input_label}) — {step.error or 'failed'}")
                _post_transcript("Workflow", msg, "observation")

            try:
                result = _wr.run_workflow(spec, on_step=on_step)
            except Exception as exc:
                _post_transcript("Writer", f"Workflow runner crashed: {exc!r}", "final")
                _post_status("● idle", None)
                return

            _post_transcript("Writer", result.summary(), "final")
            _post_status("● idle", None)

        import threading as _th
        _th.Thread(target=worker, daemon=True).start()
        return True

    # ---- Council tab ----

    def _build_council_tab(self):
        self.tab_council = ttk.Frame(self.nb)
        self.nb.add(self.tab_council, text="⚖ Council")

        # ── Backend strip ───────────────────────────────────────────────
        # Lets the user switch between Ollama and a locally-loaded GGUF
        # file without restarting the app. Settings persist to
        # vault/backend_settings.json.
        self._build_backend_strip(self.tab_council)

        # Main paned window: transcript | judge panel
        paned = tk.PanedWindow(self.tab_council, orient="horizontal",
                               bg="#1a1414", sashwidth=6, sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=6, pady=6)

        # Left: transcript
        left = ttk.Frame(paned)
        paned.add(left, minsize=500)

        ttk.Label(left, text="Transcript").pack(anchor="w")
        self.transcript = self._make_text(left, wrap="word", state="disabled")
        self._register_transcript_tags()
        sb = ttk.Scrollbar(left, command=self.transcript.yview)
        self.transcript.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.transcript.pack(fill="both", expand=True)

        # Right: judge + live stream preview.
        # In DEMO_MODE the chat is a single-personality Q&A — no panel,
        # no verdicts — so the right pane isn't useful to show. We still
        # CREATE the widgets so downstream code that calls _set_judge or
        # _vfb_show stays safe; we just don't add the pane to the layout.
        _demo = bool(getattr(branding, "DEMO_MODE", False))
        right = ttk.Frame(paned)
        if not _demo:
            paned.add(right, minsize=280)

        ttk.Label(right, text="Judge Panel").pack(anchor="w")
        self.judge_box = self._make_text(right, wrap="word", width=40, state="disabled", height=14)
        self.judge_box.pack(fill="both", expand=True)

        # ── Verdict feedback bar (shown after each deliberation) ────────────
        self._vfb_frame = ttk.Frame(right)
        self._vfb_frame.pack(fill="x", pady=(4, 0))
        self._vfb_frame.pack_forget()  # hidden until verdict arrives

        _vfb_lbl = ttk.Label(self._vfb_frame, text="Do you agree with the verdict?",
                             foreground="#d4d4d4")
        _vfb_lbl.pack(side="left", padx=(0, 6))

        self._vfb_agree_btn = ttk.Button(
            self._vfb_frame, text="✓ Agree",
            command=self._verdict_agree,
        )
        self._vfb_agree_btn.pack(side="left", padx=2)

        self._vfb_disagree_btn = ttk.Button(
            self._vfb_frame, text="✗ Disagree",
            command=self._verdict_disagree_open,
        )
        self._vfb_disagree_btn.pack(side="left", padx=2)

        # ── Disagree detail panel (hidden until Disagree clicked) ────────────
        self._vfb_detail = ttk.Frame(right)
        self._vfb_detail.pack(fill="x", pady=(2, 0))
        self._vfb_detail.pack_forget()

        ttk.Label(self._vfb_detail,
                  text="Your objection (the council will re-deliberate with it):",
                  foreground="#fab387").pack(anchor="w")
        self._vfb_text = self._make_text(self._vfb_detail, height=3, wrap="word")
        self._vfb_text.pack(fill="x")
        self._vfb_text.bind("<Control-Return>", lambda e: self._verdict_disagree_submit())

        _vfb_sub_row = ttk.Frame(self._vfb_detail)
        _vfb_sub_row.pack(fill="x", pady=(2, 0))
        ttk.Button(_vfb_sub_row, text="↩ Re-deliberate with objection",
                   command=self._verdict_disagree_submit).pack(side="left")
        ttk.Button(_vfb_sub_row, text="Cancel",
                   command=self._verdict_disagree_cancel).pack(side="left", padx=6)
        ttk.Label(_vfb_sub_row, text="Ctrl+Enter to submit",
                  foreground="#6c7086").pack(side="left")

        ttk.Label(right, text="Live Token Stream").pack(anchor="w", pady=(8, 0))
        self.stream_box = self._make_text(right, wrap="word", width=40, height=10, state="disabled")
        self.stream_box.pack(fill="both", expand=True)

        # Bottom input area
        bottom = ttk.Frame(self.tab_council)
        bottom.pack(fill="x", padx=6, pady=(0, 6))

        ttk.Label(bottom, text="Input").pack(anchor="w")
        self.input = self._make_text(bottom, wrap="word", height=4)
        self.input.pack(fill="x")
        self.input.bind("<Control-Return>", lambda e: self._send())

        btns = ttk.Frame(bottom)
        btns.pack(fill="x", pady=(4, 0))

        ttk.Button(btns, text="Send  [Ctrl+Enter]", command=self._send).pack(side="left")
        ttk.Button(btns, text="\U0001f4ca Find & Chart",
                   command=self._council_find_and_chart_button
                   ).pack(side="left", padx=6)
        ttk.Button(btns, text="\U0001f50d Look Up",
                   command=self._council_lookup_button
                   ).pack(side="left", padx=6)
        ttk.Button(btns, text="Clear", command=lambda: self._set_text(self.input, "")).pack(side="left", padx=6)
        ttk.Button(btns, text="⤓ Defer to Vault",
                   command=self._defer_to_vault).pack(side="left", padx=6)
        # Enabled only after a fast (direct-route) answer — re-asks the SAME
        # question through the full multi-role council for a prose discussion.
        self._expand_btn = ttk.Button(
            btns, text="⤢ Expand with council", state="disabled",
            command=self._council_expand_with_council)
        self._expand_btn.pack(side="left", padx=6)
        ttk.Button(btns, text="💾 Save answer",
                   command=self._save_council_answer).pack(side="left", padx=6)
        ttk.Button(btns, text="🕘 History",
                   command=self._show_question_history).pack(side="left", padx=6)
        ttk.Button(btns, text="💡 What can I ask?",
                   command=self._show_examples).pack(side="left", padx=6)

        # License / trial badge — clickable to open activation dialog.
        # In DEMO_MODE the whole element is hidden since there's no
        # licensing surface to show.
        self._license_badge_var = tk.StringVar(value="")
        if not getattr(branding, "DEMO_MODE", False):
            license_btn = tk.Button(
                btns, textvariable=self._license_badge_var,
                relief="flat", borderwidth=0, padx=8, pady=2,
                bg="#231a1a", fg="#a6e3a1", activebackground="#3a2828",
                font=("Segoe UI", 9, "bold"), cursor="hand2",
                command=lambda: activation_dialog.open_activation_dialog(
                    self, VAULT_DIR,
                    on_status_change=self._on_license_status_change,
                    blocking=False,
                ),
            )
            license_btn.pack(side="right", padx=(0, 6))

        # Manual specialist override — set to a specialist's id to force it
        # onto every query until the user picks "Auto" again.
        self._forced_specialist_id = None
        ttk.Label(btns, text="  Ask:", foreground="#7a7575").pack(side="left", padx=(10, 2))
        self._spec_pin_var = tk.StringVar(value="Auto")
        self._spec_pin_cb  = ttk.Combobox(
            btns, textvariable=self._spec_pin_var,
            state="readonly", width=22,
        )
        self._spec_pin_cb.pack(side="left")
        self._spec_pin_cb.bind("<<ComboboxSelected>>",
                               lambda _e: self._spec_pin_changed())
        # Populate now and refresh whenever the registry changes
        self._spec_pin_refresh()

        # In DEMO_MODE the chat is single-personality (Writer wears any
        # active specialist lens). The deliberation toggle is forced off
        # at run-time anyway (see _send), but defaulting it here keeps
        # the hidden state consistent should any code peek at it.
        _demo = bool(getattr(branding, "DEMO_MODE", False))
        self.var_deliberate      = tk.BooleanVar(value=not _demo)
        self.var_tools           = tk.BooleanVar(value=False)
        self.var_fill_ide        = tk.BooleanVar(value=True)
        self.var_stream          = tk.BooleanVar(value=True)
        self.var_adversarial     = tk.BooleanVar(value=False)  # T2-B: adversarial Peasant
        self.var_judge_panel     = tk.BooleanVar(value=False)  # Judge model picks panel
        self.var_robust_voices   = tk.BooleanVar(value=False)  # Robust personality voices
        # Learned user-profile injection. The checkbox is the EXPLICIT
        # bypass for sessions where the user's durable preferences don't
        # fit the current ask — unchecking skips injection on the very
        # next message while learning continues underneath. Initial
        # state honours a pre-set COUNCIL_QUIRKS_APPLY from the shell.
        self.var_use_profile     = tk.BooleanVar(
            value=ce.user_profile_apply_enabled())

        # Wire voice toggle to apply/remove voices immediately on change
        self.var_robust_voices.trace_add("write", self._on_voice_toggle)
        self.var_use_profile.trace_add("write", self._on_profile_toggle)

        # ── Toolbar row 1: core toggles ───────────────────────────────
        # Deliberation/Adversarial/Judge-panel only make sense for the
        # multi-personality build. Hidden in DEMO_MODE for a clean
        # ask-a-question UX.
        if not _demo:
            ttk.Checkbutton(btns, text="Deliberation",    variable=self.var_deliberate).pack(side="left", padx=4)
        ttk.Checkbutton(btns, text="Tools",           variable=self.var_tools).pack(side="left", padx=4)
        ttk.Checkbutton(btns, text="Fill IDE",        variable=self.var_fill_ide).pack(side="left", padx=4)
        ttk.Checkbutton(btns, text="Stream tokens",   variable=self.var_stream).pack(side="left", padx=4)
        # Visible in DEMO_MODE too — the learned profile shapes the
        # single-personality answers just the same.
        ttk.Checkbutton(btns, text="👤 Profile",      variable=self.var_use_profile).pack(side="left", padx=4)
        if not _demo:
            ttk.Checkbutton(btns, text="Adversarial",     variable=self.var_adversarial).pack(side="left", padx=4)
            ttk.Checkbutton(btns, text="Judge panel ✦",   variable=self.var_judge_panel).pack(side="left", padx=4)

        # ── Toolbar row 2: personality controls ──────────────────────
        # Robust-voices switch makes the panel members sound distinct —
        # not relevant when there's only one personality answering, so
        # it's hidden in DEMO_MODE.
        if not _demo:
            btns2 = ttk.Frame(bottom)
            btns2.pack(fill="x", pady=(2, 0))
            ttk.Label(btns2, text="Personalities:", foreground="#6c7086").pack(side="left", padx=(4,2))
            ttk.Checkbutton(btns2, text="Robust voices ✦",
                            variable=self.var_robust_voices).pack(side="left", padx=4)
            ttk.Label(btns2,
                      text="(gives each personality a distinct character and tone)",
                      foreground="#6c7086", font=("", 8)).pack(side="left", padx=4)

        # ── Council instructions bar ─────────────────────────────
        inst_row = ttk.Frame(bottom)
        inst_row.pack(fill="x", pady=(2, 0))
        ttk.Label(inst_row, text="⚡ Instruction:",
                  foreground="#fab387").pack(side="left", padx=(4, 2))
        self._inst_var  = tk.StringVar()
        self._inst_name = tk.StringVar()
        ttk.Entry(inst_row, textvariable=self._inst_name, width=18,
                  ).pack(side="left", padx=2)
        ttk.Label(inst_row, text="Name (optional)",
                  foreground="#6c7086", font=("", 8)).pack(side="left")
        inst_entry = ttk.Entry(inst_row, textvariable=self._inst_var, width=44)
        inst_entry.pack(side="left", padx=4, fill="x", expand=True)
        inst_entry.bind("<Return>", lambda e: self._apply_council_instruction())
        ttk.Button(inst_row, text="Add  [Enter]",
                   command=self._apply_council_instruction).pack(side="left", padx=2)
        ttk.Button(inst_row, text="Manage…",
                   command=self._open_instruction_manager).pack(side="left", padx=2)
        ttk.Button(inst_row, text="Content Style…",
                   command=self._open_content_style).pack(side="left", padx=2)
        self._inst_active_lbl = ttk.Label(inst_row, text="",
                                           foreground="#a6e3a1", font=("", 8))
        self._inst_active_lbl.pack(side="left", padx=6)

        # ── Save output panel (hidden until deliberation completes) ──────────
        self._save_frame = ttk.Frame(bottom)
        # not packed until output is ready
        self._last_final_text = ""
        self._last_query_text = ""
        self._last_route      = ""

        _save_hdr = ttk.Frame(self._save_frame)
        _save_hdr.pack(fill="x", pady=(4, 2))
        ttk.Label(_save_hdr, text="💾 Save output as:",
                  foreground="#d32f2f", font=("", 9, "bold")).pack(side="left", padx=4)
        ttk.Button(_save_hdr, text="📄 Text file (.txt)",
                   command=self._save_output_txt).pack(side="left", padx=4)
        ttk.Button(_save_hdr, text="📝 Markdown (.md)",
                   command=self._save_output_md).pack(side="left", padx=4)
        ttk.Button(_save_hdr, text="📐 LaTeX (.tex)",
                   command=self._save_output_latex).pack(side="left", padx=4)
        ttk.Button(_save_hdr, text="✕",
                   command=self._hide_save_buttons).pack(side="right", padx=4)

        # ── Clarification panel (hidden until a personality asks a question) ──
        self._clarif_frame = ttk.Frame(bottom)
        # Not packed yet — shown only when needed

        _clarif_hdr = ttk.Frame(self._clarif_frame)
        _clarif_hdr.pack(fill="x")
        ttk.Label(_clarif_hdr, text="🤔 A personality needs your input:",
                  foreground="#fab387", font=("", 9, "bold")).pack(side="left", padx=4)
        self._clarif_question_lbl = ttk.Label(
            self._clarif_frame, text="", foreground="#d4d4d4",
            wraplength=700, justify="left")
        self._clarif_question_lbl.pack(anchor="w", padx=8, pady=(2, 4))

        _clarif_input_row = ttk.Frame(self._clarif_frame)
        _clarif_input_row.pack(fill="x", padx=4)
        self._clarif_var = tk.StringVar()
        _clarif_entry = ttk.Entry(_clarif_input_row, textvariable=self._clarif_var, width=60)
        _clarif_entry.pack(side="left", padx=2, fill="x", expand=True)
        _clarif_entry.bind("<Return>", lambda e: self._submit_clarification())
        ttk.Button(_clarif_input_row, text="Answer  [Enter]",
                   command=self._submit_clarification).pack(side="left", padx=4)
        ttk.Button(_clarif_input_row, text="Skip",
                   command=lambda: self._submit_clarification(skip=True)).pack(side="left")

        self.status = ttk.Label(btns, text="● idle", foreground="#a6e3a1")
        self.status.pack(side="right")
        self.tps_label = ttk.Label(btns, text="", foreground="#d32f2f")
        self.tps_label.pack(side="right", padx=(0, 8))
        self._agent_label = ttk.Label(btns, text="", foreground="#a6e3a1")
        self._agent_label.pack(side="right", padx=(0, 6))

        # T2-D: Per-query model override dropdowns
        override_row = ttk.Frame(bottom)
        override_row.pack(fill="x", pady=(2, 0))
        ttk.Label(override_row, text="Override: ", foreground="#6c7086").pack(side="left")
        _backend_opts = [
            "(default)", "local_general_primary", "local_general_alt",
            "local_coder_primary", "local_coder_fast",
            "local_judge_fast", "local_peasant_fast", "local_fast",
        ]
        self._query_overrides: Dict[str, tk.StringVar] = {}
        for _role in ("writer", "coder", "intern", "skeptic", "artist"):
            ttk.Label(override_row, text=_role[:4] + ":", foreground="#6c7086").pack(side="left")
            _v = tk.StringVar(value="(default)")
            self._query_overrides[_role] = _v
            ttk.Combobox(override_row, textvariable=_v, values=_backend_opts,
                         width=16, state="readonly").pack(side="left", padx=(0, 6))

    def _register_transcript_tags(self):
        self.transcript.tag_configure("phase",   foreground=PHASE_COLOR,   font=("Consolas", 9, "italic"))
        self.transcript.tag_configure("token",   foreground=TOKEN_COLOR)
        for role, color in ROLE_COLORS.items():
            tag = f"who_{role.lower().replace('-','_').replace(' ','_')}"
            self.transcript.tag_configure(tag, foreground=color, font=("Consolas", 10, "bold"))
        self.transcript.tag_configure("who_default", foreground=DEFAULT_COLOR, font=("Consolas", 10, "bold"))
        self.transcript.tag_configure("error", foreground="#f38ba8")

    # ---- Dream3D tab ----
    # A split view: chat mirror on the left, pipeline visualization on the
    # right. The transcript here mirrors the Council transcript via
    # _append_transcript (which writes to both widgets when present). The
    # input box uses the same _send pipeline so all existing intent
    # handlers (show / modify / run workflow) work from this tab too.

    def _build_dream3d_tab(self):
        self.tab_dream3d = ttk.Frame(self.nb)
        self.nb.add(self.tab_dream3d, text="🧪 Dream3D")

        paned = tk.PanedWindow(self.tab_dream3d, orient="horizontal",
                               bg="#1a1414", sashwidth=6, sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=6, pady=6)

        # ── Left: mirrored chat ──────────────────────────────────────
        left = ttk.Frame(paned)
        paned.add(left, minsize=420)

        ttk.Label(left, text="Chat (mirrors Council)").pack(anchor="w")
        self.dream3d_transcript = self._make_text(left, wrap="word", state="disabled")
        # Reuse the same role-tag config the Council transcript uses
        try:
            for tag in self.transcript.tag_names():
                cfg = self.transcript.tag_cget(tag, "foreground")
                if cfg:
                    self.dream3d_transcript.tag_configure(tag, foreground=cfg,
                                                          font=("Consolas", 10, "bold"))
        except Exception:
            pass

        d3d_sb = ttk.Scrollbar(left, command=self.dream3d_transcript.yview)
        self.dream3d_transcript.configure(yscrollcommand=d3d_sb.set)
        d3d_sb.pack(side="right", fill="y")
        self.dream3d_transcript.pack(fill="both", expand=True)

        d3d_input_frame = ttk.Frame(left)
        d3d_input_frame.pack(fill="x", pady=(6, 0))
        self.dream3d_input = self._make_text(d3d_input_frame, wrap="word", height=3)
        self.dream3d_input.pack(fill="x")
        self.dream3d_input.bind("<Control-Return>",
                                lambda e: self._dream3d_send_from_input())
        btns = ttk.Frame(left)
        btns.pack(fill="x", pady=(4, 0))
        ttk.Button(btns, text="Send", command=self._dream3d_send_from_input).pack(side="left")
        ttk.Label(btns, text="  (Ctrl+Enter)", foreground="#7a7575").pack(side="left")

        # ── Right: pipeline picker + step viewer ─────────────────────
        right = ttk.Frame(paned)
        paned.add(right, minsize=360)

        ttk.Label(right, text="Pipelines in vault/pipelines/in/").pack(anchor="w")

        list_row = ttk.Frame(right)
        list_row.pack(fill="x", pady=(2, 0))

        self.dream3d_pipeline_list = tk.Listbox(list_row, height=8, exportselection=False)
        self.dream3d_pipeline_list.pack(side="left", fill="x", expand=True)
        plist_sb = ttk.Scrollbar(list_row, command=self.dream3d_pipeline_list.yview)
        self.dream3d_pipeline_list.configure(yscrollcommand=plist_sb.set)
        plist_sb.pack(side="left", fill="y")
        self.dream3d_pipeline_list.bind("<<ListboxSelect>>",
                                        lambda _e: self._dream3d_show_selected())
        self.dream3d_pipeline_list.bind("<Double-Button-1>",
                                        lambda _e: self._dream3d_show_selected())

        list_btns = ttk.Frame(right)
        list_btns.pack(fill="x", pady=(2, 4))
        ttk.Button(list_btns, text="↻ Refresh",
                   command=self._dream3d_refresh_pipelines).pack(side="left")
        ttk.Button(list_btns, text="Open in/ folder",
                   command=self._dream3d_open_in_folder).pack(side="left", padx=4)
        ttk.Button(list_btns, text="Open out/ folder",
                   command=self._dream3d_open_out_folder).pack(side="left", padx=4)

        # Geometry sidekick — interactive 3D cube that reads out the
        # equivalent 4×4 transformation matrix. Standalone HTML, opens
        # in the default browser; useful for sanity-checking an
        # orientation before committing it to a pipeline.
        tools_row = ttk.Frame(right)
        tools_row.pack(fill="x", pady=(0, 4))
        ttk.Label(tools_row, text="Tools:",
                  foreground="#7a7575").pack(side="left", padx=(0, 4))
        ttk.Button(tools_row, text="🧊 Transformation Cube…",
                   command=self._open_transformation_cube_tool
                   ).pack(side="left")

        ttk.Label(right, text="Pipeline visualization").pack(anchor="w", pady=(6, 0))
        self.dream3d_view = self._make_text(right, wrap="word", state="disabled")
        d3d_view_sb = ttk.Scrollbar(right, command=self.dream3d_view.yview)
        self.dream3d_view.configure(yscrollcommand=d3d_view_sb.set)
        d3d_view_sb.pack(side="right", fill="y")
        self.dream3d_view.pack(fill="both", expand=True)

        # Initial population
        self._dream3d_refresh_pipelines()

    def _dream3d_refresh_pipelines(self):
        """Re-scan vault/pipelines/in/ and populate the listbox."""
        try:
            import pipeline_scanner as _ps
            in_dir = _ps.vault_pipelines_in_dir(VAULT_DIR)
            pipelines = _ps.scan_pipelines(in_dir)
        except Exception as exc:
            self._dream3d_set_view(f"Pipeline scan failed: {exc!r}")
            return

        self._dream3d_pipelines_cache = pipelines
        self.dream3d_pipeline_list.delete(0, "end")
        if not pipelines:
            self.dream3d_pipeline_list.insert("end", "(none — drop .py / .dream3d files here)")
            self._dream3d_set_view(
                "No pipelines found.\n\n"
                "Add simplnx .py scripts or .dream3d files to:\n"
                f"  {in_dir}\n\n"
                "Then click ↻ Refresh."
            )
            return
        for pl in pipelines:
            note = f"  ({pl.format}, {len(pl.steps)} step{'s' if len(pl.steps)!=1 else ''})"
            self.dream3d_pipeline_list.insert("end", pl.name + note)

    def _dream3d_show_selected(self):
        sel = self.dream3d_pipeline_list.curselection()
        if not sel:
            return
        idx = sel[0]
        cache = getattr(self, "_dream3d_pipelines_cache", [])
        if idx < 0 or idx >= len(cache):
            return
        pl = cache[idx]
        try:
            import pipeline_scanner as _ps
            rendered = _ps.render_pipeline(pl)
        except Exception as exc:
            rendered = f"render failed: {exc!r}"
        self._dream3d_set_view(rendered)

    def _dream3d_set_view(self, text: str):
        self.dream3d_view.configure(state="normal")
        self.dream3d_view.delete("1.0", "end")
        self.dream3d_view.insert("1.0", text)
        self.dream3d_view.configure(state="disabled")

    def _dream3d_open_in_folder(self):
        import pipeline_scanner as _ps
        self._open_in_explorer(_ps.vault_pipelines_in_dir(VAULT_DIR))

    def _dream3d_open_out_folder(self):
        import pipeline_scanner as _ps
        self._open_in_explorer(_ps.vault_pipelines_out_dir(VAULT_DIR))

    def _open_transformation_cube_tool(self):
        """Open the bundled interactive 3D cube → 4×4 transformation matrix
        tool in the user's default browser. The HTML is dependency-free
        (Canvas 2D, no external libraries) so it runs fully offline.

        Useful alongside Dream3D pipelines that need a quick sanity check
        on an orientation matrix or a hand-built rotation. Lives in
        ``assets/transformation_cube.html`` so PyInstaller bundles it.
        """
        import webbrowser
        # Resolve from APP_DIR so it works in both the source checkout and
        # a PyInstaller --onedir bundle (sys._MEIPASS for --onefile, but
        # the spec ships assets/ to the bundle root).
        candidates = [
            APP_DIR / "assets" / "transformation_cube.html",
        ]
        # PyInstaller --onefile extracts to sys._MEIPASS at runtime.
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidates.append(Path(meipass) / "assets" / "transformation_cube.html")
        for html in candidates:
            if html.is_file():
                try:
                    webbrowser.open(html.resolve().as_uri())
                    self._append_transcript(
                        "Council",
                        f"Opened Transformation Cube tool ({html.name}) in "
                        "your default browser.",
                        "observation",
                    )
                except Exception as exc:
                    self._append_transcript(
                        "Council",
                        f"Couldn't launch browser for {html}: {exc!r}",
                        "observation",
                    )
                return
        self._append_transcript(
            "Council",
            "Transformation Cube tool not found — expected at "
            f"{candidates[0]}. Re-installing the assets/ folder should "
            "restore it.",
            "observation",
        )

    def _open_in_explorer(self, path: Path):
        try:
            os.startfile(str(path))  # Windows
        except Exception:
            try:
                import subprocess as _sp
                _sp.Popen(["explorer", str(path)])  # fallback
            except Exception:
                self._append_transcript(
                    "Council", f"Couldn't open folder: {path}", "observation",
                )

    def _dream3d_send_from_input(self):
        """Route the Dream3D tab's input through the Council _send flow."""
        text = self.dream3d_input.get("1.0", "end").strip()
        if not text:
            return
        # Push into the Council input widget and trigger _send so all the
        # existing intent handlers (show / modify / run workflow / chat)
        # apply uniformly.
        self._set_text(self.input, text)
        self._set_text(self.dream3d_input, "")
        self._send()
        # Refresh the pipeline list after — modifying a pipeline produces
        # a new file in out/, but in/ is unchanged. Keep refresh available
        # via the button.

    # ---- IDE tab ----

    def _build_ide_tab(self):
        self.tab_ide = ttk.Frame(self.nb)
        self.nb.add(self.tab_ide, text="💻 IDE / Runner")

        # Script name bar
        name_row = ttk.Frame(self.tab_ide)
        name_row.pack(fill="x", padx=10, pady=(8, 2))
        ttk.Label(name_row, text="Script:").pack(side="left")
        self.script_name_var = tk.StringVar(value=self.current_script_name)
        ttk.Entry(name_row, textvariable=self.script_name_var, width=40).pack(side="left", padx=6)
        ttk.Button(name_row, text="Apply", command=self._apply_script_name).pack(side="left")
        ttk.Button(name_row, text="Clear Code", command=lambda: self._set_text(self.ide_code, "")).pack(side="left", padx=6)

        paned = tk.PanedWindow(self.tab_ide, orient="horizontal", bg="#1a1414", sashwidth=6)
        paned.pack(fill="both", expand=True, padx=10, pady=4)

        left = ttk.Frame(paned)
        paned.add(left, minsize=400)

        ttk.Label(left, text="Code").pack(anchor="w")
        self.ide_code = self._make_text(left, wrap="none", font=("Consolas", 11))
        self.ide_code.pack(fill="both", expand=True)

        right = ttk.Frame(paned)
        paned.add(right, minsize=300)

        ttk.Label(right, text="Output").pack(anchor="w")
        self.ide_out = self._make_text(right, wrap="word", state="disabled")
        self.ide_out.tag_configure("stderr", foreground="#f38ba8")
        self.ide_out.tag_configure("info",   foreground="#d32f2f")
        self.ide_out.pack(fill="both", expand=True)

        btns = ttk.Frame(self.tab_ide)
        btns.pack(fill="x", padx=10, pady=(0, 8))
        ttk.Button(btns, text="▶  Run (streaming)", command=self._ide_run_stream).pack(side="left")
        ttk.Button(btns, text="Run (blocking)",     command=self._ide_run).pack(side="left", padx=6)
        ttk.Button(btns, text="Snapshot to Vault",  command=self._ide_snapshot).pack(side="left")
        ttk.Button(btns, text="Clear Output",
                   command=lambda: self._set_text(self.ide_out, "")).pack(side="left", padx=6)

    # ---- Tool Creation (Tool Forge) tab ----

    def _build_tool_forge_tab(self):
        """A tab where the local model writes a NEW tool for a described task.
        The generated code is sandbox-validated (read-only; no delete/write/
        network) and saved UNREVIEWED to App_Built_tools/. Council and Agent
        route here via 'create a tool that ...' when a needed tool is missing."""
        self.tab_forge = ttk.Frame(self.nb)
        self.nb.add(self.tab_forge, text="🛠 Tool Creation")
        self._forge_tools: list = []

        ttk.Label(
            self.tab_forge,
            text=("Describe a tool you need — the local model writes it, the "
                  "sandbox validates it (read-only; no delete / write / network "
                  "/ shell), and it's saved UNREVIEWED to App_Built_tools/. "
                  "Council & Agent route here when a tool is missing "
                  "(try: “create a tool that …”)."),
            wraplength=920, justify="left",
        ).pack(anchor="w", padx=10, pady=(8, 2))

        task_row = ttk.Frame(self.tab_forge)
        task_row.pack(fill="x", padx=10, pady=(2, 2))
        ttk.Label(task_row, text="Task:").pack(anchor="w")
        self.forge_task = self._make_text(task_row, height=3, wrap="word")
        self.forge_task.pack(fill="x")

        btns = ttk.Frame(self.tab_forge)
        btns.pack(fill="x", padx=10, pady=(2, 4))
        ttk.Button(btns, text="✨ Generate Tool",
                   command=self._forge_generate).pack(side="left")
        ttk.Button(btns, text="💾 Save Edited Code",
                   command=self._forge_save).pack(side="left", padx=6)
        ttk.Button(btns, text="▶ Run Selected",
                   command=self._forge_run).pack(side="left")
        ttk.Button(btns, text="⟳ Refresh",
                   command=self._forge_refresh_list).pack(side="left", padx=6)

        paned = tk.PanedWindow(self.tab_forge, orient="horizontal",
                               bg="#1a1414", sashwidth=6)
        paned.pack(fill="both", expand=True, padx=10, pady=4)

        left = ttk.Frame(paned)
        paned.add(left, minsize=420)
        ttk.Label(left, text="Tool code (editable — review before trusting)"
                  ).pack(anchor="w")
        self.forge_code = self._make_text(left, wrap="none",
                                          font=("Consolas", 11))
        self.forge_code.pack(fill="both", expand=True)

        right = ttk.Frame(paned)
        paned.add(right, minsize=300)
        ttk.Label(right, text="Existing app-built tools").pack(anchor="w")
        self.forge_list = tk.Listbox(right, height=8)
        self.forge_list.pack(fill="x")
        ttk.Label(right, text="Output").pack(anchor="w", pady=(6, 0))
        self.forge_out = self._make_text(right, wrap="word", state="disabled")
        self.forge_out.pack(fill="both", expand=True)

        self.forge_status = ttk.Label(self.tab_forge, text="Ready.", anchor="w")
        self.forge_status.pack(fill="x", padx=10, pady=(0, 8))
        try:
            self._forge_refresh_list()
        except Exception:
            pass

    def _forge_set_status(self, text: str):
        try:
            self.forge_status.config(text=text)
        except Exception:
            pass

    def _forge_out_write(self, text: str):
        try:
            self._set_text(self.forge_out, text)
        except Exception:
            pass

    def _forge_refresh_list(self):
        try:
            import app_built_tools as _abt
            tools = _abt.list_tools(vault_dir=VAULT_DIR)
        except Exception:
            tools = []
        self._forge_tools = tools
        try:
            self.forge_list.delete(0, "end")
            for t in tools:
                self.forge_list.insert(
                    "end",
                    f"{t.get('name')}  —  {(t.get('description') or '')[:50]}")
        except Exception:
            pass

    def _forge_generate(self):
        task = self.forge_task.get("1.0", "end").strip()
        if not task:
            self._forge_set_status("Describe a tool first.")
            return
        self._forge_set_status("Generating… the local model is writing the tool.")

        def _worker():
            import council_engine as _ce
            import tool_forge as _tf

            def _model_call(prompt):
                return _ce.local_chat(
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.2, num_predict=700, timeout=120)
            try:
                ok, msg, name, code = _tf.generate_tool(
                    task, _model_call, author="council", vault_dir=VAULT_DIR)
            except Exception as exc:
                ok, msg, name, code = False, f"error: {exc!r}", None, ""

            def _apply():
                if code:
                    self._set_text(self.forge_code, code)
                if ok:
                    self._forge_set_status(
                        f"✓ Saved '{name}' — UNREVIEWED. Select it and Run to test.")
                    self._forge_out_write(
                        f"{msg}\n\nThe tool is saved but UNREVIEWED — review the "
                        "code on the left. Select it in the list and press "
                        "'Run Selected' to try it.")
                    self._forge_refresh_list()
                else:
                    self._forge_set_status("⚠ " + str(msg)[:80])
                    self._forge_out_write(
                        str(msg) + "\n\nYou can edit the code on the left and "
                        "press 'Save Edited Code'.")
            self.after(0, _apply)

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _forge_save(self):
        code = self.forge_code.get("1.0", "end").strip()
        if not code:
            self._forge_set_status("Nothing to save — generate or paste code first.")
            return
        try:
            import tool_forge as _tf
            ok, msg, name = _tf.save_edited_tool(
                code, description="", author="user", vault_dir=VAULT_DIR)
        except Exception as exc:
            ok, msg, name = False, f"error: {exc!r}", None
        self._forge_set_status(f"✓ Saved '{name}'." if ok else "⚠ " + str(msg)[:80])
        self._forge_out_write(str(msg))
        if ok:
            self._forge_refresh_list()

    def _forge_run(self):
        sel = None
        try:
            idx = self.forge_list.curselection()
            if idx and 0 <= idx[0] < len(self._forge_tools):
                sel = self._forge_tools[idx[0]].get("name")
        except Exception:
            sel = None
        if not sel:
            self._forge_set_status("Select a tool in the list to run.")
            return
        self._forge_set_status(f"Running '{sel}'…")

        def _worker():
            import app_built_tools as _abt
            import data_index as _di
            try:
                root = _di.input_dir(VAULT_DIR)
                df, msg = _abt.run_tool(sel, {}, allowed_folders=[root],
                                        vault_dir=VAULT_DIR)
            except Exception as exc:
                df, msg = None, f"run failed: {exc!r}"

            def _apply():
                out = str(msg) + "\n"
                if df is not None:
                    try:
                        out += "\n" + df.head(30).to_string(index=False)
                    except Exception:
                        out += "\n" + str(df)[:2000]
                self._forge_out_write(out)
                self._forge_set_status("Done.")
            self.after(0, _apply)

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _forge_route(self, task: str):
        """Council/Agent (or the 'create a tool that …' command) routes a
        missing-tool request here: switch to the tab, prefill, and generate.
        Marshalled onto the UI thread (Tk isn't thread-safe)."""
        def _do():
            try:
                self.nb.select(self.tab_forge)
            except Exception:
                pass
            try:
                self._set_text(self.forge_task, task or "")
            except Exception:
                pass
            if task:
                self._forge_generate()
        try:
            self.after(0, _do)
        except Exception:
            pass

    # ---- Librarian tab ----

    def _build_librarian_tab(self):
        self.tab_lib = ttk.Frame(self.nb)
        self.nb.add(self.tab_lib, text="📚 Librarian")

        top = ttk.Frame(self.tab_lib)
        top.pack(fill="both", expand=True, padx=10, pady=10)

        left = ttk.Frame(top)
        left.pack(side="left", fill="both", expand=True)

        ttk.Label(left, text=f"Vault: {VAULT_DIR}").pack(anchor="w")

        self.vault_lb = tk.Listbox(left, bg="#231a1a", fg="#d4d4d4",
                                   selectbackground="#5a3030", relief="flat",
                                   font=("Consolas", 10))
        self.vault_lb.pack(fill="both", expand=True, pady=4)
        self.vault_lb.bind("<Double-Button-1>", lambda e: self._lib_preview())

        right = ttk.Frame(top)
        right.pack(side="right", fill="y", padx=(10, 0))

        ttk.Button(right, text="Refresh",       command=self._lib_refresh).pack(fill="x")
        ttk.Button(right, text="Preview",       command=self._lib_preview).pack(fill="x", pady=4)
        ttk.Button(right, text="Commit to Git", command=self._lib_commit).pack(fill="x")
        ttk.Button(right, text="Open Folder",   command=self._lib_open_vault).pack(fill="x", pady=4)

        self._lib_refresh()

    # ---- Sessions tab ----

    def _build_sessions_tab(self):
        self.tab_sessions = ttk.Frame(self.nb)
        self.nb.add(self.tab_sessions, text="🕓 Sessions")

        top = ttk.Frame(self.tab_sessions)
        top.pack(fill="both", expand=True, padx=10, pady=10)

        left = ttk.Frame(top)
        left.pack(side="left", fill="both", expand=True)

        ttk.Label(left, text="Past Sessions (double-click to load as prior context)").pack(anchor="w")
        _sf = ttk.Frame(left)
        _sf.pack(fill="x", pady=(2, 0))
        ttk.Label(_sf, text="🔍", foreground="#6c7086").pack(side="left")
        self._session_filter_var = tk.StringVar()
        self._session_filter_var.trace_add("write", lambda *_: self._sessions_refresh())
        ttk.Entry(_sf, textvariable=self._session_filter_var).pack(
            side="left", fill="x", expand=True, padx=(4, 0))
        self.session_lb = tk.Listbox(left, bg="#231a1a", fg="#d4d4d4",
                                     selectbackground="#5a3030", relief="flat",
                                     font=("Consolas", 10))
        self.session_lb.pack(fill="both", expand=True, pady=4)
        self.session_lb.bind("<Double-Button-1>", lambda e: self._sessions_load_prior())

        right = ttk.Frame(top)
        right.pack(side="right", fill="y", padx=(10, 0))

        ttk.Button(right, text="Refresh",           command=self._sessions_refresh).pack(fill="x")
        ttk.Button(right, text="Load as Prior",     command=self._sessions_load_prior).pack(fill="x", pady=4)
        ttk.Button(right, text="Preview Session",   command=self._sessions_preview).pack(fill="x")
        ttk.Button(right, text="Clear Prior",       command=self._sessions_clear_prior).pack(fill="x", pady=4)
        ttk.Button(right, text="Verdict History",   command=self._show_verdict_history).pack(fill="x", pady=(8,0))
        ttk.Separator(right, orient="horizontal").pack(fill="x", pady=8)
        ttk.Button(right, text="📊 Analyse Trends", command=self._sessions_analyse_trends).pack(fill="x")

        self.prior_label = ttk.Label(right, text="Prior: none", wraplength=180)
        self.prior_label.pack(anchor="w", pady=4)

        ttk.Label(top, text="Session Preview").pack(anchor="w")
        self.session_preview = self._make_text(self.tab_sessions, wrap="word", height=12, state="disabled")
        self.session_preview.pack(fill="x", padx=10, pady=(0, 10))

        self._sessions_refresh()

    # ---- Nodes tab ----

    def _build_nodes_tab(self):
        self.tab_nodes = ttk.Frame(self.nb)
        self.nb.add(self.tab_nodes, text="🖥 Nodes")

        top = ttk.Frame(self.tab_nodes)
        top.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Label(top, text="Ollama Node Status  (auto-refreshes every 15s)").pack(anchor="w")

        cols = ("host", "status", "latency", "active_models", "installed")
        self.nodes_tree = ttk.Treeview(top, columns=cols, show="headings", height=8)
        self.nodes_tree.heading("host",          text="Host")
        self.nodes_tree.heading("status",        text="Status")
        self.nodes_tree.heading("latency",       text="Latency")
        self.nodes_tree.heading("active_models", text="Active Models")
        self.nodes_tree.heading("installed",     text="Installed Models")
        self.nodes_tree.column("host",          width=220)
        self.nodes_tree.column("status",        width=80)
        self.nodes_tree.column("latency",       width=80)
        self.nodes_tree.column("active_models", width=140)
        self.nodes_tree.column("installed",     width=400)
        self.nodes_tree.pack(fill="both", expand=True, pady=4)

        self.nodes_tree.tag_configure("up",   foreground="#a6e3a1")
        self.nodes_tree.tag_configure("down", foreground="#f38ba8")

        btns = ttk.Frame(top)
        btns.pack(fill="x")
        ttk.Button(btns, text="Refresh Now", command=self._refresh_nodes_async).pack(side="left")
        self.nodes_status_label = ttk.Label(btns, text="")
        self.nodes_status_label.pack(side="left", padx=10)

        # Dispatcher hosts config
        ttk.Separator(top, orient="horizontal").pack(fill="x", pady=10)
        hrow = ttk.Frame(top)
        hrow.pack(fill="x")
        ttk.Label(hrow, text="Pi hosts (comma-separated URLs):").pack(side="left")
        self.pi_hosts_var = tk.StringVar(value=", ".join(ce.DEFAULT_PI_HOSTS))
        ttk.Entry(hrow, textvariable=self.pi_hosts_var, width=50).pack(side="left", padx=6)
        ttk.Button(hrow, text="Apply & Rebuild Dispatcher",
                   command=self._apply_pi_hosts).pack(side="left")

    # ---- Agents tab ----

    # ============================
    # Agent Jobs — autonomous, goal-driven background agents
    # ============================
    def _get_job_runner(self):
        """Lazily build the single background JobRunner (one daemon worker;
        the GGUF serialises inference so one worker is correct). Reconciles any
        job left RUNNING *or* QUEUED by a prior shutdown — its in-RAM queue is
        gone, so it can never progress; mark it failed so it isn't orphaned
        forever in the UI. MVP has no auto-resume."""
        jr = getattr(self, "_job_runner", None)
        if jr is not None:
            return jr
        import agent_jobs_runner as _ajr
        try:
            _fr = data_index.input_dir(VAULT_DIR)
        except Exception:
            _fr = VAULT_DIR
        jr = _ajr.JobRunner(vault_dir=VAULT_DIR, ui_q=self.ui_q, file_root=_fr)
        try:
            import agent_jobs as _aj
            _stale = {_aj.JobStatus.RUNNING.value, _aj.JobStatus.QUEUED.value}
            for _j in jr.store.all():
                if _j.status in _stale:
                    jr.store.set_status(_j.job_id, _aj.JobStatus.FAILED.value,
                                        "interrupted by a restart")
        except Exception:
            pass
        self._job_runner = jr
        return jr

    def _build_agent_jobs_tab(self):
        import tkinter as tk
        from tkinter import ttk
        self.tab_agent_jobs = ttk.Frame(self.nb)
        self.nb.add(self.tab_agent_jobs, text="🎯 Agent Jobs")
        self._aj_tree_iids = {}
        self._aj_iid_to_job = {}

        top = ttk.Frame(self.tab_agent_jobs)
        top.pack(fill="x", padx=10, pady=(8, 4))
        ttk.Label(top, font=("", 10, "bold"),
                  text="Give the agent a goal — it plans and runs read-only "
                       "steps in the background to achieve it.").pack(anchor="w")
        ttk.Label(top, foreground="#7a7575", justify="left",
                  text="Safe by design: it can only list / read / analyse your "
                       "data in data_in. No delete, no writes, no network — those "
                       "tools don't exist. Bounded by a step budget; cancellable.").pack(
            anchor="w", pady=(0, 4))
        self._aj_goal = self._make_text(top, wrap="word", height=3)
        self._aj_goal.pack(fill="x")
        row = ttk.Frame(top)
        row.pack(fill="x", pady=4)
        ttk.Label(row, text="Max steps:").pack(side="left")
        self._aj_maxsteps = tk.IntVar(value=6)
        ttk.Spinbox(row, from_=2, to=20, width=4,
                    textvariable=self._aj_maxsteps).pack(side="left", padx=(4, 10))
        ttk.Button(row, text="▶ Start job",
                   command=self._aj_start).pack(side="left")
        ttk.Button(row, text="⟳ Refresh",
                   command=self._aj_refresh).pack(side="left", padx=6)

        mid = ttk.Frame(self.tab_agent_jobs)
        mid.pack(fill="both", expand=True, padx=10, pady=4)
        cols = ("status", "goal", "steps")
        self._aj_tree = ttk.Treeview(mid, columns=cols, show="headings",
                                     height=8)
        for c, w in (("status", 90), ("goal", 520), ("steps", 60)):
            self._aj_tree.heading(c, text=c.title())
            self._aj_tree.column(c, width=w, anchor="w")
        self._aj_tree.pack(side="left", fill="both", expand=True)
        _sb = ttk.Scrollbar(mid, orient="vertical",
                            command=self._aj_tree.yview)
        _sb.pack(side="right", fill="y")
        self._aj_tree.configure(yscrollcommand=_sb.set)
        self._aj_tree.bind("<<TreeviewSelect>>",
                           lambda e: self._aj_show_selected())

        jb = ttk.Frame(self.tab_agent_jobs)
        jb.pack(fill="x", padx=10)
        ttk.Button(jb, text="■ Cancel", command=self._aj_cancel).pack(side="left")
        ttk.Button(jb, text="🔎 Open report",
                   command=self._aj_inspect).pack(side="left", padx=6)
        ttk.Button(jb, text="🗑 Remove finished",
                   command=self._aj_clear_finished).pack(side="left", padx=6)
        self._aj_status = tk.StringVar(value="")
        ttk.Label(jb, textvariable=self._aj_status,
                  foreground="#a6e3a1").pack(side="left", padx=10)

        ttk.Label(self.tab_agent_jobs, text="Step log").pack(anchor="w", padx=10)
        self._aj_log = self._make_text(self.tab_agent_jobs, wrap="word",
                                       height=10, state="disabled")
        self._aj_log.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        self._aj_refresh()

    def _aj_start(self):
        goal = self._aj_goal.get("1.0", "end").strip()
        if not goal:
            self._aj_status.set("Enter a goal first.")
            return
        jr = self._get_job_runner()
        jid = jr.submit(goal, max_steps=int(self._aj_maxsteps.get()))
        self._set_text(self._aj_goal, "")
        self._aj_status.set(f"Started {jid}. Watch the step log below.")
        self._aj_log_append(f"▶ {jid}: {goal}")
        self._aj_refresh()

    def _aj_selected_job_id(self):
        sel = self._aj_tree.selection()
        return self._aj_iid_to_job.get(sel[0]) if sel else None

    def _aj_cancel(self):
        jid = self._aj_selected_job_id()
        if not jid:
            self._aj_status.set("Select a job to cancel.")
            return
        self._get_job_runner().cancel(jid)
        self._aj_status.set(f"Cancelling {jid} after the current step…")

    def _aj_inspect(self):
        jid = self._aj_selected_job_id()
        if not jid:
            return
        j = self._get_job_runner().store.get(jid)
        if j is not None and j.report_path and Path(j.report_path).exists():
            self._preview_data_file(j.report_path)
        elif j is not None:
            self._aj_set_log(self._aj_render_job(j))

    def _aj_clear_finished(self):
        import agent_jobs as _aj
        jr = self._get_job_runner()
        _fin = {_aj.JobStatus.DONE.value, _aj.JobStatus.FAILED.value,
                _aj.JobStatus.CANCELLED.value}
        for j in jr.store.all():
            if j.status in _fin:
                jr.store.delete(j.job_id)
        self._aj_refresh()

    def _aj_show_selected(self):
        jid = self._aj_selected_job_id()
        if not jid:
            return
        j = self._get_job_runner().store.get(jid)
        if j is not None:
            self._aj_set_log(self._aj_render_job(j))

    def _aj_render_job(self, j):
        lines = [f"Job {j.job_id} — {j.status}", f"Goal: {j.goal}", ""]
        for s in j.steps:
            lines.append(
                f"{s.index}. {s.label}"
                + (f" — {s.observation[:180]}" if s.observation else "")
                + (f"  ⚠ {s.error}" if s.error else ""))
        if j.result_summary:
            lines += ["", "Answer:", j.result_summary]
        if j.report_path:
            lines += ["", f"Report: {j.report_path}"]
        return "\n".join(lines)

    def _aj_set_log(self, text):
        w = getattr(self, "_aj_log", None)
        if w is None:
            return
        try:
            w.configure(state="normal")
            w.delete("1.0", "end")
            w.insert("1.0", text)
            w.see("end")
            w.configure(state="disabled")
        except tk.TclError:
            pass

    def _aj_log_append(self, text):
        w = getattr(self, "_aj_log", None)
        if w is None:
            return
        try:
            w.configure(state="normal")
            w.insert("end", text + "\n")
            w.see("end")
            w.configure(state="disabled")
        except tk.TclError:
            pass

    def _aj_refresh(self):
        tree = getattr(self, "_aj_tree", None)
        if tree is None:
            return
        jr = self._get_job_runner()
        tree.delete(*tree.get_children())
        self._aj_tree_iids = {}
        self._aj_iid_to_job = {}
        for j in sorted(jr.store.all(), key=lambda x: -x.updated_ts):
            iid = tree.insert("", "end",
                              values=(j.status, j.goal[:120], len(j.steps)))
            self._aj_tree_iids[j.job_id] = iid
            self._aj_iid_to_job[iid] = j.job_id

    def _aj_update_row(self, job_id):
        tree = getattr(self, "_aj_tree", None)
        if tree is None:
            return
        # Hot per-event path: read only the 3 scalars from a single raw store
        # read, without materialising every job's dataclass graph.
        summ = self._get_job_runner().store.get_summary(job_id)
        if summ is None:
            return
        _status, _goal, _nsteps = summ
        vals = (_status, _goal[:120], _nsteps)
        iid = self._aj_tree_iids.get(job_id)
        if iid and tree.exists(iid):
            tree.item(iid, values=vals)
        else:
            iid = tree.insert("", "end", values=vals)
            self._aj_tree_iids[job_id] = iid
            self._aj_iid_to_job[iid] = job_id

    def _build_agents_tab(self):
        self.tab_agents = ttk.Frame(self.nb)
        self.nb.add(self.tab_agents, text="🤖 Agents")

        top = ttk.Frame(self.tab_agents)
        top.pack(fill="both", expand=True, padx=10, pady=10)

        # Status panel
        status_frame = ttk.LabelFrame(top, text="Agent Status")
        status_frame.pack(fill="x", pady=(0, 10))

        self.agent_status_vars = {}
        for name, available in [
            ("Coder Agent (LangGraph loop)", _CODER_AGENT_OK),
            ("Intern Agent (web research)",        _INTERN_AGENT_OK),
            ("Vault Agent (file tasks, sandboxed)",_VAULT_AGENT_OK),
            ("Sage (tunable domain expert)",          _SAGE_OK),
            ("Vault Scraper (web → vault pipeline)",   _SCRAPER_OK),
            ("Vault RAG (ChromaDB)",               _RAG_OK),
            ("Dream3D-NX Domain Patch",            _D3D_PATCH_OK),
            ("Dream3D simplnx Primer",                _DREAM3D_OK),
        ]:
            row = ttk.Frame(status_frame)
            row.pack(fill="x", padx=6, pady=2)
            color = "#a6e3a1" if available else "#f38ba8"
            icon = "✓" if available else "✗"
            ttk.Label(row, text=f"{icon} {name}", foreground=color).pack(side="left")

        # Controls
        ctrl_frame = ttk.LabelFrame(top, text="Controls")
        ctrl_frame.pack(fill="x", pady=(0, 10))

        self.var_use_coder_agent = tk.BooleanVar(value=_CODER_AGENT_OK)
        self.var_use_intern_agent     = tk.BooleanVar(value=_INTERN_AGENT_OK)
        self.var_use_rag              = tk.BooleanVar(value=_RAG_OK)

        ttk.Checkbutton(ctrl_frame, text="Use Coder coding agent (self-correcting loop)",
                        variable=self.var_use_coder_agent,
                        state="normal" if _CODER_AGENT_OK else "disabled").pack(anchor="w", padx=6, pady=2)
        ttk.Checkbutton(ctrl_frame, text="Use Intern web research agent",
                        variable=self.var_use_intern_agent,
                        state="normal" if _INTERN_AGENT_OK else "disabled").pack(anchor="w", padx=6, pady=2)
        ttk.Checkbutton(ctrl_frame, text="Use RAG context for Writer",
                        variable=self.var_use_rag,
                        state="normal" if _RAG_OK else "disabled").pack(anchor="w", padx=6, pady=2)

        btns = ttk.Frame(ctrl_frame)
        btns.pack(fill="x", padx=6, pady=4)
        ttk.Button(btns, text="Re-index Vault Now",
                   command=lambda: threading.Thread(target=self._init_rag_index, daemon=True).start()
                   ).pack(side="left")

        if _RAG_OK:
            self.rag_count_label = ttk.Label(btns, text="")
            self.rag_count_label.pack(side="left", padx=10)
            self._update_rag_count_label()
        
        # Install instructions
        install_frame = ttk.LabelFrame(top, text="Install missing dependencies")
        install_frame.pack(fill="x", pady=(0, 10))
        
        install_text = (
            # The Coder agent runs on the GGUF backend — no extra
            # install needed. LangGraph + langchain-ollama were the old
            # path; they're gone.
            "pip install crawl4ai && crawl4ai-setup          # Intern web research\n"
            "pip install chromadb sentence-transformers      # Vault RAG\n"
        )
        lbl = tk.Text(install_frame, height=4, wrap="none", font=("Consolas", 10))
        lbl.insert("1.0", install_text)
        lbl.configure(state="disabled", bg="#231a1a", fg="#d32f2f",
                      relief="flat", bd=0)
        lbl.pack(fill="x", padx=6, pady=4)

        # Live agent log
        ttk.Label(top, text="Agent Event Log").pack(anchor="w")
        self.agent_log = self._make_text(top, wrap="word", height=8, state="disabled")
        self.agent_log.tag_configure("phase",   foreground="#d32f2f")
        self.agent_log.tag_configure("result",  foreground="#a6e3a1")
        self.agent_log.tag_configure("fail",    foreground="#f38ba8")
        self.agent_log.pack(fill="x")

        # ── Sage Tuning panel ──────────────────────────────────
        ttk.Separator(self.tab_agents, orient="horizontal").pack(fill="x", padx=8, pady=6)
        if _SAGE_OK and getattr(self, "sage_agent_obj", None) is not None:
            sage_panel = sa.SageTuningPanel(
                self.tab_agents,
                sage_agent=self.sage_agent_obj,
                refresh_cb=None,
            )
            sage_panel.pack(fill="x", padx=6, pady=(0,4))
        elif not _SAGE_OK:
            ttk.Label(self.tab_agents,
                text="🧙 Sage not available — place sage_agent.py next to council_gui_engine.py",
                foreground="#6c7086",
            ).pack(padx=10, pady=4, anchor="w")

        # ── Vault Agent panel ──────────────────────────────────
        ttk.Separator(self.tab_agents, orient="horizontal").pack(fill="x", padx=8, pady=6)
        if _VAULT_AGENT_OK:
            def _get_personality(role: str):
                pm_map = {
                    "writer":     getattr(self, "writer",     None),
                    "coder": getattr(self, "coder", None),
                    "judge":      getattr(self, "judge",      None),
                    "intern":     getattr(self, "intern",     None),
                    "peasant":    getattr(self, "peasant",    None),
                }
                return pm_map.get(role)
            vault_panel = va.VaultAgentPanel(
                self.tab_agents,
                get_personality_fn=_get_personality,
                vault_dir=VAULT_DIR,
            )
            vault_panel.pack(fill="both", expand=True)
        else:
            ttk.Label(self.tab_agents,
                text="🗂  Vault Agent not available — place vault_agent.py next to council_gui_engine.py",
                foreground="#6c7086",
            ).pack(padx=10, pady=10, anchor="w")

    def _update_rag_count_label(self):
        if self.rag and hasattr(self, "rag_count_label"):
            count = self.rag.collection_count()
            self.rag_count_label.configure(text=f"Chunks indexed: {count}")

    def _agent_log_append(self, phase: str, msg: str):
        # The agent log only exists when the Agents tab is built (advanced
        # mode). In the consumer build the widget is absent, so silently
        # skip — the events still go to the transcript via the normal
        # routing in _poll_ui_queue.
        if not hasattr(self, "agent_log"):
            return
        try:
            self.agent_log.configure(state="normal")
            tag = ("result" if "PASS" in msg or "✓" in msg
                   else ("fail" if "FAIL" in msg or "error" in msg.lower()
                         else "phase"))
            self.agent_log.insert("end", f"[{phase}] {msg}\n", tag)
            self.agent_log.see("end")
            self.agent_log.configure(state="disabled")
        except tk.TclError:
            # Widget was destroyed mid-callback
            pass

    # ============================================================
    # Personal Specialists tab
    # ============================================================
    # A list view of the user's specialists with create / edit / delete
    # / test actions. Specialists are pure config — there is no per-
    # specialist data folder. The shared knowledge pool is the vault.

    # ---- Model Finder tab (US-made models that fit this hardware) ----

    def _build_model_finder_tab(self):
        """Show US-made GGUF models that run on this machine's hardware.

        Always lists a curated, offline US-only catalog ranked by fit; an
        optional checkbox augments it with a best-effort Hugging Face search
        when the machine is online. Nothing is downloaded automatically —
        the panel surfaces the repo / file so the user fetches it manually
        (the app stays offline-by-design)."""
        import tkinter as tk
        self.tab_models = ttk.Frame(self.nb)
        self.nb.add(self.tab_models, text="🇺🇸 Models")
        self._mf_results = {}     # tree-iid -> result dict

        top = ttk.Frame(self.tab_models)
        top.pack(fill="x", padx=10, pady=(8, 4))

        # Hardware summary.
        hw = {}
        try:
            import hardware_detect
            hw = hardware_detect.detect()
        except Exception:
            hw = {}
        self._mf_hw = hw
        vram = hw.get("vram_gb")
        ram = hw.get("ram_gb")
        gpu = hw.get("gpu_name") or "no GPU detected"
        hw_text = (f"Your hardware:  GPU: {gpu}   "
                   f"VRAM: {vram or '—'} GB   RAM: {ram or '—'} GB")
        ttk.Label(top, text=hw_text, font=("", 9, "bold"),
                  foreground="#cba6f7").pack(anchor="w")
        ttk.Label(
            top, justify="left", foreground="#9a9a9a",
            text=("Curated US-made GGUF models ranked by fit for this machine. "
                  "Models that fit your VRAM run on GPU; the rest run on CPU.\n"
                  "Nothing downloads automatically — pick one and fetch it from "
                  "the listed repo. US-origin is verified for the catalog; "
                  "online results are a name heuristic.")
        ).pack(anchor="w", pady=(2, 0))

        # Upgrade banner — does this machine have room for a stronger model?
        self._mf_banner = tk.StringVar(value="")
        self._mf_banner_lbl = ttk.Label(
            top, textvariable=self._mf_banner, wraplength=860,
            justify="left", font=("", 9, "bold"))
        self._mf_banner_lbl.pack(anchor="w", pady=(4, 0))

        # One-click: download the suggested upgrade and make it active.
        self._mf_top_upgrade = None       # dict of the best fitting upgrade
        self._mf_upgrade_btn = ttk.Button(
            top, text="⬇ Download & switch (no upgrade available yet)",
            state="disabled", command=self._mf_download_and_switch)
        self._mf_upgrade_btn.pack(anchor="w", pady=(4, 0))

        # Controls.
        ctl = ttk.Frame(self.tab_models)
        ctl.pack(fill="x", padx=10, pady=(2, 6))
        ttk.Label(ctl, text="Task:").pack(side="left")
        self._mf_role = tk.StringVar(value="general")
        ttk.Combobox(ctl, textvariable=self._mf_role, width=10,
                     state="readonly",
                     values=["general", "code", "tiny"]).pack(side="left", padx=6)
        self._mf_online = tk.BooleanVar(value=False)
        ttk.Checkbutton(ctl, text="Also search Hugging Face (needs internet)",
                        variable=self._mf_online).pack(side="left", padx=6)
        ttk.Button(ctl, text="🔎 Find Models",
                   command=self._mf_find).pack(side="left", padx=6)
        ttk.Button(ctl, text="⬆ Suggest upgrades",
                   command=self._mf_check_upgrades).pack(side="left", padx=2)
        self._mf_status = tk.StringVar(value="")
        ttk.Label(ctl, textvariable=self._mf_status,
                  foreground="#a6e3a1").pack(side="left", padx=8)

        # Results table.
        mid = ttk.Frame(self.tab_models)
        mid.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        cols = ("name", "org", "params", "vram", "fits", "ctx", "source")
        self._mf_tree = ttk.Treeview(mid, columns=cols, show="headings",
                                     height=14)
        headings = {
            "name": ("Model", 240), "org": ("Maker", 110),
            "params": ("Params(B)", 80), "vram": ("VRAM≈GB", 80),
            "fits": ("Fits GPU?", 80), "ctx": ("Ctx(K)", 70),
            "source": ("Source", 90),
        }
        for c in cols:
            txt, w = headings[c]
            self._mf_tree.heading(c, text=txt)
            self._mf_tree.column(c, width=w,
                                 anchor="center" if c != "name" else "w")
        vs = ttk.Scrollbar(mid, orient="vertical",
                           command=self._mf_tree.yview)
        self._mf_tree.configure(yscrollcommand=vs.set)
        self._mf_tree.pack(side="left", fill="both", expand=True)
        vs.pack(side="left", fill="y")
        self._mf_tree.tag_configure("fits", foreground="#a6e3a1")
        self._mf_tree.tag_configure("cpu", foreground="#f9e2af")
        self._mf_tree.bind("<<TreeviewSelect>>", self._mf_on_select)

        # Detail / download-info panel.
        det = ttk.LabelFrame(self.tab_models, text="Selected model")
        det.pack(fill="x", padx=10, pady=(0, 8))
        self._mf_detail = tk.Text(det, height=5, wrap="word")
        self._mf_detail.configure(state="disabled")
        self._mf_detail.pack(fill="x", padx=6, pady=4)
        drow = ttk.Frame(det)
        drow.pack(fill="x", padx=6, pady=(0, 4))
        self._mf_dl_btn = ttk.Button(drow, text="⬇ Download & install",
                                     command=self._mf_download)
        self._mf_dl_btn.pack(side="left")
        ttk.Button(drow, text="📋 Copy download info",
                   command=self._mf_copy).pack(side="left", padx=6)
        self._mf_copy_target = ""
        self._mf_selected = None      # dict of the selected model row
        self._mf_dl_cancel = False
        self._mf_downloading = False
        self._mf_auto_activate = False   # set the model active without asking

        # Populate the offline catalog + assess upgrade headroom so the tab
        # is useful before the user clicks anything. DEFER via self.after:
        # these spawn worker threads that call self.after(), and doing that
        # NOW (during tab construction, before mainloop starts) raises
        # "RuntimeError: main thread is not in main loop" — the worker fires
        # before the Tk loop exists. Scheduling from the main thread runs
        # them on the first loop iteration, by which point after() is safe.
        self.after(900, lambda: self._mf_find(initial=True))
        self.after(1100, lambda: self._mf_check_upgrades(initial=True))

    def _mf_current_model_name(self) -> str:
        """Best-effort name of the currently loaded GGUF (for size compare)."""
        import os
        p = os.environ.get("COUNCIL_GGUF_PATH", "").strip()
        if not p:
            try:
                import role_models
                p = role_models.current_loaded_path()
            except Exception:
                p = ""
        try:
            return Path(p).name if p else ""
        except Exception:
            return ""

    def _mf_check_upgrades(self, initial: bool = False):
        """Assess whether the hardware can run a stronger model than the one
        loaded; update the banner and (on explicit click) show the upgrade
        list in the table."""
        import threading as _th
        role = self._mf_role.get()
        current = self._mf_current_model_name()
        if not initial:
            self._mf_status.set("Checking headroom…")

        def _worker():
            try:
                import model_finder
                a = model_finder.assess_upgrade(
                    hardware=self._mf_hw, current_model=current, role=role)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._mf_banner.set(
                    f"Upgrade check unavailable: {exc!r}"))
                return
            self.after(0, lambda: self._mf_apply_assessment(
                a, repopulate=not initial))

        _th.Thread(target=_worker, daemon=True).start()

    def _mf_apply_assessment(self, a: dict, repopulate: bool):
        cur = a.get("current_params_b")
        curtxt = f"~{cur:g}B" if cur else "size unknown"
        name = self._mf_current_model_name() or "none set"
        can = a.get("can_upgrade")
        icon = "✅" if can else "ℹ"
        self._mf_banner.set(
            f"{icon}  Current model: {name} ({curtxt}).  {a.get('reason', '')}")
        self._mf_banner_lbl.configure(
            foreground="#a6e3a1" if can else "#9a9a9a")

        # Enable the one-click upgrade button when a fitting stronger model
        # exists; label it with the specific model so the action is clear.
        ups = a.get("upgrades") or []
        if can and ups:
            self._mf_top_upgrade = ups[0]
            self._mf_upgrade_btn.configure(
                state="normal",
                text=f"⬇ Download & switch to {ups[0].get('name', 'the upgrade')}")
        else:
            self._mf_top_upgrade = None
            self._mf_upgrade_btn.configure(
                state="disabled",
                text="⬇ Download & switch (no upgrade available)")

        if repopulate and a.get("upgrades"):
            self._mf_populate({"catalog": a["upgrades"], "online": []}, False)
            self._mf_status.set(
                f"{len(a['upgrades'])} suggested upgrade(s) that fit your hardware.")
        elif repopulate:
            self._mf_status.set(
                "No stronger model fits comfortably — showing the full catalog.")
            self._mf_find()

    def _mf_find(self, initial: bool = False):
        import threading as _th
        role = self._mf_role.get()
        online = bool(self._mf_online.get()) and not initial
        self._mf_status.set("Searching…" if online else "Loading catalog…")

        def _worker():
            try:
                import model_finder
                res = model_finder.find_models(
                    hardware=self._mf_hw, role=role, prefer_online=online)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._mf_status.set(
                    f"Model finder unavailable: {exc!r}"))
                return
            self.after(0, lambda: self._mf_populate(res, online))

        _th.Thread(target=_worker, daemon=True).start()

    def _mf_populate(self, res: dict, online_requested: bool):
        tree = self._mf_tree
        tree.delete(*tree.get_children())
        self._mf_results.clear()

        def _add(m, src_label):
            params = m.get("params_b")
            vram = m.get("vram_gb_q4", m.get("approx_vram_gb"))
            fits = m.get("fits_vram")
            ctx = m.get("context_k", "")
            iid = tree.insert(
                "", "end",
                values=(m.get("name", m.get("id", "?")),
                        m.get("org", ""),
                        f"{params:g}" if isinstance(params, (int, float)) else "",
                        f"{vram:.1f}" if isinstance(vram, (int, float)) else "",
                        "yes" if fits else "CPU",
                        ctx or "",
                        src_label),
                tags=("fits" if fits else "cpu",))
            self._mf_results[iid] = m

        for m in res.get("catalog", []):
            _add(m, "catalog")
        for m in res.get("online", []):
            _add(m, "HF")

        n_cat = len(res.get("catalog", []))
        n_on = len(res.get("online", []))
        if online_requested and not n_on:
            msg = (f"{n_cat} curated model(s). Hugging Face search returned "
                   "nothing (offline or blocked) — showing the catalog.")
        elif n_on:
            msg = f"{n_cat} curated + {n_on} from Hugging Face."
        else:
            msg = f"{n_cat} curated US-made model(s) ranked by fit."
        self._mf_status.set(msg)

    def _mf_on_select(self, _event=None):
        sel = self._mf_tree.selection()
        if not sel:
            return
        m = self._mf_results.get(sel[0])
        if not m:
            return
        self._mf_selected = m
        repo = m.get("hf_repo")
        hf_file = m.get("hf_file")
        url = m.get("url") or (f"https://huggingface.co/{repo}" if repo else "")
        lines = [f"{m.get('name', m.get('id', '?'))}  —  {m.get('org', '')}"]
        if m.get("blurb"):
            lines.append(m["blurb"])
        if m.get("license"):
            lines.append(f"License: {m['license']}")
        if repo:
            lines.append(f"Hugging Face repo: {repo}")
            if hf_file:
                lines.append(f"File: {hf_file}")
            self._mf_copy_target = f"{repo}" + (f"  ({hf_file})" if hf_file else "")
        elif url:
            lines.append(f"URL: {url}")
            self._mf_copy_target = url
        else:
            self._mf_copy_target = ""
        if m.get("source") == "huggingface" and not m.get("origin_verified", True):
            lines.append("⚠ US-origin is a name heuristic here — verify the "
                         "maker before trusting it.")
        if not m.get("fits_vram"):
            lines.append("Note: doesn't fit your VRAM — will run on CPU "
                         "(slower) via llama-cpp.")
        self._mf_detail.configure(state="normal")
        self._mf_detail.delete("1.0", "end")
        self._mf_detail.insert("1.0", "\n".join(lines))
        self._mf_detail.configure(state="disabled")

    def _mf_copy(self):
        if not self._mf_copy_target:
            self._mf_status.set("Select a model first.")
            return
        try:
            self.clipboard_clear()
            self.clipboard_append(self._mf_copy_target)
            self._mf_status.set(f"Copied: {self._mf_copy_target}")
        except Exception as exc:
            self._mf_status.set(f"Copy failed: {exc!r}")

    def _mf_download_and_switch(self):
        """One-click: download the suggested upgrade and make it active."""
        if not self._mf_top_upgrade:
            self._mf_status.set("No upgrade available to download.")
            return
        self._mf_selected = self._mf_top_upgrade
        self._mf_download(auto_activate=True)

    def _mf_download(self, auto_activate: bool = False):
        """Download the selected model's GGUF straight into the OS-appropriate
        models folder, then offer to make it the active model. The download
        is explicit + user-initiated (it doesn't change offline-by-default).

        ``auto_activate`` skips the post-download "set as active?" prompt —
        used by the upgrade button, where switching is the whole point."""
        from tkinter import messagebox
        import threading as _th

        # A click while a download is running acts as Cancel.
        if self._mf_downloading:
            self._mf_dl_cancel = True
            self._mf_status.set("Cancelling…")
            return

        m = self._mf_selected
        if not m:
            self._mf_status.set("Select a model in the table first.")
            return
        repo, hf_file = m.get("hf_repo"), m.get("hf_file")
        if not (repo and hf_file):
            self._mf_status.set(
                "This is an online search result with no single known file — "
                "use a catalog entry (it has an exact GGUF to fetch), or open "
                "the repo and pick a .gguf manually.")
            return

        try:
            import model_downloader as _md
        except Exception as exc:
            self._mf_status.set(f"Downloader unavailable: {exc!r}")
            return

        dest = _md.default_models_dir()
        size_gb = m.get("size_gb")
        free = _md.disk_free_gb(dest)
        osname = _md.detect_os()

        # Free-space guard (need the file + a little headroom).
        if size_gb and free is not None and free < size_gb * 1.1:
            messagebox.showwarning(
                "Not enough disk space",
                f"{m.get('name', hf_file)} needs ~{size_gb} GB but only "
                f"{free} GB is free on the {osname} models drive.\n\n{dest}")
            return

        size_txt = f"~{size_gb} GB" if size_gb else "unknown size"
        if not messagebox.askyesno(
                "Download model?",
                f"Download {m.get('name', hf_file)} ({size_txt}) from "
                f"Hugging Face?\n\nFrom:  {repo}\nFile:  {hf_file}\n"
                f"To ({osname}):  {dest}\nFree space:  "
                f"{free if free is not None else '?'} GB\n\n"
                "This needs an internet connection."):
            return

        self._mf_downloading = True
        self._mf_dl_cancel = False
        self._mf_auto_activate = bool(auto_activate)
        self._mf_dl_btn.configure(text="■ Cancel download")

        def _prog(done, total):
            if total:
                pct = done * 100 // total
                msg = (f"Downloading {pct}%  "
                       f"({done / 1e6:.0f} / {total / 1e6:.0f} MB)")
            else:
                msg = f"Downloading… {done / 1e6:.0f} MB"
            self.after(0, lambda: self._mf_status.set(msg))

        def _worker():
            try:
                res = _md.download_gguf(
                    repo, hf_file, dest,
                    progress=_prog,
                    should_cancel=lambda: self._mf_dl_cancel,
                    expected_size_gb=size_gb)
                self.after(0, lambda: self._mf_download_done(res, m))
            except _md.DownloadError as de:
                self.after(0, lambda de=de: self._mf_status.set(
                    str(de) + (" (partial saved — resumes next time)"
                               if "cancel" in str(de).lower() else "")))
            except Exception as exc:
                self.after(0, lambda exc=exc: self._mf_status.set(
                    f"Download failed: {exc!r}"))
            finally:
                self.after(0, self._mf_download_reset)

        _th.Thread(target=_worker, daemon=True).start()

    def _mf_download_reset(self):
        self._mf_downloading = False
        self._mf_dl_cancel = False
        self._mf_auto_activate = False
        try:
            self._mf_dl_btn.configure(text="⬇ Download & install")
        except Exception:
            pass

    def _mf_download_done(self, res: dict, m: dict):
        from tkinter import messagebox
        path = res.get("path", "")
        verb = "Already present" if res.get("skipped") else "Downloaded"
        self._mf_status.set(f"{verb}: {path}")
        # The upgrade button already committed to switching — don't ask twice.
        do_activate = self._mf_auto_activate or messagebox.askyesno(
            "Set as active model?",
            f"{verb} {m.get('name', '')}.\n\n{path}\n\n"
            "Make this the Council's active model now? "
            "(Takes effect on the next message.)")
        if do_activate:
            try:
                import onboarding
                onboarding.save_gguf_path(VAULT_DIR, path)
                try:
                    import council_engine as _ce
                    _ce.refresh_backend_config()
                except Exception:
                    pass
                self._mf_status.set(f"✅ Switched — active model is now "
                                    f"{Path(path).name}")
                # Refresh the upgrade banner against the new current model.
                self._mf_check_upgrades(initial=True)
            except Exception as exc:
                self._mf_status.set(f"Saved, but couldn't activate: {exc!r}")

    def _build_specialists_tab(self):
        self.tab_specialists = ttk.Frame(self.nb)
        self.nb.add(self.tab_specialists, text="\U0001f393 Specialists")

        # ── Header ──────────────────────────────────────────────
        hdr = ttk.Frame(self.tab_specialists)
        hdr.pack(fill="x", padx=10, pady=(8, 4))
        ttk.Label(hdr, text="Personal Specialists",
                  font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Label(hdr,
                  text="  Named lenses on your shared data. Auto-summoned when "
                       "your question matches their domain.",
                  foreground="#7a7575", font=("Segoe UI", 9)
                  ).pack(side="left", padx=8)

        # ── Two-pane layout: list (left) | detail (right) ───────
        body = ttk.PanedWindow(self.tab_specialists, orient="horizontal")
        body.pack(fill="both", expand=True, padx=10, pady=(4, 8))

        # LEFT: scrollable list of specialists
        left = ttk.Frame(body, width=260)
        body.add(left, weight=1)

        list_lbl = ttk.Frame(left)
        list_lbl.pack(fill="x", pady=(2, 2))
        ttk.Label(list_lbl, text="Active specialists",
                  font=("Segoe UI", 10, "bold")).pack(side="left")
        ttk.Button(list_lbl, text="➕ New",
                   command=self._spec_new_dialog).pack(side="right")

        list_frame = ttk.Frame(left)
        list_frame.pack(fill="both", expand=True)
        sb = ttk.Scrollbar(list_frame, orient="vertical")
        self._spec_listbox = tk.Listbox(
            list_frame, yscrollcommand=sb.set,
            bg="#231a1a", fg="#d4d4d4",
            selectbackground="#5a3030", relief="flat",
            font=("Segoe UI", 11), activestyle="none",
        )
        sb.configure(command=self._spec_listbox.yview)
        self._spec_listbox.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._spec_listbox.bind("<<ListboxSelect>>",
                                 lambda _e: self._spec_show_selected())

        # RIGHT: detail pane (editable form)
        right = ttk.Frame(body)
        body.add(right, weight=2)

        self._spec_detail_frame = ttk.Frame(right)
        self._spec_detail_frame.pack(fill="both", expand=True, padx=8, pady=4)

        # ── Footer: shared knowledge pool stats ─────────────────
        foot = ttk.LabelFrame(self.tab_specialists,
                              text="Shared knowledge pool")
        foot.pack(fill="x", padx=10, pady=(0, 8))
        self._spec_pool_stats = tk.StringVar(value="Counting…")
        ttk.Label(foot, textvariable=self._spec_pool_stats,
                  foreground="#a98a8a", font=("Segoe UI", 10),
                  ).pack(anchor="w", padx=10, pady=4)
        btnrow = ttk.Frame(foot)
        btnrow.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Button(btnrow, text="\U0001f4c2 Open data_in folder",
                   command=self._open_data_in_folder
                   ).pack(side="left")
        ttk.Button(btnrow, text="\U0001f4e4 Open data_out folder",
                   command=self._open_data_out_folder
                   ).pack(side="left", padx=4)

        # Initial population
        self._spec_refresh_list()
        self._spec_refresh_pool_stats()

    # ---- Specialist list helpers ----

    def _spec_refresh_list(self):
        """Repopulate the listbox from the registry. Preserves selection by id."""
        if not hasattr(self, "_spec_listbox"):
            return
        prev_id = None
        sel = self._spec_listbox.curselection()
        if sel:
            idx = sel[0]
            items = self.specialists.all()
            if 0 <= idx < len(items):
                prev_id = items[idx].id

        self._spec_listbox.delete(0, "end")
        items = self.specialists.all()
        for s in items:
            tag = "  ✓" if s.enabled else "  (off)"
            self._spec_listbox.insert("end", f"{s.icon}  {s.name}{tag}")

        # Restore selection
        if prev_id:
            for i, s in enumerate(items):
                if s.id == prev_id:
                    self._spec_listbox.selection_set(i)
                    self._spec_listbox.activate(i)
                    break
        elif items:
            self._spec_listbox.selection_set(0)
            self._spec_listbox.activate(0)
        self._spec_show_selected()

    def _open_data_in_folder(self):
        """Open the user's read-only input folder in the OS file manager."""
        self._open_in_filemanager(data_index.input_dir(VAULT_DIR))

    def _open_data_out_folder(self):
        """Open the app's output folder in the OS file manager."""
        self._open_in_filemanager(data_index.output_dir(VAULT_DIR))

    def _open_in_filemanager(self, path):
        """Cross-platform 'reveal in file manager'. Best-effort, never raises."""
        try:
            from pathlib import Path as _P
            target = _P(path)
            target.mkdir(parents=True, exist_ok=True)
            import subprocess, sys as _sys
            if _sys.platform == "win32":
                subprocess.Popen(["explorer", str(target)])
            elif _sys.platform == "darwin":
                subprocess.Popen(["open", str(target)])
            else:
                subprocess.Popen(["xdg-open", str(target)])
        except Exception as e:
            print(f"[OpenFolder] Could not open {path}: {e}")

    def _reveal_file(self, path):
        """Reveal a single FILE in the OS file manager (selecting it where the
        platform supports it). Unlike _open_in_filemanager this never creates
        the path — it's for existing source files behind an answer."""
        import subprocess, sys as _sys
        try:
            p = Path(path)
            if not p.exists():
                self._append_transcript(
                    "Council", f"That source file no longer exists: {p.name}",
                    "observation")
                return
            if _sys.platform == "win32":
                # explorer needs the file as ONE argument after /select,
                subprocess.Popen(f'explorer /select,"{p}"')
            elif _sys.platform == "darwin":
                subprocess.Popen(["open", "-R", str(p)])
            else:
                subprocess.Popen(["xdg-open", str(p.parent)])
        except Exception as e:
            print(f"[reveal] could not reveal {path}: {e}")

    def _resolve_source_paths(self, raw_sources):
        """Normalise a mixed list of absolute paths / vault-relative paths /
        bare filenames into existing absolute file paths, de-duplicated and
        order-preserving. Bare names are resolved against the vault input dir
        (searched recursively as a fallback). Returns a list of Path."""
        try:
            in_dir = data_index.input_dir(VAULT_DIR)
        except Exception:
            in_dir = VAULT_DIR
        out, seen = [], set()
        for s in (raw_sources or []):
            if not s or not str(s).strip():
                continue
            cand = None
            try:
                pp = Path(str(s))
                _pat = _compile_name_pattern(pp.name)
                if _pat is not None:
                    # Wildcard bare name (e.g. job_####) — first basename match.
                    for hit in in_dir.rglob("*"):
                        if hit.is_file() and _name_matches_pattern(_pat, hit.name):
                            cand = hit
                            break
                elif pp.is_absolute() and pp.exists():
                    cand = pp
                elif (in_dir / str(s)).exists():
                    cand = in_dir / str(s)
                else:
                    # bare filename — find the first match under the vault
                    name = pp.name
                    for hit in in_dir.rglob(name):
                        if hit.is_file():
                            cand = hit
                            break
            except Exception:
                cand = None
            if cand is not None:
                key = str(cand.resolve()).lower()
                if key not in seen:
                    seen.add(key)
                    out.append(cand)
        return out

    def _vmgr_instant_search(self):
        """Find vault files by NAME or CONTENT (indexed values / column names),
        instantly and with no model. Pops a clickable results list whose items
        preview on double-click."""
        term = (self._vmgr_search_var.get() or "").strip()
        if not term:
            return
        try:
            in_dir = data_index.input_dir(VAULT_DIR)
        except Exception:
            in_dir = VAULT_DIR
        results = []   # (path, reason)
        seen = set()

        def _add(path, reason):
            try:
                key = str(Path(path).resolve()).lower()
            except Exception:
                key = str(path).lower()
            if key not in seen:
                seen.add(key)
                results.append((str(path), reason))

        # 1) filename / path matches (no index needed)
        for full, reason in _search_vault_filenames(in_dir, term):
            _add(full, reason)
        # 2) content matches via the data index (best-effort; refresh lazily)
        try:
            self.data_index.refresh()
            for h in (self.data_index.search_value(term, max_per_file=1) or []):
                name = h.get("file") if isinstance(h, dict) else None
                if name:
                    hit = in_dir / name
                    _add(hit if hit.exists() else name,
                         f"contains “{term}”")
        except Exception:
            pass
        try:
            for prof, exact in (self.data_index.find_files_with_column(term)
                                or []):
                hit = in_dir / prof.name
                _add(hit if hit.exists() else prof.name, f"column “{exact}”")
        except Exception:
            pass

        self._show_file_results(f"Find: {term}", results)

    def _show_file_results(self, title, results):
        """A clickable results list of (path, reason). Double-click or Preview
        opens the model-free data preview; Open folder reveals the file."""
        import tkinter as tk
        from tkinter import ttk
        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("720x460")
        try:
            win.transient(self)
        except Exception:
            pass
        ttk.Label(win, text=f"{len(results)} match(es)", foreground="#888",
                  anchor="w").pack(fill="x", padx=8, pady=(6, 0))
        if not results:
            ttk.Label(win, text="No files match by name or indexed content. "
                      "Try a different term, or add the data in this tab.",
                      foreground="#7a7575", wraplength=680,
                      justify="left").pack(anchor="w", padx=12, pady=12)
            ttk.Button(win, text="Close",
                       command=win.destroy).pack(pady=8)
            return
        lb = tk.Listbox(win, bg="#231a1a", fg="#d4d4d4",
                        selectbackground="#5a3030", relief="flat",
                        font=("Consolas", 10))
        lb.pack(fill="both", expand=True, padx=8, pady=8)
        paths = []
        for path, reason in results:
            paths.append(path)
            lb.insert("end", f"{Path(path).name}   —   {reason}")

        def _selected():
            sel = lb.curselection()
            return paths[sel[0]] if sel else None

        lb.bind("<Double-Button-1>",
                lambda e: (_selected() and self._preview_data_file(_selected())))
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(btns, text="Preview",
                   command=lambda: (_selected()
                                    and self._preview_data_file(_selected()))
                   ).pack(side="left")
        ttk.Button(btns, text="Open containing folder",
                   command=lambda: (_selected()
                                    and self._reveal_file(_selected()))
                   ).pack(side="left", padx=6)
        ttk.Button(btns, text="Close",
                   command=win.destroy).pack(side="right")

    def _preview_data_file(self, path):
        """Quick, MODEL-FREE preview of a data file: schema (column dtypes) +
        the first rows, in a popup. No inference, so it's instant and works
        with no model loaded. Reachable from the answer source chips."""
        import tkinter as tk
        from tkinter import ttk
        p = Path(path)
        if not p.exists():
            self._append_transcript(
                "Council", f"That file no longer exists: {p.name}",
                "observation")
            return
        try:
            schema_txt, rows_txt = _data_preview_text(p)
        except Exception as e:
            schema_txt, rows_txt = (f"Preview failed: {e}", "")
        win = tk.Toplevel(self)
        win.title(f"Preview: {p.name}")
        win.geometry("900x600")
        try:
            win.transient(self)
        except Exception:
            pass
        info = ttk.Label(win, foreground="#888", anchor="w",
                         text=f"{p}  ·  {p.stat().st_size:,} bytes")
        info.pack(fill="x", padx=8, pady=(6, 0))
        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=6, pady=6)
        for label, txt in (("First rows", rows_txt), ("Schema", schema_txt)):
            fr = ttk.Frame(nb)
            nb.add(fr, text=label)
            t = self._make_text(fr, wrap="none")
            t.insert("1.0", txt or "(empty)")
            t.configure(state="disabled")
            t.pack(fill="both", expand=True)
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(btns, text="Open containing folder",
                   command=lambda: self._reveal_file(str(p))).pack(side="left")
        ttk.Button(btns, text="Close",
                   command=win.destroy).pack(side="right")

    def _render_source_chips(self, raw_sources):
        """Render clickable source-file chips under the most recent answer.
        Each chip reveals the file in the OS file manager. No-op when there are
        no resolvable sources, so non-data answers stay clean."""
        paths = self._resolve_source_paths(raw_sources)
        if not paths:
            return
        for widget in (getattr(self, "transcript", None),
                       getattr(self, "dream3d_transcript", None)):
            if widget is None:
                continue
            try:
                widget.configure(state="normal")
                widget.insert("end", "\nSources (click to preview): ", "phase")
                for pp in paths[:12]:
                    self._chip_seq = getattr(self, "_chip_seq", 0) + 1
                    tag = f"srcchip_{self._chip_seq}"
                    widget.tag_configure(tag, foreground="#61afef",
                                         underline=True)
                    widget.tag_bind(
                        tag, "<Button-1>",
                        lambda e, q=str(pp): self._preview_data_file(q))
                    widget.tag_bind(
                        tag, "<Enter>",
                        lambda e, w=widget: w.configure(cursor="hand2"))
                    widget.tag_bind(
                        tag, "<Leave>",
                        lambda e, w=widget: w.configure(cursor=""))
                    widget.insert("end", f" 📄 {pp.name} ", tag)
                if len(paths) > 12:
                    widget.insert("end", f" (+{len(paths) - 12} more)")
                widget.insert("end", "\n")
                widget.see("end")
                widget.configure(state="disabled")
            except tk.TclError:
                pass

    def _run_coach_action(self, action):
        """Execute a one-click fix offered by error coaching (see
        _coach_for_error). Best-effort; never raises into the UI loop."""
        try:
            if action == "engine":
                self._open_engine_settings()
            elif action == "models":
                if hasattr(self, "tab_models"):
                    self.nb.select(self.tab_models)
                else:
                    self._append_transcript(
                        "Council", "The Models tab isn't available in this "
                        "build.", "observation")
            elif action == "cpu":
                # Pin model + embeddings to CPU, clear the GPU-crash sentinel
                # (done inside _apply_engine_settings), then re-ask the last
                # question so the user sees it actually work.
                self._apply_engine_settings(
                    n_ctx=os.environ.get("COUNCIL_GGUF_N_CTX", ""),
                    gpu_layers="0", embed_device="cpu")
                self._append_transcript(
                    "Council", "Switched the model to CPU. Re-running your "
                    "last question…", "observation")
                q = (getattr(self, "_last_sent_query", "") or "").strip()
                if q:
                    self._set_text(self.input, q)
                    self._send()
        except Exception as e:
            print(f"[coach action] {action} failed: {e!r}")

    def _render_error_coach_button(self, coach):
        """Render the clickable one-click-fix action for a coached error as an
        underlined chip in the transcript (the plain guidance is shown
        separately as an observation)."""
        for widget in (getattr(self, "transcript", None),
                       getattr(self, "dream3d_transcript", None)):
            if widget is None:
                continue
            try:
                widget.configure(state="normal")
                widget.insert("end", "\nFix: ", "phase")
                self._chip_seq = getattr(self, "_chip_seq", 0) + 1
                tag = f"coach_{self._chip_seq}"
                widget.tag_configure(tag, foreground="#a6e3a1", underline=True)
                _act = coach.get("action", "")
                widget.tag_bind(
                    tag, "<Button-1>",
                    lambda e, a=_act: self._run_coach_action(a))
                widget.tag_bind(
                    tag, "<Enter>",
                    lambda e, w=widget: w.configure(cursor="hand2"))
                widget.tag_bind(
                    tag, "<Leave>",
                    lambda e, w=widget: w.configure(cursor=""))
                widget.insert("end",
                              f" [ {coach.get('action_label', 'Fix')} ] ", tag)
                widget.insert("end", "\n")
                widget.see("end")
                widget.configure(state="disabled")
            except tk.TclError:
                pass

    def _spec_refresh_pool_stats(self):
        """Show how many files are in the vault knowledge pool."""
        try:
            files = self._vault_data_files()
            n = len(files)
            self._spec_pool_stats.set(
                f"{n} file{'s' if n != 1 else ''} readable.  Inputs come from "
                f"data_in/ and bundled samples (read-only).  Anything the "
                f"app produces lands in data_out/ — originals are never "
                f"overwritten."
            )
        except Exception:
            self._spec_pool_stats.set(
                "Knowledge pool stats unavailable.")

    def _spec_selected(self):
        sel = self._spec_listbox.curselection()
        if not sel:
            return None
        items = self.specialists.all()
        idx = sel[0]
        return items[idx] if 0 <= idx < len(items) else None

    # ---- Detail pane ----

    def _spec_clear_detail(self):
        for w in self._spec_detail_frame.winfo_children():
            w.destroy()

    def _spec_show_selected(self):
        spec = self._spec_selected()
        self._spec_clear_detail()
        if spec is None:
            ttk.Label(self._spec_detail_frame,
                      text="Select a specialist on the left, "
                           "or click ➕ New to create one.",
                      foreground="#7a7575"
                      ).pack(anchor="nw", padx=8, pady=20)
            return
        self._spec_render_detail(spec, editable=True)

    def _spec_render_detail(self, spec, *, editable: bool):
        f = self._spec_detail_frame
        # Header
        head = ttk.Frame(f)
        head.pack(fill="x")
        ttk.Label(head, text=f"{spec.icon}  {spec.name}",
                  font=("Segoe UI", 14, "bold")
                  ).pack(side="left")
        # Enabled toggle
        en_var = tk.BooleanVar(value=spec.enabled)
        def _toggle_enabled():
            spec.enabled = bool(en_var.get())
            self.specialists.add(spec)
            self._spec_refresh_list()
            self._spec_pin_refresh()
        ttk.Checkbutton(head, text="Enabled", variable=en_var,
                        command=_toggle_enabled).pack(side="right")

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=8)

        # Description
        ttk.Label(f, text="Description", font=("Segoe UI", 10, "bold")
                  ).pack(anchor="w")
        desc_var = tk.StringVar(value=spec.description)
        ttk.Entry(f, textvariable=desc_var, width=70
                  ).pack(fill="x", pady=(2, 8))

        # Domain keywords
        ttk.Label(f, text="Domain keywords  (comma-separated)",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Label(f,
                  text="Used to auto-summon this specialist when a question "
                       "mentions one of these terms.",
                  foreground="#7a7575", font=("Segoe UI", 9), wraplength=600,
                  justify="left").pack(anchor="w")
        kw_var = tk.StringVar(value=", ".join(spec.domain_keywords))
        ttk.Entry(f, textvariable=kw_var
                  ).pack(fill="x", pady=(2, 8))

        # System prompt overlay
        ttk.Label(f, text="Lens / system prompt overlay",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Label(f,
                  text="Injected as extra context before the personality "
                       "answers. Tell it how to think, not what to know.",
                  foreground="#7a7575", font=("Segoe UI", 9), wraplength=600,
                  justify="left").pack(anchor="w")
        prompt_box = self._make_text(f, height=8)
        prompt_box.pack(fill="both", expand=True, pady=(2, 8))
        prompt_box.insert("1.0", spec.system_prompt_overlay)

        # Base personality
        bp_row = ttk.Frame(f)
        bp_row.pack(fill="x", pady=(0, 8))
        ttk.Label(bp_row, text="Base personality:",
                  font=("Segoe UI", 10, "bold")).pack(side="left")
        bp_var = tk.StringVar(value=spec.base_personality)
        ttk.Combobox(bp_row, textvariable=bp_var,
                     values=["writer", "sage", "strategist",
                             "intern", "coder", "content"],
                     state="readonly", width=14
                     ).pack(side="left", padx=8)
        ttk.Label(bp_row,
                  text="  (which personality wears this lens)",
                  foreground="#7a7575", font=("Segoe UI", 9)
                  ).pack(side="left")

        # Action buttons
        btns = ttk.Frame(f)
        btns.pack(fill="x", pady=(8, 0), side="bottom")

        def _save():
            spec.description = desc_var.get().strip()
            spec.domain_keywords = [
                k.strip() for k in kw_var.get().split(",") if k.strip()
            ]
            spec.system_prompt_overlay = prompt_box.get("1.0", "end").strip()
            spec.base_personality = bp_var.get().strip() or "writer"
            spec.enabled = bool(en_var.get())
            self.specialists.add(spec)
            self._spec_refresh_list()
            self._spec_pin_refresh()
            self._set_status("● specialist saved", "#a6e3a1")
            self.after(1500, lambda: self._set_status("● idle"))

        def _delete():
            from tkinter import messagebox
            if not messagebox.askyesno(
                "Delete specialist?",
                f"Delete '{spec.name}'?\n\n"
                f"This removes the lens but does not touch any of your data "
                f"in the vault.",
                parent=self):
                return
            self.specialists.remove(spec.id)
            self._spec_refresh_list()
            self._spec_pin_refresh()

        def _test():
            self._spec_test_dialog(spec)

        ttk.Button(btns, text="\U0001f4be Save", command=_save
                   ).pack(side="left")
        ttk.Button(btns, text="\U0001f9ea Test", command=_test
                   ).pack(side="left", padx=6)
        ttk.Button(btns, text="\U0001f5d1 Delete", command=_delete
                   ).pack(side="right")

    # ---- Create / test dialogs ----

    def _spec_new_dialog(self):
        """Quick create dialog — name + description + base; rest editable after."""
        win = tk.Toplevel(self)
        win.title("New specialist")
        win.geometry("440x280")
        win.transient(self); win.grab_set()
        try: branding.apply_window_icon(win)
        except Exception: pass

        ttk.Label(win, text="Create a new Personal Specialist",
                  font=("Segoe UI", 12, "bold")
                  ).pack(anchor="w", padx=16, pady=(14, 8))

        form = ttk.Frame(win); form.pack(fill="x", padx=16)
        ttk.Label(form, text="Name").grid(row=0, column=0, sticky="w", pady=4)
        name_var = tk.StringVar(value="")
        ttk.Entry(form, textvariable=name_var, width=32
                  ).grid(row=0, column=1, sticky="we", pady=4)
        ttk.Label(form, text="Icon (emoji)").grid(row=1, column=0, sticky="w", pady=4)
        icon_var = tk.StringVar(value="🎓")
        ttk.Entry(form, textvariable=icon_var, width=4
                  ).grid(row=1, column=1, sticky="w", pady=4)
        ttk.Label(form, text="Description").grid(row=2, column=0, sticky="w", pady=4)
        desc_var = tk.StringVar()
        ttk.Entry(form, textvariable=desc_var, width=32
                  ).grid(row=2, column=1, sticky="we", pady=4)
        form.columnconfigure(1, weight=1)

        msg_var = tk.StringVar(value="")
        ttk.Label(win, textvariable=msg_var, foreground="#f38ba8"
                  ).pack(anchor="w", padx=16, pady=4)

        def _create():
            name = name_var.get().strip()
            if not name:
                msg_var.set("Name is required.")
                return
            sid = _spec.slugify(name)
            if self.specialists.get(sid):
                msg_var.set(f"A specialist with id '{sid}' already exists.")
                return
            new_spec = _spec.Specialist(
                id=sid,
                name=name,
                icon=icon_var.get().strip() or "🎓",
                description=desc_var.get().strip(),
                domain_keywords=[],
                system_prompt_overlay=(
                    f"You are a {name.lower()}. Apply your domain expertise "
                    f"to whatever question is asked. Be specific and actionable."),
                base_personality="writer",
                enabled=True,
            )
            self.specialists.add(new_spec)
            self._spec_refresh_list()
            # Select the new one
            items = self.specialists.all()
            for i, s in enumerate(items):
                if s.id == sid:
                    self._spec_listbox.selection_clear(0, "end")
                    self._spec_listbox.selection_set(i)
                    self._spec_show_selected()
                    break
            win.destroy()

        bf = ttk.Frame(win); bf.pack(fill="x", padx=16, pady=(8, 14), side="bottom")
        ttk.Button(bf, text="Create", command=_create).pack(side="right")
        ttk.Button(bf, text="Cancel", command=win.destroy
                   ).pack(side="right", padx=6)

    def _spec_test_dialog(self, spec):
        """Pop a small dialog to try the specialist on a sample question."""
        win = tk.Toplevel(self)
        win.title(f"Test {spec.name}")
        win.geometry("620x420")
        win.transient(self)
        try: branding.apply_window_icon(win)
        except Exception: pass

        ttk.Label(win, text=f"Try {spec.icon} {spec.name} on a question",
                  font=("Segoe UI", 11, "bold")
                  ).pack(anchor="w", padx=14, pady=(12, 4))

        ttk.Label(win, text="Question:").pack(anchor="w", padx=14, pady=(4, 0))
        q_box = self._make_text(win, height=3)
        q_box.pack(fill="x", padx=14, pady=4)
        q_box.insert("1.0", "What's a good first question to ask you?")

        ttk.Label(win, text="Response:").pack(anchor="w", padx=14, pady=(8, 0))
        a_box = self._make_text(win, height=10)
        a_box.pack(fill="both", expand=True, padx=14, pady=4)
        a_box.configure(state="disabled")

        def _run():
            q = q_box.get("1.0", "end").strip()
            if not q:
                return
            run_btn.configure(state="disabled", text="Thinking…")
            a_box.configure(state="normal")
            a_box.delete("1.0", "end")
            a_box.insert("1.0", "(running…)")
            a_box.configure(state="disabled")

            def worker():
                try:
                    base = self.personalities.get(spec.base_personality) or self.writer
                    if base is None:
                        text = "(No base personality available — check your config)"
                    else:
                        text = base.respond(q, extra_context=spec.context_block())
                except Exception as e:
                    text = f"Error: {e}"
                self.after(0, lambda t=text: _show(t))

            def _show(text):
                a_box.configure(state="normal")
                a_box.delete("1.0", "end")
                a_box.insert("1.0", text)
                a_box.configure(state="disabled")
                run_btn.configure(state="normal", text="▶ Run")

            threading.Thread(target=worker, daemon=True).start()

        bf = ttk.Frame(win); bf.pack(fill="x", padx=14, pady=(0, 12), side="bottom")
        run_btn = ttk.Button(bf, text="▶ Run", command=_run)
        run_btn.pack(side="right")
        ttk.Button(bf, text="Close", command=win.destroy
                   ).pack(side="right", padx=6)

    # ---- Grapher tab ----

    def _build_grapher_tab(self):
        import tkinter as tk
        self.tab_grapher = ttk.Frame(self.nb)
        self.nb.add(self.tab_grapher, text="\U0001f4ca Grapher")
        self._grapher_live_reload_var = tk.BooleanVar(value=False)  # #8 live-reload

        # State
        self._grapher_dataset      = None
        self._grapher_spec         = None
        self._analyst_personality  = None
        self._grapher_file_paths   = []
        self._last_html_path       = None
        # Charts and exports land in vault/data_out/charts/ — the
        # write-only output area. Keep the legacy graph_output path
        # NO LONGER as the default; data_out is the single sanctioned
        # output destination per the read/write split.
        self._grapher_output_dir   = data_index.output_dir(VAULT_DIR) / "charts"
        self._grapher_plot_history  = []          # [(label, Path), ...]  — #1 history/pin
        self._grapher_transforms    = []          # [{"op":..,"cols":[],"params":{}}]  — #4
        self._grapher_live_job      = None        # after() id for live-reload  — #8
        self._grapher_overlay_ds    = None        # second DataSet for overlay  — #10
        self._grapher_overlay_paths = []          # overlay file path list  — #10

        if _GRAPHER_OK:
            try:
                self._analyst_personality = gp.AnalystPersonality(
                    personality_model=self.writer,
                    event_callback=lambda ph, msg: self.ui_q.put(("grapher_event", ph, msg)),
                )
                gp.patch_routing(ce)
            except Exception as e:
                print(f"[Grapher] Analyst init failed: {e}")

        # ── Outer horizontal split ────────────────────────────────────────────
        main_pane = tk.PanedWindow(self.tab_grapher, orient="horizontal",
                                   bg="#1a1414", sashwidth=6)
        main_pane.pack(fill="both", expand=True, padx=6, pady=6)

        # ─────────────────────────────────────────────────────────────────────
        # LEFT: file browser + scrollable settings
        # ─────────────────────────────────────────────────────────────────────
        left_outer = ttk.Frame(main_pane)
        main_pane.add(left_outer, width=340)

        # ── File browser ──────────────────────────────────────────────────────
        fb = ttk.LabelFrame(left_outer, text="Data File")
        fb.pack(fill="x", padx=4, pady=(4, 4))

        fl1 = ttk.Frame(fb)
        fl1.pack(fill="x", padx=6, pady=(4, 2))
        ttk.Label(fl1, text="Vault:", width=6).pack(side="left")
        self._grapher_file_var = tk.StringVar()
        self._grapher_file_cb  = ttk.Combobox(
            fl1, textvariable=self._grapher_file_var, width=28, state="readonly")
        self._grapher_file_cb.pack(side="left", padx=4, fill="x", expand=True)
        self._grapher_file_cb.bind("<<ComboboxSelected>>", self._grapher_load_file)
        ttk.Button(fl1, text="\u27f3", width=2,
                   command=self._grapher_refresh_files).pack(side="left")

        fl2 = ttk.Frame(fb)
        fl2.pack(fill="x", padx=6, pady=(0, 2))
        ttk.Button(fl2, text="\U0001f4c2 Browse\u2026",
                   command=self._grapher_browse_file).pack(side="left")
        ttk.Button(fl2, text="\U0001f4e6 Sample",
                   command=self._grapher_load_sample).pack(side="left", padx=(4, 0))
        ttk.Label(fl2, text="Sheet:", foreground="#a98a8a").pack(side="left", padx=(10, 2))
        self._grapher_sheet_var = tk.StringVar()
        self._grapher_sheet_cb  = ttk.Combobox(
            fl2, textvariable=self._grapher_sheet_var, width=12, state="readonly")
        self._grapher_sheet_cb.pack(side="left")
        self._grapher_sheet_cb.bind("<<ComboboxSelected>>", self._grapher_reload_sheet)

        self._grapher_status_var = tk.StringVar(value="No file loaded")
        ttk.Label(fb, textvariable=self._grapher_status_var,
                  foreground="#d32f2f", wraplength=300,
                  justify="left").pack(anchor="w", padx=6, pady=(0, 4))

        # ── Live-reload (#8) + overlay file (#10) ─────────────────────────────
        fl3 = ttk.Frame(fb)
        fl3.pack(fill="x", padx=6, pady=(0, 2))
        ttk.Checkbutton(
            fl3, text="\U0001f504 Live reload",
            variable=self._grapher_live_reload_var,
            command=self._grapher_live_toggle,
        ).pack(side="left")
        self._grapher_live_interval_var = tk.IntVar(value=5)
        ttk.Spinbox(fl3, from_=1, to=120, textvariable=self._grapher_live_interval_var,
                    width=4).pack(side="left", padx=2)
        ttk.Label(fl3, text="s", foreground="#a98a8a").pack(side="left")

        fl4 = ttk.Frame(fb)
        fl4.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Label(fl4, text="Overlay:", foreground="#a98a8a", width=7).pack(side="left")
        self._grapher_overlay_var = tk.StringVar(value="(none)")
        self._grapher_overlay_cb  = ttk.Combobox(
            fl4, textvariable=self._grapher_overlay_var, width=20, state="readonly")
        self._grapher_overlay_cb.pack(side="left", padx=4, fill="x", expand=True)
        self._grapher_overlay_cb.bind("<<ComboboxSelected>>", self._grapher_overlay_load)
        ttk.Button(fl4, text="\U0001f4c2", width=2,
                   command=self._grapher_overlay_browse).pack(side="left")

        # ── Scrollable settings ───────────────────────────────────────────────
        ctrl_frame = ttk.Frame(left_outer)
        ctrl_frame.pack(fill="both", expand=True, padx=4)

        ctrl_canvas = tk.Canvas(ctrl_frame, bg="#1a1414", highlightthickness=0)
        ctrl_scroll  = ttk.Scrollbar(ctrl_frame, orient="vertical",
                                      command=ctrl_canvas.yview)
        self._ctrl_inner = ttk.Frame(ctrl_canvas)
        self._ctrl_inner.bind(
            "<Configure>",
            lambda e: ctrl_canvas.configure(scrollregion=ctrl_canvas.bbox("all")))
        ctrl_canvas.create_window((0, 0), window=self._ctrl_inner, anchor="nw")
        ctrl_canvas.configure(yscrollcommand=ctrl_scroll.set)
        ctrl_scroll.pack(side="right", fill="y")
        ctrl_canvas.pack(side="left", fill="both", expand=True)

        def _mw_enter(e):
            ctrl_canvas.bind_all(
                "<MouseWheel>",
                lambda ev: ctrl_canvas.yview_scroll(-1 * (ev.delta // 120), "units"))
        def _mw_leave(e):
            ctrl_canvas.unbind_all("<MouseWheel>")
        ctrl_canvas.bind("<Enter>", _mw_enter)
        ctrl_canvas.bind("<Leave>", _mw_leave)

        ci = self._ctrl_inner

        # ── Plot type ─────────────────────────────────────────────────────────
        ttk.Label(ci, text="Plot Type",
                  font=("", 10, "bold")).pack(anchor="w", pady=(8, 2), padx=6)

        PLOT_GROUPS = {
            "Basic":       ["line", "bar", "scatter", "histogram", "pie", "area"],
            "Statistical": ["box", "violin", "heatmap", "correlation",
                            "distribution", "density_2d", "parallel_coords"],
            "Scientific":  ["fft", "spectrogram", "polar", "contour",
                            "surface_3d", "scatter_3d"],
            "Time Series": ["timeseries", "rolling_mean", "trend", "anomaly"],
            "Dimensional": ["pca"],
        }
        self._plot_type_var = tk.StringVar(value="line")
        for group, types in PLOT_GROUPS.items():
            ttk.Label(ci, text=group, foreground="#d32f2f",
                      font=("", 9, "bold")).pack(anchor="w", padx=8, pady=(4, 0))
            fr = ttk.Frame(ci)
            fr.pack(fill="x", padx=8)
            for col_n, pt in enumerate(types):
                ttk.Radiobutton(
                    fr, text=pt, value=pt,
                    variable=self._plot_type_var,
                    command=self._grapher_update_controls,
                ).grid(row=col_n // 2, column=col_n % 2, sticky="w", padx=2)

        ttk.Separator(ci, orient="horizontal").pack(fill="x", padx=6, pady=6)

        # ── Column selectors ──────────────────────────────────────────────────
        ttk.Label(ci, text="Columns",
                  font=("", 10, "bold")).pack(anchor="w", padx=6, pady=(0, 2))

        def _col_row(parent, label, var_name):
            fr = ttk.Frame(parent)
            fr.pack(fill="x", padx=6, pady=1)
            ttk.Label(fr, text=label, width=10).pack(side="left")
            var = tk.StringVar(value="\u2014")
            cb  = ttk.Combobox(fr, textvariable=var, width=18, state="readonly")
            cb.pack(side="left", padx=4)
            setattr(self, var_name + "_var", var)
            setattr(self, var_name + "_cb",  cb)

        _col_row(ci, "X axis:",    "_gx")
        _col_row(ci, "Y axis:",    "_gy")
        _col_row(ci, "Z axis:",    "_gz")
        _col_row(ci, "Color by:",  "_gc")
        _col_row(ci, "Size by:",   "_gs")
        # ── Faceting (#6) ────────────────────────────────────────────────────
        _col_row(ci, "Facet col:", "_gfacet_col")
        _col_row(ci, "Facet row:", "_gfacet_row")

        ttk.Label(ci, text="Multi-column (Ctrl+click):",
                  foreground="#a98a8a").pack(anchor="w", padx=6, pady=(6, 0))
        self._gcols_lb = tk.Listbox(
            ci, selectmode="multiple", height=5,
            bg="#3a2828", fg="#d4d4d4",
            selectbackground="#5a3030", exportselection=False)
        self._gcols_lb.pack(fill="x", padx=6)

        ttk.Separator(ci, orient="horizontal").pack(fill="x", padx=6, pady=6)

        # ── Options ───────────────────────────────────────────────────────────
        ttk.Label(ci, text="Options",
                  font=("", 10, "bold")).pack(anchor="w", padx=6, pady=(0, 2))

        def _spin_row(parent, label, var_name, from_, to_, default):
            fr = ttk.Frame(parent)
            fr.pack(fill="x", padx=6, pady=1)
            ttk.Label(fr, text=label, width=16).pack(side="left")
            var = tk.IntVar(value=default)
            ttk.Spinbox(fr, from_=from_, to=to_, textvariable=var, width=8).pack(side="left")
            setattr(self, var_name, var)

        def _float_row(parent, label, var_name, default):
            fr = ttk.Frame(parent)
            fr.pack(fill="x", padx=6, pady=1)
            ttk.Label(fr, text=label, width=16).pack(side="left")
            var = tk.StringVar(value=str(default))
            ttk.Entry(fr, textvariable=var, width=10).pack(side="left")
            setattr(self, var_name, var)

        def _combo_row(parent, label, var_name, values, default):
            fr = ttk.Frame(parent)
            fr.pack(fill="x", padx=6, pady=1)
            ttk.Label(fr, text=label, width=16).pack(side="left")
            var = tk.StringVar(value=default)
            ttk.Combobox(fr, textvariable=var, values=values,
                         width=14, state="readonly").pack(side="left")
            setattr(self, var_name, var)

        _spin_row(ci,  "Histogram bins:",  "_gopt_bins",        5,   500,  30)
        _spin_row(ci,  "Rolling window:",  "_gopt_window",      2,  1000,  10)
        _spin_row(ci,  "Trend degree:",    "_gopt_trend",       1,     6,   1)
        _float_row(ci, "Anomaly \u03c3:",  "_gopt_sigma",                    3.0)
        _float_row(ci, "Sample rate Hz:",  "_gopt_samplerate",               1.0)
        _spin_row(ci,  "Marker size:",     "_gopt_marker",      1,    30,   6)
        _spin_row(ci,  "Line width:",      "_gopt_lw",          1,    10,   2)
        _float_row(ci, "Opacity:",         "_gopt_opacity",                  0.85)

        _combo_row(ci, "Colour scheme:", "_gopt_cscheme",
                   ["viridis", "plasma", "inferno", "magma", "Blues", "Reds",
                    "RdBu_r", "RdYlGn", "spectral", "YlOrRd", "Greens"],
                   "viridis")
        _combo_row(ci, "Plotly theme:", "_gopt_theme",
                   ["plotly_dark", "plotly", "ggplot2", "seaborn",
                    "simple_white", "presentation"],
                   "plotly_dark")
        _float_row(ci, "Title:", "_gopt_title", "")

        ttk.Separator(ci, orient="horizontal").pack(fill="x", padx=6, pady=6)

        # ── Transform pipeline (#4) ───────────────────────────────────────────
        ttk.Label(ci, text="Transforms",
                  font=("", 10, "bold")).pack(anchor="w", padx=6, pady=(0, 2))
        self._gtransform_lb = tk.Listbox(
            ci, height=3, bg="#3a2828", fg="#a6e3a1",
            selectbackground="#5a3030", exportselection=False)
        self._gtransform_lb.pack(fill="x", padx=6)
        tf_btn_fr = ttk.Frame(ci)
        tf_btn_fr.pack(fill="x", padx=6, pady=(2, 0))
        for _op in ("normalize", "log", "standardize", "clip"):
            ttk.Button(
                tf_btn_fr, text=_op,
                command=lambda o=_op: self._grapher_transform_add(o),
            ).pack(side="left", padx=1)
        ttk.Button(tf_btn_fr, text="✕ clear",
                   command=self._grapher_transform_clear).pack(side="right")

        ttk.Separator(ci, orient="horizontal").pack(fill="x", padx=6, pady=6)

        # ── Plot / export buttons ─────────────────────────────────────────────
        ttk.Button(ci, text="\u25b6  Plot (Interactive)",
                   command=self._grapher_plot_interactive).pack(fill="x", padx=6, pady=2)

        btn_row2 = ttk.Frame(ci)
        btn_row2.pack(fill="x", padx=6, pady=2)
        ttk.Button(btn_row2, text="\U0001f4cc Pin plot",
                   command=self._grapher_pin_plot).pack(side="left", expand=True, fill="x")
        ttk.Button(btn_row2, text="\U0001f3a4 Narrate",
                   command=self._grapher_narrate).pack(side="left", padx=2)

        btn_exp = ttk.Frame(ci)
        btn_exp.pack(fill="x", padx=6, pady=2)
        ttk.Button(btn_exp, text="\U0001f5bc Export PNG",
                   command=lambda: self._grapher_export("png")).pack(
                       side="left", expand=True, fill="x")
        ttk.Button(btn_exp, text="PDF",
                   command=lambda: self._grapher_export("pdf")).pack(side="left", padx=2)
        ttk.Button(btn_exp, text="SVG",
                   command=lambda: self._grapher_export("svg")).pack(side="left")

        ttk.Separator(ci, orient="horizontal").pack(fill="x", padx=6, pady=6)

        # ── AI Assist ─────────────────────────────────────────────────────────
        ttk.Label(ci, text="AI Assist",
                  font=("", 10, "bold")).pack(anchor="w", padx=6)
        ttk.Label(ci, text="Describe what you want to visualise:",
                  foreground="#a98a8a", font=("", 9)).pack(anchor="w", padx=6)

        self._gai_prompt = tk.Text(
            ci, height=3, wrap="word",
            font=("Consolas", 10), bg="#1a1414", fg="#d4d4d4",
            insertbackground="#d4d4d4")
        self._gai_prompt.pack(fill="x", padx=6, pady=(2, 4))
        self._gai_prompt.insert("1.0", "Show me the distribution of all numeric columns")

        mode_fr = ttk.Frame(ci)
        mode_fr.pack(fill="x", padx=6, pady=(0, 4))
        ttk.Label(mode_fr, text="Mode:", width=6).pack(side="left")
        self._gai_mode_var = tk.StringVar(value="analyst")
        ttk.Radiobutton(mode_fr, text="Analyst (fast)",
                        value="analyst",
                        variable=self._gai_mode_var).pack(side="left", padx=4)
        ttk.Radiobutton(mode_fr, text="Full Council",
                        value="council",
                        variable=self._gai_mode_var).pack(side="left", padx=4)

        ttk.Button(ci, text="\U0001f916 Ask AI to plot",
                   command=self._grapher_ai_plot).pack(fill="x", padx=6, pady=2)
        ttk.Button(ci, text="\U0001f4ca Quick stats summary",
                   command=self._grapher_quick_stats).pack(fill="x", padx=6, pady=2)
        ttk.Button(ci, text="\U0001f4ca Plot council table",
                   command=self._grapher_plot_council_table).pack(fill="x", padx=6, pady=2)

        ttk.Separator(ci, orient="horizontal").pack(fill="x", padx=6, pady=6)

        # ── Presets (#2) ──────────────────────────────────────────────────────
        ttk.Label(ci, text="Presets",
                  font=("", 10, "bold")).pack(anchor="w", padx=6, pady=(0, 2))
        pr_fr = ttk.Frame(ci)
        pr_fr.pack(fill="x", padx=6, pady=(0, 2))
        self._gpreset_var = tk.StringVar(value="")
        self._gpreset_cb  = ttk.Combobox(
            pr_fr, textvariable=self._gpreset_var, width=18, state="readonly")
        self._gpreset_cb.pack(side="left", fill="x", expand=True)
        ttk.Button(pr_fr, text="\U0001f4be Save",
                   command=self._grapher_save_preset).pack(side="left", padx=2)
        ttk.Button(pr_fr, text="\U0001f4c2 Load",
                   command=self._grapher_load_preset).pack(side="left")
        self._grapher_refresh_presets()

        # ─────────────────────────────────────────────────────────────────────
        # RIGHT: plot view + stats/AI output
        # ─────────────────────────────────────────────────────────────────────
        right_pane = tk.PanedWindow(main_pane, orient="vertical",
                                    bg="#1a1414", sashwidth=5)
        main_pane.add(right_pane)

        plot_frame = ttk.Frame(right_pane)
        right_pane.add(plot_frame, height=480)

        self._grapher_plot_label_var = tk.StringVar(value="")
        ttk.Label(plot_frame, textvariable=self._grapher_plot_label_var,
                  foreground="#a6e3a1").pack(anchor="w", padx=4)

        self._grapher_web_frame = None
        if _TKWEB_OK:
            try:
                self._grapher_web_frame = tkinterweb.HtmlFrame(
                    plot_frame, messages_enabled=False)
                self._grapher_web_frame.pack(fill="both", expand=True)
            except Exception:
                self._grapher_web_frame = None

        if self._grapher_web_frame is None:
            ttk.Label(
                plot_frame,
                text="Interactive plots open in your browser.\n"
                     "Install tkinterweb for embedded view:\n"
                     "  pip install tkinterweb",
                foreground="#d32f2f", font=("", 11),
            ).pack(expand=True)
            ttk.Button(
                plot_frame, text="\U0001f310 Open last plot in browser",
                command=self._grapher_open_in_browser,
            ).pack(pady=8)

        stats_frame = ttk.LabelFrame(right_pane, text="Stats & AI Analysis")
        right_pane.add(stats_frame, height=180)
        self._grapher_stats = self._make_text(
            stats_frame, height=8, wrap="word", state="disabled")
        self._grapher_stats.pack(fill="both", expand=True, padx=4, pady=4)

        # ── Plot history (#1) ─────────────────────────────────────────────────
        hist_frame = ttk.LabelFrame(right_pane, text="\U0001f4cc Plot History")
        right_pane.add(hist_frame, height=120)
        hist_top = ttk.Frame(hist_frame)
        hist_top.pack(fill="x", padx=4, pady=2)
        ttk.Button(hist_top, text="\U0001f5c2 Open",
                   command=self._grapher_history_open).pack(side="left")
        ttk.Button(hist_top, text="\u2715 Clear history",
                   command=self._grapher_history_clear).pack(side="left", padx=4)
        self._grapher_history_lb = tk.Listbox(
            hist_frame, height=4, bg="#3a2828", fg="#d4d4d4",
            selectbackground="#5a3030", exportselection=False)
        self._grapher_history_lb.pack(fill="both", expand=True, padx=4, pady=(0, 4))
        self._grapher_history_lb.bind("<Double-Button-1>", lambda e: self._grapher_history_open())

        # ── Correlation drill (#9) ────────────────────────────────────────────
        corr_frame = ttk.LabelFrame(right_pane, text="\U0001f50e Correlation Drill")
        right_pane.add(corr_frame, height=80)
        corr_row = ttk.Frame(corr_frame)
        corr_row.pack(fill="x", padx=4, pady=4)
        ttk.Label(corr_row, text="Col A:", width=7).pack(side="left")
        self._gcorr_a_var = tk.StringVar(value="\u2014")
        self._gcorr_a_cb  = ttk.Combobox(corr_row, textvariable=self._gcorr_a_var,
                                          width=12, state="readonly")
        self._gcorr_a_cb.pack(side="left", padx=2)
        ttk.Label(corr_row, text="Col B:", width=6).pack(side="left", padx=(8, 0))
        self._gcorr_b_var = tk.StringVar(value="\u2014")
        self._gcorr_b_cb  = ttk.Combobox(corr_row, textvariable=self._gcorr_b_var,
                                          width=12, state="readonly")
        self._gcorr_b_cb.pack(side="left", padx=2)
        ttk.Button(corr_row, text="\U0001f50d Drill",
                   command=self._grapher_corr_drill).pack(side="left", padx=6)

        self._grapher_refresh_files()

    # ── Grapher helpers ───────────────────────────────────────────────────────

    def _grapher_refresh_files(self):
        """
        Populate the file dropdown from the read-only data_in/ folder.
        Used to show every CSV anywhere under VAULT_DIR; that scope was
        wide enough to pick up app-internal state files. Now restricted
        to the same input scope as the data index — vault/data_in/ —
        so the dropdown only ever shows files the user dropped in
        themselves.

        We bypass gd.scan_vault_for_data because that helper filters out
        paths whose parents start with "." — which excludes everything
        under ~/.council where the vault actually lives. We do our own
        walk to find loadable files inside data_in.
        """
        if not _GRAPHER_OK:
            return
        in_dir = data_index.input_dir(VAULT_DIR)
        in_dir.mkdir(parents=True, exist_ok=True)
        files = []
        # Hide our own folder explainer — it's not a data file the
        # user wants to load.
        _hidden_names = {"README.txt", "README.md"}
        for p in in_dir.rglob("*"):
            if not p.is_file():
                continue
            if p.name in _hidden_names:
                continue
            try:
                if gd.DataLoader.can_load(p):
                    files.append(p)
            except Exception:
                continue
        files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        labels = []
        for p in files:
            try:
                labels.append(str(p.relative_to(in_dir)))
            except ValueError:
                labels.append(str(p))
        self._grapher_file_cb["values"] = labels
        self._grapher_file_paths        = files
        n = len(files)
        self._grapher_status_var.set(
            f"{n} file{'s' if n != 1 else ''} in data_in/  "
            f"(read-only — drop CSVs there)")

    def _grapher_browse_file(self):
        import tkinter.filedialog as fd
        # Default to vault/data_in/ — that's the user's input folder
        in_dir = data_index.input_dir(VAULT_DIR)
        in_dir.mkdir(parents=True, exist_ok=True)
        path_str = fd.askopenfilename(
            title="Open data file",
            initialdir=str(in_dir),
            filetypes=[
                ("All supported",
                 "*.csv *.tsv *.xlsx *.xls *.json *.npy *.npz *.txt *.log"),
                ("CSV/TSV",  "*.csv *.tsv"),
                ("Excel",    "*.xlsx *.xls"),
                ("JSON",     "*.json"),
                ("NumPy",    "*.npy *.npz"),
                ("Text/Log", "*.txt *.log"),
                ("All files", "*.*"),
            ],
        )
        if not path_str:
            return
        p = Path(path_str)
        label    = str(p)
        existing = list(self._grapher_file_cb["values"])
        if label not in existing:
            existing.append(label)
            self._grapher_file_cb["values"] = existing
            self._grapher_file_paths.append(p)
        self._grapher_file_var.set(label)
        self._grapher_do_load(p)

    def _grapher_load_sample(self):
        """Show a small picker for the bundled sample datasets, then load."""
        from tkinter import simpledialog, messagebox
        sample_dir = Path(__file__).parent / "assets" / "sample_data"
        if not sample_dir.exists():
            messagebox.showinfo(
                "Sample data unavailable",
                "Bundled sample datasets were not found at:\n"
                f"{sample_dir}\n\n"
                "If you're running from source, make sure assets/sample_data/ "
                "exists. In a packaged build this should be present alongside "
                "the executable.",
                parent=self,
            )
            return

        samples = sorted(sample_dir.glob("*.csv"))
        if not samples:
            messagebox.showinfo("Sample data unavailable",
                                "No sample CSV files found.", parent=self)
            return

        # Tiny chooser dialog — radio list of sample names with friendly descriptions.
        descriptions = {
            "purchase_orders.csv": "800 orders across 12 months — try revenue, seasonality, AOV questions",
            "inventory.csv":       "117 SKUs with stock + holding cost — try dead-stock, supplier questions",
            "customers.csv":       "120 customers w/ segments & lifetime spend — try churn, LTV questions",
        }
        win = tk.Toplevel(self)
        win.title("Load sample dataset")
        win.geometry("520x320")
        win.transient(self)
        win.grab_set()
        try: branding.apply_window_icon(win)
        except Exception: pass

        ttk.Label(win, text="Pick a sample dataset to load into the Grapher:",
                  font=("Segoe UI", 11)).pack(anchor="w", padx=14, pady=(14, 8))
        ttk.Label(win,
                  text="These are synthetic CSVs that ship with Data's Inferno. "
                       "Use them to try the tool before loading your real data.",
                  foreground="#7a7575", wraplength=480, justify="left",
                  ).pack(anchor="w", padx=14, pady=(0, 10))

        choice = tk.StringVar(value=samples[0].name)
        for sp in samples:
            row = ttk.Frame(win)
            row.pack(fill="x", padx=14, pady=2)
            ttk.Radiobutton(row, text=sp.name, variable=choice, value=sp.name
                            ).pack(side="left")
            desc = descriptions.get(sp.name, "")
            if desc:
                ttk.Label(row, text=f"— {desc}", foreground="#a98a8a",
                          font=("Segoe UI", 9)
                          ).pack(side="left", padx=8)

        def _do_load():
            picked = sample_dir / choice.get()
            win.destroy()
            label = str(picked)
            existing = list(self._grapher_file_cb["values"])
            if label not in existing:
                existing.append(label)
                self._grapher_file_cb["values"] = existing
                self._grapher_file_paths.append(picked)
            self._grapher_file_var.set(label)
            self._grapher_do_load(picked)

        bf = ttk.Frame(win)
        bf.pack(fill="x", padx=14, pady=(16, 12), side="bottom")
        ttk.Button(bf, text="Load",   command=_do_load).pack(side="right")
        ttk.Button(bf, text="Cancel", command=win.destroy).pack(side="right", padx=6)

    def _grapher_load_file(self, event=None):
        if not _GRAPHER_OK:
            return
        sel = self._grapher_file_var.get()
        if not sel:
            return
        in_dir = data_index.input_dir(VAULT_DIR)
        for p in self._grapher_file_paths:
            # Match against the file relative to data_in/, the bare name,
            # or the full path — supports labels coming from any origin.
            try:
                rel_in = str(p.relative_to(in_dir))
            except ValueError:
                rel_in = ""
            try:
                rel_vault = str(p.relative_to(VAULT_DIR))
            except ValueError:
                rel_vault = ""
            if sel in (rel_in, rel_vault, p.name, str(p)):
                self._grapher_do_load(p)
                return

    def _grapher_do_load(self, path: Path):
        if not _GRAPHER_OK:
            return
        self._grapher_status_var.set(f"Loading {path.name}\u2026")
        ds = gd.DataLoader.load(path)
        self._grapher_dataset = ds

        if ds.load_error:
            self._grapher_status_var.set(f"\u2717 {ds.load_error}")
            return

        rows, cols = ds.shape
        self._grapher_status_var.set(
            f"\u2713 {path.name} \u2014 {rows:,} rows \u00d7 {cols} cols  [{ds.format}]")

        all_cols = ["\u2014"] + ds.all_columns
        num_cols = ["\u2014"] + ds.numeric_columns
        cat_cols = ["\u2014"] + ds.categorical_columns

        for attr, choices in [
            ("_gx", all_cols), ("_gy", num_cols),
            ("_gz", num_cols), ("_gc", all_cols), ("_gs", num_cols),
            ("_gfacet_col", cat_cols), ("_gfacet_row", cat_cols),  # #6 faceting
        ]:
            if hasattr(self, attr + "_cb"):
                getattr(self, attr + "_cb")["values"] = choices
                getattr(self, attr + "_var").set(
                    choices[0])  # default to "—"

        # Set sensible defaults for x/y
        if len(all_cols) > 1:
            self._gx_var.set(all_cols[1])
        if len(num_cols) > 1:
            self._gy_var.set(num_cols[1])

        self._gcols_lb.delete(0, "end")
        for col in ds.all_columns:
            self._gcols_lb.insert("end", col)
        for i, col in enumerate(ds.all_columns):
            if col in ds.numeric_columns:
                self._gcols_lb.selection_set(i)

        # ── Correlation drill column lists (#9)
        if hasattr(self, "_gcorr_a_cb"):
            self._gcorr_a_cb["values"] = num_cols
            self._gcorr_b_cb["values"] = num_cols
            self._gcorr_a_var.set(num_cols[1] if len(num_cols) > 1 else "\u2014")
            self._gcorr_b_var.set(num_cols[2] if len(num_cols) > 2 else "\u2014")

        # ── Overlay file column picker (#10): populate with vault files
        if hasattr(self, "_grapher_overlay_cb"):
            vault_labels = ["(none)"] + list(self._grapher_file_cb["values"])
            self._grapher_overlay_cb["values"] = vault_labels
            self._grapher_overlay_var.set("(none)")

        if ds.metadata.get("sheets"):
            sheets = ds.metadata["sheets"]
            self._grapher_sheet_cb["values"] = sheets
            self._grapher_sheet_var.set(
                ds.metadata.get("active_sheet", sheets[0]))
        else:
            self._grapher_sheet_cb["values"] = []
            self._grapher_sheet_var.set("")

        self._grapher_show_stats(ds.summary())

        # ── Auto-suggest (#3): ask analyst for a quick suggestion after load
        if self._analyst_personality and _GRAPHER_OK:
            def _auto_suggest(ds=ds):
                try:
                    suggestion = self._analyst_personality.analyse(
                        "What is the best single plot to explore this dataset? "
                        "Pick the most revealing visualisation.", ds)
                    hint = (f"\U0001f4a1 Auto-suggestion: {suggestion.spec.plot_type if suggestion.spec else '?'}"
                            f"\n{suggestion.analysis[:300]}" if suggestion.analysis else "")
                    self.ui_q.put(("grapher_autosuggest", hint, suggestion))
                except Exception:
                    pass
            import threading
            threading.Thread(target=_auto_suggest, daemon=True).start()

    def _grapher_reload_sheet(self, event=None):
        if not _GRAPHER_OK or self._grapher_dataset is None:
            return
        if self._grapher_sheet_var.get():
            self._grapher_do_load(self._grapher_dataset.source_path)

    def _grapher_update_controls(self):
        pass  # all controls visible; irrelevant ones silently ignored

    def _grapher_apply_spec_to_controls(self, spec):
        # Sync every UI control to match an AI-returned PlotSpec so the user
        # sees exactly what the AI chose and can tweak before re-plotting.
        if spec is None:
            return
        try:
            if getattr(spec, "plot_type", None):
                self._plot_type_var.set(spec.plot_type)

            ds = self._grapher_dataset

            def _try_set(attr, val):
                if not val:
                    return
                cb  = getattr(self, attr + "_cb",  None)
                var = getattr(self, attr + "_var", None)
                if cb is None or var is None:
                    return
                if val in list(cb["values"]):
                    var.set(val)

            _try_set("_gx", getattr(spec, "x_col",     None))
            _try_set("_gy", getattr(spec, "y_col",     None))
            _try_set("_gz", getattr(spec, "z_col",     None))
            _try_set("_gc", getattr(spec, "color_col", None))
            _try_set("_gs", getattr(spec, "size_col",  None))

            if getattr(spec, "columns", None) and ds:
                self._gcols_lb.selection_clear(0, "end")
                for i, col in enumerate(ds.all_columns):
                    if col in spec.columns:
                        self._gcols_lb.selection_set(i)

            if getattr(spec, "bins",              None): self._gopt_bins.set(int(spec.bins))
            if getattr(spec, "window",            None): self._gopt_window.set(int(spec.window))
            if getattr(spec, "trend_degree",      None): self._gopt_trend.set(int(spec.trend_degree))
            if getattr(spec, "anomaly_threshold", None): self._gopt_sigma.set(str(spec.anomaly_threshold))
            if getattr(spec, "fft_sample_rate",   None): self._gopt_samplerate.set(str(spec.fft_sample_rate))
            if getattr(spec, "marker_size",       None): self._gopt_marker.set(int(spec.marker_size))
            if getattr(spec, "line_width",        None): self._gopt_lw.set(int(spec.line_width))
            if getattr(spec, "opacity",           None): self._gopt_opacity.set(str(spec.opacity))
            if getattr(spec, "color_scheme",      None): self._gopt_cscheme.set(spec.color_scheme)
            if getattr(spec, "theme",             None): self._gopt_theme.set(spec.theme)
            if getattr(spec, "title",             None): self._gopt_title.set(spec.title)

        except Exception as e:
            print(f"[Grapher] apply_spec_to_controls: {e}")

    def _grapher_build_spec(self):
        if not _GRAPHER_OK:
            return None

        def _col(attr):
            v = getattr(self, attr + "_var").get()
            return v if v and v != "\u2014" else None

        ds  = self._grapher_dataset
        sel = self._gcols_lb.curselection()
        multi_cols = ([ds.all_columns[i] for i in sel
                       if i < len(ds.all_columns)] if ds else [])

        try:
            sigma     = float(self._gopt_sigma.get())
            opacity   = float(self._gopt_opacity.get())
            sr        = float(self._gopt_samplerate.get())
            title_str = self._gopt_title.get()
        except Exception:
            sigma, opacity, sr, title_str = 3.0, 0.85, 1.0, ""

        return ge.PlotSpec(
            plot_type         = self._plot_type_var.get(),
            x_col             = _col("_gx"),
            y_col             = _col("_gy"),
            z_col             = _col("_gz"),
            color_col         = _col("_gc"),
            size_col          = _col("_gs"),
            columns           = multi_cols,
            title             = title_str,
            color_scheme      = self._gopt_cscheme.get(),
            theme             = self._gopt_theme.get(),
            bins              = self._gopt_bins.get(),
            window            = self._gopt_window.get(),
            trend_degree      = self._gopt_trend.get(),
            anomaly_threshold = sigma,
            fft_sample_rate   = sr,
            marker_size       = self._gopt_marker.get(),
            line_width        = self._gopt_lw.get(),
            opacity           = opacity,
            renderer          = "plotly",
            facet_col         = _col("_gfacet_col"),   # #6 faceting
            facet_row         = _col("_gfacet_row"),   # #6 faceting
        )

    def _grapher_plot_interactive(self):
        if not _GRAPHER_OK:
            return
        if self._grapher_dataset is None:
            self._grapher_show_stats("No file loaded. Select a file first.")
            return
        spec = self._grapher_build_spec()
        if spec:
            self._grapher_render_plotly(spec, self._grapher_dataset)

    def _grapher_render_plotly(self, spec, ds):
        if not _GRAPHER_OK:
            return

        # ── Apply transforms (#4) ─────────────────────────────────────────────
        working_ds = ds
        if self._grapher_transforms and ds.df is not None:
            import copy
            working_ds = copy.copy(ds)
            working_ds.df, tf_log = ge.apply_transforms(ds.df, self._grapher_transforms)
            if hasattr(self, "_gtransform_lb"):
                self._gtransform_lb.delete(0, "end")
                for msg in tf_log:
                    self._gtransform_lb.insert("end", msg)

        # ── Multi-file overlay (#10) ──────────────────────────────────────────
        if (self._grapher_overlay_ds is not None
                and self._grapher_overlay_ds.df is not None
                and spec.y_col and spec.plot_type in ("line", "timeseries", "scatter", "area")):
            try:
                import copy
                import plotly.express as px
                import plotly.graph_objects as go
                ods = self._grapher_overlay_ds
                renderer_base = ge.PlotlyRenderer()
                html = renderer_base._overlay_render(spec, working_ds, ods)
            except Exception as e:
                renderer_base = ge.PlotlyRenderer()
                html = renderer_base.render(spec, working_ds)
        else:
            renderer_base = ge.PlotlyRenderer()
            html = renderer_base.render(spec, working_ds)

        label = f"{spec.plot_type.replace('_', ' ').title()} \u2014 {ds.name}"
        self._grapher_plot_label_var.set(label)
        self._grapher_spec = spec

        # ── Save to history (#1) ──────────────────────────────────────────────
        self._grapher_output_dir.mkdir(parents=True, exist_ok=True)
        ts    = datetime.now().strftime("%H%M%S")
        hpath = self._grapher_output_dir / f"plot_{ts}_{spec.plot_type}.html"
        hpath.write_text(html, encoding="utf-8")
        self._last_html_path = hpath
        entry_label = f"{ts} — {spec.plot_type} — {ds.name}"
        self._grapher_plot_history.append((entry_label, hpath))
        if hasattr(self, "_grapher_history_lb"):
            self._grapher_history_lb.insert("end", entry_label)
            self._grapher_history_lb.see("end")

        if self._grapher_web_frame is not None:
            self._grapher_web_frame.load_html(html)
        else:
            import webbrowser
            webbrowser.open(hpath.as_uri())
            self._grapher_show_stats(
                f"Plot opened in browser: {hpath}\n\n"
                + ge.DataAnalyser.describe(ds)[:800])

    def _grapher_open_in_browser(self):
        if self._last_html_path and self._last_html_path.exists():
            import webbrowser
            webbrowser.open(self._last_html_path.as_uri())

    def _grapher_export(self, fmt: str):
        if not _GRAPHER_OK:
            return
        if self._grapher_dataset is None:
            self._grapher_show_stats("No file loaded.")
            return
        spec = self._grapher_build_spec()
        spec.renderer = "matplotlib"
        import tkinter.filedialog as fd
        ds           = self._grapher_dataset
        default_name = f"{ds.name}_{spec.plot_type}.{fmt}"
        # Default destination: data_out/charts/ (the app's write area).
        out_dir = data_index.output_dir(VAULT_DIR) / "charts"
        out_dir.mkdir(parents=True, exist_ok=True)
        path_str     = fd.asksaveasfilename(
            title=f"Export {fmt.upper()}",
            defaultextension=f".{fmt}",
            initialdir=str(out_dir),
            initialfile=default_name,
            filetypes=[(fmt.upper(), f"*.{fmt}"), ("All files", "*.*")],
        )
        if not path_str:
            return
        mpl_r = ge.MatplotlibRenderer()
        fig   = mpl_r.render(spec, ds)
        if fig:
            saved = mpl_r.save(fig, Path(path_str))
            self._grapher_show_stats(f"\u2713 Exported: {saved}")
        else:
            self._grapher_show_stats(
                "\u2717 Export failed \u2014 matplotlib could not render this plot type.")

    def _grapher_ai_plot(self):
        if not _GRAPHER_OK:
            self._grapher_show_stats("Grapher module not available.")
            return
        if self._grapher_dataset is None:
            self._grapher_show_stats("Load a data file first.")
            return
        prompt = self._gai_prompt.get("1.0", "end").strip()
        if not prompt:
            return

        mode = getattr(self, "_gai_mode_var", None)
        mode = mode.get() if mode else "analyst"
        ds   = self._grapher_dataset

        if mode == "council":
            valid_types = (
                "line, bar, scatter, histogram, pie, area, box, violin, heatmap, "
                "correlation, distribution, density_2d, parallel_coords, fft, "
                "spectrogram, polar, contour, surface_3d, scatter_3d, "
                "timeseries, rolling_mean, trend, anomaly, pca"
            )
            council_prompt = "\n".join([
                "You are helping choose and configure a data visualisation.",
                "",
                f"Dataset: {ds.name}  ({ds.shape[0]:,} rows x {ds.shape[1]} columns)",
                f"Numeric columns: {', '.join(ds.numeric_columns)}",
                f"Categorical columns: {', '.join(ds.categorical_columns)}",
                "",
                f"User request: {prompt}",
                "",
                "Respond with:",
                "1. A short explanation of what this plot will show and why it suits the data.",
                "2. A JSON block (fenced with ```json) with these PlotSpec keys:",
                "   plot_type, x_col, y_col, columns (list), title, color_scheme, theme,",
                "   bins, window, trend_degree, anomaly_threshold, opacity",
                f"Valid plot_type values: {valid_types}",
                "Only use column names that exist in the dataset above.",
                "Set unused keys to null.",
            ])
            self._grapher_show_stats("Sending to council\u2026")

            def council_worker():
                import re, json as _json
                try:
                    response = self.writer.respond(
                        council_prompt,
                        extra_context="DATASET SUMMARY:\n" + ds.summary(),
                    )
                    m = re.search(r"```json\s*(.*?)\s*```", response, re.DOTALL)
                    if not m:
                        m = re.search(r'\{[^{}]*"plot_type"[^{}]*\}', response, re.DOTALL)
                    spec      = None
                    parse_err = ""
                    if m:
                        raw = m.group(1) if m.lastindex and m.lastindex >= 1 else m.group(0)
                        try:
                            d = _json.loads(raw)
                            if hasattr(ge.PlotSpec, "__dataclass_fields__"):
                                valid_f = set(ge.PlotSpec.__dataclass_fields__)
                                d = {k: v for k, v in d.items() if k in valid_f}
                            spec = ge.PlotSpec(**d)
                        except Exception as pe:
                            parse_err = str(pe)
                    else:
                        parse_err = "No JSON block found in council response"

                    class _R:
                        pass
                    r             = _R()
                    r.spec        = spec
                    r.analysis    = response
                    r.parse_error = parse_err
                    r.raw_json    = m.group(0) if m else ""
                    self.ui_q.put(("grapher_ai_result", r))
                except Exception as e:
                    self.ui_q.put(("grapher_event", "error", str(e)))

            import threading
            threading.Thread(target=council_worker, daemon=True).start()

        else:
            if self._analyst_personality is None:
                self._grapher_show_stats("Analyst not initialised.")
                return
            self._grapher_show_stats("Asking analyst AI\u2026")

            def analyst_worker():
                try:
                    result = self._analyst_personality.analyse(prompt, ds)
                    self.ui_q.put(("grapher_ai_result", result))
                except Exception as e:
                    self.ui_q.put(("grapher_event", "error", str(e)))

            import threading
            threading.Thread(target=analyst_worker, daemon=True).start()

    def _grapher_quick_stats(self):
        if not _GRAPHER_OK or self._grapher_dataset is None:
            self._grapher_show_stats("Load a file first.")
            return
        ds = self._grapher_dataset

        def worker():
            try:
                if self._analyst_personality:
                    text = self._analyst_personality.quick_analysis(ds)
                else:
                    text = ge.DataAnalyser.describe(ds)
                self.ui_q.put(("grapher_stats", text))
            except Exception as e:
                self.ui_q.put(("grapher_stats", f"Error: {e}"))

        import threading
        threading.Thread(target=worker, daemon=True).start()

    def _grapher_show_stats(self, text: str):
        self._grapher_stats.configure(state="normal")
        self._grapher_stats.delete("1.0", "end")
        self._grapher_stats.insert("1.0", text)
        self._grapher_stats.configure(state="disabled")

    # ── Plot history helpers (#1) ─────────────────────────────────────────────

    def _grapher_pin_plot(self):
        """Save the current plot to a timestamped file and add to history."""
        if self._last_html_path and self._last_html_path.exists():
            self._grapher_show_stats(f"\U0001f4cc Pinned: {self._last_html_path.name}")
        else:
            self._grapher_show_stats("No plot to pin — plot something first.")

    def _grapher_history_open(self):
        """Open the selected history entry in the browser."""
        if not hasattr(self, "_grapher_history_lb"):
            return
        sel = self._grapher_history_lb.curselection()
        if not sel:
            # open last
            if self._grapher_plot_history:
                _, path = self._grapher_plot_history[-1]
                import webbrowser; webbrowser.open(path.as_uri())
            return
        idx = sel[0]
        if idx < len(self._grapher_plot_history):
            _, path = self._grapher_plot_history[idx]
            if path.exists():
                import webbrowser; webbrowser.open(path.as_uri())
            else:
                self._grapher_show_stats(f"\u2717 File no longer exists: {path}")

    def _grapher_history_clear(self):
        """Clear the plot history list."""
        self._grapher_plot_history.clear()
        if hasattr(self, "_grapher_history_lb"):
            self._grapher_history_lb.delete(0, "end")

    # ── Transform pipeline helpers (#4) ──────────────────────────────────────

    def _grapher_transform_add(self, op: str):
        """Add a named transform step to the pipeline."""
        step = {"op": op, "cols": [], "params": {}}
        if op == "clip":
            step["op"]    = "clip_outliers"
            step["params"] = {"sigma": 3.0}
        self._grapher_transforms.append(step)
        if hasattr(self, "_gtransform_lb"):
            self._gtransform_lb.insert("end", op)

    def _grapher_transform_clear(self):
        """Remove all transform steps."""
        self._grapher_transforms.clear()
        if hasattr(self, "_gtransform_lb"):
            self._gtransform_lb.delete(0, "end")

    # ── Preset helpers (#2) ───────────────────────────────────────────────────

    def _grapher_preset_file(self) -> Path:
        return VAULT_DIR / "graph_presets.json"

    def _grapher_refresh_presets(self):
        """Load preset names from file and populate the combobox."""
        if not hasattr(self, "_gpreset_cb"):
            return
        try:
            pf = self._grapher_preset_file()
            if pf.exists():
                data = _json.loads(pf.read_text(encoding="utf-8"))
                self._gpreset_cb["values"] = list(data.keys())
        except Exception:
            pass

    def _grapher_save_preset(self):
        """Save current spec settings as a named preset."""
        name = simpledialog.askstring("Save Preset", "Preset name:", parent=self.root)
        if not name:
            return
        try:
            spec = self._grapher_build_spec()
            if spec is None:
                return
            preset_dict = spec.to_dict()
            pf = self._grapher_preset_file()
            data = {}
            if pf.exists():
                try:
                    data = _json.loads(pf.read_text(encoding="utf-8"))
                except Exception:
                    pass
            data[name] = preset_dict
            pf.write_text(_json.dumps(data, indent=2), encoding="utf-8")
            self._grapher_refresh_presets()
            self._gpreset_var.set(name)
            self._grapher_show_stats(f"\u2713 Preset '{name}' saved.")
        except Exception as e:
            self._grapher_show_stats(f"\u2717 Save preset error: {e}")

    def _grapher_load_preset(self):
        """Load selected preset and apply it to controls."""
        name = self._gpreset_var.get()
        if not name:
            return
        try:
            pf = self._grapher_preset_file()
            data = _json.loads(pf.read_text(encoding="utf-8"))
            if name not in data:
                self._grapher_show_stats(f"\u2717 Preset '{name}' not found.")
                return
            spec = ge.PlotSpec.from_dict(data[name])
            self._grapher_apply_spec_to_controls(spec)
            self._grapher_show_stats(f"\u2713 Preset '{name}' loaded.")
        except Exception as e:
            self._grapher_show_stats(f"\u2717 Load preset error: {e}")

    # ── Narration helper (#5) ─────────────────────────────────────────────────

    def _grapher_narrate(self):
        """Ask the Writer/Director to narrate the current plot in plain English."""
        if not _GRAPHER_OK or self._grapher_dataset is None:
            self._grapher_show_stats("No data loaded.")
            return
        spec = self._grapher_spec or self._grapher_build_spec()
        if spec is None:
            return
        ds   = self._grapher_dataset
        self._grapher_show_stats("Generating narration\u2026")

        def _worker():
            try:
                prompt = (
                    f"Narrate the following data plot in clear plain English for a non-technical audience.\n"
                    f"Dataset: {ds.name}  ({ds.shape[0]:,} rows \u00d7 {ds.shape[1]} cols)\n"
                    f"Plot type: {spec.plot_type}\n"
                    f"X axis: {spec.x_col or 'index'}  |  Y axis: {spec.y_col or '(multi)'}\n"
                    f"Color: {spec.color_col or 'none'}  |  Facet: {spec.facet_col or 'none'}\n"
                    f"Numeric columns: {', '.join(ds.numeric_columns[:6])}\n\n"
                    f"Dataset summary:\n{ds.summary()[:600]}\n\n"
                    f"Give a 3–5 sentence narration of what this chart shows, any key insights, "
                    f"and one follow-up question worth investigating."
                )
                text = self.writer.respond(prompt)
                self.ui_q.put(("grapher_stats", f"\U0001f3a4 Narration:\n\n{text}"))
            except Exception as e:
                self.ui_q.put(("grapher_stats", f"\u2717 Narration error: {e}"))

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    # ── Live reload helpers (#8) ──────────────────────────────────────────────

    def _grapher_live_toggle(self):
        """Toggle live-reload mode on/off."""
        if self._grapher_live_reload_var.get():
            self._grapher_live_schedule()
        else:
            if self._grapher_live_job is not None:
                self.root.after_cancel(self._grapher_live_job)
                self._grapher_live_job = None

    def _grapher_live_schedule(self):
        """Schedule the next live-reload tick."""
        if not self._grapher_live_reload_var.get():
            return
        try:
            interval_ms = int(self._grapher_live_interval_var.get()) * 1000
        except Exception:
            interval_ms = 5000
        self._grapher_live_job = self.root.after(interval_ms, self._grapher_live_tick)

    def _grapher_live_tick(self):
        """Reload the current file and re-render if changed."""
        self._grapher_live_job = None
        if not self._grapher_live_reload_var.get():
            return
        ds = self._grapher_dataset
        if ds and ds.source_path.exists():
            new_ds = gd.DataLoader.load(ds.source_path)
            if not new_ds.load_error and new_ds.shape != ds.shape:
                self._grapher_dataset = new_ds
                spec = self._grapher_spec or self._grapher_build_spec()
                if spec:
                    self._grapher_render_plotly(spec, new_ds)
        self._grapher_live_schedule()

    # ── Overlay helpers (#10) ─────────────────────────────────────────────────

    def _grapher_overlay_load(self, event=None):
        """Load the selected overlay file."""
        if not _GRAPHER_OK:
            return
        label = self._grapher_overlay_var.get()
        if not label or label == "(none)":
            self._grapher_overlay_ds = None
            return
        # Find path
        for p in self._grapher_file_paths:
            try:
                rel = str(p.relative_to(VAULT_DIR))
            except ValueError:
                rel = str(p)
            if rel == label or str(p) == label:
                try:
                    self._grapher_overlay_ds = gd.DataLoader.load(p)
                except Exception as e:
                    self._grapher_show_stats(f"✗ Could not load overlay {p.name}: {e}")
                    return
                self._grapher_show_stats(
                    f"\U0001f4ce Overlay loaded: {p.name}  "
                    f"({self._grapher_overlay_ds.shape[0]:,} rows)")
                return

    def _grapher_overlay_browse(self):
        """Browse for an overlay file."""
        import tkinter.filedialog as fd
        in_dir = data_index.input_dir(VAULT_DIR)
        in_dir.mkdir(parents=True, exist_ok=True)
        path_str = fd.askopenfilename(
            title="Open overlay file",
            initialdir=str(in_dir),
            filetypes=[("All supported",
                        "*.csv *.tsv *.xlsx *.xls *.json *.npy *.npz *.txt"),
                       ("All files", "*.*")],
        )
        if not path_str:
            return
        p = Path(path_str)
        try:
            self._grapher_overlay_ds = gd.DataLoader.load(p)
        except Exception as e:
            self._grapher_show_stats(f"✗ Could not load overlay {p.name}: {e}")
            return
        label = str(p)
        existing = list(self._grapher_overlay_cb["values"])
        if label not in existing:
            existing.append(label)
            self._grapher_overlay_cb["values"] = existing
            self._grapher_overlay_paths.append(p)
        self._grapher_overlay_var.set(label)
        self._grapher_show_stats(
            f"\U0001f4ce Overlay loaded: {p.name}  "
            f"({self._grapher_overlay_ds.shape[0]:,} rows)")

    # ── Correlation drill helpers (#9) ────────────────────────────────────────

    def _grapher_corr_drill(self):
        """Run correlation drill: scatter + Pearson r for two chosen columns."""
        if not _GRAPHER_OK or self._grapher_dataset is None:
            self._grapher_show_stats("Load a file first.")
            return
        col_a = self._gcorr_a_var.get()
        col_b = self._gcorr_b_var.get()
        ds    = self._grapher_dataset
        if col_a == "\u2014" or col_b == "\u2014" or col_a == col_b:
            self._grapher_show_stats("Select two different numeric columns for drill.")
            return
        if col_a not in ds.df.columns or col_b not in ds.df.columns:
            self._grapher_show_stats(f"Columns not found: {col_a}, {col_b}")
            return
        # Build scatter spec for these two columns
        spec = ge.PlotSpec(
            plot_type  = "scatter",
            x_col      = col_a,
            y_col      = col_b,
            title      = f"Correlation: {col_a} vs {col_b}",
            color_scheme = self._gopt_cscheme.get() if hasattr(self, "_gopt_cscheme") else "viridis",
            theme        = self._gopt_theme.get()   if hasattr(self, "_gopt_theme")   else "plotly_dark",
            trend_degree = 1,
            renderer   = "plotly",
        )
        self._grapher_render_plotly(spec, ds)
        # Compute Pearson r
        try:
            import pandas as pd
            clean = ds.df[[col_a, col_b]].dropna()
            r = clean[col_a].corr(clean[col_b])
            n = len(clean)
            self._grapher_show_stats(
                f"\U0001f50e Correlation Drill: {col_a} vs {col_b}\n"
                f"Pearson r = {r:.4f}  |  n = {n:,}\n"
                f"{'Strong' if abs(r) > 0.7 else 'Moderate' if abs(r) > 0.4 else 'Weak'} "
                f"{'positive' if r > 0 else 'negative'} correlation\n\n"
                + ge.DataAnalyser.describe(ds)[:400]
            )
        except Exception as e:
            self._grapher_show_stats(f"Correlation error: {e}")

    # ── Plot council table (#7) ───────────────────────────────────────────────

    def _grapher_plot_council_table(self):
        """
        Detect a markdown table in the last council answer and plot it.
        Parses the table into a DataFrame, then runs an AI auto-suggest.
        """
        if not _GRAPHER_OK:
            return
        text = getattr(self, "_last_final_text", "").strip()
        if not text:
            self._grapher_show_stats("No council answer yet — ask the council something first.")
            return
        # Look for markdown table
        import re as _re_local
        table_match = _re_local.search(
            r"(\|.+\|\s*\n\|[-| :]+\|\s*\n(?:\|.+\|\s*\n?)+)", text)
        if not table_match:
            self._grapher_show_stats(
                "\u2717 No markdown table found in last council answer.\n\n"
                "Ask the council to format data as a Markdown table first.")
            return
        raw_table = table_match.group(1)
        try:
            import pandas as pd
            import io as _io
            lines = [l.strip() for l in raw_table.strip().splitlines() if l.strip()]
            # Remove separator row
            lines = [l for l in lines if not _re_local.match(r"^\|[-| :]+\|$", l)]
            # Parse
            rows = []
            for line in lines:
                cells = [c.strip() for c in line.strip("|").split("|")]
                rows.append(cells)
            if len(rows) < 2:
                raise ValueError("Not enough rows")
            headers = rows[0]
            data    = rows[1:]
            df = pd.DataFrame(data, columns=headers)
            # Try to coerce to numeric
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(df[col])
        except Exception as e:
            self._grapher_show_stats(f"\u2717 Table parse error: {e}")
            return

        from pathlib import Path as _Path
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False,
                                         mode="w", encoding="utf-8") as f:
            df.to_csv(f, index=False)
            tmp_path = _Path(f.name)

        ds = gd.DataLoader.load(tmp_path)
        ds.name = "council_table"
        if ds.load_error:
            self._grapher_show_stats(f"\u2717 Could not load table: {ds.load_error}")
            return
        self._grapher_dataset = ds
        self._grapher_show_stats(
            f"\u2713 Council table loaded: {ds.shape[0]} rows \u00d7 {ds.shape[1]} cols\n"
            + ds.summary())
        # Auto-suggest a plot
        if self._analyst_personality:
            def _worker(ds=ds):
                try:
                    result = self._analyst_personality.analyse(
                        "What is the best plot for this data from a council response?", ds)
                    self.ui_q.put(("grapher_ai_result", result))
                except Exception:
                    pass
            import threading
            threading.Thread(target=_worker, daemon=True).start()

        # ---- Vault Manager tab ----

    def _build_vault_manager_tab(self):
        import tkinter as tk
        self.tab_vmgr = ttk.Frame(self.nb)
        self.nb.add(self.tab_vmgr, text="🗄 Vault")

        # ── Top bar ───────────────────────────────────────────
        top = ttk.Frame(self.tab_vmgr)
        top.pack(fill="x", padx=10, pady=(8, 4))

        ttk.Label(top, text="Vault:", font=("", 9, "bold")).pack(side="left")
        ttk.Label(top, text=str(VAULT_DIR),
                  foreground="#d32f2f").pack(side="left", padx=6)
        ttk.Button(top, text="📂 Open Folder",
                   command=self._vmgr_open_folder).pack(side="right")
        ttk.Button(top, text="⟳ Refresh",
                   command=self._vmgr_refresh_tree).pack(side="right", padx=4)
        ttk.Button(top, text="RAG Misses",
                   command=self._show_rag_misses).pack(side="right", padx=4)

        # ── Instant search bar (no model) ─────────────────────
        # Find vault files by NAME or by CONTENT (indexed values / columns),
        # instantly and with no inference. Results are clickable → preview.
        srch = ttk.Frame(self.tab_vmgr)
        srch.pack(fill="x", padx=10, pady=(0, 4))
        ttk.Label(srch, text="🔍 Find files:",
                  font=("", 9, "bold")).pack(side="left")
        self._vmgr_search_var = tk.StringVar()
        _se = ttk.Entry(srch, textvariable=self._vmgr_search_var, width=36)
        _se.pack(side="left", padx=6)
        _se.bind("<Return>", lambda e: self._vmgr_instant_search())
        ttk.Button(srch, text="Search",
                   command=self._vmgr_instant_search).pack(side="left")
        ttk.Label(srch, text="by file name or content — instant, no model",
                  foreground="#7a7575").pack(side="left", padx=8)

        # ── Main pane ─────────────────────────────────────────
        main = tk.PanedWindow(self.tab_vmgr, orient="horizontal",
                              bg="#1a1414", sashwidth=6)
        main.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        # ── LEFT: vault tree + controls ───────────────────────
        left = ttk.Frame(main)
        main.add(left, width=420)

        # ── Clone section ─────────────────────────────────────
        clone_lf = ttk.LabelFrame(left, text="Add GitHub Repo to Vault")
        clone_lf.pack(fill="x", padx=4, pady=(4, 6))

        r1 = ttk.Frame(clone_lf)
        r1.pack(fill="x", padx=6, pady=(4, 2))
        ttk.Label(r1, text="URL:", width=10).pack(side="left")
        self._vmgr_url_var = tk.StringVar()
        ttk.Entry(r1, textvariable=self._vmgr_url_var, width=40).pack(side="left", padx=4)

        r2 = ttk.Frame(clone_lf)
        r2.pack(fill="x", padx=6, pady=2)
        ttk.Label(r2, text="Subfolder:", width=10).pack(side="left")
        self._vmgr_subfolder_var = tk.StringVar()
        ttk.Entry(r2, textvariable=self._vmgr_subfolder_var,
                  width=20).pack(side="left", padx=4)
        ttk.Label(r2, text="Branch:", width=7).pack(side="left", padx=(10, 0))
        self._vmgr_branch_var = tk.StringVar()
        ttk.Entry(r2, textvariable=self._vmgr_branch_var,
                  width=12).pack(side="left", padx=4)

        r3 = ttk.Frame(clone_lf)
        r3.pack(fill="x", padx=6, pady=(2, 6))
        ttk.Button(r3, text="⬇  Clone Repo",
                   command=self._vmgr_clone).pack(side="left")
        ttk.Button(r3, text="🔄 Pull Updates",
                   command=self._vmgr_pull).pack(side="left", padx=6)

        # ── Import section ────────────────────────────────────
        import_lf = ttk.LabelFrame(left, text="Import Files into Vault")
        import_lf.pack(fill="x", padx=4, pady=(0, 6))

        # Row 1 — zip import
        zi1 = ttk.Frame(import_lf)
        zi1.pack(fill="x", padx=6, pady=(4, 2))
        ttk.Label(zi1, text="Zip file:", width=10).pack(side="left")
        self._vmgr_zip_var = tk.StringVar()
        ttk.Entry(zi1, textvariable=self._vmgr_zip_var,
                  width=32).pack(side="left", padx=4)
        ttk.Button(zi1, text="📂 Browse",
                   command=self._vmgr_browse_zip).pack(side="left")

        # Row 2 — zip subfolder + action
        zi2 = ttk.Frame(import_lf)
        zi2.pack(fill="x", padx=6, pady=(2, 2))
        ttk.Label(zi2, text="Subfolder:", width=10).pack(side="left")
        self._vmgr_zip_subfolder_var = tk.StringVar()
        ttk.Entry(zi2, textvariable=self._vmgr_zip_subfolder_var,
                  width=20).pack(side="left", padx=4)
        ttk.Label(zi2, text="(blank = zip name)", foreground="#5a3030").pack(side="left")

        zi3 = ttk.Frame(import_lf)
        zi3.pack(fill="x", padx=6, pady=(2, 6))
        ttk.Button(zi3, text="📦 Extract Zip to Vault",
                   command=self._vmgr_import_zip).pack(side="left")

        # Row — BATCH zip import: a folder full of .zip files, extract them all
        zb1 = ttk.Frame(import_lf)
        zb1.pack(fill="x", padx=6, pady=(2, 2))
        ttk.Label(zb1, text="Zip folder:", width=10).pack(side="left")
        self._vmgr_zipdir_var = tk.StringVar()
        ttk.Entry(zb1, textvariable=self._vmgr_zipdir_var,
                  width=32).pack(side="left", padx=4)
        ttk.Button(zb1, text="📂 Browse",
                   command=self._vmgr_browse_zip_folder).pack(side="left")
        zb2 = ttk.Frame(import_lf)
        zb2.pack(fill="x", padx=6, pady=(2, 6))
        ttk.Button(zb2, text="📦 Extract ALL zips in folder",
                   command=self._vmgr_import_zip_folder).pack(side="left")
        ttk.Label(zb2, text="(each zip → its own subfolder)",
                  foreground="#5a3030").pack(side="left", padx=6)

        # Row — folder import
        fi1 = ttk.Frame(import_lf)
        fi1.pack(fill="x", padx=6, pady=(2, 2))
        ttk.Label(fi1, text="Folder:", width=10).pack(side="left")
        self._vmgr_folder_var = tk.StringVar()
        ttk.Entry(fi1, textvariable=self._vmgr_folder_var,
                  width=32).pack(side="left", padx=4)
        ttk.Button(fi1, text="📂 Browse",
                   command=self._vmgr_browse_folder).pack(side="left")

        fi2 = ttk.Frame(import_lf)
        fi2.pack(fill="x", padx=6, pady=(2, 6))
        ttk.Button(fi2, text="📁 Copy Folder to Vault",
                   command=self._vmgr_import_folder).pack(side="left")

        # ── Index & Vectorize section ─────────────────────────
        idx_lf = ttk.LabelFrame(left, text="🔎 Index & Vectorize")
        idx_lf.pack(fill="x", padx=4, pady=(0, 6))

        ttk.Label(
            idx_lf,
            text=("Builds three retrieval layers. Run them in order:\n"
                  "  1. Keyword — fast file walk; required for search.\n"
                  "  2. Descriptions — LLM summaries; better semantics.\n"
                  "  3. Vectors — embedding model; best for fuzzy queries.\n"),
            justify="left", foreground="#9a9a9a"
        ).pack(anchor="w", padx=6, pady=(4, 0))

        self._idx_status_var = tk.StringVar(value="")
        ttk.Label(idx_lf, textvariable=self._idx_status_var,
                  foreground="#cba6f7").pack(anchor="w", padx=6)

        idx_btns = ttk.Frame(idx_lf)
        idx_btns.pack(fill="x", padx=6, pady=(2, 6))
        ttk.Button(idx_btns, text="1. Build Keyword Index",
                   command=self._vmgr_build_keyword_index
                   ).pack(side="left", padx=2)
        ttk.Button(idx_btns, text="2. Build Descriptions",
                   command=self._vmgr_build_descriptions
                   ).pack(side="left", padx=2)
        ttk.Button(idx_btns, text="3. Build Vector Embeddings",
                   command=self._vmgr_build_embeddings
                   ).pack(side="left", padx=2)

        # ── Convert Mongo BSON/JSON → model-readable section ──
        conv_lf = ttk.LabelFrame(left, text="🍃 Convert Mongo BSON / JSON")
        conv_lf.pack(fill="x", padx=4, pady=(0, 6))

        ttk.Label(
            conv_lf,
            text=("Flattens nested Mongo documents and turns ObjectId / dates /\n"
                  "Decimal128 / arrays into clean columns a model can read.\n"
                  "Source files are only read; output lands in data_in/"
                  "converted_mongo/\nso you can Index & Vectorize it above."),
            justify="left", foreground="#9a9a9a"
        ).pack(anchor="w", padx=6, pady=(4, 0))

        cv1 = ttk.Frame(conv_lf)
        cv1.pack(fill="x", padx=6, pady=(4, 2))
        ttk.Label(cv1, text="File:", width=6).pack(side="left")
        self._vmgr_mongo_var = tk.StringVar()
        ttk.Entry(cv1, textvariable=self._vmgr_mongo_var,
                  width=30).pack(side="left", padx=4)
        ttk.Button(cv1, text="📂 Browse",
                   command=self._vmgr_browse_mongo).pack(side="left")

        cv2 = ttk.Frame(conv_lf)
        cv2.pack(fill="x", padx=6, pady=(2, 2))
        self._vmgr_mongo_csv = tk.BooleanVar(value=True)
        self._vmgr_mongo_schema = tk.BooleanVar(value=True)
        self._vmgr_mongo_text = tk.BooleanVar(value=False)
        ttk.Checkbutton(cv2, text="Clean CSV",
                        variable=self._vmgr_mongo_csv).pack(side="left")
        ttk.Checkbutton(cv2, text="Schema profile",
                        variable=self._vmgr_mongo_schema).pack(side="left", padx=6)
        ttk.Checkbutton(cv2, text="Text digest",
                        variable=self._vmgr_mongo_text).pack(side="left")

        cv3 = ttk.Frame(conv_lf)
        cv3.pack(fill="x", padx=6, pady=(2, 4))
        ttk.Button(cv3, text="🍃 Convert File",
                   command=self._vmgr_convert_mongo).pack(side="left")
        ttk.Button(cv3, text="Convert ALL in vault",
                   command=lambda: self._vmgr_convert_mongo(scan_all=True)
                   ).pack(side="left", padx=6)
        ttk.Button(cv3, text="📂 Open Output",
                   command=self._vmgr_open_converted_mongo).pack(side="left")

        self._mongo_status_var = tk.StringVar(value="")
        ttk.Label(conv_lf, textvariable=self._mongo_status_var,
                  foreground="#cba6f7", wraplength=380, justify="left"
                  ).pack(anchor="w", padx=6, pady=(0, 4))

        # ── Deferred tasks (things the council couldn't do in chat) ──
        defer_lf = ttk.LabelFrame(left, text="📋 Deferred tasks")
        defer_lf.pack(fill="both", expand=False, padx=4, pady=(0, 6))
        ttk.Label(
            defer_lf, foreground="#9a9a9a", justify="left", wraplength=400,
            text="Tasks you sent here from the Council tab (⤓ Defer to Vault). "
                 "Run a summary/stats task with the full deterministic tooling, "
                 "or keep tool requests for the developer."
        ).pack(anchor="w", padx=6, pady=(4, 2))

        dcols = ("kind", "task")
        self._defer_tree = ttk.Treeview(defer_lf, columns=dcols,
                                        show="headings", height=5)
        self._defer_tree.heading("kind", text="Type")
        self._defer_tree.heading("task", text="Task")
        self._defer_tree.column("kind", width=110, anchor="w")
        self._defer_tree.column("task", width=300, anchor="w")
        self._defer_tree.pack(fill="x", padx=6, pady=(0, 2))
        self._defer_ids = {}      # tree-iid -> task id

        drow = ttk.Frame(defer_lf)
        drow.pack(fill="x", padx=6, pady=(0, 4))
        ttk.Button(drow, text="▶ Run",
                   command=self._vmgr_run_deferred).pack(side="left")
        ttk.Button(drow, text="✓ Done",
                   command=lambda: self._vmgr_set_deferred("done")
                   ).pack(side="left", padx=4)
        ttk.Button(drow, text="✗ Dismiss",
                   command=lambda: self._vmgr_set_deferred("dismissed")
                   ).pack(side="left", padx=4)
        ttk.Button(drow, text="⟳ Refresh",
                   command=self._vmgr_refresh_deferred).pack(side="left", padx=4)
        self._defer_status = tk.StringVar(value="")
        ttk.Label(defer_lf, textvariable=self._defer_status,
                  foreground="#cba6f7", wraplength=400, justify="left"
                  ).pack(anchor="w", padx=6, pady=(0, 4))
        self._vmgr_refresh_deferred()

        # ── Collections (group disparate files into a project) ──
        coll_lf = ttk.LabelFrame(left, text="📁 Collections (projects)")
        coll_lf.pack(fill="both", expand=False, padx=4, pady=(0, 6))
        ttk.Label(
            coll_lf, foreground="#9a9a9a", justify="left", wraplength=400,
            text="Group disparate files that belong together (e.g. “Job "
                 "Blue”). New… lets the council propose members from name / "
                 "value / shared-key signals; you confirm. Then the council "
                 "can pull up or summarise the whole set."
        ).pack(anchor="w", padx=6, pady=(4, 2))

        self._coll_tree = ttk.Treeview(coll_lf, columns=("name", "n"),
                                       show="headings", height=4)
        self._coll_tree.heading("name", text="Collection")
        self._coll_tree.heading("n", text="Files")
        self._coll_tree.column("name", width=300, anchor="w")
        self._coll_tree.column("n", width=60, anchor="center")
        self._coll_tree.pack(fill="x", padx=6, pady=(0, 2))
        self._coll_names = {}     # tree-iid -> collection name

        crow = ttk.Frame(coll_lf)
        crow.pack(fill="x", padx=6, pady=(0, 4))
        ttk.Button(crow, text="➕ New…",
                   command=self._vmgr_new_collection).pack(side="left")
        ttk.Button(crow, text="✎ Edit",
                   command=lambda: self._vmgr_new_collection(edit=True)
                   ).pack(side="left", padx=4)
        ttk.Button(crow, text="📊 Summarize",
                   command=self._vmgr_summarize_collection).pack(side="left", padx=4)
        ttk.Button(crow, text="✗ Delete",
                   command=self._vmgr_delete_collection).pack(side="left", padx=4)
        ttk.Button(crow, text="⟳",
                   command=self._vmgr_refresh_collections).pack(side="left", padx=4)
        self._coll_status = tk.StringVar(value="")
        ttk.Label(coll_lf, textvariable=self._coll_status, foreground="#cba6f7",
                  wraplength=400, justify="left").pack(anchor="w", padx=6, pady=(0, 4))
        self._vmgr_refresh_collections()

        # ── Scraper section ───────────────────────────────────
        scrape_lf = ttk.LabelFrame(left, text="🌐 Web Scraper")
        scrape_lf.pack(fill="x", padx=4, pady=(0, 6))

        if not _SCRAPER_OK:
            ttk.Label(scrape_lf,
                text="vault_scraper.py not found — place it next to council_gui_engine.py",
                foreground="#6c7086").pack(padx=6, pady=4, anchor="w")
        else:
            # Source selector
            sc_r1 = ttk.Frame(scrape_lf)
            sc_r1.pack(fill="x", padx=6, pady=(4, 2))
            ttk.Label(sc_r1, text="Source:", width=9).pack(side="left")
            self._scraper_source_var = tk.StringVar(value="All default sources")

            def _build_source_names():
                try:
                    names  = ["All default sources",
                               "All large sources",
                               "All sitemap sources"]
                    names += [lbl for lbl, *_ in vs.DEFAULT_SOURCES]
                    names += ["── Large sites (slow) ──"]
                    names += [lbl for lbl, *_ in vs.LARGE_SOURCES]
                    names += ["── Sitemap crawls ──"]
                    names += [lbl for lbl, *_ in vs.SITEMAP_SOURCES]
                    names += ["── GitHub raw files ──"]
                    names += [lbl for lbl, _ in vs.GITHUB_RAW_FILES]
                    return names
                except Exception:
                    return ["All default sources", "All large sources",
                            "All sitemap sources"]

            _source_names = _build_source_names()
            self._scraper_source_cb = ttk.Combobox(
                sc_r1, textvariable=self._scraper_source_var,
                values=_source_names, state="readonly", width=34)
            self._scraper_source_cb.pack(side="left", padx=4)

            def _refresh_sources():
                self._scraper_source_cb["values"] = _build_source_names()
            self.after(300, _refresh_sources)
            self.after(1500, _refresh_sources)

            # Custom URL row
            sc_r2 = ttk.Frame(scrape_lf)
            sc_r2.pack(fill="x", padx=6, pady=2)
            ttk.Label(sc_r2, text="Custom URL:", width=9).pack(side="left")
            self._scraper_url_var = tk.StringVar()
            ttk.Entry(sc_r2, textvariable=self._scraper_url_var,
                      width=34).pack(side="left", padx=4)

            # Label + depth row
            sc_r3 = ttk.Frame(scrape_lf)
            sc_r3.pack(fill="x", padx=6, pady=2)
            ttk.Label(sc_r3, text="Label:", width=9).pack(side="left")
            self._scraper_label_var = tk.StringVar(value="custom")
            ttk.Entry(sc_r3, textvariable=self._scraper_label_var,
                      width=14).pack(side="left", padx=4)
            ttk.Label(sc_r3, text="Max pages:").pack(side="left", padx=(10, 2))
            self._scraper_max_var = tk.StringVar(value="30")
            ttk.Entry(sc_r3, textvariable=self._scraper_max_var,
                      width=5).pack(side="left")

            # Options row
            sc_r4 = ttk.Frame(scrape_lf)
            sc_r4.pack(fill="x", padx=6, pady=2)
            self._scraper_skip_existing = tk.BooleanVar(value=True)
            self._scraper_dry_run       = tk.BooleanVar(value=False)
            self._scraper_no_github     = tk.BooleanVar(value=False)
            ttk.Checkbutton(sc_r4, text="Skip existing",
                variable=self._scraper_skip_existing).pack(side="left")
            ttk.Checkbutton(sc_r4, text="Dry run",
                variable=self._scraper_dry_run).pack(side="left", padx=6)
            ttk.Checkbutton(sc_r4, text="Skip GitHub files",
                variable=self._scraper_no_github).pack(side="left")

            # Action buttons + stop
            sc_r5 = ttk.Frame(scrape_lf)
            sc_r5.pack(fill="x", padx=6, pady=(4, 6))
            self._scraper_run_btn = ttk.Button(
                sc_r5, text="▶  Scrape", command=self._scraper_run)
            self._scraper_run_btn.pack(side="left")
            self._scraper_stop_btn = ttk.Button(
                sc_r5, text="■  Stop", command=self._scraper_stop,
                state="disabled")
            self._scraper_stop_btn.pack(side="left", padx=6)
            self._scraper_status = ttk.Label(
                sc_r5, text="idle", foreground="#6c7086")
            self._scraper_status.pack(side="left")

            self._scraper_running = False
            self._scraper_abort   = False

            def _on_source_change(event=None):
                sel = self._scraper_source_var.get()
                try:
                    _large_labels = (
                        {lbl for lbl, *_ in vs.LARGE_SOURCES}
                        | {lbl for lbl, *_ in vs.SITEMAP_SOURCES}
                        | {"All large sources", "All sitemap sources"}
                    )
                except Exception:
                    _large_labels = {"All large sources", "All sitemap sources"}
                if sel in _large_labels:
                    self._scraper_status.configure(
                        text="⚠ large crawl — may take 5–30 min",
                        foreground="#fab387")
                elif sel.startswith("──"):
                    self._scraper_status.configure(text="(separator)", foreground="#6c7086")
                else:
                    self._scraper_status.configure(text="idle", foreground="#6c7086")
            self._scraper_source_cb.bind("<<ComboboxSelected>>", _on_source_change)

        # ── Database Connections (read-only) ──────────────────
        # Save / list / test / remove saved SQL + Mongo connections.
        # Every read here is enforced read-only via the db_connections
        # module: client-side validator + per-dialect session hint +
        # API-design read-only (Mongo) + audit log to vault/db_audit.log.
        # See DATABASE_CONNECTIONS.md for the threat model + the
        # recommended DB-side read-only role per database.
        self._build_vmgr_db_connections_panel(left)

        # ── Data stats precompute ──────────────────────────────
        # Manual trigger for the incremental column-stats cache. Runs
        # the same sweep the background timer does, on demand, with a
        # result line in the activity log. Incremental: only unprocessed
        # CSVs are touched, so repeat clicks are near-instant.
        stats_lf = ttk.LabelFrame(left, text="📊 Data stats (precomputed)")
        stats_lf.pack(fill="x", padx=4, pady=(0, 6))
        srow = ttk.Frame(stats_lf)
        srow.pack(fill="x", padx=6, pady=4)
        ttk.Button(srow, text="🧮 Build / update stats",
                   command=self._vmgr_build_stats).pack(side="left")
        ttk.Label(srow, text="min/max/mean/… per column, cached",
                  foreground="#6c7086", font=("", 8)).pack(side="left", padx=6)

        # ── Vault tree ────────────────────────────────────────
        ttk.Label(left, text="Vault Contents",
                  font=("", 9, "bold")).pack(anchor="w", padx=4, pady=(4, 0))

        tree_frame = ttk.Frame(left)
        tree_frame.pack(fill="both", expand=True, padx=4, pady=4)

        self._vmgr_tree = ttk.Treeview(tree_frame, columns=("size", "type"),
                                        show="tree headings", selectmode="browse")
        self._vmgr_tree.heading("#0",    text="Name")
        self._vmgr_tree.heading("size",  text="Size")
        self._vmgr_tree.heading("type",  text="Type")
        self._vmgr_tree.column("#0",    width=220)
        self._vmgr_tree.column("size",  width=70,  anchor="e")
        self._vmgr_tree.column("type",  width=60,  anchor="center")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self._vmgr_tree.yview)
        self._vmgr_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._vmgr_tree.pack(fill="both", expand=True)
        self._vmgr_tree.bind("<<TreeviewSelect>>", self._vmgr_on_select)

        # Tree action buttons
        tb = ttk.Frame(left)
        tb.pack(fill="x", padx=4, pady=(0, 4))
        ttk.Button(tb, text="👁 Preview",
                   command=self._vmgr_preview).pack(side="left")
        ttk.Button(tb, text="🗑 Delete Item",
                   command=self._vmgr_delete).pack(side="left", padx=4)
        ttk.Button(tb, text="📋 Copy Path",
                   command=self._vmgr_copy_path).pack(side="left")

        # ── RIGHT: preview + log ──────────────────────────────
        right = tk.PanedWindow(main, orient="vertical",
                               bg="#1a1414", sashwidth=5)
        main.add(right)

        # File preview
        prev_frame = ttk.LabelFrame(right, text="Preview")
        right.add(prev_frame, height=340)
        self._vmgr_preview_text = self._make_text(
            prev_frame, wrap="word", height=16, state="disabled")
        self._vmgr_preview_text.pack(fill="both", expand=True, padx=4, pady=4)

        # Activity log
        log_frame = ttk.LabelFrame(right, text="Activity Log")
        right.add(log_frame, height=180)
        self._vmgr_log = self._make_text(
            log_frame, wrap="word", height=8, state="disabled")
        self._vmgr_log.pack(fill="both", expand=True, padx=4, pady=4)
        self._vmgr_log.tag_config("ok",   foreground="#a6e3a1")
        self._vmgr_log.tag_config("err",  foreground="#f38ba8")
        self._vmgr_log.tag_config("info", foreground="#d32f2f")

        # Populate tree
        self._vmgr_refresh_tree()

    # ── Database Connections panel ─────────────────────────────────────────
    # Builds the "🔌 Database Connections" LabelFrame in the Vault tab's left
    # pane. Add / list / test / remove SQL + Mongo connections through the
    # UI. All operations route through db_connections.py which enforces
    # the read-only policy (validator + session hints + API surface +
    # audit log).

    def _build_vmgr_db_connections_panel(self, parent) -> None:
        import tkinter as tk
        db_lf = ttk.LabelFrame(parent, text="🔌 Database Connections (read-only)")
        db_lf.pack(fill="x", padx=4, pady=(0, 6))

        # ── Guided setup (recommended for non-technical users) ──
        # A field-based wizard (Server / Database / Sign-in / Password)
        # that assembles + encodes the connection URL for the user and
        # tests it before saving. The manual URL row below stays for
        # power users who'd rather paste a connection string.
        guided_row = ttk.Frame(db_lf)
        guided_row.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Button(guided_row, text="➕ Connect a database (guided)…",
                   command=self._db_conn_open_wizard).pack(side="left")
        ttk.Label(guided_row, text="  ← easiest: fill in the fields",
                  foreground="#6c7086", font=("", 8)).pack(side="left")
        ttk.Separator(db_lf, orient="horizontal").pack(fill="x", padx=6, pady=4)

        # ── Add-connection row (manual URL — power users) ───────
        # Name + type + URL/URI + Save. The "type" dropdown drives the
        # placeholder we put in the URL box and where the connection is
        # stored (sql_connections.json vs mongo_connections.json).
        add_row = ttk.Frame(db_lf)
        add_row.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Label(add_row, text="Name:", width=6).pack(side="left")
        self._db_conn_name_var = tk.StringVar()
        ttk.Entry(add_row, textvariable=self._db_conn_name_var,
                   width=12).pack(side="left", padx=(2, 6))
        ttk.Label(add_row, text="Type:").pack(side="left")
        self._db_conn_type_var = tk.StringVar(value="postgresql")
        type_cb = ttk.Combobox(
            add_row, textvariable=self._db_conn_type_var,
            values=("postgresql", "mysql", "mssql", "sqlite", "duckdb",
                    "mongodb"),
            state="readonly", width=10,
        )
        type_cb.pack(side="left", padx=(2, 6))

        url_row = ttk.Frame(db_lf)
        url_row.pack(fill="x", padx=6, pady=2)
        ttk.Label(url_row, text="URL:", width=6).pack(side="left")
        self._db_conn_url_var = tk.StringVar()
        url_entry = ttk.Entry(url_row, textvariable=self._db_conn_url_var)
        url_entry.pack(side="left", fill="x", expand=True, padx=(2, 4))

        def _update_url_placeholder(*_):
            """Set a type-appropriate URL template in the entry box so
            the user can edit instead of typing from scratch. Only fills
            when the entry is empty so we don't clobber a half-typed URL."""
            cur = self._db_conn_url_var.get().strip()
            if cur:
                return
            t = self._db_conn_type_var.get()
            templates = {
                "postgresql": "postgresql://readonly_user:${PG_PASS}@host:5432/dbname",
                "mysql":      "mysql+pymysql://readonly_user:${MYSQL_PASS}@host:3306/dbname",
                "mssql":      "mssql+pyodbc://readonly_user:${MSSQL_PASS}@host:1433/dbname?driver=ODBC+Driver+17+for+SQL+Server",
                "sqlite":     "sqlite:///C:/path/to/file.db",
                "duckdb":     "duckdb:///C:/path/to/file.duckdb",
                "mongodb":    "mongodb://readonly_user:${MONGO_PASS}@host:27017/?authSource=admin",
            }
            self._db_conn_url_var.set(templates.get(t, ""))
        type_cb.bind("<<ComboboxSelected>>", _update_url_placeholder)

        btn_row = ttk.Frame(db_lf)
        btn_row.pack(fill="x", padx=6, pady=(2, 4))
        ttk.Button(btn_row, text="💾 Save",
                   command=self._db_conn_save).pack(side="left")
        ttk.Button(btn_row, text="🧪 Test",
                   command=self._db_conn_test).pack(side="left", padx=4)
        ttk.Button(btn_row, text="🗑 Remove",
                   command=self._db_conn_remove).pack(side="left")

        # ── Saved connections list ─────────────────────────────
        list_lbl = ttk.Label(
            db_lf, text="Saved connections (double-click → browse):",
            foreground="#94a3b8",
        )
        list_lbl.pack(anchor="w", padx=6, pady=(4, 0))
        self._db_conn_list = tk.Listbox(
            db_lf, height=5, bg="#231a1a", fg="#d4d4d4",
            selectbackground="#5a3030", relief="flat",
            font=("Consolas", 9),
        )
        self._db_conn_list.pack(fill="x", padx=6, pady=2)
        self._db_conn_list.bind("<Double-Button-1>",
                                 self._db_conn_browse_selected)

        # Status line — populated by the Test button and the auto-
        # refresh path. Keeps long messages (full URLs, errors) out
        # of the activity log.
        self._db_conn_status = ttk.Label(
            db_lf, text="(no connections saved yet)",
            foreground="#6c7086",
        )
        self._db_conn_status.pack(anchor="w", padx=6, pady=(2, 4))

        # Initial population
        self._db_conn_refresh_list()

    # ── Connections-panel actions ──────────────────────────────────

    def _db_conn_open_wizard(self) -> None:
        """Open the guided, field-based connection wizard. On a
        successful save it refreshes the saved-connections list."""
        try:
            import db_connect_wizard as _wiz
        except Exception as exc:
            from tkinter import messagebox
            messagebox.showerror(
                "Wizard unavailable", f"Could not load the wizard: {exc!r}")
            return

        def _on_saved(_name: str):
            try:
                self._db_conn_refresh_list()
            except Exception:
                pass

        _wiz.open_wizard(self, VAULT_DIR, on_saved=_on_saved,
                         log=self._vmgr_append)

    def _db_conn_refresh_list(self) -> None:
        """Rebuild the saved-connections listbox from the JSON files."""
        try:
            import db_connections as _db
            sql   = _db.list_sql_connections(VAULT_DIR)
            mongo = _db.list_mongo_connections(VAULT_DIR)
        except Exception as exc:
            self._db_conn_list.delete(0, "end")
            self._db_conn_status.configure(
                text=f"⚠ connection registry read failed: {exc}",
                foreground="#fab387",
            )
            return
        self._db_conn_list.delete(0, "end")
        for name in sorted(sql):
            url = sql[name]
            # Mask the password between :// and @
            masked = self._db_conn_mask(url)
            self._db_conn_list.insert("end", f"[sql]    {name}   {masked}")
        for name in sorted(mongo):
            url = mongo[name]
            masked = self._db_conn_mask(url)
            self._db_conn_list.insert("end", f"[mongo]  {name}   {masked}")
        total = len(sql) + len(mongo)
        if total == 0:
            self._db_conn_status.configure(
                text="(no connections saved yet)", foreground="#6c7086")
        else:
            self._db_conn_status.configure(
                text=f"{total} connection(s) saved — double-click a row to browse",
                foreground="#94a3b8",
            )

    @staticmethod
    def _db_conn_mask(url: str) -> str:
        """Mask any password in a connection URL for display in the
        listbox. Handles both SQLAlchemy (``foo://user:pass@host/db``)
        and Mongo (``mongodb://user:pass@host/`` + query). ${ENV_VAR}
        placeholders pass through untouched — there's nothing to mask."""
        import re as _re
        return _re.sub(
            r"(://[^:/@]+:)([^@]+)(@)",
            lambda m: m.group(1) + "***" + m.group(3),
            url,
        )

    def _db_conn_save(self) -> None:
        from tkinter import messagebox
        import db_connections as _db
        name = self._db_conn_name_var.get().strip()
        url  = self._db_conn_url_var.get().strip()
        kind = self._db_conn_type_var.get().strip().lower()
        if not name:
            messagebox.showwarning(
                "Connection name required",
                "Pick a short name for this connection (it's the handle "
                "you'll reference in analyst queries, e.g. "
                "'sales_db' → sql_query(VAULT_DIR, 'sales_db', '...')).")
            return
        if not url:
            messagebox.showwarning(
                "URL required",
                "Paste the connection URL or URI. Use ${ENV_VAR} for "
                "the password so it stays out of the JSON file:\n\n"
                "  postgresql://user:${PG_PASS}@host:5432/dbname")
            return
        # TLS / encryption posture check — surface as a yes/no
        # prompt BEFORE saving. The check is a no-op for ${ENV_VAR}-
        # protected URLs, localhost / LAN deployments, or URLs with
        # an explicit TLS hint (sslmode=require, ?tls=true, etc.).
        try:
            tls_warning = _db.check_tls_posture(url)
        except Exception:
            tls_warning = None
        if tls_warning:
            proceed = messagebox.askyesno(
                "TLS posture warning",
                tls_warning + "\n\nSave anyway?",
                default="no",
            )
            if not proceed:
                self._vmgr_append(
                    f"save aborted on TLS posture warning for {name}",
                    "info")
                return
        try:
            if kind == "mongodb":
                _db.save_mongo_connection(VAULT_DIR, name, url)
                self._vmgr_append(f"saved Mongo connection: {name}", "ok")
            else:
                _db.save_sql_connection(VAULT_DIR, name, url)
                self._vmgr_append(f"saved SQL connection: {name}", "ok")
        except ValueError as exc:
            # save_*_connection raises ValueError on empty/scheme-less
            # input — show the friendly message, not repr().
            messagebox.showerror("Save failed", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("Save failed", f"{exc!r}")
            return
        self._db_conn_name_var.set("")
        self._db_conn_url_var.set("")
        self._db_conn_refresh_list()

    def _db_conn_remove(self) -> None:
        from tkinter import messagebox
        import db_connections as _db
        sel = self._db_conn_list.curselection()
        if not sel:
            messagebox.showinfo(
                "Pick a connection",
                "Select a saved connection in the list first.")
            return
        line = self._db_conn_list.get(sel[0])
        kind, name = self._parse_listbox_line(line)
        if not name:
            return
        if not messagebox.askyesno(
                "Remove connection",
                f"Remove the {kind} connection {name!r}?\n\n"
                "The connection URL is deleted from "
                f"vault/{'mongo' if kind == 'mongo' else 'sql'}_connections.json. "
                "Your DB is not touched."):
            return
        if kind == "mongo":
            _db.remove_mongo_connection(VAULT_DIR, name)
        else:
            _db.remove_sql_connection(VAULT_DIR, name)
        self._vmgr_append(f"removed {kind} connection: {name}", "ok")
        self._db_conn_refresh_list()

    def _db_conn_test(self) -> None:
        """Probe the selected connection. SQL: SELECT 1 + table count.
        Mongo: ping + database count. Renders the result in the
        status line so the user can immediately see whether their
        connection works."""
        import db_connections as _db
        sel = self._db_conn_list.curselection()
        if not sel:
            # If no selection, also accept the name field as a test target
            name = self._db_conn_name_var.get().strip()
            kind = self._db_conn_type_var.get().strip().lower()
            if not name:
                self._db_conn_status.configure(
                    text="⚠ pick a saved connection or fill the Name field first",
                    foreground="#fab387",
                )
                return
        else:
            line = self._db_conn_list.get(sel[0])
            kind, name = self._parse_listbox_line(line)
        kind = "mongo" if kind == "mongo" else "sql"
        self._db_conn_status.configure(
            text=f"⏳ testing {kind} connection {name!r}…",
            foreground="#94a3b8",
        )
        self.update_idletasks()
        try:
            if kind == "mongo":
                result = _db.test_mongo_connection(VAULT_DIR, name)
                if result.get("ok"):
                    self._db_conn_status.configure(
                        text=(f"✓ {name} — connected · "
                               f"{result.get('databases', 0)} database(s) "
                               "visible"),
                        foreground="#a6e3a1",
                    )
                    self._vmgr_append(
                        f"tested mongo connection {name}: OK", "ok")
                else:
                    self._db_conn_status.configure(
                        text=f"✗ {name} — {result.get('error') or 'unknown error'}",
                        foreground="#f38ba8",
                    )
                    self._vmgr_append(
                        f"tested mongo connection {name}: FAILED — "
                        f"{result.get('error')}", "err")
                    _record_app_failure(
                        "db.mongo_test_failed", "db_connections",
                        str(result.get("error") or "unknown error"),
                        context={"connection": name})
            else:
                result = _db.test_sql_connection(VAULT_DIR, name)
                if result.get("ok"):
                    self._db_conn_status.configure(
                        text=(f"✓ {name} — connected ({result.get('dialect')}) · "
                               f"{result.get('tables', 0)} table(s) visible · "
                               "SELECT 1 ok"),
                        foreground="#a6e3a1",
                    )
                    self._vmgr_append(
                        f"tested sql connection {name}: OK "
                        f"({result.get('dialect')})", "ok")
                else:
                    self._db_conn_status.configure(
                        text=f"✗ {name} — {result.get('error') or 'unknown error'}",
                        foreground="#f38ba8",
                    )
                    self._vmgr_append(
                        f"tested sql connection {name}: FAILED — "
                        f"{result.get('error')}", "err")
                    _record_app_failure(
                        "db.sql_test_failed", "db_connections",
                        str(result.get("error") or "unknown error"),
                        context={"connection": name})
            # Surface TLS warning if the check returned one (independent
            # of OK/FAIL — a successful connection over plaintext is
            # exactly when the warning is most actionable).
            tls_warning = result.get("tls_warning")
            if tls_warning:
                self._vmgr_append(f"TLS posture: {tls_warning}", "info")
            # If the URL has ${ENV_VAR} placeholders that aren't set,
            # tell the user EXACTLY which ones — much friendlier than
            # an opaque auth-fail message from the driver.
            missing_env = result.get("unresolved_env_vars")
            if missing_env:
                self._vmgr_append(
                    "URL references unset environment variable(s): "
                    + ", ".join(missing_env)
                    + " — set them in your shell and re-test.",
                    "warn",
                )
        except Exception as exc:
            self._db_conn_status.configure(
                text=f"✗ test raised {exc!r}",
                foreground="#f38ba8",
            )

    def _db_conn_browse_selected(self, _event=None) -> None:
        """Double-click handler: open a separate window listing the
        tables (SQL) or collections (Mongo) of the selected
        connection. Read-only — clicking a table/collection runs a
        SELECT / find() with a 50-row cap."""
        import tkinter as tk
        from tkinter import messagebox
        import db_connections as _db
        sel = self._db_conn_list.curselection()
        if not sel:
            return
        line = self._db_conn_list.get(sel[0])
        kind, name = self._parse_listbox_line(line)
        if not name:
            return

        win = tk.Toplevel(self)
        win.title(f"{kind.upper()}: {name}")
        win.geometry("760x560")
        win.configure(bg="#1a1414")

        # Top: tables / collections list
        top = ttk.Frame(win)
        top.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Label(top, text=f"{name}  ({kind})",
                   font=("", 10, "bold")).pack(side="left")

        list_frame = ttk.Frame(win)
        list_frame.pack(fill="x", padx=6, pady=2)
        lst = tk.Listbox(list_frame, height=8, bg="#231a1a", fg="#d4d4d4",
                          selectbackground="#5a3030", relief="flat",
                          font=("Consolas", 9))
        lst.pack(fill="x")

        # Middle: export row — pull the SELECTED table/collection to a
        # file in vault/data_out/db_exports/. Read-only by construction
        # (routes through db_connections.export_*, which only READ from
        # the DB and WRITE a local file — no DB-write path exists).
        export_bar = ttk.Frame(win)
        export_bar.pack(fill="x", padx=6, pady=(2, 0))
        ttk.Label(export_bar, text="Export selected →",
                   foreground="#a6adc8").pack(side="left", padx=(0, 4))
        for _fmt, _lbl in (("csv", "CSV"), ("json", "JSON"), ("xlsx", "Excel")):
            ttk.Button(
                export_bar, text=_lbl, width=7,
                command=lambda f=_fmt: _export_selected(f),
            ).pack(side="left", padx=2)
        ttk.Label(export_bar,
                   text="(full table, read-only — never deletes)",
                   foreground="#6c7086", font=("", 8)).pack(side="left", padx=6)

        # Bottom: preview text
        preview_lf = ttk.LabelFrame(win,
                                    text="Preview (first 50 rows, read-only)")
        preview_lf.pack(fill="both", expand=True, padx=6, pady=4)
        preview = tk.Text(preview_lf, wrap="none",
                           bg="#231a1a", fg="#d4d4d4",
                           font=("Consolas", 9))
        preview.pack(fill="both", expand=True, padx=4, pady=4)

        def _populate_tables_or_collections():
            lst.delete(0, "end")
            try:
                if kind == "mongo":
                    dbs = _db.list_mongo_databases(VAULT_DIR, name)
                    for db_name in dbs:
                        cols = _db.list_mongo_collections(
                            VAULT_DIR, name, db_name)
                        for c in cols:
                            lst.insert("end", f"{db_name}.{c}")
                else:
                    tables = _db.list_sql_tables(VAULT_DIR, name)
                    for t in tables:
                        lst.insert("end", t)
            except Exception as exc:
                lst.insert("end", f"(error: {exc})")

        def _on_pick(_evt=None):
            sel2 = lst.curselection()
            if not sel2:
                return
            picked = lst.get(sel2[0])
            preview.delete("1.0", "end")
            preview.insert("end", f"loading first 50 rows of {picked}…\n")
            self.update_idletasks()
            try:
                if kind == "mongo":
                    if "." not in picked:
                        raise ValueError("expected db.collection format")
                    db_name, coll = picked.split(".", 1)
                    df = _db.read_mongo_collection(
                        VAULT_DIR, name, db_name, coll, limit=50)
                else:
                    df = _db.read_sql_table(
                        VAULT_DIR, name, picked, limit=50)
                preview.delete("1.0", "end")
                if df.empty:
                    preview.insert("end", "(no rows)\n")
                else:
                    preview.insert("end", df.to_string(index=False) + "\n")
                self._vmgr_append(
                    f"previewed {kind} {name}.{picked} — "
                    f"{len(df)} row(s)", "ok")
            except Exception as exc:
                preview.delete("1.0", "end")
                preview.insert("end", f"✗ {exc!r}\n")
                self._vmgr_append(
                    f"preview failed for {kind} {name}.{picked}: {exc}",
                    "err")

        def _export_selected(fmt: str):
            sel2 = lst.curselection()
            if not sel2:
                messagebox.showinfo(
                    "Pick a table",
                    "Select a table/collection in the list first.",
                    parent=win)
                return
            picked = lst.get(sel2[0])
            try:
                import data_index as _di
                export_dir = _di.output_dir(VAULT_DIR) / "db_exports"
            except Exception:
                export_dir = VAULT_DIR / "data_out" / "db_exports"
            dest = export_dir / _db._safe_export_name(
                f"{name}__{picked}", fmt)
            preview.delete("1.0", "end")
            preview.insert("end", f"exporting {picked} → {dest} …\n")
            self.update_idletasks()
            try:
                if kind == "mongo":
                    if "." not in picked:
                        raise ValueError("expected db.collection format")
                    db_name, coll = picked.split(".", 1)
                    info = _db.export_mongo_collection(
                        VAULT_DIR, name, db_name, coll, dest, fmt=fmt)
                else:
                    info = _db.export_sql_table(
                        VAULT_DIR, name, picked, dest, fmt=fmt)
                preview.delete("1.0", "end")
                preview.insert(
                    "end",
                    f"✓ exported {info['rows']} row(s) to:\n{info['path']}\n")
                self._vmgr_append(
                    f"exported {kind} {name}.{picked} → "
                    f"{info['rows']} row(s) ({fmt}) at {info['path']}", "ok")
                try:
                    self._open_in_explorer(Path(info["path"]).parent)
                except Exception:
                    pass
            except Exception as exc:
                preview.delete("1.0", "end")
                preview.insert("end", f"✗ export failed: {exc!r}\n")
                self._vmgr_append(
                    f"export failed for {kind} {name}.{picked}: {exc}", "err")

        lst.bind("<<ListboxSelect>>", _on_pick)
        _populate_tables_or_collections()

    @staticmethod
    def _parse_listbox_line(line: str) -> "tuple[str, str]":
        """Parse a row of the saved-connections listbox.
        Format: ``[sql]    name   url``. Returns (kind, name)."""
        if not line:
            return ("sql", "")
        kind = "mongo" if line.startswith("[mongo]") else "sql"
        # Strip the "[kind]   " prefix
        rest = line.split("]", 1)[1].strip() if "]" in line else line
        # Pull the first whitespace-separated token as the name
        parts = rest.split(None, 1)
        return (kind, parts[0] if parts else "")

    # ── Vault Scraper helpers ──────────────────────────────────────────────

    def _scraper_log(self, msg: str, tag: str = "info"):
        """Write a line to the vault activity log from the scraper."""
        self._vmgr_append(msg, tag)

    def _scraper_status_set(self, text: str, color: str = "#6c7086"):
        try:
            self._scraper_status.configure(text=text, foreground=color)
        except Exception:
            pass

    def _scraper_stop(self):
        self._scraper_abort = True
        self._scraper_status_set("stopping…", "#fab387")

    def _scraper_run(self):
        if not _SCRAPER_OK:
            return
        if self._scraper_running:
            return

        source_sel  = self._scraper_source_var.get()
        custom_url  = self._scraper_url_var.get().strip()
        if custom_url and not custom_url.startswith(("http://", "https://")):
            custom_url = "https://" + custom_url
        label       = self._scraper_label_var.get().strip() or "custom"
        skip_exist  = self._scraper_skip_existing.get()
        dry_run     = self._scraper_dry_run.get()
        no_github   = self._scraper_no_github.get()
        try:
            max_pages = int(self._scraper_max_var.get())
        except ValueError:
            max_pages = 30

        self._scraper_running = True
        self._scraper_abort   = False
        self._scraper_run_btn.configure(state="disabled")
        self._scraper_stop_btn.configure(state="normal")
        self._scraper_status_set("running…", "#a6e3a1")

        vault_dir = VAULT_DIR

        def _progress(msg: str, error: bool = False):
            """Called from worker thread — marshal to UI thread."""
            tag = "err" if error else ("ok" if "✓" in msg else "info")
            self.after(0, lambda m=msg, t=tag: self._scraper_log(m, t))

        def _worker():
            import time as _time
            total = 0
            try:
                index = vs._load_index(vault_dir)

                # ── Single custom URL ────────────────────────────
                if custom_url:
                    _progress(f"Fetching: {custom_url}")
                    # Try to give a reason on failure
                    try:
                        robots_ok = vs._can_fetch(custom_url)
                        if not robots_ok:
                            _progress(f"  ✗ Blocked by robots.txt: {custom_url}", True)
                        else:
                            html = vs._fetch(custom_url)
                            if html is None:
                                _progress(f"  ✗ Could not fetch (network error or bad URL): {custom_url}", True)
                            else:
                                text = vs._extract_text(html, custom_url)
                                if len(text) < 100:
                                    _progress(f"  ✗ Too little content extracted ({len(text)} chars): {custom_url}", True)
                                else:
                                    path = vs._write_vault(vault_dir, label, custom_url, text, index)
                                    _progress(f"  ✓ {len(text):,} chars → {path.name}")
                                    total += 1
                    except Exception as _fe:
                        _progress(f"  ✗ Error: {_fe}", True)
                    vs._save_index(vault_dir, index)

                else:
                    # Determine which tiers to run
                    _sep = source_sel.startswith("──")
                    if _sep:
                        _progress("Select a real source, not a separator line.", True)
                        return
                    _all_default = source_sel == "All default sources"
                    _all_large   = source_sel == "All large sources"
                    _all_sitemap = source_sel == "All sitemap sources"
                    _all_github  = source_sel == "All default sources"  # github included in default
                    _run_single  = not (_all_default or _all_large or _all_sitemap)

                    # ── GitHub raw files ──────────────────────────
                    if not no_github:
                        for gh_label, gh_url in vs.GITHUB_RAW_FILES:
                            if self._scraper_abort:
                                break
                            if not _all_default and not (_run_single and gh_label == source_sel):
                                continue
                            if skip_exist and gh_url in index:
                                _progress(f"[skip] {gh_label}")
                                continue
                            _progress(f"GitHub: {gh_label}")
                            ok = vs.fetch_single(
                                gh_label, gh_url, vault_dir, index,
                                raw=True, dry_run=dry_run, verbose=False,
                            )
                            _progress(f"  {'✓' if ok else '✗'} {gh_label}", not ok)
                            total += int(ok)
                            _time.sleep(0.4)

                    # ── Default doc site crawls ───────────────────
                    for src_label, seed, src_max, prefix in vs.DEFAULT_SOURCES:
                        if self._scraper_abort:
                            break
                        if not _all_default and not (_run_single and src_label == source_sel):
                            continue
                        if skip_exist and seed in index:
                            _progress(f"[skip] {src_label}")
                            continue
                        _progress(f"Crawling: {src_label}  ({seed})", False)
                        self.after(0, lambda l=src_label:
                            self._scraper_status_set(f"scraping {l}…", "#d32f2f"))
                        n = vs.crawl(
                            src_label, seed, prefix, vault_dir, index,
                            max_pages=max_pages or src_max,
                            delay=0.8, dry_run=dry_run, verbose=False,
                            progress_cb=_progress,
                            abort_cb=lambda: self._scraper_abort,
                        )
                        total += n
                        _progress(f"  ✓ {src_label}: {n} pages saved")
                        vs._save_index(vault_dir, index)
                        if self._scraper_abort:
                            break

                    # ── Large site crawls ─────────────────────────
                    for src_label, seed, src_max, prefix in vs.LARGE_SOURCES:
                        if self._scraper_abort:
                            break
                        if not _all_large and not (_run_single and src_label == source_sel):
                            continue
                        if skip_exist and seed in index:
                            _progress(f"[skip] {src_label}")
                            continue
                        _progress(f"Large crawl: {src_label}  ({seed})", False)
                        self.after(0, lambda l=src_label:
                            self._scraper_status_set(f"large: {l}…", "#fab387"))
                        n = vs.crawl(
                            src_label, seed, prefix, vault_dir, index,
                            max_pages=max_pages or src_max,
                            delay=0.7, dry_run=dry_run, verbose=False,
                            progress_cb=_progress,
                            abort_cb=lambda: self._scraper_abort,
                        )
                        total += n
                        _progress(f"  ✓ {src_label}: {n} pages saved")
                        vs._save_index(vault_dir, index)
                        if self._scraper_abort:
                            break

                    # ── Sitemap crawls ────────────────────────────
                    for src_label, sitemap_url, src_max in vs.SITEMAP_SOURCES:
                        if self._scraper_abort:
                            break
                        if not _all_sitemap and not (_run_single and src_label == source_sel):
                            continue
                        if skip_exist and sitemap_url in index:
                            _progress(f"[skip] {src_label}")
                            continue
                        _progress(f"Sitemap crawl: {src_label}", False)
                        self.after(0, lambda l=src_label:
                            self._scraper_status_set(f"sitemap: {l}…", "#cba6f7"))
                        n = vs.crawl_sitemap(
                            src_label, sitemap_url, vault_dir, index,
                            max_pages=max_pages or src_max,
                            delay=0.7, dry_run=dry_run, verbose=False,
                            progress_cb=_progress,
                            abort_cb=lambda: self._scraper_abort,
                        )
                        total += n
                        _progress(f"  ✓ {src_label}: {n} pages saved")
                        vs._save_index(vault_dir, index)
                        if self._scraper_abort:
                            break

                vs._save_index(vault_dir, index)

            except Exception as e:
                import traceback
                _progress(f"✗ Scraper error: {e}", True)
                _progress(traceback.format_exc(), True)
            finally:
                def _done():
                    self._scraper_running = False
                    self._scraper_abort   = False
                    self._scraper_run_btn.configure(state="normal")
                    self._scraper_stop_btn.configure(state="disabled")
                    status = "stopped" if self._scraper_abort else (
                        "dry-run complete" if dry_run else f"done — {total} pages"
                    )
                    self._scraper_status_set(status, "#a6e3a1")
                    self._vmgr_refresh_tree()
                self.after(0, _done)

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    # ── Vault Manager helpers ─────────────────────────────────

    def _vmgr_build_stats(self):
        """Manual trigger for the incremental column-stats precompute.
        Runs on a daemon thread (CPU/IO heavy on a cold vault) and posts
        progress + a final summary to the activity log via the UI queue."""
        self._vmgr_append("📊 building data stats (incremental — only new "
                          "files)…", "info")

        def _worker():
            try:
                def _prog(i, total, name):
                    if total and (i == total or i % 25 == 0):
                        self.ui_q.put(("agent_phase", "stats_index",
                                       f"  stats: {i}/{total} ({name})"))
                res = self._build_stats_index(on_progress=_prog)
                self.ui_q.put(("agent_phase", "stats_index",
                               f"✓ data stats ready — processed "
                               f"{res['processed']} new, {res['already_current']} "
                               f"already cached ({res['seen']} CSVs)."))
            except Exception as exc:
                self.ui_q.put(("agent_phase", "stats_index",
                               f"✗ stats build failed: {exc!r}"))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _vmgr_append(self, msg: str, tag: str = "info"):
        """Append a line to the vault manager log."""
        self._vmgr_log.configure(state="normal")
        if tag == "info" and msg.startswith("✓"):
            tag = "ok"
        elif tag == "info" and (msg.startswith("✗") or "error" in msg.lower() or "fail" in msg.lower()):
            tag = "err"
        self._vmgr_log.insert("end", msg.rstrip() + "\n", tag)
        self._vmgr_log.see("end")
        self._vmgr_log.configure(state="disabled")

    def _show_rag_misses(self):
        """Popup: vault_rag_misses.txt — queries that found no vault context."""
        miss_path = VAULT_DIR / "vault_rag_misses.txt"
        win = tk.Toplevel(self)
        win.title("RAG Miss Log — Vault Gaps")
        win.configure(bg="#1a1414")
        win.geometry("740x420")
        if not miss_path.exists():
            ttk.Label(win, text="No RAG misses recorded yet.\n"
                           "Misses are logged when RAG finds no relevant vault context.",
                      wraplength=600).pack(padx=20, pady=20)
            return
        try:
            lines = miss_path.read_text(encoding="utf-8").splitlines()
        except Exception as e:
            ttk.Label(win, text="Error reading miss log: " + str(e)).pack(padx=20, pady=20)
            return
        ttk.Label(win,
                  text=str(len(lines)) + " missed queries — these topics are not covered by your vault.",
                  foreground="#f38ba8").pack(anchor="w", padx=10, pady=(8, 2))
        ttk.Separator(win, orient="horizontal").pack(fill="x", padx=10, pady=4)
        box = self._make_text(win, height=18, wrap="word", state="normal")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 4))
        for ln in reversed(lines):
            parts = ln.split("\t", 1)
            ts    = parts[0] if len(parts) > 0 else ""
            query = parts[1] if len(parts) > 1 else ln
            box.insert("end", ts[:16] + "  ", "dim")
            box.insert("end", query + "\n")
        box.tag_configure("dim", foreground="#6c7086")
        box.configure(state="disabled")
        def _clear():
            try:
                miss_path.write_text("", encoding="utf-8")
                box.configure(state="normal")
                box.delete("1.0", "end")
                box.configure(state="disabled")
            except Exception:
                pass
        ttk.Button(win, text="Clear Miss Log", command=_clear).pack(
            anchor="e", padx=10, pady=(0, 8))

    def _vmgr_refresh_tree(self):
        """Rebuild the vault tree view."""
        self._vmgr_tree.delete(*self._vmgr_tree.get_children())
        if not VAULT_DIR.exists():
            return

        def _fmt_size(p):
            try:
                b = p.stat().st_size
                for unit in ("B", "KB", "MB"):
                    if b < 1024:
                        return f"{b:.0f} {unit}"
                    b /= 1024
                return f"{b:.1f} MB"
            except Exception:
                return ""

        # Insert top-level items
        for item in sorted(VAULT_DIR.iterdir(),
                           key=lambda p: (p.is_file(), p.name.lower())):
            if item.name.startswith("."):
                continue
            if item.is_dir():
                node = self._vmgr_tree.insert(
                    "", "end", iid=str(item),
                    text=f"📁 {item.name}", values=("", "folder"), open=False)
                # Insert children (one level deep)
                for child in sorted(item.iterdir(),
                                    key=lambda p: (p.is_file(), p.name.lower())):
                    if child.name.startswith("."):
                        continue
                    if child.is_file():
                        self._vmgr_tree.insert(
                            node, "end", iid=str(child),
                            text=f"  {child.name}",
                            values=(_fmt_size(child), child.suffix or "file"))
                    elif child.is_dir():
                        sub = self._vmgr_tree.insert(
                            node, "end", iid=str(child),
                            text=f"  📁 {child.name}", values=("", "folder"))
                        for f in sorted(child.iterdir())[:30]:
                            if f.is_file() and not f.name.startswith("."):
                                self._vmgr_tree.insert(
                                    sub, "end", iid=str(f),
                                    text=f"    {f.name}",
                                    values=(_fmt_size(f), f.suffix or "file"))
            else:
                self._vmgr_tree.insert(
                    "", "end", iid=str(item),
                    text=item.name,
                    values=(_fmt_size(item), item.suffix or "file"))

    def _vmgr_on_select(self, event=None):
        sel = self._vmgr_tree.selection()
        if not sel:
            return
        p = Path(sel[0])
        if p.is_file() and p.stat().st_size < 200_000:
            self._vmgr_show_preview(p)

    def _vmgr_show_preview(self, path: Path):
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            text = f"Cannot read file: {e}"
        self._vmgr_preview_text.configure(state="normal")
        self._vmgr_preview_text.delete("1.0", "end")
        self._vmgr_preview_text.insert("1.0", text[:8000])
        if len(text) > 8000:
            self._vmgr_preview_text.insert("end", "\n\n[truncated…]")
        self._vmgr_preview_text.configure(state="disabled")

    def _vmgr_preview(self):
        sel = self._vmgr_tree.selection()
        if not sel:
            return
        p = Path(sel[0])
        if p.is_file():
            self._vmgr_show_preview(p)

    def _vmgr_delete(self):
        import tkinter.messagebox as mb
        sel = self._vmgr_tree.selection()
        if not sel:
            return
        p = Path(sel[0])
        what = "folder and all its contents" if p.is_dir() else "file"
        if mb.askyesno("Confirm Delete", f"Delete {what}:\n{p.name}?"):
            try:
                import shutil
                if p.is_dir():
                    shutil.rmtree(p)
                else:
                    p.unlink()
                self._vmgr_append(f"✓ Deleted: {p.name}")
                self._vmgr_refresh_tree()
            except Exception as e:
                self._vmgr_append(f"✗ Delete failed: {e}")

    def _vmgr_copy_path(self):
        sel = self._vmgr_tree.selection()
        if not sel:
            return
        self.clipboard_clear()
        self.clipboard_append(sel[0])
        self._vmgr_append(f"Copied path: {sel[0]}", "info")

    def _vmgr_open_folder(self):
        import subprocess as sp
        sp.Popen(["explorer", str(VAULT_DIR)])

    # ---- Index & Vectorize button handlers ----

    def _vmgr_build_keyword_index(self):
        """Walk the vault and (re)build the keyword index. Fast — no LLM."""
        idx = _get_vault_index()
        if idx is None:
            self._idx_status_var.set("Vault index unavailable.")
            return
        self._idx_status_var.set("Walking vault…")

        def _worker():
            def _on_progress(done: int, total: int, name: str):
                # Trim very long filenames so the status bar doesn't
                # wrap or look ugly. Bounce to the UI thread via
                # self.after — Tk widgets are not thread-safe.
                short = name if len(name) <= 48 else name[:45] + "…"
                self.after(0, lambda: self._idx_status_var.set(
                    f"Indexing {done}/{total} — {short}"))
            try:
                n = idx.rebuild(progress=_on_progress)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._idx_status_var.set(
                    f"Keyword index failed: {exc!r}"))
                return
            self.after(0, lambda: self._idx_status_var.set(
                f"Keyword index built — {n} files (re)indexed, "
                f"{len(idx.records)} total."))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _vmgr_build_descriptions(self):
        """Generate per-file LLM descriptions for every record without one.
        Requires the keyword index to exist; takes ~3-10s per file on a 7B GGUF."""
        idx = _get_vault_index()
        if idx is None:
            self._idx_status_var.set("Vault index unavailable.")
            return
        try:
            idx.rebuild()
        except Exception:
            pass
        pending = sum(1 for r in idx.records.values()
                      if not r.get("description"))
        if pending == 0:
            self._idx_status_var.set(
                f"All {len(idx.records)} files already have descriptions.")
            return
        self._idx_status_var.set(
            f"Generating descriptions for {pending} files… (each ~3-10s)")

        def _worker():
            def _progress(i, total, name):
                if i % 3 == 0 or i == total:
                    self.after(0, lambda: self._idx_status_var.set(
                        f"Descriptions: {i}/{total}…"))
            try:
                n = idx.generate_descriptions(on_progress=_progress)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._idx_status_var.set(
                    f"Description build failed: {exc!r}"))
                return
            self.after(0, lambda: self._idx_status_var.set(
                f"Descriptions complete — {n} files summarized."))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _vmgr_build_embeddings(self):
        """Build vector embeddings for every record (one-time, then mtime-incremental).
        Downloads sentence-transformers model on first run (~80 MB)."""
        idx = _get_vault_index()
        if idx is None:
            self._idx_status_var.set("Vault index unavailable.")
            return
        try:
            idx.rebuild()
        except Exception:
            pass
        emb = idx.embeddings()
        if emb is None:
            self._idx_status_var.set(
                "sentence-transformers not available — pip install it first.")
            return
        self._idx_status_var.set(
            f"Embedding {len(idx.records)} files (model: {emb.model_name})…")

        def _worker():
            def _progress(i, total, name):
                if i % 10 == 0 or i == total:
                    self.after(0, lambda: self._idx_status_var.set(
                        f"Embeddings: {i}/{total}…"))
            try:
                n = idx.build_embeddings(on_progress=_progress)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._idx_status_var.set(
                    f"Embedding build failed: {exc!r}"))
                return
            stats = emb.stats()
            self.after(0, lambda: self._idx_status_var.set(
                f"Vectors ready — {stats['vectors']} files "
                f"({stats['dim']}-dim, {stats['size_kb']} KB on disk)."))

        import threading as _th
        _th.Thread(target=_worker, daemon=True).start()

    def _vmgr_clone(self):
        """Clone a GitHub repo into the vault in a background thread."""
        import threading
        url       = self._vmgr_url_var.get().strip()
        subfolder = self._vmgr_subfolder_var.get().strip() or None
        branch    = self._vmgr_branch_var.get().strip() or None

        if not url:
            self._vmgr_append("✗ Please enter a GitHub URL.", "err")
            return
        if not url.startswith("http"):
            self._vmgr_append("✗ URL must start with https://", "err")
            return

        self._vmgr_append(f"Cloning {url} …", "info")

        def worker():
            try:
                dest = _vmgr_clone_repo(
                    url,
                    vault_dir=VAULT_DIR,
                    subfolder=subfolder,
                    branch=branch,
                    log_cb=lambda m: self.ui_q.put(("vault_mgr_log", m)),
                )
                self.ui_q.put(("vault_mgr_log", f"✓ Done → {dest.name}"))
                self.ui_q.put(("vault_mgr_refresh", None))
            except Exception as e:
                self.ui_q.put(("vault_mgr_log", f"✗ Clone failed: {e}"))

        threading.Thread(target=worker, daemon=True).start()

    def _vmgr_pull(self):
        """Pull updates for the selected vault folder."""
        import threading
        sel = self._vmgr_tree.selection()
        if not sel:
            self._vmgr_append("✗ Select a repo folder in the tree first.", "err")
            return
        p = Path(sel[0])
        if not p.is_dir():
            p = p.parent
        subfolder = p.name
        clone_dir = VAULT_DIR / ".git_clones" / subfolder
        if not clone_dir.exists():
            self._vmgr_append(
                f"✗ No git clone found for '{subfolder}'. "
                "Use Clone Repo first.", "err")
            return

        self._vmgr_append(f"Pulling updates for {subfolder} …", "info")

        def worker():
            try:
                import subprocess
                r = subprocess.run(
                    ["git", "pull"], cwd=str(clone_dir),
                    capture_output=True, text=True, timeout=120)
                msg = r.stdout.strip() or r.stderr.strip() or "Done."
                self.ui_q.put(("vault_mgr_log", f"git pull: {msg}"))
                # Re-copy updated files
                rc2, url, _ = (lambda r2: (r2.returncode, r2.stdout.strip(), ""))(
                    subprocess.run(["git", "remote", "get-url", "origin"],
                                   cwd=str(clone_dir),
                                   capture_output=True, text=True, timeout=15))
                if rc2 == 0 and url:
                    _vmgr_clone_repo(
                        url, vault_dir=VAULT_DIR, subfolder=subfolder,
                        log_cb=lambda m: self.ui_q.put(("vault_mgr_log", m)),
                    )
                self.ui_q.put(("vault_mgr_log", f"✓ {subfolder} updated"))
                self.ui_q.put(("vault_mgr_refresh", None))
            except Exception as e:
                self.ui_q.put(("vault_mgr_log", f"✗ Pull failed: {e}"))

        threading.Thread(target=worker, daemon=True).start()

    # ── Vault Manager — zip / folder import ─────────────────────

    def _vmgr_browse_zip(self):
        """Open file dialog to select a zip file."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select zip file",
            filetypes=[("Zip files", "*.zip"), ("All files", "*.*")],
        )
        if path:
            self._vmgr_zip_var.set(path)
            # Auto-fill subfolder from zip name if blank
            if not self._vmgr_zip_subfolder_var.get().strip():
                stem = Path(path).stem
                self._vmgr_zip_subfolder_var.set(stem)

    def _vmgr_browse_folder(self):
        """Open directory dialog to select a folder to copy."""
        from tkinter import filedialog
        path = filedialog.askdirectory(title="Select folder to import into vault")
        if path:
            self._vmgr_folder_var.set(path)

    # ── Mongo BSON/JSON → model-readable conversion ──────────
    def _converted_mongo_dir(self):
        """Output folder for converted Mongo data: a subfolder of the vault
        INPUT dir so the clean CSVs are immediately indexable/queryable."""
        return data_index.input_dir(VAULT_DIR) / "converted_mongo"

    def _vmgr_browse_mongo(self):
        """File dialog to pick a MongoDB .bson / .json / .jsonl export."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select MongoDB BSON / JSON export",
            filetypes=[("Mongo data", "*.bson *.json *.jsonl"),
                       ("BSON", "*.bson"),
                       ("JSON / JSONL", "*.json *.jsonl"),
                       ("All files", "*.*")],
        )
        if path:
            self._vmgr_mongo_var.set(path)

    def _vmgr_open_converted_mongo(self):
        """Reveal the converted-Mongo output folder in the OS file manager."""
        d = self._converted_mongo_dir()
        try:
            d.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        self._open_in_filemanager(d)

    # ── Deferred tasks (sent from the Council tab) ──────────────
    def _vmgr_refresh_deferred(self):
        """Reload the pending deferred tasks into the Vault-tab table."""
        tree = getattr(self, "_defer_tree", None)
        if tree is None:
            return
        try:
            import deferred_tasks as _dt
            pend = _dt.DeferredTaskStore(VAULT_DIR).pending()
        except Exception as exc:
            self._defer_status.set(f"Could not load tasks: {exc!r}")
            return
        tree.delete(*tree.get_children())
        self._defer_ids = {}
        _labels = {
            "bigger_summary": "Bigger summary",
            "deeper_stats": "Deeper stats",
            "tool_request": "Tool request",
            "other": "Other",
        }
        for t in pend:
            iid = tree.insert("", "end",
                              values=(_labels.get(t.kind, t.kind), t.label()))
            self._defer_ids[iid] = t.id
        self._defer_status.set(
            f"{len(pend)} pending task(s)." if pend
            else "No pending tasks. Send some from the Council tab (⤓ Defer to Vault).")

    def _vmgr_selected_deferred(self):
        sel = self._defer_tree.selection()
        if not sel:
            self._defer_status.set("Select a task first.")
            return None
        return self._defer_ids.get(sel[0])

    def _vmgr_set_deferred(self, status: str):
        tid = self._vmgr_selected_deferred()
        if not tid:
            return
        try:
            import deferred_tasks as _dt
            store = _dt.DeferredTaskStore(VAULT_DIR)
            if status == "done":
                store.mark_done(tid)
            else:
                store.dismiss(tid)
        except Exception as exc:
            self._defer_status.set(f"Update failed: {exc!r}")
            return
        self._vmgr_refresh_deferred()

    def _vmgr_run_deferred(self):
        """Run a runnable deferred task (bigger summary / deeper stats) with
        the full deterministic tooling, write the result under
        data_in/.deferred_results/, and mark it done. Tool-requests / other
        kinds aren't auto-runnable."""
        import threading as _th
        tid = self._vmgr_selected_deferred()
        if not tid:
            return
        try:
            import deferred_tasks as _dt
            store = _dt.DeferredTaskStore(VAULT_DIR)
            task = store.get(tid)
        except Exception as exc:
            self._defer_status.set(f"Load failed: {exc!r}")
            return
        if task is None:
            self._defer_status.set("Task not found (refresh).")
            return
        if task.kind not in _dt.RUNNABLE_KINDS:
            self._defer_status.set(
                "This is a tool request / note — it's logged for the developer, "
                "not auto-runnable. Use ✓ Done when handled.")
            return
        self._defer_status.set("Running…")

        def _worker():
            try:
                import vault_analyst as _va
                in_dir = data_index.input_dir(VAULT_DIR)
                target = in_dir
                if task.folder:
                    cand = in_dir / task.folder
                    if cand.exists():
                        target = cand
                # NON-hidden folder on purpose: a ".deferred_results" dot-dir
                # is skipped by the vault index/analyst, so the council could
                # never find the saved result. This way, re-asking the same
                # question surfaces it (see the precomputed-answer route in
                # _run_analyst_step_impl).
                # Computed outputs live in data_in/derived/ (searchable, but
                # excluded from the source-data census) and are catalogued in
                # the DerivedStore with their source fingerprint for
                # staleness-safe reuse.
                import derived_results as _drv
                out_dir = _drv.derived_dir(VAULT_DIR)
                # Human-readable filename derived from the task itself, not the
                # opaque internal id. e.g. a "bigger summary of sales.csv" task
                # over folder Q3 -> "summary__Q3__bigger_summary_of_sales__a1b2.csv".
                import re as _re
                _desc = (task.question or task.folder or "deferred").strip().lower()
                _slug = _re.sub(r"[^a-z0-9]+", "_", _desc).strip("_")[:48] or "deferred"
                _fold = (_re.sub(r"[^A-Za-z0-9]+", "_", task.folder).strip("_")
                         if task.folder else "")
                _short = task.id[-4:]      # keep runs unique without being noisy
                _kindword = ("summary" if task.kind == _dt.KIND_BIGGER_SUMMARY
                             else "stats")
                _name = "__".join(p for p in (_kindword, _fold, _slug, _short) if p)
                op = out_dir / f"{_name}.csv"
                # SCOPE to the files the task is actually about. The capture
                # dialog resolves filenames from the question into task.files;
                # without this the run summarised the WHOLE folder regardless
                # of which file the user asked about ("wrong files").
                import pandas as _pd_run
                want = {f.lower() for f in (task.files or []) if f}
                named_paths = []
                if want:
                    # list_data_files = CSV ∪ Excel (superset of list_csv_files),
                    # so a CSV-only re-walk fallback was logically dead — a name
                    # that didn't match here can't match the CSV subset.
                    for _p in _va.list_data_files([target]):
                        if _p.name.lower() in want:
                            named_paths.append(_p)

                if task.kind == _dt.KIND_BIGGER_SUMMARY:
                    if named_paths:
                        # Per-COLUMN profile of each named file — a genuinely
                        # "bigger" summary than chat gave, on the right files.
                        frames = [_va.summarize_csv(p) for p in named_paths]
                        df = (_pd_run.concat(frames, ignore_index=True)
                              if frames else _va.folder_data_summary([target]))
                        summary = (f"profiled {len(named_paths)} named file(s): "
                                   + ", ".join(p.name for p in named_paths[:5]))
                    else:
                        df = _va.folder_data_summary([target])
                        summary = (f"{len(df)} file(s) profiled"
                                   + (" (named file(s) not found — used the "
                                      "whole folder)" if want else ""))
                else:   # deeper_stats
                    df = _va.folder_column_stats(VAULT_DIR, [target])
                    if want and "file" in df.columns:
                        sub = df[df["file"].str.lower().isin(want)]
                        if not sub.empty:
                            df = sub.reset_index(drop=True)
                    nfiles = int(df["file"].nunique()) if "file" in df else 0
                    summary = f"stats for {nfiles} file(s)"
                df.to_csv(op, index=False)
                store.mark_done(task.id, result_path=str(op),
                                result_summary=summary)
                # Catalogue the computed output with its SOURCE FINGERPRINT so
                # a future re-ask reuses it ONLY while the sources are unchanged
                # (staleness-safe). The precomputed-answer route in
                # _run_analyst_step_impl reads this via DerivedStore.find_fresh.
                try:
                    _srcs = ([str(p) for p in named_paths] if named_paths
                             else [str(target)])
                    _drv.DerivedStore(VAULT_DIR).record(
                        label=(task.question or task.label() or _name),
                        output=str(op), sources=_srcs,
                        operation=task.kind,
                        columns=[str(c) for c in df.columns],
                        rows=int(len(df)))
                except Exception:
                    pass
                # Refresh FIRST (it resets the status line), THEN show the
                # completion message so it isn't immediately overwritten.
                self.after(0, lambda: (
                    self._vmgr_refresh_deferred(),
                    self._defer_status.set(
                        f"Done — {summary} → {op.name} (in "
                        "data_in/derived/). Re-ask in the Council tab "
                        "to use it.")))
            except Exception as exc:
                self.after(0, lambda: self._defer_status.set(
                    f"Run failed: {exc!r}"))

        _th.Thread(target=_worker, daemon=True).start()

    # ── Collections (virtual projects over the vault) ───────────
    def _vmgr_refresh_collections(self):
        tree = getattr(self, "_coll_tree", None)
        if tree is None:
            return
        try:
            import vault_collections as _vc
            cols = _vc.CollectionStore(VAULT_DIR).all()
        except Exception as exc:
            self._coll_status.set(f"Could not load collections: {exc!r}")
            return
        tree.delete(*tree.get_children())
        self._coll_names = {}
        for c in sorted(cols, key=lambda c: c.name.lower()):
            iid = tree.insert("", "end", values=(c.name, len(c.files)))
            self._coll_names[iid] = c.name
        self._coll_status.set(
            f"{len(cols)} collection(s)." if cols else
            "No collections yet. ➕ New… groups files into a project.")

    def _vmgr_selected_collection(self):
        sel = self._coll_tree.selection()
        if not sel:
            self._coll_status.set("Select a collection first.")
            return None
        return self._coll_names.get(sel[0])

    def _vmgr_delete_collection(self):
        from tkinter import messagebox
        name = self._vmgr_selected_collection()
        if not name:
            return
        if not messagebox.askyesno(
                "Delete collection",
                f"Delete the collection “{name}”?\n(The files themselves are "
                "NOT touched — this only removes the grouping.)"):
            return
        try:
            import vault_collections as _vc
            _vc.CollectionStore(VAULT_DIR).delete(name)
        except Exception as exc:
            self._coll_status.set(f"Delete failed: {exc!r}")
            return
        self._vmgr_refresh_collections()

    def _vmgr_new_collection(self, edit: bool = False):
        """Dialog: name a project, let the council DISCOVER candidate members
        (name/value/shared-key signals), confirm/edit the set, and save it."""
        import tkinter as tk
        from tkinter import ttk, messagebox
        import vault_collections as _vc
        store = _vc.CollectionStore(VAULT_DIR)
        existing = None
        if edit:
            nm = self._vmgr_selected_collection()
            if not nm:
                return
            existing = store.get(nm)

        win = tk.Toplevel(self)
        win.title(f"Edit collection: {existing.name}" if existing
                  else "New collection")
        win.transient(self)
        win.resizable(False, False)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, foreground="#888", justify="left", wraplength=470,
                  text="Name the project, then Discover to let the council "
                       "propose members (by filename, by value match, and by "
                       "shared join keys). Remove wrong ones, add missing "
                       "files, then Save.").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Label(frm, text="Name:").grid(row=1, column=0, sticky="w", pady=3)
        name_var = tk.StringVar(value=existing.name if existing else "")
        ttk.Entry(frm, textvariable=name_var, width=32).grid(
            row=1, column=1, sticky="w")
        ttk.Button(frm, text="🔎 Discover", command=lambda: _discover()).grid(
            row=1, column=2, sticky="w", padx=6)

        ttk.Label(frm, text="Files:").grid(row=2, column=0, sticky="nw", pady=3)
        ff = ttk.Frame(frm)
        ff.grid(row=2, column=1, columnspan=2, sticky="w")
        lb = tk.Listbox(ff, height=8, width=46, selectmode="extended",
                        exportselection=False)
        lb.grid(row=0, column=0, rowspan=3, sticky="nw")
        reasons: dict = {}

        avail = []
        try:
            import os as _os_av
            _ind = data_index.input_dir(VAULT_DIR)
            for _dp, _dn, _fn in _os_av.walk(str(_ind)):
                _dn[:] = [d for d in _dn if not d.startswith(".")]
                for _f in _fn:
                    if _f.startswith("."):
                        continue
                    avail.append(str(Path(_dp, _f).relative_to(_ind)
                                     ).replace("\\", "/"))
            avail = sorted(set(avail))
        except Exception:
            avail = []
        add_var = tk.StringVar()
        addcb = ttk.Combobox(ff, textvariable=add_var, values=avail,
                             width=30, state="normal")
        addcb.grid(row=0, column=1, sticky="w", padx=(6, 0))

        def _files_now():
            return list(lb.get(0, "end"))

        def _add(_e=None):
            v = add_var.get().strip()
            if v and v not in _files_now():
                lb.insert("end", v)
            add_var.set("")
        addcb.bind("<Return>", _add)
        ttk.Button(ff, text="➕ Add", width=10, command=_add).grid(
            row=1, column=1, sticky="w", padx=(6, 0))

        def _rm():
            for i in reversed(lb.curselection()):
                lb.delete(i)
        ttk.Button(ff, text="✗ Remove", width=10, command=_rm).grid(
            row=2, column=1, sticky="nw", padx=(6, 0))

        detail = tk.StringVar(value="")
        ttk.Label(frm, textvariable=detail, foreground="#9a9a9a",
                  wraplength=470, justify="left").grid(
            row=3, column=1, columnspan=2, sticky="w")

        def _on_sel(_e=None):
            sel = lb.curselection()
            if sel:
                rel = lb.get(sel[0])
                detail.set(f"{rel} — {reasons.get(rel, 'added manually')}")
        lb.bind("<<ListboxSelect>>", _on_sel)

        if existing:
            for f in existing.files:
                lb.insert("end", f)

        def _discover():
            nm = name_var.get().strip()
            if not nm:
                messagebox.showwarning("Name first",
                                       "Enter a collection name to discover by.")
                return
            detail.set("Discovering…")

            def _w():
                try:
                    idx = getattr(self, "data_index", None)
                    props = _vc.propose_members(VAULT_DIR, nm, index=idx)
                except Exception as exc:
                    self.after(0, lambda: detail.set(f"Discover failed: {exc!r}"))
                    return

                def _apply():
                    have = set(_files_now())
                    added = 0
                    for rel, sc, rs in props:
                        reasons[rel] = ", ".join(rs)
                        if rel not in have:
                            lb.insert("end", rel)
                            have.add(rel)
                            added += 1
                    detail.set(f"Proposed {len(props)} file(s) ({added} new) — "
                               "select a row to see why; remove/add, then Save.")
                self.after(0, _apply)
            import threading as _th
            _th.Thread(target=_w, daemon=True).start()

        def _save():
            nm = name_var.get().strip()
            if not nm:
                messagebox.showwarning("Name first", "Enter a collection name.")
                return
            files = _files_now()
            if existing and existing.name != nm:
                store.rename(existing.name, nm)
            store.upsert(nm, files)
            self._coll_status.set(
                f"Saved “{nm}” — {len(files)} file(s). Ask the council: "
                f"“show me {nm}”.")
            self._vmgr_refresh_collections()
            win.destroy()

        btns = ttk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=3, sticky="e", pady=(12, 0))
        ttk.Button(btns, text="Save", command=_save).pack(side="right")
        ttk.Button(btns, text="Cancel", command=win.destroy).pack(
            side="right", padx=6)

    def _vmgr_summarize_collection(self):
        """Profile every file in the selected collection (per-column) and
        write the result to data_in/derived/collection__<name>.csv, recording
        it in the DerivedStore so a re-ask reuses it while sources are fresh."""
        import threading as _th
        name = self._vmgr_selected_collection()
        if not name:
            return
        self._coll_status.set("Summarizing…")

        def _w():
            try:
                import vault_collections as _vc
                import vault_analyst as _va
                import pandas as _pd
                import re as _re
                paths = _vc.CollectionStore(VAULT_DIR).abs_paths(name)
                if not paths:
                    self.after(0, lambda: self._coll_status.set(
                        "No existing files in that collection."))
                    return
                frames = []
                for p in paths:
                    try:
                        frames.append(_va.summarize_csv(p))
                    except Exception:
                        pass
                df = (_pd.concat(frames, ignore_index=True)
                      if frames else _pd.DataFrame())
                import derived_results as _drv
                out = _drv.derived_dir(VAULT_DIR)
                slug = _re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_") \
                    or "collection"
                op = out / f"collection__{slug}.csv"
                df.to_csv(op, index=False)
                try:
                    _drv.DerivedStore(VAULT_DIR).record(
                        label=f"summary of the {name} collection",
                        output=str(op), sources=[str(p) for p in paths],
                        operation="collection_summary",
                        columns=[str(c) for c in df.columns],
                        rows=int(len(df)))
                except Exception:
                    pass
                self.after(0, lambda: self._coll_status.set(
                    f"Summarized {len(paths)} file(s) → "
                    f"data_in/derived/{op.name}"))
            except Exception as exc:
                self.after(0, lambda: self._coll_status.set(
                    f"Summarize failed: {exc!r}"))
        _th.Thread(target=_w, daemon=True).start()

    def _vmgr_convert_mongo(self, scan_all: bool = False):
        """Convert a Mongo .bson/.json/.jsonl file (or every such file in the
        vault) into model-digestible artefacts written under
        data_in/converted_mongo/. The source is only ever read.

        Outputs (per selected checkbox):
          <stem>_clean.csv    flattened, all-scalar table (one row per doc)
          <stem>_schema.csv   field / types / presence% / example
          <stem>_digest.txt   compact key:value text digest
        """
        import threading as _th
        sel = self._vmgr_mongo_var.get().strip()
        want_csv = self._vmgr_mongo_csv.get()
        want_schema = self._vmgr_mongo_schema.get()
        want_text = self._vmgr_mongo_text.get()
        if not (want_csv or want_schema or want_text):
            self._mongo_status_var.set(
                "Pick at least one output (Clean CSV / Schema / Text digest).")
            return
        if not scan_all and not sel:
            self._mongo_status_var.set(
                "Select a .bson/.json/.jsonl file, or use “Convert ALL”.")
            return
        self._mongo_status_var.set("Converting…")

        def _worker():
            try:
                import vault_analyst as _va
                out_root = self._converted_mongo_dir()
                out_root.mkdir(parents=True, exist_ok=True)

                if scan_all:
                    in_dir = data_index.input_dir(VAULT_DIR)
                    files = []
                    for ext in ("*.bson", "*.json", "*.jsonl"):
                        files += list(in_dir.rglob(ext))
                    # Never re-convert our own outputs.
                    files = [f for f in files
                             if out_root not in f.parents and f.parent != out_root]
                else:
                    files = [Path(sel)]

                if not files:
                    self.after(0, lambda: self._mongo_status_var.set(
                        "No .bson/.json/.jsonl files found in the vault."))
                    return

                done = ok = total_rows = 0
                last_err = ""
                for fp in files:
                    try:
                        # Streaming, bounded-memory conversion — never loads
                        # the whole dump (that OOM-crashed the app on Linux).
                        summary = _va.convert_mongo_file(
                            fp, out_root,
                            want_csv=want_csv, want_schema=want_schema,
                            want_text=want_text)
                        if summary.get("docs"):
                            total_rows += summary.get("rows", 0)
                            ok += 1
                    except Exception as fe:
                        last_err = f"{fp.name}: {fe!r}"
                    done += 1
                    self.after(0, lambda d=done, t=len(files), n=fp.name:
                               self._mongo_status_var.set(
                                   f"Converting {d}/{t} — {n[:40]}"))

                tail = f"  (last error — {last_err})" if last_err else ""
                self.after(0, lambda: self._mongo_status_var.set(
                    f"Done — {ok}/{len(files)} file(s), {total_rows} rows → "
                    f"data_in/converted_mongo/.{tail}  "
                    "Run “1. Build Keyword Index” to make it searchable."))
                # Refresh the tree so the new files show up.
                self.after(0, self._vmgr_refresh_tree)
            except Exception as exc:
                self.after(0, lambda exc=exc: self._mongo_status_var.set(
                    f"Convert failed: {exc!r}"))

        _th.Thread(target=_worker, daemon=True).start()

    def _vmgr_import_zip(self):
        """Extract a zip file into a vault subfolder, keeping only indexable files."""
        import threading
        zip_path  = self._vmgr_zip_var.get().strip()
        subfolder = self._vmgr_zip_subfolder_var.get().strip()

        if not zip_path:
            self._vmgr_append("✗ Please select a zip file first.", "err")
            return
        if not Path(zip_path).exists():
            self._vmgr_append(f"✗ File not found: {zip_path}", "err")
            return

        if not subfolder:
            subfolder = Path(zip_path).stem

        self._vmgr_append(f"Extracting {Path(zip_path).name} → vault/{subfolder} …", "info")

        def worker():
            try:
                dest, copied, skipped = _vmgr_extract_zip(
                    Path(zip_path),
                    vault_dir=VAULT_DIR,
                    subfolder=subfolder,
                    log_cb=lambda m: self.ui_q.put(("vault_mgr_log", m)),
                )
                self.ui_q.put(("vault_mgr_log",
                    f"✓ Extracted {copied} files → vault/{dest.name}  ({skipped} skipped)"))
                self.ui_q.put(("vault_mgr_refresh", None))
                # Clear fields
                self._vmgr_zip_var.set("")
                self._vmgr_zip_subfolder_var.set("")
            except Exception as e:
                self.ui_q.put(("vault_mgr_log", f"✗ Extraction failed: {e}"))

        threading.Thread(target=worker, daemon=True).start()

    def _vmgr_browse_zip_folder(self):
        """Pick a folder that CONTAINS .zip files to batch-extract."""
        from tkinter import filedialog
        path = filedialog.askdirectory(
            title="Select a folder containing .zip files")
        if path:
            self._vmgr_zipdir_var.set(path)

    def _vmgr_import_zip_folder(self):
        """Extract EVERY .zip found under the chosen folder (recursively) into
        the vault's data_in/ — each zip into its own subfolder named after the
        zip — so the imported files are immediately usable by the council. One
        bad/corrupt zip is logged and skipped; the rest still import."""
        import threading
        folder = self._vmgr_zipdir_var.get().strip()
        if not folder:
            self._vmgr_append("✗ Pick a folder of zip files first.", "err")
            return
        src = Path(folder)
        if not src.exists() or not src.is_dir():
            self._vmgr_append(f"✗ Folder not found: {folder}", "err")
            return

        def worker():
            import zipfile as _zf
            # Extract into data_in/ (the analyst/index scope) so the files are
            # usable by the council, not the vault root.
            try:
                _indir = data_index.input_dir(VAULT_DIR)
                _indir.mkdir(parents=True, exist_ok=True)
            except Exception:
                _indir = VAULT_DIR
            try:
                zips = sorted(src.rglob("*.zip"))
            except Exception as e:
                self.ui_q.put(("vault_mgr_log", f"✗ Could not scan folder: {e}"))
                return
            if not zips:
                self.ui_q.put(("vault_mgr_log",
                               f"No .zip files found under {src.name}."))
                return
            self.ui_q.put(("vault_mgr_log",
                           f"Found {len(zips)} zip(s) under {src.name} — extracting…"))
            ok = total_copied = failed = 0
            used: set = set()
            for i, zp in enumerate(zips, 1):
                # Validate BEFORE extracting so a corrupt/.zip-misnamed file
                # doesn't leave an empty subfolder behind.
                if not _zf.is_zipfile(zp):
                    failed += 1
                    self.ui_q.put((
                        "vault_mgr_log",
                        f"  [{i}/{len(zips)}] ✗ {zp.name}: not a valid zip — skipped"))
                    continue
                # Unique subfolder per zip (two zips can share a stem).
                sub = zp.stem
                while sub in used:
                    sub = f"{zp.stem}_{i}"
                used.add(sub)
                try:
                    dest, copied, skipped = _vmgr_extract_zip(
                        zp, vault_dir=_indir, subfolder=sub,
                        log_cb=lambda m: self.ui_q.put(("vault_mgr_log", m)))
                    ok += 1
                    total_copied += copied
                    self.ui_q.put((
                        "vault_mgr_log",
                        f"  [{i}/{len(zips)}] ✓ {zp.name} → data_in/{dest.name} "
                        f"({copied} files, {skipped} skipped)"))
                except Exception as e:
                    failed += 1
                    self.ui_q.put((
                        "vault_mgr_log",
                        f"  [{i}/{len(zips)}] ✗ {zp.name}: {e}"))
            self.ui_q.put((
                "vault_mgr_log",
                f"Done — {ok}/{len(zips)} zip(s) extracted, {total_copied} "
                f"files total" + (f", {failed} failed" if failed else "") + "."))
            self.ui_q.put(("vault_mgr_refresh", None))
            self._vmgr_zipdir_var.set("")

        threading.Thread(target=worker, daemon=True).start()

    def _vmgr_import_folder(self):
        """Copy a local folder into the vault, keeping only indexable files."""
        import threading
        folder_path = self._vmgr_folder_var.get().strip()

        if not folder_path:
            self._vmgr_append("✗ Please select a folder first.", "err")
            return
        src = Path(folder_path)
        if not src.exists() or not src.is_dir():
            self._vmgr_append(f"✗ Folder not found: {folder_path}", "err")
            return

        subfolder = src.name
        self._vmgr_append(f"Copying {src.name} → vault/{subfolder} …", "info")

        def worker():
            try:
                dest, copied, skipped = _vmgr_copy_folder(
                    src,
                    vault_dir=VAULT_DIR,
                    subfolder=subfolder,
                    log_cb=lambda m: self.ui_q.put(("vault_mgr_log", m)),
                )
                self.ui_q.put(("vault_mgr_log",
                    f"✓ Copied {copied} files → vault/{dest.name}  ({skipped} skipped)"))
                self.ui_q.put(("vault_mgr_refresh", None))
                self._vmgr_folder_var.set("")
            except Exception as e:
                self.ui_q.put(("vault_mgr_log", f"✗ Copy failed: {e}"))

        threading.Thread(target=worker, daemon=True).start()

    # ---- Council Lens tab (#7) ----

    def _build_lens_tab(self):
        self.tab_lens = ttk.Frame(self.nb)
        self.nb.add(self.tab_lens, text="🔍 Lens")

        hdr = ttk.Frame(self.tab_lens)
        hdr.pack(fill="x", padx=10, pady=(8, 4))
        ttk.Label(hdr, text="Council Lens",
                  foreground="#d32f2f", font=("", 11, "bold")).pack(side="left")
        ttk.Label(hdr,
                  text="  Paste any content — get simultaneous parallel critique from all relevant roles",
                  foreground="#6c7086").pack(side="left")

        # Role selector
        roles_frame = ttk.LabelFrame(self.tab_lens, text="Roles to include")
        roles_frame.pack(fill="x", padx=10, pady=(0, 6))
        self._lens_role_vars: dict = {}
        _lens_defaults = {
            "writer": True, "coder": True, "sage": True,
            "peasant": True, "strategist": True, "director": True,
            "artist": False, "intern": False, "skeptic": False,
            "content": True, "musician": False,
        }
        _rf_row = ttk.Frame(roles_frame)
        _rf_row.pack(fill="x", padx=8, pady=4)
        for _rname, _default in _lens_defaults.items():
            v = tk.BooleanVar(value=_default)
            self._lens_role_vars[_rname] = v
            ttk.Checkbutton(_rf_row, text=_rname.capitalize(), variable=v).pack(side="left", padx=4)

        # Input area
        in_frame = ttk.LabelFrame(self.tab_lens, text="Content to review")
        in_frame.pack(fill="x", padx=10, pady=(0, 6))
        self._lens_input = self._make_text(in_frame, wrap="word", height=8)
        self._lens_input.pack(fill="both", expand=True, padx=6, pady=6)

        # Controls
        ctrl_row = ttk.Frame(self.tab_lens)
        ctrl_row.pack(fill="x", padx=10, pady=(0, 4))
        ttk.Button(ctrl_row, text="▶ Run Lens",  command=self._lens_run).pack(side="left")
        ttk.Button(ctrl_row, text="Clear All",   command=self._lens_clear).pack(side="left", padx=6)
        self._lens_status = ttk.Label(ctrl_row, text="", foreground="#6c7086")
        self._lens_status.pack(side="left", padx=6)

        # Output area
        out_frame = ttk.LabelFrame(self.tab_lens, text="Role critiques")
        out_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self._lens_output = self._make_text(out_frame, wrap="word", state="disabled")
        self._lens_output.pack(fill="both", expand=True, padx=6, pady=6)

    def _lens_clear(self):
        self._set_text(self._lens_input, "")
        self._lens_output.configure(state="normal")
        self._lens_output.delete("1.0", "end")
        self._lens_output.configure(state="disabled")

    def _lens_run(self):
        content = self._lens_input.get("1.0", "end").strip()
        if not content:
            return
        selected_roles = [r for r, v in self._lens_role_vars.items() if v.get()]
        if not selected_roles:
            messagebox.showinfo("Lens", "Select at least one role.")
            return

        self._lens_output.configure(state="normal")
        self._lens_output.delete("1.0", "end")
        self._lens_output.configure(state="disabled")
        self._lens_status.configure(text=f"Running {len(selected_roles)} roles in parallel…")

        def worker():
            from concurrent.futures import ThreadPoolExecutor, as_completed
            results: dict = {}
            prompt = (
                "Review the following content from your specific lens.\n"
                "Give your honest, role-specific critique — 150-250 words.\n"
                "Do NOT synthesise or defer to other roles.\n"
                "Lead with what you specifically notice, good or bad.\n\n"
                f"CONTENT:\n{content[:3000]}"
            )

            def run_role(role_name: str) -> tuple:
                model = getattr(self, role_name, None)
                if model is None:
                    return role_name, "(Role not loaded)"
                try:
                    return role_name, model.respond(prompt, max_tokens=350)
                except Exception as e:
                    return role_name, f"(Error: {e})"

            with ThreadPoolExecutor(max_workers=4) as pool:
                futures = {pool.submit(run_role, r): r for r in selected_roles}
                for fut in as_completed(futures):
                    role_name, response = fut.result()
                    self.ui_q.put(("lens_result", role_name, response))

            self.ui_q.put(("lens_done", len(selected_roles)))

        import threading as _t
        _t.Thread(target=worker, daemon=True).start()

    # ---- Vault Health Dashboard tab (#11) ----

    def _build_vault_health_tab(self):
        self.tab_vault_health = ttk.Frame(self.nb)
        self.nb.add(self.tab_vault_health, text="🗄 Vault Health")

        hdr = ttk.Frame(self.tab_vault_health)
        hdr.pack(fill="x", padx=10, pady=(8, 4))
        ttk.Label(hdr, text="Vault Health Dashboard",
                  foreground="#d32f2f", font=("", 11, "bold")).pack(side="left")

        ctrl_row = ttk.Frame(self.tab_vault_health)
        ctrl_row.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Button(ctrl_row, text="↺ Refresh",      command=self._vault_health_refresh).pack(side="left")
        ttk.Button(ctrl_row, text="📂 Open Vault",   command=self._lib_open_vault).pack(side="left", padx=6)
        ttk.Button(ctrl_row, text="📋 Open Wishlist", command=self._vault_health_open_wishlist).pack(side="left")

        # ── Three-panel layout ─────────────────────────────────────
        pw = tk.PanedWindow(self.tab_vault_health, orient="horizontal",
                            bg="#1a1414", sashwidth=5, sashrelief="flat")
        pw.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Left: memory files per personality
        left = ttk.LabelFrame(pw, text="Personality Memory Files")
        pw.add(left, minsize=200)
        cols_mem = ("role", "size", "updated")
        self._vh_mem_tree = ttk.Treeview(left, columns=cols_mem, show="headings", height=14)
        self._vh_mem_tree.heading("role",    text="Role")
        self._vh_mem_tree.heading("size",    text="Size")
        self._vh_mem_tree.heading("updated", text="Last Updated")
        self._vh_mem_tree.column("role",    width=100)
        self._vh_mem_tree.column("size",    width=70)
        self._vh_mem_tree.column("updated", width=130)
        self._vh_mem_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # Middle: vault files
        mid = ttk.LabelFrame(pw, text="Vault Files")
        pw.add(mid, minsize=240)
        cols_vf = ("name", "size", "modified")
        self._vh_vault_tree = ttk.Treeview(mid, columns=cols_vf, show="headings", height=14)
        self._vh_vault_tree.heading("name",     text="File")
        self._vh_vault_tree.heading("size",     text="Size")
        self._vh_vault_tree.heading("modified", text="Modified")
        self._vh_vault_tree.column("name",     width=180)
        self._vh_vault_tree.column("size",     width=70)
        self._vh_vault_tree.column("modified", width=130)
        self._vh_vault_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # Right: wishlist stats + project context summary
        right = ttk.LabelFrame(pw, text="Wishlist & Project Context")
        pw.add(right, minsize=200)
        self._vh_stats_box = self._make_text(right, wrap="word", state="disabled")
        self._vh_stats_box.pack(fill="both", expand=True, padx=4, pady=4)

        self._vault_health_refresh()

    def _vault_health_refresh(self):
        """Populate all three panels of the vault health dashboard."""
        import os as _os

        # ── Memory files ──────────────────────────────────────────
        self._vh_mem_tree.delete(*self._vh_mem_tree.get_children())
        all_roles = ("judge", "writer", "coder", "intern", "peasant", "artist",
                     "sage", "strategist", "librarian", "musician", "content", "director",
                     "eye", "cutter", "algorithm",
                     "_project")
        try:
            memmgr = self.writer.memory_manager
            for role in all_roles:
                p = memmgr.path_for(role)
                if p.exists():
                    sz = p.stat().st_size
                    mtime = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                    label = role if role != "_project" else "⬡ project"
                    self._vh_mem_tree.insert("", "end",
                        values=(label, _fmt_bytes(sz), mtime))
        except Exception:
            pass

        # ── Vault files ───────────────────────────────────────────
        self._vh_vault_tree.delete(*self._vh_vault_tree.get_children())
        try:
            _skip_dirs = {"conversations", "memory", ".git"}
            for item in sorted(VAULT_DIR.iterdir()):
                if item.name.startswith(".") or item.name in _skip_dirs:
                    continue
                if item.is_file():
                    sz = item.stat().st_size
                    mtime = datetime.fromtimestamp(item.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                    self._vh_vault_tree.insert("", "end",
                        values=(item.name, _fmt_bytes(sz), mtime))
        except Exception:
            pass

        # ── Wishlist + project context stats ──────────────────────
        self._vh_stats_box.configure(state="normal")
        self._vh_stats_box.delete("1.0", "end")
        try:
            wl_raw = self.librarian.get_wishlist()
            total   = sum(1 for ln in wl_raw.splitlines() if "- [" in ln)
            pending = sum(1 for ln in wl_raw.splitlines() if "- [ ]" in ln)
            filled  = sum(1 for ln in wl_raw.splitlines() if "- [x]" in ln)
            stats = (
                f"WISHLIST\n"
                f"  Total items : {total}\n"
                f"  Pending     : {pending}\n"
                f"  Filled      : {filled}\n\n"
            )
            # Project context
            memmgr = self.writer.memory_manager
            proj_path = memmgr.path_for("_project")
            if proj_path.exists():
                proj_sz = proj_path.stat().st_size
                proj_lines = len(proj_path.read_text(encoding="utf-8").splitlines())
                stats += f"PROJECT CONTEXT\n  {proj_lines} lines / {_fmt_bytes(proj_sz)}\n\n"
            else:
                stats += "PROJECT CONTEXT\n  (not yet generated)\n\n"
            # Trends
            trends_path = VAULT_DIR / "trends.md"
            if trends_path.exists():
                mtime = datetime.fromtimestamp(trends_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                stats += f"TRENDS FILE\n  Last updated: {mtime}\n"
            else:
                stats += "TRENDS FILE\n  (not yet generated)\n"
            self._vh_stats_box.insert("1.0", stats)
        except Exception as e:
            self._vh_stats_box.insert("1.0", f"(Error loading stats: {e})")
        self._vh_stats_box.configure(state="disabled")

    def _vault_health_open_wishlist(self):
        """Open the wishlist file in the default text editor."""
        try:
            wl_path = self.librarian.wishlist_path
            if not wl_path.exists():
                messagebox.showinfo("Wishlist", "No wishlist file yet.")
                return
            if sys.platform.startswith("win"):
                os.startfile(str(wl_path))  # type: ignore
            elif sys.platform == "darwin":
                subprocess.run(["open", str(wl_path)], check=False)
            else:
                subprocess.run(["xdg-open", str(wl_path)], check=False)
        except Exception as e:
            messagebox.showerror("Wishlist", str(e))

    # ---- Speech tab ----

    # ---- Changelog tab ----
    # Reads `git log` and `git show` from the repo root so users can see
    # exactly what shipped between launches. Runs git as a subprocess so a
    # broken Git install is just a status message, not a tab failure.

    # ============================
    # Diagnostics tab — system + optional-dependency status
    # ============================
    # Surfaces every optional dependency the app might use plus the
    # core environment (Python version, n_ctx, physical cores). Lets a
    # user see at a glance whether speech / embeddings / PDF / etc. are
    # available, and exactly what pip command would install each
    # missing feature. The actual install is deliberately the user's
    # job — running pip from inside a packaged .exe is dangerous (it
    # might write to the wrong python).

    def _build_diagnostics_tab(self):
        self.tab_diagnostics = ttk.Frame(self.nb)
        self.nb.add(self.tab_diagnostics, text="🔧 Diagnostics")

        # Top bar — refresh + status summary
        bar = ttk.Frame(self.tab_diagnostics)
        bar.pack(fill="x", padx=6, pady=(6, 4))
        ttk.Button(bar, text="⟳ Re-check",
                   command=self._diagnostics_refresh).pack(side="left")
        ttk.Button(bar, text="📋 Copy report",
                   command=self._diagnostics_copy_to_clipboard
                   ).pack(side="left", padx=4)
        self._diagnostics_status_var = tk.StringVar(value="")
        ttk.Label(bar, textvariable=self._diagnostics_status_var,
                  foreground="#9a9a9a").pack(side="right")

        # Scrollable text widget for the full report
        text_frame = ttk.Frame(self.tab_diagnostics)
        text_frame.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        vsb = ttk.Scrollbar(text_frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        self._diagnostics_text = tk.Text(
            text_frame, wrap="word",
            font=("Consolas", 10),
            yscrollcommand=vsb.set,
            bg="#1a1414", fg="#e6e6e6",
            insertbackground="#e6e6e6",
        )
        self._diagnostics_text.pack(fill="both", expand=True)
        vsb.config(command=self._diagnostics_text.yview)

        # Colour tags for the report — green for available, red for
        # missing, dim for the descriptive lines.
        self._diagnostics_text.tag_config("ok",       foreground="#a6e3a1")
        self._diagnostics_text.tag_config("missing",  foreground="#f38ba8")
        self._diagnostics_text.tag_config("dim",      foreground="#9a9a9a")
        self._diagnostics_text.tag_config("install",  foreground="#89b4fa")
        self._diagnostics_text.tag_config("section",
                                          font=("Consolas", 10, "bold"))

        # Render on first show
        self._diagnostics_refresh()

    def _diagnostics_refresh(self):
        """Re-run the dependency check and repaint the report."""
        try:
            import dependency_check as _dc
        except Exception as exc:
            self._diagnostics_text.configure(state="normal")
            self._diagnostics_text.delete("1.0", "end")
            self._diagnostics_text.insert("1.0",
                f"Could not load dependency_check module: {exc!r}")
            self._diagnostics_text.configure(state="disabled")
            return

        statuses = _dc.check_all()
        ok_count = sum(1 for s in statuses if s.ok)
        missing  = [s for s in statuses if not s.ok]

        self._diagnostics_text.configure(state="normal")
        self._diagnostics_text.delete("1.0", "end")

        # System summary
        self._diagnostics_text.insert("end", "System summary\n", "section")
        self._diagnostics_text.insert("end", "─" * 60 + "\n", "dim")
        for line in _dc.system_summary():
            self._diagnostics_text.insert("end", f"  {line}\n")
        self._diagnostics_text.insert("end", "\n")

        # Missing features grouped by impact — most actionable info up top
        if missing:
            self._diagnostics_text.insert(
                "end",
                f"Missing optional dependencies ({len(missing)})\n",
                "section",
            )
            self._diagnostics_text.insert("end", "─" * 60 + "\n", "dim")
            by_impact = {"high": [], "med": [], "low": []}
            for s in missing:
                by_impact.setdefault(s.spec.impact, []).append(s)
            for impact_key, label in (("high", "High impact"),
                                       ("med",  "Medium impact"),
                                       ("low",  "Low impact")):
                bucket = by_impact.get(impact_key, [])
                if not bucket:
                    continue
                self._diagnostics_text.insert("end", f"\n  {label}\n", "section")
                for s in bucket:
                    self._diagnostics_text.insert("end", f"    ✗ ", "missing")
                    self._diagnostics_text.insert("end", f"{s.spec.name}\n")
                    self._diagnostics_text.insert("end",
                        f"        {s.spec.description}\n", "dim")
                    self._diagnostics_text.insert("end",
                        f"        Missing: {', '.join(s.missing)}\n", "dim")
                    self._diagnostics_text.insert("end", "        Install: ", "dim")
                    self._diagnostics_text.insert("end",
                        f"{s.spec.install}\n", "install")
            self._diagnostics_text.insert("end", "\n")
        else:
            self._diagnostics_text.insert("end",
                "All optional dependencies are installed.\n\n", "ok")

        # Available features
        available = [s for s in statuses if s.ok]
        if available:
            self._diagnostics_text.insert(
                "end",
                f"Available optional features ({ok_count})\n",
                "section",
            )
            self._diagnostics_text.insert("end", "─" * 60 + "\n", "dim")
            for s in available:
                self._diagnostics_text.insert("end", "  ✓ ", "ok")
                self._diagnostics_text.insert("end", f"{s.spec.name}\n")
            self._diagnostics_text.insert("end", "\n")

        # Footer
        self._diagnostics_text.insert("end",
            "To install missing features, run their pip commands in the "
            "same Python environment that runs this app. Restart the "
            "app afterward to pick up the changes.\n", "dim")

        self._diagnostics_text.configure(state="disabled")
        self._diagnostics_status_var.set(
            f"✓ {ok_count} available  ·  ✗ {len(missing)} missing"
        )

    def _diagnostics_copy_to_clipboard(self):
        """Put the full plain-text report on the clipboard so users can
        paste it into a bug report or share it on Discord/email."""
        try:
            import dependency_check as _dc
            report = _dc.render_as_text()
        except Exception as exc:
            report = f"dependency_check unavailable: {exc!r}"
        try:
            self.clipboard_clear()
            self.clipboard_append(report)
            self.update()
            self._diagnostics_status_var.set("✓ Copied to clipboard")
        except Exception as exc:
            self._diagnostics_status_var.set(f"copy failed: {exc!r}")

    def _build_changelog_tab(self):
        self.tab_changelog = ttk.Frame(self.nb)
        self.nb.add(self.tab_changelog, text="📜 Changelog")

        bar = ttk.Frame(self.tab_changelog)
        bar.pack(fill="x", padx=6, pady=(6, 4))
        ttk.Button(bar, text="⟳ Refresh",
                   command=self._changelog_refresh).pack(side="left")
        ttk.Button(bar, text="📂 Open repo folder",
                   command=lambda: self._open_in_explorer(_REPO_ROOT)).pack(side="left", padx=4)
        ttk.Label(bar, text="Filter:").pack(side="left", padx=(12, 4))
        self._changelog_filter_var = tk.StringVar()
        ent = ttk.Entry(bar, textvariable=self._changelog_filter_var, width=30)
        ent.pack(side="left")
        ent.bind("<KeyRelease>", lambda _e: self._changelog_apply_filter())
        self._changelog_status_var = tk.StringVar(
            value="Click ⟳ Refresh to load commits."
        )
        ttk.Label(bar, textvariable=self._changelog_status_var,
                  foreground="#9a9a9a").pack(side="right")

        pane = tk.PanedWindow(self.tab_changelog, orient="horizontal",
                              bg="#1a1414", sashwidth=6)
        pane.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # Left: scrollable commit list
        left = ttk.Frame(pane)
        pane.add(left, width=520)
        ttk.Label(left, text="Commits (newest first)").pack(anchor="w")
        list_row = ttk.Frame(left)
        list_row.pack(fill="both", expand=True)
        self._changelog_listbox = tk.Listbox(list_row, exportselection=False,
                                              font=("Consolas", 9))
        self._changelog_listbox.pack(side="left", fill="both", expand=True)
        lb_sb = ttk.Scrollbar(list_row, command=self._changelog_listbox.yview)
        self._changelog_listbox.configure(yscrollcommand=lb_sb.set)
        lb_sb.pack(side="left", fill="y")
        self._changelog_listbox.bind("<<ListboxSelect>>",
                                     lambda _e: self._changelog_show_selected())

        # Right: commit detail (subject + full message + file stat + diff)
        right = ttk.Frame(pane)
        pane.add(right)
        ttk.Label(right, text="Commit detail").pack(anchor="w")
        self._changelog_detail = self._make_text(
            right, wrap="word", state="disabled", font=("Consolas", 9),
        )
        det_sb = ttk.Scrollbar(right, command=self._changelog_detail.yview)
        self._changelog_detail.configure(yscrollcommand=det_sb.set)
        det_sb.pack(side="right", fill="y")
        self._changelog_detail.pack(fill="both", expand=True)

        # Internal caches
        self._changelog_all_commits = []   # [(full_hash, short, date, subject), ...]
        self._changelog_listbox_index = [] # full_hashes currently visible

        # Auto-load on first build so the tab isn't blank
        self.after(200, self._changelog_refresh)

    def _changelog_refresh(self):
        """Fetch the last 100 commits and populate the listbox."""
        import subprocess
        self._changelog_status_var.set("Loading…")
        try:
            result = subprocess.run(
                ["git", "log",
                 "--pretty=format:%H|%h|%ai|%s", "-100"],
                cwd=str(_REPO_ROOT),
                capture_output=True, text=True, timeout=15,
                creationflags=(subprocess.CREATE_NO_WINDOW
                               if sys.platform == "win32"
                               and hasattr(subprocess, "CREATE_NO_WINDOW") else 0),
            )
        except FileNotFoundError:
            self._changelog_status_var.set(
                "git is not on PATH. Install Git for Windows from git-scm.com."
            )
            return
        except Exception as exc:
            self._changelog_status_var.set(f"git log failed: {exc!r}")
            return
        if result.returncode != 0:
            self._changelog_status_var.set(
                f"git log returned {result.returncode}: "
                f"{(result.stderr or '').strip()[:160]}"
            )
            return
        self._changelog_all_commits = []
        for line in (result.stdout or "").split("\n"):
            if "|" not in line:
                continue
            parts = line.split("|", 3)
            if len(parts) == 4:
                self._changelog_all_commits.append(tuple(parts))
        self._changelog_apply_filter()
        self._changelog_status_var.set(
            f"Loaded {len(self._changelog_all_commits)} commits."
        )

    def _changelog_apply_filter(self):
        """Rebuild the visible list based on the filter box (substring,
        case-insensitive against short hash + date + subject)."""
        q = (self._changelog_filter_var.get() or "").strip().lower()
        self._changelog_listbox.delete(0, "end")
        self._changelog_listbox_index = []
        for full, short, date, subject in self._changelog_all_commits:
            hay = f"{short} {date[:10]} {subject}".lower()
            if q and q not in hay:
                continue
            self._changelog_listbox_index.append(full)
            self._changelog_listbox.insert(
                "end", f"{short}  {date[:10]}  {subject[:80]}"
            )
        if self._changelog_listbox_index:
            self._changelog_listbox.selection_set(0)
            self._changelog_show_selected()
        else:
            self._changelog_set_detail("(no commits match the filter)")

    def _changelog_show_selected(self):
        sel = self._changelog_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx >= len(self._changelog_listbox_index):
            return
        full_hash = self._changelog_listbox_index[idx]
        import subprocess
        try:
            result = subprocess.run(
                ["git", "show", "--stat", "--patch", "--no-color", full_hash],
                cwd=str(_REPO_ROOT),
                capture_output=True, text=True, timeout=30,
                creationflags=(subprocess.CREATE_NO_WINDOW
                               if sys.platform == "win32"
                               and hasattr(subprocess, "CREATE_NO_WINDOW") else 0),
            )
            text = result.stdout if result.returncode == 0 else (
                f"git show failed: {result.stderr.strip()}"
            )
        except Exception as exc:
            text = f"git show error: {exc!r}"
        # Cap massive diffs so the widget stays responsive
        if len(text) > 200_000:
            text = text[:200_000] + "\n\n... (diff truncated)"
        self._changelog_set_detail(text)

    def _changelog_set_detail(self, text: str):
        self._changelog_detail.configure(state="normal")
        self._changelog_detail.delete("1.0", "end")
        self._changelog_detail.insert("1.0", text)
        self._changelog_detail.configure(state="disabled")

    def _build_speech_tab(self):
        self.tab_speech = ttk.Frame(self.nb)
        self.nb.add(self.tab_speech, text="🎙 Speech")

        top = ttk.Frame(self.tab_speech)
        top.pack(fill="both", expand=True, padx=10, pady=10)

        btns = ttk.Frame(top)
        btns.pack(fill="x")
        ttk.Button(btns, text="Record 5s",       command=self._stt_record).pack(side="left")
        ttk.Button(btns, text="Transcribe",      command=self._stt_transcribe).pack(side="left", padx=6)
        ttk.Button(btns, text="Send to Council", command=self._stt_send_to_council).pack(side="left")

        # ── #10 TTS controls ─────────────────────────────────────────────
        ttk.Separator(btns, orient="vertical").pack(side="left", fill="y", padx=10)
        ttk.Button(btns, text="🔊 Speak Last Answer",
                   command=self._tts_speak_last).pack(side="left")
        ttk.Button(btns, text="⏹ Stop",
                   command=self._tts_stop).pack(side="left", padx=4)
        self.var_tts_auto = tk.BooleanVar(value=False)
        ttk.Checkbutton(btns, text="Auto-speak answers",
                        variable=self.var_tts_auto).pack(side="left", padx=4)
        ttk.Label(btns, text="Rate:", foreground="#6c7086").pack(side="left", padx=(8, 2))
        self._tts_rate_var = tk.StringVar(value="175")
        ttk.Spinbox(btns, textvariable=self._tts_rate_var,
                    from_=80, to=300, increment=10, width=5).pack(side="left")

        ttk.Label(top, text="Transcription").pack(anchor="w", pady=(10, 0))
        self.stt_out = self._make_text(top, wrap="word")
        self.stt_out.pack(fill="both", expand=True)
        self._tts_engine = None  # lazy init

    # ---- Apothecary tab (advanced mode only) ----

    def _build_apoth_tab(self):
        """Lazily import apothecary_engine and build the Apothecary tab.

        This is only called from _build_ui when _ADVANCED_MODE is True.
        Consumer builds skip this entirely — the import never happens,
        the SSH provisioning code never loads, and the Ollama-flavoured
        UI strings the Apothecary still contains never appear anywhere
        in the bundled .exe.
        """
        global ae
        if ae is None:
            try:
                import apothecary_engine as _ae_mod
                ae = _ae_mod
            except Exception as exc:
                print(f"[Apothecary] failed to import: {exc!r}")
                return
        # Apothecary engine init — also lazy. Was previously created in
        # __init__ unconditionally; moved here so consumer mode pays
        # nothing for an engine it'll never expose.
        if not hasattr(self, "apoth"):
            self.apoth = ae.Apothecary(
                registry_path=str(REGISTRY_PATH),
                store_passwords=STORE_PASSWORDS,
            )
        self.tab_apoth = ttk.Frame(self.nb)
        self.nb.add(self.tab_apoth, text="🔧 Apothecary")
        self.apoth_console = ae.ApothecaryConsole(
            self.tab_apoth, self.apoth, ui_queue=self.ui_q,
        )
        self.apoth_console.pack(fill="both", expand=True)

    # ============================
    # Transcript helpers
    # ============================

    def _role_tag(self, who: str) -> str:
        key = who.lower().replace("-", "_").replace(" ", "_")
        tag = f"who_{key}"
        if tag in ROLE_COLORS or who in ROLE_COLORS:
            return tag
        return "who_default"

    def _set_text(self, widget: tk.Text, text: str):
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", text)
        # Editable widgets stay editable. Build the tuple from attributes
        # that actually exist — ide_code is advanced-only and missing in
        # the consumer build.
        editable = []
        for attr in ("input", "ide_code", "stt_out", "session_preview"):
            w = getattr(self, attr, None)
            if w is not None:
                editable.append(w)
        if widget not in editable:
            widget.configure(state="disabled")

    def _on_app_close(self):
        """Flush conversation logs, run the self-improvement analyzers,
        and dispose cached DB engines before the window closes. Engine
        disposal returns the pooled SQLAlchemy connections to the
        underlying DB cleanly instead of relying on socket teardown."""
        try:
            if hasattr(self, "conv_logger") and self.conv_logger:
                self.conv_logger.end_session("user_close")
        except Exception:
            pass
        # Clean shutdown ⇒ the GPU load/inference didn't crash this run, so
        # clear the GPU-crash sentinel. (A real CUDA core dump never reaches
        # this handler, so its sentinel correctly survives to force CPU next
        # launch.)
        try:
            import council_engine as _ce_close
            _ce_close.gpu_clear_attempt()
        except Exception:
            pass
        # ── Auto-analyze on close ────────────────────────────────────
        # Aggregate this session's tool gaps + failure signatures into
        # human-reviewed proposals so they accumulate without anyone
        # remembering to press the panel button. Deterministic templates
        # only (no model call — the model may already be unloaded and
        # close must stay fast); both analyzers dedup against the queue
        # so closing the app twice never writes a proposal twice.
        try:
            import tool_gap_analyzer as _tga
            from tool_registry import ToolRegistry as _TReg
            _tmp = _TReg(); _tmp.freeze()
            _gap_rep = _tga.ToolGapAnalyzer(
                _tmp.view(), threshold=2).analyze()
            _fail_rep = _tga.FailureAnalyzer(threshold=3).analyze()
            _new = _gap_rep.proposals_written + _fail_rep.proposals_written
            if _new:
                print(f"[shutdown] self-improvement: {_new} new proposal(s) "
                      "drafted — review in the Agent panel next launch.",
                      flush=True)
        except Exception:
            pass
        try:
            import db_connections as _db
            n = _db.dispose_engines()
            if n:
                print(f"[shutdown] disposed {n} cached DB engine(s)",
                      flush=True)
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass

    def _periodic_log_flush(self):
        """Drain the conversation logger every 30s so crashes don't lose logs."""
        try:
            if hasattr(self, "conv_logger") and self.conv_logger:
                self.conv_logger.flush()
        except Exception:
            pass
        try:
            self.after(30_000, self._periodic_log_flush)
        except Exception:
            pass

    # ── Filename pinning for search-headers zoom-in ──────────────────────
    # When the model's previous turn referenced a vault file by name (e.g.
    # "the closest match is orders.csv"), the user's natural follow-up is
    # "tell me more about orders.csv". We want the next turn to inject the
    # FULL [VAULT MATCH] block for that file even when search-headers
    # mode is otherwise active. The pin set is per-CouncilConsole, keyed
    # by filename (lowercased) with a turn-index expiry so a single
    # mention doesn't permanently inflate the budget.
    PIN_LIFETIME_TURNS = 3   # pin expires this many turns after creation

    def _pin_state(self) -> Dict[str, int]:
        """Lazy-init the pinned-files dict: {filename_lower: turn_added}."""
        if not hasattr(self, "_pinned_files_map") or self._pinned_files_map is None:
            self._pinned_files_map = {}
        if not hasattr(self, "_pin_turn_counter") or self._pin_turn_counter is None:
            self._pin_turn_counter = 0
        return self._pinned_files_map

    def _current_pinned_filenames(self) -> List[str]:
        """Return the list of un-expired pinned filenames. Called by the
        injection pipeline each turn; expires stale entries inline."""
        pins = self._pin_state()
        cur_turn = self._pin_turn_counter
        # Expire pins older than PIN_LIFETIME_TURNS turns.
        for name in list(pins.keys()):
            if cur_turn - pins[name] >= self.PIN_LIFETIME_TURNS:
                pins.pop(name, None)
        return list(pins.keys())

    def _update_pins_from_response(self, response_text: str) -> None:
        """Scan the model's response for filename mentions of vault
        files and pin any it sees. Called from _send after the model
        replies but BEFORE the next turn's injection. The next turn's
        injection picks up the pins and zooms the matching files into
        full [VAULT MATCH] blocks.

        We only pin names that resolve to actual vault records — no
        false positives on names the model invented.
        """
        if not response_text:
            self._pin_turn_counter = (self._pin_turn_counter or 0) + 1
            return
        try:
            idx = _get_vault_index()
        except Exception:
            idx = None
        if idx is None:
            self._pin_turn_counter = (self._pin_turn_counter or 0) + 1
            return
        # Build a set of every record name in the vault index (lower-cased)
        # for an O(1) membership check below.
        try:
            known = {str(rec.get("name", "")).lower()
                     for rec in idx.records.values() if rec.get("name")}
        except Exception:
            known = set()
        if not known:
            self._pin_turn_counter = (self._pin_turn_counter or 0) + 1
            return

        # Find every plausible filename token in the response (anything
        # containing a dot, length 3-80, alnum/punct chars). Cross-check
        # against known vault names.
        pins = self._pin_state()
        cur_turn = self._pin_turn_counter
        for m in _re.finditer(r"[\w\-]{1,60}\.[A-Za-z0-9]{1,8}", response_text):
            tok = m.group(0).strip(",.;:!?)(\"' `").lower()
            if tok and tok in known:
                pins[tok] = cur_turn
        # Advance the turn counter so the next call's expiry math is correct.
        self._pin_turn_counter = cur_turn + 1

    def _adjust_ui_scale(self, delta: float) -> None:
        """Bump the Tk scaling multiplier by `delta`. Clamped to a
        reasonable range so a stuck Ctrl+= doesn't blow up the layout."""
        new_scale = max(0.6, min(4.0, self._ui_scale + delta))
        self._apply_ui_scale(new_scale)

    def _reset_ui_scale(self) -> None:
        """Restore the auto-detected default scaling."""
        try:
            default = float(os.environ.get("COUNCIL_UI_SCALE", "0")) or (
                1.5 if _is_wsl() else (1.3 if sys.platform.startswith("linux") else 1.0)
            )
        except Exception:
            default = 1.0
        self._apply_ui_scale(default)

    def _apply_ui_scale(self, value: float) -> None:
        self._ui_scale = value
        try:
            self.tk.call("tk", "scaling", value)
        except Exception:
            pass
        # Surface the change in the title bar so the user knows the
        # keybinding actually fired. _refresh_title_with_n_ctx will
        # restore the n_ctx tag on its next tick.
        try:
            self.title(f"{self._base_title}  ·  UI scale {value:.2f}×")
        except Exception:
            pass

    def _refresh_title_with_n_ctx(self):
        """Update the window title bar with the current n_ctx + source so the
        user always sees what context window they're working with — without
        having to type 'context info' or check the launch log.

        Re-polled periodically until the model is loaded (n_ctx_status returns
        a preview value before first chat call), then once per minute as a
        cheap drift check (env-var re-evaluation, model swap, etc.).
        """
        try:
            import council_engine as _ce_title
            status = _ce_title.n_ctx_status()
            n_ctx = status.get("n_ctx", 0)
            loaded = status.get("loaded", False)
            tag = f"n_ctx={n_ctx:,}" if loaded else f"n_ctx={n_ctx:,} (preview)"
            self.title(f"{self._base_title}  ·  {tag}")
        except Exception:
            pass
        try:
            # Poll faster until the model loads, then drift-check every 60 s.
            delay = 60_000 if (_ce_title.n_ctx_status().get("loaded")) else 2_000
        except Exception:
            delay = 60_000
        try:
            self.after(delay, self._refresh_title_with_n_ctx)
        except Exception:
            pass

    def _append_transcript(self, who: str, text: str, kind: str = "final"):
        # Mirror to the per-session debug log first (write-only — the model
        # never reads conversation_logs/). This runs before the UI write
        # so even if rendering fails we still capture the event.
        try:
            if hasattr(self, "conv_logger") and self.conv_logger and kind not in ("token",):
                self.conv_logger.log_event(
                    kind=kind, who=who, text=text,
                    meta={"session": getattr(self, "session_id", "")},
                )
        except Exception:
            pass
        # Track the last user question + last final Writer answer so the
        # "Defer to Vault" action can capture the exact turn the model
        # couldn't satisfy (e.g. "give me a much bigger summary").
        try:
            if who == "User":
                self._last_user_text = text
            elif who == "Writer" and kind == "final":
                self._last_answer = text
        except Exception:
            pass

        # Provenance: capture model responses (not the user's own message,
        # not phase/token streams) so "where did X come from" can scan
        # them later.
        try:
            if (hasattr(self, "provenance") and self.provenance
                    and kind in ("final", "observation")
                    and who not in ("User",)):
                self.provenance.add_response(who, text)
        except Exception:
            pass

        tag = self._role_tag(who)   # loop-invariant — compute once, not per widget
        for widget in (getattr(self, "transcript", None),
                       getattr(self, "dream3d_transcript", None)):
            if widget is None:
                continue
            try:
                widget.configure(state="normal")
                if kind == "phase":
                    widget.insert("end", f"  {text}\n", "phase")
                elif kind == "token":
                    widget.insert("end", text, "token")
                else:
                    widget.insert("end", f"\n{who}:\n", tag)
                    widget.insert("end", text.strip() + "\n")
                widget.see("end")
                widget.configure(state="disabled")
            except tk.TclError:
                pass  # widget may have been destroyed

        if kind not in ("token", "phase", "thought"):
            self.librarian.log_event(who, text)
            self.convo_store.append(self.session_id, {"ts": now_iso(), "who": who, "text": text})

    def _append_stream_box(self, who: str, token: str):
        """Append a single token to the live stream preview box.

        Hot path — runs once per streamed token (100+/s). We do the
        cheap insert here but DEFER the expensive see("end") scroll to
        _flush_stream_box(), called once at the end of each
        _poll_ui_queue drain. The widget is left in 'normal' state
        across a drain and flipped back to 'disabled' in the flush, so
        we also avoid two configure() calls per token.
        """
        try:
            self.stream_box.configure(state="normal")
        except tk.TclError:
            return
        if who not in self._stream_buffers:
            # New speaker — add header. The dict is used only as a SET of
            # speakers-seen (this membership test); the per-token text value is
            # never read (it's .pop()/.clear()-ed), so we don't accumulate it —
            # that += was an O(N^2) string realloc on the hottest UI path.
            self._stream_buffers[who] = ""
            self.stream_box.insert("end", f"\n{who}: ", self._role_tag(who))
        self.stream_box.insert("end", token)
        self._stream_box_dirty = True

    def _flush_stream_box(self):
        """Scroll the stream box to the end and re-lock it. Called once
        per _poll_ui_queue drain when ≥1 token was appended — see
        _append_stream_box for why the per-token see() was removed."""
        if not self._stream_box_dirty:
            return
        self._stream_box_dirty = False
        try:
            self.stream_box.see("end")
            self.stream_box.configure(state="disabled")
        except tk.TclError:
            pass

    def _clear_stream_box(self):
        self._stream_buffers.clear()
        self.stream_box.configure(state="normal")
        self.stream_box.delete("1.0", "end")
        self.stream_box.configure(state="disabled")

    # ── Verdict feedback ──────────────────────────────────────────────────────

    def _vfb_show(self):
        """Show the agree/disagree bar after a verdict arrives."""
        if not hasattr(self, "_vfb_frame"):
            return
        # Reset state
        self._vfb_detail.pack_forget()
        self._vfb_frame.pack(fill="x", pady=(4, 0))
        self._vfb_agree_btn.configure(state="normal")
        self._vfb_disagree_btn.configure(state="normal")

    # ── Council instructions ─────────────────────────────────────────────

    def _apply_council_instruction(self):
        """Add a new instruction to the persistent list."""
        text = self._inst_var.get().strip()
        if not text:
            return
        name = self._inst_name.get().strip()
        entry = self._instr_mgr.add(name, text)
        self._inst_var.set("")
        self._inst_name.set("")
        self._update_inst_label()
        self._append_transcript("Council",
            f"⚡ Instruction added: [{entry['name']}] {text[:80]}", "observation")

    def _update_inst_label(self):
        """Refresh the active-count label next to the instruction bar."""
        n = self._instr_mgr.active_count()
        total = len(self._instr_mgr.all())
        if total == 0:
            self._inst_active_lbl.configure(text="")
        else:
            self._inst_active_lbl.configure(
                text=f"{n}/{total} active",
                foreground="#a6e3a1" if n > 0 else "#6c7086")

    def _open_instruction_manager(self):
        """Open the instruction list manager window."""
        win = tk.Toplevel(self)
        win.title("Council Instructions")
        win.configure(bg="#1a1414")
        win.geometry("680x460")
        win.resizable(True, True)

        ttk.Label(win,
            text="Instructions are injected into every personality on every call.",
            foreground="#6c7086").pack(anchor="w", padx=12, pady=(8, 2))

        # ── List ──────────────────────────────────────────────────
        list_frame = ttk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=12, pady=4)

        cols = ("active", "name", "text")
        tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=10)
        tree.heading("active", text="On")
        tree.heading("name",   text="Name")
        tree.heading("text",   text="Instruction text")
        tree.column("active", width=40,  anchor="center", stretch=False)
        tree.column("name",   width=140, stretch=False)
        tree.column("text",   width=460)
        sb = ttk.Scrollbar(list_frame, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)

        def _refresh():
            tree.delete(*tree.get_children())
            for e in self._instr_mgr.all():
                icon = "✓" if e["active"] else "○"
                tree.insert("", "end", iid=e["id"],
                            values=(icon, e["name"], e["text"]))

        _refresh()

        # ── Buttons ───────────────────────────────────────────────
        bf = ttk.Frame(win)
        bf.pack(fill="x", padx=12, pady=(0, 10))

        def _toggle():
            sel = tree.selection()
            if not sel:
                return
            for iid in sel:
                self._instr_mgr.toggle(iid)
            _refresh()
            self._update_inst_label()

        def _delete():
            sel = tree.selection()
            if not sel:
                return
            for iid in sel:
                self._instr_mgr.remove(iid)
            _refresh()
            self._update_inst_label()
            self._append_transcript("Council", "⚡ Instruction(s) removed.", "observation")

        def _edit():
            sel = tree.selection()
            if not sel:
                return
            iid = sel[0]
            entries = {e["id"]: e for e in self._instr_mgr.all()}
            entry = entries.get(iid)
            if not entry:
                return
            edit_win = tk.Toplevel(win)
            edit_win.title("Edit Instruction")
            edit_win.configure(bg="#1a1414")
            edit_win.geometry("540x200")

            ttk.Label(edit_win, text="Name:").pack(anchor="w", padx=12, pady=(10,2))
            name_v = tk.StringVar(value=entry["name"])
            ttk.Entry(edit_win, textvariable=name_v, width=50).pack(anchor="w", padx=12)

            ttk.Label(edit_win, text="Instruction text:").pack(anchor="w", padx=12, pady=(8,2))
            text_box = tk.Text(edit_win, height=4, wrap="word",
                               bg="#0f0c0c", fg="#d4d4d4", font=("Consolas", 9))
            text_box.insert("1.0", entry["text"])
            text_box.pack(fill="x", padx=12)

            def _save_edit():
                new_text = text_box.get("1.0", "end").strip()
                new_name = name_v.get().strip()
                if new_text:
                    self._instr_mgr.update_text(iid, new_text)
                    for e in self._instr_mgr.all():
                        if e["id"] == iid:
                            e["name"] = new_name or new_text[:40]
                    self._instr_mgr._save()
                edit_win.destroy()
                _refresh()

            ttk.Button(edit_win, text="Save", command=_save_edit).pack(pady=8)

        def _toggle_all_on():
            for e in self._instr_mgr.all():
                if not e["active"]:
                    self._instr_mgr.toggle(e["id"])
            _refresh()
            self._update_inst_label()

        def _toggle_all_off():
            for e in self._instr_mgr.all():
                if e["active"]:
                    self._instr_mgr.toggle(e["id"])
            _refresh()
            self._update_inst_label()

        ttk.Button(bf, text="Toggle On/Off", command=_toggle).pack(side="left")
        ttk.Button(bf, text="Edit",          command=_edit).pack(side="left", padx=4)
        ttk.Button(bf, text="Delete",        command=_delete).pack(side="left")
        ttk.Separator(bf, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(bf, text="All On",  command=_toggle_all_on).pack(side="left")
        ttk.Button(bf, text="All Off", command=_toggle_all_off).pack(side="left", padx=4)
        ttk.Button(bf, text="Close",   command=win.destroy).pack(side="right")

        # Double-click to toggle
        tree.bind("<Double-1>", lambda e: _toggle())

    # ── Clarification pause/resume ────────────────────────────────────────────

    def _show_clarification(self, who: str, question: str):
        """Show the clarification panel with the personality's question."""
        self._clarif_question_lbl.configure(
            text=f"{who} asks: {question}"
        )
        self._clarif_var.set("")
        self._clarif_frame.pack(fill="x", pady=(4, 0))
        self._set_status("⏸ Waiting for your answer…", "#fab387")

    def _hide_clarification(self):
        """Hide the clarification panel."""
        self._clarif_frame.pack_forget()

    def _submit_clarification(self, skip: bool = False):
        """User answered the personality's question — resume deliberation."""
        if skip:
            self._clarification_answer = "[User skipped — continue without this information]"
        else:
            ans = self._clarif_var.get().strip()
            self._clarification_answer = ans if ans else "[No answer provided]"
        # Store answer and log BEFORE resuming worker so answer_getter sees it
        ans_display = self._clarification_answer
        self._hide_clarification()
        self._append_transcript("You", ans_display, "final")
        self._pause_event.set()  # Resume the worker thread

    def _open_content_style(self):
        """Open the content style manager for cross-session creator learning."""
        if not hasattr(self, "_content_style"):
            messagebox.showinfo("Not available", "Content style manager not initialised.", parent=self)
            return

        win = tk.Toplevel(self)
        win.title("Content Style & Templates")
        win.configure(bg="#1a1414")
        win.geometry("700x560")
        win.resizable(True, True)

        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        # ── Style Preferences tab ─────────────────────────────────
        pref_tab = ttk.Frame(nb)
        nb.add(pref_tab, text="Style Preferences")

        ttk.Label(pref_tab, text="Audience description:",
                  foreground="#d32f2f").pack(anchor="w", padx=12, pady=(10,2))
        aud_v = tk.StringVar(value=self._content_style._data.get("audience", ""))
        ttk.Entry(pref_tab, textvariable=aud_v, width=60).pack(anchor="w", padx=12)

        ttk.Label(pref_tab, text="Channel tone / style:",
                  foreground="#d32f2f").pack(anchor="w", padx=12, pady=(8,2))
        tone_v = tk.StringVar(value=self._content_style._data.get("tone", ""))
        ttk.Entry(pref_tab, textvariable=tone_v, width=60).pack(anchor="w", padx=12)

        ttk.Label(pref_tab, text="Add style note (what worked, what to avoid, etc.):",
                  foreground="#d32f2f").pack(anchor="w", padx=12, pady=(8,2))
        _note_row = ttk.Frame(pref_tab)
        _note_row.pack(fill="x", padx=12)
        note_v    = tk.StringVar()
        note_cat  = tk.StringVar(value="general")
        ttk.Entry(_note_row, textvariable=note_v, width=44).pack(side="left")
        ttk.Combobox(_note_row, textvariable=note_cat, width=12,
                     values=["general","hook","pacing","tone","structure","cta"],
                     state="readonly").pack(side="left", padx=4)
        ttk.Button(_note_row, text="Add Note",
                   command=lambda: _add_note()).pack(side="left", padx=4)

        ttk.Label(pref_tab, text="Existing style notes:",
                  foreground="#6c7086").pack(anchor="w", padx=12, pady=(8,2))
        notes_box = tk.Text(pref_tab, height=8, bg="#0f0c0c", fg="#d4d4d4",
                            font=("Consolas", 9), state="disabled", relief="flat", wrap="word")
        notes_box.pack(fill="both", expand=True, padx=12, pady=(0,8))

        def _refresh_notes():
            notes_box.configure(state="normal")
            notes_box.delete("1.0", "end")
            for n in reversed(self._content_style.get_style_notes()[-20:]):
                notes_box.insert("end", "[" + n["category"] + "] " + n["note"] + "\n")
            notes_box.configure(state="disabled")

        def _add_note():
            txt = note_v.get().strip()
            if txt:
                self._content_style.add_style_note(txt, note_cat.get())
                note_v.set("")
                _refresh_notes()

        _refresh_notes()

        def _save_prefs():
            if aud_v.get().strip():
                self._content_style.set_audience(aud_v.get())
            if tone_v.get().strip():
                self._content_style.set_tone(tone_v.get())
            messagebox.showinfo("Saved", "Style preferences saved.", parent=win)

        ttk.Button(pref_tab, text="Save Preferences", command=_save_prefs).pack(pady=4)

        # ── Templates tab ─────────────────────────────────────────
        tmpl_tab = ttk.Frame(nb)
        nb.add(tmpl_tab, text="Script Templates")

        ttk.Label(tmpl_tab,
                  text="Templates are automatically selected based on your video type request.",
                  foreground="#6c7086").pack(anchor="w", padx=12, pady=(8,2))

        tmpl_list = tk.Text(tmpl_tab, height=22, bg="#0f0c0c", fg="#d4d4d4",
                            font=("Consolas", 9), state="disabled", relief="flat", wrap="word")
        tmpl_sb = ttk.Scrollbar(tmpl_tab, command=tmpl_list.yview)
        tmpl_list.configure(yscrollcommand=tmpl_sb.set)
        tmpl_sb.pack(side="right", fill="y", padx=(0,8))
        tmpl_list.pack(fill="both", expand=True, padx=(12,0), pady=(0,8))

        tmpl_list.configure(state="normal")
        for key, tmpl in self._content_style.get_templates().items():
            tmpl_list.insert("end", "▶ " + tmpl["name"] + "\n", "hdr")
            tmpl_list.insert("end", "  " + tmpl["description"] + "\n", "desc")
            for step in tmpl.get("structure", []):
                tmpl_list.insert("end", "    " + step + "\n")
            tmpl_list.insert("end", "\n")
        tmpl_list.tag_config("hdr",  foreground="#fab387", font=("Consolas", 9, "bold"))
        tmpl_list.tag_config("desc", foreground="#6c7086")
        tmpl_list.configure(state="disabled")

        ttk.Button(win, text="Close", command=win.destroy).pack(pady=(0,8))

    def _on_voice_toggle(self, *_):
        """Called whenever the Robust Voices toggle changes."""
        enabled = bool(self.var_robust_voices.get())
        ce.set_voice_mode(self.personalities, enabled)
        state = "ON — each personality now has its own voice" if enabled else "OFF — neutral mode"
        self._set_status(f"● Robust voices: {state}", "#cba6f7" if enabled else "#a6e3a1")

    def _on_profile_toggle(self, *_):
        """Explicit user-profile bypass. Unchecking parks the learned
        USER PROFILE for now — injection stops on the very next message
        (the engine checks the flag at respond() time) while quirk
        OBSERVATION keeps running, so nothing learned is lost and the
        profile keeps maturing in the background."""
        enabled = bool(self.var_use_profile.get())
        ce.set_user_profile_apply(enabled)
        if enabled:
            self._set_status("● 👤 Profile: ON — learned preferences "
                             "inform answers again", "#a6e3a1")
        else:
            self._set_status("● 👤 Profile: BYPASSED — answers ignore "
                             "learned preferences (learning continues)",
                             "#fab387")

    def _vfb_hide(self):
        """Hide the feedback bar and detail panel."""
        if hasattr(self, "_vfb_frame"):
            self._vfb_frame.pack_forget()
        if hasattr(self, "_vfb_detail"):
            self._vfb_detail.pack_forget()

    def _verdict_agree(self):
        """User agrees — log it and dismiss the bar."""
        self._append_transcript("You", "✓ Agreed with verdict.", "observation")
        # Record agreement in verdict history if the last record exists
        try:
            import json
            path = VAULT_DIR / "verdict_history.jsonl"
            if path.exists():
                lines = path.read_text(encoding="utf-8").strip().splitlines()
                if lines:
                    last = json.loads(lines[-1])
                    last["user_agreed"] = True
                    lines[-1] = json.dumps(last)
                    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        except Exception:
            pass
        self._vfb_hide()

    def _verdict_disagree_open(self):
        """Open the objection text box."""
        self._vfb_agree_btn.configure(state="disabled")
        self._vfb_disagree_btn.configure(state="disabled")
        self._vfb_text.delete("1.0", "end")
        self._vfb_detail.pack(fill="x", pady=(2, 0))
        self._vfb_text.focus_set()

    def _verdict_disagree_cancel(self):
        """Cancel and re-enable the buttons."""
        self._vfb_detail.pack_forget()
        self._vfb_agree_btn.configure(state="normal")
        self._vfb_disagree_btn.configure(state="normal")

    def _verdict_disagree_submit(self):
        """
        Take the user's objection and re-run the last query with it prepended.
        The council sees the original question plus the user's critique of the
        previous answer — forcing a fresh deliberation that addresses the objection.
        """
        objection = self._vfb_text.get("1.0", "end").strip()
        if not objection:
            return

        # Recover last query from transcript
        last_query = getattr(self, "_last_sent_query", "")

        # Record disagreement in verdict history
        try:
            import json
            path = VAULT_DIR / "verdict_history.jsonl"
            if path.exists():
                lines_vrd = path.read_text(encoding="utf-8").strip().splitlines()
                if lines_vrd:
                    last_vrd = json.loads(lines_vrd[-1])
                    last_vrd["user_agreed"]   = False
                    last_vrd["user_objection"] = objection
                    lines_vrd[-1] = json.dumps(last_vrd)
                    path.write_text("\n".join(lines_vrd) + "\n", encoding="utf-8")
        except Exception:
            pass

        self._vfb_hide()

        # Build the re-run query with objection prepended
        if last_query:
            rerun_text = (
                f"[USER OBJECTION TO PREVIOUS ANSWER]\n"
                f"{objection}\n\n"
                f"[ORIGINAL QUESTION — please re-answer addressing the objection above]\n"
                f"{last_query}"
            )
        else:
            rerun_text = (
                f"[USER OBJECTION]\n{objection}\n\n"
                "Please re-examine your previous answer and address this objection directly."
            )

        self._append_transcript("You", f"✗ Disagreed: {objection}", "observation")
        self._set_text(self.input, rerun_text)
        self._send()

    # ─────────────────────────────────────────────────────────────────────────

    def _set_judge(self, text: str):
        self.judge_box.configure(state="normal")
        self.judge_box.delete("1.0", "end")
        self.judge_box.insert("1.0", text)
        self.judge_box.configure(state="disabled")

    def _ide_print(self, text: str, tag: str = ""):
        self.ide_out.configure(state="normal")
        if tag:
            self.ide_out.insert("end", text, tag)
        else:
            self.ide_out.insert("end", text)
        self.ide_out.see("end")
        self.ide_out.configure(state="disabled")

    def _update_tps(self, who: str, tps: float):
        """Show live tokens/s in the status bar. Keeps a rolling history."""
        if not hasattr(self, "_tps_history"):
            self._tps_history = {}
        self._tps_history[who] = tps
        if hasattr(self, "tps_label") and self.tps_label.winfo_exists():
            # Show the most recent role + its speed
            parts = [k[:4] + ":" + str(v) for k, v in list(self._tps_history.items())[-3:]]
            self.tps_label.configure(text=" | ".join(parts) + " t/s")

    def _set_status(self, text: str, color: str = "#a6e3a1"):
        self.status.configure(text=text, foreground=color)

    # ============================
    # Find-relevant-data → Grapher → Analyst
    # ============================
    # When the user asks a chart-shaped question in the Council, we don't
    # want them to hunt for the right CSV manually. This helper scans the
    # vault and the bundled sample data for files that look relevant to
    # the query, hands them to the Grapher, and asks the Analyst to plot.

    # Lower-case stop-words removed from queries before scoring filenames.
    _DATA_QUERY_STOPS = {
        "show", "me", "the", "a", "an", "of", "for", "by", "in", "on", "at",
        "and", "or", "to", "with", "what", "which", "how", "many", "much",
        "do", "does", "is", "are", "was", "were", "have", "has", "had",
        "graph", "chart", "plot", "visualize", "visualise", "see", "find",
        "tell", "give", "from", "this", "that", "these", "those", "all",
        "make", "create", "draw", "produce", "us", "our", "i", "my", "your",
    }

    # Substrings that signal the user wants a chart pulled together.
    # Matched case-insensitively in the raw query.
    _DATA_QUERY_TRIGGERS = (
        "graph", "chart", "plot", "visualiz", "visualis",
        "show me", "show the", "trend", "by month", "by year",
        "by week", "by day", "by category", "compare", "histogram",
    )

    # Substrings that signal the user is asking for a value lookup or
    # a connection between files — handled by the data index instead of
    # the deliberation. Faster, deterministic, no LLM round-trip.
    _LOOKUP_QUERY_TRIGGERS = (
        "look up", "lookup", "find ", "search ", "search for",
        "where is", "who is", "who are", "whose ", "which file",
        "across files", "linked to", "connection between",
        "tell me about", "what about",
    )

    def _looks_like_lookup_question(self, query: str) -> bool:
        ql = query.lower()
        # Don't trigger lookup if it's clearly chart-shaped
        if self._looks_like_data_question(query):
            return False
        return any(t in ql for t in self._LOOKUP_QUERY_TRIGGERS)

    def _looks_like_data_question(self, query: str) -> bool:
        ql = query.lower()
        return any(t in ql for t in self._DATA_QUERY_TRIGGERS)

    def _query_keywords(self, query: str):
        """Strip punctuation + stop words, return content terms (length > 2)."""
        terms = [t.strip(".,!?;:()[]\"'").lower() for t in query.split()]
        return [t for t in terms if t and len(t) > 2 and t not in self._DATA_QUERY_STOPS]

    def _vault_data_files(self):
        """
        All loadable data files we're allowed to read. Strict read scope:
        only vault/data_in/ and the bundled samples — not the rest of
        vault/ where the app's internal state lives. This keeps inputs
        cleanly separated from anything the app might overwrite.
        """
        exts = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".npy", ".npz"}
        roots = [
            data_index.input_dir(VAULT_DIR),
            data_index.bundled_samples_dir(),
        ]
        files = []
        seen = set()
        for root in roots:
            if not root.exists():
                continue
            for p in root.rglob("*"):
                if not p.is_file() or p.suffix.lower() not in exts:
                    continue
                if p in seen:
                    continue
                seen.add(p)
                files.append(p)
        return files

    # Business synonyms — when a user types "revenue" the file probably has
    # "total" or "amount", not "revenue". Each entry is one term and the
    # column-name aliases that should also count as a hit for it.
    _SYNONYMS = {
        "revenue":  ("total", "amount", "sales", "price", "value"),
        "sales":    ("total", "amount", "revenue", "order", "value"),
        "income":   ("total", "amount", "revenue", "sales"),
        "spend":    ("total", "amount", "spent", "purchase", "lifetime"),
        "spending": ("total", "amount", "spent", "purchase"),
        "purchase": ("order", "po", "buy", "transaction"),
        "order":    ("po", "purchase", "transaction"),
        "client":   ("customer", "buyer", "account"),
        "customer": ("client", "buyer", "account"),
        "stock":    ("inventory", "qty", "quantity", "on_hand"),
        "inventory": ("stock", "qty", "quantity", "sku"),
        "product":  ("sku", "item", "name"),
        "category": ("type", "group", "class", "segment"),
        "supplier": ("vendor", "source"),
        "monthly":  ("month", "date", "ts", "timestamp"),
        "yearly":   ("year", "date", "ts", "annual"),
        "weekly":   ("week", "date", "ts"),
        "daily":    ("day", "date", "ts"),
        "trend":    ("date", "time", "ts"),
        "retention": ("last", "first", "ltv", "churn", "active"),
        "dormant":  ("last_order", "last_activity", "inactive", "churn"),
    }

    def _score_file_for_query(self, path: Path, terms) -> float:
        """
        Filename hit (×2) + header hit (×1) + synonym header hit (×0.6).
        Quick and tolerant of natural-language vs column-name vocabulary.
        """
        score = 0.0
        name = path.stem.lower()
        for t in terms:
            if t in name:
                score += 2.0

        if path.suffix.lower() in (".csv", ".tsv"):
            try:
                with path.open("r", encoding="utf-8", errors="replace") as f:
                    header = f.readline().lower()
            except Exception:
                return score
            for t in terms:
                if t in header:
                    score += 1.0
                # Synonym fallback — only if the term itself didn't already hit
                else:
                    for alt in self._SYNONYMS.get(t, ()):
                        if alt in header:
                            score += 0.6
                            break
        return score

    def _find_data_files_for_query(self, query: str, top_n: int = 3):
        terms = self._query_keywords(query)
        if not terms:
            return []

        # Boost: any file scored against the same query also benefits from
        # the active specialists' domain keywords. So if the user has a
        # Sales Specialist active, files that mention "revenue" get a
        # small extra weight even when the user typed only "monthly".
        spec_terms = []
        try:
            specs = self._resolve_active_specialists(query)
            for s in specs:
                spec_terms.extend(s.domain_keywords[:8])  # cap per specialist
        except Exception:
            specs = []

        scored = []
        for p in self._vault_data_files():
            s = self._score_file_for_query(p, terms)
            if spec_terms:
                # Add half-weight contributions from specialist keywords
                bonus = self._score_file_for_query(p, spec_terms) * 0.5
                s += bonus
            if s > 0:
                scored.append((s, p))
        scored.sort(key=lambda r: r[0], reverse=True)
        return [p for _, p in scored[:top_n]]

    def _council_find_and_chart_button(self):
        """Button handler — explicit user trigger; consumes the input like _send."""
        query = self.input.get("1.0", "end").strip()
        if not query:
            return
        self._set_text(self.input, "")
        self._append_transcript("User", query)
        self._council_find_and_chart(query)

    def _show_examples(self):
        """A 'What can I ask?' panel that raises discoverability of the app's
        capabilities. Clicking an example drops it into the council input
        (ready to send), so a new user can learn by doing."""
        import tkinter as tk
        from tkinter import ttk
        win = tk.Toplevel(self)
        win.title("What can I ask?")
        win.geometry("640x520")
        try:
            win.transient(self)
        except Exception:
            pass
        ttk.Label(win, text="Click an example to drop it into the input box.",
                  foreground="#888", anchor="w").pack(fill="x", padx=10,
                                                      pady=(8, 2))
        canvas = tk.Canvas(win, highlightthickness=0, bg="#1a1414")
        scroll = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(10, 0),
                    pady=6)
        scroll.pack(side="right", fill="y", pady=6)

        def _use(prompt):
            # Skip the parenthetical "(click ...)" guidance rows.
            if prompt.startswith("("):
                return
            self._set_text(self.input, prompt)
            win.destroy()
            try:
                self.input.focus_set()
            except Exception:
                pass

        last_cat = None
        for cat, prompt, hint in _COUNCIL_EXAMPLES:
            if cat != last_cat:
                ttk.Label(body, text=cat, font=("", 10, "bold"),
                          foreground="#f9b384").pack(anchor="w",
                                                     pady=(10, 2))
                last_cat = cat
            row = ttk.Frame(body)
            row.pack(fill="x", pady=2)
            btn_text = prompt if prompt.startswith("(") else f"→ {prompt}"
            b = ttk.Button(row, text=btn_text, width=46,
                           command=lambda p=prompt: _use(p))
            b.pack(side="left")
            if prompt.startswith("("):
                b.state(["disabled"])
            ttk.Label(row, text=hint, foreground="#7a7575",
                      wraplength=300, justify="left").pack(side="left",
                                                           padx=8)
        ttk.Button(win, text="Close",
                   command=win.destroy).pack(pady=(0, 8))

    def _show_question_history(self):
        """Browse past Council questions and re-ask one with a click. Re-asks
        are cheap + correct now that fresh derived results are reused."""
        import tkinter as tk
        from tkinter import ttk
        try:
            import question_history as _qh
            store = _qh.QuestionHistory(VAULT_DIR)
            items = store.recent(300)
        except Exception:
            store, items = None, []
        win = tk.Toplevel(self)
        win.title("Question history")
        win.geometry("680x480")
        try:
            win.transient(self)
        except Exception:
            pass
        ttk.Label(win, text=f"{len(items)} past question(s) — newest first",
                  foreground="#888", anchor="w").pack(fill="x", padx=8,
                                                      pady=(6, 0))
        lb = tk.Listbox(win, bg="#231a1a", fg="#d4d4d4",
                        selectbackground="#5a3030", relief="flat",
                        font=("Consolas", 10))
        lb.pack(fill="both", expand=True, padx=8, pady=8)
        questions = []
        for it in items:
            q = str(it.get("q", "")).replace("\n", " ").strip()
            questions.append(q)
            lb.insert("end", q[:160] if q else "(blank)")
        if not items:
            lb.insert("end", "(no questions yet — ask something in the "
                      "Council tab)")

        def _selected():
            sel = lb.curselection()
            if not sel or sel[0] >= len(questions):
                return None
            return questions[sel[0]]

        def _reask():
            q = _selected()
            if not q:
                return
            win.destroy()
            self._set_text(self.input, q)
            self._send()

        lb.bind("<Double-Button-1>", lambda e: _reask())
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(btns, text="Re-ask", command=_reask).pack(side="left")

        def _copy():
            q = _selected()
            if q:
                try:
                    self.clipboard_clear()
                    self.clipboard_append(q)
                except Exception:
                    pass

        ttk.Button(btns, text="Copy", command=_copy).pack(side="left", padx=6)

        def _clear():
            from tkinter import messagebox
            if store is not None and messagebox.askyesno(
                    "Clear history",
                    "Remove all saved questions? This cannot be undone.",
                    parent=win):
                store.clear()
                lb.delete(0, "end")
                questions.clear()

        ttk.Button(btns, text="Clear all", command=_clear).pack(side="left",
                                                                padx=6)
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right")

    def _save_council_answer(self):
        """Export the most recent answer (question + answer + any result table
        + sources) to a Markdown report. Defaults into data_in/derived/ so it
        sits with the other computed outputs; the user picks the final path."""
        from tkinter import filedialog, messagebox
        answer = (getattr(self, "_last_answer", "") or "").strip()
        if not answer:
            messagebox.showinfo(
                "Nothing to save",
                "Ask a question first — there's no answer to save yet.")
            return
        question = ((getattr(self, "_last_user_text", "") or "")
                    or getattr(self, "_last_sent_query", "") or "").strip()
        table = (getattr(self, "_last_answer_table", "") or "").strip()
        sources = self._resolve_source_paths(
            getattr(self, "_last_turn_sources", []))
        slug = _re.sub(r"[^a-z0-9]+", "_", question.lower()).strip("_")[:48] \
            or "answer"
        try:
            import derived_results as _drv
            default_dir = _drv.derived_dir(VAULT_DIR)
        except Exception:
            default_dir = VAULT_DIR
        path = filedialog.asksaveasfilename(
            title="Save answer report",
            initialdir=str(default_dir),
            initialfile=f"answer__{slug}.md",
            defaultextension=".md",
            filetypes=[("Markdown report", "*.md"), ("Text", "*.txt"),
                       ("All files", "*.*")])
        if not path:
            return
        md = _build_answer_report_md(question, answer, table, sources)
        try:
            Path(path).write_text(md, encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Save failed", f"{e!r}")
            return
        self._append_transcript(
            "Council", f"Saved the answer to {Path(path).name}.",
            "observation")
        self._render_source_chips([path])

    def _council_expand_with_council(self):
        """Re-ask the last fast (direct-route) question through the FULL
        multi-role council. The fast-answer short-circuit gave an instant
        deterministic answer; this trades that speed for a prose discussion
        when the user wants it. One-shot bypass via _force_full_council."""
        q = (getattr(self, "_last_fast_question", "") or "").strip()
        if not q:
            return
        try:
            self._expand_btn.configure(state="disabled")
        except Exception:
            pass
        self._force_full_council = True
        self._last_fast_question = ""
        self._set_text(self.input, q)
        self._send()

    # ============================
    # Look Up — cross-file value/column search
    # ============================
    # Treat the input as either a literal value (lookup) or a column
    # name (column membership) and surface every file that mentions it,
    # including cross-references between files. The Council itself isn't
    # invoked — this is a deterministic scan of indexed data, useful for
    # answers that don't need a deliberation.

    def _council_lookup_button(self):
        """Button handler — runs the lookup against the input box value."""
        query = self.input.get("1.0", "end").strip()
        if not query:
            return
        self._set_text(self.input, "")
        self._append_transcript("User", query)
        self._council_run_lookup(query)

    def _council_run_lookup(self, query: str):
        """Execute the lookup, post results to transcript, and pop a results window."""
        # Refresh the index lazily — picks up new files that arrived in
        # the vault since startup (e.g. user just dropped a CSV in).
        self._append_transcript("Council",
            "Looking up across indexed files…", "observation")
        try:
            self.data_index.refresh()
        except Exception as e:
            self._append_transcript("Council",
                f"Could not refresh data index: {e}", "final")
            return

        # Value + column-name search, driven by the query's CONTENT terms
        # rather than the whole raw sentence. search_value tests
        # `needle in cell`, so passing "who bought promethium in Q3" never
        # substring-matches a single cell — we extract keywords first and
        # union the per-term hits, merged by file. The popup shows both value
        # and column matches; the more useful set bubbles to the top.
        terms = self._query_keywords(query) or [query.strip()]
        terms = [t for t in dict.fromkeys(terms) if t]   # dedupe, keep order

        merged: dict = {}
        for term in terms:
            for h in self.data_index.search_value(term, max_per_file=25):
                cur = merged.get(h["path"])
                if cur is None:
                    merged[h["path"]] = {
                        "file":          h["file"],
                        "path":          h["path"],
                        "row_count":     h["row_count"],
                        "matched_count": len(h["rows"]),
                        "column_hits":   list(h["column_hits"]),
                        "rows":          list(h["rows"]),
                        "_rowkeys":      {tuple(sorted(r.items())) for r in h["rows"]},
                    }
                else:
                    for c in h["column_hits"]:
                        if c not in cur["column_hits"]:
                            cur["column_hits"].append(c)
                    for r in h["rows"]:
                        rk = tuple(sorted(r.items()))
                        if rk not in cur["_rowkeys"] and len(cur["rows"]) < 25:
                            cur["_rowkeys"].add(rk)
                            cur["rows"].append(r)
        value_hits = sorted(merged.values(), key=lambda r: -len(r["rows"]))
        for h in value_hits:
            h["matched_count"] = len(h["rows"])
            h.pop("_rowkeys", None)

        col_seen: set = set()
        col_hits: list = []
        for term in terms:
            for prof, exact in self.data_index.find_files_with_column(term):
                key = (getattr(prof, "name", str(prof)), exact)
                if key not in col_seen:
                    col_seen.add(key)
                    col_hits.append((prof, exact))
        relationships = self.data_index.find_relationships()

        if not value_hits and not col_hits:
            self._append_transcript("Council",
                f"No matches for {query!r} in any indexed file. "
                "Try the Vault tab to add the data first, or use Grapher → "
                "Sample to explore the bundled demo CSVs.", "final")
            return

        # Summarise into the transcript — one line per hit
        lines: list = []
        if value_hits:
            lines.append("Value matches:")
            for h in value_hits[:6]:
                cols = ", ".join(h["column_hits"][:4])
                lines.append(f"  • {h['file']} — {h['matched_count']} row(s)  (in: {cols})")
        if col_hits:
            lines.append("Files with that column:")
            for prof, exact in col_hits[:6]:
                lines.append(f"  • {prof.name} — column “{exact}”")
        self._append_transcript("Council", "\n".join(lines), "observation")

        # Open the detailed results window
        self._lookup_show_window(query, value_hits, col_hits, relationships)

    def _lookup_show_window(self, query, value_hits, col_hits, relationships):
        """Detailed read-only window showing matches + connection map."""
        win = tk.Toplevel(self)
        win.title(f"Look Up — {query}")
        win.geometry("780x520")
        try: branding.apply_window_icon(win)
        except Exception: pass
        try:
            t = branding.get_theme("dark")
            win.configure(bg=t["bg"])
            fg, bg, abg, mfg = t["fg"], t["bg"], t["panel_bg"], t["muted_fg"]
        except Exception:
            fg, bg, abg, mfg = "#d4d4d4", "#1a1414", "#231a1a", "#7a7575"

        # Header
        head = tk.Frame(win, bg=abg)
        head.pack(fill="x")
        tk.Label(head, text=f"\U0001f50d  Look Up", font=("Segoe UI", 14, "bold"),
                 bg=abg, fg=fg).pack(side="left", padx=14, pady=10)
        tk.Label(head, text=f"query: {query}",
                 bg=abg, fg=mfg, font=("Consolas", 10)).pack(side="left", padx=8)

        # Tabs: Matches | Files with column | Relationships
        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        # ---- Matches tab ---------------------------------------
        matches_tab = ttk.Frame(nb)
        nb.add(matches_tab, text=f"Matches  ({sum(h['matched_count'] for h in value_hits)})")
        if not value_hits:
            tk.Label(matches_tab,
                     text=f"No rows contain the value {query!r} in any indexed file.",
                     bg=bg, fg=mfg, wraplength=720, justify="left",
                     ).pack(anchor="w", padx=12, pady=12)
        else:
            txt = self._make_text(matches_tab, height=24, wrap="none")
            txt.pack(fill="both", expand=True, padx=8, pady=8)
            for h in value_hits:
                txt.insert("end",
                    f"━━ {h['file']}  ({h['matched_count']} row(s) "
                    f"of {h['row_count']:,} match in: "
                    f"{', '.join(h['column_hits'])}) ━━\n")
                # Render up to 5 rows per file
                for row in h["rows"][:5]:
                    parts = "  |  ".join(
                        f"{k}={v}" for k, v in row.items() if v
                    )
                    txt.insert("end", "  " + parts[:240] + "\n")
                if len(h["rows"]) > 5:
                    txt.insert("end",
                        f"  …and {h['matched_count'] - 5} more (raise max_per_file to see all)\n")
                txt.insert("end", "\n")
            txt.configure(state="disabled")

        # ---- Column tab ----------------------------------------
        col_tab = ttk.Frame(nb)
        nb.add(col_tab, text=f"Column Match  ({len(col_hits)})")
        if not col_hits:
            tk.Label(col_tab, text=f"No file has a column named like {query!r}.",
                     bg=bg, fg=mfg, wraplength=720, justify="left",
                     ).pack(anchor="w", padx=12, pady=12)
        else:
            txt = self._make_text(col_tab, height=24)
            txt.pack(fill="both", expand=True, padx=8, pady=8)
            for prof, exact in col_hits:
                col = next((c for c in prof.columns if c.name == exact), None)
                samples = ", ".join(col.sample_values[:6]) if col else ""
                txt.insert("end",
                    f"• {prof.name}\n"
                    f"    column:    {exact}  ({col.inferred_type if col else 'text'})\n"
                    f"    samples:   {samples}\n"
                    f"    distinct:  {col.distinct_count if col else 0}\n\n")
            txt.configure(state="disabled")

        # ---- Relationships tab --------------------------------
        rel_tab = ttk.Frame(nb)
        nb.add(rel_tab, text=f"Connections  ({len(relationships)})")
        if not relationships:
            tk.Label(rel_tab,
                     text="No cross-file column matches detected. Add more "
                          "data files to the vault to enable relationship "
                          "detection.",
                     bg=bg, fg=mfg, wraplength=720, justify="left",
                     ).pack(anchor="w", padx=12, pady=12)
        else:
            txt = self._make_text(rel_tab, height=24)
            txt.pack(fill="both", expand=True, padx=8, pady=8)
            txt.insert("end",
                "Columns that appear in 2+ files — likely foreign-key links:\n\n")
            for r in relationships:
                files = " · ".join(f["name"] for f in r["files"])
                examples = ", ".join(r["examples"][:4])
                txt.insert("end",
                    f"• {r['column']}\n"
                    f"    files:    {files}\n"
                    f"    examples: {examples}\n\n")
            txt.configure(state="disabled")

        # Footer — close
        bf = ttk.Frame(win)
        bf.pack(fill="x", padx=8, pady=(0, 8), side="bottom")
        ttk.Button(bf, text="Close", command=win.destroy).pack(side="right")

    def _council_find_and_chart(self, query: str = "") -> bool:
        """
        Find data files relevant to `query`, switch to the Grapher, load the
        top match, populate the AI-plot prompt with the original query, and
        ask the Analyst to chart it. Returns True if a file was found.
        """
        if not query:
            query = self.input.get("1.0", "end").strip()
        if not query:
            return False

        files = self._find_data_files_for_query(query, top_n=3)
        if not files:
            self._append_transcript(
                "Council",
                "I couldn't find any data files in the vault that match that "
                "question. Try the Vault tab to add a CSV, or use Grapher → "
                "Sample to explore with the bundled demo data first.",
                "final",
            )
            return False

        # Echo what we found into the transcript so the user can see the trail
        bullet_list = "\n  • ".join(p.name for p in files)
        self._append_transcript(
            "Council",
            f"Found {len(files)} relevant data file"
            f"{'s' if len(files) != 1 else ''}:\n  • {bullet_list}\n\n"
            f"Loading the top match into the Grapher…",
            "observation",
        )

        # Switch to Grapher and queue the work after the tab is realised
        if hasattr(self, "tab_grapher"):
            self.nb.select(self.tab_grapher)
        self.after(150, lambda fs=files, q=query: self._grapher_load_and_ask(fs, q))
        return True

    # ============================
    # Personal Specialists — runtime resolution
    # ============================

    def _spec_pin_refresh(self):
        """Repopulate the Ask: combobox values from the current registry."""
        if not hasattr(self, "_spec_pin_cb"):
            return
        items = self.specialists.all(enabled_only=True)
        values = ["Auto"] + [f"{s.icon} {s.name}" for s in items]
        self._spec_pin_cb["values"] = values
        # Keep current selection if still valid; otherwise reset to Auto
        if self._spec_pin_var.get() not in values:
            self._spec_pin_var.set("Auto")
            self._forced_specialist_id = None

    def _spec_pin_changed(self):
        """Handler for the Ask: dropdown."""
        choice = self._spec_pin_var.get()
        if choice == "Auto" or not choice:
            self._forced_specialist_id = None
            return
        # Find the matching specialist by display label
        for s in self.specialists.all():
            if f"{s.icon} {s.name}" == choice:
                self._forced_specialist_id = s.id
                return
        # Fallback — unknown choice, revert to auto
        self._forced_specialist_id = None
        self._spec_pin_var.set("Auto")

    def _resolve_active_specialists(self, query: str):
        """
        Decide which specialists are active for `query`. Returns a list of
        Specialist objects (may be empty). Logic:
          • If the user pinned a specialist via the manual dropdown, that
            takes precedence.
          • Otherwise, match keywords against the registry and return up to
            3 most relevant.
        """
        # Manual override
        forced_id = getattr(self, "_forced_specialist_id", None)
        if forced_id:
            spec = self.specialists.get(forced_id)
            return [spec] if spec else []

        # Auto-match — list of (specialist, score) tuples, already ranked
        matches = self.specialists.match(query, max_specialists=3)
        return [s for (s, _score) in matches]

    def _build_specialist_overlay(self, specs) -> str:
        """
        Compose context blocks from one or more specialists into a single
        injection string. Multi-specialist queries get a "you are
        deliberating with N specialists" header so the model knows it
        should reconcile perspectives rather than only adopting one.
        """
        if not specs:
            return ""
        if len(specs) == 1:
            return specs[0].context_block()
        header = (
            f"MULTI-SPECIALIST DELIBERATION  ({len(specs)} lenses active)\n"
            f"Apply each lens to the question; reconcile when they conflict.\n"
            f"Be explicit about which lens each part of the answer comes from.\n"
        )
        body = "\n\n──────────────────────────\n\n".join(
            s.context_block() for s in specs
        )
        return header + "\n" + body

    def _grapher_load_and_ask(self, files, query: str):
        """
        Helper: register `files` in the Grapher's combobox, load the first,
        and prompt the Analyst with `query`.
        """
        if not _GRAPHER_OK:
            self._append_transcript(
                "Council",
                "The Grapher module isn't available — install matplotlib + plotly "
                "to enable chart generation.", "final")
            return

        # Register all candidate files in the file dropdown so the user can
        # switch between them with one click if the top pick wasn't right.
        existing = list(getattr(self, "_grapher_file_cb", None) and
                        self._grapher_file_cb["values"] or [])
        for f in files:
            label = str(f)
            if label not in existing:
                existing.append(label)
                self._grapher_file_paths.append(f)
        if hasattr(self, "_grapher_file_cb"):
            self._grapher_file_cb["values"] = existing

        top = files[0]
        self._grapher_file_var.set(str(top))
        self._grapher_do_load(top)

        # Populate the AI-plot input with the user's question and run the
        # Analyst. We give the dataset a moment to finish loading first.
        def _ask_analyst():
            try:
                self._gai_prompt.delete("1.0", "end")
                self._gai_prompt.insert("1.0", query)
                # Force the analyst path (not the heavier "council" mode)
                if hasattr(self, "_gai_mode_var"):
                    try: self._gai_mode_var.set("analyst")
                    except Exception: pass
                self._grapher_ai_plot()
            except Exception as e:
                self._append_transcript("Council",
                    f"Analyst couldn't process the query: {e}", "final")

        self.after(400, _ask_analyst)

    # ============================
    # Main send logic
    # ============================

    def _maybe_suggest_model_swap(self, user_text: str) -> None:
        """Ask the swap advisor whether a specialist model would help this
        task enough to be worth the cost; if so, prompt the user once per
        target per session. On yes: LOCAL target swaps the model
        (GPU-gated); REMOTE target enables node dispatch so the dispatcher
        routes that role's model to the reachable node."""
        from tkinter import messagebox
        import os as _os
        try:
            import swap_advisor as _adv
            import role_models as _rm
        except Exception:
            return

        # Cheap, I/O-free gate: advise() returns None when classify_task does,
        # so skip the RoleModelRegistry disk read + dispatcher probe entirely
        # for the common case where the task matches no specialist.
        try:
            if _adv.classify_task(user_text) is None:
                return
        except Exception:
            pass

        # Build the advisor's view of the world from current state.
        current = _os.path.basename(_os.environ.get("COUNCIL_GGUF_PATH", "")) \
            or "the current model"
        assignments = {}
        try:
            assignments = _rm.RoleModelRegistry(VAULT_DIR).all()
        except Exception:
            pass
        # Remote specialists: roles whose assigned model is installed on a
        # reachable node. Derived from the dispatcher's probe (best-effort,
        # no extra network here — uses cached probe state).
        remote_specialists = {}
        try:
            if _os.environ.get("COUNCIL_REMOTE_NODES", "").strip().lower() in ("1", "true", "yes", "on"):
                statuses = [s for s in self.dispatcher.probe_all() if getattr(s, "reachable", False)]
                for role, model in assignments.items():
                    for s in statuses:
                        if any(model.split("/")[-1].split(".")[0] in m
                               for m in getattr(s, "installed_models", [])):
                            remote_specialists[role] = {
                                "model": model,
                                "node": getattr(s, "host", "a node"),
                                "label": f"{role} model on {getattr(s, 'host', 'a node')}"}
                            break
        except Exception:
            pass

        sugg = _adv.advise(
            user_text, current_model=current,
            role_assignments=assignments,
            remote_specialists=remote_specialists,
            gpu_swap_enabled=_rm.gpu_swap_enabled(),
        )
        if sugg is None:
            return
        # Don't nag: at most one prompt per (role, target_kind) per session.
        seen = getattr(self, "_swap_suggested", None)
        if seen is None:
            seen = self._swap_suggested = set()
        key = (sugg.role, sugg.target_kind, sugg.target_model)
        if key in seen:
            return
        seen.add(key)

        if not messagebox.askyesno(
                "Use a specialist model?",
                f"{sugg.reason}\n\nSwitch to {sugg.target_label}?\n"
                f"Cost: {sugg.est_cost}.",
                parent=self):
            return
        if sugg.target_kind == "remote":
            _os.environ["COUNCIL_REMOTE_NODES"] = "1"
            self._append_transcript(
                "Council", f"Routing {sugg.role} tasks to {sugg.target_label} "
                "(remote node). Local model stays loaded.", "observation")
        else:
            # swap_to_role() resets the engine's model singleton, so the
            # next inference lazy-loads the specialist — no personality
            # rebuild needed (they call the global GGUF singleton).
            res = _rm.swap_to_role(sugg.role, VAULT_DIR)
            if res.get("swapped"):
                self._append_transcript(
                    "Council", f"Switched to {sugg.target_label} "
                    f"({_os.path.basename(res.get('model', ''))}) — loads on the "
                    "next message.", "observation")
            else:
                self._append_transcript(
                    "Council", f"Couldn't switch ({res.get('reason')}); "
                    "continuing with the current model.", "observation")

    def _defer_to_vault(self):
        """Capture the current/last request as a DEFERRED TASK — something the
        council couldn't do easily in-chat — so the Vault tab can run it with
        the heavyweight deterministic tooling (or log a tool request). The
        question is pre-filled from the input box or the last user turn."""
        import tkinter as tk
        from tkinter import ttk, messagebox
        import deferred_tasks as _dt

        prefill = ""
        try:
            prefill = self.input.get("1.0", "end").strip()
        except Exception:
            prefill = ""
        if not prefill:
            prefill = getattr(self, "_last_user_text", "") or ""

        win = tk.Toplevel(self)
        win.title("Defer to Vault")
        win.transient(self)
        win.resizable(False, False)
        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, foreground="#888", justify="left", wraplength=440,
                  text="Save a task the council couldn't do easily in chat. "
                       "The Vault tab can run summaries/stats with the full "
                       "deterministic tooling, or hold tool requests for the "
                       "developer.").grid(row=0, column=0, columnspan=2,
                                          sticky="w", pady=(0, 10))

        ttk.Label(frm, text="What do you want done?").grid(
            row=1, column=0, sticky="nw", pady=3)
        q_txt = tk.Text(frm, width=46, height=3, wrap="word")
        q_txt.insert("1.0", prefill)
        q_txt.grid(row=1, column=1, sticky="w")

        ttk.Label(frm, text="Type:").grid(row=2, column=0, sticky="w", pady=3)
        kind_var = tk.StringVar(value="Bigger summary")
        kind_map = {
            "Bigger summary": _dt.KIND_BIGGER_SUMMARY,
            "Deeper stats": _dt.KIND_DEEPER_STATS,
            "Tool request (for the developer)": _dt.KIND_TOOL_REQUEST,
            "Other": _dt.KIND_OTHER,
        }
        ttk.Combobox(frm, textvariable=kind_var, width=32, state="readonly",
                     values=list(kind_map.keys())).grid(
            row=2, column=1, sticky="w")

        ttk.Label(frm, text="Note (optional):").grid(
            row=3, column=0, sticky="w", pady=3)
        note_var = tk.StringVar()
        ttk.Entry(frm, textvariable=note_var, width=46).grid(
            row=3, column=1, sticky="w")

        ttk.Label(frm, text="Folder (optional, under data_in):").grid(
            row=4, column=0, sticky="w", pady=3)
        folder_var = tk.StringVar()
        ttk.Entry(frm, textvariable=folder_var, width=46).grid(
            row=4, column=1, sticky="w")

        # ── Files to work on ─────────────────────────────────────
        # AUTO-DETECTED from the message (often wrong/incomplete), but the
        # user must confirm the scope: edit the list, add more files (a job
        # can span several files), or explicitly choose the whole folder.
        ttk.Label(frm, text="Files to work on\n(auto-detected — fix / add):",
                  justify="left").grid(row=5, column=0, sticky="nw", pady=3)
        files_frame = ttk.Frame(frm)
        files_frame.grid(row=5, column=1, sticky="w")

        files_lb = tk.Listbox(files_frame, height=4, width=34,
                              selectmode="extended", exportselection=False)
        files_lb.grid(row=0, column=0, rowspan=3, sticky="nw")
        # Auto-detect pre-fill from the question.
        try:
            import vault_analyst as _va_pf
            _pf = _va_pf.resolve_filename_hints(
                prefill, [data_index.input_dir(VAULT_DIR)])
            for _f in dict.fromkeys(r.name for _t, r in _pf if r is not None):
                files_lb.insert("end", _f)
        except Exception:
            pass
        # Every (non-hidden) file under data_in, for the Add dropdown.
        _avail = []
        try:
            import os as _os_av
            _ind = data_index.input_dir(VAULT_DIR)
            for _dp, _dn, _fn in _os_av.walk(str(_ind)):
                _dn[:] = [d for d in _dn
                          if not d.startswith(".") and d != "deferred_results"]
                _avail.extend(f for f in _fn if not f.startswith("."))
            _avail = sorted(set(_avail))
        except Exception:
            _avail = []
        add_var = tk.StringVar()
        add_cb = ttk.Combobox(files_frame, textvariable=add_var, values=_avail,
                              width=26, state="normal")
        add_cb.grid(row=0, column=1, sticky="w", padx=(6, 0))

        def _files_now():
            return list(files_lb.get(0, "end"))

        def _add_file(_e=None):
            v = add_var.get().strip()
            if v and v not in _files_now():
                files_lb.insert("end", v)
            add_var.set("")
        add_cb.bind("<Return>", _add_file)
        ttk.Button(files_frame, text="➕ Add", width=10,
                   command=_add_file).grid(row=1, column=1, sticky="w",
                                           padx=(6, 0))

        def _remove_sel():
            for i in reversed(files_lb.curselection()):
                files_lb.delete(i)
        ttk.Button(files_frame, text="✗ Remove", width=10,
                   command=_remove_sel).grid(row=2, column=1, sticky="nw",
                                             padx=(6, 0))

        whole_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            frm, variable=whole_var,
            text="Apply to the WHOLE folder instead (ignore the file list)"
        ).grid(row=6, column=1, sticky="w", pady=(2, 0))

        def _save():
            q = q_txt.get("1.0", "end").strip()
            if not q:
                messagebox.showwarning("Nothing to defer",
                                       "Describe what you want done first.")
                return
            if whole_var.get():
                files = []                       # explicit whole-folder choice
            else:
                files = [f.strip() for f in _files_now() if f.strip()]
                if not files:
                    messagebox.showwarning(
                        "Choose the files",
                        "Add at least one file to work on, or tick "
                        "“Apply to the WHOLE folder”.")
                    return
            try:
                _dt.DeferredTaskStore(VAULT_DIR).add(
                    kind=kind_map.get(kind_var.get(), _dt.KIND_OTHER),
                    question=q,
                    answer_excerpt=getattr(self, "_last_answer", "") or "",
                    files=files, folder=folder_var.get().strip(),
                    note=note_var.get().strip())
            except Exception as exc:
                messagebox.showerror("Save failed", f"{exc!r}")
                return
            try:
                self._append_transcript(
                    "Writer", "Saved to the Vault tab's deferred tasks — open "
                    "🗄 Vault → “Deferred tasks” to run it.", "observation")
            except Exception:
                pass
            # Refresh the Vault panel if it's been built.
            try:
                self._vmgr_refresh_deferred()
            except Exception:
                pass
            win.destroy()

        btns = ttk.Frame(frm)
        btns.grid(row=7, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(btns, text="Save", command=_save).pack(side="right")
        ttk.Button(btns, text="Cancel", command=win.destroy).pack(
            side="right", padx=6)

    def _send(self):
        # Licensing gate — skipped entirely in DEMO_MODE. In product
        # builds, this blocks new deliberations when trial expired and
        # no license active; past sessions are still readable.
        if (not getattr(branding, "DEMO_MODE", False)
                and not self._can_run_deliberation()):
            from tkinter import messagebox
            choice = messagebox.askyesno(
                "Activation required",
                f"{getattr(self, '_license_status', {}).get('message', 'License required')}\n\n"
                "You can still view your past sessions. To run new "
                "deliberations, activate a license now?",
                parent=self,
            )
            if choice:
                activation_dialog.open_activation_dialog(
                    self, VAULT_DIR,
                    on_status_change=self._on_license_status_change,
                    blocking=False,
                )
            return

        # Hide verdict feedback bar on new query
        self._vfb_hide()
        user_text = self.input.get("1.0", "end").strip()
        if not user_text:
            return
        # One-shot "Expand with council" bypass: capture + clear it here at the
        # very top so it can't leak past an earlier fast-path return (chart /
        # lookup) onto an unrelated later question.
        _force_full = bool(getattr(self, "_force_full_council", False))
        self._force_full_council = False
        # Reset per-turn provenance so a later answer never shows last turn's
        # source chips (set again after this turn's injection runs).
        self._last_turn_sources = []
        self._last_answer_table = ""   # set if this turn renders an analyst table
        self._last_sent_query = user_text   # saved for verdict disagree re-run
        self._set_text(self.input, "")
        self._append_transcript("User", user_text)
        # Log to the per-vault question history (browse + re-ask later).
        try:
            import question_history as _qh
            _qh.QuestionHistory(VAULT_DIR).add(user_text)
        except Exception:
            pass
        self._clear_stream_box()
        # ── Specialist-model suggestion ─────────────────────────────────
        # If this task looks like a fit for an assigned specialist and the
        # benefit is judged worth the cost, ask the user (once per target
        # per session) before proceeding. Best-effort; never blocks send.
        try:
            self._maybe_suggest_model_swap(user_text)
        except Exception as _adv_exc:
            print(f"[advisor] skipped: {_adv_exc!r}")
        # ── Pipeline intents (show / modify a Dream3D pipeline) ─────────
        # The handlers manage their own status: list/show are synchronous
        # (they set idle themselves on completion); modify is async (it
        # leaves status on "editing…" until the worker finishes).
        if self._handle_pipeline_intent(user_text):
            return

        # ── Workflow intents (run a sequence of pipelines) ──────────────
        # "run workflow A, B, C [on <dir> [per-file|per-step]]"
        if self._handle_workflow_intent(user_text):
            return  # status set inside (worker thread keeps status updated)

        # ── Vault ergonomic intents (stats, dupes, history) ─────────────
        if self._handle_vault_tools_intent(user_text):
            self._set_status("● idle")
            return

        # ── File-listing SAFETY NET ──────────────────────────────────────
        # When the user asks for a list of files in a folder — in ANY
        # phrasing — and the folder is a real path on disk, bypass the
        # model entirely and use the deterministic `_list_files_response`.
        # Background: `_LIST_FILES_RE` only catches strict phrasings like
        # "list files in X". Looser phrasings ("what's in X", "give me a
        # list of files in X", "can you show me everything in X") fell
        # through to the model, which would inject a [FOLDER:] block and
        # then paraphrase it — adding, dropping, or inventing files. The
        # deterministic handler cannot hallucinate.
        if self._LISTING_INTENT_RE.search(user_text):
            for _p_str in _extract_file_paths(user_text):
                try:
                    _pp = Path(_p_str.strip())
                except Exception:
                    continue
                if _pp.is_dir():
                    # Found a real directory mentioned in a listing-shaped
                    # question. Route to the deterministic handler.
                    self._list_files_response(str(_pp))
                    self._set_status("● idle")
                    return

        # ── 'forget X' command: reject prior fuzzy matches ─────────────
        # User can type "forget rockstar" or "forget rockstar, witcher" to
        # exclude those tokens from future fuzzy expansion. Persists per-vault.
        _forget_match = _re.match(r"^\s*forget\s+(.+)$", user_text, _re.IGNORECASE)
        if _forget_match:
            tokens = [t.strip() for t in _re.split(r"[,\s]+", _forget_match.group(1)) if t.strip()]
            idx = _get_vault_index()
            if idx is not None and tokens:
                added = idx.add_to_fuzzy_denylist(tokens)
                if added:
                    self._append_transcript(
                        "Council",
                        "OK — these will no longer be used as fuzzy matches: "
                        + ", ".join(added),
                        "observation",
                    )
                else:
                    self._append_transcript(
                        "Council",
                        "(Those terms were already on the fuzzy denylist.)",
                        "observation",
                    )
            else:
                self._append_transcript(
                    "Council",
                    "(Vault index unavailable — could not record fuzzy rejection.)",
                    "observation",
                )
            self._set_status("● idle")
            return

        # Inject file contents before routing so every council member sees the data.
        # Keep the ORIGINAL text for trigger detection — the augmented version
        # contains injected CSV/JSON content whose values can spuriously match
        # routing triggers (e.g. "plot" appearing in a game description sends
        # everything to the Grapher).
        original_user_text = user_text

        # ── Task-memo condense (runs FIRST, very fast) ─────────────────────
        # Re-condense the per-session memo from this turn's question so
        # the writer always sees a fresh goal+constraints+forbidden block
        # at the TOP of its context, even when the rest of the window is
        # eaten by file dumps. Post the result to the transcript so the
        # user can spot a misread before the model answers.
        try:
            def _condense_call(prompt: str) -> str:
                # Small fast call; temperature 0 for stable extraction.
                import council_engine as _ce
                return _ce.local_chat(
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.0,
                    num_predict=240,
                    timeout=45,
                )
            _memo = self.task_memory.update(original_user_text, llm_call=_condense_call)
        except Exception as _memo_exc:
            print(f"[TaskMemory] update failed: {_memo_exc!r}")
            _memo = self.task_memory.current()
        if _memo and not _memo.is_empty():
            self._append_transcript(
                "Council",
                self.task_memory.render_transcript_line(),
                "observation",
            )
        _task_memo_block = self.task_memory.render_injection_block() or None

        # ── Analyst step ────────────────────────────────────────────────
        # The analyst computes deterministic answers from data ("how many",
        # "what's the sum"); when it succeeds, the injection layer reduces
        # vault-match candidates from 5 to 1 (#7) so the answer doesn't
        # fight for budget with speculative matches. When it FAILS (e.g.
        # NameError in the generated code), we inject a failure block that
        # explicitly refuses any invented numeric answer instead of falling
        # back to model freeform — that's how the cross-machine
        # hallucination used to creep back in.
        _analyst_block, _analyst_err, _analyst_notices = _run_analyst_step(original_user_text)
        # Surface resolver notices FIRST so the user sees what the analyst
        # interpreted before the result / error message — that order makes
        # it obvious when the analyst grabbed the wrong file or scope.
        # Notices include two kinds:
        #   • Plain observation strings (resolver hints, scope notes, etc.)
        #   • A special "__ANALYST_TABLE__:<...>" payload carrying the
        #     user-facing rendering of the analyst's full DataFrame. We
        #     surface that as an "observation" too but with a clear
        #     "Analyst result table:" header so the user sees the
        #     actual numbers, not just the model's prose summary.
        # A direct-route headline (file-count / summary / stats / precomputed
        # / collection) that the fast-answer short-circuit below can render as
        # the final answer WITHOUT any model call — that's the O1 speed win.
        _fast_answer = None
        _analyst_sources: list = []
        for _note in (_analyst_notices or []):
            if isinstance(_note, str) and _note.startswith("__ANALYST_TABLE__:"):
                _table = _note[len("__ANALYST_TABLE__:"):]
                # Stash for "Save answer" (exported into the report).
                self._last_answer_table = _table
                self._append_transcript(
                    "Council",
                    "Analyst result table:\n" + _table,
                    "observation",
                )
            elif isinstance(_note, str) and _note.startswith("__ANALYST_ANSWER__:"):
                _fast_answer = _note[len("__ANALYST_ANSWER__:"):]
            elif isinstance(_note, str) and _note.startswith("__ANALYST_SOURCES__:"):
                try:
                    import json as _js_src
                    _analyst_sources = list(
                        _js_src.loads(_note[len("__ANALYST_SOURCES__:"):]) or [])
                except Exception:
                    _analyst_sources = []
            else:
                self._append_transcript("Council", _note, "observation")
        if _analyst_block and not _analyst_err:
            self._append_transcript("Council", "Computing from data...", "observation")
        elif _analyst_err:
            self._append_transcript(
                "Council",
                f"⚠ Analyst tried to compute the answer but failed: "
                f"{_analyst_err}\n"
                f"The model will be told to refuse inventing a number; "
                f"try rephrasing or check the file/column name.",
                "observation",
            )

        # ── Inject file/folder/vault context with token-aware caps ────────
        # The new pipeline (a) tags each block with its token cost, (b) caps
        # each block at n_ctx/8 tokens with a visible head+tail trim marker,
        # (c) drops the lowest-priority blocks (vault matches) when the
        # cumulative cost would exceed the safe input budget. The breakdown
        # dict is stashed for the `context info` chat intent.
        try:
            _n_ctx = ce.get_n_ctx()
        except Exception:
            _n_ctx = 4096
        # Resolve the current pinned-file set (filenames the model
        # referenced in a recent prior turn). Pins are stamped with the
        # turn-index they were created on and expire after a few turns
        # so the budget isn't permanently bloated by a one-off "show
        # me X" reference. The pin store lives on self; this turn's
        # number is just an incrementing counter.
        _pinned_files = self._current_pinned_filenames()
        augmented, fuzzy_matches, _injection_breakdown = _inject_file_contents(
            user_text, analyst_block=_analyst_block, n_ctx=_n_ctx,
            task_memo_block=_task_memo_block,
            pinned_files=_pinned_files,
        )
        self._last_injection_breakdown = _injection_breakdown
        # Surface defensive-wrapper failures to the transcript. When the
        # injection pipeline's top-level try/except fires, the breakdown
        # carries an `injection_error` field — without surfacing it, the
        # user just sees a generic answer and has no idea context
        # retrieval failed silently.
        _inj_err = _injection_breakdown.get("injection_error") if isinstance(_injection_breakdown, dict) else None
        if _inj_err:
            self._append_transcript(
                "Council",
                f"⚠ Context retrieval failed with an unexpected error: "
                f"{_inj_err}\n"
                f"The model has been told to treat this turn as no-data-"
                f"provided and will refuse to invent specific values.",
                "observation",
            )
        # ``was_injected`` is the legacy gate that suppresses the fast-path
        # chart/lookup handlers below. The OLD pipeline only injected vault
        # hits + explicit files (analyst was a separate post-step), so this
        # flag meant "the model is being asked to reason about ALREADY
        # FETCHED data — don't divert to the chart path." Now that analyst
        # is folded into the injection pipeline, we have to recompute the
        # flag to EXCLUDE analyst-only injections, otherwise a plain "how
        # many X" question would block the chart fast path even though no
        # vault context exists yet.
        _labels = [lbl for lbl, _ in _injection_breakdown.get("costs", [])]
        _has_vault = any(lbl.startswith("[VAULT MATCH") for lbl in _labels)
        _has_file = any(lbl.startswith("[FILE:") for lbl in _labels)
        _has_folder = any(lbl.startswith("[FOLDER:") for lbl in _labels)
        _has_nodata = any(lbl.startswith("[NO DATA AVAILABLE") for lbl in _labels)
        was_injected = bool(_has_vault or _has_file or _has_folder or _has_nodata)
        # ── Provenance: the source files behind this turn's answer ──────────
        # Assembled from (a) the analyst's own sources (derived result inputs,
        # collection members), (b) vault-match / file / folder injection
        # labels, and (c) files the user named explicitly. Rendered as
        # clickable chips under the final answer (fast path + deliberation).
        # Explicit paths the user named — computed ONCE and reused at all three
        # post-injection sites below (source assembly, injected-names notice,
        # provenance). original_user_text == user_text here (user_text isn't
        # reassigned to the augmented text until later), and _extract_file_paths
        # is a pure function of its text, so one call is byte-identical to three
        # (each did a full unicodedata scan + regex + per-path resolve()).
        try:
            _explicit_paths_turn = _extract_file_paths(original_user_text)
        except Exception:
            _explicit_paths_turn = []
        _turn_sources: list = list(_analyst_sources)
        for _lbl in _labels:
            _m = _SOURCE_LABEL_RE.match(_lbl)
            if _m:
                _turn_sources.append(_m.group(1).strip())
        _turn_sources.extend(_explicit_paths_turn)
        self._last_turn_sources = _turn_sources
        if was_injected:
            # Filter out empty names (Path("/").name == "" etc) so we don't
            # render "Reading: , , " on edge-case paths.
            injected_names = [n for n in (Path(p).name for p in _explicit_paths_turn) if n]
            if injected_names:
                self._append_transcript("Council", "Reading: " + ", ".join(injected_names), "observation")
            elif _has_vault:
                self._append_transcript("Council", "Reviewing matching vault files...", "observation")
            # If only [ANALYST RESULT] was injected, the earlier "Computing
            # from data..." line already informed the user — no extra
            # transcript noise needed.
        if fuzzy_matches:
            _lines = ["Fuzzy matches used (treated as your spelling):"]
            for orig, suggestions in fuzzy_matches.items():
                pairs = ", ".join(f"{m} ({r:.2f})" for m, r in suggestions)
                _lines.append(f"  • {orig} → {pairs}")
            _lines.append(
                "If any are wrong, type:  forget WORD  "
                "(comma-separated for multiple) — they won't be used again."
            )
            self._append_transcript("Council", "\n".join(_lines), "observation")
        # Surface dropped blocks so the user knows what got cut to fit budget.
        _dropped = _injection_breakdown.get("dropped") or []
        if _dropped:
            _msg_lines = [
                f"⚠ Dropped {len(_dropped)} lower-priority block"
                f"{'s' if len(_dropped) != 1 else ''} to fit n_ctx="
                f"{_injection_breakdown.get('n_ctx', 0):,}:"
            ]
            for _label, _cost in _dropped:
                _msg_lines.append(f"  • {_label}  (~{_cost:,} tokens)")
            _msg_lines.append(
                "(Analyst result, task memo, and the NO-DATA marker are "
                "never dropped. Explicit files / folder listings / vault "
                "summary CAN be dropped on cumulative budget overflow — "
                "raise COUNCIL_GGUF_N_CTX to keep them. Type 'context "
                "info' for the full token breakdown.)"
            )
            self._append_transcript("Council", "\n".join(_msg_lines), "observation")
        user_text = augmented

        # ── Context-window check ──────────────────────────────────────────
        # llama-cpp-python silently truncates any prompt that doesn't fit in
        # `n_ctx` — it clips the tail and proceeds, which is the #1 cause of
        # "the model hallucinated even though I gave it the file." Warn the
        # user as soon as we cross 80% of the configured window so they can
        # raise COUNCIL_GGUF_N_CTX before sending another query. We don't
        # block the request — the warning goes to the transcript and the
        # call still goes through (truncated output is still useful for
        # diagnosing what got clipped).
        try:
            _budget = ce.context_budget_report(user_text)
            self._last_context_budget = _budget
            _used = _budget["input_tokens"]
            _ctx  = _budget["n_ctx"]
            _safe = _budget["safe_input"]
            if _budget["over_window"]:
                _maxh = _budget.get("model_max_ctx") or 0
                _hint = (
                    f" The model supports up to {_maxh} tokens — raising "
                    f"COUNCIL_GGUF_N_CTX to {min(max(16384, _ctx * 4), _maxh)} "
                    f"would fit this prompt."
                ) if _maxh and _maxh > _ctx else (
                    " Raise COUNCIL_GGUF_N_CTX (e.g. to 16384 or 32768) "
                    "before launch."
                )
                self._append_transcript(
                    "Council",
                    f"⚠ Prompt is ~{_used:,} tokens but n_ctx is only {_ctx:,}. "
                    f"The model will silently truncate the tail and may miss "
                    f"the injected file/analyst data.{_hint}",
                    "observation",
                )
            elif _used > _safe:
                _pct = _budget["pct_of_window"]
                self._append_transcript(
                    "Council",
                    f"⚠ Prompt is ~{_used:,} tokens ({_pct:.0f}% of "
                    f"n_ctx={_ctx:,}). Reply space is tight — consider "
                    f"raising COUNCIL_GGUF_N_CTX for fuller answers.",
                    "observation",
                )
        except Exception as _e:
            print(f"[ContextBudget] check failed: {_e!r}")

        # ── Provenance: remember what we showed the model this turn ──
        # Lets the user ask "where did $5,000 come from?" after the reply.
        try:
            if hasattr(self, "provenance"):
                import provenance as _prov_mod
                injected_records = []
                # Recorded file paths (explicit user-given paths) — reuse the
                # once-computed list; _read_file_for_injection is now memoized
                # so this doesn't re-parse files injection already read.
                for p_str in _explicit_paths_turn:
                    block = _read_file_for_injection(p_str)
                    if block:
                        injected_records.append(_prov_mod.InjectedBlock(
                            file_name=Path(p_str).name,
                            file_path=p_str,
                            block=block,
                        ))
                # If the analyst step ran, treat its result as another source
                if _analyst_block:
                    injected_records.append(_prov_mod.InjectedBlock(
                        file_name="(analyst result)",
                        file_path="(in-session)",
                        block=_analyst_block,
                    ))
                self.provenance.record_turn(
                    user_text=original_user_text,
                    augmented_text=user_text,
                    injected_files=injected_records,
                )
        except Exception as _e:
            print(f"[Provenance] record failed: {_e!r}")

        # ── Fast path: chart-shaped questions go straight to the Grapher ──
        # Skip this entirely when we just injected vault context — the user is
        # asking the council to reason about the data, not graph it.
        if not was_injected and self._looks_like_data_question(original_user_text):
            handled = self._council_find_and_chart(original_user_text)
            if handled:
                self._set_status("● grapher", "#a6e3a1")
                return
            # else fall through to regular deliberation

        # Look-up shaped questions skip the deliberation in favour of a
        # fast deterministic search across the data index. Skipped when we
        # already injected vault context for the same reason as above.
        if not was_injected and self._looks_like_lookup_question(original_user_text):
            self._council_run_lookup(original_user_text)
            self._set_status("● lookup", "#a6e3a1")
            return

        # ── Fast path: a deterministic analyst answer needs no council ─────
        # When the analyst direct-routed (file count, data/stats summary, a
        # fresh precomputed/derived result, or a named collection), the answer
        # is already computed exactly — running the full multi-role
        # deliberation + Writer synthesis on top only adds model latency (the
        # O1 slowness). Render the answer immediately and offer an explicit
        # "Expand with the council" button for a fuller, prose discussion.
        # One-shot bypass: the Expand button set _force_full_council, captured
        # into _force_full at the top of _send.
        if (_fast_answer and not _analyst_err and not _force_full
                and getattr(self, "_fast_answers_enabled", True)):
            self._append_transcript("Writer", _fast_answer, "final")
            _fsrc = getattr(self, "_last_turn_sources", None)
            if _fsrc:
                self._render_source_chips(_fsrc)
            # Remember the question so the Expand button can re-ask it through
            # the full council without the user retyping anything.
            self._last_fast_question = original_user_text
            try:
                self._expand_btn.configure(state="normal")
            except Exception:
                pass
            self._append_transcript(
                "Council",
                "Answered directly from the data (no council deliberation, so "
                "it's instant). Click “⤢ Expand with council” for a fuller "
                "discussion.",
                "observation",
            )
            self._set_status("● direct", "#a6e3a1")
            return

        # ── Personal Specialists: detect which (if any) to summon ──────
        # In single-voice mode we skip specialists entirely — only the Writer
        # is heard, and the Writer already has every piece of injected context.
        if _single_voice_mode():
            self._active_specialists = []
        else:
            self._active_specialists = self._resolve_active_specialists(original_user_text)
            if self._active_specialists:
                names = ", ".join(f"{s.icon} {s.name}" for s in self._active_specialists)
                self._append_transcript(
                    "Council",
                    f"Consulting: {names}",
                    "observation",
                )

        # ── Phase 1: Route (keyword-based, instant) ───────────────────────
        # In single-voice mode, force the writer route and don't surface the
        # judge's routing decision to the user.
        if _single_voice_mode():
            route = "writer"
            self._set_status("● writer…", "#fab387")
        else:
            route = self.judge.route(user_text)
            self._set_judge(f"Route: {route}\n")
            self._set_status(f"● {route}…", "#fab387")

        # Tabs only present in advanced mode — fall through to deliberation
        # if the user is on the consumer build (no exposed admin tabs).
        if route == "apothecary":
            if hasattr(self, "tab_apoth"):
                self._append_transcript("Judge", "Routing to Apothecary tab.", "final")
                self.nb.select(self.tab_apoth)
                self._set_status("● idle")
                return
            # else: deliberation handles maintenance question conversationally
        if route == "speech":
            self._append_transcript("Judge", "Routing to Speech tab.", "final")
            self.nb.select(self.tab_speech)
            self._set_status("● idle")
            return
        if route == "librarian":
            if hasattr(self, "tab_lib"):
                self._append_transcript("Judge", "Routing to Librarian tab.", "final")
                self.nb.select(self.tab_lib)
                self._set_status("● idle")
                return
        if route == "ide":
            if hasattr(self, "tab_ide"):
                self.nb.select(self.tab_ide)

        # ── Phase 2: Panel selection (before worker spawns) ────────────
        # In single-voice mode the panel is just the Writer — no other roles
        # deliberate, the Writer synthesizes directly from the injected context.
        if _single_voice_mode():
            _keyword_panel, _synth_role = ["writer"], "writer"
        else:
            _keyword_panel, _synth_role = _panel_for_route(route)
            if self.var_judge_panel.get():
                # Judge model picks roles — brief synchronous call in main thread.
                self._set_status("● judge routing…", "#cba6f7")
                try:
                    _judge_chosen = self.judge.choose_panel(user_text)
                    if _judge_chosen:
                        _keyword_panel = _judge_chosen
                        self._set_judge(f"Judge panel: {_judge_chosen}\n")
                        self._set_status(f"● panel: {_judge_chosen}…", "#fab387")
                except Exception as _jpe:
                    self._set_judge(f"Judge panel failed ({_jpe}), using keyword panel\n")

        # ── Personality-lead override ────────────────────────────
        # If the user names a specific personality ("with the writer as lead",
        # "focus on coder", "have the sage take the lead"), reorder the
        # panel so that personality goes first and becomes the synth role.
        _LEAD_ALIASES = {
            "writer":     "writer",
            "coder": "coder",
            "intern":     "intern",
            "peasant":    "peasant",
            "artist":     "artist",
            "skeptic":    "skeptic",
            "sage":       "sage",
            "strategist": "strategist",
            "librarian":  "librarian",
            "musician":   "musician",
            "content":    "content",  "content creator": "content",
            "director":   "director",
        }
        _lead_role = None
        _t_lower = user_text.lower()
        _lead_phrases = [
            "as the lead", "as lead", "take the lead", "as main", "as the main",
            "as focus", "focus on", "with the", "have the", "personality as",
            "personality to", "personality taking",
        ]
        if any(ph in _t_lower for ph in _lead_phrases):
            for alias, role in _LEAD_ALIASES.items():
                if alias in _t_lower:
                    _lead_role = role
                    break

        if _lead_role and _lead_role in self.personalities:
            # Put lead role first, keep others, set as synth
            _others = [r for r in _keyword_panel if r != _lead_role]
            _keyword_panel = [_lead_role] + _others
            _synth_role = _lead_role
            self._set_judge("Lead role override: " + _lead_role + "\n")

        # _resolved_panel is a closure variable — the worker reads it directly.
        # It is set once here and never mutated inside the worker.
        _resolved_panel: list = _keyword_panel
        _resolved_synth: str  = _synth_role

        use_stream = bool(self.var_stream.get())

        def _token_cb(who: str, token: str):
            """Called from the worker thread — post to UI queue."""
            if token.startswith("\x00tps:"):
                # T3-C: timing sentinel from done packet
                try:
                    _tps_val = float(token[5:])
                    self.ui_q.put(("tps_update", who, _tps_val))
                except Exception:
                    pass
                return
            if use_stream:
                self.ui_q.put(("stream_token", who, token))

        def worker():
            # ── Crash-recovery sentinel ─────────────────────────────
            # Mark this session as in-flight; clear on clean exit. If the
            # app crashes between these two calls, find_orphaned_sessions()
            # will surface it on next launch so the user can resume.
            try:
                self.convo_store.mark_session_active(self.session_id, user_text)
            except Exception:
                pass
            try:
                use_deliberation  = bool(self.var_deliberate.get())  if hasattr(self, "var_deliberate")  else True
                use_adversarial   = bool(self.var_adversarial.get()) if hasattr(self, "var_adversarial") else False

                # DEMO_MODE forces single-personality direct mode regardless
                # of the toggle — the home build is "ask a question, get an
                # answer", not a multi-AI deliberation.
                if getattr(branding, "DEMO_MODE", False):
                    use_deliberation = False

                # ── Fast path: Deliberation toggle OFF ─────────────────
                # Skips the full orchestrator → direct Writer response with
                # any active Personal Specialist lens applied as extra
                # context. Quick Q&A, no panel debate.
                if not use_deliberation:
                    self.ui_q.put(("agent_phase", "direct",
                                   "▶ Direct mode"))
                    extra = ""
                    active_specs = list(getattr(self, "_active_specialists", []) or [])
                    if active_specs:
                        try:
                            extra = self._build_specialist_overlay(active_specs)
                        except Exception:
                            extra = ""
                    answer = self.writer.respond(user_text, extra_context=extra)
                    ev = AgentEvent("Writer", "final", answer)
                    self.ui_q.put(("live_event", ev))
                    self.ui_q.put(("judge_final", ""))
                    self.ui_q.put(("done", None))
                    return

                enable_tools = bool(self.var_tools.get())
                use_tp_agent  = bool(self.var_use_coder_agent.get()) if hasattr(self, "var_use_coder_agent") else False
                use_in_agent  = bool(self.var_use_intern_agent.get())     if hasattr(self, "var_use_intern_agent")     else False
                use_rag       = bool(self.var_use_rag.get())              if hasattr(self, "var_use_rag")              else False
                tools = _make_tools(self.runner, self.librarian, VAULT_DIR) if enable_tools else {}

                # ── Librarian proactive briefing ───────────────────
                # Searches vault on multiple angles BEFORE deliberation
                # so every personality gets relevant context, not just Writer.
                lib_brief = {"raw": "", "peasant": "", "summary": "", "sources": [], "found": False}
                rag_context = ""
                if use_rag and self.rag:
                    try:
                        lib_brief = _librarian_brief(
                            self.rag,
                            user_text,
                            log_cb=lambda m: self.ui_q.put(("agent_phase", "rag_search", m)),
                            max_chars=5500,  # desktop: 8K context window allows larger briefing
                        )
                        rag_context = lib_brief["raw"]
                        if lib_brief["found"]:
                            self.ui_q.put(("agent_phase", "rag_search",
                                           f"RAG retrieved context ({len(rag_context)} chars)"))
                        else:
                            self.ui_q.put(("agent_phase", "rag_search",
                                           "RAG: no relevant vault context found"))
                    except Exception as e:
                        self.ui_q.put(("agent_phase", "rag_search", f"RAG error: {e}"))

                # Log queries that returned no vault context
                if use_rag and not lib_brief.get("found", True):
                    try:
                        _miss_path = VAULT_DIR / "vault_rag_misses.txt"
                        with open(_miss_path, "a", encoding="utf-8") as _mf:
                            _mf.write(now_iso() + "\t" + user_text[:200].replace("\n", " ") + "\n")
                    except Exception:
                        pass

                # ── Coder agent wrapper ───────────────────────
                class _CoderWrapper:
                    """Makes CoderAgent look like a ModelAgent for the orchestrator."""
                    display_name = "Coder"
                    def __init__(self_, agent):
                        self_.agent = agent
                        # Expose .model so orchestrator rebuttal/cross-fire can call .model.respond()
                        self_.model = agent.model
                    def _make_token_cb(self_):
                        return None
                    def act(self_, ctx):
                        state = self_.agent.run(ctx.user_text)
                        evs = [AgentEvent("Coder", "final",
                                          f"{state.final_code}\n\n{state.explanation}")]
                        if state.passed:
                            evs.insert(0, AgentEvent("Coder", "observation",
                                                     f"✓ Passed in {state.attempt} attempt(s)"))
                        else:
                            evs.insert(0, AgentEvent("Coder", "observation",
                                                     f"⚠ Best effort after {state.attempt} attempts"))
                        ctx.shared.setdefault("coder_code", state.final_code)
                        return evs

                # ── Intern agent wrapper ───────────────────────────
                class _InternWrapper:
                    """Makes InternAgent look like a ModelAgent for the orchestrator."""
                    display_name = "Intern"
                    def __init__(self_, agent):
                        self_.agent = agent
                        # Expose .model so orchestrator rebuttal/cross-fire can call .model.respond()
                        self_.model = agent.model
                    def _make_token_cb(self_):
                        return None
                    def act(self_, ctx):
                        draft = self_.agent.run(ctx.user_text)
                        evs = [AgentEvent("Intern", "final", draft.draft)]
                        if draft.needed_research and draft.research:
                            evs.insert(0, AgentEvent("Intern", "observation",
                                                     f"🔎 Researched: {draft.research.query}\n"
                                                     f"Sources: {', '.join(draft.research.urls_tried[:3])}"))
                        return evs

                # ── Build agents dict ──────────────────────────────
                # Gate agent wrappers by resolved panel membership.
                # If coder/intern are not in the panel, build them as plain
                # ModelAgents — no Dream3D static analyser, no web research loop.
                # This is the critical guard: "What is the sun?" will never build
                # a CoderWrapper because "coder" won't be in _resolved_panel.
                _tp_in_panel = "coder" in _resolved_panel
                _in_in_panel = "intern"     in _resolved_panel

                if use_tp_agent and self.coder_agent and _tp_in_panel:
                    coder_slot = _CoderWrapper(self.coder_agent)
                else:
                    coder_slot = ModelAgent("Coder", self.coder,
                                                 tools=tools, enable_tools=enable_tools,
                                                 token_callback=_token_cb)

                if use_in_agent and self.intern_agent and _in_in_panel:
                    intern_slot = _InternWrapper(self.intern_agent)
                else:
                    intern_slot = ModelAgent("Intern", self.intern,
                                            tools=tools, enable_tools=enable_tools,
                                            token_callback=_token_cb)

                # ── Inject Librarian briefing into every personality ──
                # Each role gets context appropriate to its job:
                #   Writer / Coder / Intern / Artist → full vault briefing
                #   Peasant → targeted briefing that guides specific questions
                #
                # We patch .respond() on each model for the duration of this
                # request, then restore originals after deliberation.
                _patched_models = {}

                # Prepend any active council-wide instructions to every patch
                _ci = self._instr_mgr.active_text()

                def _patch_model(model, extra: str, role_name: str):
                    """Wrap model.respond to prepend vault context for this request."""
                    if model is None or not extra:
                        return
                    orig = model.respond
                    # Store (model, original_fn) so the finally block can restore
                    # without needing to re-resolve the model by role name.
                    _patched_models[role_name] = (model, orig)
                    def _patched(prompt, **kwargs):
                        existing = kwargs.get("extra_context", "")
                        _ci_block = ("COUNCIL INSTRUCTIONS:\n" + _ci + "\n\n") if _ci else ""
                        kwargs["extra_context"] = (_ci_block + extra + "\n\n" + existing).strip()
                        return orig(prompt, **kwargs)
                    model.respond = _patched

                # Always apply council instructions even if no vault context
                if _ci and not lib_brief["found"]:
                    _ci_block = "COUNCIL INSTRUCTIONS:\n" + _ci
                    for _pm in [self.writer, self.coder, self.intern,
                                self.artist, self.sage, self.strategist,
                                self.content, self.director]:
                        if _pm is not None:
                            _patch_model(_pm, _ci_block, "ci_only")
                    if self.skeptic is not None:
                        _patch_model(self.skeptic, _ci_block, "ci_only")

                # Personal Specialists — patch the lens overlay onto each
                # specialist's base personality. If two specialists share a
                # base (e.g. both default to "writer"), they both end up in
                # the overlay and the model is told to reconcile them.
                _active_specs = list(getattr(self, "_active_specialists", []) or [])
                if _active_specs:
                    # Group by base personality so each model gets a single
                    # combined overlay rather than being patched repeatedly.
                    by_base = {}
                    for sp in _active_specs:
                        by_base.setdefault(sp.base_personality or "writer", []).append(sp)
                    for base_role, sp_list in by_base.items():
                        target_model = self.personalities.get(base_role) or self.writer
                        overlay = self._build_specialist_overlay(sp_list)
                        if target_model is not None and overlay:
                            _patch_model(target_model, overlay,
                                          f"specialist_{base_role}")

                if lib_brief["found"]:
                    # ── Run Librarian personality over raw RAG results ──────────
                    # The Librarian interprets the raw vault context, ranks it,
                    # produces a structured ACCESS LIST for all personalities,
                    # and emits WISHLIST_ENTRY lines for any gaps it finds.
                    _lib_personality = getattr(self, "librarian_personality", None)
                    _librarian_briefing = lib_brief["raw"]  # fallback: raw context
                    if _lib_personality is not None:
                        try:
                            # Inject current wishlist so Librarian avoids duplicate entries
                            _current_wishlist = self.librarian.get_wishlist()
                            _lib_query = (
                                f"VAULT QUERY: {user_text[:300]}\n\n"
                                f"RAW VAULT RESULTS:\n{lib_brief['raw'][:3000]}\n\n"
                                f"VAULT WISHLIST (already logged — avoid duplicates):\n"
                                f"{_current_wishlist[:1000]}"
                            )
                            _lib_response = _lib_personality.respond(_lib_query)

                            # Parse and persist any WISHLIST_ENTRY lines the Librarian emitted
                            for _wline in _lib_response.splitlines():
                                if _wline.strip().startswith("WISHLIST_ENTRY"):
                                    # Format: WISHLIST_ENTRY | <who> | <topic> | <reason>
                                    _parts = [p.strip() for p in _wline.split("|")]
                                    if len(_parts) >= 4:
                                        self.librarian.log_gap(
                                            who=_parts[1],
                                            topic=_parts[2],
                                            reason=_parts[3],
                                        )

                            # Strip WISHLIST_ENTRY and PANEL_ADD lines before
                            # sending the briefing to other models — those are
                            # system-level directives, not council context.
                            _panel_additions: list = []
                            _clean_lines = []
                            for _ln in _lib_response.splitlines():
                                _stripped = _ln.strip()
                                if _stripped.startswith("WISHLIST_ENTRY"):
                                    pass  # already handled above
                                elif _stripped.startswith("PANEL_ADD:"):
                                    # ── #6 Dynamic panel expansion ──────────
                                    _suggested = _stripped.split(":", 1)[-1].strip().lower()
                                    _valid_addable = {
                                        "writer", "coder", "intern", "sage",
                                        "strategist", "artist", "musician",
                                        "content", "director", "peasant",
                                    }
                                    if (_suggested in _valid_addable
                                            and _suggested not in _resolved_panel
                                            and getattr(self, _suggested, None) is not None):
                                        _panel_additions.append(_suggested)
                                else:
                                    _clean_lines.append(_ln)
                            _clean_response = "\n".join(_clean_lines)
                            # Use Librarian's structured output as the briefing
                            _librarian_briefing = (
                                "LIBRARIAN ACCESS LIST:\n"
                                + _clean_response
                            )
                            if _panel_additions:
                                self.ui_q.put(("agent_phase", "librarian",
                                               f"Panel expanded: +{', '.join(_panel_additions)}"))
                            self.ui_q.put(("agent_phase", "librarian",
                                           f"Librarian indexed {len(lib_brief['sources'])} sources"))
                        except Exception as _le:
                            self.ui_q.put(("agent_phase", "librarian",
                                           f"Librarian indexing failed ({_le}), using raw context"))

                    # Inject Librarian briefing only into roles that actually use vault.
                    # Roles with use_vault="none" would discard it immediately in respond() —
                    # skipping them avoids constructing context that gets thrown away.
                    _full_vault_roles  = ("writer", "coder", "sage", "strategist",
                                          "content", "director", "librarian")
                    _lite_vault_roles  = ("peasant",)   # gets targeted briefing below
                    # intern, artist, musician, skeptic → use_vault="none", skip entirely

                    for _role_name in _full_vault_roles:
                        _pm = getattr(self, _role_name, None)
                        if _pm is not None:
                            _patch_model(_pm, _librarian_briefing, _role_name)
                    # Peasant gets the shorter targeted briefing
                    _patch_model(self.peasant, lib_brief["peasant"], "peasant")

                    # Also patch underlying models in agent wrappers
                    if hasattr(coder_slot, "model") and coder_slot.model is not self.coder:
                        _patch_model(coder_slot.model, _librarian_briefing, "coder_agent_model")
                    if hasattr(intern_slot, "model") and intern_slot.model is not self.intern:
                        _patch_model(intern_slot.model, _librarian_briefing, "intern_agent_model")

                agents = {
                    "writer":     ModelAgent("Writer",   self.writer,  enable_tools=False, token_callback=_token_cb),
                    "peasant":    ModelAgent("Peasant",  self.peasant, enable_tools=False, token_callback=_token_cb),
                    "intern":     intern_slot,
                    "coder": coder_slot,
                    "artist":     ModelAgent("Artist",   self.artist,  enable_tools=False, token_callback=_token_cb),
                }
                if self.skeptic is not None:
                    agents["skeptic"] = ModelAgent("Skeptic", self.skeptic, enable_tools=False, token_callback=_token_cb)
                # Sage: use SageAgent wrapper if available so it gets knowledge injection
                if getattr(self, "sage_agent_obj", None) is not None:
                    agents["sage"] = ModelAgent("Sage", self.sage_agent_obj.model,
                                                enable_tools=False, token_callback=_token_cb)
                elif self.sage is not None:
                    agents["sage"] = ModelAgent("Sage", self.sage, enable_tools=False, token_callback=_token_cb)
                # Strategist — add if personality model is initialised
                if getattr(self, "strategist", None) is not None:
                    agents["strategist"] = ModelAgent("Strategist", self.strategist, enable_tools=False, token_callback=_token_cb)
                if getattr(self, "content", None) is not None:
                    agents["content"] = ModelAgent("Content", self.content, enable_tools=False, token_callback=_token_cb)
                if getattr(self, "director", None) is not None:
                    agents["director"] = ModelAgent("Director", self.director, enable_tools=False, token_callback=_token_cb)

                # T2-D: Apply per-query model overrides
                _override_map = {
                    "writer": self.writer, "coder": self.coder,
                    "intern": self.intern, "artist": self.artist,
                    "skeptic": self.skeptic,
                }
                _override_restores = {}
                if hasattr(self, "_query_overrides"):
                    for _orole, _ovar in self._query_overrides.items():
                        _oval = _ovar.get()
                        if _oval != "(default)" and _orole in _override_map:
                            _model = _override_map[_orole]
                            if _model is not None:
                                _override_restores[_orole] = _model.backend_key
                                _model.backend_key = _oval

                def _ev_cb(ev: AgentEvent):
                    self.ui_q.put(("live_event", ev))

                # Panel was resolved in the main thread before this worker started.
                # Use it directly — but add any Librarian-recommended expansions.
                _lib_expanded = _resolved_panel + locals().get("_panel_additions", [])
                panel      = [p for p in _lib_expanded if p in agents]
                synth_role = _resolved_synth
                if not panel:
                    panel = ["writer", "peasant"]  # safe minimum fallback

                def _clarif_cb(who: str, question: str):
                    self.ui_q.put(("clarification_needed", who, question))

                def _answer_getter() -> str:
                    ans = self._clarification_answer
                    self._clarification_answer = ""
                    return ans

                orch = DeliberationOrchestrator(
                    judge_model=self.judge, agents=agents,
                    max_rounds=1, debate_turns=1,
                    # Adaptive gates in the orchestrator handle the rest:
                    #   confidence ≥8 on round 0 → accepts NEEDS_WORK, exits early
                    #   confidence ≤2 on round 0 → escalates max_rounds to 2
                    #   Verdict: PASS always exits immediately
                    event_callback=_ev_cb,
                    clarification_cb=_clarif_cb,
                    pause_event=self._pause_event,
                    answer_getter=_answer_getter,
                )
                _code_routes = {"ide", "coder", "intern"}
                _content_routes = {"content", "writer", "chat", "musician"}
                _query_mode  = "technical" if route in _code_routes else "conversational"
                _latex_mode  = _detect_latex_request(user_text)
                # Inject content style memory and templates for content routes
                if route in ("content", "director") and getattr(self, "_content_style", None):
                    _style_ctx = self._content_style.build_context_block(user_text)
                    if _style_ctx:
                        for _style_role in ("content", "director"):
                            _style_pm = getattr(self, _style_role, None)
                            if _style_pm is not None:
                                _patch_model(_style_pm, _style_ctx, f"{_style_role}_style")
                        self.ui_q.put(("agent_phase", "content_style",
                                       "Content style memory + template injected"))
                # ── #5 Per-route temperature adaptation ──────────────────────────
                # Temperatures are fixed per role at build time but routes have very
                # different needs. Save originals, apply overrides, restore in finally.
                _temp_overrides = ce._ROUTE_TEMP_OVERRIDES.get(route, {})
                _temp_restores: dict = {}
                for _tr_role, _tr_temp in _temp_overrides.items():
                    _tr_model = getattr(self, _tr_role, None)
                    if _tr_model is not None and _tr_role in panel:
                        _temp_restores[_tr_role] = _tr_model.temperature
                        _tr_model.temperature = _tr_temp

                _orch_extra  = {
                    "peasant_adversarial": use_adversarial,
                    "query_mode":          _query_mode,
                    "latex_mode":          _latex_mode,
                }
                # If LaTeX requested, prepend instruction to writer context
                if _latex_mode and self.writer:
                    _latex_inst = (
                        "LATEX MODE: The user has requested LaTeX output.\n"
                        "Write your response as clean prose — do NOT include LaTeX markup.\n"
                        "The system will automatically wrap your content in a LaTeX document.\n"
                        "Focus on well-structured content: clear sections, proper paragraphs.\n"
                        "Do not use markdown headers or bullet points — use prose paragraphs.\n"
                    )
                    _patch_model(self.writer, _latex_inst, "writer_latex")

                # ── #4 Director→Writer style-locked pipeline ─────────────────────
                # When the director route is active AND the user is asking for a
                # script/draft, run Director first for a style brief, then inject
                # that brief into Writer before the deliberation begins.
                # This means Writer gets an authoritative style fingerprint to lock
                # onto, not just generic content memory — the result sounds like the
                # user because it IS structured around their patterns.
                _script_keywords = {
                    "script", "write", "draft", "narrate", "narration",
                    "intro", "outro", "hook", "voiceover", "voice over",
                    "episode", "video", "youtube", "short",
                }
                _is_script_request = (
                    route == "director"
                    and getattr(self, "director", None) is not None
                    and self.writer is not None
                    and any(kw in user_text.lower() for kw in _script_keywords)
                )
                if _is_script_request:
                    try:
                        self.ui_q.put(("agent_phase", "director",
                                       "Director analysing style for script lock…"))
                        _style_brief_prompt = (
                            "DIRECTOR STYLE BRIEF — pre-deliberation pass.\n\n"
                            "The Writer is about to draft a script based on the user's request below.\n"
                            "Your job: produce a concise STYLE BRIEF (150–250 words) the Writer can\n"
                            "use as a hard constraint. Cover:\n"
                            "  • Sentence length and rhythm (short punchy? long flowing?)\n"
                            "  • Energy arc (how does this creator open / build / close?)\n"
                            "  • Verbal tics, phrases, or transitions they use\n"
                            "  • Humour style and register (dry, self-deprecating, enthusiastic?)\n"
                            "  • Anything the Writer must NOT do to stay in voice\n\n"
                            "Draw on your style memory. Be specific — cite patterns, not generalities.\n"
                            "End with: STYLE LOCKED.\n\n"
                            f"USER REQUEST:\n{user_text}"
                        )
                        _director_brief = self.director.respond(
                            _style_brief_prompt, max_tokens=350
                        )
                        if _director_brief.strip():
                            _brief_injection = (
                                "DIRECTOR STYLE BRIEF (apply as hard constraint):\n"
                                + _director_brief.strip()
                            )
                            _patch_model(self.writer, _brief_injection, "director_style_brief")
                            self.ui_q.put(("agent_phase", "director",
                                           "Style brief injected into Writer"))
                    except Exception as _dsb_e:
                        self.ui_q.put(("agent_phase", "director",
                                       f"Style brief skipped ({_dsb_e})"))

                events = orch.run(user_text, panel=panel, synth=synth_role,
                                  extra_ctx=_orch_extra)

                # Extract final and critique
                final_text = next((e.text for e in reversed(events) if e.who == "Writer" and e.kind == "final"), "")
                # Strip code blocks from conversational responses
                final_text = _filter_final(final_text, route, user_text)
                last_critique = next((e.text for e in reversed(events) if e.who == "Judge" and e.kind == "observation"), "")

                # IDE fill — prefer Coder agent code if available
                if route == "ide" and self.var_fill_ide.get():
                    # Check if Coder agent produced code directly
                    tp_code = None
                    for ev in reversed(events):
                        if ev.who == "Coder" and ev.kind == "final":
                            extracted = _extract_code_block(ev.text)
                            if extracted:
                                tp_code = extracted
                            break

                    if tp_code:
                        code = tp_code
                        base = _safe_script_basename(user_text.splitlines()[0][:60])
                    else:
                        json_prompt = (
                            "Return JSON ONLY — keys: filename, code.\n"
                            "filename: short descriptive snake_case name ending in .py\n"
                            "code: complete runnable python script\n\n"
                            f"USER REQUEST:\n{user_text}\n\n"
                            f"PROPOSAL:\n{final_text}\n"
                        )
                        json_resp = self.writer.respond(json_prompt)
                        fn, code = _parse_script_json(json_resp)
                        if not code:
                            code = _extract_code_block(final_text) or final_text
                        base = _safe_script_basename(fn) if fn else _safe_script_basename(user_text.splitlines()[0][:60])

                    self.ui_q.put(("set_script_name", base))
                    self.ui_q.put(("ide_fill", code))

                # Memory update -- always, not just on PASS.
                # Failed deliberations teach roles what to avoid.
                _deliberation_passed = "Verdict: PASS" in (last_critique or "")
                # Harvest low-confidence gaps from orchestrator shared context
                _low_conf_gaps = getattr(getattr(orch, "_last_ctx", None), "shared", {}).get("_low_conf_gaps", [])
                self.ui_q.put(("memory_update", user_text, final_text, last_critique,
                               _deliberation_passed, _low_conf_gaps))

                # T1-C: Confidence score — compute first so the persisted record has it
                _conf = 0
                for _ev in reversed(events):
                    if _ev.who == "Judge" and _ev.kind == "observation" and "confidence" in _ev.text:
                        try:
                            import json as _cj
                            _cobj = _cj.loads(_ev.text.split("Ranking:\n", 1)[-1].strip())
                            _conf = int(_cobj.get("confidence", 0))
                        except Exception:
                            pass
                        break

                # Persist verdict record for history dashboard (now with real confidence)
                self.ui_q.put(("verdict_record", {
                    "ts":         now_iso(),
                    "session_id": self.session_id,
                    "route":      route,
                    "query":      user_text[:120],
                    "confidence": _conf,
                    "passed":     _deliberation_passed,
                    "rounds":     sum(1 for e in events
                                     if e.who == "Judge" and e.kind == "observation"
                                     and "Verdict:" in e.text),
                }))

                # T1-B: Session naming
                try:
                    _sname = self.judge.name_session(user_text, last_critique or "")
                    if _sname:
                        self.ui_q.put(("session_rename", _sname))
                except Exception:
                    pass

                self.ui_q.put(("verdict_confidence", _conf))

                self.ui_q.put(("judge_final", last_critique))
                self.ui_q.put(("save_output_ready", final_text, user_text, route))
                self.ui_q.put(("done", None))

            except Exception as e:
                import traceback
                self.ui_q.put(("error", traceback.format_exc()))
            finally:
                # Always restore patched model.respond — even on exception.
                # _patched_models now stores (model, original_fn) directly, so
                # restoration is unconditional and doesn't rely on lookup maps.
                for _rname, _entry in locals().get("_patched_models", {}).items():
                    try:
                        _model, _orig_fn = _entry
                    except (TypeError, ValueError):
                        continue   # legacy single-fn entry — skip rather than crash
                    if _model is not None:
                        try: _model.respond = _orig_fn
                        except Exception: pass
                for _orole, _orig_key in locals().get("_override_restores", {}).items():
                    _om = locals().get("_override_map", {}).get(_orole)
                    if _om is not None:
                        try: _om.backend_key = _orig_key
                        except Exception: pass
                # Restore per-route temperature adjustments
                for _tr_role, _orig_temp in locals().get("_temp_restores", {}).items():
                    _tr_m = getattr(self, _tr_role, None)
                    if _tr_m is not None:
                        try: _tr_m.temperature = _orig_temp
                        except Exception: pass
                # Clear crash-recovery sentinel — deliberation reached its
                # finally block, so by definition the app didn't crash mid-flight.
                try:
                    self.convo_store.mark_session_done(self.session_id)
                except Exception:
                    pass
                # Clear per-query specialist state so it doesn't leak into
                # the next deliberation. Manual pin (_forced_specialist_id)
                # is intentionally NOT cleared — the user explicitly chose it.
                try:
                    self._active_specialists = []
                except Exception:
                    pass

        threading.Thread(target=worker, daemon=True).start()

    # ============================
    # Queue polling
    # ============================

    def _poll_ui_queue(self):
        try:
            while True:
                item = self.ui_q.get_nowait()
                kind = item[0]

                if kind == "live_event":
                    _, ev = item
                    if ev.kind == "phase":
                        self._append_transcript(ev.who, ev.text, "phase")
                        # Clear stream box so each agent shows cleanly
                        self._clear_stream_box()
                        # Update active-agent indicator in toolbar
                        if hasattr(self, "_agent_label"):
                            _skip = {"Orchestrator", "Librarian", ""}
                            _atext = ("▶ " + ev.who) if ev.who not in _skip else ""
                            self._agent_label.configure(text=_atext)
                    elif ev.kind == "final":
                        # Finalised — clear stream buffer for this speaker
                        self._stream_buffers.pop(ev.who, None)
                        self._append_transcript(ev.who, ev.text, "final")
                        # Provenance chips under the Writer's final answer —
                        # the source files that fed this turn (assembled in
                        # _send as self._last_turn_sources).
                        # Only do the source-path resolution when there ARE
                        # sources — skips the work entirely for non-data answers.
                        _srcs = getattr(self, "_last_turn_sources", None)
                        if ev.who == "Writer" and _srcs:
                            try:
                                self._render_source_chips(_srcs)
                            except Exception:
                                pass
                    elif ev.kind in ("observation", "action"):
                        self._append_transcript(ev.who, ev.text, ev.kind)
                    # "thought" and "token" events are lightweight; skip transcript

                elif kind == "stream_token":
                    _, who, token = item
                    self._append_stream_box(who, token)

                elif kind == "tps_update":
                    _, who, tps = item
                    self._update_tps(who, tps)

                elif kind == "judge_final":
                    _, txt = item
                    if txt:
                        self._set_judge(txt)
                        # T3-B: highlight required changes if present
                        import council_engine as _ce
                        _changes = _ce.JudgeModel.parse_required_changes(txt)
                        if _changes:
                            _chg_block = ("\n" + "─" * 32 + "\n"
                                          + "REQUIRED CHANGES:\n"
                                          + "\n".join("  • " + c for c in _changes)
                                          + "\n")
                            self._set_judge(txt + _chg_block)

                elif kind == "job_status":
                    # Agent-job lifecycle event from the background JobRunner.
                    _, job_id, status = item
                    self._aj_update_row(job_id)
                    self._aj_log_append(f"• {job_id}: {status}")

                elif kind == "job_step":
                    _, job_id, step = item
                    self._aj_update_row(job_id)
                    _obs = (step.get("observation") or "")
                    self._aj_log_append(
                        f"  {job_id} · step {step.get('index')} · "
                        f"{step.get('label')}"
                        + (f" — {_obs[:160]}" if _obs else ""))

                elif kind == "job_done":
                    _, job_id, status, summary = item
                    self._aj_update_row(job_id)
                    self._aj_log_append(
                        f"✓ {job_id}: {status}"
                        + (f" — {str(summary)[:200]}" if summary else ""))

                elif kind == "ide_fill":
                    # IDE tab is advanced-only; ignore the fill event in
                    # consumer builds. Code still appears in the transcript.
                    if hasattr(self, "ide_code"):
                        _, code = item
                        try:
                            self.ide_code.delete("1.0", "end")
                            self.ide_code.insert("1.0", code)
                        except tk.TclError:
                            pass

                elif kind == "set_script_name":
                    _, base = item
                    base = _safe_script_basename(base)
                    self.current_script_name = base
                    if hasattr(self, "script_name_var"):
                        try: self.script_name_var.set(base)
                        except tk.TclError: pass
                    self._append_transcript("Librarian", f"Script named: {base}.py", "final")

                elif kind == "memory_update":
                    _, user_text, final_text, critique, *_rest = item
                    _passed_flag  = _rest[0] if len(_rest) > 0 else True
                    _lc_gaps      = _rest[1] if len(_rest) > 1 else []
                    self._do_memory_update(user_text, final_text, critique,
                                           passed=_passed_flag, low_conf_gaps=_lc_gaps)

                elif kind == "agent_phase":
                    _, phase, msg = item
                    self._agent_log_append(phase, msg)
                    # Also echo notable events to transcript
                    if any(x in msg for x in ("✓", "✗", "⚠", "Researched", "RAG")):
                        self._append_transcript("Agent", msg, "observation")
                    if hasattr(self, "rag_count_label"):
                        self._update_rag_count_label()

                elif kind == "grapher_event":
                    _, phase, msg = item
                    if hasattr(self, "_grapher_stats"):
                        self._grapher_show_stats(f"[{phase}] {msg}")

                elif kind == "grapher_stats":
                    _, text = item
                    if hasattr(self, "_grapher_stats"):
                        self._grapher_show_stats(text)

                elif kind == "grapher_ai_result":
                    _, result = item
                    if not hasattr(self, "_grapher_stats"):
                        pass
                    elif result.parse_error:
                        self._grapher_show_stats(
                            f"✗ AI parse error: {result.parse_error}\n\n"
                            + getattr(result, "analysis", ""))
                    else:
                        if result.analysis:
                            self._grapher_show_stats(result.analysis)
                        if result.spec and self._grapher_dataset:
                            self._grapher_apply_spec_to_controls(result.spec)
                            self._grapher_render_plotly(result.spec, self._grapher_dataset)

                elif kind == "grapher_autosuggest":
                    # #3 auto-suggest after file load
                    _, hint, suggestion = item
                    if hasattr(self, "_grapher_stats") and hint:
                        self._grapher_show_stats(hint)
                    if (suggestion and getattr(suggestion, "spec", None)
                            and self._grapher_dataset):
                        self._grapher_apply_spec_to_controls(suggestion.spec)

                elif kind == "vault_mgr_log":
                    _, msg = item
                    if hasattr(self, "_vmgr_log"):
                        self._vmgr_append(msg)

                elif kind == "vault_mgr_refresh":
                    if hasattr(self, "_vmgr_tree"):
                        self._vmgr_refresh_tree()

                elif kind == "apoth_out":
                    _, text = item
                    self._append_transcript("Apothecary", text, "final")

                elif kind == "ide_stdout":
                    _, text = item
                    self._ide_print(text)

                elif kind == "ide_stderr":
                    _, text = item
                    self._ide_print(text, "stderr")

                elif kind == "ide_info":
                    _, text = item
                    self._ide_print(text, "info")

                elif kind == "nodes_result":
                    _, statuses = item
                    self._populate_nodes_tree(statuses)
                    if hasattr(self, "nodes_status_label"):
                        try:
                            self.nodes_status_label.configure(
                                text=f"Last updated: {now_iso()}")
                        except tk.TclError:
                            pass  # widget destroyed mid-callback
                    # Schedule next auto-refresh
                    if self._node_refresh_id:
                        self.after_cancel(self._node_refresh_id)
                    self._node_refresh_id = self.after(15_000, self._refresh_nodes_async)

                elif kind == "stt_out":
                    _, text = item
                    self.stt_out.delete("1.0", "end")
                    self.stt_out.insert("1.0", text)

                elif kind == "verdict_confidence":
                    _, conf = item
                    self._last_confidence = conf
                    _bar = "█" * conf + "░" * (10 - conf)
                    _lbl = "HIGH" if conf >= 7 else "MED" if conf >= 4 else "LOW"
                    self._set_judge("Confidence: " + str(conf) + "/10  [" + _bar + "]  " + _lbl + "\n")

                elif kind == "verdict_record":
                    _, record = item
                    self._save_verdict_record(record)

                elif kind == "session_rename":
                    _, new_name = item
                    self._rename_session(new_name)

                elif kind == "trend_result":
                    _, trend_text = item
                    # Show trends in a popup
                    _tw = tk.Toplevel(self)
                    _tw.title("Cross-Session Trends")
                    _tw.configure(bg="#1a1414")
                    _tw.geometry("700x400")
                    _trend_box = self._make_text(_tw, wrap="word")
                    _trend_box.pack(fill="both", expand=True, padx=10, pady=10)
                    _trend_box.insert("1.0", trend_text)
                    _trend_box.configure(state="disabled")
                    self._append_transcript("Librarian",
                        "Cross-session trends generated → saved to vault/trends.md", "final")

                elif kind == "lens_result":
                    _, role_name, response = item
                    if hasattr(self, "_lens_output"):
                        self._lens_output.configure(state="normal")
                        sep = "─" * 40
                        self._lens_output.insert("end",
                            f"\n{sep}\n🔍 {role_name.upper()}\n{sep}\n{response}\n")
                        self._lens_output.configure(state="disabled")
                        self._lens_output.see("end")

                elif kind == "lens_done":
                    _, count = item
                    if hasattr(self, "_lens_status"):
                        self._lens_status.configure(text=f"Done — {count} roles responded.")

                elif kind == "clarification_needed":
                    _, who, question = item
                    self._pause_event.clear()  # pause the worker
                    self._show_clarification(who, question)

                elif kind == "save_output_ready":
                    _, _final, _query, _route = item
                    self._last_final_text  = _final
                    self._last_query_text  = _query
                    self._last_route       = _route
                    # Scan this turn's final answer for vault filename
                    # mentions and pin them for the next turn's
                    # injection. The next user message can then say
                    # "tell me more about orders.csv" and the injector
                    # will zoom the file into a full [VAULT MATCH]
                    # block even when search-headers mode is active.
                    try:
                        self._update_pins_from_response(_final)
                    except Exception as _pin_err:
                        print(f"[pin] update failed: {_pin_err!r}")
                    # Auto-save if LaTeX was requested
                    if _detect_latex_request(_query):
                        self._auto_save_latex(_final, _query)
                    # Show save buttons in transcript
                    self._show_save_buttons()

                elif kind == "done":
                    self._set_status("● idle")
                    if hasattr(self, "_agent_label"):
                        self._agent_label.configure(text="")
                    # Show verdict feedback bar
                    self._vfb_show()
                    # Auto-speak if enabled (#10)
                    _final_for_tts = getattr(self, "_last_final_text", "")
                    if _final_for_tts:
                        self._tts_auto_speak(_final_for_tts)

                elif kind == "error":
                    _, msg = item
                    _coach = _coach_for_error(msg)
                    if _coach:
                        # Friendly guidance + a one-click fix, with only the
                        # LAST traceback line for the curious (full trace is in
                        # conversation_logs). Non-technical users shouldn't face
                        # a raw Python traceback.
                        _last = (str(msg).strip().splitlines() or ["error"])[-1]
                        self._append_transcript(
                            "Council",
                            "⚠ That didn't work. " + _coach["plain"]
                            + f"\n(Technical detail: {_last[:200]})",
                            "observation")
                        self._render_error_coach_button(_coach)
                        self._set_status("● needs attention", "#f9e2af")
                    else:
                        self._append_transcript("ERROR", msg, "final")
                        self.transcript.tag_add("error", "end-2l", "end")
                        self._set_status("● error", "#f38ba8")

        except queue.Empty:
            pass
        # Coalesced scroll: if any stream tokens were appended during
        # this drain, do the single deferred see("end") + re-lock now
        # instead of once per token inside the loop above.
        self._flush_stream_box()
        self.after(50, self._poll_ui_queue)

    # ============================
    # Actions
    # ============================

    def _apply_script_name(self):
        nm = _safe_script_basename(self.script_name_var.get())
        self.current_script_name = nm
        self.script_name_var.set(nm)
        self._append_transcript("Librarian", f"Script name: {nm}.py", "final")

    def _new_session(self):
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._append_transcript("Librarian", f"New session started: {self.session_id}", "final")
        self._sessions_refresh()

    def _save_verdict_record(self, record: dict):
        """Append one verdict record to the persistent history file."""
        import json as _j
        try:
            VERDICT_HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(VERDICT_HISTORY_PATH, "a", encoding="utf-8") as _f:
                _f.write(_j.dumps(record, ensure_ascii=False) + "\n")
        except Exception as e:
            self._append_transcript("Librarian", "Verdict history write failed: " + str(e), "final")

    def _load_verdict_history(self, last_n: int = 200) -> list:
        """Load the last N verdict records from the history file."""
        import json as _j
        if not VERDICT_HISTORY_PATH.exists():
            return []
        records = []
        try:
            lines = VERDICT_HISTORY_PATH.read_text(encoding="utf-8").splitlines()
            for ln in lines[-last_n:]:
                try:
                    records.append(_j.loads(ln))
                except Exception:
                    pass
        except Exception:
            pass
        return records

    # ── Output save methods ──────────────────────────────────────────────────

    def _show_save_buttons(self):
        """Show the save-output panel below the input area."""
        if hasattr(self, "_save_frame"):
            self._save_frame.pack(fill="x", pady=(4, 0))

    def _hide_save_buttons(self):
        if hasattr(self, "_save_frame"):
            self._save_frame.pack_forget()

    def _get_save_filename(self, ext: str) -> Optional[str]:
        """Prompt user for a filename stem, return full name with ext."""
        from tkinter import simpledialog as _sd
        # Suggest a name from the query
        _raw = self._last_query_text[:40].strip()
        _raw = __import__("re").sub(r"[^\w\s-]", "", _raw).strip()
        _suggested = __import__("re").sub(r"\s+", "_", _raw).lower() or "council_output"
        name = _sd.askstring(
            "Save output",
            f"Filename (without extension):",
            initialvalue=_suggested,
            parent=self,
        )
        if not name:
            return None
        name = name.strip().rstrip(".")
        if not name:
            return None
        return name + ext

    def _save_output_txt(self):
        """Save the last Writer output as a plain .txt file to the vault."""
        if not self._last_final_text:
            self._append_transcript("Librarian", "No output to save yet.", "observation")
            return
        fname = self._get_save_filename(".txt")
        if not fname:
            return
        # Strip any markdown formatting for plain text
        import re as _re
        plain = _re.sub(r"\*\*(.+?)\*\*", r"\1", self._last_final_text)
        plain = _re.sub(r"\*(.+?)\*",   r"\1", plain)
        plain = _re.sub(r"^#{1,6}\s+",   "",     plain, flags=_re.MULTILINE)
        plain = _re.sub(r"`{1,3}",        "",     plain)
        saved = self.librarian.save_text(fname, plain)
        self._append_transcript(
            "Librarian",
            f"✓ Saved as {fname} ({len(plain)} chars) → vault",
            "final"
        )
        self._offer_disk_save(fname, plain)

    def _save_output_md(self):
        """Save the last Writer output as a .md file to the vault."""
        if not self._last_final_text:
            self._append_transcript("Librarian", "No output to save yet.", "observation")
            return
        fname = self._get_save_filename(".md")
        if not fname:
            return
        md = f"# {self._last_query_text[:80]}\n\n{self._last_final_text}"
        self.librarian.save_text(fname, md)
        self._append_transcript("Librarian", f"✓ Saved as {fname} → vault", "final")
        self._offer_disk_save(fname, md)

    def _save_output_latex(self):
        """Wrap the last output in a LaTeX document and save as .tex."""
        if not self._last_final_text:
            self._append_transcript("Librarian", "No output to save yet.", "observation")
            return
        fname = self._get_save_filename(".tex")
        if not fname:
            return
        title = self._last_query_text[:60].strip() or "Council Output"
        tex = _wrap_latex(title, self._last_final_text)
        self.librarian.save_text(fname, tex)
        self._append_transcript(
            "Librarian",
            f"✓ Saved as {fname} (LaTeX document) → vault",
            "final"
        )
        self._offer_disk_save(fname, tex)

    def _auto_save_latex(self, text: str, query: str):
        """Automatically save as LaTeX when user explicitly requested it."""
        import re as _re
        _raw = _re.sub(r"[^\w\s-]", "", query[:40]).strip()
        fname = _re.sub(r"\s+", "_", _raw).lower() or "council_latex"
        fname = fname + ".tex"
        title = query[:60].strip()
        tex = _wrap_latex(title, text)
        self.librarian.save_text(fname, tex)
        self._append_transcript(
            "Librarian",
            f"✓ Auto-saved LaTeX: {fname} → vault  (click 📐 to save to disk)",
            "final"
        )

    def _offer_disk_save(self, vault_name: str, content: str):
        """Ask user if they want to also save to a location on disk."""
        from tkinter import filedialog as _fd
        if not messagebox.askyesno(
            "Also save to disk?",
            f"Saved to vault as {vault_name}.\n\nAlso save a copy somewhere on disk?",
            parent=self,
        ):
            return
        ext = "." + vault_name.rsplit(".", 1)[-1] if "." in vault_name else ".txt"
        _filetypes = {
            ".txt":  [("Text files", "*.txt"), ("All files", "*.*")],
            ".md":   [("Markdown files", "*.md"), ("All files", "*.*")],
            ".tex":  [("LaTeX files", "*.tex"), ("All files", "*.*")],
        }
        path = _fd.asksaveasfilename(
            parent=self,
            title="Save output to disk",
            defaultextension=ext,
            initialfile=vault_name,
            filetypes=_filetypes.get(ext, [("All files", "*.*")]),
        )
        if path:
            try:
                with open(path, "w", encoding="utf-8") as _f:
                    _f.write(content)
                self._append_transcript(
                    "Librarian", f"✓ Also saved to: {path}", "observation"
                )
            except Exception as _e:
                self._append_transcript(
                    "Librarian", f"✗ Disk save failed: {_e}", "observation"
                )

    def _export_session_md(self):
        """
        Render the current session from ConversationStore into a Markdown file
        and save it to the vault. Covers all stored turns (up to 200).
        """
        turns = self.convo_store.load_last(self.session_id, n=200)
        if not turns:
            self._append_transcript("Librarian", "Nothing to export — session is empty.", "final")
            return

        lines = [
            "# Council Deliberation Export",
            "",
            "**Session:** " + self.session_id,
            "**Exported:** " + __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "",
            "---",
            "",
        ]

        _section_roles = {"Judge", "Orchestrator"}
        _peasant_roles = {"Peasant"}  # Q&A block formatting
        _phase_marker = "\u25b6"  # the triangle used by _phase()

        for turn in turns:
            who  = turn.get("who", "?")
            text = turn.get("text", "").strip()
            ts   = turn.get("ts", "")
            if not text:
                continue
            # Phase lines get lighter formatting
            if text.startswith(_phase_marker):
                lines.append("> " + text)
                lines.append("")
            elif who in _section_roles:
                lines.append("## " + who)
                if ts:
                    lines.append("*" + ts + "*")
                lines.append("")
                lines.append(text)
                lines.append("")
                lines.append("---")
                lines.append("")
            elif who in _peasant_roles:
                lines.append("### Peasant Q&A")
                if ts:
                    lines.append("*" + ts + "*")
                lines.append("")
                for ln in text.splitlines():
                    stripped = ln.strip()
                    if stripped.startswith("[q:") or stripped.startswith("⚠"):
                        lines.append("> " + stripped)
                    elif "?" in stripped and stripped:
                        lines.append("**" + stripped + "**")
                    elif stripped:
                        lines.append(stripped)
                lines.append("")
            else:
                lines.append("### " + who)
                if ts:
                    lines.append("*" + ts + "*")
                lines.append("")
                lines.append(text)
                lines.append("")

        md_text = "\n".join(lines)
        filename = "export_" + self.session_id + ".md"
        saved_path = self.librarian.save_text(filename, md_text)
        self._append_transcript(
            "Librarian",
            "Session exported: " + filename + " (" + str(len(turns)) + " turns, "
            + str(len(md_text)) + " chars)",
            "final"
        )

    def _rename_session(self, new_name: str):
        """Rename the current session if it still has a raw timestamp ID."""
        import re as _re
        if not _re.match(r"^\d{8}_\d{6}$", self.session_id):
            return
        ok = self.convo_store.rename_session(self.session_id, new_name)
        if ok:
            old = self.session_id
            self.session_id = new_name
            for model in self.personalities.values():
                if hasattr(model, "session_id") and model.session_id == old:
                    model.session_id = new_name
            self._append_transcript("Librarian", "Session named: " + new_name, "final")
            self._sessions_refresh()

    # ── Background deliberation queue (T2-A) ─────────────────────────────

    def _show_bg_queue(self):
        """Open/raise the background deliberation queue window."""
        if hasattr(self, "_bg_win") and self._bg_win.winfo_exists():
            self._bg_win.lift()
            return
        win = tk.Toplevel(self)
        win.title("Background Deliberation Queue")
        win.configure(bg="#1a1414")
        win.geometry("620x440")
        self._bg_win = win
        if not hasattr(self, "_bg_queue"):
            self._bg_queue = []
        ttk.Label(win, text="Queue items (one prompt per line or load from vault):").pack(anchor="w", padx=8, pady=(8,2))
        self._bg_input = self._make_text(win, height=6, wrap="word")
        self._bg_input.pack(fill="x", padx=8)
        btn_row = ttk.Frame(win)
        btn_row.pack(fill="x", padx=8, pady=4)
        ttk.Button(btn_row, text="Add to Queue",     command=self._bg_queue_add).pack(side="left")
        ttk.Button(btn_row, text="Load Vault Items", command=self._bg_queue_from_vault).pack(side="left", padx=6)
        ttk.Button(btn_row, text="Run Queue Now",    command=self._bg_queue_run).pack(side="left", padx=6)
        ttk.Button(btn_row, text="Clear",            command=self._bg_queue_clear).pack(side="left")
        # Options row: full council toggle + repeat interval
        opt_row = ttk.Frame(win)
        opt_row.pack(fill="x", padx=8, pady=(0,2))
        if not hasattr(self, "_bg_full_council"):
            self._bg_full_council = tk.BooleanVar(value=True)
        if not hasattr(self, "_bg_repeat_mins"):
            self._bg_repeat_mins = tk.StringVar(value="0")
        ttk.Checkbutton(opt_row, text="Full council deliberation",
                         variable=self._bg_full_council).pack(side="left")
        ttk.Label(opt_row, text="  Repeat every (mins, 0=off):", foreground="#6c7086").pack(side="left")
        ttk.Entry(opt_row, textvariable=self._bg_repeat_mins, width=5).pack(side="left", padx=(2,0))
        ttk.Label(win, text="Queue & Results:").pack(anchor="w", padx=8)
        self._bg_log = self._make_text(win, height=12, wrap="word", state="disabled")
        self._bg_log.pack(fill="both", expand=True, padx=8, pady=(0,8))
        self._bg_log_append("Queue ready. Items: " + str(len(self._bg_queue)))

    def _bg_log_append(self, msg: str):
        if not hasattr(self, "_bg_log") or not self._bg_log.winfo_exists():
            return
        self._bg_log.configure(state="normal")
        self._bg_log.insert("end", msg + "\n")
        self._bg_log.see("end")
        self._bg_log.configure(state="disabled")

    def _bg_queue_add(self):
        if not hasattr(self, "_bg_queue"):
            self._bg_queue = []
        text = self._bg_input.get("1.0", "end").strip()
        if not text:
            return
        items = [l.strip() for l in text.splitlines() if l.strip()]
        self._bg_queue.extend(items)
        self._bg_log_append("Added " + str(len(items)) + " item(s). Total: " + str(len(self._bg_queue)))
        self._bg_input.delete("1.0", "end")

    def _bg_queue_from_vault(self):
        """Load vault .md/.txt items as queue prompts."""
        if not hasattr(self, "_bg_queue"):
            self._bg_queue = []
        added = 0
        for name in self.librarian.list_items():
            if name.endswith((".md", ".txt")):
                try:
                    text = self.librarian.read_text(name).strip()
                    if text:
                        self._bg_queue.append("Review vault item: " + name + "\n\n" + text[:800])
                        added += 1
                except Exception:
                    pass
        self._bg_log_append("Loaded " + str(added) + " vault item(s) into queue.")

    def _bg_queue_clear(self):
        if hasattr(self, "_bg_queue"):
            self._bg_queue.clear()
        self._bg_log_append("Queue cleared.")

    def _bg_queue_run(self):
        """Run all queued items sequentially in a daemon thread.
        When full council mode is on, runs the real DeliberationOrchestrator
        (no UI streaming -- results saved directly to vault).
        Falls back to writer.respond() if full council is off.
        Results are saved to vault and surfaced in the BG log.
        Repeat-every-N-minutes: reloads vault items and re-runs on a timer.
        """
        if not hasattr(self, "_bg_queue") or not self._bg_queue:
            self._bg_log_append("Queue is empty.")
            return
        import threading as _th, time as _t, re as _re
        import json as _bj
        full_council = getattr(self, "_bg_full_council", None)
        use_full = bool(full_council.get()) if full_council is not None else True
        repeat_var = getattr(self, "_bg_repeat_mins", None)
        try:
            repeat_mins = float(repeat_var.get()) if repeat_var else 0
        except Exception:
            repeat_mins = 0
        items = list(self._bg_queue)
        self._bg_queue.clear()
        mode_label = "full council" if use_full else "writer-only"
        self._bg_log_append("Starting BG run (" + mode_label + "): " + str(len(items)) + " item(s)...")

        def _run_one(prompt: str, idx: int, total: int, run_id: str) -> str:
            """Run a single prompt; returns the result text."""
            if not use_full:
                return self.writer.respond(prompt)
            # Full deliberation -- no token streaming, no UI updates mid-flight
            route = self.judge.route(prompt)
            panel, synth_role = _panel_for_route(route)
            agents = {
                "writer":     ModelAgent("Writer",     self.writer,     enable_tools=False),
                "peasant":    ModelAgent("Peasant",    self.peasant,    enable_tools=False),
                "intern":     ModelAgent("Intern",     self.intern,     enable_tools=False),
                "coder": ModelAgent("Coder", self.coder, enable_tools=False),
                "artist":     ModelAgent("Artist",     self.artist,     enable_tools=False),
            }
            if self.skeptic is not None:
                agents["skeptic"] = ModelAgent("Skeptic", self.skeptic, enable_tools=False)
            panel = [p for p in panel if p in agents] or ["intern", "coder", "artist"]
            orch = DeliberationOrchestrator(
                judge_model=self.judge, agents=agents,
                max_rounds=2, debate_turns=1,  # leaner for background -- 1 cross-fire turn
            )
            events = orch.run(prompt, panel=panel, synth=synth_role)
            result = next((e.text for e in reversed(events)
                           if e.who == "Writer" and e.kind == "final"), "")
            critique = next((e.text for e in reversed(events)
                             if e.who == "Judge" and e.kind == "observation"), "")
            verdict = "PASS" if "Verdict: PASS" in critique else "NEEDS_WORK"
            conf = 0
            try:
                rank_raw = next((e.text.split("Ranking:\n", 1)[-1].strip()
                    for e in reversed(events)
                    if e.who == "Judge" and "confidence" in e.text), "")
                if rank_raw:
                    conf = int(_bj.loads(rank_raw).get("confidence", 0))
            except Exception:
                pass
            return result + "\n\n---\nVerdict: " + verdict + " | Confidence: " + str(conf) + "/10"

        def _run_bg(items_to_run):
            total = len(items_to_run)
            for i, prompt in enumerate(items_to_run, 1):
                run_id = "bg_" + _re.sub(r"[^\w_]", "", prompt[:30].replace(" ", "_"))
                self.after(0, lambda i=i, t=total, p=prompt: self._bg_log_append(
                    "[" + str(i) + "/" + str(t) + "] " + p[:60] + "..."
                ))
                try:
                    result = _run_one(prompt, i, total, run_id)
                    import time as _ti
                    stamp = str(int(_ti.time()))
                    fname = run_id + "_" + stamp + ".md"
                    self.librarian.save_text(fname,
                        "# BG Deliberation\n"
                        "## Prompt\n" + prompt + "\n\n"
                        "## Result\n" + result
                    )
                    self.after(0, lambda fn=fname: self._bg_log_append("  Saved: " + fn))
                except Exception as e:
                    self.after(0, lambda i=i, e=e: self._bg_log_append(
                        "  ERROR item " + str(i) + ": " + str(e)))
            self.after(0, lambda: self._bg_log_append("Run complete (" + mode_label + ")."))

        def _scheduler_loop():
            """Run once immediately, then repeat every repeat_mins if > 0."""
            _run_bg(items)
            if repeat_mins > 0:
                import time as _t2
                self.after(0, lambda: self._bg_log_append(
                    "Repeating in " + str(repeat_mins) + " min(s). Close the BG window to cancel."))
                _t2.sleep(repeat_mins * 60)
                # On repeat: reload vault items fresh
                repeat_items = []
                for name in self.librarian.list_items():
                    if name.endswith((".md", ".txt")) and not name.startswith("bg_"):
                        try:
                            text = self.librarian.read_text(name).strip()
                            if text:
                                repeat_items.append("Review and update vault item: " + name + "\n\n" + text[:600])
                        except Exception:
                            pass
                if repeat_items:
                    _run_bg(repeat_items)

        _th.Thread(target=_scheduler_loop, daemon=True, name="bg-deliberation").start()

    def _start_config_watcher(self):
        """Daemon thread: watches pins.json and hot-reloads on change."""
        import threading as _th, time as _t
        _last_mtime = [0.0]
        def _watch():
            while True:
                try:
                    if PINS_PATH.exists():
                        mtime = PINS_PATH.stat().st_mtime
                        if mtime != _last_mtime[0]:
                            if _last_mtime[0] != 0.0:
                                self.after(0, self._hot_reload_pins)
                            _last_mtime[0] = mtime
                except Exception:
                    pass
                _t.sleep(2.0)
        _th.Thread(target=_watch, daemon=True, name="config-watcher").start()

    def _hot_reload_pins(self):
        """Called on UI thread when pins.json changes -- rebuilds personalities.
        _reloading_pins lock prevents concurrent rebuilds on rapid saves.
        """
        if getattr(self, "_reloading_pins", False):
            return
        self._reloading_pins = True
        try:
            pins = ce.load_personality_pins(PINS_PATH)
            self.personalities = ce.build_personalities(
                pins=pins, vault_dir=VAULT_DIR, session_id=self.session_id,
                trace=True, dispatcher=self.dispatcher,
                prior_session_id=self.prior_session_id,
            )
            self._unpack_personalities()
            self._append_transcript("Librarian", "pins.json changed -- personalities hot-reloaded.", "final")
        except Exception as e:
            self._append_transcript("Librarian", "Hot-reload failed: " + str(e), "final")
        finally:
            self._reloading_pins = False

    def _edit_pins(self):
        current = "{}"
        try:
            if PINS_PATH.exists():
                current = PINS_PATH.read_text(encoding="utf-8")
        except Exception:
            pass
        new = simpledialog.askstring(
            "Personality Pins JSON",
            "Pin personalities to backend keys.\n\n"
            "Valid keys: local_general_primary, local_general_alt, local_coder_primary,\n"
            "local_coder_fast, local_judge_fast, local_peasant_fast\n\n"
            "Example:\n{\"coder\":\"local_coder_primary\",\"writer\":\"local_general_primary\"}",
            initialvalue=current, parent=self,
        )
        if new is None:
            return
        try:
            obj = _json.loads(new)
            if not isinstance(obj, dict):
                raise ValueError("Pins must be a JSON object.")
            PINS_PATH.write_text(_json.dumps(obj, indent=2), encoding="utf-8")
            pins = ce.load_personality_pins(PINS_PATH)
            self.personalities = ce.build_personalities(
                pins=pins, vault_dir=VAULT_DIR, session_id=self.session_id,
                trace=True, dispatcher=self.dispatcher, prior_session_id=self.prior_session_id,
            )
            self._unpack_personalities()
            self._append_transcript("Librarian", f"Pins updated: {PINS_PATH}", "final")
        except Exception as e:
            messagebox.showerror("Invalid JSON", str(e))

    def _do_memory_update(self, user_text: str, final_text: str, critique: str,
                          passed: bool = True, low_conf_gaps: list = None):
        try:
            memmgr = self.writer.memory_manager
            if memmgr is None:
                return
            outcome = "PASS" if passed else "NEEDS_WORK/max-rounds"
            roles = [
                ("intern",      self.intern),
                ("coder",  self.coder),
                ("peasant",     self.peasant),
                ("artist",      self.artist),
                ("writer",      self.writer),
                ("judge",       self.judge),
            ]
            # Add optional personalities if they exist
            for _opt_role in ("skeptic", "sage", "strategist", "musician", "content",
                              "director", "eye", "cutter", "algorithm"):
                _opt_model = getattr(self, _opt_role, None)
                if _opt_model is not None:
                    roles.append((_opt_role, _opt_model))
            for role_name, role_model in roles:
                ce.update_role_memory_after_pass(
                    role_name=role_name, role_model=role_model,
                    memory_manager=memmgr, user_text=user_text,
                    final_answer=final_text, judge_critique=critique,
                    passed=passed,
                )

            # ── #1 Shared project memory: observer roles write cross-session facts ──
            # Runs after role memory so observers can reference their fresh role memory.
            # Uses only one observer role per deliberation (first available in priority order).
            _observer_priority = ["coder", "sage", "strategist", "director"]
            for _obs_name in _observer_priority:
                _obs_model = getattr(self, _obs_name, None)
                if _obs_model is not None:
                    try:
                        ce.update_project_memory_after_pass(
                            role_name=_obs_name, role_model=_obs_model,
                            memory_manager=memmgr, user_text=user_text,
                            final_answer=final_text, judge_critique=critique,
                            passed=passed,
                        )
                    except Exception:
                        pass
                    break  # one observer per deliberation is enough

            # ── User-quirks layer: observe always, influence only when
            # mature. One small extraction call per deliberation; the
            # compiled USER PROFILE block stays empty until the maturity
            # gates pass (≥N distinct sessions, each quirk corroborated
            # in ≥K sessions) so a young profile can't poison answers.
            try:
                import user_quirks as _uq

                def _quirk_call(prompt: str) -> str:
                    return ce.local_chat(
                        messages=[{"role": "user", "content": prompt}],
                        temperature=0.0, num_predict=200, timeout=45,
                    )

                _was_active = bool(
                    (memmgr.read(ce._USER_PROFILE_KEY) or "").strip())
                _qstatus = _uq.update_after_deliberation(
                    user_text, getattr(self, "session_id", "unknown"),
                    _quirk_call, memory_manager=memmgr,
                )
                _now_active = bool(
                    (memmgr.read(ce._USER_PROFILE_KEY) or "").strip())
                if _now_active and not _was_active:
                    self._append_transcript(
                        "Librarian",
                        "👤 User profile is now ACTIVE — "
                        f"{_qstatus['quirks_confirmed']} preference(s) "
                        f"confirmed across {_qstatus['sessions_observed']} "
                        "sessions now inform every personality. "
                        "(COUNCIL_QUIRKS_ENABLE=0 disables.)",
                        "final",
                    )
                elif (_qstatus.get("enabled")
                        and not _qstatus.get("active")
                        and _qstatus.get("observed_now")):
                    self._append_transcript(
                        "Librarian",
                        "👤 User-profile learning: observing only — "
                        f"{_qstatus['sessions_observed']}/"
                        f"{_qstatus['sessions_required']} sessions before "
                        "anything is applied.",
                        "observation",
                    )
            except Exception as _uq_exc:
                print(f"[quirks] update skipped: {_uq_exc!r}")

            self._append_transcript("Librarian", f"Role memories updated ({outcome}).", "final")

            # ── #3 Proactive wishlist surfacing ───────────────────────────────────
            # After every memory update, check for open wishlist items and surface
            # the top items so the user knows what data to hunt down.
            try:
                _wishlist_raw = self.librarian.get_wishlist()
                _pending = [
                    ln.strip() for ln in _wishlist_raw.splitlines()
                    if ln.strip().startswith("- [ ]")
                ]
                if _pending:
                    _top = _pending[:5]
                    _wl_msg = (
                        "📋 **Vault wishlist — top items still missing:**\n"
                        + "\n".join(_top)
                        + (f"\n…and {len(_pending) - len(_top)} more." if len(_pending) > 5 else "")
                    )
                    self._append_transcript("Librarian", _wl_msg, "final")
            except Exception:
                pass

            # ── #8 Low-confidence gap logging ─────────────────────────────────
            # Roles that self-reported low confidence get their gaps persisted
            # to the wishlist so the user knows what vault data would help them.
            if low_conf_gaps:
                for _gap in low_conf_gaps:
                    try:
                        self.librarian.log_gap(
                            who=_gap.get("who", "?"),
                            topic=_gap.get("topic", "unknown"),
                            reason=_gap.get("reason", "low self-reported confidence"),
                        )
                    except Exception:
                        pass

        except Exception as e:
            self._append_transcript("Librarian", f"Memory update failed: {e}", "final")

    # ---- IDE actions ----

    # Patterns flagged before code execution. If any match, the user must
    # explicitly approve via the trust gate. Generated by an LLM, executed
    # blindly is the #1 risk in agentic dev tools.
    _IDE_RISKY_PATTERNS = [
        (r'\bos\.system\b',                "shell command execution"),
        (r'\bsubprocess\.\w+',             "subprocess calls"),
        (r'\bshutil\.(rmtree|move)\b',     "directory delete/move"),
        (r'\bos\.(remove|unlink|rmdir)\b', "file/directory deletion"),
        (r'\.unlink\(\)',                  "Path.unlink (file deletion)"),
        (r'\beval\s*\(',                   "dynamic code evaluation"),
        (r'\bexec\s*\(',                   "dynamic code execution"),
        (r'\b__import__\s*\(',             "dynamic imports"),
        (r'\brequests\.(get|post|put|delete|patch)\b', "outbound HTTP"),
        (r'\burllib\.(request|urlopen)',   "outbound HTTP"),
        (r'\bsocket\.(socket|connect)',    "raw socket access"),
        (r'\bos\.environ\[',               "environment variable access"),
        (r'\bpickle\.(load|loads)\b',      "pickle deserialisation (RCE risk)"),
    ]

    def _ide_scan_risky(self, code: str):
        """Return list of (pattern_label, line_number) for any risky pattern hits."""
        import re as _re
        hits = []
        for line_no, line in enumerate(code.splitlines(), start=1):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            for pat, label in self._IDE_RISKY_PATTERNS:
                if _re.search(pat, line):
                    hits.append((label, line_no, stripped[:120]))
                    break  # one flag per line is enough
        return hits

    def _ide_trust_gate(self, code: str) -> bool:
        """
        Show a confirmation dialog if code contains risky patterns. Trust
        decisions are session-scoped (hashed code is remembered until restart).
        Returns True if execution should proceed.
        """
        import hashlib
        # Initialise per-session trust set
        if not hasattr(self, "_ide_trusted_hashes"):
            self._ide_trusted_hashes: set = set()
        h = hashlib.sha256(code.encode("utf-8", errors="replace")).hexdigest()
        if h in self._ide_trusted_hashes:
            return True

        hits = self._ide_scan_risky(code)
        if not hits:
            return True   # no risky patterns — run freely

        # Build a clear, scannable message
        lines = ["This script contains operations that can affect your system:\n"]
        for label, ln, snippet in hits[:10]:
            lines.append(f"  • Line {ln}: {label}")
            lines.append(f"      {snippet}")
        if len(hits) > 10:
            lines.append(f"  …and {len(hits) - 10} more")
        lines.append("\nReview the script carefully before running.")
        lines.append("\nRun this script anyway?")

        from tkinter import messagebox
        choice = messagebox.askyesno(
            "Confirm code execution",
            "\n".join(lines),
            parent=self,
            icon="warning",
            default="no",
        )
        if choice:
            self._ide_trusted_hashes.add(h)
            return True
        return False

    def _ide_run(self):
        code = self.ide_code.get("1.0", "end")
        if not code.strip():
            return
        if not self._ide_trust_gate(code):
            self.ui_q.put(("ide_info", "[Cancelled by user]\n"))
            return

        def worker():
            fname = f"{_safe_script_basename(self.current_script_name)}.py"
            self.ui_q.put(("ide_info", f"[{fname}] Running…\n"))
            try:
                rc, out, err, path = self.runner.run_code(code, filename_hint=fname, timeout_s=120)
                self.ui_q.put(("ide_info", f"[rc={rc}]\n"))
                if out:
                    self.ui_q.put(("ide_stdout", out))
                if err:
                    self.ui_q.put(("ide_stderr", err))
            except Exception as e:
                self.ui_q.put(("ide_stderr", f"Runner error: {e}\n"))

        threading.Thread(target=worker, daemon=True).start()

    def _ide_run_stream(self):
        code = self.ide_code.get("1.0", "end")
        if not code.strip():
            return
        if not self._ide_trust_gate(code):
            self.ui_q.put(("ide_info", "[Cancelled by user]\n"))
            return

        def worker():
            fname = f"{_safe_script_basename(self.current_script_name)}.py"
            self.ui_q.put(("ide_info", f"[{fname}] Running (streaming)…\n"))
            try:
                rc, path = self.runner.run_code_streaming(
                    code, filename_hint=fname, timeout_s=120,
                    stdout_callback=lambda l: self.ui_q.put(("ide_stdout", l)),
                    stderr_callback=lambda l: self.ui_q.put(("ide_stderr", l)),
                )
                self.ui_q.put(("ide_info", f"\n[{path.name}] Exited rc={rc}\n"))
            except Exception as e:
                self.ui_q.put(("ide_stderr", f"Runner error: {e}\n"))

        threading.Thread(target=worker, daemon=True).start()

    def _ide_snapshot(self):
        code = self.ide_code.get("1.0", "end")
        if not code.strip():
            return
        label = _safe_script_basename(self.current_script_name)
        path = self.librarian.snapshot_code(code, label=label)
        self._append_transcript("Librarian", f"Snapshot saved: {path}", "final")
        self._lib_refresh()

    # ---- Librarian actions ----

    def _lib_refresh(self):
        self.vault_lb.delete(0, "end")
        for name in self.librarian.list_items():
            self.vault_lb.insert("end", name)

    def _lib_preview(self):
        sel = self.vault_lb.curselection()
        if not sel:
            return
        name = self.vault_lb.get(sel[0])
        try:
            txt = self.librarian.read_text(name)
            win = tk.Toplevel(self)
            win.title(f"Vault: {name}")
            win.geometry("800x600")
            t = self._make_text(win, wrap="word")
            t.insert("1.0", txt)
            t.configure(state="disabled")
            t.pack(fill="both", expand=True, padx=6, pady=6)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _lib_commit(self):
        msg = simpledialog.askstring("Commit Message", "Message:", parent=self)
        if not msg:
            return
        ok, out = self.librarian.git_commit_all(msg)
        self._append_transcript("Librarian", f"{'OK' if ok else 'FAIL'}: {out}", "final")

    def _lib_open_vault(self):
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(VAULT_DIR))   # type: ignore
            elif sys.platform == "darwin":
                subprocess.run(["open", str(VAULT_DIR)], check=False)
            else:
                subprocess.run(["xdg-open", str(VAULT_DIR)], check=False)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ---- Sessions actions ----

    def _sessions_refresh(self):
        self.session_lb.delete(0, "end")
        _vh_lookup: dict = {}
        for r in self._load_verdict_history(last_n=500):
            sid = r.get("session_id", "")
            if sid:
                existing = _vh_lookup.get(sid)
                if existing is None or r.get("confidence", 0) > existing.get("confidence", 0):
                    _vh_lookup[sid] = r
        _filter = getattr(self, "_session_filter_var", None)
        _ftext  = _filter.get().strip().lower() if _filter else ""
        for sid in self.convo_store.list_sessions():
            if _ftext and _ftext not in sid.lower():
                continue
            record = _vh_lookup.get(sid)
            if record:
                conf  = record.get("confidence", 0)
                badge = "✓" if record.get("passed") else "✗"
                label = f"{sid:<30}  [{conf}/10 {badge}]"
            else:
                label = sid
            self.session_lb.insert("end", label)

    def _sessions_load_prior(self):
        sel = self.session_lb.curselection()
        if not sel:
            return
        sid = self.session_lb.get(sel[0]).split("  [")[0].strip()
        self.prior_session_id = sid
        self.prior_label.configure(text=f"Prior: {sid}")

        # ── #2 Session-end summary: generate once, reuse forever ────────────────
        # If no generated summary exists for this session yet, create one now
        # using the writer model. Subsequent loads hit the cached .summary.md file.
        # We use a background thread so the UI doesn't block.
        def _maybe_generate_summary():
            try:
                convo_store = self.writer.conversation_store
                if convo_store is None:
                    return
                existing = convo_store.load_generated_summary(sid)
                if existing:
                    return  # already generated
                turns = convo_store.load_last(sid, n=40)
                if not turns:
                    return
                # Build a raw transcript excerpt for the writer to compress
                raw_lines = []
                for t in turns:
                    who = t.get("who", "?")
                    text = t.get("text", "")[:600]
                    raw_lines.append(f"{who}: {text}")
                raw_excerpt = "\n".join(raw_lines)
                summary_prompt = (
                    f"Summarise this past council session in 200–350 words.\n"
                    f"Focus on: what the user was trying to achieve, key decisions made, "
                    f"what worked, what was unresolved, and anything worth remembering for "
                    f"future sessions.\n"
                    f"Write in plain prose — no bullet points, no headers.\n\n"
                    f"SESSION TRANSCRIPT (most recent {len(turns)} turns):\n"
                    f"{raw_excerpt}"
                )
                summary_text = self.writer.respond(summary_prompt, max_tokens=450)
                if summary_text.strip():
                    convo_store.save_session_summary(sid, summary_text)
                    self._append_transcript(
                        "Librarian", f"Session summary generated for '{sid}'.", "final"
                    )
            except Exception as _se:
                self._append_transcript(
                    "Librarian", f"Summary generation skipped: {_se}", "final"
                )

        import threading as _threading
        _threading.Thread(target=_maybe_generate_summary, daemon=True).start()

        # Rebuild personalities with new prior context
        pins = ce.load_personality_pins(PINS_PATH)
        self.personalities = ce.build_personalities(
            pins=pins, vault_dir=VAULT_DIR, session_id=self.session_id,
            trace=True, dispatcher=self.dispatcher, prior_session_id=sid,
        )
        self._unpack_personalities()
        self._append_transcript("Librarian", f"Prior session loaded: {sid}", "final")

    def _sessions_clear_prior(self):
        self.prior_session_id = None
        self.prior_label.configure(text="Prior: none")
        pins = ce.load_personality_pins(PINS_PATH)
        self.personalities = ce.build_personalities(
            pins=pins, vault_dir=VAULT_DIR, session_id=self.session_id,
            trace=True, dispatcher=self.dispatcher, prior_session_id=None,
        )
        self._unpack_personalities()
        self._append_transcript("Librarian", "Prior session cleared.", "final")

    def _sessions_analyse_trends(self):
        """Run cross-session trend analysis using Sage (or Writer as fallback)."""
        trend_model = getattr(self, "sage", None) or self.writer
        if trend_model is None:
            messagebox.showinfo("Trends", "No model available for trend analysis.")
            return
        self._append_transcript("Librarian", "Analysing cross-session trends…", "phase")

        def worker():
            try:
                memmgr = self.writer.memory_manager
                if memmgr is None:
                    self.ui_q.put(("agent_phase", "sage", "Trend analysis: no memory manager"))
                    return
                result = ce.generate_cross_session_trends(
                    trend_model=trend_model,
                    memory_manager=memmgr,
                    vault_dir=VAULT_DIR,
                )
                self.ui_q.put(("trend_result", result))
            except Exception as e:
                self.ui_q.put(("trend_result", f"(Error: {e})"))

        import threading as _t
        _t.Thread(target=worker, daemon=True).start()

    def _show_verdict_history(self):
        """Popup: summary table of all past verdicts from verdict_history.jsonl."""
        records = self._load_verdict_history()
        win = tk.Toplevel(self)
        win.title("Verdict History")
        win.configure(bg="#1a1414")
        win.geometry("900x500")
        if records:
            total    = len(records)
            passed   = sum(1 for r in records if r.get("passed"))
            avg_conf = round(sum(r.get("confidence", 0) for r in records) / total, 1)
            by_route = {}
            for r in records:
                rt = r.get("route", "?")
                by_route.setdefault(rt, {"total": 0, "passed": 0, "conf": []})
                by_route[rt]["total"]  += 1
                by_route[rt]["passed"] += int(r.get("passed", False))
                by_route[rt]["conf"].append(r.get("confidence", 0))
            route_parts = [
                rt + ":" + str(v["passed"]) + "/" + str(v["total"])
                + " (conf " + str(round(sum(v["conf"])/len(v["conf"]), 1)) + ")"
                for rt, v in sorted(by_route.items())
            ]
            summary = (
                str(total) + " deliberations  |  "
                + str(passed) + " passed (" + str(round(passed/total*100)) + "%)  |  "
                + "avg confidence " + str(avg_conf) + "/10\n"
                + "By route:  " + "   ".join(route_parts)
            )
        else:
            summary = "No verdict history yet. Run a deliberation to start tracking."
        ttk.Label(win, text=summary, wraplength=860, justify="left").pack(
            anchor="w", padx=10, pady=(8, 4))
        ttk.Separator(win, orient="horizontal").pack(fill="x", padx=10, pady=4)
        cols = ("ts", "session", "route", "conf", "passed", "rounds", "query")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=18)
        for col, hdr, w in [
            ("ts","Time",140), ("session","Session",180), ("route","Route",80),
            ("conf","Conf",50), ("passed","Pass",45), ("rounds","Rds",35), ("query","Query",340),
        ]:
            tree.heading(col, text=hdr)
            tree.column(col, width=w, stretch=(col == "query"))
        tree.tag_configure("pass", foreground="#a6e3a1")
        tree.tag_configure("fail", foreground="#f38ba8")
        tree.tag_configure("med",  foreground="#fab387")
        for r in reversed(records):
            _pass = r.get("passed", False)
            _conf = r.get("confidence", 0)
            tag   = "pass" if _pass else ("med" if _conf >= 4 else "fail")
            tree.insert("", "end", values=(
                r.get("ts","")[:16], r.get("session_id","")[:22],
                r.get("route",""), str(_conf)+"/10",
                "✓" if _pass else "✗",
                str(r.get("rounds", 0)), r.get("query","")[:80],
            ), tags=(tag,))
        sb = ttk.Scrollbar(win, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=(10,0), pady=(0,10))

    def _sessions_preview(self):
        sel = self.session_lb.curselection()
        if not sel:
            return
        sid = self.session_lb.get(sel[0]).split("  [")[0].strip()
        summary = self.convo_store.load_session_summary(sid, max_turns=20)
        self._set_text(self.session_preview, summary or "(empty session)")

    # ---- Node status ----

    def _refresh_nodes_async(self):
        # The Nodes tab only exists in advanced mode. In the consumer
        # build the dispatcher still works, but there's no UI to update —
        # so skip the probe entirely. This is also called on startup
        # (self.after(2000, ...)) so the guard is critical.
        if not hasattr(self, "tab_nodes"):
            return
        def worker():
            statuses = self.dispatcher.probe_all()
            self.ui_q.put(("nodes_result", statuses))
        threading.Thread(target=worker, daemon=True).start()
        if hasattr(self, "nodes_status_label"):
            try:
                self.nodes_status_label.configure(text="Probing…")
            except tk.TclError:
                pass

    def _populate_nodes_tree(self, statuses: List[ce.NodeStatus]):
        # No-op when the Nodes tab isn't built (consumer mode)
        if not hasattr(self, "nodes_tree"):
            return
        for row in self.nodes_tree.get_children():
            self.nodes_tree.delete(row)
        for s in statuses:
            status_str = "● up" if s.reachable else "✕ down"
            latency_str = f"{s.latency_ms:.0f} ms" if s.reachable else "—"
            active_str = ", ".join(s.active_model_names) if s.active_model_names else ("none" if s.reachable else "—")
            installed_str = ", ".join(s.installed_models[:6]) if s.installed_models else "—"
            if len(s.installed_models) > 6:
                installed_str += f" (+{len(s.installed_models)-6} more)"
            tag = "up" if s.reachable else "down"
            self.nodes_tree.insert("", "end",
                values=(s.host, status_str, latency_str, active_str, installed_str),
                tags=(tag,))

    def _apply_pi_hosts(self):
        raw = self.pi_hosts_var.get()
        hosts = [h.strip() for h in raw.split(",") if h.strip()]
        self.dispatcher = ce.build_dispatcher(extra_hosts=hosts)
        pins = ce.load_personality_pins(PINS_PATH)
        self.personalities = ce.build_personalities(
            pins=pins, vault_dir=VAULT_DIR, session_id=self.session_id,
            trace=True, dispatcher=self.dispatcher, prior_session_id=self.prior_session_id,
        )
        self._unpack_personalities()
        self._append_transcript("Librarian", f"Dispatcher rebuilt with hosts: {hosts}", "final")
        self._refresh_nodes_async()

    # ---- Speech actions ----

    def _stt_record(self):
        ok, msg = self.speech.ready()
        if not ok:
            messagebox.showwarning("Speech→Text", msg)
            return
        wav_path = TMP_DIR / "latest.wav"

        def worker():
            try:
                ok2, msg2 = self.speech.record_wav(wav_path, seconds=5)
                self.ui_q.put(("stt_out", msg2 if ok2 else f"ERROR: {msg2}"))
            except Exception as e:
                self.ui_q.put(("error", str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _stt_transcribe(self):
        wav_path = TMP_DIR / "latest.wav"
        if not wav_path.exists():
            messagebox.showwarning("Speech→Text", "No recording found. Record first.")
            return

        def worker():
            try:
                ok, text = self.speech.transcribe(wav_path)
                self.ui_q.put(("stt_out", text if ok else f"ERROR: {text}"))
            except Exception as e:
                self.ui_q.put(("error", str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _stt_send_to_council(self):
        text = self.stt_out.get("1.0", "end").strip()
        if not text:
            return
        self._set_text(self.input, text)
        self.nb.select(self.tab_council)
        self._send()

    # ---- TTS actions (#10) ----

    def _tts_get_engine(self):
        """Lazily initialise pyttsx3 TTS engine. Returns None if unavailable."""
        if self._tts_engine is not None:
            return self._tts_engine
        try:
            import pyttsx3 as _pyttsx3
            eng = _pyttsx3.init()
            self._tts_engine = eng
            return eng
        except Exception:
            return None

    def _tts_speak_last(self):
        """Speak the last Writer final answer in a background thread."""
        text = getattr(self, "_last_final_text", "").strip()
        if not text:
            self._append_transcript("Librarian", "Nothing to speak yet.", "final")
            return
        def worker():
            eng = self._tts_get_engine()
            if eng is None:
                self.ui_q.put(("agent_phase", "speech",
                               "TTS unavailable — install pyttsx3 (pip install pyttsx3)"))
                return
            try:
                # Robust rate lookup: tolerates missing var or non-numeric content
                rate = 175
                rv = getattr(self, "_tts_rate_var", None)
                if rv is not None:
                    try:
                        rate = int(rv.get())
                    except (TypeError, ValueError):
                        rate = 175
                eng.setProperty("rate", rate)
                eng.say(text[:4000])  # cap at 4000 chars for sane length
                eng.runAndWait()
            except Exception as _te:
                self.ui_q.put(("agent_phase", "speech", f"TTS error: {_te}"))
        import threading as _t
        _t.Thread(target=worker, daemon=True).start()

    def _tts_stop(self):
        """Stop TTS playback."""
        eng = getattr(self, "_tts_engine", None)
        if eng is not None:
            try:
                eng.stop()
            except Exception:
                pass

    def _tts_auto_speak(self, text: str):
        """Called after each deliberation if auto-speak is enabled."""
        if not getattr(self, "var_tts_auto", None) or not self.var_tts_auto.get():
            return
        def worker():
            eng = self._tts_get_engine()
            if eng is None:
                return
            try:
                rate = int(self._tts_rate_var.get()) if hasattr(self, "_tts_rate_var") else 175
                eng.setProperty("rate", rate)
                eng.say(text[:4000])
                eng.runAndWait()
            except Exception:
                pass
        import threading as _t
        _t.Thread(target=worker, daemon=True).start()


# ============================================================
# Utilities
# ============================================================

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _fmt_bytes(n: int) -> str:
    """Human-readable byte size."""
    for unit in ("B", "KB", "MB"):
        if n < 1024:
            return f"{n}{unit}"
        n //= 1024
    return f"{n}MB"


# Hard reference to the live root window. Under an interactive host
# (Spyder / IPython / Jupyter) the kernel already runs a Tk event loop,
# so app.mainloop() returns immediately instead of blocking; main()
# then returns and its `app` local would go out of scope, letting the
# withdrawn root get garbage-collected before the splash ever reveals
# it — the "splash shows then everything vanishes" bug. Keeping the
# reference at module scope pins the window for the kernel's loop.
_APP = None


def _under_interactive_host() -> bool:
    """True when running inside an IPython / Spyder / Jupyter kernel
    rather than a plain ``python …`` launch or the frozen .exe.

    Such hosts run their OWN Qt event loop and worker threads in the
    same process, which makes two things from the normal startup path
    unsafe:
      • pumping a Tk splash with update_idletasks() while the host's Qt
        loop is also live, and
      • initialising torch / CUDA on a BACKGROUND thread (the RAG
        indexer) — off-main-thread CUDA init segfaults the kernel about
        30 s in, which is the "no tabs then the kernel dies" report.

    Under an interactive host we take a simpler, deterministic path: no
    splash, reveal the window immediately, and skip the auto-start of
    the torch-loading RAG thread. Vault keyword search still works;
    full-document semantic RAG can be built on demand from the Vault
    tab. The normal .exe / CLI launch is unaffected.
    """
    try:
        from IPython import get_ipython  # type: ignore
        if get_ipython() is not None:
            return True
    except Exception:
        pass
    return ("spyder_kernels" in sys.modules) or ("spyder" in sys.modules)


def main():
    global _APP

    # Install system-level crash hooks BEFORE the GUI exists so any error
    # during startup is captured. Tk hook gets wired after the root window
    # is created (inside CouncilConsole.__init__).
    def _on_crash(crash_path):
        # Defer dialog to the Tk main loop — show_dialog needs a parent.
        try:
            if _APP and _APP.winfo_exists():
                _APP.after(0, lambda: crash_reporter.show_dialog(_APP, crash_path))
        except Exception:
            pass
    crash_reporter.install(VAULT_DIR, on_crash=_on_crash)

    # The splash is created and pumped INSIDE CouncilConsole.__init__ so
    # the spinning cog covers the entire construction (heavy model /
    # personality wiring + the ~15-tab build). The root is withdrawn the
    # whole time. By the time CouncilConsole() returns, construction is
    # done and the cog has been turning; here we just enforce a minimum
    # on-screen time, then dismiss the splash and reveal the window.
    app = CouncilConsole()
    _APP = app                      # pin against GC under interactive hosts
    crash_reporter.install_tk_hook(app, VAULT_DIR, on_crash=_on_crash)

    # Reveal is made bulletproof with independent guarantees, because a
    # withdrawn root that never re-appears is a dead app: an idempotent
    # _reveal(), the splash dismiss's on_done (normal path), AND a
    # backstop timer — so an event-loop quirk can't strand the window.
    _MIN_SPLASH_MS = 1500
    _revealed = {"done": False}

    def _reveal():
        if _revealed["done"]:
            return
        _revealed["done"] = True
        try:
            app.deiconify()
            app.lift()
            app.focus_force()
        except Exception as e:
            # Do NOT swallow silently — a failed reveal means a blank
            # session, and the reason needs to reach the console.
            print(f"[Splash] reveal failed, showing window directly: {e!r}")
            try:
                app.deiconify()
            except Exception:
                pass

    def _finish_splash_and_reveal():
        sp = getattr(app, "_splash", None)
        if sp is not None:
            try:
                sp.dismiss(on_done=_reveal)   # destroy cog, then reveal
                return
            except Exception:
                pass
        _reveal()

    if getattr(app, "_interactive_host", False):
        # No splash under Spyder/IPython — __init__ already deiconified;
        # this just guarantees it and avoids relying on deferred timers
        # the host's loop may pump only intermittently.
        _reveal()
    else:
        try:
            import time as _time
            started = getattr(app, "_splash_started", _time.monotonic())
            elapsed_ms = (_time.monotonic() - started) * 1000.0
            remaining = max(0, int(_MIN_SPLASH_MS - elapsed_ms))
            # During `remaining`, the splash's own after-loop spins the cog
            # smoothly (the Tk loop is live now). Then dismiss + reveal.
            app.after(remaining, _finish_splash_and_reveal)
            # Backstop — guarantees the window appears even if the dismiss
            # path or its on_done never completes.
            app.after(remaining + 1500, _reveal)
        except Exception as e:
            print(f"[Splash] reveal scheduling failed: {e}")
            _reveal()

    # Always run our own blocking loop — the standard way every Tk app
    # runs, including inside Spyder/IPython. The earlier "skip mainloop
    # if the host pumps Tk" heuristic was wrong: when skipped, main()
    # returns and the script ends, but the live root, its pending
    # after() timers (reveal, _poll_ui_queue), and the torch-loading RAG
    # daemon thread were left in an undefined state — the reveal never
    # fired ("no tabs") and the process segfaulted once torch finished
    # loading (~30s, "kernel dies"). Owning the loop on the main thread
    # keeps everything deterministic. Under Spyder this makes the kernel
    # "busy" until the window is closed — which is correct for a GUI app.
    app.mainloop()


def _purge_stale_pycache():
    """
    Defensive clean-up against a Python gotcha: if a .pyc file ends up
    NEWER than its .py source (can happen after some git operations,
    file-system mtime resets, or editor weirdness), Python keeps using
    the stale bytecode and a fix won't take effect.

    On launch, scan our own folder for any pyc/py pairs where the pyc is
    newer, and delete the offending pyc. The next import recompiles.
    Costs ~50 ms; saves an entire debugging session when something
    goes weird.

    Note: by the time this runs, council_gui_engine.pyc itself has
    already been loaded — so we can't unstick our own module from
    here. But every other module imported by main() benefits, AND
    deleting our stale .pyc means the user's NEXT launch is correct
    even if THIS one isn't.
    """
    try:
        from pathlib import Path as _P
        root = _P(__file__).resolve().parent
        cache_root = root / "__pycache__"
        if not cache_root.exists():
            return
        cleared = 0
        for pyc in cache_root.glob("*.pyc"):
            stem = pyc.stem.split(".cpython")[0]
            src = root / f"{stem}.py"
            if not src.exists():
                continue
            try:
                if pyc.stat().st_mtime > src.stat().st_mtime + 0.5:
                    pyc.unlink()
                    cleared += 1
            except OSError:
                pass
        if cleared:
            print(f"[Cache] Cleared {cleared} stale .pyc file(s) — "
                  f"if a fix still seems missing, restart once more.")
    except Exception:
        pass


if __name__ == "__main__":
    _purge_stale_pycache()
    print("=" * 60)
    print(f" {branding.PRODUCT_NAME} v{branding.VERSION}")
    print(f" {branding.PRODUCT_TAGLINE}")
    print("=" * 60)
    print()
    main()
