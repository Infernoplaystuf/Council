"""
Vault analyst — pandas-backed helper functions over the user's vault, plus a
sandboxed runner that lets the council compute real answers instead of guessing
from sample rows.

The shape is taken from a local-RAG reference script:
  - A library of typed helper functions covering the most common CSV questions
    (count, average, std, sum, min/max, filter, groupby, correlation, etc.).
  - A code generator that asks the model to express a question as a single
    pandas expression assigning to `result_df`.
  - An AST + regex validator that blocks anything the council shouldn't run
    (network, filesystem writes, shells, eval/exec).
  - A locked-down `exec` with a restricted builtins map and a `safe_import`
    allowlist.

The council uses this from the Writer's pre-synthesis step: if the user is
asking something computational ("how many", "what percentage", "average X"),
the helpers run first and their result is injected as text the Writer can
quote directly.
"""

from __future__ import annotations

import ast
import io
import json
import os
import re
import traceback
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Any, Iterable, List, Optional, Tuple

# Process-wide LRU cache for parsed DataFrames. The per-file inventory
# helpers below (csv_inventory, count_rows_per_csv, average_numeric_
# column_per_csv, etc.) used to call pd.read_csv from cold for every
# question — a single user turn that triggers two helpers re-reads each
# CSV twice. Routing through df_cache gives in-turn cache hits for free.
try:
    from df_cache import cached_read_csv as _cached_read_csv
except Exception:
    _cached_read_csv = None  # graceful — fall back to pd.read_csv per call


def _read_csv_cached(path, **kw):
    """Wrapper that prefers df_cache.cached_read_csv but degrades to a
    bare pandas read when the cache is unavailable. Returns the same
    DataFrame regardless — callers see no behavioural difference."""
    if _cached_read_csv is not None:
        try:
            return _cached_read_csv(path, **kw)
        except Exception:
            # Defensive: if the cache misbehaves for any reason, fall
            # through to the uncached path so the analyst still answers.
            pass
    import pandas as _pd_inner
    return _pd_inner.read_csv(path, **kw)


try:
    import pandas as pd
except Exception as exc:  # pragma: no cover - imported at runtime
    raise RuntimeError(
        "vault_analyst requires pandas. Install with `conda install -c "
        "conda-forge pandas` in the council env."
    ) from exc

try:
    import numpy as np
except Exception:
    np = None  # numpy is optional but useful in user code


# ============================================================
# Path helpers
# ============================================================

def normalize_data_folders(data_folders: Any) -> List[Path]:
    """Accept str / Path / iterable; return a deduped list of existing dirs."""
    if data_folders is None:
        return []
    if isinstance(data_folders, (str, Path)):
        raw = [data_folders]
    else:
        raw = list(data_folders)

    folders: List[Path] = []
    seen: set[str] = set()
    for item in raw:
        try:
            p = Path(item).expanduser().resolve()
        except Exception:
            continue
        if p.is_dir() and str(p).lower() not in seen:
            folders.append(p)
            seen.add(str(p).lower())
    return folders


def is_path_inside_allowed_folders(path: str | Path, allowed_folders: Any) -> bool:
    try:
        p = Path(path).expanduser().resolve()
    except Exception:
        return False
    for folder in normalize_data_folders(allowed_folders):
        try:
            p.relative_to(folder)
            return True
        except ValueError:
            continue
    return False


def first_matching_root(path: Path, roots: Any) -> Optional[Path]:
    path = Path(path).expanduser().resolve()
    for root in normalize_data_folders(roots):
        try:
            path.relative_to(root)
            return root
        except ValueError:
            continue
    return None


def safe_relative_path(path: Path, root: Optional[Path]) -> str:
    try:
        if root is not None:
            return str(Path(path).resolve().relative_to(Path(root).resolve()))
    except Exception:
        pass
    return str(Path(path).resolve())


def _drop_protected(paths: List[Path], data_folder: Any) -> List[Path]:
    """Filter out any file under a protected vault subdir (conversation_logs etc.).

    Resolves each folder once and each path once, rather than re-resolving both
    on every (path, folder) pair inside is_protected_path — same kept/dropped
    set and same input order. An un-resolvable path is KEPT (matching
    is_protected_path returning False), never crashing the walk.
    """
    try:
        from conversation_logger import _PROTECTED_SUBDIRS_LC
    except Exception:
        # Older conversation_logger without the hoisted set — fall back to the
        # per-item helper (still correct, just slower).
        try:
            from conversation_logger import is_protected_path
        except Exception:
            return paths
        folders = normalize_data_folders(data_folder)
        return [p for p in paths
                if not any(is_protected_path(p, vd) for vd in folders)]

    folders = normalize_data_folders(data_folder)
    res_folders: List[Path] = []
    for vd in folders:
        try:
            res_folders.append(Path(vd).expanduser().resolve())
        except Exception:
            pass  # unresolvable folder: skip it (is_protected_path -> False)
    out: List[Path] = []
    for p in paths:
        try:
            rp = Path(p).expanduser().resolve()
        except Exception:
            out.append(p)   # unresolvable path is kept (matches False return)
            continue
        protected = False
        for fr in res_folders:
            try:
                rel = rp.relative_to(fr)
            except ValueError:
                continue
            parts = rel.parts
            if parts and parts[0].lower() in _PROTECTED_SUBDIRS_LC:
                protected = True
                break
        if not protected:
            out.append(p)
    return out


def list_csv_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """All CSVs under the given folder(s), deduped. Protected subdirs
    (conversation logs) are excluded — they're for the user only."""
    folders = normalize_data_folders(data_folder)
    out: List[Path] = []
    for folder in folders:
        if recursive:
            out.extend(sorted(folder.rglob("*.csv")))
        else:
            out.extend(sorted(folder.glob("*.csv")))

    deduped: dict[str, Path] = {}
    for path in out:
        try:
            _rp = path.resolve()          # resolve once (was twice per path)
            deduped[str(_rp).lower()] = _rp
        except Exception:
            pass
    return _drop_protected(sorted(deduped.values()), folders)


_EXCEL_GLOBS = ("*.xlsx", "*.xls", "*.xlsm")


def list_excel_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """All Excel workbooks under the given folder(s), deduped."""
    folders = normalize_data_folders(data_folder)
    out: List[Path] = []
    for folder in folders:
        for pattern in _EXCEL_GLOBS:
            if recursive:
                out.extend(sorted(folder.rglob(pattern)))
            else:
                out.extend(sorted(folder.glob(pattern)))
    deduped: dict[str, Path] = {}
    for path in out:
        try:
            _rp = path.resolve()          # resolve once (was twice per path)
            deduped[str(_rp).lower()] = _rp
        except Exception:
            pass
    return _drop_protected(sorted(deduped.values()), folders)


def list_data_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """CSV + Excel files combined — for helpers that want either."""
    return sorted(set(list_csv_files(data_folder, recursive=recursive))
                  | set(list_excel_files(data_folder, recursive=recursive)))


# Internal/cache dirs that aren't user data — never counted as "files".
_CENSUS_SKIP_DIRS = {".stats_cache", ".vault_index", "__pycache__",
                     ".git", "conversation_logs",
                     # App-generated COMPUTED outputs — not source data, so a
                     # file-count / data-summary should not inflate with them.
                     "derived", "deferred_results"}


def folder_file_counts(data_folder: Any, recursive: bool = True
                       ) -> Dict[str, Any]:
    """Cheap file census — total count + breakdown by extension — WITHOUT
    reading any file contents. Answers "how many files are in <folder>"
    deterministically and in a tiny prompt, so it never overflows a small
    context window the way model code-gen would. Skips hidden entries and
    the app's internal cache dirs. Returns
    {"total": int, "folders": int, "by_ext": {ext: count}}.
    """
    folders = normalize_data_folders(data_folder)
    total = 0
    nfolders = 0
    by_ext: Dict[str, int] = {}
    # os.walk (not Path.rglob) so we can (a) PRUNE internal/hidden dirs
    # instead of descending into them, and (b) survive a permission-denied
    # or broken-symlink dir via the onerror callback — rglob would raise
    # mid-iteration and abort the whole census (real risk under WSL /mnt).
    import os as _os
    for root in folders:
        rp = Path(root)
        if not rp.exists():
            continue
        if not recursive:
            try:
                for p in rp.iterdir():
                    if p.name in _CENSUS_SKIP_DIRS or p.name.startswith("."):
                        continue
                    try:
                        if p.is_dir():
                            nfolders += 1
                        elif p.is_file():
                            total += 1
                            ext = p.suffix.lower() or "(no extension)"
                            by_ext[ext] = by_ext.get(ext, 0) + 1
                    except OSError:
                        continue
            except OSError:
                continue
            continue
        for dirpath, dirnames, filenames in _os.walk(
                str(rp), onerror=lambda _e: None):
            # Prune internal/hidden subdirs in place so os.walk skips them.
            dirnames[:] = [d for d in dirnames
                           if d not in _CENSUS_SKIP_DIRS and not d.startswith(".")]
            nfolders += len(dirnames)
            for fn in filenames:
                if fn.startswith("."):
                    continue
                total += 1
                ext = Path(fn).suffix.lower() or "(no extension)"
                by_ext[ext] = by_ext.get(ext, 0) + 1
    return {"total": total, "folders": nfolders, "by_ext": by_ext}


# ============================================================
# Messy-data helpers — robust CSV / multi-table Excel / reshape
# ============================================================

_ENCODING_CANDIDATES = ("utf-8", "utf-8-sig", "cp1252", "latin-1")
_SEPARATOR_CANDIDATES = (",", ";", "\t", "|")


def _sniff_csv_separator(sample: str) -> str:
    """Pick the separator that gives the most consistent column count."""
    best_sep, best_consistency = ",", 0
    for sep in _SEPARATOR_CANDIDATES:
        lines = [ln for ln in sample.splitlines() if ln.strip()][:50]
        if len(lines) < 2:
            continue
        counts = [ln.count(sep) for ln in lines]
        if not counts or max(counts) == 0:
            continue
        # Pick the separator where the mode column count repeats most
        from collections import Counter as _C
        mode_count, mode_freq = _C(counts).most_common(1)[0]
        if mode_count > 0 and mode_freq > best_consistency:
            best_sep, best_consistency = sep, mode_freq
    return best_sep


def _detect_header_row(path: Path, encoding: str, sep: str,
                       max_check: int = 20) -> int:
    """Find the first row that looks like a real CSV header.

    Heuristic: scan up to `max_check` rows; the header is the first row
    whose column count matches the *modal* column count of the rows
    that follow it. Skips title/banner rows where row 1 is a single
    cell of text and the real data starts at row 3 or 4.
    """
    try:
        with open(path, encoding=encoding, errors="replace", newline="") as fh:
            rows: list[list[str]] = []
            import csv as _csv
            reader = _csv.reader(fh, delimiter=sep)
            for i, row in enumerate(reader):
                rows.append(row)
                if i >= max_check + 10:
                    break
    except Exception:
        return 0
    if not rows:
        return 0
    # Median column count of the bottom half (data rows are usually
    # consistent; title rows are 1-cell anomalies).
    from statistics import median as _median
    later_lens = [len(r) for r in rows[max_check:max_check + 20]] or \
                 [len(r) for r in rows[max(0, len(rows)//2):]]
    target = int(_median(later_lens)) if later_lens else 1
    for i, row in enumerate(rows[:max_check]):
        if len(row) == target and any(c.strip() for c in row):
            return i
    return 0


def read_csv_robust(path: Any) -> tuple[pd.DataFrame, dict]:
    """Read a CSV that may have any of: a non-utf-8 encoding, a separator
    other than comma, a title/banner row before the headers, or trailing
    summary rows.

    Returns (df, diagnostics) where diagnostics has keys:
      encoding, separator, header_row_index, footer_rows_stripped.
    """
    p = Path(path)
    diag: dict = {"encoding": None, "separator": None,
                  "header_row_index": 0, "footer_rows_stripped": 0}
    # 1. Encoding sniff — read first 8 KB
    raw = b""
    try:
        with open(p, "rb") as fh:
            raw = fh.read(8192)
    except Exception:
        return pd.DataFrame(), diag
    encoding = "utf-8"
    for enc in _ENCODING_CANDIDATES:
        try:
            raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    diag["encoding"] = encoding
    # 2. Separator sniff on a decoded sample
    sample = raw.decode(encoding, errors="replace")
    sep = _sniff_csv_separator(sample)
    diag["separator"] = sep
    # 3. Header-row detect
    header_row = _detect_header_row(p, encoding, sep)
    diag["header_row_index"] = header_row
    # 4. Read with the determined params
    try:
        df = pd.read_csv(p, encoding=encoding, sep=sep,
                         skiprows=header_row, on_bad_lines="skip")
    except Exception:
        return pd.DataFrame(), diag
    # 5. Strip trailing summary rows
    df, n_stripped = _strip_summary_rows_inplace(df)
    diag["footer_rows_stripped"] = n_stripped
    return df, diag


_SUMMARY_PATTERNS = (
    "total", "grand total", "sub total", "subtotal", "sum",
    "average", "avg", "mean", "count",
)


def _strip_summary_rows_inplace(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Drop trailing rows whose first non-null column matches a summary
    word. Returns (cleaned_df, rows_dropped)."""
    if df is None or df.empty:
        return df, 0
    n_dropped = 0
    # Walk from the bottom up
    for _ in range(min(10, len(df))):
        last = df.iloc[-1]
        try:
            first_val = next((str(v) for v in last.values
                              if v is not None and str(v).strip()),
                             "")
        except Exception:
            break
        if first_val.strip().lower() in _SUMMARY_PATTERNS or \
           any(p in first_val.strip().lower() for p in _SUMMARY_PATTERNS):
            df = df.iloc[:-1]
            n_dropped += 1
        else:
            break
    return df.reset_index(drop=True), n_dropped


def strip_summary_rows(
    df: pd.DataFrame,
    *,
    patterns: tuple = _SUMMARY_PATTERNS,
) -> pd.DataFrame:
    """Public wrapper — drop trailing summary rows. Use custom patterns
    if your data uses different terminology."""
    global _SUMMARY_PATTERNS
    saved = _SUMMARY_PATTERNS
    try:
        _SUMMARY_PATTERNS = tuple(p.lower() for p in patterns)
        df2, _ = _strip_summary_rows_inplace(df)
        return df2
    finally:
        _SUMMARY_PATTERNS = saved


def find_data_block(path_or_df: Any) -> dict:
    """Locate the actual tabular data region in a messy CSV/Excel sheet.

    For a file: reads it, detects header row + extent of consistent rows.
    For a DataFrame: scans for the bounding box of non-null cells.

    Returns a dict with: start_row, end_row, start_col, end_col, n_rows,
    n_cols. Rows/cols are 0-based.
    """
    if isinstance(path_or_df, pd.DataFrame):
        df = path_or_df
    else:
        p = Path(path_or_df)
        if p.suffix.lower() == ".csv":
            df, _ = read_csv_robust(p)
        elif p.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
            df = pd.read_excel(p)
        else:
            return {"error": "unsupported file type"}
    if df is None or df.empty:
        return {"start_row": 0, "end_row": 0, "start_col": 0, "end_col": 0,
                "n_rows": 0, "n_cols": 0}
    non_null = df.notna()
    rows_with_data = non_null.any(axis=1)
    cols_with_data = non_null.any(axis=0)
    if not rows_with_data.any() or not cols_with_data.any():
        return {"start_row": 0, "end_row": 0, "start_col": 0, "end_col": 0,
                "n_rows": 0, "n_cols": 0}
    start_row = int(rows_with_data.idxmax())
    end_row = int(rows_with_data[::-1].idxmax())
    col_indices = [i for i, has in enumerate(cols_with_data) if has]
    start_col = col_indices[0]
    end_col = col_indices[-1]
    return {
        "start_row": start_row, "end_row": end_row,
        "start_col": start_col, "end_col": end_col,
        "n_rows": end_row - start_row + 1,
        "n_cols": end_col - start_col + 1,
    }


def read_excel_smart_tables(
    path: Any,
    sheet: Any = 0,
    *,
    gap_tolerance: int = 1,
    min_cells: int = 4,
    expand_merged: bool = True,
) -> List[Dict[str, Any]]:
    """Robust multi-table reader for messy spreadsheets.

    Handles: multiple tables stacked vertically OR side-by-side, gaps
    inside tables, merged cells (expanded so each underlying cell gets
    the top-left value), inconsistent row widths (short rows padded),
    leading/trailing blank rows/cols (trimmed), and header rows that
    don't start at row 1 (auto-detected per block).

    `gap_tolerance` is how many consecutive blank rows OR columns are
    treated as "still part of the same table". 1 means single-cell
    blanks won't split a table; 0 means any blank breaks it.

    `min_cells` filters out tiny noise blocks (a stray label in a
    corner won't be returned as its own "table").

    Returns a list of dicts:
      {
        "df":        pandas DataFrame,
        "top_left":  (row, col)  — 0-based cell coords in the sheet,
        "n_rows":    int,
        "n_cols":    int,
        "header_row": row offset within the block,
      }
    """
    p = Path(path)
    try:
        import openpyxl as _oxl
        wb = _oxl.load_workbook(p, read_only=False, data_only=True)
    except Exception:
        # Fall back to pandas's all-in-one read
        return [{"df": pd.read_excel(p, sheet_name=sheet),
                 "top_left": (0, 0), "n_rows": 0, "n_cols": 0,
                 "header_row": 0}]
    try:
        if isinstance(sheet, int) and 0 <= sheet < len(wb.worksheets):
            ws = wb.worksheets[sheet]
        elif isinstance(sheet, str) and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return []

        # Materialize the cell grid as a 2D list of values, expanding
        # merged ranges so every cell in the merged region holds the
        # top-left value (otherwise only one cell has a value).
        grid: List[List[Any]] = [
            [None] * max_col for _ in range(max_row)
        ]
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                grid[r - 1][c - 1] = cell.value

        if expand_merged:
            for rng in ws.merged_cells.ranges:
                anchor = grid[rng.min_row - 1][rng.min_col - 1]
                for r in range(rng.min_row, rng.max_row + 1):
                    for c in range(rng.min_col, rng.max_col + 1):
                        if r - 1 < max_row and c - 1 < max_col:
                            if grid[r - 1][c - 1] in (None, "", " "):
                                grid[r - 1][c - 1] = anchor
    finally:
        try: wb.close()
        except Exception: pass

    # Build occupancy mask
    occ = [
        [(v is not None and str(v).strip() != "") for v in row]
        for row in grid
    ]

    # Flood-fill connected components with gap tolerance.
    # We dilate the occupancy mask by `gap_tolerance` cells so blanks
    # smaller than the tolerance get bridged.
    R, C = len(occ), (len(occ[0]) if occ else 0)
    if R == 0 or C == 0:
        return []
    dilated = [[False] * C for _ in range(R)]
    g = max(0, int(gap_tolerance))
    for r in range(R):
        for c in range(C):
            if occ[r][c]:
                for dr in range(-g, g + 1):
                    for dc in range(-g, g + 1):
                        rr, cc = r + dr, c + dc
                        if 0 <= rr < R and 0 <= cc < C:
                            dilated[rr][cc] = True

    # 8-connectivity flood-fill on the dilated mask to find clusters
    visited = [[False] * C for _ in range(R)]
    clusters: List[List[Tuple[int, int]]] = []
    for r in range(R):
        for c in range(C):
            if not dilated[r][c] or visited[r][c]:
                continue
            stack = [(r, c)]
            cluster: List[Tuple[int, int]] = []
            while stack:
                rr, cc = stack.pop()
                if not (0 <= rr < R and 0 <= cc < C):
                    continue
                if visited[rr][cc] or not dilated[rr][cc]:
                    continue
                visited[rr][cc] = True
                if occ[rr][cc]:                    # only the real cells count
                    cluster.append((rr, cc))
                for dr in (-1, 0, 1):
                    for dc in (-1, 0, 1):
                        if dr == 0 and dc == 0:
                            continue
                        stack.append((rr + dr, cc + dc))
            if len(cluster) >= min_cells:
                clusters.append(cluster)

    out: List[Dict[str, Any]] = []
    for cluster in clusters:
        rows = [p[0] for p in cluster]
        cols = [p[1] for p in cluster]
        r0, r1 = min(rows), max(rows)
        c0, c1 = min(cols), max(cols)

        # Slice the grid; pad to a rectangle.
        sub: List[List[Any]] = []
        for r in range(r0, r1 + 1):
            row_slice = grid[r][c0:c1 + 1]
            # Pad short rows with None to keep rectangular
            if len(row_slice) < (c1 - c0 + 1):
                row_slice = row_slice + [None] * (c1 - c0 + 1 - len(row_slice))
            sub.append(row_slice)

        # Header-row detection within the cluster: pick the FIRST row
        # whose filled-cell count equals the modal filled-cell count
        # of rows below it. That's typically the header.
        fill_counts = [sum(1 for v in row if v not in (None, "", " "))
                       for row in sub]
        if len(fill_counts) > 1:
            tail_mode = max(set(fill_counts[1:]), key=fill_counts[1:].count)
            header_idx = 0
            for i, fc in enumerate(fill_counts):
                if fc == tail_mode and any(
                    v is not None and str(v).strip() for v in sub[i]
                ):
                    header_idx = i
                    break
        else:
            header_idx = 0

        # Build DataFrame
        try:
            header = [
                str(v).strip() if v is not None else f"col_{i}"
                for i, v in enumerate(sub[header_idx])
            ]
            # Dedupe column names (Excel allows duplicates; pandas doesn't)
            seen: Dict[str, int] = {}
            unique_header: List[str] = []
            for h in header:
                base = h if h else "(blank)"
                if base in seen:
                    seen[base] += 1
                    unique_header.append(f"{base}.{seen[base]}")
                else:
                    seen[base] = 0
                    unique_header.append(base)

            body = sub[header_idx + 1:]
            df = pd.DataFrame(body, columns=unique_header)
            # Trim entirely empty trailing rows / cols inside the block
            df = df.dropna(axis=1, how="all")
            df = df.dropna(axis=0, how="all").reset_index(drop=True)
            if df.empty:
                continue
            out.append({
                "df":         df,
                "top_left":   (r0, c0),
                "n_rows":     len(df),
                "n_cols":     len(df.columns),
                "header_row": header_idx,
            })
        except Exception:
            continue

    # Order clusters top-to-bottom, then left-to-right
    out.sort(key=lambda d: (d["top_left"][0], d["top_left"][1]))
    return out


def read_excel_all_tables(path: Any, sheet: Any = 0) -> List[pd.DataFrame]:
    """Find every separate tabular block in one Excel sheet.

    Many workbooks stack multiple small tables on one sheet separated by
    blank rows. This function returns each block as its own DataFrame
    with its detected header row. Useful when `pd.read_excel` only gives
    you the first table.
    """
    p = Path(path)
    try:
        import openpyxl as _oxl
        wb = _oxl.load_workbook(p, read_only=True, data_only=True)
    except Exception:
        return [pd.read_excel(p, sheet_name=sheet)]
    try:
        if isinstance(sheet, int) and 0 <= sheet < len(wb.worksheets):
            ws = wb.worksheets[sheet]
        elif isinstance(sheet, str) and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]

        # Build a 2D occupancy grid: row_idx -> set of col indices
        # where cells have values.
        rows_with_data: list[set] = []
        for row in ws.iter_rows(values_only=True):
            cells = {i for i, v in enumerate(row)
                     if v is not None and str(v).strip()}
            rows_with_data.append(cells)

        # Find contiguous runs of non-empty rows separated by 1+ blank rows.
        runs: list[tuple[int, int]] = []   # (start_row, end_row) 0-based
        i = 0
        while i < len(rows_with_data):
            if rows_with_data[i]:
                start = i
                while i < len(rows_with_data) and rows_with_data[i]:
                    i += 1
                runs.append((start, i - 1))
            else:
                i += 1
    finally:
        try: wb.close()
        except Exception: pass

    # For each run, read that slice with pandas and treat the first row
    # as the header.
    out: list[pd.DataFrame] = []
    for (r_start, r_end) in runs:
        if r_end - r_start < 1:
            continue   # 1-row "table" is probably a title, not data
        try:
            df = pd.read_excel(
                p, sheet_name=sheet, header=r_start,
                nrows=r_end - r_start,
            )
            # Drop fully-empty columns
            df = df.dropna(axis=1, how="all")
            if not df.empty:
                out.append(df)
        except Exception:
            continue
    return out or [pd.read_excel(p, sheet_name=sheet)]


def unpivot_year_columns(
    df: pd.DataFrame,
    id_cols: Optional[List[str]] = None,
    *,
    year_min: int = 1900,
    year_max: int = 2100,
    value_name: str = "value",
) -> pd.DataFrame:
    """Convert a wide-form DataFrame with year columns into long form.

    Auto-detects which columns are years (4-digit integers in
    [year_min, year_max]). Non-year columns are treated as id_vars
    unless explicit `id_cols` is provided.

    Example:
      country | 2020 | 2021 | 2022
      US      | 100  | 120  | 140

    Becomes:
      country | year | value
      US      | 2020 | 100
      US      | 2021 | 120
      US      | 2022 | 140
    """
    if df is None or df.empty:
        return df
    year_cols: list[str] = []
    for c in df.columns:
        cs = str(c).strip()
        if cs.isdigit() and year_min <= int(cs) <= year_max:
            year_cols.append(c)
    if not year_cols:
        return df  # nothing to unpivot
    if id_cols is None:
        id_cols = [c for c in df.columns if c not in year_cols]
    melted = df.melt(id_vars=id_cols, value_vars=year_cols,
                     var_name="year", value_name=value_name)
    melted["year"] = pd.to_numeric(melted["year"], errors="coerce").astype("Int64")
    return melted


def read_table(path: Any, *, sheet: Optional[str] = None) -> pd.DataFrame:
    """Read a tabular file into a DataFrame. Supported formats:
       .csv / .tsv / .csv.gz / .xlsx / .xls / .xlsm / .parquet
    For Excel, `sheet` picks which tab (default = first sheet)."""
    p = Path(path)
    name = p.name.lower()
    suf = p.suffix.lower()
    if suf in (".xlsx", ".xls", ".xlsm"):
        return pd.read_excel(p, sheet_name=sheet if sheet else 0)
    if suf == ".tsv":
        return pd.read_csv(p, sep="\t")
    if suf == ".parquet":
        return pd.read_parquet(p)
    if name.endswith(".csv.gz") or (suf == ".gz" and p.stem.lower().endswith(".csv")):
        return pd.read_csv(p, compression="infer")
    return pd.read_csv(p)


def list_sqlite_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """All SQLite database files under the given folder(s), deduped."""
    folders = normalize_data_folders(data_folder)
    out: List[Path] = []
    patterns = ("*.db", "*.sqlite", "*.sqlite3")
    for folder in folders:
        for pat in patterns:
            if recursive:
                out.extend(sorted(folder.rglob(pat)))
            else:
                out.extend(sorted(folder.glob(pat)))
    deduped: dict[str, Path] = {}
    for path in out:
        try:
            _rp = path.resolve()          # resolve once (was twice per path)
            deduped[str(_rp).lower()] = _rp
        except Exception:
            pass
    return _drop_protected(sorted(deduped.values()), folders)


def list_sqlite_tables(path: Any) -> List[str]:
    """Return the list of table names in a SQLite database (read-only)."""
    import sqlite3
    con = sqlite3.connect(f"file:{Path(path)}?mode=ro", uri=True)
    try:
        cur = con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        )
        return [r[0] for r in cur.fetchall() if r and r[0]]
    finally:
        con.close()


def read_sqlite_table(path: Any, table: str, *, limit: Optional[int] = None) -> pd.DataFrame:
    """Read a SQLite table into a DataFrame. Read-only connection so the
    analyst sandbox can never write back."""
    import sqlite3
    qname = '"' + str(table).replace('"', '""') + '"'
    sql = f"SELECT * FROM {qname}"
    if limit:
        sql += f" LIMIT {int(limit)}"
    con = sqlite3.connect(f"file:{Path(path)}?mode=ro", uri=True)
    try:
        return pd.read_sql_query(sql, con)
    finally:
        con.close()


def list_parquet_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """All Parquet files under the given folder(s), deduped."""
    folders = normalize_data_folders(data_folder)
    out: List[Path] = []
    for folder in folders:
        if recursive:
            out.extend(sorted(folder.rglob("*.parquet")))
        else:
            out.extend(sorted(folder.glob("*.parquet")))
    return _drop_protected(sorted({p.resolve() for p in out}), folders)


def list_duckdb_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """All DuckDB databases under the given folder(s), deduped."""
    folders = normalize_data_folders(data_folder)
    out: List[Path] = []
    for folder in folders:
        if recursive:
            out.extend(sorted(folder.rglob("*.duckdb")))
        else:
            out.extend(sorted(folder.glob("*.duckdb")))
    return _drop_protected(sorted({p.resolve() for p in out}), folders)


def list_bson_files(data_folder: Any, recursive: bool = True) -> List[Path]:
    """All BSON (MongoDB) files under the given folder(s), deduped."""
    folders = normalize_data_folders(data_folder)
    out: List[Path] = []
    for folder in folders:
        if recursive:
            out.extend(sorted(folder.rglob("*.bson")))
        else:
            out.extend(sorted(folder.glob("*.bson")))
    return _drop_protected(sorted({p.resolve() for p in out}), folders)


# ============================================================
# DuckDB helpers
# ============================================================

def _import_duckdb():
    try:
        import duckdb
        return duckdb
    except ImportError as exc:
        raise RuntimeError(
            "DuckDB support needs the `duckdb` package. Install with: "
            f"`pip install duckdb` (original: {exc})"
        ) from exc


def _import_sqlalchemy():
    try:
        import sqlalchemy
        return sqlalchemy
    except ImportError as exc:
        raise RuntimeError(
            "Remote SQL connections need SQLAlchemy. Install with: "
            f"`pip install sqlalchemy` (original: {exc}). For specific "
            "databases also install the matching driver: psycopg2-binary "
            "(Postgres), PyMySQL (MySQL), or pyodbc (MSSQL)."
        ) from exc


def list_duckdb_tables(path: Any) -> List[str]:
    """Return all table names in a DuckDB database (read-only)."""
    duckdb = _import_duckdb()
    con = duckdb.connect(str(Path(path)), read_only=True)
    try:
        return [r[0] for r in con.execute(
            "SELECT table_name FROM duckdb_tables() ORDER BY table_name"
        ).fetchall() if r and r[0]]
    finally:
        con.close()


def read_duckdb_table(path: Any, table: str, *, limit: Optional[int] = None) -> pd.DataFrame:
    """Read a DuckDB table as a DataFrame (read-only)."""
    duckdb = _import_duckdb()
    qname = '"' + str(table).replace('"', '""') + '"'
    sql = f"SELECT * FROM {qname}"
    if limit:
        sql += f" LIMIT {int(limit)}"
    con = duckdb.connect(str(Path(path)), read_only=True)
    try:
        return con.execute(sql).fetch_df()
    finally:
        con.close()


def duckdb_query(path: Any, sql: str) -> pd.DataFrame:
    """Run an arbitrary read-only SQL query against a DuckDB file.

    DuckDB shines here — it can SELECT directly from CSV/Parquet/JSON
    files with `read_csv('path')` / `read_parquet('path')` inside the
    SQL, so this single helper unlocks SQL queries over your vault.
    """
    duckdb = _import_duckdb()
    con = duckdb.connect(str(Path(path)), read_only=True)
    try:
        return con.execute(sql).fetch_df()
    finally:
        con.close()


# ============================================================
# BSON helpers (MongoDB)
# ============================================================

def read_bson_documents(path: Any) -> List[Dict[str, Any]]:
    """Decode a .bson file into a list of documents (dicts).

    Raises RuntimeError with an install hint when pymongo (the source
    of the `bson` module) isn't available, so users don't get a
    misleading 'No module named bson' that they might misread as a
    file-corruption error.
    """
    try:
        import bson
    except ImportError as exc:
        raise RuntimeError(
            "Reading .bson files needs pymongo (provides the `bson` "
            "module). Install with: `pip install pymongo` "
            f"(original: {exc})"
        ) from exc
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"BSON file not found: {p}")
    with open(p, "rb") as fh:
        data = fh.read()
    return list(bson.decode_all(data))


def read_bson_as_df(path: Any) -> pd.DataFrame:
    """Decode a .bson file into a flat DataFrame. Nested fields stay
    as their native Python types in each column."""
    docs = read_bson_documents(path)
    return pd.json_normalize(docs) if docs else pd.DataFrame()


# ============================================================
# MongoDB -> model-digestible conversion
# ============================================================
# Mongo BSON / JSON is nested and full of ObjectId / datetime / arrays of
# sub-documents that a language model reads poorly. The helpers below flatten
# + coerce documents into clean scalars (see mongo_normalize.py). They accept
# a .bson file, a Mongo-export .json / .jsonl file, OR a list of dicts you
# already have (e.g. from read_mongo_collection), so the same clean view is
# available whatever the source.

def read_json_documents(path: Any) -> List[Dict[str, Any]]:
    """Read a JSON / JSONL / NDJSON file into a list of documents.

    Handles the three shapes a Mongo export takes:
      * a JSON array of objects                ``[ {...}, {...} ]``
      * one JSON object                        ``{...}``  (-> single-doc list)
      * JSON-Lines / NDJSON (one object/line)  ``{...}\\n{...}\\n``
    MongoDB *Extended JSON* type wrappers (``$oid`` / ``$date`` / …) are left
    intact here and coerced later by the normaliser, so no bson is needed.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"JSON file not found: {p}")
    text = p.read_text(encoding="utf-8", errors="replace").strip()
    if not text:
        return []
    # Try a single JSON value (array or object) first.
    try:
        obj = json.loads(text)
        if isinstance(obj, list):
            return [d for d in obj if isinstance(d, dict)]
        if isinstance(obj, dict):
            return [obj]
    except json.JSONDecodeError:
        pass
    # Fall back to JSON-Lines: one document per non-blank line.
    docs: List[Dict[str, Any]] = []
    for line in text.splitlines():
        line = line.strip().rstrip(",")
        if not line or line in ("[", "]"):
            continue
        try:
            d = json.loads(line)
            if isinstance(d, dict):
                docs.append(d)
        except json.JSONDecodeError:
            continue
    return docs


def mongo_documents_to_frame(docs: Any, *,
                             max_array_items: int = 25,
                             max_array_chars: int = 300) -> pd.DataFrame:
    """Flatten + coerce a list of Mongo documents into a clean, all-scalar
    DataFrame: dotted columns for nested fields, ObjectId/dates as strings,
    arrays collapsed to bounded cells. Feed this to the model instead of
    raw documents."""
    import mongo_normalize as _mn
    return _mn.documents_to_frame(
        docs, max_array_items=max_array_items, max_array_chars=max_array_chars)


def bson_to_clean_frame(path: Any, **kw) -> pd.DataFrame:
    """Read a .bson file straight into the model-digestible DataFrame."""
    return mongo_documents_to_frame(read_bson_documents(path), **kw)


def json_to_clean_frame(path: Any, **kw) -> pd.DataFrame:
    """Read a Mongo-export .json / .jsonl file into the model-digestible
    DataFrame (Extended-JSON ``$oid`` / ``$date`` wrappers coerced)."""
    return mongo_documents_to_frame(read_json_documents(path), **kw)


def mongo_schema_profile(docs: Any, *, sample: int = 1000) -> pd.DataFrame:
    """One row per field after flattening: field, types seen, how many docs
    carry it, presence %, and an example value. The most token-cheap way to
    show a model a collection's structure before it queries the data."""
    import mongo_normalize as _mn
    import pandas as _pd
    if isinstance(docs, (str, Path)):
        p = Path(docs)
        docs = (read_bson_documents(p) if p.suffix.lower() == ".bson"
                else read_json_documents(p))
    return _pd.DataFrame(_mn.infer_schema(docs, sample=sample))


def mongo_documents_to_text(docs: Any, *, max_docs: int = 50,
                            max_value_chars: int = 160,
                            include_schema: bool = True) -> str:
    """Render documents as a compact, flat ``key: value`` text block (with an
    optional schema header) for direct insertion into a model prompt."""
    import mongo_normalize as _mn
    if isinstance(docs, (str, Path)):
        p = Path(docs)
        docs = (read_bson_documents(p) if p.suffix.lower() == ".bson"
                else read_json_documents(p))
    return _mn.documents_to_text(
        docs, max_docs=max_docs, max_value_chars=max_value_chars,
        include_schema=include_schema)


def _mongo_json_byte_budget() -> int:
    """Max bytes a single JSON ARRAY file may be before conversion refuses
    (it must be loaded whole). Streaming formats (.bson / .jsonl) are
    unaffected. Override with COUNCIL_MONGO_MAX_JSON_MB."""
    ov = os.environ.get("COUNCIL_MONGO_MAX_JSON_MB", "").strip()
    if ov:
        try:
            return max(8, int(ov)) * 1024 * 1024
        except ValueError:
            pass
    cap = 256 * 1024 * 1024
    try:
        import psutil
        avail = int(psutil.virtual_memory().available * 0.20)
        if avail > 0:
            return min(cap, avail)
    except Exception:
        pass
    return cap


def convert_mongo_file(src: Any, out_dir: Any, *,
                       want_csv: bool = True, want_schema: bool = True,
                       want_text: bool = False,
                       max_docs: Optional[int] = None) -> Dict[str, Any]:
    """Stream-convert ONE .bson/.json/.jsonl file into model-digestible
    files under ``out_dir`` with BOUNDED memory (safe on a huge dump). See
    mongo_normalize.stream_convert_file. Returns a summary dict."""
    import mongo_normalize as _mn
    return _mn.stream_convert_file(
        src, out_dir, want_csv=want_csv, want_schema=want_schema,
        want_text=want_text, max_docs=max_docs,
        max_json_bytes=_mongo_json_byte_budget())


def mongo_explode_array(docs: Any, record_path: str, *,
                        meta: Optional[List[str]] = None) -> pd.DataFrame:
    """One row per element of the array at ``record_path`` (dotted), carrying
    chosen top-level ``meta`` fields down — the tidy/tabular view of an
    array-of-subdocuments field (e.g. order ``line_items``)."""
    import mongo_normalize as _mn
    if isinstance(docs, (str, Path)):
        p = Path(docs)
        docs = (read_bson_documents(p) if p.suffix.lower() == ".bson"
                else read_json_documents(p))
    return _mn.explode_documents(docs, record_path, meta=meta)


# ============================================================
# SQLAlchemy bridge — remote SQL databases (Postgres / MySQL / MSSQL / etc.)
# ============================================================

# Connection registry persists in vault/sql_connections.json. Plain JSON
# so the user can edit it by hand. NEVER stores raw passwords — users
# put credentials in env vars and reference them like
# `postgresql://user:${PGPASS}@host/db`.
#
# Storage + URL resolution + audit log all moved to db_connections.py
# so the SQL and Mongo bridges share the same registry / log / env-var
# expander. The functions below are thin re-exports kept here so any
# code that previously imported list_sql_connections / save_sql_connection
# from vault_analyst keeps working — no behaviour change.


def list_sql_connections(vault_dir: Any) -> Dict[str, str]:
    """Return saved ``{name: url}``. URLs still contain ``${ENV_VAR}``
    placeholders; resolution happens at connect time."""
    import db_connections as _db
    return _db.list_sql_connections(vault_dir)


def save_sql_connection(vault_dir: Any, name: str, url: str) -> None:
    """Save a SQLAlchemy connection URL. Use ``${ENV_VAR}`` placeholders
    for passwords — they expand at connect time."""
    import db_connections as _db
    _db.save_sql_connection(vault_dir, name, url)


def remove_sql_connection(vault_dir: Any, name: str) -> bool:
    """Drop a saved SQL connection. Returns True if it existed."""
    import db_connections as _db
    return _db.remove_sql_connection(vault_dir, name)


def save_mongo_connection(vault_dir: Any, name: str, uri: str) -> None:
    """Save a MongoDB URI (``mongodb://...``). Use ``${ENV_VAR}``
    placeholders for passwords."""
    import db_connections as _db
    _db.save_mongo_connection(vault_dir, name, uri)


def remove_mongo_connection(vault_dir: Any, name: str) -> bool:
    """Drop a saved Mongo connection. Returns True if it existed."""
    import db_connections as _db
    return _db.remove_mongo_connection(vault_dir, name)


def list_sql_tables(vault_dir: Any, conn_name: str) -> List[str]:
    """Inspect a named connection and return its table names.
    Read-only — uses db_connections layer with full audit trail."""
    import db_connections as _db
    return _db.list_sql_tables(vault_dir, conn_name)


def read_sql_table(
    vault_dir: Any, conn_name: str, table: str,
    *, limit: Optional[int] = 10000,
) -> pd.DataFrame:
    """Pull a remote SQL table into a DataFrame using a saved
    connection. Read-only by construction (SELECT only). Default
    10K-row limit; pass ``limit=None`` to lift it."""
    import db_connections as _db
    return _db.read_sql_table(vault_dir, conn_name, table, limit=limit)


def sql_query(vault_dir: Any, conn_name: str, sql: str) -> pd.DataFrame:
    """Run an arbitrary SQL query through a saved connection. The
    query is VALIDATED to be a single read-only statement before
    dispatch — DDL / DML keywords (INSERT, UPDATE, DELETE, DROP,
    TRUNCATE, ALTER, CREATE, GRANT, REVOKE, MERGE, …) are rejected
    even with comment-cloaking, and multi-statement payloads are
    rejected too. See db_connections._validate_select_only."""
    import db_connections as _db
    return _db.sql_query(vault_dir, conn_name, sql)


# ── MongoDB — read-only by API design ──────────────────────────────
# Same convention as the SQL bridge: URLs in vault/mongo_connections.json,
# ${ENV_VAR} placeholders supported. The wrappers below expose ONLY
# find / aggregate / count / distinct / list_* — no insert / update /
# delete / drop methods are reachable from the analyst sandbox. The
# aggregation validator also rejects $out, $merge, $function,
# $accumulator, $where which can bypass a read-only role.

def list_mongo_connections(vault_dir: Any) -> Dict[str, str]:
    """Return saved ``{name: mongodb_uri}``."""
    import db_connections as _db
    return _db.list_mongo_connections(vault_dir)


def list_mongo_databases(vault_dir: Any, conn_name: str) -> List[str]:
    """List databases visible to the saved Mongo connection."""
    import db_connections as _db
    return _db.list_mongo_databases(vault_dir, conn_name)


def list_mongo_collections(
    vault_dir: Any, conn_name: str, db_name: str,
) -> List[str]:
    """List collections in a database."""
    import db_connections as _db
    return _db.list_mongo_collections(vault_dir, conn_name, db_name)


def read_mongo_collection(
    vault_dir: Any,
    conn_name: str,
    db_name: str,
    collection: str,
    *,
    query: Optional[Dict[str, Any]] = None,
    projection: Optional[Dict[str, Any]] = None,
    limit: Optional[int] = 10000,
    skip: int = 0,
    sort: Optional[List] = None,
) -> pd.DataFrame:
    """Mongo find() → DataFrame. Default 10K-row hard limit; pass
    limit=None to lift (audit-logged with a WARN tag)."""
    import db_connections as _db
    return _db.read_mongo_collection(
        vault_dir, conn_name, db_name, collection,
        query=query, projection=projection,
        limit=limit, skip=skip, sort=sort,
    )


def mongo_aggregate(
    vault_dir: Any,
    conn_name: str,
    db_name: str,
    collection: str,
    pipeline: List[Dict[str, Any]],
    *,
    allow_disk_use: bool = False,
) -> pd.DataFrame:
    """Run a Mongo aggregation pipeline. Pipeline is validated —
    $out / $merge / $function / $accumulator / $where stages
    are rejected."""
    import db_connections as _db
    return _db.mongo_aggregate(
        vault_dir, conn_name, db_name, collection, pipeline,
        allow_disk_use=allow_disk_use,
    )


def mongo_count(
    vault_dir: Any, conn_name: str, db_name: str, collection: str,
    *, query: Optional[Dict[str, Any]] = None,
) -> int:
    """count_documents on a collection."""
    import db_connections as _db
    return _db.mongo_count(vault_dir, conn_name, db_name, collection,
                            query=query)


def mongo_distinct(
    vault_dir: Any, conn_name: str, db_name: str, collection: str,
    field: str, *, query: Optional[Dict[str, Any]] = None,
) -> List[Any]:
    """distinct values of `field` in a collection."""
    import db_connections as _db
    return _db.mongo_distinct(vault_dir, conn_name, db_name, collection,
                               field, query=query)


def read_excel_sheets(path: Any) -> Dict[str, pd.DataFrame]:
    """Read every sheet of an Excel workbook into a dict {sheet_name: df}."""
    p = Path(path)
    return pd.read_excel(p, sheet_name=None)


def detect_excel_header_rows(path: Any, sheet: Any = None) -> int:
    """Auto-detect how many header rows an Excel sheet has by inspecting
    merged-cell ranges via openpyxl. Returns 1 for plain CSV-style
    headers, 2 when row 1 has merged groups + row 2 has sub-columns,
    3 for two levels of grouping.

    `sheet` accepts a sheet NAME (str) or 0-based INDEX (int). When None
    or out of range, falls back to the first sheet.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(Path(path), read_only=False, data_only=True)
    except Exception:
        return 1
    try:
        ws = None
        if isinstance(sheet, int) and 0 <= sheet < len(wb.worksheets):
            ws = wb.worksheets[sheet]
        elif isinstance(sheet, str) and sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]
        merged_rows: set = set()
        for rng in ws.merged_cells.ranges:
            if rng.max_col - rng.min_col >= 1:
                for r in range(rng.min_row, min(rng.max_row, 3) + 1):
                    merged_rows.add(r)
        if not merged_rows:
            return 1
        return max(2, min(max(merged_rows) + 1, 3))
    finally:
        try: wb.close()
        except Exception: pass


def read_excel_with_merged_headers(
    path: Any,
    sheet: Any = 0,
    *,
    header_rows: Optional[int] = None,
    flatten: bool = True,
    sep: str = " / ",
) -> pd.DataFrame:
    """Read an Excel sheet where the top row(s) are merged GROUP headers
    above the real column-name row.

    Without this helper, `pd.read_excel(path)` collapses such files
    into useless 'Unnamed: 0', 'Unnamed: 1' columns. With it, each
    output column is named like 'Site A / energy' so the model can
    find "the highest value in each energy column" by filtering on
    the suffix.

    Args:
      path:         workbook path
      sheet:        sheet name or 0-based index (default: first sheet)
      header_rows:  number of header rows to consume; auto-detected
                    via openpyxl merged-cell ranges when None
      flatten:      if True (default), MultiIndex columns are joined
                    with `sep` into single strings
      sep:          separator for flattened multi-headers
    """
    p = Path(path)
    if header_rows is None:
        # Pass the sheet identifier through verbatim — detect_excel_header_rows
        # accepts both string names and integer indexes.
        header_rows = detect_excel_header_rows(p, sheet)
    if header_rows <= 1:
        return pd.read_excel(p, sheet_name=sheet)
    df = pd.read_excel(p, sheet_name=sheet, header=list(range(header_rows)))
    if not flatten or not isinstance(df.columns, pd.MultiIndex):
        return df
    new_cols = []
    for tup in df.columns:
        parts = [str(x) for x in tup
                 if x is not None and not str(x).startswith("Unnamed:")]
        new_cols.append(sep.join(parts) if parts else "(unnamed)")
    df.columns = new_cols
    return df


def excel_inventory(
    data_folder: Any,
    recursive: bool = True,
    max_files: Optional[int] = None,
) -> pd.DataFrame:
    """One row per (workbook, sheet): sheet name, row count, column count,
    column-list preview. Mirrors csv_inventory for Excel."""
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    xls = list_excel_files(folders, recursive=recursive)
    if max_files is not None:
        xls = xls[:max_files]
    for p in xls:
        root = first_matching_root(p, folders)
        try:
            xl = pd.ExcelFile(str(p))
        except Exception as exc:
            rows.append({
                "workbook":      p.name,
                "relative_path": safe_relative_path(p, root),
                "sheet":         "",
                "status":        f"open error: {exc}",
                "columns":       "",
                "column_count":  None,
                "row_count":     None,
            })
            continue
        for sname in xl.sheet_names:
            try:
                df = xl.parse(sname, nrows=5)
                rows.append({
                    "workbook":      p.name,
                    "relative_path": safe_relative_path(p, root),
                    "sheet":         sname,
                    "status":        "ok",
                    "columns":       ", ".join(map(str, df.columns)),
                    "column_count":  len(df.columns),
                    "row_count":     None,  # nrows=5 only; full count via separate helper
                })
            except Exception as exc:
                rows.append({
                    "workbook":      p.name,
                    "relative_path": safe_relative_path(p, root),
                    "sheet":         sname,
                    "status":        f"sheet read error: {exc}",
                    "columns":       "",
                    "column_count":  None,
                    "row_count":     None,
                })
    return pd.DataFrame(rows)


# ============================================================
# Column resolution (case- and substring-tolerant)
# ============================================================

def find_column_case_insensitive(df: pd.DataFrame, column_name: str) -> Optional[str]:
    target = str(column_name).strip().lower()
    for col in df.columns:
        if str(col).strip().lower() == target:
            return col
    return None


def find_columns_contains(df: pd.DataFrame, text: str) -> List[str]:
    target = str(text).strip().lower()
    return [col for col in df.columns if target in str(col).strip().lower()]


# ============================================================
# Cache-backed column stats (precomputed via stats_cache)
# ============================================================

def cached_column_stats(vault_dir: Any, path: Any) -> Dict[str, Any]:
    """Return precomputed column stats for one CSV from the stats cache,
    computing + storing them on a miss. Fast on repeat — the whole point
    of the cache. Returns {} if stats_cache is unavailable."""
    try:
        import stats_cache
        return stats_cache.StatsCache(vault_dir).get(path) or {}
    except Exception:
        return {}


def folder_column_stats(vault_dir: Any, data_folder: Any,
                        recursive: bool = True, *,
                        csv_files: Optional[List[Path]] = None
                        ) -> "pd.DataFrame":
    """One row per (file, column) with the precomputed stats — count,
    missing, min, max, mean, std, sum (numeric) / n_unique, top (text) —
    served from the cache (computed + stored on first sight of a file).
    Use this instead of re-reading files when a question asks for column
    statistics. The ``notes`` column flags files that already carry their
    own summary (a Total/Mean row or a stat-named column).

    ``csv_files``: an already-materialised CSV list (e.g. computed by the
    caller for a cache key). When given it's used verbatim instead of walking
    the folder again; None (default) preserves the original walk."""
    rows: List[Dict[str, Any]] = []
    try:
        import stats_cache
        cache = stats_cache.StatsCache(vault_dir)
    except Exception:
        return pd.DataFrame(rows)
    folders = normalize_data_folders(data_folder)
    _files = (csv_files if csv_files is not None
              else list_csv_files(folders, recursive=recursive))
    for fp in _files:
        st = cache.get(fp) or {}
        root = first_matching_root(fp, folders)
        rel = safe_relative_path(fp, root)
        sd = st.get("self_describing") or {}
        note = ("file carries its own summary"
                if (sd.get("summary_columns") or sd.get("summary_rows"))
                else "")
        cs = st.get("column_stats") or {}
        if not cs:
            rows.append({"file": fp.name, "relative_path": rel,
                         "column": None, "notes": note or "no stats"})
            continue
        for col, s in cs.items():
            rows.append({
                "file": fp.name, "relative_path": rel, "column": col,
                "dtype": s.get("dtype"), "count": s.get("count"),
                "missing": s.get("missing"), "min": s.get("min"),
                "max": s.get("max"), "mean": s.get("mean"),
                "std": s.get("std"), "sum": s.get("sum"),
                "n_unique": s.get("n_unique"), "top": s.get("top"),
                "notes": note,
            })
    return pd.DataFrame(rows)


# ============================================================
# Per-CSV aggregate helpers
# Each returns a DataFrame with one row per input CSV.
# ============================================================

def folder_data_summary(
    data_folder: Any,
    recursive: bool = True,
    max_files: Optional[int] = None,
    include_image_metadata: bool = True,
    *,
    csv_files: Optional[List[Path]] = None,
) -> pd.DataFrame:
    """A "true data summary" of every data file in `data_folder`.

    Returns ONE row per file with the columns the Council's data
    analytics queries actually want to see — without making the model
    write per-file code for every question:

        file              file basename
        relative_path     path relative to its search root
        type              csv / tsv / parquet / xlsx / json / sqlite /
                          duckdb / bson / image / text / source / unknown
        size_kb           file size in KB
        rows              row count (for tabular formats; null otherwise)
        columns           column count
        column_names      comma-joined first 12 column names
        dtypes            comma-joined dtype shorthand (int, float, str, ...)
        missing_pct       overall missing-value rate as a percentage (0-100)
        numeric_cols      count of numeric columns
        date_cols         count of date / datetime columns
        sample_value      one representative value from the first non-empty
                          column (useful for ID-shape sniffing)
        notes             short freeform diagnostic (errors, warnings)

    Designed for the Council intent "give me a true data summary of
    files in the <subfolder> folder". The model's analyst step can
    call it with a single line:

        result_df = folder_data_summary(DATA_FOLDERS)

    and the resulting DataFrame is the answer.
    """
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []

    # Build a deduped file list across the supported tabular + structured
    # formats. Image files are reported but not deeply parsed.
    file_paths: list[Path] = []
    seen: set = set()
    # Reuse a caller-supplied CSV list (e.g. already walked for a cache key)
    # instead of walking for CSVs a second time; None preserves the walk.
    _csv_collector = ((lambda: csv_files) if csv_files is not None
                      else (lambda: list_csv_files(folders, recursive=recursive)))
    for collector in (
        _csv_collector,
        lambda: list_excel_files(folders, recursive=recursive),
        lambda: list_parquet_files(folders, recursive=recursive),
        lambda: list_sqlite_files(folders, recursive=recursive),
        lambda: list_duckdb_files(folders, recursive=recursive),
        lambda: list_bson_files(folders, recursive=recursive),
    ):
        try:
            for p in collector():
                key = str(p.resolve()).lower()
                if key not in seen:
                    seen.add(key)
                    file_paths.append(p)
        except Exception:
            continue

    # Also pick up JSON, plain-text, and image files via folder walks —
    # these aren't in the collector helpers but are part of "what's in
    # this folder" from the user's perspective.
    _EXTRA_EXT = {".json", ".jsonl", ".ndjson", ".txt", ".md",
                  ".yaml", ".yml", ".xml",
                  ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp",
                  ".tiff", ".tif"}
    for folder in folders:
        try:
            walker = folder.rglob("*") if recursive else folder.glob("*")
            for p in walker:
                if not p.is_file():
                    continue
                if p.suffix.lower() not in _EXTRA_EXT:
                    continue
                key = str(p.resolve()).lower()
                if key not in seen:
                    seen.add(key)
                    file_paths.append(p)
        except Exception:
            continue

    if max_files is not None:
        file_paths = file_paths[:max_files]

    for fp in file_paths:
        root = first_matching_root(fp, folders)
        rec: dict[str, Any] = {
            "file":          fp.name,
            "relative_path": safe_relative_path(fp, root),
            "type":          "unknown",
            "size_kb":       0.0,
            "rows":          None,
            "columns":       None,
            "column_names":  "",
            "dtypes":        "",
            "missing_pct":   None,
            "numeric_cols":  None,
            "date_cols":     None,
            "sample_value":  "",
            "notes":         "",
        }
        try:
            rec["size_kb"] = round(fp.stat().st_size / 1024, 1)
        except Exception:
            pass

        suf = fp.suffix.lower()
        try:
            if suf in (".csv", ".tsv"):
                # Read only a head-sample for the column stats (keeps memory
                # flat across hundreds of files); count the exact rows with
                # a cheap byte scan. Reading every file in full was the
                # cause of the OOM/slow crash on 200+ CSV summaries.
                sep = "\t" if suf == ".tsv" else ","
                sample = _read_csv_cached(fp, sep=sep,
                                          nrows=_SUMMARY_SAMPLE_ROWS)
                exact_rows = _count_csv_rows_fast(fp)
                _fill_tabular_summary(
                    rec, sample, "csv" if suf == ".csv" else "tsv",
                    exact_rows=exact_rows,
                    sampled=(exact_rows < 0 or exact_rows > len(sample)),
                )
            elif suf == ".parquet":
                df = pd.read_parquet(fp)
                _fill_tabular_summary(rec, df, "parquet")
            elif suf in (".xlsx", ".xls", ".xlsm"):
                # Use the first sheet for the headline summary; multi-sheet
                # files are noted in `notes`.
                xl = pd.ExcelFile(fp)
                first_sheet = xl.sheet_names[0] if xl.sheet_names else None
                if first_sheet is None:
                    rec["type"] = "excel"
                    rec["notes"] = "no sheets"
                else:
                    df = pd.read_excel(fp, sheet_name=first_sheet)
                    _fill_tabular_summary(rec, df, "excel")
                    if len(xl.sheet_names) > 1:
                        rec["notes"] = (f"first sheet '{first_sheet}'; "
                                        f"file has {len(xl.sheet_names)} sheets")
            elif suf in (".json", ".jsonl", ".ndjson"):
                # JSONL/NDJSON load as records; plain JSON we tabularise
                # via json_normalize if it's a list of dicts.
                rec["type"] = "json" if suf == ".json" else "jsonl"
                try:
                    if suf == ".json":
                        import json as _json
                        data = _json.loads(fp.read_text(encoding="utf-8",
                                                         errors="replace"))
                        if isinstance(data, list) and data and isinstance(data[0], dict):
                            df = pd.json_normalize(data)
                            _fill_tabular_summary(rec, df, "json")
                        elif isinstance(data, dict):
                            rec["columns"] = len(data)
                            rec["column_names"] = ", ".join(
                                list(data.keys())[:12])
                            rec["notes"] = "single object (not a list of records)"
                    else:
                        df = pd.read_json(fp, lines=True)
                        _fill_tabular_summary(rec, df, "jsonl")
                except Exception as exc:
                    rec["notes"] = f"json parse failed: {exc}"
            elif suf in (".db", ".sqlite", ".sqlite3"):
                rec["type"] = "sqlite"
                try:
                    tables = list_sqlite_tables(fp)
                    rec["columns"] = len(tables)
                    rec["column_names"] = ", ".join(tables[:12])
                    rec["notes"] = f"{len(tables)} table(s)"
                except Exception as exc:
                    rec["notes"] = f"sqlite read failed: {exc}"
            elif suf == ".duckdb":
                rec["type"] = "duckdb"
                try:
                    tables = list_duckdb_tables(fp)
                    rec["columns"] = len(tables)
                    rec["column_names"] = ", ".join(tables[:12])
                    rec["notes"] = f"{len(tables)} table(s)"
                except Exception as exc:
                    rec["notes"] = f"duckdb read failed: {exc}"
            elif suf == ".bson":
                rec["type"] = "bson"
                try:
                    docs = read_bson_documents(fp)
                    rec["rows"] = len(docs)
                    if docs:
                        keys = list(docs[0].keys()) if isinstance(docs[0], dict) else []
                        rec["columns"] = len(keys)
                        rec["column_names"] = ", ".join(map(str, keys[:12]))
                except Exception as exc:
                    rec["notes"] = f"bson read failed: {exc}"
            elif suf in (".png", ".jpg", ".jpeg", ".gif", ".bmp",
                          ".webp", ".tiff", ".tif"):
                rec["type"] = "image"
                if include_image_metadata:
                    try:
                        from PIL import Image  # type: ignore[import]
                        with Image.open(fp) as img:
                            rec["columns"] = img.width
                            rec["rows"]    = img.height
                            rec["column_names"] = f"mode={img.mode}"
                            rec["notes"]   = f"{img.width}×{img.height} {img.format or ''}"
                    except Exception as exc:
                        rec["notes"] = f"image (PIL unavailable: {exc})"
                else:
                    rec["notes"] = "image"
            elif suf in (".txt", ".md", ".yaml", ".yml", ".xml"):
                rec["type"] = "text"
                try:
                    text = fp.read_text(encoding="utf-8", errors="replace")
                    rec["rows"] = text.count("\n")
                    sample = text[:80].replace("\n", " ").strip()
                    rec["sample_value"] = sample
                except Exception as exc:
                    rec["notes"] = f"text read failed: {exc}"
            else:
                rec["type"] = "other"
        except Exception as exc:
            rec["notes"] = f"profile error: {exc}"
        rows.append(rec)

    return pd.DataFrame(rows)


# Rows to sample when summarising a file. A summary needs column names,
# dtypes, a missing-value feel and a sample value — none of which require
# the WHOLE file. Reading only the head keeps memory flat and time low so
# folder_data_summary scales to hundreds of files without OOM.
_SUMMARY_SAMPLE_ROWS = 2000


def _count_csv_rows_fast(path: Any) -> int:
    """Exact data-row count for a CSV/TSV via a buffered byte scan — no
    pandas parse, bounded memory. Returns newline-count minus the header
    (floored at 0), or -1 if it can't be read. Treats a missing trailing
    newline as a final row. (A file with quoted embedded newlines would
    over-count slightly; that's an acceptable approximation for a
    hundreds-of-files overview and far cheaper than a full parse.)"""
    try:
        n = 0
        last = b"\n"
        with open(path, "rb") as fh:
            while True:
                chunk = fh.read(1024 * 1024)
                if not chunk:
                    break
                n += chunk.count(b"\n")
                last = chunk[-1:]
        if last != b"\n":          # file didn't end in a newline
            n += 1
        return max(0, n - 1)       # drop the header line
    except Exception:
        return -1


def _fill_tabular_summary(rec: dict, df: "pd.DataFrame", type_label: str,
                          *, exact_rows: "Optional[int]" = None,
                          sampled: bool = False) -> None:
    """Populate the tabular-summary fields on ``rec`` from ``df``.

    When ``df`` is only a head-sample of a larger file, pass
    ``exact_rows`` (the true row count, counted cheaply elsewhere) and
    ``sampled=True``; the row count is then reported exactly while the
    column stats are derived from the sample and flagged in ``notes``."""
    try:
        rec["type"]         = type_label
        rec["rows"]         = int(exact_rows) if (exact_rows is not None and exact_rows >= 0) else int(len(df))
        rec["columns"]      = int(len(df.columns))
        rec["column_names"] = ", ".join(map(str, df.columns[:12]))
        # Dtype shorthand: keep the names short — int / float / str /
        # datetime / bool / object — readable in a one-row summary.
        def _short_dtype(dt) -> str:
            s = str(dt)
            if s.startswith("int"):
                return "int"
            if s.startswith("float"):
                return "float"
            if "datetime" in s:
                return "datetime"
            if s == "bool":
                return "bool"
            return "str"
        types = [_short_dtype(d) for d in df.dtypes[:12]]
        rec["dtypes"] = ", ".join(types)
        rec["numeric_cols"] = int(
            df.select_dtypes(include=["number"]).shape[1])
        try:
            rec["date_cols"] = int(
                df.select_dtypes(include=["datetime", "datetimetz"]).shape[1])
        except Exception:
            rec["date_cols"] = 0
        # Missing-value rate — total NaNs / total cells, ×100.
        total_cells = max(1, len(df) * len(df.columns))
        rec["missing_pct"] = round(
            float(df.isna().sum().sum()) / total_cells * 100, 1)
        # Representative value — first non-null in the first column.
        if len(df) and len(df.columns):
            try:
                non_null = df[df.columns[0]].dropna()
                if len(non_null):
                    sv = non_null.iloc[0]
                    rec["sample_value"] = str(sv)[:60]
            except Exception:
                pass
        # Flag that column stats came from a head-sample, not the full
        # file, so the row count (exact) and the stats (sampled) aren't
        # silently conflated.
        if sampled:
            note = f"stats sampled from first {len(df):,} rows"
            rec["notes"] = (rec["notes"] + "; " + note) if rec["notes"] else note
    except Exception as exc:
        rec["notes"] = f"summary error: {exc}"


def csv_inventory(
    data_folder: Any,
    recursive: bool = True,
    max_files: Optional[int] = None,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    csvs = list_csv_files(folders, recursive=recursive)
    if max_files is not None:
        csvs = csvs[:max_files]

    for csv_path in csvs:
        root = first_matching_root(csv_path, folders)
        try:
            df_head = _read_csv_cached(csv_path, nrows=5)
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "root_folder": str(root) if root else "",
                "status": "ok",
                "columns": ", ".join(map(str, df_head.columns)),
                "column_count": len(df_head.columns),
            })
        except Exception as exc:
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "root_folder": str(root) if root else "",
                "status": f"error: {exc}",
                "columns": "",
                "column_count": None,
            })
    return pd.DataFrame(rows)


def count_rows_per_csv(data_folder: Any, recursive: bool = True) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": "ok",
                "row_count": len(df),
                "column_count": len(df.columns),
            })
        except Exception as exc:
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": f"error: {exc}",
                "row_count": None,
                "column_count": None,
            })
    return pd.DataFrame(rows)


def _numeric_aggregate_per_csv(
    data_folder: Any,
    column_name: str,
    *,
    op: str,
    exclude_zero: bool,
    recursive: bool,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            actual_col = find_column_case_insensitive(df, column_name)
            if actual_col is None:
                rows.append({
                    "csv": csv_path.name,
                    "relative_path": safe_relative_path(csv_path, root),
                    "status": f"missing column: {column_name}",
                    "column_used": None,
                    "value": None,
                    "rows_used": 0,
                    "total_rows": len(df),
                })
                continue

            values = pd.to_numeric(df[actual_col], errors="coerce")
            mask = values.notna()
            if exclude_zero:
                mask = mask & (values != 0)
            filtered = values[mask]

            if op == "mean":
                value = float(filtered.mean()) if filtered.size else None
            elif op == "std":
                value = float(filtered.std()) if filtered.size >= 2 else None
            elif op == "sum":
                value = float(filtered.sum()) if filtered.size else 0.0
            elif op == "min":
                value = float(filtered.min()) if filtered.size else None
            elif op == "max":
                value = float(filtered.max()) if filtered.size else None
            else:
                value = None

            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": "ok",
                "column_used": actual_col,
                "value": value,
                "rows_used": int(filtered.count()),
                "total_rows": len(df),
            })
        except Exception as exc:
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": f"error: {exc}",
                "column_used": None,
                "value": None,
                "rows_used": 0,
                "total_rows": None,
            })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.rename(columns={"value": op})
    return out


def average_numeric_column_per_csv(data_folder, column_name, *, exclude_zero=False, recursive=True):
    return _numeric_aggregate_per_csv(data_folder, column_name, op="mean",
                                       exclude_zero=exclude_zero, recursive=recursive)


def std_numeric_column_per_csv(data_folder, column_name, *, exclude_zero=False, recursive=True):
    return _numeric_aggregate_per_csv(data_folder, column_name, op="std",
                                       exclude_zero=exclude_zero, recursive=recursive)


def sum_numeric_column_per_csv(data_folder, column_name, *, exclude_zero=False, recursive=True):
    return _numeric_aggregate_per_csv(data_folder, column_name, op="sum",
                                       exclude_zero=exclude_zero, recursive=recursive)


def min_max_numeric_column_per_csv(data_folder, column_name, *, exclude_zero=False, recursive=True):
    """Returns both min and max in the same frame."""
    mn = _numeric_aggregate_per_csv(data_folder, column_name, op="min",
                                     exclude_zero=exclude_zero, recursive=recursive)
    mx = _numeric_aggregate_per_csv(data_folder, column_name, op="max",
                                     exclude_zero=exclude_zero, recursive=recursive)
    if mn.empty:
        return mn
    merged = mn[["csv", "relative_path", "status", "column_used", "rows_used", "total_rows"]].copy()
    merged["min"] = mn["min"]
    merged["max"] = mx["max"]
    return merged


def numeric_summary_per_csv(
    data_folder: Any,
    column_name: Optional[str] = None,
    *,
    exclude_zero: bool = False,
    recursive: bool = True,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            if column_name:
                actual_col = find_column_case_insensitive(df, column_name)
                if actual_col is None:
                    rows.append({"csv": csv_path.name, "status": f"missing column: {column_name}"})
                    continue
                candidate_cols = [actual_col]
            else:
                candidate_cols = list(df.columns)

            for col in candidate_cols:
                values = pd.to_numeric(df[col], errors="coerce")
                mask = values.notna()
                if exclude_zero:
                    mask = mask & (values != 0)
                filtered = values[mask]
                if filtered.empty:
                    continue
                rows.append({
                    "csv": csv_path.name,
                    "relative_path": safe_relative_path(csv_path, root),
                    "column": col,
                    "count": int(filtered.count()),
                    "mean": float(filtered.mean()),
                    "median": float(filtered.median()),
                    "std": float(filtered.std()) if filtered.size >= 2 else None,
                    "min": float(filtered.min()),
                    "max": float(filtered.max()),
                })
        except Exception as exc:
            rows.append({"csv": csv_path.name, "status": f"error: {exc}"})
    return pd.DataFrame(rows)


def count_matching_rows_per_csv(
    data_folder: Any,
    column_name: str,
    *,
    equals: Any = None,
    contains: Optional[str] = None,
    case_sensitive: bool = False,
    recursive: bool = True,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            actual_col = find_column_case_insensitive(df, column_name)
            if actual_col is None:
                # also try substring match against header names
                similar = find_columns_contains(df, column_name)
                if similar:
                    actual_col = similar[0]
            if actual_col is None:
                rows.append({
                    "csv": csv_path.name,
                    "relative_path": safe_relative_path(csv_path, root),
                    "status": f"missing column: {column_name}",
                    "match_count": None,
                    "total_rows": len(df),
                })
                continue

            series = df[actual_col]
            if equals is not None:
                if case_sensitive:
                    mask = series.astype(str) == str(equals)
                else:
                    mask = series.astype(str).str.lower() == str(equals).lower()
            elif contains is not None:
                mask = series.astype(str).str.contains(
                    str(contains), case=case_sensitive, na=False, regex=False,
                )
            else:
                mask = pd.Series([False] * len(df), index=df.index)

            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": "ok",
                "column_used": actual_col,
                "match_count": int(mask.sum()),
                "total_rows": len(df),
                "match_pct": round(100 * mask.sum() / max(1, len(df)), 2),
            })
        except Exception as exc:
            rows.append({
                "csv": csv_path.name,
                "status": f"error: {exc}",
                "match_count": None,
                "total_rows": None,
            })
    return pd.DataFrame(rows)


def filter_rows_across_csvs(
    data_folder: Any,
    column_name: str,
    *,
    equals: Any = None,
    contains: Optional[str] = None,
    case_sensitive: bool = False,
    recursive: bool = True,
    max_rows_per_csv: Optional[int] = None,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    frames: list[pd.DataFrame] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            actual_col = find_column_case_insensitive(df, column_name)
            if actual_col is None:
                continue
            series = df[actual_col]
            if equals is not None:
                if case_sensitive:
                    mask = series.astype(str) == str(equals)
                else:
                    mask = series.astype(str).str.lower() == str(equals).lower()
            elif contains is not None:
                mask = series.astype(str).str.contains(
                    str(contains), case=case_sensitive, na=False, regex=False,
                )
            else:
                mask = pd.Series([False] * len(df), index=df.index)

            sub = df.loc[mask].copy()
            if max_rows_per_csv is not None:
                sub = sub.head(max_rows_per_csv)
            if not sub.empty:
                sub.insert(0, "source_csv", csv_path.name)
                sub.insert(1, "relative_path", safe_relative_path(csv_path, root))
                frames.append(sub)
        except Exception:
            continue
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def groupby_mean_per_csv(
    data_folder: Any,
    group_col: str,
    value_col: str,
    *,
    exclude_zero: bool = False,
    recursive: bool = True,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    frames: list[pd.DataFrame] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            actual_group = find_column_case_insensitive(df, group_col)
            actual_value = find_column_case_insensitive(df, value_col)
            if actual_group is None or actual_value is None:
                frames.append(pd.DataFrame([{
                    "csv": csv_path.name,
                    "status": f"missing column(s): group={group_col}, value={value_col}",
                }]))
                continue

            values = pd.to_numeric(df[actual_value], errors="coerce")
            work = df.copy()
            work[actual_value] = values
            work = work[work[actual_value].notna()]
            if exclude_zero:
                work = work[work[actual_value] != 0]
            if work.empty:
                frames.append(pd.DataFrame([{
                    "csv": csv_path.name,
                    "status": "no valid numeric rows",
                }]))
                continue
            grouped = (
                work.groupby(actual_group, dropna=False)[actual_value]
                .agg(["count", "mean", "min", "max"])
                .reset_index()
            )
            grouped.insert(0, "csv", csv_path.name)
            grouped.insert(1, "relative_path", safe_relative_path(csv_path, root))
            grouped = grouped.rename(columns={actual_group: "group"})
            frames.append(grouped)
        except Exception as exc:
            frames.append(pd.DataFrame([{"csv": csv_path.name, "status": f"error: {exc}"}]))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ============================================================
# Text cleaning + date extraction
# ============================================================

def clean_string_column(
    df: pd.DataFrame,
    column: str,
    *,
    lowercase: bool = True,
    strip: bool = True,
    collapse_whitespace: bool = True,
    nfkc: bool = True,
) -> pd.DataFrame:
    """Normalize a string column in place: optional lowercase, strip
    whitespace, collapse runs of internal whitespace, and apply NFKC
    Unicode normalization. Returns the same df with the column updated."""
    col = column if column in df.columns else find_column_case_insensitive(df, column)
    if col is None:
        return df
    s = df[col].astype(str)
    if nfkc:
        import unicodedata as _ud
        s = s.map(lambda x: _ud.normalize("NFKC", x))
    if strip:
        s = s.str.strip()
    if collapse_whitespace:
        s = s.str.replace(r"\s+", " ", regex=True)
    if lowercase:
        s = s.str.lower()
    df[col] = s
    return df


def extract_dates(
    df: pd.DataFrame,
    column: str,
    *,
    new_column: Optional[str] = None,
    errors: str = "coerce",
) -> pd.DataFrame:
    """Parse a string column into pandas datetimes. By default writes
    the parsed value back into `column`; pass `new_column` to keep the
    original. Unparseable values become NaT."""
    col = column if column in df.columns else find_column_case_insensitive(df, column)
    if col is None:
        return df
    parsed = pd.to_datetime(df[col], errors=errors)
    target = new_column or col
    df[target] = parsed
    return df


# ============================================================
# Top-N / value-counts wrapper
# ============================================================

def top_n_per_csv(
    data_folder: Any,
    column: str,
    *,
    n: int = 10,
    recursive: bool = True,
) -> pd.DataFrame:
    """Top-N value counts for `column` in each CSV. Returns one long
    frame: csv, relative_path, rank, value, count."""
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
        except Exception as exc:
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": f"read error: {exc}",
            })
            continue
        col = find_column_case_insensitive(df, column)
        if col is None:
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": f"missing column: {column}",
            })
            continue
        vc = df[col].astype(str).value_counts(dropna=False).head(int(n))
        for rank, (value, count) in enumerate(vc.items(), start=1):
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "rank": rank,
                "value": str(value),
                "count": int(count),
            })
    return pd.DataFrame(rows)


# ============================================================
# Schema documentation
# ============================================================

def compare_schemas(file_a: Any, file_b: Any) -> pd.DataFrame:
    """Diff the column schemas of two CSV / Excel files.

    Returns a DataFrame with one row per column showing whether it's
    `only_in_a`, `only_in_b`, or `in_both`, plus the dtypes from each
    side when present. Useful for spotting renamed columns,
    dropped columns, or type drift between snapshots.
    """
    def _headers_and_types(path: Any):
        p = Path(path)
        suf = p.suffix.lower()
        try:
            if suf in (".xlsx", ".xls", ".xlsm"):
                df = pd.read_excel(p, nrows=50)
            else:
                df = read_table(p)
                df = df.head(50)
        except Exception:
            return [], {}
        return list(df.columns), {str(c): str(df[c].dtype) for c in df.columns}

    cols_a, types_a = _headers_and_types(file_a)
    cols_b, types_b = _headers_and_types(file_b)
    set_a = {str(c) for c in cols_a}
    set_b = {str(c) for c in cols_b}
    rows: List[Dict[str, Any]] = []
    for col in sorted(set_a | set_b):
        in_a = col in set_a
        in_b = col in set_b
        if in_a and in_b:
            status = ("type_changed" if types_a.get(col) != types_b.get(col)
                      else "in_both")
        elif in_a:
            status = "only_in_a"
        else:
            status = "only_in_b"
        rows.append({
            "column":  col,
            "status":  status,
            "dtype_a": types_a.get(col, ""),
            "dtype_b": types_b.get(col, ""),
        })
    return pd.DataFrame(rows)


# Type-inference patterns (beyond pandas's literal dtype).
_DATE_HINT_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}|^\d{1,2}/\d{1,2}/\d{2,4}|^\d{1,2}\.\d{1,2}\.\d{2,4}"
)
_DATETIME_HINT_RE = re.compile(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}")
_PERCENT_HINT_RE = re.compile(r"^-?\d+(?:\.\d+)?%$")
_CURRENCY_HINT_RE = re.compile(r"^[\$€£¥₹]\s?\d|^\d+(?:[,\d]+)?(?:\.\d+)?\s?(?:USD|EUR|GBP)")
_BOOL_HINT_RE = re.compile(r"^(true|false|yes|no|y|n|1|0)$", re.IGNORECASE)
_EMAIL_HINT_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
_URL_HINT_RE = re.compile(r"^https?://", re.IGNORECASE)
_PHONE_HINT_RE = re.compile(r"^\+?\d[\d\s().-]{6,}\d$")
_UUID_HINT_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)


def _infer_column_kind(s: pd.Series, *, sample_size: int = 200) -> str:
    """Return a richer "kind" label for a column beyond its raw dtype.

    Categories: date, datetime, percent, currency, boolean, email, url,
    phone, uuid, integer, float, categorical, text.
    """
    sample = s.dropna().astype(str).head(sample_size)
    if sample.empty:
        return "empty"

    if pd.api.types.is_bool_dtype(s):
        return "boolean"
    if pd.api.types.is_integer_dtype(s):
        return "integer"
    if pd.api.types.is_float_dtype(s):
        non_null = s.dropna()
        if not non_null.empty:
            try:
                # Safe whole-number check — % 1 == 0 works for any float
                if ((non_null % 1) == 0).all():
                    return "integer (stored as float)"
            except Exception:
                pass
        return "float"
    if pd.api.types.is_datetime64_any_dtype(s):
        return "datetime"

    n = len(sample)

    def _frac_matching(pat) -> float:
        return sum(1 for v in sample if pat.match(v.strip())) / n

    if _frac_matching(_DATETIME_HINT_RE) > 0.7:
        return "datetime"
    if _frac_matching(_DATE_HINT_RE) > 0.7:
        return "date"
    if _frac_matching(_PERCENT_HINT_RE) > 0.7:
        return "percent"
    if _frac_matching(_CURRENCY_HINT_RE) > 0.7:
        return "currency"
    if _frac_matching(_EMAIL_HINT_RE) > 0.7:
        return "email"
    if _frac_matching(_URL_HINT_RE) > 0.7:
        return "url"
    if _frac_matching(_PHONE_HINT_RE) > 0.7:
        return "phone"
    if _frac_matching(_UUID_HINT_RE) > 0.7:
        return "uuid"
    if _frac_matching(_BOOL_HINT_RE) > 0.95:
        return "boolean"
    # Try numeric coercion
    coerced = pd.to_numeric(sample, errors="coerce")
    if coerced.notna().mean() > 0.9:
        valid = coerced.dropna()
        try:
            if not valid.empty and ((valid % 1) == 0).all():
                return "integer (text-encoded)"
        except Exception:
            pass
        return "float (text-encoded)"
    # Categorical if low cardinality
    nunique = s.nunique(dropna=True)
    if 1 < nunique <= max(20, int(len(s) * 0.05)):
        return "categorical"
    return "text"


def column_type_inferences(path: Any) -> pd.DataFrame:
    """Infer a richer type label for every column in a CSV/Excel.

    Returns a DataFrame with: column, dtype, inferred_kind, non_null,
    null_pct, unique. The inferred_kind goes beyond pandas's dtype to
    label date/datetime/percent/currency/boolean/email/url/phone/uuid/
    integer/float/categorical/text — useful for deciding which helper
    to apply next or for catching encoding mistakes (a column showing
    `object` dtype but `email` kind, etc.).
    """
    p = Path(path)
    try:
        df = read_table(p)
    except Exception as exc:
        return pd.DataFrame([{"column": "", "error": str(exc)}])
    rows: List[Dict[str, Any]] = []
    total = max(1, len(df))
    for col in df.columns:
        s = df[col]
        rows.append({
            "column":         str(col),
            "dtype":          str(s.dtype),
            "inferred_kind":  _infer_column_kind(s),
            "non_null":       int(s.count()),
            "null_pct":       round(100 * int(s.isna().sum()) / total, 2),
            "unique":         int(s.nunique(dropna=True)),
        })
    return pd.DataFrame(rows)


def schema_doc_from_csv(path: Any) -> str:
    """Generate a Markdown schema doc from summarize_csv output.

    Useful for onboarding teammates onto a new dataset without writing
    docs by hand. Returns a Markdown string the caller can save."""
    p = Path(path)
    try:
        prof = summarize_csv(p)
    except Exception as exc:
        return f"# {p.name}\n\nCould not profile file: {exc}\n"
    # summarize_csv returns a 1-row diagnostic frame (cols: csv/column/status,
    # NO 'dtype') for an empty/unreadable file rather than raising. Detect
    # that shape so we don't KeyError on the per-column 'dtype' access below.
    if prof.empty or "dtype" not in prof.columns:
        note = ""
        if not prof.empty and "status" in prof.columns:
            note = f" ({prof.iloc[0]['status']})"
        return f"# {p.name}\n\nEmpty or unreadable.{note}\n"

    # summarize_csv emits exactly one profile row per column (we only reach
    # here when prof is the real profile — the empty/unreadable shape returned
    # above). So the column count is len(prof); no second read of the file.
    n_cols = len(prof)
    try:
        nrows_total = int(prof["non_null"].max() + prof["null_pct"].max() / 100 * prof["non_null"].max())
    except Exception:
        nrows_total = None

    md: list[str] = [
        f"# {p.name}",
        "",
        f"_Schema documentation, auto-generated from `summarize_csv`._",
        "",
        f"- Path: `{p}`",
        f"- Columns: {n_cols}",
    ]
    if nrows_total:
        md.append(f"- Approx rows: {nrows_total}")
    md.append("")
    md.append("| Column | Type | Non-null | Null % | Unique | Notes |")
    md.append("|---|---|---:|---:|---:|---|")
    for _, row in prof.iterrows():
        notes_parts = []
        if row.get("mean") is not None:
            try:
                notes_parts.append(f"mean={float(row['mean']):.2f}")
                notes_parts.append(f"min={row['min']}")
                notes_parts.append(f"max={row['max']}")
            except Exception:
                pass
        elif row.get("top_values"):
            notes_parts.append("top: " + str(row["top_values"]))
        md.append("| `{c}` | {dt} | {nn} | {npct} | {u} | {notes} |".format(
            c=row["column"],
            dt=row["dtype"],
            nn=row["non_null"],
            npct=row["null_pct"],
            u=row["unique"],
            notes=", ".join(notes_parts),
        ))
    md.append("")
    return "\n".join(md)


# ============================================================
# Data-quality checks
# ============================================================

def detect_data_quality_issues(df: pd.DataFrame) -> pd.DataFrame:
    """Run a battery of common data-quality checks against `df` and
    return a one-row-per-issue DataFrame.

    Checks performed:
      - duplicate rows
      - all-null columns
      - mixed dtypes within a single column (object cols where >5% of
        non-null values fail to coerce to numeric while others succeed)
      - leading/trailing whitespace in string columns
      - suspiciously round numeric columns (mean/std all integer)
      - encoding artifacts (presence of Unicode replacement char U+FFFD)
      - inconsistent date formats in date-like columns
    """
    rows: list[dict[str, Any]] = []

    def _issue(severity, where, kind, message):
        rows.append({
            "severity": severity, "column": where, "kind": kind,
            "message": message,
        })

    if df is None or df.empty:
        _issue("info", "(table)", "empty", "DataFrame is empty")
        return pd.DataFrame(rows)

    # 1. Duplicate rows
    try:
        dup = int(df.duplicated().sum())
        if dup > 0:
            _issue("warning", "(rows)", "duplicate_rows",
                   f"{dup} duplicate row(s)")
    except Exception:
        pass

    for col in df.columns:
        s = df[col]
        # 2. All-null column
        if s.notna().sum() == 0:
            _issue("warning", str(col), "all_null", "every value is null")
            continue

        # 3. Mixed numeric / non-numeric within object column
        if s.dtype == "object":
            non_null = s.dropna()
            if not non_null.empty:
                coerced = pd.to_numeric(non_null, errors="coerce")
                num_ok = coerced.notna().sum()
                num_total = len(non_null)
                if 0 < num_ok < num_total and num_ok / num_total > 0.05:
                    _issue("warning", str(col), "mixed_types",
                           f"{num_ok}/{num_total} values parse as numeric — "
                           f"column may have mixed numeric/text entries")

        # 4. Leading / trailing whitespace
        if s.dtype == "object":
            try:
                trimmable = s.dropna().astype(str).map(
                    lambda x: x != x.strip()
                ).sum()
                if trimmable > 0:
                    _issue("info", str(col), "whitespace",
                           f"{trimmable} value(s) have leading/trailing whitespace")
            except Exception:
                pass

        # 5. Encoding artifacts (U+FFFD)
        if s.dtype == "object":
            try:
                bad = s.dropna().astype(str).str.contains("�").sum()
                if bad > 0:
                    _issue("warning", str(col), "encoding",
                           f"{bad} value(s) contain U+FFFD (replacement char) — "
                           f"likely a decoding issue upstream")
            except Exception:
                pass

        # 6. Suspiciously round numerics (entire column = integers stored as float)
        if pd.api.types.is_float_dtype(s):
            non_null = s.dropna()
            if not non_null.empty and (non_null == non_null.astype(int)).all():
                _issue("info", str(col), "float_with_no_decimals",
                       "all values are whole numbers — consider an int dtype")

    # 7. Date columns with inconsistent formats — detect on object cols
    #    whose values include common date separators but parse to >2 distinct
    #    inferred formats.
    for col in df.columns:
        s = df[col]
        if s.dtype != "object":
            continue
        sample = s.dropna().astype(str).head(50)
        if sample.empty:
            continue
        looks_dateish = sample.str.contains(r"\d{2,4}[-/.\\]\d{1,2}").sum()
        if looks_dateish < max(3, len(sample) // 4):
            continue
        formats = set()
        for v in sample:
            # crude format inference — replace digits with 'd' so we group by shape
            formats.add(re.sub(r"\d", "d", v))
            if len(formats) > 5:
                break
        if len(formats) > 2:
            _issue("warning", str(col), "mixed_date_format",
                   f"date-like column shows {len(formats)} distinct value shapes")

    return pd.DataFrame(rows)


def detect_data_quality_issues_per_csv(
    data_folder: Any,
    *,
    recursive: bool = True,
) -> pd.DataFrame:
    """Run detect_data_quality_issues across every CSV in the folder.
    Returns a long frame with a leading `csv` column."""
    folders = normalize_data_folders(data_folder)
    frames: list[pd.DataFrame] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        try:
            df = _read_csv_cached(csv_path)
            issues = detect_data_quality_issues(df)
            if not issues.empty:
                issues.insert(0, "csv", csv_path.name)
                frames.append(issues)
        except Exception as exc:
            frames.append(pd.DataFrame([{
                "csv": csv_path.name,
                "severity": "error",
                "column": "",
                "kind": "read_error",
                "message": str(exc),
            }]))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ============================================================
# CSV splitter
# ============================================================

def split_csv_by_column(
    path: Any,
    column: str,
    output_dir: Any,
    *,
    sanitize: bool = True,
) -> pd.DataFrame:
    """Split a CSV into per-value files. For each distinct value in
    `column`, writes <output_dir>/<stem>_<value>.csv with only the
    matching rows. Returns a summary DataFrame (one row per output)."""
    p = Path(path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    df = pd.read_csv(p)
    col = column if column in df.columns else find_column_case_insensitive(df, column)
    if col is None:
        return pd.DataFrame([{"status": f"missing column: {column}"}])

    summary: list[dict[str, Any]] = []
    for value, sub in df.groupby(col, dropna=False):
        vname = str(value)
        if sanitize:
            vname = re.sub(r"[^A-Za-z0-9._-]+", "_", vname).strip("_") or "blank"
        out_path = out_dir / f"{p.stem}_{vname}.csv"
        n = 2
        while out_path.exists():
            out_path = out_dir / f"{p.stem}_{vname}_v{n}.csv"
            n += 1
        sub.to_csv(out_path, index=False)
        summary.append({
            "value":     str(value),
            "rows":      int(len(sub)),
            "output":    str(out_path),
        })
    return pd.DataFrame(summary)


def pivot_per_csv(
    data_folder: Any,
    index_col: str,
    columns_col: str,
    value_col: str,
    *,
    agg: str = "sum",
    recursive: bool = True,
) -> pd.DataFrame:
    """Pivot each CSV in the folder: rows=index_col, cols=columns_col,
    values=value_col aggregated by `agg`. Returns one long frame with
    a leading `csv` column distinguishing rows from different files."""
    folders = normalize_data_folders(data_folder)
    frames: list[pd.DataFrame] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            ix = find_column_case_insensitive(df, index_col)
            cx = find_column_case_insensitive(df, columns_col)
            vx = find_column_case_insensitive(df, value_col)
            if not (ix and cx and vx):
                frames.append(pd.DataFrame([{
                    "csv": csv_path.name,
                    "status": f"missing column(s): index={index_col}, "
                              f"columns={columns_col}, values={value_col}",
                }]))
                continue
            df[vx] = pd.to_numeric(df[vx], errors="coerce")
            df = df[df[vx].notna()]
            if df.empty:
                frames.append(pd.DataFrame([{
                    "csv": csv_path.name,
                    "status": "no numeric values to pivot",
                }]))
                continue
            piv = (df.pivot_table(
                       index=ix, columns=cx, values=vx, aggfunc=agg,
                       fill_value=0)
                     .reset_index())
            piv.insert(0, "csv", csv_path.name)
            piv.insert(1, "relative_path", safe_relative_path(csv_path, root))
            frames.append(piv)
        except Exception as exc:
            frames.append(pd.DataFrame([{
                "csv": csv_path.name,
                "status": f"error: {exc}",
            }]))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def correlation_per_csv(
    data_folder: Any,
    x_col: str,
    y_col: str,
    *,
    exclude_zero: bool = False,
    recursive: bool = True,
) -> pd.DataFrame:
    folders = normalize_data_folders(data_folder)
    rows: list[dict[str, Any]] = []
    for csv_path in list_csv_files(folders, recursive=recursive):
        root = first_matching_root(csv_path, folders)
        try:
            df = _read_csv_cached(csv_path)
            actual_x = find_column_case_insensitive(df, x_col)
            actual_y = find_column_case_insensitive(df, y_col)
            if actual_x is None or actual_y is None:
                rows.append({
                    "csv": csv_path.name,
                    "status": f"missing column(s): x={x_col}, y={y_col}",
                    "correlation": None,
                    "rows_used": 0,
                })
                continue
            x = pd.to_numeric(df[actual_x], errors="coerce")
            y = pd.to_numeric(df[actual_y], errors="coerce")
            mask = x.notna() & y.notna()
            if exclude_zero:
                mask = mask & (x != 0) & (y != 0)
            rows_used = int(mask.sum())
            corr = x[mask].corr(y[mask]) if rows_used >= 2 else None
            rows.append({
                "csv": csv_path.name,
                "relative_path": safe_relative_path(csv_path, root),
                "status": "ok" if rows_used >= 2 else "not enough valid rows",
                "x_column_used": actual_x,
                "y_column_used": actual_y,
                "correlation": float(corr) if corr is not None else None,
                "rows_used": rows_used,
                "total_rows": len(df),
            })
        except Exception as exc:
            rows.append({
                "csv": csv_path.name,
                "status": f"error: {exc}",
                "correlation": None,
                "rows_used": 0,
            })
    return pd.DataFrame(rows)


# ============================================================
# CSV-level helpers (single file)
# ============================================================

def summarize_csv(path: Any) -> pd.DataFrame:
    """One-call profile of a single CSV.

    Returns a DataFrame with one row per column: dtype, null %, unique
    count, top values (for categoricals), mean/std/min/max (for numerics).
    Saves the model from hand-rolling 15 lines of pandas.
    """
    p = Path(path)
    # An empty / header-less file raises EmptyDataError ("No columns to
    # parse"); a malformed one raises ParserError. Degrade to a one-row
    # diagnostic frame instead of crashing — these helpers are called over
    # whole folders, where one bad file shouldn't abort the batch.
    try:
        df = pd.read_csv(p)
    except Exception as exc:
        return pd.DataFrame([{
            "csv": p.name, "column": None,
            "status": f"could not read: {type(exc).__name__}: {exc}",
        }])
    if df.shape[1] == 0:
        return pd.DataFrame([{
            "csv": p.name, "column": None, "status": "empty file (no columns)",
        }])
    rows: List[Dict[str, Any]] = []
    total = max(1, len(df))
    for col in df.columns:
        s = df[col]
        nulls = int(s.isna().sum())
        rec: Dict[str, Any] = {
            "csv":          p.name,
            "column":       str(col),
            "dtype":        str(s.dtype),
            "non_null":     int(s.count()),
            "null_pct":     round(100 * nulls / total, 2),
            "unique":       int(s.nunique(dropna=True)),
        }
        # Try numeric coercion to see if we have aggregatable values
        coerced = pd.to_numeric(s, errors="coerce")
        if coerced.notna().sum() >= max(3, total * 0.5):
            valid = coerced.dropna()
            rec["mean"]   = float(valid.mean())
            rec["std"]    = float(valid.std()) if valid.size >= 2 else None
            rec["min"]    = float(valid.min())
            rec["max"]    = float(valid.max())
            rec["median"] = float(valid.median())
            rec["top_values"] = ""
        else:
            # Categorical / text — show top values
            try:
                vc = s.astype(str).value_counts(dropna=True).head(5)
                rec["top_values"] = ", ".join(
                    f"{idx}({cnt})" for idx, cnt in vc.items()
                )
            except Exception:
                rec["top_values"] = ""
            rec["mean"] = rec["std"] = rec["min"] = rec["max"] = rec["median"] = None
        rows.append(rec)
    return pd.DataFrame(rows)


def detect_outliers(
    df: pd.DataFrame,
    column: str,
    *,
    method: str = "iqr",
    threshold: float = 1.5,
) -> pd.DataFrame:
    """Return rows where `column` is an outlier.

    method='iqr'  — value outside [Q1 - k*IQR, Q3 + k*IQR] with k=threshold
    method='zscore' — |z| > threshold
    Adds an `outlier_score` column showing how far out the value is.
    """
    if column not in df.columns:
        actual = find_column_case_insensitive(df, column)
        if actual is None:
            return pd.DataFrame()
        column = actual
    values = pd.to_numeric(df[column], errors="coerce")
    work = df.loc[values.notna()].copy()
    work[column] = values[values.notna()]
    if work.empty:
        return work

    method = (method or "iqr").lower()
    if method == "zscore":
        mu = work[column].mean()
        sd = work[column].std() or 1.0
        z = (work[column] - mu) / sd
        mask = z.abs() > threshold
        work["outlier_score"] = z.abs()
    else:  # iqr
        q1 = work[column].quantile(0.25)
        q3 = work[column].quantile(0.75)
        iqr = q3 - q1
        lower = q1 - threshold * iqr
        upper = q3 + threshold * iqr
        mask = (work[column] < lower) | (work[column] > upper)
        # Vectorized element-wise max(lower-v, v-upper) / iqr. The old
        # .apply(lambda) ran one Python call per row (O(n) interpreter
        # round-trips); the row-wise pd.concat(...).max(axis=1) runs in
        # C. No numpy dependency (np is optional in this module).
        denom = max(iqr, 1e-9)
        col = work[column]
        work["outlier_score"] = pd.concat(
            [lower - col, col - upper], axis=1
        ).max(axis=1) / denom
    return work.loc[mask].sort_values("outlier_score", ascending=False)


def time_series_resample(
    df: pd.DataFrame,
    date_col: str,
    value_col: str,
    *,
    freq: str = "M",
    agg: str = "sum",
) -> pd.DataFrame:
    """Aggregate `value_col` over `date_col` at the given frequency.

    freq follows pandas offset aliases: 'D' (day), 'W' (week), 'M'
    (month), 'Q' (quarter), 'Y' (year). `agg` is any pandas agg name
    ('sum', 'mean', 'min', 'max', 'count', 'median').
    """
    dc = date_col if date_col in df.columns else find_column_case_insensitive(df, date_col)
    vc = value_col if value_col in df.columns else find_column_case_insensitive(df, value_col)
    if dc is None or vc is None:
        return pd.DataFrame()
    work = df.copy()
    work[dc] = pd.to_datetime(work[dc], errors="coerce")
    work[vc] = pd.to_numeric(work[vc], errors="coerce")
    work = work[work[dc].notna() & work[vc].notna()]
    if work.empty:
        return work
    grouped = work.set_index(dc)[vc].resample(freq).agg(agg)
    out = grouped.reset_index()
    out.columns = [dc, f"{vc}_{agg}"]
    return out


# ============================================================
# Cross-CSV helpers
# ============================================================

def join_csvs_on_column(
    left: Any,
    right: Any,
    on: str,
    *,
    how: str = "inner",
    suffixes: tuple = ("_left", "_right"),
) -> pd.DataFrame:
    """Read two CSVs and join them on a column.

    `left` and `right` are CSV paths. `on` is the column name (case-
    insensitive). `how` is any pandas join kind: 'inner', 'left',
    'right', 'outer'.
    """
    df_l = pd.read_csv(left) if not isinstance(left, pd.DataFrame) else left
    df_r = pd.read_csv(right) if not isinstance(right, pd.DataFrame) else right
    col_l = on if on in df_l.columns else find_column_case_insensitive(df_l, on)
    col_r = on if on in df_r.columns else find_column_case_insensitive(df_r, on)
    if col_l is None or col_r is None:
        return pd.DataFrame()
    # Rename to a common column name so merge works cleanly
    df_l = df_l.rename(columns={col_l: on})
    df_r = df_r.rename(columns={col_r: on})
    return pd.merge(df_l, df_r, on=on, how=how, suffixes=suffixes)


def compare_two_csvs(
    a: Any,
    b: Any,
    *,
    on: Optional[str] = None,
) -> pd.DataFrame:
    """Row-level diff between two CSV snapshots.

    With `on`: rows are matched by that key column. Result has a `status`
    column with 'added' / 'removed' / 'changed' / 'same' and per-column
    columns showing left/right values where they differ.

    Without `on`: returns rows present in exactly one CSV (treats every
    row as a fingerprint). Faster but less informative.
    """
    df_a = pd.read_csv(a) if not isinstance(a, pd.DataFrame) else a
    df_b = pd.read_csv(b) if not isinstance(b, pd.DataFrame) else b

    if on:
        col_a = on if on in df_a.columns else find_column_case_insensitive(df_a, on)
        col_b = on if on in df_b.columns else find_column_case_insensitive(df_b, on)
        if col_a is None or col_b is None:
            return pd.DataFrame([{"status": "error",
                                  "message": f"join column {on!r} missing"}])
        df_a = df_a.rename(columns={col_a: on}).set_index(on)
        df_b = df_b.rename(columns={col_b: on}).set_index(on)
        all_keys = df_a.index.union(df_b.index)
        rows: List[Dict[str, Any]] = []
        common_cols = [c for c in df_a.columns if c in df_b.columns]
        for k in all_keys:
            in_a = k in df_a.index
            in_b = k in df_b.index
            if in_a and not in_b:
                rows.append({on: k, "status": "removed"})
            elif in_b and not in_a:
                rows.append({on: k, "status": "added"})
            else:
                ra = df_a.loc[k]
                rb = df_b.loc[k]
                diffs = {}
                for c in common_cols:
                    va = ra[c] if not isinstance(ra, pd.DataFrame) else ra[c].iloc[0]
                    vb = rb[c] if not isinstance(rb, pd.DataFrame) else rb[c].iloc[0]
                    if pd.isna(va) and pd.isna(vb):
                        continue
                    if va != vb:
                        diffs[f"{c}_a"] = va
                        diffs[f"{c}_b"] = vb
                rec = {on: k, "status": "changed" if diffs else "same"}
                rec.update(diffs)
                rows.append(rec)
        return pd.DataFrame(rows)
    # No key — set-style diff
    sa = set(map(tuple, df_a.astype(str).values.tolist()))
    sb = set(map(tuple, df_b.astype(str).values.tolist()))
    added   = [list(t) + ["added"]   for t in sb - sa]
    removed = [list(t) + ["removed"] for t in sa - sb]
    cols = list(df_a.columns) + ["status"]
    return pd.DataFrame(added + removed, columns=cols)


def preview_csv_inventory(folders: Any, max_files: int = 25, max_cols: int = 30) -> str:
    """Compact, prompt-friendly listing of CSVs and their columns."""
    normalized = normalize_data_folders(folders)
    csvs = list_csv_files(normalized)
    lines = [f"CSV files found: {len(csvs)}"]
    for csv_path in csvs[:max_files]:
        root = first_matching_root(csv_path, normalized)
        try:
            df_head = _read_csv_cached(csv_path, nrows=5)
            cols = list(df_head.columns)
            shown_cols = cols[:max_cols]
            more = "" if len(cols) <= max_cols else f" ... (+{len(cols) - max_cols} more)"
            lines.append(f"- {csv_path.name}")
            lines.append(f"  relative_path: {safe_relative_path(csv_path, root)}")
            lines.append(f"  columns: {shown_cols}{more}")
        except Exception as exc:
            lines.append(f"- {csv_path.name}  (error reading columns: {exc})")
    if len(csvs) > max_files:
        lines.append(f"... {len(csvs) - max_files} more CSV files not shown.")
    return "\n".join(lines)


# ============================================================
# Sandboxed code execution
# ============================================================

DANGEROUS_PATTERNS = [
    r"\bimport\s+os\b",
    r"\bimport\s+sys\b",
    r"\bimport\s+subprocess\b",
    r"\bimport\s+shutil\b",
    r"\bimport\s+socket\b",
    r"\bimport\s+requests\b",
    r"\bimport\s+urllib\b",
    r"\bfrom\s+os\b",
    r"\bfrom\s+sys\b",
    r"\bfrom\s+subprocess\b",
    r"\bfrom\s+shutil\b",
    r"\bfrom\s+socket\b",
    r"\bfrom\s+requests\b",
    r"\bfrom\s+urllib\b",
    r"\beval\s*\(",
    r"\bexec\s*\(",
    r"\bcompile\s*\(",
    r"\binput\s*\(",
    r"\bopen\s*\([^)]*,\s*['\"][wa+x]",
    r"\.to_csv\s*\(",
    r"\.to_excel\s*\(",
    r"\.to_parquet\s*\(",
    r"\.to_pickle\s*\(",
    r"\.unlink\s*\(",
    r"\.rmdir\s*\(",
    r"\.remove\s*\(",
    r"\.rename\s*\(",
    r"\brmtree\s*\(",
    r"\bsystem\s*\(",
    r"\bpopen\s*\(",
]


# Attribute names the sandbox refuses — at BOTH the AST validator and the
# sandboxed getattr (so `getattr(obj, "write_text")` can't dodge the AST check).
# pathlib.Path is exposed for read-side path construction, so its WRITE/CREATE
# surface must be blocked here or model code could write arbitrary files
# (Path("/x").write_text(...), .touch(), .mkdir(), .open("w"), …) — a real
# arbitrary-write escape. Deletion / process / DB-write names are here too.
# NB: names that collide with common, legitimate pandas/str methods
# (replace, to_json/to_html/to_string which return strings, drop, …) are
# deliberately NOT listed.
_SANDBOX_FORBIDDEN_ATTRS = frozenset({
    # deletion / rename (data-loss)
    "unlink", "rmdir", "remove", "rename", "rmtree", "removedirs",
    # serialise-to-disk / DB writes (write-only forms)
    "to_csv", "to_excel", "to_parquet", "to_pickle", "to_sql",
    "to_feather", "to_hdf", "to_stata",
    # pathlib / filesystem write + create surface
    "write_text", "write_bytes", "touch", "mkdir", "makedirs", "mknod",
    "symlink_to", "hardlink_to", "chmod", "lchmod", "open", "fdopen",
    # process / shell
    "system", "popen", "Popen",
})
# Introspection dunders that enable classic eval-sandbox escapes (reach os via
# the class hierarchy, or grab __globals__ / __builtins__).
_SANDBOX_FORBIDDEN_DUNDERS = frozenset({
    "__subclasses__", "__bases__", "__mro__", "__base__", "__globals__",
    "__subclasshook__", "__builtins__", "__code__", "__closure__",
    "__reduce__", "__reduce_ex__", "__getattribute__",
})


def validate_generated_code(code: str) -> Tuple[bool, str]:
    if not code.strip():
        return False, "Generated code is empty."
    for pattern in DANGEROUS_PATTERNS:
        if re.search(pattern, code, flags=re.IGNORECASE):
            return False, f"Blocked unsafe pattern: {pattern}"
    try:
        tree = ast.parse(code)
    except SyntaxError as exc:
        return False, f"Syntax error: {exc}"

    forbidden_import_roots = {
        "os", "sys", "subprocess", "shutil", "socket",
        "requests", "urllib", "http", "ftplib",
    }
    forbidden_calls = {"eval", "exec", "compile", "input"}
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                root = alias.name.split(".")[0]
                if root in forbidden_import_roots:
                    return False, f"Forbidden import: {alias.name}"
        if isinstance(node, ast.ImportFrom):
            root = (node.module or "").split(".")[0]
            if root in forbidden_import_roots:
                return False, f"Forbidden import from: {node.module}"
        # Any attribute access to an escape dunder — even without a call —
        # e.g. `x.__globals__[...]`, `type(1).__mro__`.
        if isinstance(node, ast.Attribute) and node.attr in _SANDBOX_FORBIDDEN_DUNDERS:
            return False, f"Forbidden attribute: {node.attr}"
        if isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id in forbidden_calls:
                return False, f"Forbidden call: {node.func.id}"
            if isinstance(node.func, ast.Attribute) and node.func.attr in _SANDBOX_FORBIDDEN_ATTRS:
                return False, f"Forbidden method: {node.func.attr}"
    return True, "ok"


def _safe_import(name, globals=None, locals=None, fromlist=(), level=0):
    # scipy added for the SPC / engineering / stats helpers — it has
    # no filesystem, network, or subprocess surface; it's a pure
    # numerical library and a safe addition to the sandbox allowlist.
    allowed_roots = {
        "pandas", "pathlib", "numpy", "math", "re", "json",
        "statistics", "collections",
        "scipy",          # SPC capability indices, ANOVA, FFT, normality tests
    }
    root = name.split(".")[0]
    if root not in allowed_roots:
        raise ImportError(f"Import blocked by sandbox: {name}")
    return __import__(name, globals, locals, fromlist, level)


# Readers that pull a whole file into a DataFrame. Looping any of these
# over hundreds of files (then pd.concat-ing) is the one way model-written
# analyst code can exhaust memory and OOM-crash the app — every other
# sandbox helper aggregates file-by-file. We cap the cumulative bytes
# such reads hold within a single execution.
_BUDGETED_READERS = frozenset({
    "read_csv", "read_table", "read_excel", "read_parquet",
    "read_json", "read_feather", "read_orc", "read_fwf",
})


def _analyst_read_budget_bytes() -> int:
    """Bytes a single analyst execution may hold across whole-file reads.
    Override with COUNCIL_ANALYST_READ_BUDGET_MB. Default scales to the
    machine: min(1.5 GiB, 40% of available RAM) — low enough on a
    memory-capped box (e.g. WSL) to REFUSE before the OS OOM-killer fires,
    generous enough that normal multi-file queries never trip it."""
    ov = os.environ.get("COUNCIL_ANALYST_READ_BUDGET_MB", "").strip()
    if ov:
        try:
            # Respect an explicit override down to a small sane floor.
            return max(8, int(ov)) * 1024 * 1024
        except ValueError:
            pass
    cap = 1536 * 1024 * 1024
    try:
        import psutil
        avail = int(psutil.virtual_memory().available * 0.40)
        if avail > 0:
            return min(cap, avail)
    except Exception:
        pass
    return cap


class _BudgetedPandas:
    """Proxy around the real pandas module for the analyst sandbox.

    Delegates everything to pandas EXCEPT the whole-file readers, which it
    wraps to accrue the in-memory size of every frame they return. Once the
    running total crosses the budget it raises a CATCHABLE MemoryError —
    refusing the 201st read in a `pd.concat([read_csv(f) for f in files])`
    loop *before* the allocation that would OOM-kill the process. The error
    points the model at the bounded per-file helpers instead.
    """

    def __init__(self, real, budget_bytes: int, state: dict) -> None:
        self._real = real
        self._budget = int(budget_bytes)
        self._state = state          # {"used": int}

    def _account(self, df):
        try:
            n = int(df.memory_usage(deep=True).sum())
        except Exception:
            n = 0
        self._state["used"] += n
        if self._state["used"] > self._budget:
            used_mb = self._state["used"] // (1024 * 1024)
            cap_mb = self._budget // (1024 * 1024)
            raise MemoryError(
                f"Analyst read budget exceeded (~{used_mb} MB held, cap "
                f"{cap_mb} MB). This query loads too many whole files at "
                "once. Use the bounded per-file helpers instead — e.g. "
                "column_stats(), numeric_summary_per_csv(), "
                "average_numeric_column_per_csv(), count_rows_per_csv() — "
                "which aggregate file-by-file without holding every file "
                "in memory.")
        return df

    def __getattr__(self, name):
        # __getattr__ only fires for names not found normally, so the
        # instance attrs (_real/_budget/_state) never route through here.
        real_attr = getattr(self._real, name)
        if name in _BUDGETED_READERS and callable(real_attr):
            def _guarded(*a, **k):
                return self._account(real_attr(*a, **k))
            _guarded.__name__ = name
            return _guarded
        return real_attr


def execute_pandas_code(
    code: str,
    allowed_folders: List[Path],
) -> Tuple[Optional[pd.DataFrame], str]:
    ok, msg = validate_generated_code(code)
    if not ok:
        return None, f"SAFETY CHECK FAILED: {msg}"

    normalized_folders = normalize_data_folders(allowed_folders)
    if not normalized_folders:
        return None, "No allowed folders configured for the analyst."

    # Per-execution memory budget for whole-file reads. The proxy below
    # guards `pd`; the import hook hands the SAME proxy back for any
    # `import pandas` so model code can't sidestep the cap by re-importing.
    _read_state = {"used": 0}
    _budgeted_pd = _BudgetedPandas(pd, _analyst_read_budget_bytes(), _read_state)

    def _guarded_import(name, globals=None, locals=None, fromlist=(), level=0):
        if name.split(".")[0] == "pandas":
            return _budgeted_pd
        return _safe_import(name, globals, locals, fromlist, level)

    def _guarded_getattr(obj, name, *default):
        # Close the getattr bypass: the AST validator blocks `x.write_text(...)`
        # by name, but `getattr(x, "write_text")(...)` would sidestep it. Refuse
        # the same forbidden names + escape dunders here too.
        if isinstance(name, str) and (
                name in _SANDBOX_FORBIDDEN_ATTRS
                or name in _SANDBOX_FORBIDDEN_DUNDERS):
            raise AttributeError(
                f"sandbox: access to attribute {name!r} is blocked")
        return getattr(obj, name, *default)

    safe_builtins = {
        "__import__": _guarded_import,
        # Numeric / collection core
        "abs": abs, "all": all, "any": any, "bool": bool, "dict": dict,
        "enumerate": enumerate, "float": float, "int": int,
        "isinstance": isinstance, "issubclass": issubclass,
        "len": len, "list": list, "max": max, "min": min, "print": print,
        "range": range, "round": round, "set": set, "frozenset": frozenset,
        "sorted": sorted, "str": str, "sum": sum, "tuple": tuple, "zip": zip,
        # Attribute / introspection — model frequently writes
        # `getattr(df, "shape")`, `type(x).__name__`, etc. Without these
        # in builtins the sandbox raises NameError mid-snippet and the
        # analyst silently falls back to model freeform (= wrong answer).
        "getattr": _guarded_getattr, "hasattr": hasattr,
        "type": type, "repr": repr, "format": format, "hash": hash,
        "id": id, "callable": callable, "vars": vars, "dir": dir,
        # Iterator helpers
        "iter": iter, "next": next, "reversed": reversed,
        "map": map, "filter": filter, "slice": slice,
        # Numeric / char helpers
        "divmod": divmod, "pow": pow, "ord": ord, "chr": chr,
        "bin": bin, "hex": hex, "oct": oct,
        "bytes": bytes, "bytearray": bytearray, "complex": complex,
        # Literals — needed inside exec because __builtins__ replaced
        "True": True, "False": False, "None": None,
        # Common exception classes for except-clauses
        "Exception": Exception, "ValueError": ValueError, "TypeError": TypeError,
        "KeyError": KeyError, "IndexError": IndexError,
        "AttributeError": AttributeError, "ZeroDivisionError": ZeroDivisionError,
        "StopIteration": StopIteration, "OverflowError": OverflowError,
        "LookupError": LookupError, "ArithmeticError": ArithmeticError,
        "RuntimeError": RuntimeError, "NotImplementedError": NotImplementedError,
    }

    globals_dict: dict[str, Any] = {
        "__builtins__": safe_builtins,
        "pd": _budgeted_pd,
        "Path": Path,
        "DATA_FOLDERS": [str(p) for p in normalized_folders],
        "DATA_FOLDER": str(normalized_folders[0]),
        "list_csv_files": list_csv_files,
        "find_column_case_insensitive": find_column_case_insensitive,
        "find_columns_contains": find_columns_contains,
        "csv_inventory": csv_inventory,
        "folder_data_summary": folder_data_summary,
        # Cache-backed precomputed column stats — fast on repeat (served
        # from <vault>/.stats_cache, computed + stored on first sight).
        # vault_dir=None lets stats_cache resolve the vault root itself so
        # the location is the same regardless of the scope folder passed.
        #   column_stats(folder=None) -> one row per (file, column)
        #   file_stats(path)          -> dict of stats for one CSV
        "column_stats": (lambda folder=None: folder_column_stats(
            None, folder if folder is not None
            else [str(p) for p in normalized_folders])),
        "file_stats": (lambda path: cached_column_stats(None, path)),
        "count_rows_per_csv": count_rows_per_csv,
        "average_numeric_column_per_csv": average_numeric_column_per_csv,
        "std_numeric_column_per_csv": std_numeric_column_per_csv,
        "sum_numeric_column_per_csv": sum_numeric_column_per_csv,
        "min_max_numeric_column_per_csv": min_max_numeric_column_per_csv,
        "numeric_summary_per_csv": numeric_summary_per_csv,
        "count_matching_rows_per_csv": count_matching_rows_per_csv,
        "filter_rows_across_csvs": filter_rows_across_csvs,
        "groupby_mean_per_csv": groupby_mean_per_csv,
        "correlation_per_csv": correlation_per_csv,
        # Newer single-file + cross-file helpers
        "summarize_csv":          summarize_csv,
        "detect_outliers":        detect_outliers,
        "time_series_resample":   time_series_resample,
        "join_csvs_on_column":    join_csvs_on_column,
        "compare_two_csvs":       compare_two_csvs,
        # Excel + CSV-or-Excel helpers
        "list_excel_files":       list_excel_files,
        "list_data_files":        list_data_files,
        "read_table":             read_table,
        "read_excel_sheets":      read_excel_sheets,
        "excel_inventory":        excel_inventory,
        "detect_excel_header_rows":      detect_excel_header_rows,
        "read_excel_with_merged_headers": read_excel_with_merged_headers,
        # Messy-data helpers
        "read_csv_robust":        read_csv_robust,
        "find_data_block":        find_data_block,
        "read_excel_all_tables":  read_excel_all_tables,
        "read_excel_smart_tables": read_excel_smart_tables,
        "strip_summary_rows":     strip_summary_rows,
        "unpivot_year_columns":   unpivot_year_columns,
        # Schema diff + smart type inference
        "compare_schemas":        compare_schemas,
        "column_type_inferences": column_type_inferences,
        # Parquet / SQLite + pivot
        "list_parquet_files":     list_parquet_files,
        "list_sqlite_files":      list_sqlite_files,
        "list_sqlite_tables":     list_sqlite_tables,
        "read_sqlite_table":      read_sqlite_table,
        "pivot_per_csv":          pivot_per_csv,
        # Text cleaning + top-N + schema doc
        "clean_string_column":    clean_string_column,
        "extract_dates":          extract_dates,
        "top_n_per_csv":          top_n_per_csv,
        "schema_doc_from_csv":    schema_doc_from_csv,
        # Data quality + splitter
        "detect_data_quality_issues":         detect_data_quality_issues,
        "detect_data_quality_issues_per_csv": detect_data_quality_issues_per_csv,
        "split_csv_by_column":                split_csv_by_column,
        # DuckDB
        "list_duckdb_files":   list_duckdb_files,
        "list_duckdb_tables":  list_duckdb_tables,
        "read_duckdb_table":   read_duckdb_table,
        "duckdb_query":        duckdb_query,
        # BSON / MongoDB
        "list_bson_files":     list_bson_files,
        "read_bson_documents": read_bson_documents,
        "read_bson_as_df":     read_bson_as_df,
        # Mongo -> model-digestible: flatten nested docs, coerce ObjectId /
        # dates / arrays into clean scalars. Work on a .bson / .json / .jsonl
        # path OR a list of docs (e.g. from read_mongo_collection).
        "read_json_documents":     read_json_documents,
        "mongo_documents_to_frame": mongo_documents_to_frame,
        "bson_to_clean_frame":      bson_to_clean_frame,
        "json_to_clean_frame":      json_to_clean_frame,
        "mongo_schema_profile":     mongo_schema_profile,
        "mongo_documents_to_text":  mongo_documents_to_text,
        "mongo_explode_array":      mongo_explode_array,
        # SQLAlchemy bridge
        "list_sql_connections": list_sql_connections,
        "list_sql_tables":      list_sql_tables,
        "read_sql_table":       read_sql_table,
        "sql_query":            sql_query,
        # MongoDB — read-only by API design. Pipeline validator
        # blocks $out / $merge / $function / $accumulator / $where.
        "list_mongo_connections": list_mongo_connections,
        "list_mongo_databases":   list_mongo_databases,
        "list_mongo_collections": list_mongo_collections,
        "read_mongo_collection":  read_mongo_collection,
        "mongo_aggregate":        mongo_aggregate,
        "mongo_count":            mongo_count,
        "mongo_distinct":         mongo_distinct,
    }
    if np is not None:
        globals_dict["np"] = np

    # Domain helpers — SPC (Gate A), engineering + stats (Gate B).
    # The register_helpers entry point keeps the wiring centralised so
    # the sandbox surface for new analytic capabilities lands in one
    # place, and a registration failure in one helper module (missing
    # scipy on a CPU-only bundle, say) doesn't take the others down.
    try:
        import analyst_helpers as _ah
        _ah.register_helpers(globals_dict)
    except Exception as _ah_exc:
        import sys as _sys_dbg
        print(f"[analyst] domain helpers not registered: {_ah_exc!r}",
              file=_sys_dbg.stderr)

    # CRITICAL: pass `globals_dict` as BOTH globals and locals. When
    # exec(code, globals, locals) is called with *different* dicts, Python
    # treats the top-level code as if it were inside a class body. That
    # scoping rule means dict/list comprehensions and nested functions
    # CANNOT see top-level variables (free-var lookup in nested function
    # scopes skips the class-like enclosing scope, per PEP 227 / 3104).
    #
    # In practice this surfaces as confusing `NameError: name '<x>' is not
    # defined` whenever the model writes the very natural pattern:
    #
    #     name = "rating"
    #     non_zero = [v for v in df[name] if v != 0]
    #     result_df = pd.DataFrame({"avg": [sum(non_zero)/len(non_zero)]})
    #
    # The iterable `df[name]` evaluates fine in the enclosing class-scope,
    # but a body reference to `name` inside the comprehension fails. We
    # use a single namespace so the model's hand-rolled pandas snippets
    # behave the same as if they ran in a normal module.
    stdout_buf = io.StringIO()
    stderr_buf = io.StringIO()

    try:
        with redirect_stdout(stdout_buf), redirect_stderr(stderr_buf):
            exec(code, globals_dict)
    except Exception:
        log = "EXECUTION ERROR:\n" + traceback.format_exc()
        if stdout_buf.getvalue():
            log += "\nSTDOUT:\n" + stdout_buf.getvalue()
        if stderr_buf.getvalue():
            log += "\nSTDERR:\n" + stderr_buf.getvalue()
        return None, log

    result_df = globals_dict.get("result_df")

    log_parts: list[str] = []
    if stdout_buf.getvalue():
        log_parts.append("STDOUT:\n" + stdout_buf.getvalue())
    if stderr_buf.getvalue():
        log_parts.append("STDERR:\n" + stderr_buf.getvalue())

    if result_df is None:
        log_parts.append("Code did not create `result_df`.")
        return None, "\n\n".join(log_parts) if log_parts else "no result_df"

    if isinstance(result_df, pd.Series):
        result_df = result_df.reset_index()
    if isinstance(result_df, dict):
        result_df = pd.DataFrame([result_df])
    if not isinstance(result_df, pd.DataFrame):
        log_parts.append(f"result_df is not a DataFrame; got {type(result_df)}")
        return None, "\n\n".join(log_parts)

    if not log_parts:
        log_parts.append("ok")
    return result_df, "\n\n".join(log_parts)


# ============================================================
# Code generation prompt
# ============================================================

def build_pandas_code_prompt(
    question: str,
    data_folders: List[Path],
    inventory: str,
    filename_hints: Optional[str] = None,
    subfolder_scope: Optional[Path] = None,
) -> str:
    folder_lines = "\n".join(f"- {p}" for p in data_folders)
    # When the analyst pre-resolved filename references or restricted the
    # search scope, surface that context to the model BEFORE the inventory
    # so a fuzzy reference like "sales report" is mapped to the real file
    # and the model never gets a chance to invent a wrong path.
    scope_note = ""
    if subfolder_scope is not None:
        scope_note = (
            "\nSCOPE — The user mentioned a subfolder. The analyst has "
            "restricted DATA_FOLDERS to:\n"
            f"  {subfolder_scope}\n"
            "Read files from this scope only. Do NOT scan the broader "
            "vault.\n"
        )
    hint_note = ""
    if filename_hints:
        hint_note = "\n" + filename_hints + "\n"
    return f"""You are writing a single pandas snippet that answers the user's question
about CSV files in their vault. Output ONLY executable Python code — no
markdown fences, no commentary, no explanation.

User question:
{question}
{scope_note}{hint_note}
Available data folders:
{folder_lines}

CSV inventory (column names taken from real files):
{inventory}
""" + """
Available variables:
- DATA_FOLDERS  (list[str]) — pass this as `data_folder=DATA_FOLDERS` to helpers
- DATA_FOLDER   (str)       — first folder, for legacy single-folder calls
- pd            (pandas)
- Path          (pathlib.Path)
- np            (numpy, may be None)

Available helper functions (prefer these over raw pandas — they handle
case-insensitive column matching, NaN/zero filtering, and multi-CSV scans):
  list_csv_files(data_folder, recursive=True)
  csv_inventory(data_folder, recursive=True, max_files=None)
  folder_data_summary(data_folder, recursive=True, max_files=None)
    THE GO-TO HELPER FOR "GIVE ME A DATA SUMMARY" QUERIES. Returns
    one row per file across CSV / TSV / Parquet / Excel / JSON /
    SQLite / DuckDB / BSON / image / text formats with:
       file, relative_path, type, size_kb, rows, columns,
       column_names, dtypes, missing_pct, numeric_cols, date_cols,
       sample_value, notes
    Use this when the user asks for "a true data summary", "what's
    in the folder", "describe the files", "schema overview", or
    "inventory of the subfolder X" — it answers the WHOLE question
    in one call instead of needing per-file code.
  count_rows_per_csv(data_folder, recursive=True)
  find_column_case_insensitive(df, column_name)
  find_columns_contains(df, text)
  average_numeric_column_per_csv(data_folder, column_name, exclude_zero=False)
  std_numeric_column_per_csv(data_folder, column_name, exclude_zero=False)
  sum_numeric_column_per_csv(data_folder, column_name, exclude_zero=False)
  min_max_numeric_column_per_csv(data_folder, column_name, exclude_zero=False)
  numeric_summary_per_csv(data_folder, column_name=None, exclude_zero=False)
  count_matching_rows_per_csv(data_folder, column_name, equals=None, contains=None,
                              case_sensitive=False)
  filter_rows_across_csvs(data_folder, column_name, equals=None, contains=None,
                          case_sensitive=False, max_rows_per_csv=None)
  groupby_mean_per_csv(data_folder, group_col, value_col, exclude_zero=False)
  correlation_per_csv(data_folder, x_col, y_col, exclude_zero=False)
  summarize_csv(path)                          # one-row-per-column profile
  detect_outliers(df, column, method="iqr", threshold=1.5)
  time_series_resample(df, date_col, value_col, freq="M", agg="sum")
  join_csvs_on_column(left_path, right_path, on, how="inner")
  compare_two_csvs(a_path, b_path, on=None)   # row-level diff snapshot

Excel + cross-format support:
  list_excel_files(data_folder, recursive=True)
  list_data_files(data_folder, recursive=True)  # CSV + Excel combined
  read_table(path, sheet=None)                  # CSV, TSV, .csv.gz,
                                                # XLSX, Parquet -> df
  read_excel_sheets(path)                       # all sheets -> {name: df}
  excel_inventory(data_folder, recursive=True)
  detect_excel_header_rows(path, sheet=None)    # auto-detect merged headers
  read_excel_with_merged_headers(path, sheet=0, header_rows=None, flatten=True)

Messy data — when CSV / Excel files don't follow the "row 1 is headers,
rest is data" convention:
  read_csv_robust(path)                # auto-detect encoding + separator,
                                       # skip title/banner rows, strip
                                       # trailing summary rows. Returns
                                       # (df, diagnostics).
  find_data_block(path_or_df)          # bounding box of the actual data
                                       # region (start_row, end_row,
                                       # start_col, end_col).
  read_excel_all_tables(path, sheet)   # simple multi-table reader
                                       # (blank-row separation only)
  read_excel_smart_tables(path, sheet, gap_tolerance=1, min_cells=4)
    Robust version. Use this when read_excel_all_tables misses
    tables. Handles: vertical AND horizontal gaps between tables,
    merged cells in the body (expanded), inconsistent row widths
    (short rows padded), title/note rows above the real header
    (auto-detected per cluster). Returns a list of
    {df, top_left, n_rows, n_cols, header_row} dicts.
  strip_summary_rows(df, patterns=...) # drop trailing "Total" rows.
  unpivot_year_columns(df, id_cols)    # wide-form (..., 2020, 2021, 2022)
                                       # -> long form (country, year, value).
    USE THIS for workbooks where the top row(s) are merged group headers
    above the real column-name row (e.g. "Site A | Site B | Site C" merged
    over rows of "energy, voltage, current" sub-columns). Without it,
    pd.read_excel produces "Unnamed: 0", "Unnamed: 1" columns and the
    data is unusable. With it, you get columns named "Site A / energy",
    "Site B / energy", etc. — filter by the suffix to operate on each
    sub-column across groups.

Parquet + SQLite:
  list_parquet_files(data_folder, recursive=True)
  list_sqlite_files(data_folder, recursive=True)
  list_sqlite_tables(path)                      # list[str] of table names
  read_sqlite_table(path, table, limit=None)    # df from a SQLite table

Pivot tables:
  pivot_per_csv(data_folder, index_col, columns_col, value_col, agg="sum")

Text + dates + value counts + schema doc:
  clean_string_column(df, column, lowercase=True, strip=True, ...)
  extract_dates(df, column, new_column=None, errors="coerce")
  top_n_per_csv(data_folder, column, n=10)     # value_counts per file
  schema_doc_from_csv(path)                    # -> markdown string

Data quality + splitter:
  detect_data_quality_issues(df)
    issues table: duplicates, all-null cols, mixed types, whitespace,
    encoding artifacts, suspicious round numerics, mixed date formats
  detect_data_quality_issues_per_csv(folder)
  split_csv_by_column(path, column, output_dir)
    writes one CSV per distinct value of `column`

DuckDB (file-based analytical SQL):
  list_duckdb_files(folder, recursive=True)
  list_duckdb_tables(path)
  read_duckdb_table(path, table, limit=None)
  duckdb_query(path, sql)
    DuckDB can SELECT directly from CSV/Parquet/JSON files inside the
    SQL — e.g. duckdb_query(db, "SELECT * FROM read_csv('data.csv')").

BSON / MongoDB:
  list_bson_files(folder, recursive=True)
  read_bson_documents(path)    # -> list[dict]
  read_bson_as_df(path)        # -> DataFrame (raw; nested cells kept)
  read_json_documents(path)    # -> list[dict] from .json / .jsonl / NDJSON

Mongo -> model-digestible (PREFER these for BSON/JSON — they flatten nested
docs and turn ObjectId / dates / Decimal128 / arrays into clean scalars, so
the model isn't fed raw ObjectId and array-of-subdocument noise). Each accepts
a .bson / .json / .jsonl PATH or a list of docs (e.g. from read_mongo_collection):
  bson_to_clean_frame(path)              # .bson  -> clean flat DataFrame
  json_to_clean_frame(path)              # .json/.jsonl -> clean flat DataFrame
  mongo_documents_to_frame(docs)         # list[dict] -> clean flat DataFrame
  mongo_schema_profile(docs_or_path)     # -> field, types, present_pct, example
  mongo_documents_to_text(docs_or_path)  # -> compact key:value text for a prompt
  mongo_explode_array(docs_or_path, "line_items", meta=["_id","name"])
                                         # one row per array element (tidy view)

Remote SQL via SQLAlchemy (READ-ONLY — see notes below):
  list_sql_connections(vault_dir)    # -> {name: url}
  list_sql_tables(vault_dir, conn_name)
  read_sql_table(vault_dir, conn_name, table, limit=10000)
  sql_query(vault_dir, conn_name, sql)
  • Connections live in vault/sql_connections.json with ${ENV_VAR}
    placeholders for passwords.
  • Every SQL string passed to sql_query is VALIDATED before
    dispatch: only single SELECT / WITH / EXPLAIN / SHOW / DESCRIBE
    statements are allowed. Multi-statement payloads (`;`), DML/DDL
    keywords (INSERT / UPDATE / DELETE / DROP / TRUNCATE / ALTER /
    CREATE / GRANT / …), and comment-cloaked writes are rejected.
  • read_sql_table defaults to 10K-row hard cap; pass limit=None
    to lift (audit-logged loudly).

Remote MongoDB (READ-ONLY by API design):
  list_mongo_connections(vault_dir)            # -> {name: uri}
  list_mongo_databases(vault_dir, conn)
  list_mongo_collections(vault_dir, conn, db)
  read_mongo_collection(vault_dir, conn, db, coll, query=None,
                        projection=None, limit=10000, skip=0, sort=None)
  mongo_aggregate(vault_dir, conn, db, coll, pipeline,
                  allow_disk_use=False)
  mongo_count(vault_dir, conn, db, coll, query=None)
  mongo_distinct(vault_dir, conn, db, coll, field, query=None)
  • Connections live in vault/mongo_connections.json with ${ENV_VAR}
    placeholders.
  • Only find / aggregate / count / distinct are reachable from
    these helpers — no insert / update / delete / drop methods are
    exposed.
  • Aggregation pipelines are VALIDATED — $out / $merge / $function
    / $accumulator / $where stages are rejected (they can write or
    run server-side JS).
  • Default 10K-row hard cap on read_mongo_collection.

Audit log:
  Every database query lands in vault/db_audit.log as one JSONL
  record per query — timestamp, connection name, query (truncated
  to 500 chars), result row count, duration. Forensic, not
  preventive — but if anything ever leaks through, the log tells
  you exactly what happened.

Manufacturing / SPC helpers (analyst_helpers.spc):
  process_capability(series, lsl=None, usl=None, subgroup_size=None,
                     column=None)
    → dict with Cp / Cpk (short-term, needs subgroup_size) and
      Pp / Ppk (long-term) plus a normality test result. ALWAYS
      check normality_ok before reporting Cpk — non-normal data
      makes Cpk misleading. Returns warnings the model should
      mention. Either lsl or usl may be None for one-sided specs.
      For a multi-column DataFrame / CSV, pass column='<name>'.
  control_chart_limits(series, chart_type='xbar', subgroup_size=None,
                       column=None)
    → dict with center, UCL, LCL. Supports 'xbar', 'r', 'i'
      (individuals), 'mr' (moving range), 'p' (proportion),
      'np' (count). Constants hardcoded from NIST/ASTM.

For manufacturing data with specification limits, USE process_capability
and check normality_ok before reporting Cpk values.

Rules:
- Assign the final answer to a DataFrame named `result_df`.
- No filesystem writes, no network, no subprocess, no eval/exec.
- Prefer helpers over hand-rolled pandas when one fits.
- PREFER AGGREGATIONS OVER RAW ROWS. The result is shown to a
  language model with a tight context budget, so:
  • `df.describe()`, `df.groupby(...).agg(...)`, `value_counts()`,
    `df[col].mean()/sum()/count()`, scalar answers — strongly
    preferred.
  • Return raw rows only if the user explicitly asked for "rows",
    "examples", or named specific records they want to see.
  • When raw rows are warranted, cap at a small head/tail
    (e.g. `df.head(10)` or top-N by ranking) and let the summary
    do the heavy lifting.
- Output one or two summary lines if a single scalar suffices —
  e.g. `result_df = pd.DataFrame([{'metric': 'total_rows', 'value': N}])`.

Examples:

Q: How many rows are in each CSV?
result_df = count_rows_per_csv(DATA_FOLDERS)

Q: Give me a true data summary of the files in this folder.
Q: What's in this subfolder?  /  Describe the files in here.
Q: Overview / inventory / profile of the data.
# `folder_data_summary` answers all of these in one call. The result
# has rows / columns / dtypes / missing% / sample value per file —
# everything the user actually means by "summary" without making
# the model hand-roll per-file logic.
result_df = folder_data_summary(DATA_FOLDERS)

Q: What is the average rating for each CSV, ignoring zeros?
result_df = average_numeric_column_per_csv(DATA_FOLDERS, column_name="rating", exclude_zero=True)

Q: Count games whose all_platforms column contains PlayStation 5.
result_df = count_matching_rows_per_csv(DATA_FOLDERS, column_name="all_platforms",
                                        contains="PlayStation 5")

Q: What percentage of games are on PlayStation vs Xbox?
# Two count_matching_rows queries concatenated — match_pct column is the answer.
ps = count_matching_rows_per_csv(DATA_FOLDERS, column_name="all_platforms",
                                 contains="PlayStation")
xb = count_matching_rows_per_csv(DATA_FOLDERS, column_name="all_platforms",
                                 contains="Xbox")
ps.insert(0, "platform", "PlayStation")
xb.insert(0, "platform", "Xbox")
result_df = pd.concat([ps, xb], ignore_index=True)

Q: Show all rows where the developer is Rockstar Games.
result_df = filter_rows_across_csvs(DATA_FOLDERS, column_name="developers",
                                    contains="Rockstar Games", max_rows_per_csv=50)

Q: What is the highest value in each "energy" column of merged_sites.xlsx?
# Merged-header workbook — top row has groups "Site A", "Site B", ...,
# row 2 has "energy, voltage, current" under each. Use the merged-aware
# reader so columns become "Site A / energy", "Site B / energy", etc.
df = read_excel_with_merged_headers("merged_sites.xlsx", sheet=0)
energy_cols = [c for c in df.columns if c.lower().endswith("/ energy")]
result_df = pd.DataFrame({
    "column": energy_cols,
    "max":    [pd.to_numeric(df[c], errors="coerce").max() for c in energy_cols],
})

Q: What is the correlation between metacritic and user_rating?
result_df = correlation_per_csv(DATA_FOLDERS, x_col="metacritic", y_col="user_rating")

Now answer the user question. Output only the code, no fences.
"""


def extract_python_code(model_text: str) -> str:
    """Strip optional ```python``` fences from model output."""
    text = (model_text or "").strip()
    fence = re.search(r"```(?:python)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL)
    if fence:
        return fence.group(1).strip()
    return text


# ============================================================
# Computational question detection
# ============================================================

_COMPUTE_KEYWORDS = (
    "how many", "count", "number of", "total",
    "average", "mean", "median", "std", "standard deviation",
    "sum ", "min ", "max ", "minimum", "maximum",
    "percent", "percentage", "%", "ratio", "proportion",
    "compare", " vs ", " versus ",
    "correlation", "corr ",
    "group by", "groupby", "by month", "by year", "by category",
    "filter", "rows with", "rows where", "rows that",
    "which file", "which files", "across files", "across the files",
    # Data-summary intents — the Council tab needs to handle queries
    # like "give me a true data summary of files in the sales/
    # subfolder" through the analyst pipeline (not freeform text from
    # the model, which can't actually count rows or read schemas).
    "summary of files", "summary of the files",
    "data summary", "true data summary",
    # Stats-summary intents — route to the bounded, cache-backed
    # folder_column_stats path (NOT model code-gen, which over 200+
    # files OOMs the app). Bare "stats"/"statistics" is inherently a
    # computational ask, so it belongs in the analyst pipeline.
    "stats", "statistics", "summary of stats", "stats summary",
    "summary statistics", "column stats", "min max mean", "min/max/mean",
    "summarize the files", "summarize files",
    "summarize the data", "summarize this folder",
    "describe the files", "describe these files", "describe the data",
    "overview of files", "overview of the files",
    "overview of the folder", "overview of the data",
    "what's in", "whats in", "what is in",
    "schema of", "schemas of", "schemas in",
    # Schema-style questions about a single file. These look like
    # freeform chat ("what columns…") but they're answerable from the
    # vault index alone — route them through the analyst so the model
    # doesn't hallucinate column names.
    "what columns", "which columns", "list columns",
    "what fields", "which fields", "list fields",
    "what headers", "which headers", "list headers",
    "column names", "field names", "header names",
    "profile the", "profile this",
    "inventory", "inventory of", "file inventory",
    "audit the", "audit this", "data quality",
    # Ranking / extremum questions — top-N, highest, lowest, biggest.
    "top ", "bottom ", "highest", "lowest", "biggest", "smallest",
    "most ", "least ", "largest", "rank by",
)


def looks_computational(query: str) -> bool:
    """Heuristic: does the user's question want a numeric/aggregate answer?"""
    q = (query or "").lower()
    return any(kw in q for kw in _COMPUTE_KEYWORDS)


# ============================================================
# Filename + subfolder hint resolution
# ============================================================
# Two helpers used by the analyst step BEFORE it builds the model
# prompt:
#
#   resolve_subfolder_hint(query, base_folder) — if the user said
#       something like "in the test_data folder" / "inside projects/"
#       and a real subfolder matches, return its path so the analyst
#       restricts its scope. The inventory the model sees becomes
#       smaller and more relevant, eliminating cross-folder accidents
#       where the model grabs an unrelated CSV by name match.
#
#   resolve_filename_hints(query, allowed_folders) — extract any
#       filename-shaped or filename-like tokens from the user's
#       question and fuzzy-match each against the actual file listing.
#       Returns (user_token → resolved_path) pairs so the prompt
#       builder can tell the model "the user said 'sales report' but
#       the actual file is 'Sales_Q3_2024.csv'."
#
# Both are intentionally cheap and dependency-free (regex + difflib
# only) — they run on every analyst call and must not slow the
# pipeline down.

import difflib as _difflib
import re as _re_va


# Tokens that look like real English words rather than file references —
# excluded from fuzzy filename matching so "average" / "column" / etc.
# don't accidentally resolve to a similarly-spelled CSV.
_COMMON_WORD_DENYLIST = frozenset({
    "average", "column", "columns", "value", "values", "data",
    "table", "tables", "row", "rows", "file", "files", "folder",
    "folders", "vault", "report", "reports", "summary", "summaries",
    "result", "results", "answer", "calculate", "compute", "find",
    "show", "list", "give", "tell", "what", "where", "which", "many",
    "much", "total", "count", "average", "sum", "mean", "median",
    "min", "max", "the", "this", "that", "these", "those", "and",
    "but", "with", "without", "from", "into", "onto", "over", "under",
    "above", "below", "near", "containing", "contains", "include",
    "includes", "exclude", "excludes", "excluding", "including",
    "specific", "general", "across", "every", "each", "all", "any",
    "have", "has", "are", "was", "were", "been", "being", "for",
})

_FILENAME_TOKEN_RE = _re_va.compile(
    # Tokens that look like filenames: word chars / dashes / underscores
    # / dots, with a recognised data-file extension.
    r"\b([\w\-\.]+\.(?:csv|tsv|xlsx?|xlsm|parquet|json|sqlite3?|db|"
    r"duckdb|bson|h5|hdf5|d3dpipeline|gz))\b",
    _re_va.IGNORECASE,
)

_QUOTED_TOKEN_RE = _re_va.compile(
    # Anything in single, double, or backtick quotes — these are usually
    # the user trying to refer to a specific file by name.
    r"""['"`]([^'"`\n]{2,80})['"`]""",
)

_SUBFOLDER_HINT_RE = _re_va.compile(
    # "(in|inside|under|within|from) [the] [<noun>] <name>"
    # where <noun> is folder/directory/subfolder/subdirectory (optional).
    r"\b(?:in|inside|under|within|from)\s+"
    r"(?:my\s+|the\s+|our\s+)?"
    r"(?:folder\s+|directory\s+|dir\s+|subfolder\s+|subdirectory\s+)?"
    # Capture group: a path-like token. Allowed chars cover:
    #   word chars, dot, dash, slash, backslash — plus ':' so Windows
    #   absolute paths ("C:\Users\...") and '~' for home shortcuts
    #   ("~/data") are captured intact instead of being chopped at the
    #   first colon/tilde. Length-bounded so we don't swallow the whole
    #   sentence on edge cases.
    r"([A-Za-z0-9_~][A-Za-z0-9_\-./\\:~]{0,120})"
    r"(?:\s+(?:folder|directory|subfolder|subdirectory))?\b",
    _re_va.IGNORECASE,
)


def _extract_candidate_filename_tokens(query: str) -> List[str]:
    """Pull every plausible filename reference out of the user's text.

    Returns a list of strings in the order they appear, deduplicated
    case-insensitively. Catches:
      • Explicit filenames with extensions (sales.csv, Q3_data.xlsx)
      • Quoted strings ("Sales Report", 'Q3-2024')
    Strips trailing punctuation. Common English words are filtered out.
    """
    if not query:
        return []
    seen: set = set()
    out: list = []

    # 1) Explicit filename-with-extension matches
    for m in _FILENAME_TOKEN_RE.finditer(query):
        tok = m.group(1).strip(".,;:!?)(\"' `")
        key = tok.lower()
        if key and key not in seen and key not in _COMMON_WORD_DENYLIST:
            seen.add(key)
            out.append(tok)

    # 2) Quoted strings — could be a filename without an extension, or
    #    a multi-word reference like "Q3 sales"
    for m in _QUOTED_TOKEN_RE.finditer(query):
        tok = m.group(1).strip()
        # Strip likely sentence punctuation but preserve internal dots
        tok = tok.strip(",;:!?)(`")
        key = tok.lower()
        if (key and key not in seen
                and key not in _COMMON_WORD_DENYLIST
                and len(key) >= 3):
            seen.add(key)
            out.append(tok)

    return out


def _fuzzy_match_filename(token: str, candidates: List[Path],
                          cutoff: float = 0.55) -> Optional[Path]:
    """Pick the best fuzzy match of `token` against `candidates` paths.

    Matches against each candidate's filename (with AND without
    extension) and picks the highest similarity score above the cutoff.
    Returns None if no candidate clears the bar — better to surface
    "no match" to the user than to confidently pick a wrong file.

    For very short tokens (< 6 chars, like "a.csv" or "x") the
    SequenceMatcher ratio is too noisy — "a.csv" vs "b.csv" yields
    0.8 by character overlap alone, which would silently grab the
    wrong file. We raise the cutoff to 0.85 for tokens of length 5
    and require exact/substring match for tokens of length ≤ 4.
    """
    if not token or not candidates:
        return None
    token_lc = token.lower()
    # Exact substring match wins immediately — most natural user
    # behaviour ("sales.csv" should always pick the file containing
    # 'sales.csv' even if difflib quirks would prefer something else).
    for p in candidates:
        if token_lc == p.name.lower() or token_lc == p.stem.lower():
            return p
    for p in candidates:
        if token_lc in p.name.lower():
            return p

    # Very short tokens: refuse to fuzzy-match. The ratio is too noisy
    # and a wrong silent pick is worse than no pick.
    if len(token_lc) <= 4:
        return None
    # Short-but-not-tiny tokens (5 chars): demand a higher confidence.
    effective_cutoff = max(cutoff, 0.85) if len(token_lc) == 5 else cutoff

    # Fall back to difflib ratio against both .name and .stem forms.
    best_score = 0.0
    best_path: Optional[Path] = None
    for p in candidates:
        for cand in (p.name.lower(), p.stem.lower()):
            ratio = _difflib.SequenceMatcher(None, token_lc, cand).ratio()
            if ratio > best_score:
                best_score = ratio
                best_path = p
    return best_path if best_score >= effective_cutoff else None


def resolve_filename_hints(query: str, allowed_folders: List[Path]
                           ) -> List[Tuple[str, Optional[Path]]]:
    """Extract candidate filename tokens from `query` and fuzzy-resolve
    each one against the data inventory of `allowed_folders`.

    Returns a list of ``(user_token, resolved_path_or_None)`` in input
    order. ``resolved_path`` is None when no file cleared the fuzzy
    cutoff — the caller passes this through to the prompt so the model
    is *told* the user's reference didn't match anything and refrains
    from inventing a file.
    """
    tokens = _extract_candidate_filename_tokens(query)
    if not tokens:
        return []
    # Build the inventory once — every supported data-file type.
    inventory: list = []
    try:
        inventory.extend(list_data_files(allowed_folders, recursive=True))
    except Exception:
        pass
    try:
        inventory.extend(list_parquet_files(allowed_folders, recursive=True))
    except Exception:
        pass
    try:
        inventory.extend(list_sqlite_files(allowed_folders, recursive=True))
    except Exception:
        pass
    try:
        inventory.extend(list_duckdb_files(allowed_folders, recursive=True))
    except Exception:
        pass
    try:
        inventory.extend(list_bson_files(allowed_folders, recursive=True))
    except Exception:
        pass
    # Dedupe by resolved path
    seen_paths: set = set()
    unique_inv: list = []
    for p in inventory:
        try:
            key = str(p.resolve()).lower()
        except Exception:
            key = str(p).lower()
        if key not in seen_paths:
            seen_paths.add(key)
            unique_inv.append(p)
    return [(tok, _fuzzy_match_filename(tok, unique_inv)) for tok in tokens]


def resolve_subfolder_hint(query: str, base_folder: Path
                           ) -> Optional[Path]:
    """If the user mentioned a subfolder name and it exists under
    `base_folder`, return the resolved path. Otherwise None.

    Matches phrases like "in the projects folder" / "inside test_data"
    / "from Q3_2024/" against the actual immediate-child directories
    of `base_folder` (case-insensitive, dash/underscore-tolerant).
    Falls back to fuzzy matching when no exact match exists, with a
    higher cutoff (0.7) than filenames — folder names are typically
    short and a sloppy match here would scope the analyst to the wrong
    data entirely.
    """
    if not query or not base_folder:
        return None
    try:
        base_folder = Path(base_folder)
        if not base_folder.is_dir():
            return None
    except Exception:
        return None

    # Build the candidate list: every directory under base_folder, up to
    # 2 levels deep. We recurse so "in the Q3 folder" matches both
    # `data_in/Q3/` and `data_in/projects/Q3/`.
    children: list = []
    try:
        for p in base_folder.iterdir():
            if p.is_dir():
                children.append(p)
                # One more level
                try:
                    for q in p.iterdir():
                        if q.is_dir():
                            children.append(q)
                except Exception:
                    pass
    except Exception:
        return None
    if not children:
        return None

    def _norm(s: str) -> str:
        return s.lower().replace("-", "_").replace(" ", "_")

    # Try each match position in the query and resolve to a real folder
    for m in _SUBFOLDER_HINT_RE.finditer(query):
        cand = m.group(1).strip().strip("'\"`")
        # Don't strip leading slashes from absolute paths ("/home/...") but
        # do strip path separators that snuck in via the regex tail.
        if cand and not cand.startswith(("/", "\\")) and not (
            len(cand) >= 2 and cand[1] == ":"
        ):
            cand = cand.strip("/\\")
        if not cand or cand.lower() in _COMMON_WORD_DENYLIST:
            continue
        # Skip tokens that look like filenames (they have data-file extensions)
        if "." in cand and cand.rsplit(".", 1)[1].lower() in (
            "csv", "tsv", "xlsx", "xls", "xlsm", "parquet", "json",
            "sqlite", "sqlite3", "db", "duckdb", "bson", "h5", "hdf5",
            "d3dpipeline", "gz",
        ):
            continue

        # Absolute / home-relative paths: if the user pasted a full path
        # AND it lives under base_folder AND it's a real directory,
        # return it directly. Without this, "in C:\Users\me\.council\
        # vault\data_in\Q3" silently fell through to fuzzy matching
        # against immediate children of data_in/, which never resolves.
        try:
            abs_cand = Path(cand).expanduser()
        except Exception:
            abs_cand = None
        if abs_cand is not None and abs_cand.is_absolute() and abs_cand.is_dir():
            try:
                abs_cand.resolve().relative_to(base_folder.resolve())
                return abs_cand.resolve()
            except (ValueError, OSError):
                # Outside base_folder — don't expose paths the analyst
                # has no business reading.
                pass

        cand_norm = _norm(cand)

        # Exact-name match first
        for child in children:
            if _norm(child.name) == cand_norm:
                return child

        # Path-prefix match ("projects/Q3" → data_in/projects/Q3).
        # Strip the absolute-path prefix if the user gave one inside
        # base_folder but we somehow missed the abs-path branch above.
        if "/" in cand or "\\" in cand:
            walk_parts = [p for p in _re_va.split(r"[/\\]", cand) if p]
            # Skip Windows drive letter ("C:") if present
            if walk_parts and len(walk_parts[0]) == 2 and walk_parts[0][1] == ":":
                walk_parts = walk_parts[1:]
            sub = base_folder
            for part in walk_parts:
                next_match = None
                try:
                    for child in sub.iterdir():
                        if child.is_dir() and _norm(child.name) == _norm(part):
                            next_match = child
                            break
                except Exception:
                    break
                if not next_match:
                    sub = None
                    break
                sub = next_match
            if sub is not None and sub.is_dir() and sub != base_folder:
                return sub

        # Fuzzy fallback — higher cutoff than filenames (folder
        # mismatches are more costly than file mismatches).
        best_score = 0.0
        best_dir: Optional[Path] = None
        for child in children:
            ratio = _difflib.SequenceMatcher(
                None, cand_norm, _norm(child.name)
            ).ratio()
            if ratio > best_score:
                best_score = ratio
                best_dir = child
        if best_score >= 0.7 and best_dir is not None:
            return best_dir

    return None


def format_filename_hints(hints: List[Tuple[str, Optional[Path]]],
                          base_folder: Optional[Path] = None,
                          max_hints: int = 8) -> str:
    """Render filename-hint pairs into a short block for the prompt.

    The block is appended to `build_pandas_code_prompt` output so the
    model uses the resolved filenames instead of guessing.

    Cap the number of hints rendered at `max_hints` so a verbose
    question with 20 quoted phrases doesn't blow the prompt budget.
    Excess hints are summarised in a single trailer line.
    """
    if not hints:
        return ""
    lines = ["NOTE — Resolved file references from the user's question:"]
    rendered = hints[:max_hints]
    for tok, resolved in rendered:
        if resolved is None:
            lines.append(
                f'  • "{tok}"  → NO MATCH in inventory. Do NOT invent '
                f'a path for this; if the question requires it, the '
                f'computation cannot proceed.'
            )
        else:
            try:
                rel = (resolved.relative_to(base_folder)
                       if base_folder else resolved)
            except Exception:
                rel = resolved
            lines.append(f'  • "{tok}"  → "{rel}"')
    if len(hints) > max_hints:
        extra = len(hints) - max_hints
        lines.append(
            f"  ... ({extra} more file reference{'s' if extra != 1 else ''} "
            f"not shown — use the resolved names above, refuse to invent.)"
        )
    return "\n".join(lines)


# ============================================================
# DataFrame -> prompt-friendly text
# ============================================================

def format_result_for_prompt(
    df: pd.DataFrame,
    *,
    max_rows: int = 30,
    max_chars: int = 4000,
    max_tokens: Optional[int] = None,
    count_tokens: Optional[Any] = None,
) -> str:
    """Render an analyst DataFrame for prompt injection.

    Truncation strategy (head + middle + tail) — pandas-style summary
    rows ("Total" / "Mean" / "..." that helpers like ``describe()`` and
    ``agg()`` append) tend to live at the TAIL of the result. Head-only
    truncation throws them away. We keep the head AND the tail and elide
    the middle so the model sees both the first rows and the
    aggregations.

    Budget knobs:
      - ``max_rows`` / ``max_chars`` — legacy char-based caps (kept as
        a hard backstop).
      - ``max_tokens`` / ``count_tokens`` — token-aware cap when the
        caller has the tokenizer in hand. When ``max_tokens`` is given,
        we re-render with progressively fewer rows until the result
        fits, or fall back to head+tail char truncation if even one
        row exceeds the budget.
    """
    if df is None or df.empty:
        return "(analyst returned an empty result)"

    def _render_head_tail(d, n_head: int, n_tail: int) -> str:
        total = len(d)
        if n_head + n_tail >= total:
            return d.to_string(index=False)
        head = d.head(n_head)
        tail = d.tail(n_tail)
        head_txt = head.to_string(index=False)
        # tail.to_string() repeats the header row; strip it so the
        # output reads continuously.
        tail_txt = tail.to_string(index=False)
        tail_lines = tail_txt.split("\n", 1)
        tail_body = tail_lines[1] if len(tail_lines) > 1 else tail_lines[0]
        omitted = total - n_head - n_tail
        return (head_txt + "\n... (" + str(omitted) + " row"
                + ("s" if omitted != 1 else "") + " omitted from middle)\n"
                + tail_body)

    # Token-aware path — try progressively smaller head+tail splits
    # until we fit. We bias toward keeping the TAIL (where summary rows
    # live) by giving it slightly more rows than the head.
    if max_tokens is not None and count_tokens is not None:
        for n_head, n_tail in [(20, 20), (12, 12), (8, 8), (5, 5),
                               (3, 3), (2, 2), (1, 1)]:
            text = _render_head_tail(df, n_head, n_tail)
            if count_tokens(text) <= max_tokens:
                return text
        # Even one head + one tail row blows the budget — fall through
        # to the char-cap path below as a last resort.

    # Char-based path (legacy default).
    n_head = min(max_rows, max(1, len(df) - max_rows // 3))
    n_tail = min(max_rows // 3, max(1, len(df) - n_head))
    text = _render_head_tail(df, n_head, n_tail)
    if len(text) > max_chars:
        # Last-resort char trim. Keep head + last few lines so summary
        # rows still come through even when the row count is huge.
        keep_head = int(max_chars * 0.7)
        keep_tail = max_chars - keep_head - 64
        text = (text[:keep_head] + "\n... (truncated middle)\n"
                + text[-max(keep_tail, 0):]) if keep_tail > 0 else text[:max_chars]
    return text
